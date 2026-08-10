#!/usr/bin/env python3
"""Focused integration coverage for dashboard catalog and image readiness gates."""

from __future__ import annotations

import base64
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import catalog_repository
import dashboard_app


VALID_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def _write_catalog(path: Path, *, title: str = "Original title") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listings"
    sheet.cell(row=4, column=2, value="product-04")
    sheet.cell(row=4, column=8, value=title)
    workbook.save(path)
    workbook.close()


class DashboardCatalogPatchIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.shop_id = "templystudios"
        self.workbook = self.root / "shops" / self.shop_id / "Etsy_SEO_Generator.xlsx"
        _write_catalog(self.workbook)
        self.client = TestClient(dashboard_app.app, base_url="http://127.0.0.1:8090")
        self.token_patch = patch.object(dashboard_app, "_DASHBOARD_MUTATION_TOKEN", "integration-token")
        self.token_patch.start()

    def tearDown(self) -> None:
        self.token_patch.stop()
        self.temp_dir.cleanup()

    def _mutation_headers(self) -> dict[str, str]:
        return {
            "Host": "127.0.0.1:8090",
            "Origin": "http://127.0.0.1:8090",
            dashboard_app._DASHBOARD_MUTATION_TOKEN_HEADER: "integration-token",
        }

    def _read_title(self) -> str:
        workbook = load_workbook(self.workbook, read_only=True)
        try:
            return str(workbook["Listings"].cell(row=4, column=8).value)
        finally:
            workbook.close()

    def test_canonical_patch_uses_repository_and_updates_intended_field(self) -> None:
        with patch.object(dashboard_app, "BASE_DIR", self.root), patch.object(
            dashboard_app, "_active_shop_id", self.shop_id
        ), patch.object(
            dashboard_app.catalog_repository,
            "apply_catalog_update",
            wraps=dashboard_app.catalog_repository.apply_catalog_update,
        ) as repository_update:
            response = self.client.patch(
                "/api/products/4",
                headers=self._mutation_headers(),
                json={
                    "title": "Updated title",
                    "custom_product_field": "must remain an arbitrary input field",
                    "_expected_hash": catalog_repository._catalog_hash(self.workbook),
                },
            )

        self.assertEqual(200, response.status_code, response.text)
        repository_update.assert_called_once()
        self.assertEqual("Updated title", self._read_title())

    def test_stale_patch_precondition_returns_409_without_overwrite(self) -> None:
        original_hash = catalog_repository._catalog_hash(self.workbook)
        with patch.object(dashboard_app, "BASE_DIR", self.root), patch.object(
            dashboard_app, "_active_shop_id", self.shop_id
        ):
            response = self.client.patch(
                "/api/products/4",
                headers=self._mutation_headers(),
                json={"title": "Should not be written", "_expected_hash": "stale-hash"},
            )

        self.assertEqual(409, response.status_code, response.text)
        self.assertEqual(original_hash, catalog_repository._catalog_hash(self.workbook))
        self.assertEqual("Original title", self._read_title())

    def test_explicit_custom_workbook_keeps_legacy_writer_behavior(self) -> None:
        custom_workbook = self.root / "imported" / "custom.xlsx"
        _write_catalog(custom_workbook)

        dashboard_app.save_to_excel(4, {"title": "Custom update"}, excel_path=custom_workbook)

        workbook = load_workbook(custom_workbook, read_only=True)
        try:
            self.assertEqual("Custom update", workbook["Listings"].cell(row=4, column=8).value)
        finally:
            workbook.close()


class DashboardImageReadinessIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.shop_root = self.root / "shops" / "templystudios"
        self.image_dir = self.shop_root / "product-04" / "images"
        self.image_dir.mkdir(parents=True)
        self.client = TestClient(dashboard_app.app, base_url="http://127.0.0.1:8090")
        self.token_patch = patch.object(dashboard_app, "_DASHBOARD_MUTATION_TOKEN", "integration-token")
        self.token_patch.start()

    def tearDown(self) -> None:
        self.token_patch.stop()
        self.temp_dir.cleanup()

    def _mutation_headers(self) -> dict[str, str]:
        return {
            "Host": "127.0.0.1:8090",
            "Origin": "http://127.0.0.1:8090",
            dashboard_app._DASHBOARD_MUTATION_TOKEN_HEADER: "integration-token",
        }

    def _upload(self, *uploads: tuple[str, bytes]) -> dict:
        files = [
            ("files", (name, payload, "image/png"))
            for name, payload in uploads
        ]
        with patch.object(
            dashboard_app,
            "get_product_by_row",
            return_value={"folder": "product-04"},
        ), patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_root):
            response = self.client.post(
                "/api/products/4/images",
                headers=self._mutation_headers(),
                files=files,
            )
        return response

    def test_valid_upload_is_readiness_checked_then_saved_with_safe_name(self) -> None:
        response = self._upload(("../hero.png", VALID_PNG))

        self.assertEqual(200, response.status_code, response.text)
        self.assertEqual(["hero.png"], response.json()["saved"])
        self.assertEqual(VALID_PNG, (self.image_dir / "hero.png").read_bytes())

    def test_corrupt_upload_has_no_partial_final_write(self) -> None:
        response = self._upload(("valid.png", VALID_PNG), ("corrupt.png", b"not-an-image"))

        self.assertEqual(400, response.status_code, response.text)
        self.assertIn("readiness", response.json()["detail"].lower())
        self.assertEqual([], sorted(path.name for path in self.image_dir.iterdir()))

    def test_image_push_is_blocked_by_invalid_local_asset(self) -> None:
        (self.image_dir / "broken.png").write_bytes(b"not-an-image")
        product = {
            "row": 4,
            "folder": "product-04",
            "etsy_url": "https://www.etsy.com/listing/123456789",
        }
        with patch.object(dashboard_app, "BASE_DIR", self.root), patch.object(
            dashboard_app, "_active_shop_id", "templystudios"
        ), patch.object(dashboard_app, "get_product_by_row", return_value=product):
            response = self.client.post(
                "/api/products/4/push-to-etsy",
                headers=self._mutation_headers(),
                json={"fields": ["images"]},
            )

        self.assertEqual(400, response.status_code, response.text)
        self.assertIn("readiness", response.json()["detail"].lower())

    def test_metadata_only_push_does_not_run_image_readiness_gate(self) -> None:
        product = {
            "row": 4,
            "folder": "product-04",
            "etsy_url": "https://www.etsy.com/listing/123456789",
        }

        class FakeJobStore:
            def create_or_get_deduplicated_job(self, **kwargs):
                return ({"job_id": "job-1", "status": "queued", "created_at": 1}, True)

        def fake_create_task(coroutine):
            coroutine.close()
            return object()

        with patch.object(dashboard_app, "BASE_DIR", self.root), patch.object(
            dashboard_app, "_active_shop_id", "templystudios"
        ), patch.object(dashboard_app, "get_product_by_row", return_value=product), patch.object(
            dashboard_app, "_get_job_store", return_value=FakeJobStore()
        ), patch.object(dashboard_app, "_etsy_update_shop_is_busy", return_value=False
        ), patch.object(dashboard_app.asyncio, "create_task", side_effect=fake_create_task), patch.object(
            dashboard_app, "_register_background_task"
        ), patch.object(dashboard_app, "AssetReadinessEngine") as readiness_engine:
            response = self.client.post(
                "/api/products/4/push-to-etsy",
                headers=self._mutation_headers(),
                json={"fields": ["title"]},
            )

        self.assertEqual(200, response.status_code, response.text)
        readiness_engine.assert_not_called()


if __name__ == "__main__":
    unittest.main()
