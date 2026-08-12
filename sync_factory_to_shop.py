#!/usr/bin/env python3
"""
sync_factory_to_shop.py - Synchronize and watch Etsy Image Factory output and import to dashboard.

Usage:
  python3 sync_factory_to_shop.py [--shop templystudios] [--watch] [--poll 5]
"""
import argparse
import json
import time
import shutil
import re
import urllib.parse
from pathlib import Path
import openpyxl

from shop_asset_workflow import (
    copy_image_with_watermark,
    get_watermark_text,
)

BASE_DIR = Path(__file__).parent
ACTIVE_SHOP_FILE = BASE_DIR / "active_shop.txt"
FACTORY_SHOP_ID = "templystudios"
SRC_DIR = BASE_DIR / "shops" / FACTORY_SHOP_ID
IMG_EXTS = {".png", ".jpg", ".jpeg", ".webp"}
SRC_IMAGE_DIR = "etsy-images-chatgpt-final"
SRC_ZIP_DIR = "source-zip"
FACTORY_EXCLUDED_SOURCE_NAMES = {
    "_deleted_products",
    "_asset_quarantine",
    "_failed_local_imports",
    "master_products",
}
PRODUCT_FOLDER_RE = re.compile(r"^product-(\d+)$")
CONFIG_FILE = BASE_DIR / "shops_config.json"


def load_active_shop_id() -> str:
    try:
        return ACTIVE_SHOP_FILE.read_text(encoding="utf-8").strip()
    except OSError:
        return ""


def get_active_shop_id() -> str:
    shop_id = load_active_shop_id()
    if shop_id in {"templystudios", "daisyflowdigital"}:
        return shop_id
    return ""


def assert_temple_shop_active() -> bool:
    return get_active_shop_id() == FACTORY_SHOP_ID


def _factory_file_is_usable(path: Path, suffixes: set[str] | None = None) -> bool:
    if path.is_symlink():
        return False
    if not path.is_file() or path.name == ".DS_Store":
        return False
    if suffixes is not None and path.suffix.lower() not in suffixes:
        return False
    try:
        stat = path.stat()
    except OSError:
        return False
    return stat.st_size > 0 and getattr(stat, "st_blocks", 1) > 0


def _is_supported_factory_source_dir(source_dir: Path) -> bool:
    if (
        not source_dir.is_dir()
        or source_dir.is_symlink()
        or source_dir.name.startswith(".")
        or source_dir.name in FACTORY_EXCLUDED_SOURCE_NAMES
        or PRODUCT_FOLDER_RE.fullmatch(source_dir.name) is not None
    ):
        return False
    images_dir = source_dir / SRC_IMAGE_DIR
    zip_dir = source_dir / SRC_ZIP_DIR
    return images_dir.is_dir() and zip_dir.is_dir()


def load_shop_config(shop_id: str) -> dict:
    try:
        with CONFIG_FILE.open("r", encoding="utf-8") as f:
            return json.load(f).get(shop_id, {})
    except (OSError, json.JSONDecodeError):
        return {}


def clean_html(text: str) -> str:
    return (text or "").replace("&amp;", "&").replace("&quot;", '"').replace("&#39;", "'").strip()

def normalize_name(name: object) -> str:
    text = str(name or "").lower()
    text = text.replace("-", " ").replace("_", " ").strip()
    return re.sub(r"\s+", " ", text)

def get_factory_images(folder: Path) -> list[Path]:
    images_dir = folder / SRC_IMAGE_DIR
    if not images_dir.is_dir():
        return []
    return sorted(
        [f for f in images_dir.iterdir() if _factory_file_is_usable(f, IMG_EXTS)],
        key=lambda p: (p.name != "01_hero_image.png", p.name.lower())
    )

def get_factory_files(folder: Path) -> list[Path]:
    files_dir = folder / SRC_ZIP_DIR
    if not files_dir.is_dir():
        return []
    return sorted(
        [f for f in files_dir.iterdir() if _factory_file_is_usable(f, {".zip"})],
        key=lambda p: p.name.lower()
    )

def is_folder_stable(folder: Path, wait_secs: int = 5) -> bool:
    """Checks if files size/mtime haven't changed for wait_secs seconds."""
    def snapshot(f):
        return {str(p): (p.stat().st_size, p.stat().st_mtime) for p in f.rglob("*") if p.is_file()}
    try:
        s1 = snapshot(folder)
        time.sleep(wait_secs)
        s2 = snapshot(folder)
        return s1 == s2
    except Exception:
        return False

def sync_shop(shop_id: str):
    if not assert_temple_shop_active() or shop_id != FACTORY_SHOP_ID:
        print(f"❌ Refusing sync: factory can only sync for shop {FACTORY_SHOP_ID}")
        return
    dst_dir = BASE_DIR / "shops" / shop_id
    shop_config = load_shop_config(shop_id)
    factory_dir = SRC_DIR
    excel_path = dst_dir / "Etsy_SEO_Generator.xlsx"

    if not dst_dir.exists():
        print(f"❌ Shop directory not found: {dst_dir}")
        return
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return
    if not factory_dir.exists():
        print(f"❌ Source directory not found: {factory_dir}")
        return

    print(f"🔄 Starting sync for shop {shop_id}...")
    print(f"   Source: {factory_dir}")
    print(f"   Destination: {dst_dir}")
    watermark = get_watermark_text(shop_id, shop_config)
    print(f"   Watermark: {watermark}")

    # Load spreadsheet
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Listings"]

    # Build mapping of existing products by their keyword only for valid product folders.
    imported_map = {}
    for r in range(4, ws.max_row + 1):
        folder = ws.cell(row=r, column=2).value
        kw = ws.cell(row=r, column=3).value
        if not PRODUCT_FOLDER_RE.fullmatch(str(folder or "")):
            continue
        folder = str(folder).strip()
        product_dir = dst_dir / folder
        if (
            not product_dir.is_dir()
            or product_dir.is_symlink()
            or product_dir.resolve().parent != dst_dir.resolve()
        ):
            continue
        if kw and str(kw).strip():
            # Match by normalized keyword (Column C)
            key = normalize_name(kw)
            if key not in imported_map:
                imported_map[normalize_name(kw)] = (str(folder), r)

    existing_numbers = []
    for item in dst_dir.glob("product-*"):
        if not item.is_dir():
            continue
        try:
            existing_numbers.append(int(item.name.split("-", 1)[1]))
        except (IndexError, ValueError):
            continue
    next_product_num = max(existing_numbers, default=0) + 1

    # Scan factory directories with exact source layout
    factory_folders = [d for d in factory_dir.iterdir() if _is_supported_factory_source_dir(d)]

    wb_modified = False

    for src_folder in sorted(factory_folders, key=lambda d: d.name.lower()):
        images = get_factory_images(src_folder)
        files = get_factory_files(src_folder)

        if not images or not files:
            continue  # empty factory folder, skip

        if files:
            source_keyword = files[0].stem.replace("-", " ").replace("_", " ").strip()
        else:
            source_keyword = src_folder.name.replace("-", " ").replace("_", " ").strip()
        folder_keyword = normalize_name(source_keyword)

        matched = imported_map.get(folder_keyword)
        if matched:
            folder_name, row_num = matched
            print(f"  ⏭ Already imported {src_folder.name} as {folder_name} (row {row_num}); skipping to avoid overwrite.")
            continue
        else:
            # Always allocate a collision-free folder number in the target shop.
            # Never reuse the generic master_products folder ID.
            while (dst_dir / f"product-{next_product_num:02d}").exists():
                next_product_num += 1
            folder_name = f"product-{next_product_num:02d}"
            next_product_num += 1
            prod_path = dst_dir / folder_name
            img_dst = prod_path / "images"
            file_dst = prod_path / "files"
            img_dst.mkdir(parents=True, exist_ok=True)
            file_dst.mkdir(parents=True, exist_ok=True)

            copied_imgs = 0
            for f in images:
                copy_image_with_watermark(f, img_dst / f.name, watermark)
                copied_imgs += 1

            copied_files = 0
            file_stems = []
            for f in files:
                shutil.copy2(f, file_dst / f.name)
                copied_files += 1
                file_stems.append(f.stem)

            # Build keyword
            keyword = source_keyword

            # Find empty row in Excel B column or append
            target_row = None
            for r in range(4, ws.max_row + 2):
                val = ws.cell(row=r, column=2).value
                if val is None or str(val).strip() == "":
                    target_row = r
                    break
            if not target_row:
                target_row = ws.max_row + 1

            # Write Excel row
            ws.cell(row=target_row, column=1, value=target_row - 3)      # A: STT
            ws.cell(row=target_row, column=2, value=folder_name)        # B: Folder
            ws.cell(row=target_row, column=3, value=keyword)            # C: Keywords
            ws.cell(row=target_row, column=5, value=4.99)               # E: Price
            ws.cell(row=target_row, column=11, value=999)               # K: Quantity
            ws.cell(row=target_row, column=12, value="I did")           # L: Who Made
            ws.cell(row=target_row, column=13, value="2020_2026")       # M: When Made
            ws.cell(row=target_row, column=14, value="⏳ Chờ đăng")       # N: Status
            ws.cell(row=target_row, column=15, value="Digital Planner")  # O: Section

            print(f"  🆕 Imported {src_folder.name} -> {folder_name} (Row {target_row}) | Keyword: {keyword}")
            # Update map to avoid duplicate rows within this run
            imported_map[normalize_name(keyword)] = (folder_name, target_row)
            wb_modified = True

    if wb_modified:
        wb.save(excel_path)
        print("💾 Excel spreadsheet updated and saved.")
    else:
        print("✅ Sync completed. No spreadsheet changes needed.")

def _collect_source_signature(source_dir: Path) -> tuple[tuple[str, int, int], ...]:
    signature: list[tuple[str, int, int]] = []
    for f in get_factory_images(source_dir) + get_factory_files(source_dir):
        try:
            s = f.stat()
            signature.append((f.name, int(s.st_size), int(s.st_mtime_ns)))
        except OSError:
            continue
    return tuple(sorted(signature))


def get_dir_fingerprint(directory: Path) -> dict:
    """Return map of source slug -> (name,size,mtime_ns) tuples for strict layout only."""
    fingerprint = {}
    if not directory.exists():
        return fingerprint
    for item in directory.iterdir():
        if not _is_supported_factory_source_dir(item):
            continue
        fingerprint[item.name] = _collect_source_signature(item)
    return fingerprint

def watch_folders(shop_id: str, poll_interval: int):
    if not assert_temple_shop_active() or shop_id != FACTORY_SHOP_ID:
        print(f"❌ Refusing watcher: factory can only watch for shop {FACTORY_SHOP_ID}")
        return
    factory_dir = SRC_DIR
    print(f"👁 Watcher started. Monitoring: {factory_dir}")
    print(f"   Press Ctrl+C to stop.")
    
    # Pre-populate fingerprint
    known_state = get_dir_fingerprint(factory_dir)
    
    while True:
        try:
            time.sleep(poll_interval)
            if not assert_temple_shop_active() or shop_id != FACTORY_SHOP_ID:
                print(f"⚠️ Active shop is no longer {FACTORY_SHOP_ID}; waiting for switch back before syncing...")
                continue
            current_state = get_dir_fingerprint(factory_dir)
            
            # Check for additions or modifications
            changed = False
            for folder, mtime in current_state.items():
                if folder not in known_state or mtime != known_state[folder]:
                    # Folder is new or has new files
                    folder_path = factory_dir / folder
                    print(f"🔔 Detected change in Image Factory folder: {folder} — waiting to stabilize...")
                    if is_folder_stable(folder_path, wait_secs=5):
                        print(f"📈 Folder {folder} is stable. Triggering sync...")
                        sync_shop(shop_id)
                        changed = True
                    else:
                        print(f"⏳ Folder {folder} is still writing/copying files...")
            
            if changed:
                known_state = get_dir_fingerprint(factory_dir)
            else:
                # Update known state to avoid looping alerts if stable check fails once
                known_state = current_state
                
        except KeyboardInterrupt:
            print("\n👋 Watcher stopped.")
            break
        except Exception as e:
            print(f"❌ Watcher error: {e}")
            time.sleep(poll_interval)

def main():
    parser = argparse.ArgumentParser(description="Sync Image Factory to Dashboard Shop")
    parser.add_argument("--shop", default="templystudios", help="Active shop ID")
    parser.add_argument("--watch", action="store_true", help="Run in daemon watcher mode")
    parser.add_argument("--poll", type=int, default=5, help="Poll interval in seconds")
    args = parser.parse_args()
    requested_shop = (args.shop or "").strip().lower()
    if requested_shop != FACTORY_SHOP_ID:
        print(f"❌ Không cho phép --shop khác {FACTORY_SHOP_ID} cho workflow này.")
        return

    # Initial sync
    if not assert_temple_shop_active():
        print("❌ Active shop mismatch: active shop must be templystudios before sync.")
        return

    sync_shop(args.shop)

    if args.watch:
        watch_folders(args.shop, args.poll)

if __name__ == "__main__":
    main()
