#!/usr/bin/env python3
"""Focused regression tests for merge-safe digital hash guard."""

from __future__ import annotations

import tempfile
import hashlib
import json
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from etsy_catalog import merge_safe_duplicates


def make_catalog(group_id: str, folders: list[str]) -> dict:
    return {
        "counts": {"total": 1},
        "duplicate_groups": [
            {
                "group_id": group_id,
                "match_type": "matching digital asset signature",
                "match_value": "unit test guard",
                "asset_names": [],
                "folders": folders,
                "listing_ids": [],
                "safe_to_merge": True,
                "records": [
                    {"folder": folder, "row": 4 + idx}
                    for idx, folder in enumerate(folders)
                ],
            }
        ],
    }


class TestMergeDigitalGuard(unittest.TestCase):
    def _write_file(self, target: Path, content: bytes) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(content)

    def _make_workbook(self, excel_path: Path, rows: list[dict[str, str | int | None]] | None = None) -> None:
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listings"
        row_data = rows or [
            {"row": 4, "folder": "product-a"},
            {"row": 5, "folder": "product-b"},
        ]
        for item in row_data:
            row_num = int(item.get("row", 0))
            ws.cell(row=row_num, column=2).value = item.get("folder")
            if item.get("status") is not None:
                ws.cell(row=row_num, column=14).value = item.get("status")
            if item.get("etsy_url") is not None:
                ws.cell(row=row_num, column=16).value = item.get("etsy_url")
            if item.get("sku") is not None:
                ws.cell(row=row_num, column=18).value = item.get("sku")
        wb.save(excel_path)

    def _run_merge(
        self,
        base_dir: Path,
        catalog: dict,
        group_id: str,
        safe_hash_side_effect=None,
    ) -> dict:
        excel_path = base_dir / "shops" / "daisyflowdigital" / "Etsy_SEO_Generator.xlsx"
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        self._make_workbook(excel_path)
        catalog_patcher = patch("etsy_catalog.build_unified_catalog", return_value=catalog)
        hash_patcher = None
        if safe_hash_side_effect is not None:
            hash_patcher = patch("etsy_catalog._safe_hash", side_effect=safe_hash_side_effect)
        if hash_patcher is not None:
            with catalog_patcher, hash_patcher:
                return merge_safe_duplicates(base_dir, "daisyflowdigital", excel_path, [group_id])
        with catalog_patcher:
            return merge_safe_duplicates(base_dir, "daisyflowdigital", excel_path, [group_id])

    def test_merge_guard_blocks_non_matching_digital_sets(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            folder_a = shop_dir / "product-a" / "files"
            folder_b = shop_dir / "product-b" / "files"
            self._write_file(folder_a / "a.pdf", b"same")
            self._write_file(folder_b / "b.pdf", b"different")

            catalog = make_catalog("group-guard-block", ["product-a", "product-b"])
            result = self._run_merge(base_dir, catalog, "group-guard-block")

            self.assertEqual([], result["merged"])
            self.assertEqual(1, len(result.get("skipped", [])))
            self.assertIn("hash set mismatch", result["skipped"][0]["reason"])
            self.assertTrue((shop_dir / "product-a").is_dir())
            self.assertTrue((shop_dir / "product-b").is_dir())

    def test_canonical_transfer_maps_etsy_data_and_quarantines_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            excel_path = shop_dir / "Etsy_SEO_Generator.xlsx"
            self._make_workbook(excel_path, rows=[
                {
                    "row": 4,
                    "folder": "product-669",
                    "etsy_url": "https://www.etsy.com/listing/451234",
                    "status": "✅ Đã đăng",
                    "sku": "dd_product_669",
                },
                {"row": 5, "folder": "product-04"},
            ])

            self._write_file(shop_dir / "product-669" / "files" / "shared.pdf", b"same")
            self._write_file(shop_dir / "product-04" / "files" / "shared.pdf", b"same")
            self._write_file(shop_dir / "product-04" / "images" / "hero.jpg", b"B" * 2048)

            catalog = make_catalog("group-transfer", ["product-669", "product-04"])
            catalog["duplicate_groups"][0]["records"][0]["listing_id"] = "451234"
            catalog["duplicate_groups"][0]["records"][0]["total_bytes"] = 5
            catalog["duplicate_groups"][0]["records"][1]["total_bytes"] = 4096

            catalog_patcher = patch("etsy_catalog.build_unified_catalog", return_value=catalog)
            with catalog_patcher:
                result = merge_safe_duplicates(base_dir, "daisyflowdigital", excel_path, ["group-transfer"])

            self.assertEqual(1, len(result.get("merged", [])))
            self.assertEqual([], result.get("skipped", []))
            merged = result["merged"][0]
            self.assertEqual("product-04", merged["canonical_folder"])
            self.assertEqual(["product-669"], merged["moved_folders"])

            manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
            quarantine = Path(manifest["quarantine"])
            self.assertTrue((shop_dir / "product-04").is_dir())
            self.assertFalse((shop_dir / "product-669").exists())
            self.assertTrue((quarantine / "product-669").is_dir())

            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Listings"]
            self.assertEqual("https://www.etsy.com/listing/451234", ws.cell(row=5, column=16).value)
            self.assertEqual("✅ Đã đăng", ws.cell(row=5, column=14).value)
            self.assertEqual("dd_product_04", ws.cell(row=5, column=18).value)
            self.assertIsNone(ws.cell(row=4, column=2).value)
            self.assertIsNone(ws.cell(row=4, column=16).value)
            self.assertIsNone(ws.cell(row=4, column=18).value)
            self.assertIn("merged_into=product-04", str(ws.cell(row=4, column=17).value or ""))

    def test_merge_does_not_copy_duplicate_files_with_same_bytes_and_different_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            folder_a = shop_dir / "product-a" / "files"
            folder_b = shop_dir / "product-b" / "files"
            self._write_file(folder_a / "a.pdf", b"same-deliverable")
            self._write_file(folder_b / "b.pdf", b"same-deliverable")
            self._write_file(shop_dir / "product-a" / "images" / "cover-a.txt", b"same")
            self._write_file(shop_dir / "product-b" / "images" / "cover-b.txt", b"same")

            catalog = make_catalog("group-allow-same-hash-different-name", ["product-a", "product-b"])
            catalog["duplicate_groups"][0]["records"][0]["total_bytes"] = 5000
            catalog["duplicate_groups"][0]["records"][1]["total_bytes"] = 100
            result = self._run_merge(base_dir, catalog, "group-allow-same-hash-different-name")

            self.assertEqual(1, len(result.get("merged", [])))
            self.assertEqual([], result.get("skipped", []))
            self.assertEqual("product-a", result["merged"][0]["canonical_folder"])
            self.assertEqual(["product-b"], result["merged"][0]["moved_folders"])

            merged_folder = shop_dir / "product-a" / "files"
            deliverables = sorted(path.name for path in merged_folder.iterdir())
            self.assertEqual(["a.pdf"], deliverables)
            self.assertFalse((shop_dir / "product-b").exists())
            self.assertTrue((shop_dir / "product-a" / "images" / "cover-b.txt").exists())
            manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
            self.assertTrue((Path(manifest["quarantine"]) / "product-b").is_dir())

    def test_merge_guard_allows_matching_digital_sets(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            folder_a = shop_dir / "product-a" / "files"
            folder_b = shop_dir / "product-b" / "files"
            self._write_file(folder_a / "a.pdf", b"shared")
            self._write_file(folder_b / "b.pdf", b"shared")

            catalog = make_catalog("group-guard-allow", ["product-a", "product-b"])
            result = self._run_merge(base_dir, catalog, "group-guard-allow")

            self.assertEqual(1, len(result.get("merged", [])))
            self.assertEqual([], result.get("skipped", []))
            self.assertFalse((shop_dir / "product-a").is_dir())
            self.assertTrue((shop_dir / "product-b").is_dir())
            self.assertEqual("product-a", result["merged"][0]["moved_folders"][0])

    def test_merge_guard_blocks_merge_when_extra_unrelated_zip_exists(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            folder_a = shop_dir / "product-a" / "files"
            folder_b = shop_dir / "product-b" / "files"
            self._write_file(folder_a / "shared.pdf", b"same-content")
            self._write_file(folder_b / "shared.pdf", b"same-content")
            self._write_file(folder_a / "extra.zip", b"unrelated-zip")

            catalog = make_catalog("group-guard-extra-zip", ["product-a", "product-b"])
            result = self._run_merge(base_dir, catalog, "group-guard-extra-zip")

            self.assertEqual([], result["merged"])
            self.assertEqual(1, len(result.get("skipped", [])))
            self.assertIn("hash set mismatch", result["skipped"][0]["reason"])
            self.assertTrue((shop_dir / "product-a").is_dir())
            self.assertTrue((shop_dir / "product-b").is_dir())

    def test_merge_guard_blocks_merge_when_hash_raises_oserror(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            folder_a = shop_dir / "product-a" / "files"
            folder_b = shop_dir / "product-b" / "files"
            self._write_file(folder_a / "ok.pdf", b"stable")
            self._write_file(folder_b / "bad.pdf", b"fail")

            def failing_hash(path: Path) -> str:
                if path.name == "bad.pdf":
                    raise OSError("read fail")
                return hashlib.sha256(path.read_bytes()).hexdigest()

            catalog = make_catalog("group-guard-oserror", ["product-a", "product-b"])
            result = self._run_merge(
                base_dir,
                catalog,
                "group-guard-oserror",
                safe_hash_side_effect=failing_hash,
            )

            self.assertEqual([], result["merged"])
            self.assertEqual(1, len(result.get("skipped", [])))
            skipped = result["skipped"][0]
            failed = sorted((p for paths in skipped["failed_relative_paths"].values() for p in paths))
            self.assertIn("files/bad.pdf", failed)

    def test_merge_guard_blocks_same_filename_different_bytes_same_size(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            folder_a = shop_dir / "product-a" / "files"
            folder_b = shop_dir / "product-b" / "files"
            self._write_file(folder_a / "shared.pdf", b"AAAAB")
            self._write_file(folder_b / "shared.pdf", b"AAABC")

            catalog = make_catalog("group-guard-bytes", ["product-a", "product-b"])
            result = self._run_merge(base_dir, catalog, "group-guard-bytes")

            self.assertEqual([], result["merged"])
            self.assertEqual(1, len(result.get("skipped", [])))
            self.assertIn("hash set mismatch", result["skipped"][0]["reason"])
            self.assertTrue((shop_dir / "product-a").is_dir())
            self.assertTrue((shop_dir / "product-b").is_dir())

    def test_empty_group_filter_merges_nothing(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            self._write_file(shop_dir / "product-a" / "files" / "a.pdf", b"shared")
            self._write_file(shop_dir / "product-b" / "files" / "b.pdf", b"shared")
            catalog = make_catalog("group-empty-filter", ["product-a", "product-b"])
            excel_path = shop_dir / "Etsy_SEO_Generator.xlsx"
            self._make_workbook(excel_path)
            with patch("etsy_catalog.build_unified_catalog", return_value=catalog):
                result = merge_safe_duplicates(base_dir, "daisyflowdigital", excel_path, [])
            self.assertEqual([], result["merged"])
            self.assertTrue((shop_dir / "product-a").is_dir())
            self.assertTrue((shop_dir / "product-b").is_dir())

    def test_source_source_collision_is_preflighted_without_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            for folder in ("product-a", "product-b", "product-c"):
                self._write_file(shop_dir / folder / "files" / "shared.pdf", b"same-deliverable")
            self._write_file(shop_dir / "product-a" / "images" / "cover.jpg", b"image-a")
            self._write_file(shop_dir / "product-b" / "images" / "cover.jpg", b"image-b")
            catalog = make_catalog("group-source-collision", ["product-a", "product-b", "product-c"])
            catalog["duplicate_groups"][0]["records"][2]["listing_id"] = "123"
            result = self._run_merge(base_dir, catalog, "group-source-collision")
            self.assertEqual([], result["merged"])
            self.assertEqual(1, len(result["skipped"]))
            self.assertIn("collision", result["skipped"][0]["reason"])
            for folder in ("product-a", "product-b", "product-c"):
                self.assertTrue((shop_dir / folder).is_dir())

    def test_all_deliverable_extensions_participate_in_full_match(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base_dir = Path(tmpdir)
            shop_dir = base_dir / "shops" / "daisyflowdigital"
            self._write_file(shop_dir / "product-a" / "files" / "shared.pdf", b"same")
            self._write_file(shop_dir / "product-b" / "files" / "shared.pdf", b"same")
            self._write_file(shop_dir / "product-a" / "files" / "extra.svg", b"svg-deliverable")
            catalog = make_catalog("group-extra-svg", ["product-a", "product-b"])
            result = self._run_merge(base_dir, catalog, "group-extra-svg")
            self.assertEqual([], result["merged"])
            self.assertEqual(1, len(result["skipped"]))
            self.assertTrue((shop_dir / "product-a").is_dir())
            self.assertTrue((shop_dir / "product-b").is_dir())


if __name__ == "__main__":
    unittest.main()
