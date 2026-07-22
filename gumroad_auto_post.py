"""
Gumroad Auto Draft Poster
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Dùng Chrome thật (tránh bot detection)
• Đọc data từ sheet "Gumroad" trong Etsy_SEO_Generator.xlsx
• Upload: ảnh cover + preview gallery + file PDF
• Điền: title, description, price, tags, permalink, summary
• Lưu draft (unpublished) — anh publish tay sau khi review
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Cách dùng:
  python3 gumroad_auto_post.py              # 5 sản phẩm đầu
  python3 gumroad_auto_post.py --batch 3    # 3 sản phẩm
  python3 gumroad_auto_post.py --batch 5 --skip 5   # bỏ qua 5 SP đầu

Bước đầu tiên: Chạy python3 gumroad_setup_excel.py để tạo sheet "Gumroad"
"""

import asyncio, sys, subprocess, argparse, re
from pathlib import Path

# ── Auto-install deps ──────────────────────────────────────────────────────────
def ensure_deps():
    pkgs = {"openpyxl": "openpyxl", "playwright": "playwright"}
    for mod, pkg in pkgs.items():
        try:
            __import__(mod)
        except ImportError:
            print(f"▶ Cài {pkg}...")
            subprocess.run([sys.executable, "-m", "pip", "install", pkg, "--quiet"], check=True)

ensure_deps()

import openpyxl
from playwright.async_api import async_playwright

BASE_DIR    = Path(__file__).parent
EXCEL_FILE  = BASE_DIR / "Etsy_SEO_Generator.xlsx"
BROWSER_DIR = BASE_DIR / ".browser-session"
CHROME_PATH = Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")
GUMROAD_URL = "https://app.gumroad.com"

DEFAULT_BATCH = 5

# ── Helpers: Text ──────────────────────────────────────────────────────────────
def trim_title(title: str, max_len: int = 100) -> str:
    if len(title) <= max_len:
        return title
    cut = title[:max_len].rsplit(" ", 1)[0].rstrip(",|;- ")
    return cut

def make_permalink(folder: str, keywords: str) -> str:
    """Tạo URL slug từ keyword đầu tiên, fallback về folder name."""
    kw = str(keywords or "").strip()
    if kw:
        first = kw.split(",")[0].strip().lower()
        slug  = re.sub(r"[^a-z0-9\-]", "-", first)
        slug  = re.sub(r"-+", "-", slug).strip("-")
        if slug:
            return slug[:60]
    return folder.lower().replace("_", "-")

# ── Read Excel ─────────────────────────────────────────────────────────────────
def read_products(shop_id: str, batch: int = DEFAULT_BATCH, skip: int = 0):
    excel_file = BASE_DIR / "shops" / shop_id / "Etsy_SEO_Generator.xlsx"
    if not excel_file.exists():
        print(f"❌ Không tìm thấy Excel tại: {excel_file}")
        return [], None, None, 0
        
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    if "Gumroad" not in wb.sheetnames:
        print(f"\n⚠  Sheet 'Gumroad' chưa có trong Excel của shop '{shop_id}'!")
        print(f"   Chạy trước: python3 gumroad_setup_excel.py --shop {shop_id}\n")
        return [], wb, None, 0

    ws = wb["Gumroad"]
    all_products = []

    for row_num, row in enumerate(ws.iter_rows(min_row=4, max_row=100, values_only=True), start=4):
        cols = (list(row) + [None] * 16)[:16]
        # Col: A=stt B=folder C=keywords D=notes E=price F=tags G=_ H=title I=desc J=tags_text K=summary L=permalink M=_ N=status O=_ P=_
        _, folder, keywords, _, price, _, _, title, description, tags_text, summary, permalink_col, _, status, _, _ = cols

        if not folder or not title or str(title).startswith("←"):
            continue
        if status and "Đã đăng" in str(status):
            continue

        img_dir  = BASE_DIR / "shops" / shop_id / str(folder) / "images"
        file_dir = BASE_DIR / "shops" / shop_id / str(folder) / "files"
        img_exts = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

        img_paths = (
            sorted([str(f) for f in img_dir.iterdir() if f.suffix.lower() in img_exts])[:10]
            if img_dir.exists() else []
        )
        pdf_paths = (
            sorted([str(f) for f in file_dir.glob("*.pdf")])
            if file_dir.exists() else []
        )

        cover_path    = img_paths[0] if img_paths else None
        preview_paths = img_paths[1:] if len(img_paths) > 1 else []

        kw_str    = str(keywords or "")
        permalink = str(permalink_col or "").strip() or make_permalink(str(folder), kw_str)

        all_products.append({
            "folder":        str(folder),
            "title":         trim_title(str(title)),
            "description":   str(description or ""),
            "price":         float(price) if price else 0.0,
            "tags":          str(tags_text or ""),
            "summary":       str(summary or "")[:250],
            "permalink":     permalink,
            "cover_path":    cover_path,
            "preview_paths": preview_paths,
            "pdf_paths":     pdf_paths,
            "row":           row_num,
        })

    return all_products[skip: skip + batch], wb, ws, len(all_products)

def save_status(shop_id: str, wb, ws, row: int, text: str):
    excel_file = BASE_DIR / "shops" / shop_id / "Etsy_SEO_Generator.xlsx"
    ws.cell(row=row, column=14, value=text)
    wb.save(excel_file)

# ── Playwright helpers ─────────────────────────────────────────────────────────
async def smart_fill(page, selector: str, value: str, timeout: int = 6000) -> bool:
    try:
        el = page.locator(selector).first
        await el.wait_for(state="visible", timeout=timeout)
        await el.scroll_into_view_if_needed()
        await page.wait_for_timeout(300)
        await el.click(click_count=3)
        await page.wait_for_timeout(200)
        await el.fill(str(value))
        await page.wait_for_timeout(300)
        return True
    except Exception:
        return False

async def fill_rich_text(page, text: str) -> bool:
    """
    Điền vào rich-text editor (Tiptap / ProseMirror / Trix).
    """
    try:
        # Tiptap / ProseMirror editor
        el = page.locator('.ProseMirror').first
        if await el.count() > 0:
            # Chuyển plain text thành HTML paragraphs và lọc bỏ khoảng trống dư thừa
            lines = [l.strip() for l in text.split("\n")]
            cleaned_lines = []
            for line in lines:
                if line:
                    cleaned_lines.append(f"<p>{line}</p>")
                else:
                    if not cleaned_lines or cleaned_lines[-1] != "<p></p>":
                        cleaned_lines.append("<p></p>")
            html_content = "".join(cleaned_lines)
            
            success = await page.evaluate('''(html) => {
                const el = document.querySelector('.ProseMirror');
                if (el) {
                    if (el.editor) {
                        el.editor.commands.setContent(html);
                        return true;
                    }
                    el.innerHTML = html;
                    el.dispatchEvent(new Event('input', { bubbles: true }));
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                    return true;
                }
                return false;
            }''', html_content)

            if success:
                return True
    except Exception:
        pass

    try:
        # Thử Trix editor
        trix = page.locator("trix-editor").first
        if await trix.count() > 0 and await trix.is_visible():
            await trix.click()
            await page.wait_for_timeout(400)
            await page.evaluate(
                """(text) => {
                    const ed = document.querySelector('trix-editor');
                    if (ed && ed.editor) {
                        ed.editor.loadHTML('');
                        ed.editor.insertString(text);
                    }
                }""",
                text
            )
            await page.wait_for_timeout(500)
            return True
    except Exception:
        pass

    # Fallback: contenteditable div
    try:
        ce = page.locator('[contenteditable="true"]').first
        if await ce.count() > 0 and await ce.is_visible():
            await ce.click()
            await page.wait_for_timeout(300)
            await page.keyboard.press("Meta+a")
            await page.keyboard.press("Delete")
            await ce.type(text[:5000], delay=5)
            await page.wait_for_timeout(500)
            return True
    except Exception:
        pass

    return False

async def fill_by_label(page, label_text: str, value: str) -> bool:
    """
    Điền vào input dựa trên tên label một cách chính xác (tránh dynamic ID).
    """
    try:
        label_loc = page.locator(f'label:has-text("{label_text}")').first
        if await label_loc.count() > 0:
            for_attr = await label_loc.get_attribute("for")
            if for_attr:
                success = await page.evaluate("""(id, val) => {
                    const el = document.getElementById(id);
                    if (el) {
                        el.value = val;
                        el.dispatchEvent(new Event('input', { bubbles: true }));
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                        return true;
                    }
                    return false;
                }""", for_attr, str(value))
                if success:
                    return True
    except Exception:
        pass
    return False


async def try_file_chooser(page, btn_selector: str, files) -> bool:
    """Bấm nút → bắt file chooser dialog → set files."""
    btn = page.locator(btn_selector).first
    if await btn.count() == 0 or not await btn.is_visible():
        return False
    try:
        async with page.expect_file_chooser(timeout=8000) as fc_info:
            await btn.click()
        fc = await fc_info.value
        await fc.set_files(files)
        return True
    except Exception:
        return False

# ── Upload: Cover ──────────────────────────────────────────────────────────────
async def upload_cover(page, cover_path):
    if not cover_path:
        return
    try:
        # Check if the "Computer files" label/tab is already visible
        comp_label = page.locator('label:has-text("Computer files"), label:has-text("Computer")').first
        if await comp_label.count() == 0 or not await comp_label.is_visible():
            # Otherwise click the main "Upload images or videos" button first
            upload_btn = page.locator('button:has-text("Upload images or videos")').first
            if await upload_btn.count() > 0 and await upload_btn.is_visible():
                await upload_btn.click()
                await page.wait_for_timeout(1000)
        
        comp_label = page.locator('label:has-text("Computer files"), label:has-text("Computer")').first
        if await comp_label.count() > 0:
            async with page.expect_file_chooser(timeout=15000) as fc_info:
                await comp_label.click()
            fc = await fc_info.value
            await fc.set_files(cover_path)
            await page.wait_for_timeout(3000)
            print(f"  🖼  Cover ✓")
            return
            
        # Fallback: direct input
        fi = page.locator('input[type="file"]').first
        if await fi.count() > 0:
            await fi.set_input_files(cover_path, timeout=15000)
            await page.wait_for_timeout(3000)
            print(f"  🖼  Cover (fallback) ✓")
            return

        print("  ⚠ Cover: không tìm được upload area — bỏ qua")
    except Exception as e:
        print(f"  ⚠ Cover: {e}")

# ── Upload: Thumbnail ──────────────────────────────────────────────────────────
async def upload_thumbnail(page, thumbnail_path):
    if not thumbnail_path:
        return
    try:
        # Try direct single file input first (Thumbnail does not have multiple attribute)
        thumb_input = page.locator('input[type="file"]:not([multiple])').first
        if await thumb_input.count() > 0:
            await thumb_input.set_input_files(thumbnail_path)
            await page.wait_for_timeout(3000)
            print(f"  🖼  Thumbnail ✓")
            return
            
        # Try label upload under Thumbnail heading
        thumb_label = page.locator('label:has-text("Upload")').first
        if await thumb_label.count() > 0:
            async with page.expect_file_chooser(timeout=15000) as fc_info:
                await thumb_label.click()
            fc = await fc_info.value
            await fc.set_files(thumbnail_path)
            await page.wait_for_timeout(3000)
            print(f"  🖼  Thumbnail ✓")
            return
            
        print("  ⚠ Thumbnail: không tìm được upload area — bỏ qua")
    except Exception as e:
        print(f"  ⚠ Thumbnail: {e}")

# ── Upload: Preview gallery ────────────────────────────────────────────────────
async def upload_previews(page, preview_paths: list):
    if not preview_paths:
        return
    try:
        for btn_sel in [
            'button:has-text("Add preview")',
            'button:has-text("Upload preview")',
            'button:has-text("Add images")',
            'button:has-text("Add more")',
            '[data-testid*="preview"] button',
            '[class*="preview"] button',
        ]:
            if await try_file_chooser(page, btn_sel, preview_paths):
                await page.wait_for_timeout(4000)
                print(f"  🖼  {len(preview_paths)} preview ✓")
                return

        # Fallback: input[type=file] thứ hai (sau cover)
        all_fi = page.locator('input[type="file"]')
        if await all_fi.count() > 1:
            try:
                inp = all_fi.nth(1)
                await inp.set_input_files(preview_paths, timeout=15000)
                await page.wait_for_timeout(3000)
                print(f"  🖼  {len(preview_paths)} preview (fallback) ✓")
                return
            except Exception:
                pass

        print(f"  ⚠ Preview: không tìm được upload area — bỏ qua")
    except Exception as e:
        print(f"  ⚠ Preview: {e}")

# ── Upload: PDF content files ──────────────────────────────────────────────────
async def upload_files(page, pdf_paths: list):
    if not pdf_paths:
        return
    try:
        # Try direct file input on the Content page
        file_input = page.locator('input[type="file"][name="file"]').first
        if await file_input.count() > 0:
            await file_input.set_input_files(pdf_paths)
            print(f"  ⏳ Đợi upload PDF ({Path(pdf_paths[0]).name})...")
            # Wait up to 30 seconds for the file to be uploaded
            pdf_name = Path(pdf_paths[0]).name
            for _ in range(30):
                await page.wait_for_timeout(1000)
                body_text = await page.inner_text("body")
                if pdf_name in body_text or pdf_name.replace(".pdf", "") in body_text:
                    break
            print(f"  📎 {len(pdf_paths)} PDF ✓")
            return

        # Fallback chooser
        for btn_sel in [
            'button:has-text("Upload your files")',
            'button:has-text("Upload files")',
            'button:has-text("Add files")',
        ]:
            if await try_file_chooser(page, btn_sel, pdf_paths):
                await page.wait_for_timeout(8000)
                print(f"  📎 {len(pdf_paths)} PDF (fallback) ✓")
                return

        print("  ⚠ PDF: không tìm được upload area — anh upload thủ công nhé!")
    except Exception as e:
        print(f"  ⚠ PDF: {e}")

# ── Fill one listing ───────────────────────────────────────────────────────────
async def fill_product(page, product: dict) -> bool:
    print(f"\n{'─'*55}")
    print(f"  📦 {product['folder']} | cover: {'✓' if product['cover_path'] else '✗'} | {len(product['pdf_paths'])} PDF")
    print(f"     {product['title'][:65]}...")

    # ── Bước 1: Tạo sản phẩm mới ────────────────────────────────────────────
    await page.goto(f"{GUMROAD_URL}/products/new", wait_until="domcontentloaded")
    await page.wait_for_timeout(4000)

    if "sign_in" in page.url or "login" in page.url:
        print("  ❌ Cần đăng nhập Gumroad — xem hướng dẫn bên dưới")
        return False

    # Gumroad mới: modal/form name + price rồi mới redirect sang edit page
    name_filled = False
    for sel in [
        'input[name="name"]',
        'input[placeholder*="name" i]',
        'input[placeholder*="product name" i]',
        'input[id*="name" i]',
    ]:
        if await smart_fill(page, sel, product["title"], timeout=5000):
            print(f"  📝 Title ✓")
            name_filled = True
            break

    if not name_filled:
        # Thử input[type=text] đầu tiên
        if await smart_fill(page, 'input[type="text"]', product["title"], timeout=4000):
            print(f"  📝 Title (fallback) ✓")
            name_filled = True

    # ── Price ────────────────────────────────────────────────────────────────
    price_filled = False
    for sel in [
        'input[name="price"]',
        'input[placeholder*="price" i]',
        'input[id*="price" i]',
        'input[type="number"]',
    ]:
        if await smart_fill(page, sel, f"{product['price']:.2f}", timeout=4000):
            print(f"  💲 ${product['price']:.2f} ✓")
            price_filled = True
            break

    await page.wait_for_timeout(500)

    # Submit bước đầu (wizard "Create product")
    for btn_sel in [
        'button:has-text("Next")',
        'button:has-text("Create product")',
        'button:has-text("Create")',
        'form button[type="submit"]',
    ]:
        btn = page.locator(btn_sel).first
        if await btn.count() > 0 and await btn.is_visible():
            await btn.click()
            await page.wait_for_timeout(4000)
            print(f"  ➡️  Tạo sản phẩm ✓")
            break

    # Chờ redirect sang trang edit
    await page.wait_for_timeout(3000)
    current_url = page.url
    print(f"  🔗 URL hiện tại: {current_url}")

    # Chuyển hướng trực tiếp sang Product Tab (/edit)
    edit_base_url = current_url.split("/edit")[0] + "/edit"
    print(f"  🔗 Chuyển sang Product Tab: {edit_base_url}")
    await page.goto(edit_base_url, wait_until="domcontentloaded")
    await page.wait_for_timeout(4000)

    # ── Bước 2: Điền chi tiết trên trang edit (Product Tab) ────────────────
    # Title / Name
    await fill_by_label(page, "Name", product["title"])

    # Description
    if product["description"]:
        await page.wait_for_timeout(500)
        if await fill_rich_text(page, product["description"]):
            print(f"  📄 Description ✓")
        else:
            for sel in ['textarea[name*="description"]', 'textarea[id*="description"]']:
                if await smart_fill(page, sel, product["description"], timeout=4000):
                    print(f"  📄 Description ✓")
                    break

    # Permalink / Custom Path
    if product["permalink"]:
        await page.wait_for_timeout(500)
        if await fill_by_label(page, "URL", product["permalink"]):
            print(f"  🔗 Permalink: /{product['permalink']} ✓")

    # Summary
    if product["summary"]:
        await page.wait_for_timeout(500)
        if await fill_by_label(page, "Summary", product["summary"]):
            print(f"  📋 Summary ✓")

    # Price / Amount
    await fill_by_label(page, "Amount", f"{product['price']:.2f}")
    await page.wait_for_timeout(500)

    # Tags
    if product["tags"]:
        try:
            tag_list = [t.strip() for t in product["tags"].split(",") if t.strip()][:10]
            for sel in [
                'input[name="tags"]',
                'input[placeholder*="tag" i]',
                'input[id*="tags" i]',
                'input[placeholder*="categories" i]',
            ]:
                tag_input = page.locator(sel).first
                if await tag_input.count() > 0 and await tag_input.is_visible():
                    filled = 0
                    for tag in tag_list:
                        await tag_input.fill(tag)
                        await page.wait_for_timeout(300)
                        await tag_input.press("Enter")
                        await page.wait_for_timeout(400)
                        filled += 1
                    if filled:
                        print(f"  🏷  {filled} tags ✓")
                    break
        except Exception as e:
            print(f"  ⚠ Tags: {e}")
    await page.wait_for_timeout(500)

    # Upload Cover & Thumbnail
    if product["cover_path"]:
        await upload_cover(page, product["cover_path"])
        await page.wait_for_timeout(1000)
        await upload_thumbnail(page, product["cover_path"])
        await page.wait_for_timeout(1000)

    # Lưu thay đổi của Product Tab
    print("  💾 Lưu Product Tab...")
    save_btn = page.locator('button:has-text("Save changes"), button:has-text("Save")').first
    if await save_btn.count() > 0 and await save_btn.is_visible():
        await save_btn.click()
        await page.wait_for_timeout(3000)

    # ── Bước 3: Chuyển sang Content Tab và upload PDF ───────────────────────
    content_url = edit_base_url + "/content"
    print(f"  🔗 Chuyển sang Content Tab: {content_url}")
    await page.goto(content_url, wait_until="domcontentloaded")
    await page.wait_for_timeout(4000)

    # Upload files
    if product["pdf_paths"]:
        await upload_files(page, product["pdf_paths"])
        await page.wait_for_timeout(2000)

    # ── Bước 4: Lưu Product Draft cuối cùng ──────────────────────────────────
    print("  💾 Lưu Content Tab và kết thúc...")
    saved = False
    for btn_sel in [
        'button:has-text("Save changes")',
        'button:has-text("Save")',
    ]:
        btn = page.locator(btn_sel).first
        if await btn.count() > 0 and await btn.is_visible():
            btn_text = (await btn.inner_text()).strip().lower()
            if "publish" in btn_text or "enable" in btn_text:
                continue
            await btn.click()
            await page.wait_for_timeout(3500)
            print("  💾 Saved draft successfully! ✅")
            saved = True
            break

    if not saved:
        try:
            await page.keyboard.press("Control+s")
            await page.wait_for_timeout(2500)
            print("  💾 Saved (Ctrl+S) ✅")
            saved = True
        except Exception:
            pass

    if not saved:
        print("  ⚠ Không tìm thấy nút Save — anh lưu thủ công trên trình duyệt nhé!")
        return False

    return True

# ── Main ───────────────────────────────────────────────────────────────────────
async def main():
    parser = argparse.ArgumentParser(description="Gumroad Auto Draft Poster")
    parser.add_argument("--shop", type=str, required=True, help="ID của shop hiện tại (ví dụ: templystudios)")
    parser.add_argument("--batch", type=int, default=DEFAULT_BATCH,
                        help=f"Số sản phẩm mỗi lần chạy (mặc định {DEFAULT_BATCH})")
    parser.add_argument("--skip", type=int, default=0,
                        help="Bỏ qua N sản phẩm đầu trong danh sách chờ")
    args = parser.parse_args()

    products, wb, ws, total = read_products(shop_id=args.shop, batch=args.batch, skip=args.skip)

    if not products or ws is None:
        print("\n⚠  Không có sản phẩm nào cần đăng.\n")
        return

    print(f"\n{'='*55}")
    print(f"  🟢  Gumroad Auto Draft Poster")
    print(f"  📦 Shop: {args.shop} | {total} tổng | Batch: {args.batch} | Skip: {args.skip}")
    print(f"{'='*55}")
    for p in products:
        cover = "🖼" if p["cover_path"] else "—"
        print(f"   • {p['folder']} {cover} | {len(p['pdf_paths'])} PDF | {p['title'][:45]}...")
    print()

    # Dọn Chrome lock files
    BROWSER_DIR.mkdir(exist_ok=True)
    for lf in ["SingletonLock", "SingletonCookie", "SingletonSocket"]:
        try:
            (BROWSER_DIR / lf).unlink(missing_ok=True)
        except Exception:
            pass

    async with async_playwright() as pw:
        browser = None
        ctx = None
        
        try:
            print("⏳ Đang thử kết nối tới Chrome đang mở (cổng 9222)...")
            browser = await pw.chromium.connect_over_cdp("http://127.0.0.1:9222")
            ctx = browser.contexts[0]
            print("✅ Đã kết nối thành công tới Chrome đang mở!")
        except Exception:
            print("ℹ️ Chrome debug cổng 9222 không mở. Đang khởi chạy Chrome session mới...")
            launch_kw = dict(
                user_data_dir=str(BROWSER_DIR),
                headless=False,
                args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
                viewport=None,
            )
            if CHROME_PATH.exists():
                launch_kw["executable_path"] = str(CHROME_PATH)
                print("  🌐 Dùng Google Chrome thật")
            else:
                print("  🌐 Dùng Chromium")
            ctx = await pw.chromium.launch_persistent_context(**launch_kw)

        page = ctx.pages[0] if ctx.pages else await ctx.new_page()

        # ── Kiểm tra đăng nhập ───────────────────────────────────────────────
        await page.goto(f"{GUMROAD_URL}/dashboard", wait_until="domcontentloaded")
        await page.wait_for_timeout(4000)

        if "sign_in" in page.url or "login" in page.url:
            print("\n  ⚠  Chưa đăng nhập Gumroad!")
            print("     → Đăng nhập trên trình duyệt vừa mở, rồi nhấn Enter tại đây.")
            input()
            await page.goto(f"{GUMROAD_URL}/dashboard", wait_until="domcontentloaded")
            await page.wait_for_timeout(4000)

        print("  ✅ Đã vào Gumroad Dashboard!\n")

        success = failed = 0
        for i, product in enumerate(products, 1):
            print(f"\n[{i}/{len(products)}]", end="")
            try:
                ok = await fill_product(page, product)
                if ok:
                    success += 1
                    save_status(args.shop, wb, ws, product["row"], "✅ Đã đăng draft")
                else:
                    failed += 1
                    save_status(args.shop, wb, ws, product["row"], "❌ Lỗi")
            except Exception as e:
                print(f"\n  ❌ {e}")
                failed += 1
                save_status(args.shop, wb, ws, product["row"], f"❌ {str(e)[:40]}")

            if i < len(products):
                wait_sec = 6
                print(f"\n  ⏳ Nghỉ {wait_sec}s trước sản phẩm tiếp theo...")
                await asyncio.sleep(wait_sec)

        # ── Tổng kết ─────────────────────────────────────────────────────────
        print(f"\n{'='*55}")
        print(f"  ✅ Thành công : {success}/{len(products)}")
        print(f"  ❌ Thất bại  : {failed}/{len(products)}")
        remaining = total - args.skip - len(products)
        if remaining > 0:
            next_skip = args.skip + len(products)
            print(f"\n  📌 Còn {remaining} sản phẩm. Lần sau chạy:")
            print(f"     python3 gumroad_auto_post.py --shop {args.shop} --batch {args.batch} --skip {next_skip}")
        print(f"{'='*55}\n")
        await ctx.close()

if __name__ == "__main__":
    asyncio.run(main())
