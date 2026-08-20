import copy
import fcntl
import threading
import hashlib
import json
import re
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from cloud_asset_store import CloudAssetError, LocalRemote, ProductLock, canonical_json_bytes
from scripts.daisy_folder_renumber import (
    MigrationError,
    HeldLocks,
    _is_catalog_writer_command,
    apply_plan,
    build_plan,
    configured_remote,
    make_mapping,
    prepare_cloud_move,
    product_name,
)


class DaisyFolderRenumberTests(unittest.TestCase):
    def test_offline_guard_recognizes_all_known_catalog_writers(self):
        commands = (
            "/opt/homebrew/bin/rclone copy source destination",
            "rclone copy source destination",
            "/opt/homebrew/bin/python3 /repo/dashboard_app.py",
            "python3 dashboard_app.py",
            "/opt/homebrew/bin/python3 /repo/etsy_shop_sync.py --shop daisyflowdigital",
            "python3 cloud_asset_cli.py status",
            "/opt/homebrew/bin/python3 /repo/backup_etsy_to_drive.py",
            "python3 bulk_create_unmapped_listings.py --shop daisyflowdigital",
            "/opt/homebrew/bin/python3 /repo/cloud_asset_cli.py status",
            "/opt/homebrew/bin/python3 /repo/image_factory_watcher.py",
            "/opt/homebrew/bin/python3 /repo/social_auto_post.py",
            "/opt/homebrew/bin/python3 /repo/gumroad_auto_post.py",
        )
        for command in commands:
            with self.subTest(command=command):
                self.assertTrue(_is_catalog_writer_command(command))
        self.assertFalse(_is_catalog_writer_command("node /repo/test_dashboard_stats_ui.js"))

    def test_lock_contender_cannot_unlink_an_existing_migration_lock(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary))
            plan = self._plan(fixture)
            lock_path = fixture["shop"] / ".daisy-folder-renumber.lock"
            with lock_path.open("x+") as owner:
                fcntl.flock(owner.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                with self.assertRaises(FileExistsError):
                    with HeldLocks(plan):
                        self.fail("contender unexpectedly acquired the migration lock")
                self.assertTrue(lock_path.exists())
                with self.assertRaises(FileExistsError):
                    with HeldLocks(plan):
                        self.fail("third contender unexpectedly acquired the migration lock")
                fcntl.flock(owner.fileno(), fcntl.LOCK_UN)

    def test_all_product_locks_are_created_and_held_until_migration_exits(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary))
            plan = self._plan(fixture)
            missing_lock = fixture["shop"] / "product-08" / ".cloud-assets.lock"
            missing_lock.unlink()
            with HeldLocks(plan):
                self.assertTrue(missing_lock.is_file())
                result = []

                def compete():
                    try:
                        with ProductLock(fixture["shop"] / "product-08", timeout_seconds=0.05):
                            result.append("acquired")
                    except CloudAssetError:
                        result.append("blocked")

                contender = threading.Thread(target=compete)
                contender.start()
                contender.join(timeout=1)
                self.assertEqual(result, ["blocked"])
            with ProductLock(fixture["shop"] / "product-08", timeout_seconds=0.05):
                pass

    def test_mapping_sorts_physical_folders_and_is_contiguous(self):
        mappings = make_mapping(("product-01", "product-07", "product-428"), token="unit")
        self.assertEqual(
            [(item.old_name, item.new_name) for item in mappings],
            [("product-01", "product-01"), ("product-07", "product-02"), ("product-428", "product-03")],
        )
        self.assertEqual(product_name(99), "product-99")
        self.assertEqual(product_name(100), "product-100")
        self.assertTrue(mappings[1].temporary_name.startswith(".daisy-folder-renumber-unit-"))

    def test_dry_run_plan_is_immutable_and_cloud_only_folder_is_valid(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary), cloud=True, cloud_only=True)
            before = self._tree_hashes(fixture["root"])
            plan = self._plan(fixture)
            after = self._tree_hashes(fixture["root"])
            self.assertEqual(before, after)
            self.assertEqual(len(plan.mappings), 3)
            self.assertEqual(plan.unregistered_old, ("product-08",))
            self.assertEqual([(m.old_name, m.new_name) for m in plan.cloud_moves], [("product-08", "product-03")])

    def test_workbook_only_column_b_changes_and_stale_reference_is_cleared(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary))
            plan = self._plan(fixture)
            original = load_workbook(fixture["workbook"], data_only=False)
            backup_values = {
                (row, column): original["Listings"].cell(row, column).value
                for row in range(1, original["Listings"].max_row + 1)
                for column in range(1, original["Listings"].max_column + 1)
                if column != 2
            }
            original.close()
            backup, counts = apply_plan(
                plan, fixture["root"] / "backup", remote=fixture["remote"], runtime_checker=lambda _: None
            )
            self.assertTrue((backup / "mapping_manifest.json").is_file())
            self.assertEqual(counts["workbook_bindings"], 2)
            workbook = load_workbook(fixture["workbook"], data_only=False)
            sheet = workbook["Listings"]
            self.assertEqual(sheet.cell(4, 2).value, "product-01")
            self.assertEqual(sheet.cell(5, 2).value, "product-02")
            self.assertIsNone(sheet.cell(6, 2).value)
            for key, value in backup_values.items():
                self.assertEqual(sheet.cell(*key).value, value)
            self.assertEqual(sheet.cell(4, 18).value, "remote_sku_keep_01")
            self.assertEqual(sheet.cell(6, 18).value, "remote_sku_keep_stale")
            workbook.close()
            self.assertTrue((fixture["shop"] / "product-03").is_dir())
            self.assertFalse((fixture["shop"] / "product-08").exists())

    def test_semantically_empty_serialized_cells_do_not_false_positive(self):
        """Scanner-materialized and serialized empty B cells survive semantically."""
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary))
            self._inject_empty_numeric_cell(fixture["workbook"], "B2")
            plan = self._plan(fixture)
            _, counts = apply_plan(
                plan,
                fixture["root"] / "backup",
                remote=fixture["remote"],
                runtime_checker=lambda _: None,
            )
            self.assertEqual(counts["workbook_bindings"], 2)
            workbook = load_workbook(fixture["workbook"], data_only=False)
            self.assertIsNone(workbook["Listings"]["B2"].value)
            self.assertIsNone(workbook["Listings"]["B7"].value)
            self.assertEqual(workbook["Listings"]["C7"].value, "keep-sparse-row")
            workbook.close()

    def test_local_failure_rolls_back_folders_workbook_map_and_cloud_state(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary), cloud=True)
            plan = self._plan(fixture)
            before = self._tree_hashes(fixture["root"])
            old_pointer_path = "assets/v1/shops/daisyflowdigital/product-08/current.json"
            old_pointer = fixture["remote"].read_bytes(old_pointer_path)
            with patch(
                "scripts.daisy_folder_renumber._update_local_files",
                side_effect=RuntimeError("injected local failure"),
            ):
                with self.assertRaisesRegex(MigrationError, "automatic rollback completed"):
                    apply_plan(
                        plan,
                        fixture["root"] / "backup",
                        remote=fixture["remote"],
                        runtime_checker=lambda _: None,
                    )
            after = self._tree_hashes(fixture["root"], exclude_prefix="backup")
            after = {key: value for key, value in after.items() if not key.startswith("remote/")}
            expected = {
                key: value for key, value in before.items()
                if not key.startswith("backup/") and not key.startswith("remote/")
            }
            self.assertEqual(after, expected)
            self.assertEqual(fixture["remote"].read_bytes(old_pointer_path), old_pointer)
            self.assertTrue(fixture["remote"].path_exists(
                "assets/v1/shops/daisyflowdigital/product-03/current.json"
            ))
            self.assertTrue((fixture["shop"] / "product-08").is_dir())
            journals = list((fixture["root"] / "backup").glob("*/migration_journal.json"))
            self.assertEqual(json.loads(journals[0].read_text())["status"], "rolled_back")

    def test_successful_cloud_apply_passes_postflight_and_updates_local_identity(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary), cloud=True, cloud_only=True)
            plan = self._plan(fixture)
            _, counts = apply_plan(
                plan, fixture["root"] / "backup", remote=fixture["remote"], runtime_checker=lambda _: None
            )
            self.assertEqual(counts["cloud_moves"], 1)
            state = json.loads((fixture["shop"] / "product-03" / ".cloud-assets.json").read_text())
            self.assertEqual(state["product"]["product"], "product-03")
            self.assertEqual(state["current_manifest"]["product"]["key"],
                             "shops/daisyflowdigital/product-03")

    def test_postflight_rejects_missing_local_cloud_state_update(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary), cloud=True, cloud_only=True)
            plan = self._plan(fixture)

            def skip_cloud_state_updates(active_plan):
                workbook = load_workbook(active_plan.workbook_path, data_only=False)
                sheet = workbook["Listings"]
                for edit in active_plan.workbook_edits:
                    sheet.cell(edit.row, 2).value = edit.new_value
                workbook.save(active_plan.workbook_path)
                workbook.close()
                active_plan.source_map_path.write_text(
                    json.dumps(active_plan.updated_source_map, ensure_ascii=False, indent=2) + "\n",
                    encoding="utf-8",
                )

            with patch(
                "scripts.daisy_folder_renumber._update_local_files",
                side_effect=skip_cloud_state_updates,
            ):
                with self.assertRaisesRegex(MigrationError, "automatic rollback completed"):
                    apply_plan(
                        plan,
                        fixture["root"] / "backup",
                        remote=fixture["remote"],
                        runtime_checker=lambda _: None,
                    )
            self.assertTrue((fixture["shop"] / "product-08").is_dir())

    def test_cloud_migration_rewrites_identity_preserves_old_and_is_idempotent(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary), cloud=True, cloud_only=True)
            plan = self._plan(fixture)
            move = plan.cloud_moves[0]
            old_pointer = fixture["remote"].read_bytes(
                f"{move.old_identity.remote_prefix}/current.json"
            )
            first = prepare_cloud_move(fixture["remote"], move)
            self.assertEqual(first["remote"], "copied_verified")
            second = prepare_cloud_move(fixture["remote"], move)
            self.assertEqual(second["remote"], "reused_verified")
            self.assertEqual(
                fixture["remote"].read_bytes(f"{move.old_identity.remote_prefix}/current.json"),
                old_pointer,
            )
            pointer = json.loads(
                fixture["remote"].read_bytes(f"{move.new_identity.remote_prefix}/current.json")
            )
            manifest = json.loads(
                fixture["remote"].read_bytes(
                    f"{move.new_identity.remote_prefix}/revisions/{move.revision}/manifest.json"
                )
            )
            self.assertEqual(pointer["product"], "shops/daisyflowdigital/product-03")
            self.assertEqual(manifest["product"]["product"], "product-03")
            self.assertEqual(
                pointer["manifest_sha256"], hashlib.sha256(canonical_json_bytes(manifest)).hexdigest()
            )

    def test_conflicting_remote_retry_fails_closed(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary), cloud=True)
            move = self._plan(fixture).cloud_moves[0]
            prepare_cloud_move(fixture["remote"], move)
            fixture["remote"].write_bytes(
                f"{move.new_identity.remote_prefix}/current.json", b"{}\n", overwrite=True
            )
            with self.assertRaisesRegex(MigrationError, "conflicting new remote prefix"):
                prepare_cloud_move(fixture["remote"], move)

    def test_remote_retry_completes_verified_revision_missing_its_pointer(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary), cloud=True, cloud_only=True)
            move = self._plan(fixture).cloud_moves[0]
            prepare_cloud_move(fixture["remote"], move)
            current = (
                fixture["root"] / "remote" / move.new_identity.remote_prefix / "current.json"
            )
            current.unlink()
            result = prepare_cloud_move(fixture["remote"], move)
            self.assertEqual(result["remote"], "completed_partial_verified")
            pointer = json.loads(current.read_text(encoding="utf-8"))
            self.assertEqual(pointer["product"], move.new_identity.key)

    def test_repository_config_overrides_stale_remote_defaults(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            (root / "cloud_asset_store.config.json").write_text(
                json.dumps({
                    "remote": "fixture_remote",
                    "parent_id": "fixtureParent_1234567890",
                    "rclone_bin": "/fixture/rclone",
                }),
                encoding="utf-8",
            )
            remote = configured_remote(root)
            self.assertEqual(remote.remote, "fixture_remote")
            self.assertEqual(remote.parent_id, "fixtureParent_1234567890")
            self.assertEqual(remote.rclone_bin, "/fixture/rclone")

    def test_duplicate_workbook_binding_fails_closed(self):
        with tempfile.TemporaryDirectory() as temporary:
            fixture = self._fixture(Path(temporary))
            workbook = load_workbook(fixture["workbook"])
            workbook["Listings"].cell(6, 2).value = "product-01"
            workbook.save(fixture["workbook"])
            workbook.close()
            with self.assertRaisesRegex(MigrationError, "duplicate workbook folder reference"):
                self._plan(fixture)

    def _plan(self, fixture):
        return build_plan(
            repo_root=fixture["root"],
            expected_folders=3,
            expected_bindings=2,
            expected_stale=1,
            expected_stale_names=None,
            expected_source_map_updates=1,
            expected_cloud_moves=None,
            token="fixture",
        )

    def _fixture(self, root: Path, *, cloud: bool = False, cloud_only: bool = False):
        shop = root / "shops" / "daisyflowdigital"
        shop.mkdir(parents=True)
        (shop / ".catalog_workbook.lock").touch()
        for name in ("product-01", "product-03", "product-08"):
            folder = shop / name
            folder.mkdir()
            (folder / ".cloud-assets.lock").touch()
            if not (cloud_only and name == "product-08"):
                (folder / "asset.bin").write_bytes((name + "\n").encode())
            (folder / "zero-byte-placeholder.bin").touch()

        workbook_path = shop / "Etsy_SEO_Generator.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Listings"
        sheet.append(["banner"])
        sheet.append(["instructions"])
        sheet.append(["STT", "Folder"] + [None] * 15 + ["SKU"])
        sheet.append([1, "product-01", "keep-a"] + [None] * 14 + ["remote_sku_keep_01"])
        sheet.append([2, "product-03", "keep-b"] + [None] * 14 + ["remote_sku_keep_03"])
        sheet.append([3, "product-404", "keep-stale"] + [None] * 14 + ["remote_sku_keep_stale"])
        # Keep a populated row whose B cell is blank. workbook_plan scans
        # B4:max_row and therefore materializes sparse B7 in memory; openpyxl
        # may omit that semantically empty node again when saving.
        sheet.append([4, None, "keep-sparse-row"] + [None] * 14 + ["remote_sku_keep_sparse"])
        workbook.save(workbook_path)
        workbook.close()
        source_map = root / "product_source_map.json"
        source_map.write_text(json.dumps({
            "current": {"daisyflowdigital": "product-08", "templystudios": "product-08"},
            "historical": "product-08",
        }) + "\n", encoding="utf-8")
        remote_root = root / "remote"
        remote = LocalRemote(remote_root)

        if cloud:
            old_key = "shops/daisyflowdigital/product-08"
            revision = "20260812T120000Z-fixture"
            file_bytes = b"fixture"
            image_bytes = b"image"
            manifest = {
                "schema": 1,
                "type": "etsy-cloud-asset-manifest",
                "product": {
                    "scope": "shops", "shop": "daisyflowdigital",
                    "product": "product-08", "key": old_key,
                },
                "revision": revision,
                "created_at": "2026-08-12T12:00:00Z",
                "files": [
                    {"path": "files/customer.pdf", "role": "file", "size": len(file_bytes),
                     "sha256": hashlib.sha256(file_bytes).hexdigest()},
                    {"path": "images/gallery.png", "role": "image", "size": len(image_bytes),
                     "sha256": hashlib.sha256(image_bytes).hexdigest()},
                ],
                "counts": {"files": 1, "file_count": 1, "file_bytes": len(file_bytes),
                           "images": 1, "image_count": 1, "image_bytes": len(image_bytes),
                           "preview": 0, "preview_count": 0, "preview_bytes": 0,
                           "total": 2, "total_count": 2, "total_bytes": len(file_bytes) + len(image_bytes)},
            }
            manifest_data = canonical_json_bytes(manifest)
            digest = hashlib.sha256(manifest_data).hexdigest()
            state = {
                "schema": 1, "state": "CLOUD_ONLY" if cloud_only else "CLOUD_VERIFIED",
                "product": manifest["product"], "current_revision": revision,
                "current_manifest": manifest, "current_manifest_sha256": digest,
                "history": [],
            }
            (shop / "product-08" / ".cloud-assets.json").write_bytes(canonical_json_bytes(state))
            revision_root = remote_root / "assets/v1" / old_key / "revisions" / revision
            (revision_root / "files").mkdir(parents=True)
            (revision_root / "images").mkdir(parents=True)
            (revision_root / "files/customer.pdf").write_bytes(file_bytes)
            (revision_root / "images/gallery.png").write_bytes(image_bytes)
            (revision_root / "manifest.json").write_bytes(manifest_data)
            if not cloud_only:
                (shop / "product-08" / "files").mkdir()
                (shop / "product-08" / "images").mkdir()
                (shop / "product-08" / "files/customer.pdf").write_bytes(file_bytes)
                (shop / "product-08" / "images/gallery.png").write_bytes(image_bytes)
            pointer = {
                "schema": 1, "type": "etsy-cloud-current-pointer", "product": old_key,
                "revision": revision,
                "revision_path": f"assets/v1/{old_key}/revisions/{revision}",
                "manifest_sha256": digest, "verified_at": "2026-08-12T12:00:00Z",
            }
            current = remote_root / "assets/v1" / old_key / "current.json"
            current.parent.mkdir(parents=True, exist_ok=True)
            current.write_bytes(canonical_json_bytes(pointer))

        return {
            "root": root, "shop": shop, "workbook": workbook_path,
            "source_map": source_map, "remote": remote,
        }

    @staticmethod
    def _tree_hashes(root: Path, exclude_prefix: str = ""):
        result = {}
        for path in sorted(root.rglob("*")):
            relative = path.relative_to(root).as_posix()
            if exclude_prefix and relative.startswith(exclude_prefix):
                continue
            if path.is_file():
                result[relative] = hashlib.sha256(path.read_bytes()).hexdigest()
            elif path.is_dir():
                result[relative + "/"] = "dir"
        return result

    @staticmethod
    def _inject_empty_numeric_cell(path: Path, coordinate: str) -> None:
        """Insert the exact harmless XML shape present in the live workbook."""
        temporary = path.with_name("injected.xlsx")
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temporary, "w") as target:
            for info in source.infolist():
                data = source.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    data, count = re.subn(
                        rb'(<row r="2"[^>]*>.*?)(</row>)',
                        rb'\1' + f'<c r="{coordinate}" t="n"></c>'.encode() + rb'\2',
                        data,
                        count=1,
                    )
                    if count != 1:
                        raise AssertionError("could not inject empty numeric workbook cell")
                target.writestr(info, data)
        temporary.replace(path)


if __name__ == "__main__":
    unittest.main()
