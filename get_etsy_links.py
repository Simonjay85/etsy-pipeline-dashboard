"""
Etsy Link Fetcher
• Đăng nhập bằng session Chrome đã lưu (giống etsy_auto_post.py)
• Vào Shop Manager → Listings → lấy URL + tên từng listing đã publish
• Lưu link vào cột P của Etsy_SEO_Generator.xlsx
• Chạy: python3 get_etsy_links.py
"""
import asyncio, sys, subprocess
from pathlib import Path

def ensure_deps():
    for mod, pkg in {"openpyxl": "openpyxl", "playwright": "playwright"}.items():
        try:
            __import__(mod)
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", pkg, "--quiet"], check=True)

ensure_deps()

import openpyxl
from playwright.async_api import async_playwright

BASE_DIR    = Path(__file__).parent
EXCEL_FILE  = BASE_DIR / "shops" / "templystudios" / "Etsy_SEO_Generator.xlsx"
BROWSER_DIR = BASE_DIR / ".browser-session"
CHROME_PATH = Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")

async def main():
    print("\n" + "="*55)
    print("  🔗 Etsy Link Fetcher")
    print("="*55)

    BROWSER_DIR.mkdir(exist_ok=True)

    async with async_playwright() as pw:
        launch_kw = dict(
            user_data_dir=str(BROWSER_DIR),
            headless=False,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
            viewport=None,
        )
        if CHROME_PATH.exists():
            launch_kw["executable_path"] = str(CHROME_PATH)

        ctx  = await pw.chromium.launch_persistent_context(**launch_kw)
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()

        # Kiểm tra đăng nhập
        await page.goto("https://www.etsy.com/your/shops/me/tools/listings",
                        wait_until="domcontentloaded")
        await page.wait_for_timeout(4000)

        if "signin" in page.url or "join" in page.url:
            print("\n  ⚠  Chưa đăng nhập Etsy! Vui lòng đăng nhập trên trình duyệt vừa mở.")
            input("  → Nhấn Enter sau khi đăng nhập xong...\n")
            await page.goto("https://www.etsy.com/your/shops/me/tools/listings",
                            wait_until="domcontentloaded")
            await page.wait_for_timeout(4000)

        print("  ✅ Đã vào Shop Manager\n")

        # ── Lấy tất cả listing (active + draft) ──────────────────────────────
        all_listings = []

        for status_filter in ["active", "draft"]:
            url = f"https://www.etsy.com/your/shops/me/tools/listings?status={status_filter}"
            await page.goto(url, wait_until="domcontentloaded")
            await page.wait_for_timeout(3000)

            # Scroll để load hết
            for _ in range(5):
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(1000)

            # Lấy listing items
            rows = await page.query_selector_all('[data-listing-id], .listing-row, [class*="listing-card"]')

            if not rows:
                # Thử selector khác
                rows = await page.query_selector_all('a[href*="/listing/"]')

            count = 0
            seen_ids = set()

            # Lấy tất cả edit links chứa title và listing ID
            links = await page.query_selector_all('a[href*="/listing-editor/edit/"]')
            for link in links:
                href = await link.get_attribute("href")
                if not href:
                    continue
                # Trích xuất listing_id từ href (ví dụ: https://www.etsy.com/.../edit/4509070606)
                parts = href.split("/edit/")
                if len(parts) < 2:
                    continue
                listing_id = parts[1].split("?")[0].split("/")[0]
                if not listing_id.isdigit():
                    continue

                full_url = f"https://www.etsy.com/listing/{listing_id}"
                if full_url in seen_ids:
                    continue

                # Lấy và chuẩn hóa title
                raw_text = await link.inner_text()
                lines = [line.strip() for line in raw_text.split("\n") if line.strip()]
                title = ""
                for line in lines:
                    if line in ["Digital", "Video", "Editing options"]:
                        continue
                    if "in stock" in line.lower() or "auto-renews" in line.lower() or "$" in line:
                        continue
                    title = line
                    break
                if not title and lines:
                    title = lines[0]
                title = title[:80].strip()

                if title:
                    seen_ids.add(full_url)
                    all_listings.append({
                        "title": title,
                        "url": full_url,
                        "status": status_filter
                    })
                    count += 1

            # Fallback sang data-listing-id nếu phương pháp trên không tìm thấy
            if count == 0:
                listing_cards = await page.query_selector_all('[data-listing-id]')
                for card in listing_cards:
                    listing_id = await card.get_attribute("data-listing-id")
                    if not listing_id or not listing_id.isdigit():
                        continue
                    url = f"https://www.etsy.com/listing/{listing_id}"
                    if url in seen_ids:
                        continue
                    title_el = await card.query_selector('h3, [class*="title"], p[class*="title"]')
                    raw_text = await title_el.inner_text() if title_el else f"Listing {listing_id}"
                    lines = [line.strip() for line in raw_text.split("\n") if line.strip()]
                    title = ""
                    for line in lines:
                        if line in ["Digital", "Video", "Editing options"]:
                            continue
                        if "in stock" in line.lower() or "auto-renews" in line.lower() or "$" in line:
                            continue
                        title = line
                        break
                    if not title and lines:
                        title = lines[0]
                    title = title[:80].strip()

                    seen_ids.add(url)
                    all_listings.append({
                        "title": title,
                        "url": url,
                        "status": status_filter
                    })
                    count += 1

            print(f"  📋 {status_filter.upper()}: {count} listings")

        if not all_listings:
            print("\n  ⚠  Không tìm thấy listing nào.")
            print("  → Hãy đảm bảo bạn đã publish ít nhất 1 sản phẩm trên Etsy.")
            await ctx.close()
            return

        # ── Hiển thị kết quả ─────────────────────────────────────────────────
        print(f"\n  Tổng: {len(all_listings)} listings\n")
        for i, l in enumerate(all_listings, 1):
            status_icon = "✅" if l["status"] == "active" else "📝"
            print(f"  {i:2}. {status_icon} {l['title'][:50]}")
            print(f"      {l['url']}")

        # ── Lưu vào Excel ─────────────────────────────────────────────────────
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Listings"]

        # Tiêu đề cột P (col 16)
        ws.cell(row=3, column=16, value="Etsy URL")

        saved = 0
        matched_urls = set()
        for row_idx in range(4, ws.max_row + 1):
            title_cell = ws.cell(row=row_idx, column=8).value  # cột H = Title
            if not title_cell:
                continue
            title_str = str(title_cell).strip()[:60]

            # Match với listing từ Etsy (so sánh title)
            best_match = None
            best_score = 0
            for l in all_listings:
                if l["url"] in matched_urls:
                    continue
                # So sánh overlap từ
                words_excel = set(title_str.lower().split())
                words_etsy  = set(l["title"].lower().split())
                overlap = len(words_excel & words_etsy)
                score = overlap / max(len(words_excel), 1)
                if score > best_score and score >= 0.75:
                    best_score = score
                    best_match = l

            if best_match:
                ws.cell(row=row_idx, column=16, value=best_match["url"])
                matched_urls.add(best_match["url"])
                saved += 1
                print(f"\n  ✅ Row {row_idx} → {best_match['url'][:60]} (Score: {best_score:.2f})")
            else:
                # Clear Column P if no match found on Etsy to prevent false associations
                ws.cell(row=row_idx, column=16, value=None)

        wb.save(EXCEL_FILE)
        print(f"\n  💾 Đã lưu {saved} links vào Excel (cột P)")

        # ── Xuất file links.txt ────────────────────────────────────────────────
        links_file = BASE_DIR / "etsy_links.txt"
        with open(links_file, "w", encoding="utf-8") as f:
            f.write("ETSY LISTING LINKS\n")
            f.write("="*60 + "\n\n")
            for i, l in enumerate(all_listings, 1):
                status_icon = "✅ ACTIVE" if l["status"] == "active" else "📝 DRAFT"
                f.write(f"{i}. [{status_icon}] {l['title']}\n")
                f.write(f"   {l['url']}\n\n")

        print(f"\n  📄 Đã xuất: etsy_links.txt")
        print(f"{'='*55}\n")

        await ctx.close()

if __name__ == "__main__":
    asyncio.run(main())
