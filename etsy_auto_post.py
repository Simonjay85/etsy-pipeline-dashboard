"""
Etsy Auto Draft Poster
• Dùng Chrome thật (tránh bot detection)
• Hỗ trợ form mới dạng tab của Etsy
• Tự upload ảnh + PDF
• Tự dịch 9 ngôn ngữ (title + description + tags)
• Chọn Shop Section từ Excel
• Delay hợp lý để tránh lỗi
Dùng: python3 etsy_auto_post.py [--batch 5] [--skip N]
"""
from __future__ import annotations

import argparse
import asyncio
from contextlib import contextmanager
import filecmp
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
import re
import unicodedata
from difflib import SequenceMatcher

# ── Auto-install ───────────────────────────────────────────────────────────────
def ensure_deps():
    pkgs = {"openpyxl": "openpyxl", "playwright": "playwright", "deep_translator": "deep-translator"}
    for mod, pkg in pkgs.items():
        try:
            __import__(mod)
        except ImportError:
            print(f"▶ Cài {pkg}...")
            subprocess.run([sys.executable, "-m", "pip", "install", pkg, "--quiet"], check=True)
    try:
        from google import genai
    except ImportError:
        print("▶ Cài google-genai...")
        subprocess.run([sys.executable, "-m", "pip", "install", "google-genai", "--quiet"], check=True)

def initialize_poster_runtime() -> None:
    """Chuẩn bị dependency chỉ khi poster thực sự được chạy (không chạy khi import)."""
    ensure_deps()

if __name__ == "__main__":
    initialize_poster_runtime()

import openpyxl
from playwright.async_api import TimeoutError as PlaywrightTimeoutError, async_playwright
from deep_translator import GoogleTranslator
from etsy_browser_session import (
    is_session_ready as is_etsy_session_ready,
    resolve_etsy_session,
)
from cloud_asset_store import CloudAssetError, CloudAssetStore
from cloud_asset_store_config import load_config as load_cloud_asset_config

try:
    from google import genai
    from google.genai import types
    has_genai = True
except ImportError:
    has_genai = False

BASE_DIR    = Path(__file__).parent
SHOP_DIR    = BASE_DIR
EXCEL_FILE  = SHOP_DIR / "Etsy_SEO_Generator.xlsx"
CHROME_PATH = Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")
CONFIG_FILE = BASE_DIR / "shops_config.json"
SHOPS      = {}
CLOUD_ASSET_STORE: CloudAssetStore | None = None
POST_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
POST_FILE_EXTS = {".pdf", ".zip"}
CDP_CONNECT_ATTEMPTS = 3
CDP_CONNECT_TIMEOUT_MS = 10000
CDP_CONNECT_RETRY_DELAYS = (0.25, 0.50)
SHOP_MANAGER_LISTINGS_URL = "https://www.etsy.com/your/shops/me/tools/listings"
# Shop Manager can take longer than the default Playwright navigation timeout
# while an already-authenticated CDP page is waking up. Keep this bounded and
# retry only the same page/session; callers still perform the sign-in and shop
# identity gates after navigation returns.
SHOP_MANAGER_NAVIGATION_TIMEOUT_MS = 60000
SHOP_MANAGER_NAVIGATION_ATTEMPTS = 2
SHOP_MANAGER_NAVIGATION_RETRY_DELAY_MS = 1000


def get_cloud_asset_store() -> CloudAssetStore:
    global CLOUD_ASSET_STORE
    if CLOUD_ASSET_STORE is None:
        config = load_cloud_asset_config(BASE_DIR)
        CLOUD_ASSET_STORE = CloudAssetStore(
            repo_root=config.repo_root,
            remote=config.remote,
            parent_id=config.parent_id,
            rclone_bin=config.rclone_bin,
            cache_root=config.cache_root,
            lock_timeout_seconds=config.lock_timeout_seconds,
            success_ttl_seconds=config.success_ttl_seconds,
            failure_ttl_seconds=config.failure_ttl_seconds,
            offload_age_days=config.offload_age_days,
        )
    return CLOUD_ASSET_STORE


def resolve_product_asset_paths(
    product: dict,
    shop_id: str,
    store: CloudAssetStore | None = None,
) -> dict:
    """Resolve verified local/cache paths before the Etsy editor is opened."""

    folder = str(product.get("folder") or "").strip()
    if not re.fullmatch(r"product-\d+", folder):
        raise RuntimeError(f"❌ Product folder không hợp lệ: {folder}")
    configured_shop_root = Path(SHOP_DIR)
    if configured_shop_root.name == str(shop_id).strip() and configured_shop_root.parent.name == "shops":
        product_root = configured_shop_root / folder
    else:
        product_root = BASE_DIR / "shops" / str(shop_id).strip() / folder
    if not product_root.is_dir() or product_root.is_symlink():
        raise RuntimeError(f"❌ Không tìm thấy product folder an toàn: {product_root}")
    asset_store = store or get_cloud_asset_store()
    store_repo_root = getattr(asset_store, "repo_root", None)
    if store_repo_root is not None:
        try:
            product_root.absolute().relative_to(Path(store_repo_root).absolute())
        except ValueError:
            # Unit/in-process callers may intentionally point SHOP_DIR at a
            # temporary checkout. Preserve local-only compatibility while
            # keeping CloudAssetStore's canonical path validation active.
            local_repo_root = product_root.parents[2]
            asset_store = CloudAssetStore(
                repo_root=local_repo_root,
                remote_store=getattr(asset_store, "remote", None),
                cache_root=local_repo_root / "output" / "cloud-cache",
                lock_timeout_seconds=getattr(asset_store, "lock_timeout_seconds", 30.0),
                success_ttl_seconds=getattr(asset_store.cache, "success_ttl_seconds", 24 * 60 * 60),
                failure_ttl_seconds=getattr(asset_store.cache, "failure_ttl_seconds", 7 * 24 * 60 * 60),
                offload_age_days=getattr(asset_store, "offload_age_days", 7),
            )
    try:
        resolution = asset_store.resolve_asset_root(product_root)
    except (CloudAssetError, OSError, ValueError, TypeError, KeyError) as exc:
        raise RuntimeError(f"❌ Không hydrate được asset {shop_id}/{folder}: {exc}") from exc

    if not isinstance(resolution, dict):
        raise RuntimeError(f"❌ Asset resolver trả về kết quả không hợp lệ cho {shop_id}/{folder}")
    source = str(resolution.get("source") or "")
    if source not in {"local", "cloud-cache"}:
        raise RuntimeError(f"❌ Asset resolver không xác nhận được nguồn cho {shop_id}/{folder}")
    asset_root = Path(str(resolution.get("asset_root") or product_root))
    if asset_root.is_symlink() or not asset_root.is_dir():
        raise RuntimeError(f"❌ Asset root không an toàn cho {shop_id}/{folder}: {asset_root}")
    allowed_root = product_root if source == "local" else asset_root

    def verified_paths(raw_paths, allowed_suffixes: set[str]) -> list[str]:
        if not isinstance(raw_paths, (list, tuple)):
            raise RuntimeError(f"❌ Danh sách asset không hợp lệ cho {shop_id}/{folder}")
        verified = []
        for raw_path in raw_paths:
            path = Path(str(raw_path))
            if path.is_symlink() or not path.is_file():
                raise RuntimeError(f"❌ Asset path không tồn tại/an toàn cho {shop_id}/{folder}: {path}")
            try:
                path.absolute().relative_to(allowed_root.absolute())
            except ValueError as exc:
                raise RuntimeError(f"❌ Asset path nằm ngoài vùng đã xác minh: {path}") from exc
            if path.suffix.lower() in allowed_suffixes:
                verified.append(str(path))
        return sorted(verified)

    image_paths = verified_paths(resolution.get("image_paths", []), POST_IMAGE_EXTS)[:10]
    digital_paths = verified_paths(resolution.get("file_paths", []), POST_FILE_EXTS)
    if resolution.get("source") == "cloud-cache" and not image_paths:
        raise RuntimeError(f"❌ Cloud asset {shop_id}/{folder} đã hydrate nhưng không có ảnh hợp lệ")
    product["image_paths"] = image_paths
    product["pdf_paths"] = digital_paths
    product["asset_root"] = str(resolution.get("asset_root") or product_root)
    product["_cloud_asset_resolution"] = resolution
    return resolution


def mark_product_asset_operation_success(
    product: dict,
    store: CloudAssetStore | None = None,
) -> dict:
    resolution = product.get("_cloud_asset_resolution")
    if not resolution:
        return {"ok": True, "marked": False}
    return (store or get_cloud_asset_store()).mark_hydration_cleanup_eligible(resolution)

def _load_shop_configs(config_file: Path = CONFIG_FILE) -> dict:
    if not config_file.exists():
        return {}
    try:
        with config_file.open("r", encoding="utf-8") as f:
            loaded = json.load(f)
            if isinstance(loaded, dict):
                return loaded
    except Exception:
        pass
    return {}

SHOPS = _load_shop_configs()

def _expand_user_path(value: str, home_dir: Path) -> Path:
    raw = str(value or "").strip()
    if not raw:
        return Path(raw)
    if raw.startswith("~"):
        return Path(str(home_dir) + raw[1:])
    return Path(raw)

def resolve_browser_session_dir(
    shop_id: str,
    *,
    config: dict | None = None,
    base_dir: Path | None = None,
    home_dir: Path | None = None,
) -> Path:
    target_shop = str(shop_id or "").strip()
    target_shop_key = target_shop.lower()
    shops_config = config if config is not None else SHOPS
    target_shop_cfg = shops_config.get(target_shop, {}) if isinstance(shops_config, dict) else {}
    raw_session = str(target_shop_cfg.get("browser_session", "")).strip()
    base_root = base_dir or BASE_DIR
    home_root = home_dir or Path.home()

    if raw_session:
        return _expand_user_path(raw_session, home_root)

    legacy_session = base_root / ".browser-session"
    if target_shop_key == "templystudios" and legacy_session.exists():
        return legacy_session

    return home_root / f".etsy_browser_session_{target_shop}"

def _is_profile_in_use(profile_dir: Path) -> tuple[bool, Path | None]:
    for lock_name in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
        lock_path = profile_dir / lock_name
        if lock_path.exists():
            return True, lock_path
    return False, None


async def _open_poster_context(pw, shop_id: str, browser_dir: Path):
    """Reuse only the verified login browser for this exact poster profile."""
    session = resolve_etsy_session(BASE_DIR, SHOPS, shop_id)
    expected_profile = session.profile_dir.resolve()
    if browser_dir.resolve() != expected_profile:
        raise RuntimeError(
            f"❌ Profile Etsy không khớp cấu hình poster: {browser_dir} != {expected_profile}"
        )

    if is_etsy_session_ready(session):
        last_error: Exception | None = None
        browser = None
        for attempt in range(CDP_CONNECT_ATTEMPTS):
            try:
                browser = await pw.chromium.connect_over_cdp(
                    session.cdp_url, timeout=CDP_CONNECT_TIMEOUT_MS
                )
            except PlaywrightTimeoutError as exc:
                last_error = exc
                if attempt + 1 >= CDP_CONNECT_ATTEMPTS:
                    break
                if not await asyncio.to_thread(is_etsy_session_ready, session):
                    raise RuntimeError(
                        f"❌ CDP đang bận hoặc không phản hồi cho shop {shop_id}, "
                        f"profile {session.profile_dir}, port {session.debug_port}. "
                        "Hãy mở lại phiên đăng nhập đúng profile rồi thử lại."
                    ) from exc
                await asyncio.sleep(CDP_CONNECT_RETRY_DELAYS[min(attempt, len(CDP_CONNECT_RETRY_DELAYS) - 1)])
                continue
            break

        if browser is None:
            raise RuntimeError(
                f"❌ CDP đang bận hoặc không phản hồi cho shop {shop_id}, "
                f"profile {session.profile_dir}, port {session.debug_port}. "
                "Hãy mở lại phiên đăng nhập đúng profile rồi thử lại."
            ) from last_error
        if not browser.contexts:
            raise RuntimeError("❌ Chrome Etsy đúng profile nhưng không có browser context")
        ctx = browser.contexts[0]
        page = await ctx.new_page()
        print(f"  🌐 Dùng lại Chrome Etsy đã đăng nhập (port {session.debug_port})")
        return ctx, page, False

    in_use, lock_path = _is_profile_in_use(browser_dir)
    if in_use and lock_path is not None:
        raise RuntimeError(
            f"❌ Chrome profile đang bị khóa ({lock_path.name}) nhưng không khớp "
            f"session Etsy đã xác minh cho shop {shop_id}. Đóng Chrome đó rồi mở lại "
            "bằng nút “Đăng nhập Etsy cho Post”."
        )

    launch_kw = dict(
        user_data_dir=str(browser_dir),
        headless=False,
        args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
        viewport=None,
    )
    if CHROME_PATH.exists():
        launch_kw["executable_path"] = str(CHROME_PATH)
        print("  🌐 Dùng Google Chrome thật")
    else:
        print("  🌐 Dùng Chromium")
    ctx = await pw.chromium.launch_persistent_context(**launch_kw)
    page = ctx.pages[0] if ctx.pages else await ctx.new_page()
    return ctx, page, True

# ── Config ─────────────────────────────────────────────────────────────────────
FILL_TRANSLATIONS = True
DEFAULT_BATCH     = 5
DRAFT_LISTINGS_URL = "https://www.etsy.com/your/shops/me/tools/listings/page:1,state:draft"
DRAFT_FILTER_POLL_MS = 250
DRAFT_FILTER_CHECK_TIMEOUT_MS = 1200
DRAFT_FILTER_FORCE_ATTEMPTS = 10
UNVERIFIED_DRAFT_URL_SENTINEL = "<URL chưa xác minh>"
DRAFT_DUPLICATE_CHECK_FAILED_SENTINEL = "<DRAFT_DUPLICATE_CHECK_FAILED>"

LANGUAGES = [
    ("nl", "Dutch",      0),
    ("fr", "French",     1),
    ("de", "German",     2),
    ("it", "Italian",    3),
    ("ja", "Japanese",   4),
    ("pl", "Polish",     5),
    ("pt", "Portuguese", 6),
    ("ru", "Russian",    7),
    ("es", "Spanish",    8),
]

def trim_title(title: str, max_len: int = 140) -> str:
    """Cắt title tối đa max_len ký tự, không cắt giữa từ và dọn dẹp ký tự đặc biệt."""
    import re, html
    
    # Giải mã thực thể HTML (ví dụ: &amp; -> &, &quot; -> ", etc.)
    title = html.unescape(title)
    
    # Helper to replace subsequent occurrences of a character
    def replace_subsequent(s: str, char: str, replacement: str) -> str:
        parts = s.split(char)
        if len(parts) > 2:
            return parts[0] + char + replacement.join(parts[1:])
        return s

    # Giữ lại tối đa 1 ký tự đặc biệt theo quy định của Etsy, các ký tự thừa sẽ được chuyển đổi an toàn
    title = replace_subsequent(title, "&", " and ")
    title = replace_subsequent(title, "%", " percent")
    title = replace_subsequent(title, ":", " -")
    
    # Thu gọn nhiều khoảng trắng liên tiếp
    title = re.sub(r'\s+', ' ', title).strip()

    if len(title) <= max_len:
        return title
    cut = title[:max_len].rsplit(" ", 1)[0].rstrip(",|;- ")
    return cut

# ── Alt text generator ────────────────────────────────────────────────────────
def generate_alt_texts(title: str, keywords: str, count: int) -> list:
    """Tạo alt text cho từng ảnh dựa vào title và keywords (tối đa 250 ký tự/ảnh)."""
    clean_title = title[:200].strip()
    kw_list = [k.strip() for k in keywords.split(",") if k.strip()]

    suffixes = [
        "instant download",
        "printable PDF",
        "digital download",
        "printable template",
        "digital planner",
        "PDF download",
        "printable download",
        "digital file",
        "instant printable",
        "editable template",
    ]

    alt_texts = []
    for i in range(count):
        if i == 0:
            # Ảnh đầu tiên dùng full title
            alt_texts.append(clean_title[:250])
        else:
            suffix = suffixes[i % len(suffixes)]
            if kw_list:
                kw = kw_list[i % len(kw_list)]
                text = f"{kw} - {suffix}"
            else:
                text = f"{clean_title[:200]} - {suffix}"
            alt_texts.append(text[:250])
    return alt_texts

# ── Tag cleaner ───────────────────────────────────────────────────────────────
def clean_tags(raw: str, max_tags: int = 13) -> list:
    """Chuẩn hoá danh sách tag: mỗi tag 1-20 ký tự, tối đa max_tags tag duy nhất (case-insensitive)."""
    result = []
    seen = set()
    for t in str(raw or "").split(","):
        t = t.strip()
        if not t:
            continue
        if 1 <= len(t) <= 20:
            tag_str = t
        else:
            tag_str = t[:20].rsplit(" ", 1)[0].rstrip(",- ")
        if 1 <= len(tag_str) <= 20:
            key = tag_str.lower()
            if key not in seen:
                seen.add(key)
                result.append(tag_str)
    return result[:max_tags]


def normalize_category_text(value: str) -> str:
    """Chuẩn hóa chữ cho so sánh danh mục."""
    return re.sub(r"\s+", " ", str(value or "").lower().replace("✓", " ").strip())


_CATEGORY_OPTION_METADATA_SUFFIXES = ("physical or digital", "physical", "digital")


def _strip_metadata_suffix(category: str) -> tuple[str, str | None]:
    normalized = normalize_category_text(category)
    for suffix in _CATEGORY_OPTION_METADATA_SUFFIXES:
        for suffix_token in (f" {suffix}", f" ({suffix})"):
            if normalized.endswith(suffix_token):
                base = normalized[: -len(suffix_token)].rstrip()
                if base:
                    return base, suffix
    return normalized, None


_CATEGORY_DIGITAL_ONLY_METADATA_RE = re.compile(
    r"(?:\(\s*digital\s*\)|\bdigital\b)\s*[.!?]*$",
    re.IGNORECASE,
)


def _category_option_is_digital_only_metadata(option_text: str) -> bool:
    """Return true only for an unambiguous terminal Digital metadata marker."""
    normalized = normalize_category_text(option_text)
    if not normalized or re.search(r"\bphysical\b", normalized, re.IGNORECASE):
        return False

    match = _CATEGORY_DIGITAL_ONLY_METADATA_RE.search(normalized)
    if not match:
        return False

    prefix = normalized[:match.start()].rstrip()
    if not prefix:
        return False

    # A terminal breadcrumb leaf named Digital is not metadata evidence. The
    # plain suffix form must be attached to the category label itself; a
    # parenthesized marker is independently unambiguous.
    marker = match.group(0).lstrip()
    if not marker.startswith("(") and prefix.endswith((">", "/", "\\")):
        return False
    return True


def category_option_matches(target_category: str, option_text: str) -> bool:
    """Khớp chính xác theo option label hoặc leaf breadcrumb, có xử lý suffix metadata."""
    target_norm = normalize_category_text(target_category)
    if not target_norm:
        return False

    option_norm = normalize_category_text(option_text)
    if not option_norm:
        return False

    target_base, _ = _strip_metadata_suffix(target_norm)
    target_candidates = {target_norm, target_base}

    option_leaf_norm = normalize_category_text(extract_category_leaf(option_text))
    option_candidates = {option_norm, option_leaf_norm}
    option_without_prefix = re.sub(r"^\+\s*", "", option_norm)
    if option_without_prefix:
        option_candidates.add(option_without_prefix)
        for suffix in _CATEGORY_OPTION_METADATA_SUFFIXES:
            for suffix_token in (f" {suffix}", f" ({suffix})"):
                if option_without_prefix.endswith(suffix_token):
                    base = option_without_prefix[: -len(suffix_token)].rstrip()
                    if base:
                        option_candidates.add(base)
    option_parts = [part.strip() for part in re.split(r"\s*>\s*|\s*/\s*|\\", option_text) if part.strip()]
    if len(option_parts) >= 2:
        suffix_candidate = normalize_category_text(option_parts[-1])
        if suffix_candidate in _CATEGORY_OPTION_METADATA_SUFFIXES:
            option_candidates.add(normalize_category_text(option_parts[-2]))

    for option_candidate in option_candidates:
        if not option_candidate:
            continue
        if option_candidate in target_candidates:
            return True
        option_base, _ = _strip_metadata_suffix(option_candidate)
        if option_base and option_base in target_candidates:
            return True
    return False


def extract_category_leaf(category_value: str) -> str:
    """Lấy leaf name từ chuỗi danh mục dạng 'A > B > C' hoặc path."
    """
    category_value = str(category_value or "").strip()
    if not category_value:
        return ""
    # Ưu tiên tách bằng các delimiter phổ biến của breadcrumb
    parts = re.split(r"\s*>\s*|\s*/\s*|\\", category_value)
    leaf = parts[-1] if parts else category_value
    return leaf.replace("✓", "").strip()


def infer_listing_category(product: dict) -> str:
    """Suy luận danh mục từ nội dung sản phẩm (title/keywords/tags/description)."""
    combined = " ".join([
        str(product.get("title", "")),
        str(product.get("keywords", "")),
        str(product.get("tags", "")),
        str(product.get("description", "")),
    ]).lower()
    norm = re.sub(r"[^a-z0-9]+", " ", combined)

    if re.search(r"\b(svg|dxf|eps|vector|cut file|cricut|silhouette)\b", norm):
        return "Cutting Machine Files"

    # STL/3MF and explicit digital 3D model/file language map to the Etsy leaf
    # Craft Supplies & Tools > Patterns & How To > Craft Machine Files >
    # 3D Printer Files. A bare "3D" mention is accepted only when it also
    # appears in a verified downloadable path; this keeps physical products
    # and generic 3D styling fail-closed.
    digital_paths = [str(path) for path in (product.get("pdf_paths") or [])]
    digital_norm = re.sub(r"[^a-z0-9]+", " ", " ".join(digital_paths).lower())
    three_d_resource_stems = (
        r"\b(stl|3mf|obj|fbx|blend|gltf|glb|step|stp|iges|igs)\b|"
        r"\b3d\s+(model|models|file|files|resource|resources|design|designs)\b|"
        r"\b3d\s+print(?:able|ing)?\s+(file|files|model|models)\b|"
        r"\b3d\s+printer\s+(file|files|model|models)\b"
    )
    has_verified_bare_3d_download = (
        bool(digital_paths)
        and re.search(r"\b3d\b", norm)
        and re.search(r"\b3d\b", digital_norm)
    )

    if re.search(three_d_resource_stems, f"{norm} {digital_norm}") or has_verified_bare_3d_download:
        return "3D Printer Files"

    if re.search(r"\b(resume|curriculum)\b|\bcv\b", norm):
        return "Résumé Templates"

    planner_stems = (
        r"\b(planner|planners|journal|journals|journaling|calendar|calendars|"
        r"workbook|workbooks|worksheet|worksheets|checklist|checklists|tracker|trackers)\b"
    )
    if re.search(planner_stems, norm):
        return "Planner Templates"

    kdp_bundle_stems = r"\b(kdp|low[-\s]*content|book[-\s]*interior|book[-\s]*interiors)\b"
    kdp_template_stems = r"\b(template|templates|bundle|pack|kit)\b"
    if re.search(kdp_bundle_stems, norm) and re.search(kdp_template_stems, norm):
        return "Planner Templates"

    ai_template_stems = (
        r"\b("
        r"ai\s*commands?|"
        r"ai\s+prompt\s+guide|"
        r"prompt\s+guide|"
        r"prompt\s+commands?|"
        r"etsy\s+sellers?|etsy\s+seller|"
        r"chatgpt|"
        r"prompt\s+pack|"
        r"prompt\s+resource"
        r")\b"
    )
    if re.search(ai_template_stems, norm):
        return "Guides & How Tos"

    # Imported rows in the shop workbook may have an empty Category column
    # while still carrying the canonical Digital Planner section. These rows
    # are valid digital template listings, so keep category selection
    # deterministic instead of failing with "Không suy luận được danh mục".
    section_norm = re.sub(
        r"[^a-z0-9]+",
        " ",
        str(product.get("section", "")).lower(),
    ).strip()
    if section_norm in {"digital planner", "digital template", "digital templates"}:
        return "Planner Templates"

    return ""


def resolve_listing_category(product: dict) -> str:
    """Giải quyết danh mục cuối cùng để chọn: ưu tiên category trong workbook, nếu trống thì suy luận."""
    explicit = extract_category_leaf(product.get("category", ""))
    if explicit:
        return explicit
    return infer_listing_category(product)


def normalize_search_text(value: str) -> str:
    """Chuẩn hóa chuỗi cho so khớp tiêu đề/lưu trữ."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip())


def normalize_sku(value: str) -> str:
    """Chuẩn hóa SKU để so sánh chính xác."""
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


async def _force_draft_filter(page):
    """Điều hướng về drafts và ép filter về Draft status an toàn trước khi đọc listings."""
    print("  🔧 Ép bộ lọc Draft để đọc đúng danh sách.")
    await page.goto(DRAFT_LISTINGS_URL, wait_until="domcontentloaded")
    for _ in range(24):
        status_state = await _collect_status_radio_state(page)
        if status_state["radio_count"] > 0 and status_state["draft_indexes"]:
            break
        await page.wait_for_timeout(250)

    status_state = await _collect_status_radio_state(page)
    if status_state["radio_count"] == 0:
        raise RuntimeError("Không tìm thấy bộ lọc Draft trên danh sách listings.")

    all_status_locator = page.locator('input[name="item_status"]')
    for _ in range(DRAFT_FILTER_FORCE_ATTEMPTS):
        status_state = await _collect_status_radio_state(page)
        if _is_exact_draft_filter(status_state):
            return

        if not status_state["draft_indexes"]:
            await page.wait_for_timeout(DRAFT_FILTER_POLL_MS)
            continue

        # Ưu tiên radio Draft đang enabled để tránh lỗi intercept/pointer events.
        draft_radio_index = None
        for idx in status_state["draft_indexes"]:
            if status_state["radios"][idx]["enabled"]:
                draft_radio_index = idx
                break

        if draft_radio_index is None:
            await page.wait_for_timeout(DRAFT_FILTER_POLL_MS)
            continue

        draft_radio = all_status_locator.nth(draft_radio_index)
        try:
            await draft_radio.check(force=True, timeout=DRAFT_FILTER_CHECK_TIMEOUT_MS)
        except Exception as err:
            raise RuntimeError(f"Không thể bật radio Draft: {err}")

        for _ in range(3):
            status_state = await _collect_status_radio_state(page)
            if _is_exact_draft_filter(status_state):
                return
            await page.wait_for_timeout(DRAFT_FILTER_POLL_MS)

    raise RuntimeError("Không xác nhận được filter Draft đã được bật.")


def _is_exact_draft_filter(state: dict) -> bool:
    checked = state.get("checked_values", [])
    return len(checked) == 1 and checked[0] == "draft"


async def _collect_status_radio_state(page) -> dict:
    """Thu thập trạng thái nhóm filter item_status trước khi quét listing."""
    status_locator = page.locator('input[name="item_status"]')
    status_count = await status_locator.count()
    radios = []
    draft_indexes = []
    for i in range(status_count):
        loc = status_locator.nth(i)
        raw_value = await loc.get_attribute("value")
        value = (raw_value or "").strip().lower()
        checked = bool(await loc.is_checked())
        enabled = bool(await loc.is_enabled())
        radio = {
            "value": value,
            "checked": checked,
            "enabled": enabled,
        }
        radios.append(radio)
        if value == "draft":
            draft_indexes.append(i)
    checked_values = sorted([r["value"] for r in radios if r.get("checked") and r.get("value")])

    return {
        "radio_count": status_count,
        "radios": radios,
        "draft_indexes": draft_indexes,
        "checked_count": len(checked_values),
        "checked_values": checked_values,
    }


async def _ensure_draft_filter_after_grid(page) -> None:
    status_state = await _collect_status_radio_state(page)
    if not _is_exact_draft_filter(status_state):
        raise RuntimeError(
            "Lọc Draft không còn chính xác sau khi grid ổn định. "
            "Dừng để tránh xác minh trùng lặp sai."
        )


def _normalize_listing_field(value) -> str:
    v = str(value or "").strip()
    return v


def _is_meaningful_title(title: str) -> bool:
    """Loại các text UI không phải tiêu đề sản phẩm thực tế."""
    norm = normalize_category_text(title)
    if not norm:
        return False
    if len(norm) <= 2:
        return False
    banned = {
        "edit",
        "view",
        "save",
        "edit listing",
        "preview",
        "delete",
        "listings",
    }
    return norm not in banned and norm != "actions"


def _pick_best_title(existing_title: str, incoming_title: str) -> str:
    """
    Chọn tiêu đề đáng tin nhất khi cùng một listing ID bị trùng.
    Ưu tiên meaningful và dài hơn.
    """
    cur = existing_title or ""
    inc = incoming_title or ""

    cur_norm = normalize_category_text(cur)
    inc_norm = normalize_category_text(inc)
    cur_meaningful = _is_meaningful_title(cur_norm)
    inc_meaningful = _is_meaningful_title(inc_norm)

    if cur_meaningful != inc_meaningful:
        return inc if inc_meaningful else cur

    if inc_meaningful and len(inc_norm) > len(cur_norm):
        return inc

    return cur


def _dedupe_draft_cards(cards: list[dict]) -> list[dict]:
    """
    Deduplicate listings by ID and merge sparse fields (title/SKU) so we don't
    drop information due to one imperfect DOM snapshot.
    """
    merged: dict[str, dict] = {}
    for raw in cards:
        if not isinstance(raw, dict):
            continue
        card_id = _normalize_listing_field(raw.get("id"))
        if not card_id:
            continue

        existing = merged.get(card_id)
        if existing is None:
            merged[card_id] = {**raw}
            merged[card_id]["id"] = card_id
            continue

        for key, value in raw.items():
            if key == "id":
                continue
            if key == "title":
                existing["title"] = _pick_best_title(str(existing.get("title", "")), str(value))
                continue

            existing_norm = _normalize_listing_field(existing.get(key))
            incoming_norm = _normalize_listing_field(value)
            if not existing_norm and incoming_norm:
                existing[key] = value

        if _normalize_listing_field(existing.get("status")).lower() != "draft":
            incoming_status = _normalize_listing_field(raw.get("status")).lower()
            if incoming_status == "draft":
                existing["status"] = raw.get("status")

    return list(merged.values())


def _get_media_thumbnail_selectors() -> list[str]:
    """Primary selectors for uploaded listing photos (delete/remove controls)."""
    return [
        'button[data-testid="image-delete-button"]',
        '[data-testid*="photo" i] button[aria-label*="Remove" i]',
        '[data-testid*="photo" i] button[aria-label*="Delete" i]',
    ]


def _get_media_thumbnail_fallback_selectors() -> list[str]:
    """Legacy thumbnail tiles — may include the empty upload slot."""
    return [
        "button.le-aspect-ratio--square",
    ]


PHOTO_UPLOAD_BATCH_SIZE = 5
# The observed Etsy editor can take roughly two minutes to settle a five-photo
# batch. Keep this bounded while allowing that slow, but finite, upload window.
PHOTO_UPLOAD_WAIT_MS = 180000
PHOTO_UPLOAD_POLL_MS = 500


async def _count_listing_image_thumbs(page) -> int:
    """Đếm ảnh listing đã upload trên UI.

    Prefer delete/remove buttons. Square aspect-ratio tiles can include Etsy's
    empty upload slot, which falsely inflates exact-count checks (e.g. 11 vs 10).
    """
    primary_counts: list[int] = []
    for sel in _get_media_thumbnail_selectors():
        try:
            primary_counts.append(await page.locator(sel).count())
        except Exception:
            continue
    primary = max(primary_counts or [0])
    if primary > 0:
        return primary

    fallback_counts: list[int] = []
    for sel in _get_media_thumbnail_fallback_selectors():
        try:
            fallback_counts.append(await page.locator(sel).count())
        except Exception:
            continue
    return max(fallback_counts or [0])


async def _read_pending_photo_uploads(page) -> dict | None:
    """Read visible, media-scoped upload progress from the live Etsy DOM.

    Etsy can render the thumbnail before its upload has finished.  This read
    intentionally looks only below the photo/media surface and only at visible
    alert/status/progress evidence, so a matching thumbnail count cannot be
    mistaken for a settled upload.
    """
    script = r"""
        () => {
            const isVisible = (node) => {
                if (!node || !node.isConnected) return false;
                const style = window.getComputedStyle(node);
                return node.getClientRects().length > 0
                    && style.display !== 'none'
                    && style.visibility !== 'hidden'
                    && style.opacity !== '0';
            };

            const anchors = Array.from(document.querySelectorAll(
                'input[name="listing-media-upload"], '
                + 'input[type="file"][accept*="image" i], '
                + '[data-testid*="listing-media" i], '
                + '[data-testid*="photo" i], '
                + '[class*="photo" i], [class*="media" i]'
            ));
            const roots = new Set();
            for (const anchor of anchors) {
                roots.add(anchor);
                const root = anchor.closest(
                    'section, form, [role="group"], '
                    + '[data-testid*="listing-media" i], '
                    + '[data-testid*="photo" i], [class*="photo" i], '
                    + '[class*="media" i]'
                );
                if (root && root !== document.body && root !== document.documentElement) {
                    roots.add(root);
                }
            }

            const mediaRoots = Array.from(roots).filter(isVisible);
            const inMediaSurface = (node) => mediaRoots.some((root) =>
                root === node || root.contains(node) || node.contains(root)
            );
            const candidates = new Set();
            const candidateSelector = [
                '[role="alert"]', '[role="status"]', '[aria-live]',
                '[role="progressbar"]', 'progress', '[aria-busy="true"]',
                '[data-state="loading"]', '[data-state="uploading"]',
                '[data-state="processing"]', '[class*="upload" i]',
                '[class*="processing" i]', '[class*="loading" i]',
                '[class*="progress" i]'
            ].join(', ');
            for (const root of mediaRoots) {
                if (root.matches(candidateSelector)) candidates.add(root);
                for (const node of root.querySelectorAll(candidateSelector + ', span, p')) {
                    if (node.matches('span, p')) {
                        const text = (node.textContent || '').replace(/\s+/g, ' ').trim();
                        if (!/^image\s+is\s+uploading\b/i.test(text)
                            && !/\b(?:uploading|processing)\b/i.test(text)) continue;
                    }
                    candidates.add(node);
                }
            }

            const pendingCandidates = [];
            for (const node of candidates) {
                if (!isVisible(node) || !inMediaSurface(node)) continue;
                const text = [
                    node.textContent || '',
                    node.getAttribute('aria-label') || '',
                    node.getAttribute('title') || '',
                    node.getAttribute('data-state') || ''
                ].join(' ').replace(/\s+/g, ' ').trim();
                const progressNow = Number(node.getAttribute('aria-valuenow'));
                const progressMax = Number(node.getAttribute('aria-valuemax'));
                const completeProgress = Number.isFinite(progressNow)
                    && Number.isFinite(progressMax) && progressMax > 0
                    && progressNow >= progressMax;
                const state = String(node.getAttribute('data-state') || '').toLowerCase();
                const isPending = !completeProgress
                    && !/^(complete|completed|success)$/.test(state)
                    && (
                        /^image\s+is\s+uploading\b/i.test(text)
                        || /\b(?:uploading|processing)\b/i.test(text)
                        || node.getAttribute('aria-busy') === 'true'
                        || node.matches('progress, [role="progressbar"]')
                        || /^(loading|uploading|processing)$/.test(state)
                );
                if (!isPending) continue;
                pendingCandidates.push(node);
            }

            // Prefer the most specific visible evidence to avoid counting a
            // wrapper and its nested alert as two separate uploads.
            const pending = pendingCandidates.filter((node) =>
                !pendingCandidates.some((child) => child !== node && node.contains(child))
            );

            return {
                readable: mediaRoots.length > 0,
                pending: pending.length > 0,
                pendingCount: pending.length,
                evidence: pending.slice(0, 5).map((node) =>
                    (node.textContent || node.getAttribute('aria-label') || '')
                        .replace(/\s+/g, ' ').trim()
                )
            };
        }
    """
    try:
        state = await page.evaluate(script)
    except Exception:
        return None
    if not isinstance(state, dict):
        return None
    if not isinstance(state.get("readable"), bool) or not isinstance(state.get("pending"), bool):
        return None
    return state


async def _wait_for_expected_image_count(page, expected_count: int, exact: bool = False,
                                        timeout_ms: int = PHOTO_UPLOAD_WAIT_MS,
                                        log_progress: bool = False) -> bool:
    checks = max(1, timeout_ms // PHOTO_UPLOAD_POLL_MS)
    last_count = -1
    for _ in range(checks):
        try:
            current = await _count_listing_image_thumbs(page)
        except Exception:
            current = 0

        pending_state = await _read_pending_photo_uploads(page)
        # A count is not a readiness signal while Etsy still shows a visible
        # upload/processing alert or progress indicator. If the live DOM read
        # cannot be verified, fail closed instead of risking a duplicate batch.
        settled = bool(pending_state and pending_state.get("readable") and not pending_state.get("pending"))

        if exact:
            if current == expected_count and settled:
                return True
        elif current >= expected_count and settled:
            return True

        if current != last_count:
            if log_progress and current >= 0:
                print(f"  ⏳ Ảnh trên UI: {current}/{expected_count}")
            last_count = current
        await page.wait_for_timeout(PHOTO_UPLOAD_POLL_MS)

    return False


_EXACT_ADD_PHOTOS_INPUT_SELECTORS = (
    # Prefer an input nested in the exact accessible photo label when Etsy
    # renders the label as the input's parent.
    'label[aria-label="Add photos"] input[name="listing-media-upload"]',
    'label[aria-label="Add photos"] input[type="file"]',
    # Current Etsy editor markup may associate the input through label[for]
    # instead of nesting it. XPath keeps this association exact and avoids
    # selecting the first same-name input, which is the video surface.
    'xpath=//input[@name="listing-media-upload" and @id=//label[@aria-label="Add photos"]/@for]',
    'xpath=//input[@type="file" and @id=//label[@aria-label="Add photos"]/@for]',
    'xpath=//input[@name="listing-media-upload" and @id=//label[normalize-space(.)="Add photos"]/@for]',
    'xpath=//input[@type="file" and @id=//label[normalize-space(.)="Add photos"]/@for]',
    # Also cover exact visible text and sibling/nested label layouts without
    # broad text matching such as :has-text("Add photo").
    'xpath=//label[@aria-label="Add photos"]/following-sibling::input[@name="listing-media-upload"][1]',
    'xpath=//label[@aria-label="Add photos"]/following-sibling::input[@type="file"][1]',
    'xpath=//label[normalize-space(.)="Add photos"]/following-sibling::input[@name="listing-media-upload"][1]',
    'xpath=//label[normalize-space(.)="Add photos"]/following-sibling::input[@type="file"][1]',
    'xpath=//label[@aria-label="Add photos"]//input[@type="file"]',
    'xpath=//label[normalize-space(.)="Add photos"]//input[@type="file"]',
)


async def _find_exact_add_photos_input(page):
    """Find only the file input explicitly associated with Add photos."""
    for selector in _EXACT_ADD_PHOTOS_INPUT_SELECTORS:
        try:
            candidate = page.locator(selector).first
            if await candidate.count() > 0:
                return candidate, selector
        except Exception:
            continue
    return None, None


async def _upload_listing_photos(page, paths: list[str]) -> None:
    """Upload listing photos using Etsy listing-media input + file-chooser fallbacks."""
    if not paths:
        return

    exact_photo_input, exact_photo_selector = await _find_exact_add_photos_input(page)
    if exact_photo_input is not None:
        try:
            await exact_photo_input.wait_for(state="attached", timeout=15000)
            await exact_photo_input.set_input_files(paths, timeout=60000)
            return
        except Exception as direct_err:
            # Never fall through to a same-name broad selector after an exact
            # Add photos input was found: that fallback can target Add videos.
            raise RuntimeError(
                f"Input Add photos tồn tại nhưng upload ảnh thất bại ({exact_photo_selector}): {direct_err}"
            ) from direct_err

    upload_selectors = [
        '[data-testid="empty-photo-thumbnail"] input[name="listing-media-upload"]',
        'input[name="listing-media-upload"]',
        'input[type="file"][accept*="image"]',
        'input[type="file"][accept*="jpeg"]',
        'label[for="listing-photos"] ~ * input[type="file"]',
        'input[type="file"]',
    ]
    for sel in upload_selectors:
        fi = page.locator(sel).first
        if await fi.count() == 0:
            continue
        try:
            await fi.wait_for(state="attached", timeout=15000)
            await fi.set_input_files(paths, timeout=60000)
            return
        except Exception as direct_err:
            print(f"  ⚠️ Input upload ảnh thất bại ({sel}): {direct_err}")

    add_photo_selectors = [
        'button:has-text("Add photos")',
        'button:has-text("Add photo")',
        'button:has-text("Upload photos")',
        'button:has-text("Upload photo")',
        'button[aria-label*="Add photo" i]',
        'button[aria-label*="Upload photo" i]',
        '[role="button"]:has-text("Add photos")',
        '[role="button"]:has-text("Add photo")',
        'label:has-text("Add photos")',
        'label:has-text("Add photo")',
    ]
    for sel in add_photo_selectors:
        add_btn = page.locator(sel).first
        if await add_btn.count() == 0:
            continue
        try:
            if not await add_btn.is_visible():
                continue
            await add_btn.scroll_into_view_if_needed()
            async with page.expect_file_chooser(timeout=15000) as fc_info:
                await add_btn.click()
            await fc_info.value.set_files(paths)
            return
        except Exception as chooser_err:
            print(f"  ⚠️ File chooser ảnh thất bại ({sel}): {chooser_err}")

    fi = page.locator('input[type="file"]').last
    if await fi.count() > 0:
        await fi.set_input_files(paths, timeout=60000)
        return

    raise RuntimeError("Không tìm thấy nút/input upload ảnh phù hợp trên form Etsy.")


async def _upload_listing_photos_until_count(
    page,
    paths: list[str],
    *,
    expected_total: int,
    exact: bool = True,
    batch_size: int = PHOTO_UPLOAD_BATCH_SIZE,
    timeout_ms: int = PHOTO_UPLOAD_WAIT_MS,
) -> int:
    """Upload unique paths in bounded batches until the UI count matches."""
    unique_paths = []
    seen_paths = set()
    for path in (paths or []):
        path = str(path) if path else ""
        if path and path not in seen_paths:
            seen_paths.add(path)
            unique_paths.append(path)
    paths = unique_paths[:10]
    if not paths:
        return await _count_listing_image_thumbs(page)

    expected_total = max(1, min(int(expected_total), 10))
    batch_size = max(1, min(int(batch_size), 10))

    submitted_paths: set[str] = set()
    next_path = 0
    while next_path < len(paths):
        before = await _count_listing_image_thumbs(page)
        if before >= expected_total:
            break

        needed = expected_total - before
        batch = paths[next_path:next_path + min(batch_size, needed)]
        batch = [path for path in batch if path not in submitted_paths]
        if not batch:
            break
        batch_target = min(expected_total, before + len(batch))
        print(f"  📤 Upload ảnh đợt {next_path // batch_size + 1}: {len(batch)} file (UI {before} → mục tiêu {batch_target})")
        await _upload_listing_photos(page, batch)
        submitted_paths.update(batch)
        next_path += len(batch)
        settled = await _wait_for_expected_image_count(
            page,
            expected_count=batch_target,
            exact=False,
            timeout_ms=timeout_ms,
            log_progress=True,
        )
        if not settled:
            pending_state = await _read_pending_photo_uploads(page)
            if pending_state is None or not pending_state.get("readable"):
                raise RuntimeError(
                    "Không xác minh được trạng thái settle của upload ảnh sau "
                    f"{timeout_ms / 1000:.0f}s; dừng trước khi upload bổ sung."
                )
            if pending_state.get("pending"):
                evidence = ", ".join(pending_state.get("evidence") or [])
                raise RuntimeError(
                    "Etsy vẫn còn ảnh đang upload/processing sau "
                    f"{timeout_ms / 1000:.0f}s ({pending_state.get('pendingCount', 0)} pending"
                    f"{': ' + evidence if evidence else ''}); "
                    "dừng, không upload trùng path."
                )
            # Count is settled but Etsy did not accept the full batch. Continue
            # only with paths never submitted; never replay the same batch.
            print("  ⚠️ Batch đã settle nhưng số ảnh còn thiếu; chỉ dùng path chưa gửi.")

    current = await _count_listing_image_thumbs(page)

    if exact:
        ok = current == expected_total
    else:
        ok = current >= expected_total
    if not ok:
        raise RuntimeError(
            f"Số ảnh trên UI chưa đạt kỳ vọng. Expected_count={expected_total}"
            f", actual={current}, exact={exact}."
        )

    print(f"  📷 {current} ảnh ✓")
    return current


def _pick_draft_card_id(cards: list[dict], product: dict) -> str | None:
    """
    Chọn listing ID an toàn:
      1) khớp SKU exact nếu có,
      2) khớp title exact đã normalize,
      3) chỉ fuzzy nhất (không ambiguous), tỷ lệ cao.
    """
    target_title = normalize_search_text(product.get("title", ""))
    target_sku = normalize_sku(product.get("sku"))
    if not cards:
        return None

    draft_cards = []
    for c in cards:
        status = (c.get("status") or "").lower()
        if status and status != "draft":
            continue
        draft_cards.append(c)

    # Ưu tiên khớp SKU chính xác
    if target_sku:
        sku_exact = [c for c in draft_cards if normalize_sku(c.get("sku", "")) == target_sku]
        if len(sku_exact) == 1:
            return str(sku_exact[0]["id"])
        if len(sku_exact) > 1:
            print(f"  ⚠️ Trùng lặp SKU trong draft theo index ({len(sku_exact)} kết quả): {target_sku}")
            return None

    # Khớp title exact sau khi normalize
    if target_title:
        exact_title = [c for c in draft_cards if normalize_search_text(c.get("title", "")) == target_title]
        if len(exact_title) == 1:
            return str(exact_title[0]["id"])
        if len(exact_title) > 1:
            print(f"  ⚠️ Có {len(exact_title)} draft khớp tiêu đề exact; tránh tự động chọn để không sai.")
            return None

    # Fuzzy khớp conservatively, chỉ nhận khi best đủ áp đảo runner-up.
    best_match = None
    best_ratio = 0.0
    best_ratio_second = 0.0
    for c in draft_cards:
        title_norm = normalize_search_text(c.get("title", ""))
        if not title_norm or not target_title:
            continue
        ratio = SequenceMatcher(None, target_title, title_norm).ratio()
        if ratio <= 0.92:
            continue
        if ratio > best_ratio:
            best_ratio_second = best_ratio
            best_ratio = ratio
            best_match = c
        elif ratio > best_ratio_second:
            best_ratio_second = ratio

    if best_match is None:
        return None

    if (best_ratio - best_ratio_second) < 0.03:
        print(
            f"  ⚠️ Fuzzy khớp tiêu đề không đủ khoảng cách an toàn (best={best_ratio:.2f}, runner-up={best_ratio_second:.2f}); bỏ qua xác minh URL."
        )
        return None

    return str(best_match["id"])


async def _editor_product_signature_matches(page, product: dict) -> bool:
    """Chỉ tin ID trên URL nếu title/SKU trong editor khớp sản phẩm."""
    target_title = normalize_search_text(product.get("title", ""))
    target_sku = normalize_sku(product.get("sku"))
    try:
        sig = await page.evaluate(r'''() => {
            const titleEl = document.querySelector('#listing-title-input, textarea[name="title"]');
            const skuEl = document.querySelector('#listing-sku-input, input[name="sku"], [data-testid="sku-input"]');
            return {
                title: (titleEl && (titleEl.value || titleEl.innerText) || "").trim(),
                sku: (skuEl && skuEl.value || "").trim()
            };
        }''')
        editor_title = normalize_search_text(sig.get("title", ""))
        editor_sku = normalize_sku(sig.get("sku", ""))
        if not editor_title or editor_title != target_title:
            return False
        if target_sku:
            return editor_sku == target_sku and bool(editor_sku)
        return True
    except Exception:
        return False


async def _collect_draft_cards(page):
    """Lấy raw card data từ Drafts list sau khi đã ép filter."""
    await _force_draft_filter(page)
    await _wait_for_draft_grid_stable(page)
    await _ensure_draft_filter_after_grid(page)
    cards = await page.evaluate(r'''() => {
        let items = [];
        let anchors = Array.from(document.querySelectorAll('a[href*="/listing-editor/edit/"]'));
        for (let a of anchors) {
            let href = a.getAttribute('href') || '';
            let match = href.match(/\/listing-editor\/edit\/(\d+)/);
            if (!match) continue;
            let id = match[1];
            let row = a.closest('tr') || a.closest('[class*="card"]') || a.closest('div');
            let title = (a.innerText || '').trim();
            if (!title && row) {
                let titleEl = row.querySelector('[class*="title"], [class*="name"], h3, h4, [data-testid*="title"]');
                if (titleEl) title = (titleEl.innerText || '').trim();
            }
            if (!title && row) {
                let textNodes = Array.from(row.querySelectorAll('a,span,div')).map(el => (el.innerText || '').trim()).filter(Boolean);
                if (textNodes.length > 0) title = textNodes[0];
            }
            let rowText = row ? (row.innerText || '').toLowerCase() : '';
            let sku = '';
            if (rowText) {
                let skuMatch = rowText.match(/\bsku\b\s*[:#]?\s*([a-z0-9][a-z0-9_-]{2,})/i);
                if (skuMatch) sku = skuMatch[1];
            }
            if (!id) continue;
            items.push({ id: id, title: title || "", sku: sku || "", status: 'draft' });
        }
        return items;
    }''')
    return _dedupe_draft_cards(cards)


async def _wait_for_draft_grid_stable(
    page,
    max_wait_ms: int = 15000,
    pause_ms: int = 250,
    min_settle_ms: int = 400,
    stable_repeats: int = 6,
    empty_stable_repeats: int = 3,
    min_stable_wait_ms: int = 1500,
) -> None:
    """
    Chờ danh sách draft ổn định sau khi bật filter:
    - không loading
    - hoặc marker rỗng draft rõ ràng
    - hoặc tập ID không đổi liên tiếp cho danh sách không rỗng.
    """
    await page.wait_for_timeout(min_settle_ms)
    end_at = asyncio.get_running_loop().time() + max_wait_ms / 1000

    last_id_tuple = None
    stable_count = 0
    empty_count = 0
    min_stable_until = asyncio.get_running_loop().time() + min_stable_wait_ms / 1000
    state_script = r'''() => {
        const __etsyDraftGridStateMarker = true;
        const anchors = Array.from(document.querySelectorAll('a[href*="/listing-editor/edit/"]'));
        const re = /\/listing-editor\/edit\/(\d+)/;
        const ids = Array.from(new Set(
            anchors.map((a) => {
                const href = a.getAttribute("href") || "";
                const match = href.match(re);
                return match && match[1] ? match[1] : "";
            }).filter(Boolean)
        )).sort();

        const loadingNodes = Array.from(document.querySelectorAll(
            '[aria-busy="true"], .wt-spinner, .wt-loading, .loading-spinner, [data-state="loading"]'
        ));
        const loading = loadingNodes.some((n) => {
            const style = window.getComputedStyle(n);
            return n && n.offsetParent !== null && style && style.visibility !== "hidden" && style.display !== "none";
        });

        const emptyStateMarkers = Array.from(document.querySelectorAll('.wt-empty-state, [role="status"], [data-testid*="empty"]'));
        const emptyState = emptyStateMarkers.some((el) => {
            if (!el || el.offsetParent === null) return false;
            const text = (el.textContent || "").toLowerCase();
            return (
                text.includes("no listings") ||
                text.includes("no results") ||
                text.includes("you have no drafts") ||
                text.includes("nothing to show")
            );
        });

        return { loading, ids, emptyState };
    }'''

    while True:
        if asyncio.get_running_loop().time() > end_at:
            raise RuntimeError(
                "Quá thời gian chờ draft grid ổn định. "
                "Không thể xác nhận filter Draft an toàn trước khi quét listings."
            )

        try:
            state = await page.evaluate(state_script)
        except Exception:
            state = {"loading": True, "ids": [], "emptyState": False}

        ids = state.get("ids", [])
        if not isinstance(ids, list):
            ids = list(ids) if ids is not None else []
        id_tuple = tuple(sorted({str(i).strip() for i in ids if str(i).strip()}))
        loading = bool(state.get("loading", True))
        empty_state = bool(state.get("emptyState", False))
        now = asyncio.get_running_loop().time()

        if not loading:
            if id_tuple:
                if last_id_tuple is not None and id_tuple == last_id_tuple:
                    stable_count += 1
                    empty_count = 0
                    if stable_count >= stable_repeats and now >= min_stable_until:
                        return
                else:
                    stable_count = 1
                    last_id_tuple = id_tuple
                    empty_count = 0
            elif empty_state:
                empty_count += 1
                if empty_count >= empty_stable_repeats and now >= min_stable_until:
                    return
            else:
                empty_count = 0
                last_id_tuple = None
                stable_count = 0
        else:
            last_id_tuple = None
            stable_count = 0
            empty_count = 0

        await page.wait_for_timeout(pause_ms)


def _get_save_button_selector(explicit_edit: bool = False) -> str:
    """Return the appropriate save button selector for create vs explicit edit flow."""
    if explicit_edit:
        return (
            'button:has-text("Save changes"):visible, '
            'button:has-text("Save as draft"):visible, '
            'button:has-text("Save Draft"):visible, '
            'button:has-text("Save draft"):visible'
        )

    return (
        'button:has-text("Save draft"):visible, '
        'button:has-text("Save as draft"):visible, '
        'button:has-text("Save Draft"):visible'
    )
# ── Translate ──────────────────────────────────────────────────────────────────
def translate_text(text: str, lang: str, max_chars=4800) -> str:
    if not text: return ""
    try:
        if len(text) <= max_chars:
            return GoogleTranslator(source="en", target=lang).translate(text) or text
        parts, chunk = [], ""
        for para in text.split("\n\n"):
            if len(chunk) + len(para) + 2 > max_chars:
                if chunk:
                    parts.append(GoogleTranslator(source="en", target=lang).translate(chunk) or chunk)
                chunk = para
            else:
                chunk = (chunk + "\n\n" + para).strip()
        if chunk:
            parts.append(GoogleTranslator(source="en", target=lang).translate(chunk) or chunk)
        return "\n\n".join(parts)
    except Exception as e:
        print(f"    ⚠ dịch {lang}: {e}")
        return text

def translate_tag(tag: str, lang: str) -> str:
    """Dịch 1 tag, giữ ≤20 ký tự."""
    try:
        translated = GoogleTranslator(source="en", target=lang).translate(tag) or tag
        translated = translated.strip()
        if len(translated) > 20:
            short = translated[:20].rsplit(" ", 1)[0].rstrip(",- ")
            if 1 <= len(short) <= 20:
                translated = short
            else:
                translated = translated[:20].rstrip(",- ")
        return translated if 1 <= len(translated) <= 20 else tag[:20].rstrip(",- ")
    except Exception:
        return tag[:20].rstrip(",- ")

def get_sku_prefix(shop_id: str) -> str:
    shop_id_lower = str(shop_id or "templystudios").lower()
    if "temply" in shop_id_lower:
        return "TS"
    elif "daisy" in shop_id_lower:
        return "dd"
    else:
        return str(shop_id or "TS")[:2].upper()

def generate_sku(shop_id: str, folder_name: str) -> str:
    prefix = get_sku_prefix(shop_id)
    import re
    clean_folder = "".join(c if c.isalnum() else "_" for c in folder_name).lower()
    clean_folder = re.sub(r'_+', '_', clean_folder).strip('_')
    return f"{prefix}_{clean_folder}"


def _normalize_requested_products(raw_products: list[str] | None) -> list[str]:
    """Chuẩn hoá --products: hỗ trợ repeat + comma-separated; giữ đúng thứ tự, bỏ trùng."""
    if not raw_products:
        return []

    def _normalize_one_folder(raw: str) -> str:
        folder = str(raw).strip()
        if not folder:
            raise ValueError("folder trống không hợp lệ")
        if not re.fullmatch(r"product-\d+", folder):
            raise ValueError(f"Định dạng folder không hợp lệ: {folder}")
        return folder

    normalized = []
    seen = set()
    for raw in raw_products:
        if raw is None:
            continue
        items = str(raw).split(",")
        for item in items:
            folder = _normalize_one_folder(item)
            if not folder:
                continue
            if folder in seen:
                continue
            normalized.append(folder)
            seen.add(folder)
    return normalized


def _normalize_selected_products(raw_items: list[str] | None) -> list[tuple[int, str]]:
    if not raw_items:
        return []
    normalized = []
    seen_rows = set()
    seen_folders = set()
    for raw in raw_items:
        match = re.fullmatch(r"(\d+):(product-\d+)", str(raw or "").strip())
        if not match:
            raise ValueError(f"Cặp row:folder không hợp lệ: {raw}")
        row = int(match.group(1))
        folder = match.group(2)
        if row < 4:
            raise ValueError(f"Row không hợp lệ: {row}")
        if row in seen_rows or folder in seen_folders:
            raise ValueError(f"Cặp row/folder bị lặp: {raw}")
        normalized.append((row, folder))
        seen_rows.add(row)
        seen_folders.add(folder)
    return normalized


_LISTING_REFERENCE_RE = re.compile(r"/(?:listing/|listing-editor/edit/)(\d+)")


def _listing_reference_has_id(value: object) -> bool:
    text = str(value or "").strip()
    return bool(_LISTING_REFERENCE_RE.search(text) or re.fullmatch(r"\d+", text))


# ── Read Excel ─────────────────────────────────────────────────────────────────
def read_products(
    batch=DEFAULT_BATCH,
    skip=0,
    product_folder=None,
    shop_id="templystudios",
    product_folders: list[str] | None = None,
    selected_products: list[tuple[int, str]] | None = None,
    store: CloudAssetStore | None = None,
):
    selected_products = list(selected_products or [])
    if product_folders is None:
        product_folders = []
    if product_folder:
        product_folders = [str(product_folder)] + product_folders
    requested_folders = _normalize_requested_products(product_folders)
    selected_by_row = {row: folder for row, folder in selected_products}
    selected_order = {row: index for index, (row, _) in enumerate(selected_products)}
    requested_set = set(requested_folders)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Listings"]
    all_products = []
    max_row = max(ws.max_row, 4)
    for row_num, row in enumerate(ws.iter_rows(min_row=4, max_row=max_row, values_only=True), start=4):
        cols = (list(row) + [None]*19)[:19]
        _, folder, _, _, price, category, _, title, description, tags, qty, _, when_made, status, section = cols[:15]
        etsy_reference = cols[15] if len(cols) > 15 else ""
        sku = cols[17] if len(cols) > 17 else ""
        if not folder or not title or str(title).startswith("←"):
            continue
        if selected_products:
            expected_folder = selected_by_row.get(row_num)
            if expected_folder is None:
                continue
            if str(folder) != expected_folder:
                raise RuntimeError(
                    f"❌ Row {row_num} đã đổi folder từ {expected_folder} thành {folder}, dừng để tránh đăng nhầm"
                )
        # If targeting a specific product, ignore status filter for that product
        if selected_products or requested_folders:
            if str(folder) not in requested_set:
                if not selected_products:
                    continue
            if status and "Đã đăng" in str(status):
                raise RuntimeError(
                    f"❌ {folder} đã có trạng thái '{status}', dừng để tránh đăng trùng"
                )
            if _listing_reference_has_id(etsy_reference):
                raise RuntimeError(
                    f"❌ {folder} đã có Etsy listing ID/URL, dừng để tránh đăng trùng"
                )
        elif status and "Đã đăng" in str(status):
            continue

        # Auto-generate default SKU if empty
        if not sku or not str(sku).strip():
            sku = generate_sku(shop_id, str(folder))

        clean_title = trim_title(str(title))
        clean_keywords = str(cols[2] or "")  # column C = keywords

        all_products.append({
            "folder":      folder,
            "title":       clean_title,
            "description": str(description or ""),
            "price":       float(str(price)) if price is not None and str(price).strip() else 4.99,
            "category":    str(category or ""),
            "tags":        str(tags or ""),
            "qty":         int(float(str(qty))) if qty is not None and str(qty).strip() else 999,
            "when_made":   str(when_made) if when_made else "2020_2026",
            "section":     str(section).strip() if section else "",
            "sku":         str(sku).strip(),
            "image_paths": [],
            "pdf_paths":   [],
            "row":         row_num,
            "keywords":    clean_keywords,
            "alt_texts":   [],
        })

    if selected_products:
        all_products = sorted(all_products, key=lambda item: selected_order[item["row"]])
    elif requested_folders:
        folder_index = {folder: idx for idx, folder in enumerate(requested_folders)}
        all_products = sorted(all_products, key=lambda item: folder_index[item["folder"]])
    else:
        all_products = all_products[skip: skip + batch]

    # Resolve only products that will actually reach the browser. This keeps
    # batch/skip selection from contacting cloud storage for unrelated rows,
    # while still failing closed before any Etsy navigation.
    for product in all_products:
        resolution = resolve_product_asset_paths(product, shop_id, store=store)
        product["_cloud_asset_resolution"] = resolution
        product["alt_texts"] = generate_alt_texts(
            product["title"], product["keywords"], len(product["image_paths"])
        )

    return all_products, wb, ws, len(all_products)

def save_status(wb, ws, row, text, url=None):
    ws.cell(row=row, column=14, value=text)
    if url:
        ws.cell(row=row, column=16, value=url)
    wb.save(EXCEL_FILE)

# ── Helpers ────────────────────────────────────────────────────────────────────
async def click_tab(page, *names):
    """Click tab theo tên (thử nhiều selector)."""
    for name in names:
        for sel in [
            f'[role="tab"]:has-text("{name}")',
            f'button:has-text("{name}")',
            f'a:has-text("{name}")',
            f'li:has-text("{name}")',
            f'[class*="tab"]:has-text("{name}")',
        ]:
            el = page.locator(sel).first
            if await el.count() > 0 and await el.is_visible():
                await el.click()
                await page.wait_for_timeout(1500)
                return True
    return False

async def detect_form_type(page):
    """Phát hiện form Etsy dạng tab mới hay trang dài cũ."""
    new_form_selectors = [
        'a:has-text("Item Details")',
        'a:has-text("Pricing & Shipping")',
        'button:has-text("Item Details")',
        '[href*="details"]',
        '[href*="pricing"]',
    ]
    for sel in new_form_selectors:
        if await page.locator(sel).count() > 0:
            return "tabs"
    if await page.locator('#listing-title-input, textarea[name="title"]').count() > 0:
        return "single"
    return "tabs"

async def smart_fill(page, selector, value, timeout=6000):
    """Điền giá trị vào input/textarea, hỗ trợ React."""
    try:
        el = page.locator(selector).first
        await el.wait_for(state="visible", timeout=timeout)
        await el.scroll_into_view_if_needed()
        await page.wait_for_timeout(300)
        await el.click(click_count=3)
        await page.wait_for_timeout(200)
        await el.fill(str(value))
        await page.wait_for_timeout(300)
        return True
    except Exception as e:
        return False

async def dismiss_alerts(page):
    """Đóng các popup/alert của Etsy nếu có."""
    try:
        for sel in [
            '[role="alert"] button[aria-label*="close" i]',
            '[role="alert"] button[aria-label*="dismiss" i]',
            '.wt-alert button',
            '[data-clg-id="WtAlert"] button',
        ]:
            btns = page.locator(sel)
            cnt = await btns.count()
            for i in range(cnt):
                try:
                    btn = btns.nth(i)
                    if await btn.is_visible():
                        await btn.click()
                        await page.wait_for_timeout(500)
                except Exception:
                    pass
        await page.wait_for_timeout(600)
    except Exception:
        pass

async def clear_and_fill(el, value):
    """Clear và điền giá trị — tương thích mọi phiên bản Playwright."""
    await el.scroll_into_view_if_needed()
    await el.click(click_count=3)
    await el.fill(value)

# ── Fill alt text cho từng ảnh ────────────────────────────────────────────────
# ── Visual Alt Text Generator ──────────────────────────────────────────────────
def generate_visual_alt_text(img_path, title, keywords):
    """
    Sử dụng Gemini 2.5 Flash thông qua Vertex AI để phân tích hình ảnh và sinh Alt Text chuẩn SEO.
    Nếu thất bại hoặc không có thư viện, trả về None để dùng cơ chế fallback dựa trên text.
    """
    if not has_genai:
        print("      ⚠️ Không có thư viện google-genai. Sử dụng Text SEO fallback.")
        return None

    try:
        # Đọc ảnh dưới dạng bytes
        with open(img_path, "rb") as f:
            img_bytes = f.read()

        # Xác định mime_type
        suffix = Path(img_path).suffix.lower()
        mime_type = "image/png"
        if suffix in [".jpg", ".jpeg"]:
            mime_type = "image/jpeg"
        elif suffix == ".webp":
            mime_type = "image/webp"
        elif suffix == ".gif":
            mime_type = "image/gif"

        # Khởi tạo client Vertex AI
        client = genai.Client(vertexai=True, project="temply-ai-lab", location="us-central1")

        prompt = f"""
Analyze the provided product image and generate a highly descriptive, SEO-optimized Alt Text for Etsy.

Instructions:
1. Describe the layout, visual design, text, and specific pages or features shown in the image in detail.
2. Incorporate 1-2 relevant SEO keywords from the following list: {keywords}.
3. The description must be natural, engaging, and extremely helpful for visually impaired buyers.
4. Keep the length strictly under 250 characters.
5. Do NOT use generic phrases like "image of" or "screenshot of".
6. Output ONLY the raw alt text, no quotes, no markdown, no comments, just the plain text.

Listing Title: {title}
"""
        part = types.Part.from_bytes(
            data=img_bytes,
            mime_type=mime_type
        )

        import time
        max_retries = 3
        backoff_sec = 3
        response = None
        for attempt in range(max_retries + 1):
            try:
                response = client.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=[part, prompt]
                )
                break
            except Exception as e:
                err_msg = str(e).lower()
                if ("429" in err_msg or "resource exhausted" in err_msg or "resource_exhausted" in err_msg) and attempt < max_retries:
                    print(f"      ⚠️ Gặp lỗi 429 (Resource Exhausted). Đợi {backoff_sec}s và thử lại #{attempt + 1}...")
                    time.sleep(backoff_sec)
                    backoff_sec *= 2
                else:
                    raise e

        text = (response.text or "").strip()
        if text.startswith('"') and text.endswith('"'):
            text = text[1:-1].strip()
        if text.startswith("'") and text.endswith("'"):
            text = text[1:-1].strip()

        if len(text) > 250:
            text = text[:247] + "..."

        return text
    except Exception as e:
        print(f"      ❌ Lỗi khi phân tích ảnh qua Gemini: {e}")
        return None

# ── Fill alt text cho từng ảnh ────────────────────────────────────────────────
async def fill_image_alt_texts(page, product):
    """Điền alt text cho từng ảnh sau khi upload lên Etsy."""
    title = product.get("title", "")
    keywords = product.get("keywords", "")
    image_paths = product.get("image_paths", [])

    print(f"  ✍️ Bắt đầu điền Alt Text cho các ảnh...")
    await page.wait_for_timeout(3000)  # Đợi ảnh load/upload hoàn tất

    # Định vị các nút thumbnail hình vuông đại diện cho từng ảnh trên lưới
    thumb_btns = page.locator('button.le-aspect-ratio--square')
    cnt = await thumb_btns.count()
    print(f"    🔍 Tìm thấy {cnt} ảnh trên giao diện Etsy.")

    if cnt == 0:
        print("    ⚠️ Không tìm thấy ảnh nào trên giao diện để điền Alt Text.")
        return

    # Sinh danh sách Alt Text dạng văn bản dự phòng (fallback)
    fallback_alts = generate_alt_texts(title, keywords, cnt)
    filled = 0

    for i in range(cnt):
        print(f"\n    🎬 Xử lý ảnh #{i+1}...")
        btn = thumb_btns.nth(i)
        
        try:
            # 1. Hover và click thumbnail
            await btn.scroll_into_view_if_needed()
            await page.wait_for_timeout(400)
            await btn.hover()
            await page.wait_for_timeout(400)
            await btn.click(force=True)
            await page.wait_for_timeout(2000)

            # 2. Đợi nút "+ Alt text" hoặc "Alt text" trong dialog trung gian xuất hiện và click
            alt_btn = page.locator('div[role="dialog"] button:has-text("Alt text"), button:has-text("Alt text"), button:has-text("+ Alt text")').first
            try:
                await alt_btn.wait_for(state="visible", timeout=3000)
                await alt_btn.click()
                await page.wait_for_timeout(1500)
            except Exception as alt_btn_ex:
                print(f"      ❌ Không tìm thấy nút mở Alt Text trong dialog: {alt_btn_ex}")
                try:
                    await page.keyboard.press("Escape")
                except:
                    pass
                await page.wait_for_timeout(1000)
                continue

            # 3. Tạo nội dung Alt Text (Thử Gemini trước, nếu lỗi/thiếu ảnh thì dùng Fallback)
            alt_val = None
            if i < len(image_paths):
                img_path = image_paths[i]
                print(f"      📸 Đang phân tích ảnh cục bộ qua Gemini: {Path(img_path).name}...")
                try:
                    alt_val = await asyncio.wait_for(
                        asyncio.to_thread(generate_visual_alt_text, img_path, title, keywords),
                        timeout=15.0
                    )
                except asyncio.TimeoutError:
                    print("      ⚠️ Gemini API call timed out after 15 seconds. Using Text SEO fallback.")
                    alt_val = None
                except Exception as g_err:
                    print(f"      ⚠️ Lỗi khi gọi Gemini: {g_err}. Using Text SEO fallback.")
                    alt_val = None
            
            if not alt_val:
                alt_val = fallback_alts[i]
                print(f"      ⚠️ Sử dụng Alt Text văn bản SEO (fallback): '{alt_val[:50]}...'")
            else:
                print(f"      ✨ Sinh Alt Text bằng Gemini thành công: '{alt_val}'")

            # 4. Tìm textarea trong Alt Text modal và điền
            alt_input = page.locator('div[role="dialog"]:visible textarea, div[class*="dialog"]:visible textarea, [class*="modal"]:visible textarea').first
            await alt_input.wait_for(state="visible", timeout=3000)
            await alt_input.click()
            await alt_input.fill(alt_val)
            await page.wait_for_timeout(500)

            # 5. Click Apply để lưu Alt Text hiện tại
            apply_btn = page.locator('div[role="dialog"]:visible button:has-text("Apply")').first
            await apply_btn.wait_for(state="visible", timeout=2000)
            await apply_btn.click()
            await page.wait_for_timeout(500)
            
            # Đợi modal điền Alt Text đóng hoàn toàn
            await alt_input.wait_for(state="hidden", timeout=5000)

            # 6. Click Done để lưu và đóng dialog ảnh hiện tại
            done_btn = page.locator('div[role="dialog"]:visible button:has-text("Done")').first
            await done_btn.wait_for(state="visible", timeout=2000)
            await done_btn.click(force=True)
            await page.wait_for_timeout(1000)
            
            print(f"      ✓ Đã điền và lưu Alt Text cho ảnh #{i+1} thành công!")
            filled += 1

        except Exception as e:
            print(f"      ❌ Lỗi khi xử lý ảnh #{i+1}: {e}")
            try:
                # Tránh kẹt modal: Nhấn Escape 2 lần để đóng các dialog đang mở
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(1000)
                await page.keyboard.press("Escape")
            except:
                pass
            await page.wait_for_timeout(1000)

    if filled > 0:
        print(f"  🖼️  Alt text: Đã điền thành công {filled}/{cnt} ảnh ✓")
    else:
        print("  ⚠️ Alt text: không điền được ảnh nào — bỏ qua")


# ── Tab: Photo & Video ─────────────────────────────────────────────────────────
async def fill_photo_tab(page, product, explicit_edit: bool = False):
    await click_tab(page, "Photo & Video", "Photos")
    await page.wait_for_timeout(800)
    paths = list(product.get("image_paths") or [])[:10]
    if not paths:
        return

    before = await _count_listing_image_thumbs(page)
    if not explicit_edit:
        expected_total = len(paths)
        exact_expected = True
    else:
        local_target = len(paths)
        expected_total = max(before, local_target)
        exact_expected = False

    await _upload_listing_photos_until_count(
        page,
        paths,
        expected_total=expected_total,
        exact=exact_expected,
    )

    # Điền alt text sau khi ảnh đã upload xong
    await fill_image_alt_texts(page, product)

# ── Tab: Category ──────────────────────────────────────────────────────────────
class DigitalListingTypeError(RuntimeError):
    """The Etsy editor could not be proven to be in Digital listing mode."""


async def _is_locator_visible(locator, *, visibility_timeout_ms=None) -> bool:
    if visibility_timeout_ms is None:
        return bool(await locator.is_visible())
    try:
        return bool(
            await asyncio.wait_for(
                locator.is_visible(),
                timeout=max(0, float(visibility_timeout_ms)) / 1000,
            )
        )
    except Exception:
        return False


async def _visible_locator_items(locator, *, visibility_timeout_ms=None):
    items = []
    for index in range(await locator.count()):
        item = locator.nth(index)
        if await _is_locator_visible(item, visibility_timeout_ms=visibility_timeout_ms):
            items.append(item)
    return items


_LISTING_TYPE_CONTROL_SELECTORS = (
    '#category-mixed-listing-type',
    '[id*="category-mixed-listing-type" i]',
    '[id*="mixed-listing-type" i]',
    '[data-testid*="listing-type" i][aria-haspopup]',
    '[data-testid*="physical-digital" i][aria-haspopup]',
    '[role="button"][aria-haspopup][aria-label]',
    '[role="combobox"][aria-haspopup][aria-label]',
    '[role="combobox"][aria-label*="physical" i][aria-label*="digital" i]',
    '[role="combobox"][aria-label*="listing" i][aria-label*="type" i]',
    '[role="button"][aria-haspopup][aria-label*="listing" i][aria-label*="type" i]',
    'button[aria-haspopup="listbox"]',
    'button[aria-haspopup="menu"]',
    'button[aria-haspopup][aria-label*="physical" i][aria-label*="digital" i]',
    'button[aria-haspopup][aria-label*="listing type" i]',
)

_LISTING_TYPE_CONTROL_LABEL_RE = re.compile(
    r"\b(?:physical|digital|mixed)\b.*\b(?:listing|type|listing-type|item)\b|\b(?:listing|type|listing-type|item)\b.*\b(?:physical|digital|mixed)\b",
    re.IGNORECASE,
)
# Etsy's current editor exposes the Physical/Digital control with this
# accessible name, without putting either value in its CSS attributes. Keep
# this deliberately exact (apart from whitespace and trailing punctuation) so
# a generic button/combobox cannot be mistaken for the listing-type control.
_LISTING_TYPE_ACCESSIBLE_NAME_RE = re.compile(
    r"^\s*what\s+type\s+of\s+item\s+is\s+it\s*[!?.,:;…]*\s*$",
    re.IGNORECASE,
)
_LISTING_TYPE_DISPLAY_VALUES = {
    "digital",
    "digital download",
    "digital file",
    "digital files",
    "digital item",
    "digital listing",
    "downloadable",
}


def _normalize_listing_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


async def _get_control_text_candidates(control) -> list[str]:
    candidates = []
    try:
        candidates.append(await control.inner_text())
    except Exception:
        candidates.append("")
    # Native <select> controls expose their selected option through
    # input_value(), while get_attribute("value") may be unset in the DOM.
    # Not every candidate is an input/select, so tolerate unsupported locator
    # types and keep the exact-value check below as the gate.
    try:
        input_value = getattr(control, "input_value", None)
        if input_value is not None:
            candidates.append(await input_value())
    except Exception:
        pass
    for attr in ("id", "name", "value", "aria-label", "role", "data-testid", "data-test-id"):
        try:
            value = await control.get_attribute(attr)
        except Exception:
            value = None
        if value:
            candidates.append(value)
    return [_normalize_listing_text(value) for value in candidates if _normalize_listing_text(value)]


async def _control_signature(control) -> str:
    parts = []
    for attr in ("id", "name", "data-testid", "data-test-id"):
        try:
            value = await control.get_attribute(attr)
        except Exception:
            value = None
        if value:
            parts.append(_normalize_listing_text(value))
    if not parts:
        try:
            parts.append(_normalize_listing_text(await control.inner_text()))
        except Exception:
            parts.append(_normalize_listing_text(repr(control)))
    return "::".join(parts) or _normalize_listing_text(repr(control))


def _looks_like_listing_type_control_text(value: str) -> bool:
    normalized = _normalize_listing_text(value)
    if not normalized:
        return False
    if normalized == "digital":
        return False
    if _LISTING_TYPE_CONTROL_LABEL_RE.search(normalized):
        return True
    if (
        ("physical" in normalized and "digital" in normalized and ("listing" in normalized or "type" in normalized or "item" in normalized))
        or ("mixed" in normalized and ("listing" in normalized or "type" in normalized or "item" in normalized))
        or ("listing-type" in normalized)
        or ("listing type" in normalized)
    ):
        return True
    return False


def _is_affirmative_digital_display(value: str) -> bool:
    return _normalize_listing_text(value) in _LISTING_TYPE_DISPLAY_VALUES


async def _is_listing_type_control(control) -> bool:
    for candidate in await _get_control_text_candidates(control):
        if _looks_like_listing_type_control_text(candidate):
            return True
    return False


async def _control_has_digital_listing_readback(control) -> bool:
    for candidate in await _get_control_text_candidates(control):
        if _is_affirmative_digital_display(candidate):
            return True
    return False


async def _locate_candidate_listing_type_controls(page) -> list:
    def selector_has_listing_signature(selector: str) -> bool:
        lower_selector = selector.casefold()
        return (
            "listing-type" in lower_selector
            or "listing type" in lower_selector
            or "mixed-listing-type" in lower_selector
            or "mixed listing type" in lower_selector
            or (
                "physical" in lower_selector
                and "digital" in lower_selector
                and "data-testid" in lower_selector
            )
        )

    controls: list = []
    collected: set[str] = set()
    for selector in _LISTING_TYPE_CONTROL_SELECTORS:
        try:
            found = await _visible_locator_items(page.locator(selector))
        except Exception:
            found = []
        for control in found:
            signature = await _control_signature(control)
            if signature in collected:
                continue
            collected.add(signature)
            if await _is_listing_type_control(control):
                controls.append(control)
                continue
            if not selector_has_listing_signature(selector):
                continue
            candidates = await _get_control_text_candidates(control)
            if any("physical" in value and "digital" in value for value in candidates):
                controls.append(control)

    for selector in (
        '[role="button"][aria-label*="physical" i][aria-label*="digital" i]',
        '[role="combobox"][aria-label*="physical" i][aria-label*="digital" i]',
        'button[aria-haspopup="listbox"][aria-label*="listing" i]',
    ):
        try:
            found = await _visible_locator_items(page.locator(selector))
        except Exception:
            found = []
        for control in found:
            signature = await _control_signature(control)
            if signature in collected:
                continue
            collected.add(signature)
            candidates = await _get_control_text_candidates(control)
            if any("physical" in value and "digital" in value for value in candidates):
                controls.append(control)

    # The current Etsy editor uses an accessible role/name pair rather than a
    # descriptive id or aria-label. Search both possible widget roles, but
    # only with the exact anchored label above. The caller still requires an
    # independently affirmative, exact Digital readback after selection.
    for role in ("combobox", "button"):
        try:
            found = await _visible_locator_items(
                page.get_by_role(role, name=_LISTING_TYPE_ACCESSIBLE_NAME_RE)
            )
        except Exception:
            found = []
        for control in found:
            signature = await _control_signature(control)
            if signature in collected:
                continue
            collected.add(signature)
            controls.append(control)

    return controls


async def _wait_for_listing_type_control(page, *, timeout_ms: int = 8000, poll_ms: int = 250):
    attempts = max(1, int(timeout_ms) // int(poll_ms) + 1)
    for attempt in range(attempts):
        controls = await _locate_candidate_listing_type_controls(page)
        if controls:
            if len(controls) == 1:
                return controls[0]
            raise DigitalListingTypeError(
                "Phát hiện nhiều hơn 1 điều khiển listing type có thể trùng nhau"
            )
        if attempt + 1 < attempts:
            await page.wait_for_timeout(poll_ms)
    return None


async def _option_text_candidates(option) -> list[str]:
    values = []
    try:
        values.append(await option.inner_text())
    except Exception:
        values.append("")
    for attr in ("value", "aria-label"):
        try:
            value = await option.get_attribute(attr)
        except Exception:
            value = None
        if value:
            values.append(value)
    deduped = []
    seen = set()
    for value in values:
        text = _normalize_listing_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        deduped.append(text)
    return deduped


async def _option_has_affirmative_digital_display(option) -> bool:
    """Check each option label/value independently for an exact Digital value."""
    return any(
        _is_affirmative_digital_display(candidate)
        for candidate in await _option_text_candidates(option)
    )


async def _wait_for_listing_type_listbox(page, *, timeout_ms: int = 5000, poll_ms: int = 250):
    popup_name_re = re.compile(r"(?:physical|digital|mixed).*listing type|listing type.*(?:physical|digital)", re.IGNORECASE)
    attempts = max(1, int(timeout_ms) // int(poll_ms) + 1)
    for attempt in range(attempts):
        for role in ("listbox", "menu"):
            try:
                named = page.get_by_role(role, name=popup_name_re)
                candidates = await _visible_locator_items(named)
            except Exception:
                candidates = []

            if not candidates:
                try:
                    candidates = await _visible_locator_items(page.get_by_role(role))
                except Exception:
                    candidates = []

            for candidate in candidates:
                options = await _visible_locator_items(candidate.get_by_role("option"))
                if not options:
                    continue
                if await _is_listing_type_control(candidate):
                    return candidate
                digital_options = [
                    option for option in options
                    if await _option_has_affirmative_digital_display(option)
                ]
                if len(digital_options) == 1:
                    return candidate
        if attempt + 1 < attempts:
            await page.wait_for_timeout(poll_ms)
    return None


async def _select_digital_listing_type_dropdown_if_present(page) -> bool:
    """Select and verify the current Etsy Physical/Digital listbox control."""
    controls = await _locate_candidate_listing_type_controls(page)
    if not controls:
        controls = []
    if not controls:
        single = await _wait_for_listing_type_control(page)
        controls = [single] if single is not None else []
    if not controls:
        return False
    if len(controls) != 1:
        raise DigitalListingTypeError(
            f"Tìm thấy {len(controls)} dropdown loại listing đang hiển thị; không thể chọn Digital an toàn"
        )

    control = controls[0]
    current_is_digital = await _control_has_digital_listing_readback(control)
    if not current_is_digital:
        await control.click()
        await page.wait_for_timeout(300)

        listbox = await _wait_for_listing_type_listbox(page, timeout_ms=5000, poll_ms=250)
        if listbox is None:
            raise DigitalListingTypeError(
                "Cần đúng 1 listbox listing type sau khi mở dropdown"
            )

        digital_options = []
        for idx in range(5):
            candidates = await _visible_locator_items(listbox.get_by_role("option"))
            digital_options = []
            for candidate in candidates:
                if await _option_has_affirmative_digital_display(candidate):
                    digital_options.append(candidate)
            if digital_options:
                break
            if idx + 1 < 5:
                await page.wait_for_timeout(250)

        if len(digital_options) != 1:
            raise DigitalListingTypeError(
                f"Cần đúng 1 tùy chọn Digital trong listing-type listbox, tìm thấy {len(digital_options)}"
            )
        await digital_options[0].click()
        await page.wait_for_timeout(500)

    for _ in range(5):
        if await _control_has_digital_listing_readback(control):
            return True
        await page.wait_for_timeout(200)
        final_text = _normalize_listing_text(await control.inner_text())

    raise DigitalListingTypeError(
        "Dropdown loại listing không giữ trạng thái Digital (đang hiển thị: "
        f"'{final_text or 'trống'}')"
    )


async def _legacy_listing_type_radio_semantics(radio) -> dict[str, object]:
    raw = await radio.evaluate("""el => {
        const labels = el.labels ? Array.from(el.labels) : [];
        return {
            value: el.value || '',
            ariaLabel: el.getAttribute('aria-label') || '',
            labels: labels.map(label => label.innerText || label.textContent || ''),
        };
    }""")
    if not isinstance(raw, dict):
        return {"value": "", "aria_label": "", "labels": []}

    def normalize(value) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip().casefold()

    raw_labels = raw.get("labels")
    labels = raw_labels if isinstance(raw_labels, list) else []
    return {
        "value": normalize(raw.get("value")),
        "aria_label": normalize(raw.get("ariaLabel")),
        "labels": [normalize(label) for label in labels if normalize(label)],
    }


def _is_affirmative_digital_radio_semantics(semantics: dict[str, object]) -> bool:
    """Accept only an independently affirmative, exact Digital radio label/value."""
    allowed = {
        "digital",
        "digital download",
        "digital file",
        "digital files",
        "digital item",
        "digital listing",
        "downloadable",
    }
    fields = [semantics.get("value"), semantics.get("aria_label")]
    labels = semantics.get("labels")
    if isinstance(labels, list):
        fields.extend(labels)
    normalized_fields = [str(field or "").strip().casefold() for field in fields]
    combined = " ".join(normalized_fields)
    if re.search(
        r"\bphysical\b|\bnot\s+(?:an?\s+)?digital\b|\bnon[- ]digital\b|\bno\s+digital\b",
        combined,
    ):
        return False
    return any(field in allowed for field in normalized_fields)


async def _select_and_verify_legacy_digital_radio(page) -> bool:
    """Select a legacy Digital radio by semantics, independent of DOM order."""
    radios = page.locator('input[name="listing_type_options_group"]')
    radio_count = await radios.count()
    if radio_count == 0:
        return False

    digital_radios = []
    for index in range(radio_count):
        radio = radios.nth(index)
        semantics = await _legacy_listing_type_radio_semantics(radio)
        if _is_affirmative_digital_radio_semantics(semantics):
            digital_radios.append((index, radio))

    if len(digital_radios) != 1:
        raise DigitalListingTypeError(
            f"Cần đúng 1 legacy radio Digital, tìm thấy {len(digital_radios)}"
        )

    digital_index, digital = digital_radios[0]
    if not await digital.is_checked():
        if not await digital.is_enabled():
            raise DigitalListingTypeError("Legacy radio Digital đang bị vô hiệu hóa")
        await digital.check(force=True, timeout=5000)
        await page.wait_for_timeout(500)

    checked_indexes = []
    for index in range(radio_count):
        radio = radios.nth(index)
        if await radio.is_checked():
            checked_indexes.append(index)
    if checked_indexes != [digital_index] or not await digital.is_checked():
        raise DigitalListingTypeError("Legacy radio không giữ duy nhất trạng thái Digital")

    final_semantics = await _legacy_listing_type_radio_semantics(digital)
    if not _is_affirmative_digital_radio_semantics(final_semantics):
        raise DigitalListingTypeError(
            "Legacy radio đã chọn không còn mang ngữ nghĩa Digital khẳng định"
        )
    return True


async def select_and_verify_digital_listing_type(page) -> str:
    """Fail closed unless the current or legacy editor proves Digital state."""
    if await _select_digital_listing_type_dropdown_if_present(page):
        return "dropdown"
    if await _select_and_verify_legacy_digital_radio(page):
        return "legacy_radio"
    raise DigitalListingTypeError("Không tìm thấy điều khiển loại listing Physical/Digital")


DIGITAL_UPLOAD_SURFACE_WAIT_MS = 12000
DIGITAL_UPLOAD_SURFACE_POLL_MS = 250
DIGITAL_ADD_FILE_SCROLL_TIMEOUT_MS = 1500
DIGITAL_ADD_FILE_VISIBILITY_TIMEOUT_MS = 1500
DIGITAL_ADD_FILE_CLICK_TIMEOUT_MS = 10000
DIGITAL_SURFACE_LIVENESS_TIMEOUT_MS = 1500
DIGITAL_CATEGORY_SURFACE_SETTLE_MS = 2500
DIGITAL_SURFACE_STABILITY_HOLD_MS = 500
DIGITAL_SURFACE_STABILITY_POLL_MS = 100
# Large customer files can remain in Etsy's Loading state while the upload
# endpoint is still processing. Keep response capture bounded, but give it the
# same 90-second budget as the positive UI upload/readback contract below.
DIGITAL_FILE_UPLOAD_RESPONSE_TIMEOUT_MS = 90000
DIGITAL_SAVED_DRAFT_READBACK_TIMEOUT_MS = 15000
DIGITAL_SAVED_DRAFT_READBACK_POLL_MS = 250
_DIGITAL_FILES_NAME_RE = re.compile(r"^Digital files?$", re.IGNORECASE)
_ADD_FILE_NAME_RE = re.compile(r"^Add files?$", re.IGNORECASE)
_IMAGE_ONLY_ACCEPT_TOKENS = {
    "image/*", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".avif", ".svg",
}


def _normalize_ui_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _accepts_only_images(accept: str | None) -> bool:
    """Return True for an input that cannot be used for customer files."""
    tokens = [token.strip().casefold() for token in str(accept or "").split(",") if token.strip()]
    if not tokens:
        return False
    return all(
        token.startswith("image/") or token in _IMAGE_ONLY_ACCEPT_TOKENS
        for token in tokens
    )


async def _input_has_digital_scope(inp, *, container=None) -> bool:
    """Require a non-image input to be a descendant of the resolved container."""
    try:
        accept = await inp.get_attribute("accept")
    except Exception:
        accept = None
    if _accepts_only_images(accept):
        return False
    if container is None:
        # A page-global input is never enough evidence. Callers must first
        # resolve the bounded Digital files container and pass it explicitly.
        return False
    try:
        input_handle = await inp.element_handle()
        container_handle = await container.element_handle()
        if input_handle is None or container_handle is None:
            return False
        return bool(await input_handle.evaluate(
            "(el, root) => Boolean(root && root.contains(el))",
            container_handle,
        ))
    except Exception:
        return False


async def _find_scoped_customer_file_inputs(page, *, container=None):
    """Return only non-image file inputs inside one resolved container."""
    if container is None:
        return []
    try:
        all_inputs = container.locator('input[type="file"]')
        count = await all_inputs.count()
    except Exception:
        return []
    scoped = []
    for idx in range(count):
        inp = all_inputs.nth(idx)
        if await _input_has_digital_scope(inp, container=container):
            scoped.append((idx, inp))
    return scoped


async def _find_visible_digital_files_container(page):
    """Resolve one visible bounded Digital files container, never a whole form."""
    try:
        stable = page.locator("#field-digitalFiles")
        stable_items = await _visible_locator_items(stable)
        if len(stable_items) == 1:
            return stable_items[0]
    except Exception:
        pass

    # A region is already bounded. Prefer it over reconstructing a scope from
    # a heading, and fail closed if Etsy renders multiple visible exact regions.
    try:
        regions = page.get_by_role("region", name=_DIGITAL_FILES_NAME_RE)
        visible_regions = await _visible_locator_items(regions)
        if len(visible_regions) == 1:
            return visible_regions[0]
    except Exception:
        pass

    # If only the heading is exposed, climb to the nearest bounded semantic
    # container. Deliberately exclude <form>: it may contain Personalization,
    # gallery, and unrelated file inputs.
    try:
        headings = page.get_by_role("heading", name=_DIGITAL_FILES_NAME_RE)
        visible_headings = await _visible_locator_items(headings)
        if len(visible_headings) != 1:
            return None
        bounded = visible_headings[0].locator(
            "xpath=ancestor::*[self::section or self::fieldset or @role='region'][1]"
        )
        bounded_items = await _visible_locator_items(bounded)
        if len(bounded_items) == 1:
            return bounded_items[0]
    except Exception:
        pass
    return None


async def _find_exact_add_file_affordance(
    page,
    *,
    container=None,
    visibility_timeout_ms=None,
):
    """Find Add file/Add files only inside the resolved Digital files container."""
    if container is None:
        container = await _find_visible_digital_files_container(page)
    if container is None:
        return None
    try:
        button = container.get_by_role("button", name=_ADD_FILE_NAME_RE)
        visible_buttons = await _visible_locator_items(
            button,
            visibility_timeout_ms=visibility_timeout_ms,
        )
        if visible_buttons:
            return visible_buttons[0]
    except Exception:
        pass

    # Labels and role buttons are included because Etsy has rendered both
    # variants over time. Text/aria-label is checked exactly; no has-text
    # substring selector is allowed here.
    try:
        candidates = container.locator('label, [role="button"]')
        for idx in range(await candidates.count()):
            candidate = candidates.nth(idx)
            if not await _is_locator_visible(
                candidate,
                visibility_timeout_ms=visibility_timeout_ms,
            ):
                continue
            try:
                text = _normalize_ui_text(await candidate.inner_text())
            except Exception:
                text = ""
            try:
                aria_label = _normalize_ui_text(await candidate.get_attribute("aria-label"))
            except Exception:
                aria_label = ""
            if _ADD_FILE_NAME_RE.fullmatch(text) or _ADD_FILE_NAME_RE.fullmatch(aria_label):
                return candidate
    except Exception:
        pass
    return None


async def _inspect_digital_upload_surface(page) -> dict[str, object] | None:
    """Resolve one positive upload surface within one bounded container."""
    container = await _find_visible_digital_files_container(page)
    if container is None:
        return None

    affordance = await _find_exact_add_file_affordance(page, container=container)
    if affordance is not None:
        return {"kind": "add_file", "locator": container, "control": affordance}

    scoped_inputs = await _find_scoped_customer_file_inputs(page, container=container)
    if scoped_inputs:
        index, inp = scoped_inputs[0]
        return {
            "kind": "customer_file_input",
            "index": index,
            "locator": container,
            "control": inp,
        }
    # The bounded region itself is still valid positive evidence while Etsy's
    # React controls are rendering. The uploader will continue to fail closed
    # if neither a scoped Add file control nor a scoped customer-file input
    # appears before mutation.
    return {"kind": "digital_files", "locator": container}


async def _resolve_digital_upload_surface(
    page,
    *,
    timeout_ms: int = DIGITAL_UPLOAD_SURFACE_WAIT_MS,
    poll_ms: int = DIGITAL_UPLOAD_SURFACE_POLL_MS,
) -> dict[str, object]:
    """Boundedly wait for React to render a positively identified upload surface."""
    poll_ms = max(1, int(poll_ms))
    attempts = max(1, int(timeout_ms) // poll_ms + 1)
    for attempt in range(attempts):
        surface = await _inspect_digital_upload_surface(page)
        if surface is not None:
            return surface
        if attempt + 1 < attempts:
            await page.wait_for_timeout(poll_ms)
    raise DigitalListingTypeError(
        "Không tìm thấy bề mặt upload customer files đã xác minh "
        "(Digital files, Add file(s), hoặc input customer-file có scope)"
    )


async def _establish_stable_digital_upload_surface(
    page,
    *,
    initial_surface=None,
    settle_ms: int = DIGITAL_CATEGORY_SURFACE_SETTLE_MS,
    hold_ms: int = DIGITAL_SURFACE_STABILITY_HOLD_MS,
    poll_ms: int = DIGITAL_SURFACE_STABILITY_POLL_MS,
):
    """Wait for category React work to settle, then hold one live surface briefly."""
    settle_ms = max(0, int(settle_ms))
    hold_ms = max(0, int(hold_ms))
    poll_ms = max(1, int(poll_ms))
    if settle_ms:
        await page.wait_for_timeout(settle_ms)

    surface = initial_surface
    if surface is None:
        try:
            surface = await asyncio.wait_for(
                _find_visible_digital_files_container(page),
                timeout=DIGITAL_SURFACE_LIVENESS_TIMEOUT_MS / 1000,
            )
        except asyncio.TimeoutError:
            return None
    if surface is None:
        return None

    deadline = asyncio.get_running_loop().time() + hold_ms / 1000
    while True:
        if not await _is_likely_live_surface(
            page,
            locator=surface,
            attempts=1,
            wait_ms=0,
        ):
            return None
        remaining_ms = int(max(0, (deadline - asyncio.get_running_loop().time()) * 1000))
        if remaining_ms <= 0:
            return surface
        await page.wait_for_timeout(min(poll_ms, remaining_ms))


async def _verify_digital_files_region(page) -> None:
    """Backward-compatible name for the shared positive upload-surface check."""
    await _resolve_digital_upload_surface(page)


async def _select_category_with_exact_readback(page, search_term: str, *, cat_input=None) -> str:
    """Select one exact category option and verify readback, returning the clicked text."""
    if not search_term:
        raise RuntimeError("search_term for category selection cannot be empty")
    if cat_input is None:
        cat_input = page.locator(
            '#category-field-search, '
            '#listing-editor_category-search-typeahead, '
            'input[placeholder*="Examples:"], '
            'input.wt-input.le-category-search__input[placeholder*="Type to search" i], '
            'input[placeholder*="category" i], '
            'input[role="combobox"][placeholder*="Type to search" i], '
            'input[aria-label*="category" i]'
        ).first

    # When a committed category is present, the input may be read-only and a
    # #field-category overlay can block direct clicks. Clear it first when the
    # explicit control is available, then proceed with exact search input.
    clear_button = page.locator('#field-category button[aria-label="Clear"]').first
    clear_attempts = 3
    for attempt in range(clear_attempts):
        has_clear = False
        try:
            has_clear = (await clear_button.count()) > 0 and await clear_button.is_visible()
        except Exception:
            has_clear = False
        if not has_clear:
            if attempt == 0:
                break
            await page.wait_for_timeout(120)
            continue
        try:
            await clear_button.scroll_into_view_if_needed()
            await clear_button.click(force=True)
            await page.wait_for_timeout(180)
            break
        except Exception as clear_error:
            if attempt + 1 >= clear_attempts:
                raise RuntimeError(
                    "Không bấm được nút Clear danh mục đã chọn trước khi tìm kiếm lại."
                ) from clear_error
            await page.wait_for_timeout(120)
            continue

    await cat_input.wait_for(state="visible", timeout=6000)
    await cat_input.click()
    await cat_input.fill("")
    await cat_input.fill(search_term)

    target_norm = normalize_category_text(search_term)
    option_text_sample_limit = 8
    observed_option_texts: list[str] = []
    selected_option_text = ""
    selected = False
    options_selector = '[role="option"], li[class*="option"], li[class*="result"], div[role="option"]'
    for _ in range(15):
        options = page.locator(options_selector)
        option_count = await options.count()
        for idx in range(option_count):
            opt = options.nth(idx)
            if not await opt.count() > 0:
                continue
            if not await opt.is_visible():
                continue
            opt_text = (await opt.inner_text()) or ""
            if opt_text and len(observed_option_texts) < option_text_sample_limit:
                if opt_text not in observed_option_texts:
                    observed_option_texts.append(opt_text)
            if category_option_matches(search_term, opt_text):
                await opt.click()
                selected_option_text = opt_text
                selected = True
                break
        if selected:
            break
        await page.wait_for_timeout(300)

    if not selected:
        visible_sample = ", ".join(observed_option_texts[:option_text_sample_limit]) or "<không có tùy chọn nào>"
        raise RuntimeError(
            f"Không tìm thấy tùy chọn danh mục chính xác cho '{search_term}'. "
            f"Giá trị gợi ý không khớp (target_norm='{target_norm}'). "
            f"Đã quan sát: [{visible_sample}]"
        )

    # Verify selected category is visibly reflected in UI
    matched = False
    for _ in range(12):
        current_text = ""
        try:
            current_text = (await cat_input.input_value()).strip()
        except Exception:
            pass
        if category_option_matches(search_term, current_text):
            matched = True
            break
        await page.wait_for_timeout(300)

    if not matched:
        raise RuntimeError(f"Không xác nhận được danh mục đã chọn: '{search_term}'.")

    return selected_option_text


async def _prime_digital_listing_type_via_planner_templates(
    page,
    target_search_term: str,
    *,
    planner_category: str = "Planner Templates",
    cat_input=None,
) -> None:
    """Prime Digital listing type via a mixed category path when direct digital-only path misses upload surface."""
    if not target_search_term:
        raise DigitalListingTypeError("Thiếu category target để prime listing type")

    # Move to known mixed category used by this shop, select Digital via the
    # exact listing-type helper, then re-select the original category.
    print("  🧪 Chọn tạm Planner Templates để mở Digital control")
    await _select_category_with_exact_readback(page, planner_category, cat_input=cat_input)
    if not await _select_digital_listing_type_dropdown_if_present(page):
        raise DigitalListingTypeError(
            "Không chọn được Digital sau khi chuyển qua category tạm Planner Templates"
        )
    print("  🧪 Đã kiểm tra Digital control")
    await _select_category_with_exact_readback(page, target_search_term, cat_input=cat_input)
    print(f"  🧪 Đã chọn lại mục tiêu danh mục: {target_search_term}")


async def fill_category_tab(page, product):
    await click_tab(page, "Category")
    await page.wait_for_timeout(800)

    # The current editor scopes category choices by this Physical/Digital
    # dropdown. Select Digital before searching so the category result set is
    # digital-aware. Legacy forms expose their radios later in Item Options.
    category_listing_type_control = await _select_digital_listing_type_dropdown_if_present(page)

    category_str = resolve_listing_category(product)
    if not category_str:
        raise RuntimeError(
            f"Không suy luận được danh mục cho '{product.get('title', '').strip()}'. "
            f"Vui lòng điền cột Category."
        )
    if product.get("category"):
        print(f"  📂 Sử dụng category từ workbook: '{category_str}'")
    else:
        print(f"  💡 Cột Category trống. Tự động suy luận danh mục: '{category_str}'")

    search_term = extract_category_leaf(category_str).strip()
    if not search_term:
        raise RuntimeError(f"Category sau khi chuẩn hóa rỗng: '{category_str}'")

    print(f"  🔍 Tìm kiếm danh mục cho term: '{search_term}'...")

    cat_input = page.locator(
        '#category-field-search, '
        '#listing-editor_category-search-typeahead, '
        'input[placeholder*="Examples:"], '
        'input.wt-input.le-category-search__input[placeholder*="Type to search" i], '
        'input[placeholder*="category" i], '
        'input[role="combobox"][placeholder*="Type to search" i], '
        'input[aria-label*="category" i]'
    ).first
    selected_option_text = await _select_category_with_exact_readback(
        page,
        search_term,
        cat_input=cat_input,
    )

    has_customer_files = bool(product.get("pdf_paths"))
    if (
        not has_customer_files
        and not product.get("_digital_listing_type_verified")
        and _category_option_is_digital_only_metadata(selected_option_text)
    ):
        product["_digital_listing_type_verified"] = "category_digital_metadata"

    if category_listing_type_control:
        if not await _select_digital_listing_type_dropdown_if_present(page):
            raise DigitalListingTypeError(
                "Dropdown loại listing biến mất sau khi chọn category; không thể xác minh Digital"
            )
        product["_digital_listing_type_verified"] = "dropdown"

    print(f"  📂 Category: {search_term} ✓")

    # Etsy keeps the customer-file surface mounted on the Category view after
    # the digital category is selected.  Item Options/Item Details navigation
    # can unmount that region permanently for the current editor instance, so
    # upload while the positively verified Category surface is still present.
    # The marker is written only after upload_digital_files completes its
    # positive UI read-back contract.
    if has_customer_files and not product.get("_digital_files_uploaded"):
        initial_surface = await _find_visible_digital_files_container(page)
        surface = await _establish_stable_digital_upload_surface(
            page,
            initial_surface=initial_surface,
        )
        verified_surface = None
        if surface is not None:
            verified_surface = {"locator": surface}
            print("  ✅ Surface Digital files ổn định sau khi chọn Category")
        elif initial_surface is not None:
            print("  ⚠ Surface Digital files transient — chọn lại category chính xác")
            await _select_category_with_exact_readback(
                page,
                search_term,
                cat_input=cat_input,
            )
            surface = await _establish_stable_digital_upload_surface(page)
            if surface is not None:
                verified_surface = {"locator": surface}
                print("  ✅ Surface Digital files ổn định sau khi chọn lại category")
        if surface is None:
            await _prime_digital_listing_type_via_planner_templates(
                page,
                search_term,
                cat_input=cat_input,
            )
            surface = await _establish_stable_digital_upload_surface(page)
            if surface is not None:
                verified_surface = {"locator": surface}
                print("  ✅ Surface Digital files ổn định sau khi prime Planner Templates")
            if surface is None:
                if not category_listing_type_control and not await _select_digital_listing_type_dropdown_if_present(page):
                    raise DigitalListingTypeError(
                        "Sau khi prime qua Planner Templates vẫn không hiển thị được "
                        "Digital files region"
                    )
        if surface is not None or category_listing_type_control:
            product["_digital_listing_type_verified"] = "digital_files"

        await upload_digital_files(
            page,
            product,
            verified_surface=verified_surface,
        )
        product["_digital_files_uploaded"] = True

# ── Tab: Item Details ──────────────────────────────────────────────────────────
async def fill_item_details_tab(page, product):
    await dismiss_alerts(page)
    if not await _click_verified_item_details_tab(page):
        raise DigitalListingTypeError(
            "Không tìm thấy tab Item Details chính xác"
        )
    await page.wait_for_timeout(1200)
    await dismiss_alerts(page)

    # Etsy lazy-mounts the Physical/Digital control in Item Details. Prove the
    # listing type while this tab is active, before Item Options can unmount it.
    if not product.get("_digital_listing_type_verified"):
        product["_digital_listing_type_verified"] = await select_and_verify_digital_listing_type(page)
        print("  💻 Listing type: Digital ✓")

    # Title
    if await smart_fill(page, '#listing-title-input, textarea[name="title"]', product["title"]):
        print("  📝 Title ✓")
    await page.wait_for_timeout(500)

    # Description
    if await smart_fill(page, '#listing-description-textarea, textarea[name="description"]',
                        product["description"], timeout=6000):
        print("  📄 Description ✓")
    await page.wait_for_timeout(500)

    # Tags (EN)
    try:
        await dismiss_alerts(page)
        # Delete existing tags first to make room
        try:
            await page.evaluate('''() => {
                let removeButtons = Array.from(document.querySelectorAll('span.wt-tag button, [data-testid="tag-pill"] button, button[aria-label*="remove" i], button[aria-label*="delete" i], button[class*="remove" i]'));
                for (let btn of removeButtons) {
                    btn.click();
                }
            }''')
            await page.wait_for_timeout(1000)
        except Exception as tag_clear_ex:
            print(f"  ⚠ Lỗi xóa tag cũ: {tag_clear_ex}")

        tag_list = clean_tags(product["tags"])
        filled = skipped = 0
        for tag in tag_list:
            await dismiss_alerts(page)
            tag_el = page.locator('#listing-tags-input, input[placeholder*="tag" i]').first
            if await tag_el.is_visible() and await tag_el.is_editable():
                try:
                    await tag_el.fill(tag, timeout=3000)
                    await page.wait_for_timeout(300)
                    await tag_el.press("Enter")
                    await page.wait_for_timeout(600)
                    err = page.locator('.wt-alert--error-01, [role="alert"]').first
                    if await err.count() > 0 and await err.is_visible():
                        await dismiss_alerts(page)
                        skipped += 1
                    else:
                        filled += 1
                except Exception as fill_ex:
                    print(f"      ⚠️ Không điền được tag '{tag}': {fill_ex}")
                    skipped += 1
        msg = f"  🏷  {filled} tags ✓"
        if skipped: msg += f" ({skipped} bị reject/bỏ qua)"
        print(msg)
    except Exception as e:
        print(f"  ⚠ tags: {e}")
    await page.wait_for_timeout(500)

    # Shop Section
    section_name = product.get("section", "").strip()
    if section_name:
        try:
            await dismiss_alerts(page)
            # Tìm dropdown Shop Section
            section_found = False
            for sel in [
                'select[name*="section" i]',
                'select[id*="section" i]',
                '[data-testid*="section"] select',
            ]:
                sel_el = page.locator(sel).first
                if await sel_el.count() > 0 and await sel_el.is_visible():
                    await sel_el.select_option(label=section_name)
                    await page.wait_for_timeout(800)
                    print(f"  📁 Section: {section_name} ✓")
                    section_found = True
                    break

            if not section_found:
                # Thử click vào combobox rồi chọn option
                for btn_sel in [
                    'button[id*="section" i]',
                    '[data-testid*="section"] button',
                    'div[class*="section"] button',
                ]:
                    btn = page.locator(btn_sel).first
                    if await btn.count() > 0 and await btn.is_visible():
                        await btn.click()
                        await page.wait_for_timeout(800)
                        opt = page.locator(f'[role="option"]:has-text("{section_name}"), li:has-text("{section_name}")').first
                        if await opt.count() > 0:
                            await opt.click()
                            await page.wait_for_timeout(800)
                            print(f"  📁 Section: {section_name} ✓")
                            section_found = True
                            break

            if not section_found:
                print(f"  ⚠ Section: không tìm thấy dropdown — bỏ qua")
        except Exception as e:
            print(f"  ⚠ section: {e}")

    # Category's Digital files root can be transient while Etsy finishes
    # taxonomy/prefilled-data rendering. Reconcile again on the stable Item
    # Details surface immediately before Item Options/Save, without uploading
    # a duplicate when the exact bounded readback already contains the files.
    if product.get("pdf_paths"):
        await _reconcile_digital_files_on_item_details(page, product)

# ── Tab: Item Options ──────────────────────────────────────────────────────────
async def fill_item_options_tab(page, product):
    await click_tab(page, "Item Options", "Options")
    await page.wait_for_timeout(1000)

    if not product.get("_digital_listing_type_verified"):
        product["_digital_listing_type_verified"] = await select_and_verify_digital_listing_type(page)
    print("  💻 Listing type: Digital ✓")

    # Who made = I did
    try:
        who = page.locator('input[name="whoMade"]').first
        if await who.count() > 0:
            handle = await who.element_handle()
            if handle:
                await page.evaluate("el => el.click()", handle)
                await page.wait_for_timeout(600)
                print("  👤 Who made: I did ✓")
    except Exception as e:
        print(f"  ⚠ whoMade: {e}")

    # What is it = Finished product
    try:
        supply = page.locator('input[name="isSupply"]').first
        if await supply.count() > 0:
            handle = await supply.element_handle()
            if handle:
                await page.evaluate("el => el.click()", handle)
                await page.wait_for_timeout(600)
                print("  🏷 Finished product ✓")
    except Exception as e:
        print(f"  ⚠ isSupply: {e}")

    # When made
    try:
        when = page.locator('#when-made-select').first
        if await when.count() > 0 and await when.is_visible():
            await when.select_option(value=product["when_made"])
            await page.wait_for_timeout(600)
            print(f"  📅 When made ✓")
    except Exception as e:
        print(f"  ⚠ whenMade: {e}")

    # How is this digital content created? → "Created by me"
    await page.wait_for_timeout(800)
    try:
        # Tìm radio button "Created by me"
        created_by_me = False
        for sel in [
            'input[type="radio"][value*="created_by_me" i]',
            'input[type="radio"][value*="human" i]',
            'input[type="radio"][value*="manually" i]',
        ]:
            rb = page.locator(sel).first
            if await rb.count() > 0:
                handle = await rb.element_handle()
                if handle:
                    await page.evaluate("el => el.click()", handle)
                    await page.wait_for_timeout(600)
                    print("  ✍️  Digital content: Created by me ✓")
                created_by_me = True
                break

        if not created_by_me:
            # Tìm theo label text
            for label_text in ["Created by me", "I created this"]:
                label = page.locator(f'label:has-text("{label_text}")').first
                if await label.count() > 0 and await label.is_visible():
                    await label.click()
                    await page.wait_for_timeout(600)
                    print(f"  ✍️  Digital content: {label_text} ✓")
                    created_by_me = True
                    break

        if not created_by_me:
            # Thử radio đầu tiên trong nhóm liên quan
            radios = page.locator('input[type="radio"]')
            cnt = await radios.count()
            for i in range(cnt):
                r = radios.nth(i)
                # Tìm radio gần text "Created by me"
                parent = r.locator('xpath=../..')
                txt = await parent.inner_text()
                if "created" in txt.lower() or "Created" in txt:
                    handle = await r.element_handle()
                    if handle:
                        await page.evaluate("el => el.click()", handle)
                        await page.wait_for_timeout(600)
                        print("  ✍️  Digital content: Created by me ✓")
                    created_by_me = True
                    break
    except Exception as e:
        print(f"  ⚠ digital content created: {e}")

    # Category normally uploads while Etsy's Digital files region is mounted.
    # Keep an exact-tab, fail-closed fallback for UI variants where Category
    # did not expose that surface; never upload a second time after a verified
    # Category upload.
    if product["pdf_paths"] and not product.get("_digital_files_uploaded"):
        if not await _click_verified_item_details_tab(page):
            raise DigitalListingTypeError(
                "Không tìm thấy tab Item Details chính xác sau Item Options"
            )
        await page.wait_for_timeout(1000)
        await upload_digital_files(page, product)
        product["_digital_files_uploaded"] = True

# ── Tab: Pricing & Shipping ────────────────────────────────────────────────────
async def fill_pricing_tab(page, product):
    await dismiss_alerts(page)
    await click_tab(page, "Pricing & Shipping", "Pricing")
    await page.wait_for_timeout(1200)
    await dismiss_alerts(page)

    if await smart_fill(page, '#listing-price-input, [data-testid="price-input"], input[name="price"]',
                        f"{product['price']:.2f}"):
        print(f"  💲 Price: ${product['price']:.2f} ✓")
    else:
        print(f"  ⚠ Price: không điền được — kiểm tra thủ công")
    await page.wait_for_timeout(500)

    if await smart_fill(page, '#listing-quantity-input, input[name="quantity"]', str(product["qty"])):
        print(f"  🔢 Qty: {product['qty']} ✓")

    if "sku" in product and product["sku"]:
        if await smart_fill(page, '#listing-sku-input, input[name="sku"], [data-testid="sku-input"]', str(product["sku"])):
            print(f"  🔑 SKU: {product['sku']} ✓")

# ── Upload Digital Files ───────────────────────────────────────────────────────
DIGITAL_FILE_UPLOAD_WAIT_MS = 90000
DIGITAL_FILE_UPLOAD_POLL_MS = 1000
DIGITAL_FILE_UPLOAD_STABLE_READS = 2


def _normalize_customer_filename(value) -> str:
    """Normalize local/UI filenames across Etsy's whitespace sanitization."""
    name = Path(str(value or "")).name
    ascii_name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", ascii_name.casefold())


def _canonicalize_customer_filename_representation(value) -> str:
    """Collapse one file's filename text and action-label representations."""
    text = _normalize_ui_text(value)
    if not text:
        return ""
    action_words = r"remove|delete|download|uploaded|completed"
    text = re.sub(
        rf"^(?:{action_words})\b(?:\s+file)?[\s:–—-]+",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        rf"[\s:–—-]+(?:{action_words})\b(?:\s+file)?$",
        "",
        text,
        flags=re.IGNORECASE,
    )
    return _normalize_customer_filename(text)


def _canonicalize_customer_filename_representations(values) -> list[str]:
    """Dedupe filename/action-label representations by logical file identity."""
    identities = []
    seen = set()
    for value in values or []:
        identity = _canonicalize_customer_filename_representation(value)
        if identity and identity not in seen:
            seen.add(identity)
            identities.append(identity)
    return identities


def _customer_file_alias_groups(source_paths, staged_paths=None):
    """Build one source/staged alias group for each logical customer file."""
    source_values = [Path(str(path)) for path in (source_paths or [])]
    staged_values = [Path(str(path)) for path in (staged_paths or [])]
    groups = []
    for index, source in enumerate(source_values):
        aliases = {
            _normalize_customer_filename(source.name),
        }
        staged = staged_values[index] if index < len(staged_values) else None
        if staged is not None:
            aliases.add(_normalize_customer_filename(staged.name))
        groups.append({
            "source": source.name,
            "staged": staged.name if staged is not None else "",
            "aliases": sorted(alias for alias in aliases if alias),
        })
    return groups


def _add_customer_receipt_aliases(alias_groups, receipts):
    """Attach a complete, positional receipt set or fail closed."""
    groups = [dict(group) for group in (alias_groups or [])]
    for group in groups:
        group["aliases"] = _canonicalize_customer_filename_representations(
            group.get("aliases") or []
        )
    normalized_receipts = []
    receipt_ids = set()
    for receipt in receipts or []:
        if not isinstance(receipt, dict):
            continue
        try:
            file_id = int(receipt.get("fileId", receipt.get("file_id")))
        except (TypeError, ValueError):
            continue
        if file_id <= 0:
            continue
        if file_id in receipt_ids:
            raise DigitalListingTypeError(
                "Customer-file response lặp fileId; dừng để tránh map sai logical source"
            )
        receipt_ids.add(file_id)
        receipt_name = _canonicalize_customer_filename_representation(
            receipt.get("name")
        )
        if not receipt_name:
            raise DigitalListingTypeError(
                "Customer-file response có fileId hợp lệ nhưng thiếu filename"
            )
        normalized_receipts.append(receipt_name)
    if not normalized_receipts:
        if groups:
            raise DigitalListingTypeError(
                "Customer-file response không có receipt hợp lệ cho logical source files"
            )
        return groups
    if len(normalized_receipts) != len(groups):
        raise DigitalListingTypeError(
            "Customer-file receipt cardinality không khớp logical source files"
        )
    # The response preserves upload order and has one receipt per staged
    # source, so non-equivalent names still map deterministically by position.
    for index, receipt_name in enumerate(normalized_receipts):
        groups[index]["aliases"] = sorted(
            set(groups[index].get("aliases") or []) | {receipt_name}
        )
    return groups


_ETSY_CUSTOMER_FILENAME_RE = re.compile(r"^[A-Za-z0-9._-]{3,70}$")


def _etsy_safe_customer_filename(
    source_name,
    reserved_names=None,
    reserved_readback_names=None,
) -> str:
    """Return a deterministic Etsy-safe customer-file basename."""
    source = Path(str(source_name or "")).name
    extension = Path(source).suffix.lower()
    if extension not in POST_FILE_EXTS:
        raise DigitalListingTypeError(
            f"Customer file '{source}' có phần mở rộng không được hỗ trợ"
        )

    stem = Path(source).stem
    ascii_stem = unicodedata.normalize("NFKD", stem).encode("ascii", "ignore").decode("ascii")
    ascii_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", ascii_stem)
    ascii_stem = re.sub(r"-+", "-", ascii_stem).strip("._-") or "file"

    names = {str(value).casefold() for value in (reserved_names or ())}
    readback_names = {
        _normalize_customer_filename(value)
        for value in (reserved_readback_names or ())
        if _normalize_customer_filename(value)
    }

    def candidate(number: int) -> str:
        suffix = "" if number == 1 else f"-{number}"
        stem_budget = max(1, 70 - len(extension) - len(suffix))
        candidate_stem = ascii_stem[:stem_budget].rstrip("._-") or "file"
        candidate_stem = candidate_stem[:max(1, 70 - len(extension) - len(suffix))]
        return f"{candidate_stem}{suffix}{extension}"

    number = 1
    while True:
        value = candidate(number)
        if not _ETSY_CUSTOMER_FILENAME_RE.fullmatch(value):
            raise DigitalListingTypeError(
                f"Không tạo được basename Etsy hợp lệ cho customer file '{source}'"
            )
        normalized = _normalize_customer_filename(value)
        if value.casefold() not in names and normalized not in readback_names:
            return value
        number += 1


@contextmanager
def _stage_etsy_customer_files(source_paths):
    """Copy customer files into one isolated directory with safe basenames."""
    paths = [Path(str(path)) for path in (source_paths or [])]
    seen_sources = set()
    for source in paths:
        source_name = source.name
        try:
            source_key = source.resolve(strict=True)
        except (FileNotFoundError, OSError) as exc:
            raise DigitalListingTypeError(
                f"Customer file '{source_name}' không tồn tại hoặc không đọc được"
            ) from exc
        if source.is_symlink() or not source.is_file():
            raise DigitalListingTypeError(
                f"Customer file '{source_name}' phải là regular file, không phải symlink"
            )
        if source_key in seen_sources:
            raise DigitalListingTypeError(
                f"Customer file '{source_name}' bị lặp; không thể upload an toàn"
            )
        seen_sources.add(source_key)
        if source.suffix.lower() not in POST_FILE_EXTS:
            raise DigitalListingTypeError(
                f"Customer file '{source_name}' có phần mở rộng không được hỗ trợ"
            )

    stage_root = Path(tempfile.mkdtemp(prefix="etsy-customer-files-"))
    reserved_names = set()
    reserved_readback_names = set()
    staged_paths = []
    try:
        for source in paths:
            target_name = _etsy_safe_customer_filename(
                source.name,
                reserved_names=reserved_names,
                reserved_readback_names=reserved_readback_names,
            )
            target = stage_root / target_name
            try:
                shutil.copy2(source, target)
                if target.is_symlink() or not target.is_file():
                    raise OSError("staged file is not a regular file")
                if target.stat().st_size != source.stat().st_size:
                    raise OSError("staged file size mismatch")
                if not filecmp.cmp(source, target, shallow=False):
                    raise OSError("staged file bytes mismatch")
            except Exception as exc:
                raise DigitalListingTypeError(
                    f"Không stage được customer file '{source.name}' ({type(exc).__name__})"
                ) from exc

            reserved_names.add(target_name.casefold())
            reserved_readback_names.add(_normalize_customer_filename(target_name))
            staged_paths.append(target)
            if target_name != source.name:
                print(f"  🧾 Customer filename staged: '{source.name}' → '{target_name}'")
        yield staged_paths
    finally:
        shutil.rmtree(stage_root, ignore_errors=True)


async def _read_digital_file_upload_state(
    page,
    *,
    surface: dict[str, object] | None = None,
) -> dict[str, object]:
    """Read customer-file names only after the shared surface is verified."""
    if surface is None:
        surface = await _resolve_digital_upload_surface(page)
    container = surface.get("locator") if isinstance(surface, dict) else None
    if container is None:
        raise DigitalListingTypeError("Không có container Digital files đã xác minh để đọc readback")

    # Evaluate on the exact resolved container. Do not rebuild roots from the
    # document: a filename in Personalization or another form section must not
    # satisfy this upload contract.
    state = await container.evaluate(r'''root => {
        const isVisible = (el) => {
            if (!el || !el.isConnected) return false;
            const style = window.getComputedStyle(el);
            return el.offsetParent !== null && style.display !== 'none' && style.visibility !== 'hidden';
        };
        const values = new Set();
        const filePattern = /[^\n\r]{0,180}\.(?:pdf|zip)\b/ig;
        const canonicalizeFilenameRepresentation = (value) => {
            let text = String(value || '').replace(/\s+/g, ' ').trim();
            if (!text) return '';
            const actions = '(?:remove|delete|download|uploaded|completed)';
            text = text.replace(
                new RegExp('^' + actions + '\\b(?:\\s+file)?[\\s:–—-]+', 'i'),
                ''
            );
            text = text.replace(
                new RegExp('[\\s:–—-]+' + actions + '\\b(?:\\s+file)?$', 'i'),
                ''
            );
            return text.trim();
        };
        const addMatches = (value) => {
            const text = String(value || '').replace(/\s+/g, ' ').trim();
            if (!text) return;
            const matches = text.match(filePattern) || [];
            for (const match of matches) {
                const clean = canonicalizeFilenameRepresentation(match);
                if (clean) values.add(clean);
            }
        };

        let pending = false;
        let failed = false;
        let completedCount = 0;
        const nodes = [root, ...root.querySelectorAll(
            '[data-filename], [data-file-name], [data-testid*="file" i], '
            + '[class*="file-name" i], [class*="filename" i], [aria-label], [title], li, p, span'
        )];
        for (const node of nodes) {
            if (!isVisible(node)) continue;
            addMatches(node.getAttribute && node.getAttribute('data-filename'));
            addMatches(node.getAttribute && node.getAttribute('data-file-name'));
            addMatches(node.getAttribute && node.getAttribute('aria-label'));
            addMatches(node.getAttribute && node.getAttribute('title'));
            if (node.childElementCount === 0 || node === root) addMatches(node.textContent);

            const text = (node.textContent || '').replace(/\s+/g, ' ').trim();
            const stateText = [
                node.getAttribute && node.getAttribute('data-state'),
                node.getAttribute && node.getAttribute('aria-label'),
                node.className,
                text,
            ].filter(Boolean).join(' ').toLowerCase();
            if (/\b(upload failed|failed to upload|could not upload|couldn't upload|upload error)\b/.test(stateText)) {
                failed = true;
            }
            if (/\b(upload complete|uploaded|complete|completed|success)\b/.test(stateText)) {
                completedCount += 1;
            }
        }

        const pendingNodes = root.querySelectorAll(
            '[aria-busy="true"], progress, [role="progressbar"], '
            + '[data-state="loading"], [data-state="uploading"], [data-state="processing"], '
            + '[class*="spinner" i], [class*="loading" i]'
        );
        for (const node of pendingNodes) {
            if (!isVisible(node)) continue;
            const now = Number(node.getAttribute('aria-valuenow'));
            const max = Number(node.getAttribute('aria-valuemax'));
            const completeProgress = Number.isFinite(now) && Number.isFinite(max) && max > 0 && now >= max;
            const stateText = String(node.getAttribute('data-state') || '').toLowerCase();
            if (!completeProgress && !/^(complete|completed|success)$/.test(stateText)) pending = true;
        }
        for (const node of root.querySelectorAll('span, p, [role="status"]')) {
            if (!isVisible(node) || node.childElementCount > 0) continue;
            if (/^(uploading|processing|scanning)(?:\.{0,3}|\s+.*)$/i.test((node.textContent || '').trim())) {
                pending = true;
            }
        }

        const names = Array.from(values);
        return {
            hasRegion: isVisible(root),
            names,
            count: names.length,
            pending,
            failed,
            completedCount,
        };
    }''')
    if not isinstance(state, dict):
        raise DigitalListingTypeError("Etsy không trả về trạng thái Digital files hợp lệ")
    raw_names = state.get("names")
    state["names"] = _canonicalize_customer_filename_representations(
        raw_names if isinstance(raw_names, list) else []
    )
    state["count"] = len(state["names"])
    return state


def _upload_readback_matches(candidate: str, expected: str) -> bool:
    """Allow only the exact normalized filename plus common UI action wrappers."""
    if not candidate or not expected:
        return False
    if candidate == expected:
        return True
    action_prefixes = ("remove", "delete", "download", "uploaded", "completed")
    action_suffixes = ("remove", "delete", "download", "uploaded", "completed")
    return any(candidate == prefix + expected for prefix in action_prefixes) or any(
        candidate == expected + suffix for suffix in action_suffixes
    )


async def _wait_for_uploaded_digital_files(
    page,
    expected_paths,
    *,
    surface: dict[str, object] | None = None,
    timeout_ms: int = DIGITAL_FILE_UPLOAD_WAIT_MS,
    poll_ms: int = DIGITAL_FILE_UPLOAD_POLL_MS,
    stable_reads: int = DIGITAL_FILE_UPLOAD_STABLE_READS,
) -> dict[str, object]:
    """Fail closed until every requested file has stable, completed UI read-back."""
    expected_names = [Path(str(path)).name for path in (expected_paths or [])]
    expected_by_normalized = {
        _normalize_customer_filename(name): name for name in expected_names
    }
    if not expected_names or "" in expected_by_normalized:
        raise DigitalListingTypeError("Danh sách customer files cần upload không hợp lệ")
    if len(expected_by_normalized) != len(expected_names):
        raise DigitalListingTypeError(
            "Tên customer files trùng nhau sau khi chuẩn hóa; không thể xác minh upload an toàn"
        )

    poll_ms = max(1, int(poll_ms))
    stable_reads = max(1, int(stable_reads))
    attempts = max(stable_reads, max(1, int(timeout_ms) // poll_ms + 1))
    last_signature = None
    stable_count = 0
    last_state: dict[str, object] = {}
    last_missing = list(expected_names)
    if surface is None:
        try:
            surface = await _resolve_digital_upload_surface(page)
        except DigitalListingTypeError:
            raise
        except Exception as exc:
            raise DigitalListingTypeError(
                f"Không xác minh được bề mặt customer-file upload trên Etsy: {type(exc).__name__}"
            ) from exc

    for attempt in range(attempts):
        try:
            state = await _read_digital_file_upload_state(page, surface=surface)
        except DigitalListingTypeError:
            raise
        except Exception as exc:
            raise DigitalListingTypeError(
                "Không đọc được trạng thái customer-file upload từ Etsy "
                f"({type(exc).__name__})"
            ) from exc

        last_state = state
        raw_names = state.get("names")
        names = raw_names if isinstance(raw_names, list) else []
        actual_normalized = {
            _normalize_customer_filename(name) for name in names
            if _normalize_customer_filename(name)
        }
        last_missing = [
            original
            for expected, original in expected_by_normalized.items()
            if not any(
                _upload_readback_matches(candidate, expected)
                for candidate in actual_normalized
            )
        ]
        try:
            observed_count = int(state.get("count", len(actual_normalized)))
        except (TypeError, ValueError):
            observed_count = len(actual_normalized)
        signature = (
            tuple(sorted(actual_normalized)),
            observed_count,
            int(state.get("completedCount", 0) or 0),
            bool(state.get("pending")),
            bool(state.get("failed")),
        )
        ready = (
            bool(state.get("hasRegion"))
            and not last_missing
            and not bool(state.get("pending"))
            and not bool(state.get("failed"))
            and observed_count >= len(expected_names)
        )
        if ready:
            stable_count = stable_count + 1 if signature == last_signature else 1
            if stable_count >= stable_reads:
                return state
        else:
            stable_count = 0
        last_signature = signature
        if attempt + 1 < attempts:
            await page.wait_for_timeout(poll_ms)

    state_summary = {
        "missing": last_missing,
        "observed_names": last_state.get("names", []),
        "pending": bool(last_state.get("pending")),
        "failed": bool(last_state.get("failed")),
        "has_region": bool(last_state.get("hasRegion")),
    }
    raise DigitalListingTypeError(
        "Không xác minh được toàn bộ customer files đã upload hoàn tất trên Etsy: "
        f"{state_summary}"
    )


def _is_customer_file_upload_response(response) -> bool:
    """Match only Etsy's customer-file upload response for this editor flow."""
    try:
        request = response.request
        method = str(getattr(request, "method", "") or "").upper()
        url = str(getattr(response, "url", "") or "")
        status = int(getattr(response, "status", 0) or 0)
    except Exception:
        return False
    return (
        method == "POST"
        and status == 200
        and re.search(
            r"/api/v3/ajax/shop/\d+/mission-control/listing-editor/files(?:[/?]|$)",
            url,
        ) is not None
    )


async def _capture_customer_file_upload_response(page, operation) -> list[dict[str, object]]:
    """Run one upload operation and retain its exact server file receipt."""
    expect_response = getattr(page, "expect_response", None)
    if not callable(expect_response):
        raise DigitalListingTypeError(
            "Không có response contract cho customer-file upload; dừng trước khi Save"
        )

    try:
        async with expect_response(
            _is_customer_file_upload_response,
            timeout=DIGITAL_FILE_UPLOAD_RESPONSE_TIMEOUT_MS,
        ) as response_info:
            await operation()
        response = await response_info.value
        payload = await response.json()
    except DigitalListingTypeError:
        raise
    except Exception as exc:
        raise DigitalListingTypeError(
            "Không nhận được response 200 có fileId từ customer-file upload "
            f"({type(exc).__name__})"
        ) from exc

    status = int(getattr(response, "status", 0) or 0)
    if status != 200:
        raise DigitalListingTypeError(
            f"Customer-file upload trả về HTTP {status}, không thể xác minh fileId"
        )

    raw_items = payload if isinstance(payload, list) else [payload]
    receipts = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        file_id = item.get("fileId", item.get("file_id"))
        try:
            valid_file_id = int(file_id)
        except (TypeError, ValueError):
            continue
        if valid_file_id <= 0:
            continue
        receipts.append({
            "fileId": valid_file_id,
            "name": str(item.get("name") or "").strip(),
            "type": str(item.get("type") or "").strip(),
            "url": str(item.get("url") or "").strip(),
            "status": status,
        })
    if not receipts:
        raise DigitalListingTypeError(
            "Customer-file upload response không có fileId hợp lệ"
        )
    return receipts


def _digital_file_state_matches_paths(state: dict[str, object], expected_paths) -> bool:
    """Return true only when the bounded UI reports every file completed."""
    if not isinstance(state, dict):
        return False
    raw_names = state.get("names")
    actual_names = [
        _normalize_customer_filename(name)
        for name in (raw_names if isinstance(raw_names, list) else [])
        if _normalize_customer_filename(name)
    ]
    expected_names = [
        _normalize_customer_filename(Path(str(path)).name)
        for path in (expected_paths or [])
    ]
    try:
        count = int(state.get("count", len(actual_names)) or 0)
    except (TypeError, ValueError):
        count = len(actual_names)
    return (
        bool(state.get("hasRegion"))
        and not bool(state.get("pending"))
        and not bool(state.get("failed"))
        and count >= len(expected_names)
        and all(
            any(_upload_readback_matches(candidate, expected) for candidate in actual_names)
            for expected in expected_names
        )
    )


async def _reconcile_digital_files_on_item_details(page, product) -> None:
    """Persist customer files on the stable Item Details editor surface."""
    expected_paths = product.get("pdf_paths") or []
    if not expected_paths:
        return

    try:
        initial_surface = await asyncio.wait_for(
            _find_visible_digital_files_container(page),
            timeout=DIGITAL_UPLOAD_SURFACE_WAIT_MS / 1000,
        )
        surface_locator = await asyncio.wait_for(
            _establish_stable_digital_upload_surface(
                page,
                initial_surface=initial_surface,
            ),
            timeout=DIGITAL_UPLOAD_SURFACE_WAIT_MS / 1000,
        )
    except asyncio.TimeoutError as exc:
        raise DigitalListingTypeError(
            "Không xác minh được bề mặt Digital files ổn định trên Item Details "
            f"trong {DIGITAL_UPLOAD_SURFACE_WAIT_MS}ms"
        ) from exc
    if surface_locator is None:
        raise DigitalListingTypeError(
            "Không xác minh được bề mặt Digital files ổn định trên Item Details; "
            "dừng trước khi Save"
        )

    surface = {"locator": surface_locator}
    try:
        current_state = await asyncio.wait_for(
            _read_digital_file_upload_state(page, surface=surface),
            timeout=DIGITAL_UPLOAD_SURFACE_WAIT_MS / 1000,
        )
    except DigitalListingTypeError:
        raise
    except Exception as exc:
        raise DigitalListingTypeError(
            "Không đọc được customer-file state trên Item Details "
            f"({type(exc).__name__})"
        ) from exc

    if _digital_file_state_matches_paths(current_state, expected_paths):
        print("  ✅ Customer files đã có trên Item Details ổn định — không upload trùng")
        product["_digital_files_uploaded"] = True
        return

    if bool(current_state.get("pending")) or bool(current_state.get("failed")):
        raise DigitalListingTypeError(
            "Customer-file state trên Item Details đang pending/failed; "
            "dừng để tránh upload trùng"
        )

    print("  🔁 Customer files chưa có trên Item Details ổn định — upload/reconcile")
    await upload_digital_files(page, product, verified_surface=surface)
    product["_digital_files_uploaded"] = True


async def _read_saved_draft_digital_file_names(
    page,
    *,
    alias_groups=None,
) -> list[str]:
    """Poll the exact saved Digital files DOM until it is readable and complete."""
    deadline = asyncio.get_running_loop().time() + (
        DIGITAL_SAVED_DRAFT_READBACK_TIMEOUT_MS / 1000
    )
    last_reason = "surface chưa hiển thị"
    while True:
        remaining = deadline - asyncio.get_running_loop().time()
        if remaining <= 0:
            break
        try:
            surface_locator = await asyncio.wait_for(
                _find_visible_digital_files_container(page),
                timeout=remaining,
            )
            if surface_locator is None:
                last_reason = "surface chưa hiển thị"
            else:
                state_remaining = deadline - asyncio.get_running_loop().time()
                if state_remaining <= 0:
                    break
                state = await asyncio.wait_for(
                    _read_digital_file_upload_state(
                        page,
                        surface={"locator": surface_locator},
                    ),
                    timeout=state_remaining,
                )
                if not isinstance(state, dict) or not state.get("hasRegion"):
                    last_reason = "surface không còn visible"
                elif bool(state.get("pending")):
                    last_reason = "upload pending"
                elif bool(state.get("failed")):
                    last_reason = "upload failed"
                else:
                    names = [
                        _normalize_customer_filename(name)
                        for name in (state.get("names") if isinstance(state.get("names"), list) else [])
                        if _normalize_customer_filename(name)
                    ]
                    if names and (
                        alias_groups is None
                        or _saved_names_match_alias_groups(names, alias_groups)
                    ):
                        return names
                    last_reason = "filename readback chưa hoàn tất"
        except asyncio.TimeoutError:
            break
        except DigitalListingTypeError as exc:
            last_reason = str(exc)
        except Exception as exc:
            last_reason = f"{type(exc).__name__}"

        remaining = deadline - asyncio.get_running_loop().time()
        if remaining <= 0:
            break
        await page.wait_for_timeout(
            min(DIGITAL_SAVED_DRAFT_READBACK_POLL_MS, max(1, int(remaining * 1000)))
        )
    raise DigitalListingTypeError(
        "Không xác minh được bounded Digital files DOM sau khi Save "
        f"trong {DIGITAL_SAVED_DRAFT_READBACK_TIMEOUT_MS}ms ({last_reason})"
    )


def _saved_names_match_alias_groups(saved_names, alias_groups) -> bool:
    """Require an injective saved-DOM filename match for each logical source."""
    normalized_saved = _canonicalize_customer_filename_representations(saved_names)
    groups = [
        set(_canonicalize_customer_filename_representations(group.get("aliases") or []))
        for group in (alias_groups or [])
        if isinstance(group, dict) and group.get("aliases")
    ]
    if not groups or len(normalized_saved) < len(groups):
        return False

    candidates = []
    for aliases in groups:
        candidates.append([
            index for index, saved in enumerate(normalized_saved)
            if any(_upload_readback_matches(saved, alias) for alias in aliases)
        ])
    if any(not indexes for indexes in candidates):
        return False

    def assign(group_index, used):
        if group_index >= len(candidates):
            return True
        for saved_index in candidates[group_index]:
            if saved_index in used:
                continue
            if assign(group_index + 1, used | {saved_index}):
                return True
        return False

    return assign(0, set())


async def _verify_saved_draft_digital_files(page, product) -> None:
    """Fail closed unless every logical source file is present in saved DOM."""
    alias_groups = product.get("_digital_file_alias_groups") or []
    if not alias_groups:
        alias_groups = _customer_file_alias_groups(product.get("pdf_paths") or [])

    receipts = product.get("_digital_file_upload_receipts") or []
    if receipts:
        alias_groups = _add_customer_receipt_aliases(alias_groups, receipts)
    saved_names = await _read_saved_draft_digital_file_names(
        page,
        alias_groups=alias_groups,
    )
    if not _saved_names_match_alias_groups(saved_names, alias_groups):
        raise DigitalListingTypeError(
            "Draft đã Save nhưng bounded Digital files DOM không chứa đủ customer files"
        )
    print("  ✅ Draft readback: bounded Digital files DOM có đủ customer files")


async def _discover_new_draft_id(page, product: dict) -> str | None:
    """Find one post-save draft outside the pre-create baseline."""
    if "_draft_ids_before_create" not in product:
        return None
    baseline_ids = {
        str(listing_id).strip()
        for listing_id in product.get("_draft_ids_before_create", [])
        if str(listing_id).strip()
    }

    for attempt in range(2):
        cards = await _collect_draft_cards(page)
        new_cards = [
            card for card in cards
            if isinstance(card, dict)
            and str(card.get("id", "")).strip()
            and str(card.get("id", "")).strip() not in baseline_ids
        ]
        if len(new_cards) == 1:
            return str(new_cards[0]["id"])
        # Filter against the baseline before applying title/SKU matching; an
        # old draft with the same title must not win this lookup.
        matched_id = _pick_draft_card_id(new_cards, product)
        if matched_id:
            return matched_id
        if attempt == 0:
            await page.wait_for_timeout(3000)
    return None


def _known_listing_id_from_edit_url(edit_url: str | None) -> str | None:
    """Extract the exact existing editor ID for an explicit edit flow."""
    match = re.search(r"/listing-editor/edit/(\d+)(?:[/?#]|$)", str(edit_url or ""))
    return match.group(1) if match else None


async def _click_verified_item_details_tab(page) -> bool:
    """Fallback only to the exact Item Details tab; never broad-click Details."""
    try:
        tabs = page.get_by_role("tab", name=re.compile(r"^Item Details$", re.IGNORECASE))
        visible = await _visible_locator_items(tabs)
        if len(visible) == 1:
            await visible[0].click()
            await page.wait_for_timeout(500)
            return True
    except Exception:
        pass

    # Some Etsy variants expose tabs as plain links/buttons. Keep the exact
    # accessible/text match and reject the generic Details label entirely.
    try:
        candidates = page.locator('a, button, [role="tab"]')
        matches = []
        for idx in range(await candidates.count()):
            candidate = candidates.nth(idx)
            if not await candidate.is_visible():
                continue
            try:
                text = _normalize_ui_text(await candidate.inner_text())
            except Exception:
                text = ""
            try:
                aria_label = _normalize_ui_text(await candidate.get_attribute("aria-label"))
            except Exception:
                aria_label = ""
            if re.fullmatch(r"Item Details", text, re.IGNORECASE) or re.fullmatch(
                r"Item Details", aria_label, re.IGNORECASE
            ):
                matches.append(candidate)
        if len(matches) == 1:
            await matches[0].click()
            await page.wait_for_timeout(500)
            return True
    except Exception:
        pass
    return False


async def _is_likely_live_surface(
    page,
    *,
    locator: object | None,
    attempts: int = 3,
    wait_ms: int = 120,
) -> bool:
    """Bounded liveness probe on a bounded locator.

    Prefer exact surface checks without re-running the broad resolve flow.
    """
    if locator is None or not hasattr(locator, "is_visible"):
        return False

    for _ in range(max(1, int(attempts))):
        try:
            visible = await asyncio.wait_for(
                locator.is_visible(),
                timeout=DIGITAL_SURFACE_LIVENESS_TIMEOUT_MS / 1000,
            )
            if bool(visible):
                return True
        except Exception:
            pass
        await page.wait_for_timeout(max(0, int(wait_ms)))
    return False


async def upload_digital_files(page, product, *, verified_surface=None):
    if not product["pdf_paths"]:
        return
    if verified_surface is not None and not isinstance(verified_surface, dict):
        raise DigitalListingTypeError("Đầu vào bề mặt upload đã xác minh không đúng kiểu dữ liệu")
    try:
        with _stage_etsy_customer_files(product["pdf_paths"]) as staged_paths:
            upload_receipts: list[dict[str, object]] = []
            alias_groups = _customer_file_alias_groups(
                product["pdf_paths"],
                staged_paths,
            )
            product["_digital_file_alias_groups"] = alias_groups
            # Kiểm tra kích thước file
            for path in product["pdf_paths"]:
                size_mb = Path(path).stat().st_size / (1024 * 1024)
                if size_mb > 20:
                    print(f"  ❌ CẢNH BÁO: File '{Path(path).name}' ({size_mb:.2f} MB) vượt quá giới hạn 20MB của Etsy!")
                    print("     • Etsy chỉ cho phép upload file dưới 20MB.")
                    print("     • Hướng xử lý: Hãy nén file PDF lại, hoặc đổi sang upload 1 file PDF/TXT hướng dẫn có chứa link tải từ Google Drive.")
            surface: dict[str, object]
            if verified_surface is None:
                await dismiss_alerts(page)
                # Keep the current tab when React has already rendered a positive
                # upload surface. Only a missing surface permits an exact tab fallback.
                try:
                    surface = await _resolve_digital_upload_surface(page)
                except DigitalListingTypeError:
                    if not await _click_verified_item_details_tab(page):
                        raise DigitalListingTypeError(
                            "Không có bề mặt upload customer files trong view hiện tại "
                            "và không tìm thấy tab Item Details chính xác"
                        )
                    surface = await _resolve_digital_upload_surface(page)
            else:
                if "locator" not in verified_surface:
                    raise DigitalListingTypeError("Bề mặt Digital files đã xác minh thiếu locator")
                surface = verified_surface
                print("  ✅ Reuse bề mặt Digital files đã xác minh từ Category")
                container = surface.get("locator")
                if not await _is_likely_live_surface(
                    page,
                    locator=container,
                ):
                    raise DigitalListingTypeError(
                        "Bề mặt Digital files đã xác minh đã biến mất"
                    )

            # Step 2: Scroll the exact Add file affordance when one exists.
            if verified_surface is None:
                add_btn = await _find_exact_add_file_affordance(
                    page,
                    container=surface["locator"],
                )
            else:
                try:
                    add_btn = await asyncio.wait_for(
                        _find_exact_add_file_affordance(
                            page,
                            container=surface["locator"],
                            visibility_timeout_ms=DIGITAL_ADD_FILE_VISIBILITY_TIMEOUT_MS,
                        ),
                        timeout=DIGITAL_ADD_FILE_VISIBILITY_TIMEOUT_MS / 1000,
                    )
                except asyncio.TimeoutError:
                    print(
                        "  ⚠ Add file discovery vượt quá "
                        f"{DIGITAL_ADD_FILE_VISIBILITY_TIMEOUT_MS}ms — tiếp tục input scoped"
                    )
                    add_btn = None
            if add_btn is not None:
                try:
                    if verified_surface is None:
                        # Preserve the legacy caller behavior, including its
                        # default Playwright timeout and failure semantics.
                        await add_btn.scroll_into_view_if_needed()
                    else:
                        await add_btn.scroll_into_view_if_needed(
                            timeout=DIGITAL_ADD_FILE_SCROLL_TIMEOUT_MS
                        )
                except PlaywrightTimeoutError:
                    if verified_surface is None:
                        raise
                    print(
                        "  ⚠ Không scroll được 'Add file' trong "
                        f"{DIGITAL_ADD_FILE_SCROLL_TIMEOUT_MS}ms — tiếp tục surface đã xác minh"
                    )
                else:
                    if verified_surface is None:
                        await page.wait_for_timeout(500)

            # Step 3: Use the file chooser, then require positive UI read-back.
            btn_disabled = False
            if add_btn is not None:
                if verified_surface is None:
                    # Preserve legacy caller behavior and its default timeout.
                    btn_disabled = await add_btn.is_disabled()

            if add_btn is not None and (verified_surface is not None or not btn_disabled):
                chooser_set_files = False
                try:
                    async with page.expect_file_chooser(timeout=10000) as fc_info:
                        if verified_surface is None:
                            # Preserve legacy caller behavior and its default timeout.
                            await add_btn.click()
                        else:
                            await asyncio.wait_for(
                                add_btn.click(),
                                timeout=DIGITAL_ADD_FILE_CLICK_TIMEOUT_MS / 1000,
                            )
                    file_chooser = await fc_info.value
                    upload_receipts = await _capture_customer_file_upload_response(
                        page,
                        lambda: file_chooser.set_files(staged_paths),
                    )
                    chooser_set_files = True
                    print(f"  ⏳ Đợi xác minh {len(staged_paths)} customer file(s) trên Etsy...")
                    await _wait_for_uploaded_digital_files(
                        page,
                        staged_paths,
                        surface=surface,
                    )
                    product["_digital_file_upload_receipts"] = upload_receipts
                    product["_digital_file_alias_groups"] = _add_customer_receipt_aliases(
                        alias_groups,
                        upload_receipts,
                    )
                    print(f"  📎 {len(staged_paths)} file(s) ✓ (đã xác minh trên UI)")
                    return
                except DigitalListingTypeError:
                    raise
                except Exception as fc_err:
                    if chooser_set_files:
                        raise DigitalListingTypeError(
                            "Đã gửi customer files nhưng không xác minh được upload "
                            f"({type(fc_err).__name__})"
                        ) from fc_err
                    print(
                        "  ⚠ File chooser failed "
                        f"({type(fc_err).__name__}) — trying direct input..."
                    )
            elif add_btn is not None:
                print("  ℹ️  'Add file' button disabled — using direct input fallback")
            else:
                print("  ℹ️  No exact 'Add file' affordance — using scoped input fallback")

            # Step 4: Fallback — a direct input is successful only after the same
            # complete filename read-back contract passes.
            direct_errors = []
            for idx, inp in await _find_scoped_customer_file_inputs(
                page,
                container=surface["locator"],
            ):
                try:
                    upload_receipts = await _capture_customer_file_upload_response(
                        page,
                        lambda: inp.set_input_files(staged_paths, timeout=15000),
                    )
                    await _wait_for_uploaded_digital_files(
                        page,
                        staged_paths,
                        surface=surface,
                    )
                    product["_digital_file_upload_receipts"] = upload_receipts
                    product["_digital_file_alias_groups"] = _add_customer_receipt_aliases(
                        alias_groups,
                        upload_receipts,
                    )
                    print(
                        f"  📎 {len(staged_paths)} file(s) ✓ "
                        f"(đã xác minh trên UI, input #{idx})"
                    )
                    return
                except DigitalListingTypeError:
                    raise
                except Exception as direct_err:
                    direct_errors.append(f"input #{idx}: {type(direct_err).__name__}")
                    continue

            detail = "; ".join(direct_errors) if direct_errors else "không tìm thấy file input"
            raise DigitalListingTypeError(
                f"Không upload được customer files bằng file chooser hoặc direct input ({detail})"
            )
    except DigitalListingTypeError:
        raise
    except Exception as e:
        raise DigitalListingTypeError(
            f"Customer-file upload thất bại ({type(e).__name__})"
        ) from e




# ── Translations ───────────────────────────────────────────────────────────────
async def fill_translations(page, product):
    print("  🌐 Đang dịch và điền Translations...")
    try:
        h = page.locator('text="Translations"').first
        if await h.count() > 0:
            await h.scroll_into_view_if_needed()
            await page.wait_for_timeout(800)
    except Exception:
        pass

    en_tags = clean_tags(product["tags"])

    for lang_code, lang_name, idx in LANGUAGES:
        try:
            # Click tab ngôn ngữ
            clicked = False
            for sel in [f'button:has-text("{lang_name}")',
                        f'[role="tab"]:has-text("{lang_name}")',
                        f'li:has-text("{lang_name}")']:
                el = page.locator(sel).first
                if await el.count() > 0 and await el.is_visible():
                    await el.scroll_into_view_if_needed()
                    await el.click()
                    await page.wait_for_timeout(1000)
                    clicked = True
                    break
            if not clicked:
                print(f"    ⚠ tab {lang_name} không thấy")
                continue

            # Dịch title và description
            trans_title = trim_title(translate_text(product["title"], lang_code))
            trans_desc  = translate_text(product["description"], lang_code)

            # Điền title
            for sel in [f'textarea[name="translations.{idx}.title"]',
                        f'#field-translations-{idx}-title-input',
                        f'textarea[id*="translations"][id*="{idx}"][id*="title"]',
                        f'textarea[id*="{lang_code}"][id*="title"]']:
                el = page.locator(sel).first
                if await el.count() > 0 and await el.is_visible():
                    await clear_and_fill(el, trans_title)
                    await page.wait_for_timeout(400)
                    break

            # Điền description
            for sel in [f'textarea[name="translations.{idx}.description"]',
                        f'#listing-{lang_code}-translation-description-textarea',
                        f'textarea[id*="{lang_code}"][id*="description"]',
                        f'textarea[id*="translations"][id*="{idx}"][id*="description"]']:
                el = page.locator(sel).first
                if await el.count() > 0 and await el.is_visible():
                    await clear_and_fill(el, trans_desc)
                    await page.wait_for_timeout(400)
                    break

            # Điền tags dịch
            try:
                # Delete existing translation tags for this language
                try:
                    await page.evaluate('''() => {
                        let activePanel = document.querySelector('[role="tabpanel"]:not([class*="hidden"]), [role="tabpanel"]:not([style*="display: none"])');
                        if (activePanel) {
                            let removeButtons = Array.from(activePanel.querySelectorAll('span.wt-tag button, [data-testid="tag-pill"] button, button[aria-label*="remove" i], button[aria-label*="delete" i], button[class*="remove" i]'));
                            for (let btn of removeButtons) {
                                btn.click();
                            }
                        }
                    }''')
                    await page.wait_for_timeout(500)
                except Exception:
                    pass

                tag_input_sel = (
                    f'input[name="translations.{idx}.tags"], '
                    f'input[id*="{lang_code}"][id*="tag"], '
                    f'input[id*="translations"][id*="{idx}"][id*="tag"], '
                    f'input[placeholder*="tag" i]'
                )
                tag_input = page.locator(tag_input_sel).first
                if await tag_input.count() > 0 and await tag_input.is_visible() and await tag_input.is_editable():
                    seen_trans_tags = set()
                    try:
                        existing_pills = await page.evaluate('''() => {
                            let activePanel = document.querySelector('[role="tabpanel"]:not([class*="hidden"]), [role="tabpanel"]:not([style*="display: none"])');
                            if (!activePanel) return [];
                            let pills = Array.from(activePanel.querySelectorAll('span.wt-tag, [data-testid="tag-pill"]'));
                            return pills.map(p => (p.innerText || '').trim().toLowerCase()).filter(Boolean);
                        }''')
                        for ep in (existing_pills or []):
                            seen_trans_tags.add(ep)
                    except Exception:
                        pass

                    tags_filled = 0
                    for tag in en_tags:
                        trans_tag = translate_tag(tag, lang_code)
                        if not trans_tag:
                            continue
                        norm_tag = trans_tag.strip().lower()
                        if norm_tag in seen_trans_tags:
                            continue
                        seen_trans_tags.add(norm_tag)

                        try:
                            await tag_input.fill(trans_tag, timeout=3000)
                            await page.wait_for_timeout(300)
                            await tag_input.press("Enter")
                            await page.wait_for_timeout(500)
                            tags_filled += 1
                        except Exception as fill_ex:
                            print(f"      ⚠️ Không điền được trans tag '{trans_tag}': {fill_ex}")
                    if tags_filled:
                        print(f"    🌍 {lang_name} ✓ ({tags_filled} tags)")
                    else:
                        print(f"    🌍 {lang_name} ✓")
                else:
                    print(f"    🌍 {lang_name} ✓")
            except Exception as te:
                print(f"    🌍 {lang_name} ✓ (tags: {te})")

        except Exception as e:
            print(f"    ⚠ {lang_name}: {e}")

        await page.wait_for_timeout(500)

# ── Fill One Listing ───────────────────────────────────────────────────────────
async def check_duplicate_draft(page, product: dict) -> bool:
    """Quét Drafts và kiểm tra trùng lặp theo SKU/tiêu đề cho đúng sản phẩm."""
    try:
        print("  🔍 Đang kiểm tra trùng lặp trên Etsy Drafts...")
        drafts = await _collect_draft_cards(page)
        product["_draft_ids_before_create"] = sorted({
            str(card.get("id", "")).strip()
            for card in drafts
            if isinstance(card, dict) and str(card.get("id", "")).strip()
        })
        title = str(product.get("title", ""))
        if not title:
            return False

        matched_id = _pick_draft_card_id(drafts, product)
        if matched_id:
            print(f"  ⚠️ Trùng lặp phát hiện: listing {matched_id}")
            return True

        print("  ✓ Không phát hiện trùng lặp trên Etsy.")
        return False
    except Exception as e:
        print(f"  ⚠ Lỗi kiểm tra trùng lặp: {e}")
        return DRAFT_DUPLICATE_CHECK_FAILED_SENTINEL


async def get_newly_created_listing_url(page, product: dict):
    """
    Lấy URL listing vừa tạo từ editor hoặc danh sách Drafts theo cơ chế fail-safe.
    Trả về string URL (e.g. 'https://www.etsy.com/listing/4509048784'),
    hoặc UNVERIFIED_DRAFT_URL_SENTINEL nếu chưa xác minh được.
    """
    import re
    target_title = str(product.get("title", ""))

    # Phương pháp 0: Dùng Listing ID đã capture từ URL redirect khi lưu draft.
    # Đây là nguồn tin cậy nhất vì Etsy redirect sang edit/<id> ngay sau khi lưu.
    captured_id = product.get("_captured_listing_id_from_redirect")
    if captured_id and str(captured_id).strip():
        lid = str(captured_id).strip()
        # Đợi trang editor load xong rồi xác thực chữ ký sản phẩm
        await page.wait_for_timeout(2000)
        for _retry in range(3):
            if await _editor_product_signature_matches(page, product):
                print(f"  🎯 Lấy được Listing ID từ redirect URL (đã xác thực): {lid}")
                return f"https://www.etsy.com/listing/{lid}"
            await page.wait_for_timeout(1500)
        # Nếu không xác thực được chữ ký nhưng ID đã capture từ redirect thì vẫn đáng tin
        print(f"  🎯 Dùng Listing ID từ redirect URL (không xác thực được chữ ký, vẫn dùng): {lid}")
        return f"https://www.etsy.com/listing/{lid}"

    # Phương pháp 1: Dùng trực tiếp URL hiện tại nếu đang ở trang listing editor/listing
    # và editor đang khớp đúng sản phẩm hiện tại.
    if target_title:
        current_url = page.url
        match = re.search(r'(?:edit|listing)/(\d+)', current_url)
        if match:
            lid = match.group(1)
            # Đợi trang load rồi thử xác thực
            await page.wait_for_timeout(2000)
            for _retry in range(3):
                if await _editor_product_signature_matches(page, product):
                    print(f"  🎯 Lấy được Listing ID từ URL editor (đã xác thực sản phẩm): {lid}")
                    return f"https://www.etsy.com/listing/{lid}"
                await page.wait_for_timeout(1500)
            # URL có chứa edit/<id> ngay sau khi lưu → rất có thể là đúng listing
            print(f"  🎯 Dùng Listing ID từ URL editor (không xác thực chữ ký, vẫn dùng): {lid}")
            return f"https://www.etsy.com/listing/{lid}"

    # Phương pháp 2: Quét Drafts để chọn đúng listing vừa tạo.
    # Đợi một chút để Etsy index draft mới trước khi quét.
    await page.wait_for_timeout(3000)
    try:
        links = await _collect_draft_cards(page)
        baseline_ids = {
            str(listing_id).strip()
            for listing_id in product.get("_draft_ids_before_create", [])
            if str(listing_id).strip()
        }
        candidate_links = [
            card for card in links
            if not baseline_ids
            or str(card.get("id", "")).strip() not in baseline_ids
        ]
        # Never let an existing baseline draft win title/SKU matching.
        matched_id = _pick_draft_card_id(candidate_links, product)
        if matched_id:
            print(f"  🎯 Quét Drafts chọn được Listing ID duy nhất: {matched_id}")
            return f"https://www.etsy.com/listing/{matched_id}"

        if "_draft_ids_before_create" in product:
            current_ids = {
                str(card.get("id", "")).strip()
                for card in links
                if isinstance(card, dict) and str(card.get("id", "")).strip()
            }
            new_ids = sorted(current_ids - baseline_ids)
            if len(new_ids) == 1:
                new_id = new_ids[0]
                print(f"  🎯 Xác minh được Listing ID mới duy nhất so với baseline Drafts: {new_id}")
                return f"https://www.etsy.com/listing/{new_id}"
            if len(new_ids) > 1:
                print(f"  ⚠ Có {len(new_ids)} Draft ID mới; không tự động chọn để tránh sai listing.")
    except Exception as e:
        print(f"  ⚠ Lỗi khi quét tìm Listing ID: {e}")

    # Phương pháp 3: Retry quét Drafts thêm 1 lần sau khi đợi thêm
    try:
        print("  🔄 Thử quét Drafts lần 2 sau 5s...")
        await page.wait_for_timeout(5000)
        links = await _collect_draft_cards(page)
        baseline_ids = {
            str(listing_id).strip()
            for listing_id in product.get("_draft_ids_before_create", [])
            if str(listing_id).strip()
        }
        candidate_links = [
            card for card in links
            if not baseline_ids
            or str(card.get("id", "")).strip() not in baseline_ids
        ]
        matched_id = _pick_draft_card_id(candidate_links, product)
        if matched_id:
            print(f"  🎯 Quét Drafts lần 2 chọn được Listing ID: {matched_id}")
            return f"https://www.etsy.com/listing/{matched_id}"
    except Exception as e:
        print(f"  ⚠ Lỗi khi quét Drafts lần 2: {e}")

    print("  ⚠ Không xác minh được listing URL sau lưu draft.")
    return UNVERIFIED_DRAFT_URL_SENTINEL


async def fill_listing(page, product, edit_url=None):
    print(f"\n{'─'*55}")
    print(f"  📦 {product['folder']} | {len(product['image_paths'])} ảnh | {len(product['pdf_paths'])} PDF")
    print(f"     {product['title'][:60]}...")

    # Early local validations to save time and provide clear feedback
    early_errors = []
    if product.get("price", 0) < 0.20:
        early_errors.append(f"Giá bán ${product.get('price', 0):.2f} không hợp lệ (tối thiểu $0.20)")
    
    for path in product.get("pdf_paths", []):
        size_mb = Path(path).stat().st_size / (1024 * 1024)
        if size_mb > 20:
            early_errors.append(f"File PDF '{Path(path).name}' ({size_mb:.2f} MB) vượt quá giới hạn 20MB")
            
    if early_errors:
        err_msg = " | ".join(early_errors)
        print(f"  ❌ Lỗi validate sớm: {err_msg}")
        return False, err_msg

    # Never trust a marker supplied by a prior attempt or caller. This run must
    # prove the editor's Digital state before it is allowed to reach Save.
    product.pop("_digital_listing_type_verified", None)
    product.pop("_digital_files_uploaded", None)
    product.pop("_digital_file_upload_receipts", None)
    product.pop("_digital_file_alias_groups", None)

    if edit_url:
        print(f"  📝 Đang sửa listing tại {edit_url}...")
        await page.goto(edit_url, wait_until="domcontentloaded")
    else:
        # ── Kiểm tra trùng draft trên Etsy trước khi tạo mới ──────────────────
        is_dup = await check_duplicate_draft(page, product)
        if is_dup == DRAFT_DUPLICATE_CHECK_FAILED_SENTINEL:
            return False, "Không thể xác minh trùng lặp draft trước khi tạo mới, dừng để tránh lỗi trùng"
        if is_dup:
            return "DUPLICATE"   # Caller sẽ mark Excel là đã đăng

        await page.goto("https://www.etsy.com/your/shops/me/listing-editor/create",
                        wait_until="domcontentloaded")
    await page.wait_for_timeout(5000)

    if "listing-editor" not in page.url:
        print(f"  ❌ Không vào được editor. URL hiện tại: {page.url}")
        return False, "Không vào được Etsy listing editor"

    await page.wait_for_timeout(2500)


    form_type = await detect_form_type(page)
    print(f"  📋 Form: {form_type}")

    if form_type == "tabs":
        await fill_photo_tab(page, product, explicit_edit=bool(edit_url))
        await page.wait_for_timeout(1000)

        await fill_category_tab(page, product)
        await page.wait_for_timeout(1000)

        await fill_item_details_tab(page, product)
        await page.wait_for_timeout(1000)

        await fill_item_options_tab(page, product)
        await page.wait_for_timeout(1000)

        await fill_pricing_tab(page, product)
        await page.wait_for_timeout(1000)

    else:
        # ── Form trang dài cũ (fallback) ──────────────────────────────────
        product["_digital_listing_type_verified"] = await select_and_verify_digital_listing_type(page)
        print("  💻 Digital ✓")
        if product["pdf_paths"]:
            raise DigitalListingTypeError(
                "Form Etsy trang dài không có luồng upload customer files đã xác minh; "
                "dừng trước khi Save"
            )

        ok = await smart_fill(page, '#listing-title-input, textarea[name="title"]', product["title"])
        print(f"  📝 Title {'✓' if ok else '⚠ không điền được'}")
        await page.wait_for_timeout(500)
        ok = await smart_fill(page, '#listing-description-textarea, textarea[name="description"]',
                              product["description"], timeout=6000)
        print(f"  📄 Description {'✓' if ok else '⚠ không điền được'}")
        await page.wait_for_timeout(500)

        try:
            tags = clean_tags(product["tags"])
            filled = 0
            for tag in tags:
                el = page.locator('#listing-tags-input').first
                if await el.is_visible():
                    await el.fill(tag); await el.press("Enter")
                    await page.wait_for_timeout(400); filled += 1
            print(f"  🏷  {filled} tags ✓")
        except Exception as e: print(f"  ⚠ tags: {e}")

        ok = await smart_fill(page, '#listing-price-input, [data-testid="price-input"]', f"{product['price']:.2f}")
        print(f"  💲 Price {'✓' if ok else '⚠ không điền được'}")
        ok = await smart_fill(page, '#listing-quantity-input, input[name="quantity"]', str(product["qty"]))
        print(f"  🔢 Qty {'✓' if ok else '⚠ không điền được'}")

        if "sku" in product and product["sku"]:
            ok = await smart_fill(page, '#listing-sku-input, input[name="sku"], [data-testid="sku-input"]', str(product["sku"]))
            print(f"  🔑 SKU: {product['sku']} {'✓' if ok else '⚠ không điền được'}")

        if product["image_paths"]:
            before_count = await _count_listing_image_thumbs(page)
            expected = before_count + len(product["image_paths"][:10])
            await _upload_listing_photos_until_count(
                page,
                product["image_paths"],
                expected_total=expected,
                exact=False,
            )

    if not product.get("_digital_listing_type_verified"):
        raise DigitalListingTypeError(
            "Chưa xác minh được listing ở trạng thái Digital; dừng trước khi Save"
        )

    # Translations (chạy cuối, sau khi đã điền hết)
    if FILL_TRANSLATIONS:
        await page.wait_for_timeout(1000)
        await fill_translations(page, product)

    # Save listing
    await page.wait_for_timeout(1500)
    try:
        btn = page.locator(_get_save_button_selector(explicit_edit=bool(edit_url))).first
        await btn.wait_for(state="visible", timeout=10000)
        await btn.click()
        
        print("  ⏳ Đang đợi Etsy xử lý lưu bản nháp...")
        
        # Check for validation errors and wait for redirect
        errors_found = []
        saved_successfully = False
        _redirect_listing_id = None  # Capture listing ID from URL redirect
        
        for _wait in range(30):
            await page.wait_for_timeout(1000)
            
            # Evaluate visible validation errors using JS
            try:
                found_msgs = await page.evaluate('''() => {
                    const list = [];
                    const selectors = [
                        '.wt-validation__message:visible',
                        '.wt-alert--error:visible',
                        '[class*="error-message"]:visible',
                        '[class*="validation-message"]:visible',
                        '[role="alert"]:visible',
                        'span.wt-text-danger',
                        'p.wt-text-danger',
                        '[class*="wt-validation__message"]'
                    ];
                    for (const sel of selectors) {
                        try {
                            const els = document.querySelectorAll(sel);
                            for (const el of els) {
                                if (el.getBoundingClientRect().height > 0) {
                                    const txt = el.innerText.trim();
                                    if (txt && !list.includes(txt) && !txt.toLowerCase().includes("loading") && txt.length < 250) {
                                        list.push(txt);
                                    }
                                }
                            }
                        } catch(e) {}
                    }
                    return list;
                }''')
                for msg in found_msgs:
                    if msg not in errors_found:
                        errors_found.append(msg)
            except Exception:
                pass
            
            if errors_found:
                break
                
            # Check if URL changed (saved successfully)
            if "edit/" in page.url or "listings" in page.url:
                saved_successfully = True
                # Capture listing ID from redirect URL (e.g. listing-editor/edit/12345678)
                _id_match = re.search(r'(?:edit|listing)/(\d+)', page.url)
                if _id_match:
                    _redirect_listing_id = _id_match.group(1)
                break

        if not saved_successfully and not errors_found:
            # Check if there is some generic error text on screen
            try:
                body_text = await page.inner_text("body")
                for err_word in ["error", "required", "invalid", "lỗi", "bắt buộc", "cannot"]:
                    if err_word in body_text.lower():
                        errors_found.append(f"Phát hiện từ khóa lỗi '{err_word}'")
            except Exception:
                pass

        if errors_found or not saved_successfully:
            print(f"  ❌ Save draft failed! saved_successfully={saved_successfully}")
            for err in errors_found:
                print(f"     • {err}")
            # Take screenshot to help debug
            try:
                screenshot_path = str(BASE_DIR / "save_draft_failure.png")
                await page.screenshot(path=screenshot_path, full_page=True)
                print(f"  📸 Đã chụp ảnh màn hình lỗi tại: {screenshot_path}")
            except Exception as ss_ex:
                print(f"  ⚠ Không thể chụp ảnh màn hình: {ss_ex}")
            err_reason = " | ".join(errors_found) if errors_found else "Lưu bản nháp thất bại"
            return False, err_reason

        if product.get("pdf_paths"):
            # Re-open the exact saved editor before reporting success. The
            # pre-save DOM is not sufficient evidence: Etsy can unmount the
            # Category root and discard an upload that was never persisted.
            if not _redirect_listing_id:
                current_id = re.search(r'(?:edit|listing)/(\d+)', page.url)
                if current_id:
                    _redirect_listing_id = current_id.group(1)
            if not _redirect_listing_id and edit_url:
                _redirect_listing_id = _known_listing_id_from_edit_url(edit_url)
            if not _redirect_listing_id and "tools/listings" in page.url:
                _redirect_listing_id = await _discover_new_draft_id(page, product)
            if not _redirect_listing_id or not hasattr(page, "goto"):
                raise DigitalListingTypeError(
                    "Không xác định được editor ID để readback digitalFiles sau khi Save"
                )
            exact_saved_editor_url = (
                "https://www.etsy.com/your/shops/me/listing-editor/edit/"
                f"{_redirect_listing_id}"
            )
            await page.goto(
                exact_saved_editor_url,
                wait_until="domcontentloaded",
                timeout=DIGITAL_SAVED_DRAFT_READBACK_TIMEOUT_MS,
            )
            await page.wait_for_timeout(1000)
            await _verify_saved_draft_digital_files(page, product)

        print("  💾 Saved as draft ✅")
        # Trích xuất và trả về URL của listing vừa tạo để lưu vào Excel
        if _redirect_listing_id:
            product["_captured_listing_id_from_redirect"] = _redirect_listing_id
        new_url = await get_newly_created_listing_url(page, product)
        if new_url:
            return new_url
        return True
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"  ❌ Save draft: {e}")
        try:
            screenshot_path = str(BASE_DIR / "save_draft_exception.png")
            await page.screenshot(path=screenshot_path, full_page=True)
            print(f"  📸 Đã chụp ảnh màn hình exception tại: {screenshot_path}")
        except Exception:
            pass
        return False, f"Exception: {str(e)[:100]}"

def _is_etsy_signin_page(url: str, content: str) -> bool:
    url_lower = (url or "").lower()
    content_lower = (content or "").lower()
    return ("signin" in url_lower or "join" in url_lower or "access is temporarily" in content_lower)


async def _navigate_to_shop_manager(page) -> None:
    """Enter Shop Manager on the exact existing Etsy page/session.

    A warmed-up authenticated Chrome/CDP session can occasionally exceed the
    default 30-second Playwright navigation timeout even though the page is
    still usable. Retry that same navigation a bounded number of times, while
    letting every non-timeout error propagate unchanged. Authentication and
    shop identity are deliberately checked by the caller after this helper.
    """
    last_timeout: PlaywrightTimeoutError | None = None
    for attempt in range(SHOP_MANAGER_NAVIGATION_ATTEMPTS):
        try:
            await page.goto(
                SHOP_MANAGER_LISTINGS_URL,
                wait_until="domcontentloaded",
                timeout=SHOP_MANAGER_NAVIGATION_TIMEOUT_MS,
            )
            return
        except PlaywrightTimeoutError as exc:
            last_timeout = exc
            if attempt + 1 >= SHOP_MANAGER_NAVIGATION_ATTEMPTS:
                break
            await page.wait_for_timeout(SHOP_MANAGER_NAVIGATION_RETRY_DELAY_MS)

    raise RuntimeError(
        "❌ Không thể vào Shop Manager sau "
        f"{SHOP_MANAGER_NAVIGATION_ATTEMPTS} lần thử (timeout "
        f"{SHOP_MANAGER_NAVIGATION_TIMEOUT_MS}ms mỗi lần)."
    ) from last_timeout


def _expected_etsy_shop_slug(shop_id: str) -> str:
    shop_config = SHOPS.get(shop_id, {})
    shop_link = str(shop_config.get("etsy_link") or "").strip()
    match = re.search(r"(?:etsy\.com)?/shop/([^/?#]+)", shop_link, flags=re.IGNORECASE)
    return (match.group(1) if match else shop_id).strip().lower()


async def _assert_shop_manager_identity(page, shop_id: str) -> None:
    expected_slug = _expected_etsy_shop_slug(shop_id)
    current_url = str(page.url or "").lower()
    if f"/shop/{expected_slug}" in current_url or f"/shops/{expected_slug}/" in current_url:
        return

    public_shop_hrefs = await page.locator('a[href*="/shop/"]').evaluate_all(
        "(links) => links.map((link) => link.href || link.getAttribute('href') || '')"
    )
    observed_slugs = set()
    for href in public_shop_hrefs or []:
        match = re.search(r"(?:etsy\.com)?/shop/([^/?#]+)", str(href), flags=re.IGNORECASE)
        if match:
            observed_slugs.add(match.group(1).strip().lower())
    if expected_slug in observed_slugs:
        return

    observed = ", ".join(sorted(observed_slugs)) or "không xác định"
    raise RuntimeError(
        f"❌ Phiên Chrome không xác minh được đúng shop '{shop_id}' "
        f"(cần {expected_slug}, thấy {observed}). Dừng để tránh đăng nhầm shop."
    )


def _exit_code_for_failed_products(failed_count: int) -> int:
    return 1 if failed_count > 0 else 0


# ── Main ───────────────────────────────────────────────────────────────────────
async def main():
    global EXCEL_FILE, SHOP_DIR
    parser = argparse.ArgumentParser(description="Etsy Auto Poster")
    parser.add_argument("--batch", type=int, default=DEFAULT_BATCH,
                        help=f"Số sản phẩm mỗi lần chạy (mặc định {DEFAULT_BATCH})")
    parser.add_argument("--skip", type=int, default=0,
                        help="Bỏ qua N sản phẩm đầu (để chạy tiếp từ giữa chừng)")
    parser.add_argument("--product", type=str, default=None,
                        help="Chạy 1 sản phẩm cụ thể theo folder name (vd: product-41)")
    parser.add_argument(
        "--products",
        action="append",
        default=None,
        help="Chạy nhiều sản phẩm theo folder name (vd: --products product-1 --products product-3 hoặc --products product-1,product-3)",
    )
    parser.add_argument(
        "--selected-product",
        action="append",
        default=None,
        help="Chạy đúng cặp row:folder đã chọn (vd: --selected-product 42:product-39)",
    )
    parser.add_argument("--shop", type=str, default="templystudios",
                        help="Mã shop cần chạy")
    parser.add_argument("--edit-url", type=str, default=None,
                        help="Edit an existing listing URL instead of creating a new one")
    args = parser.parse_args()

    SHOP_DIR = BASE_DIR / "shops" / args.shop
    EXCEL_FILE = SHOP_DIR / "Etsy_SEO_Generator.xlsx"

    try:
        requested_products = _normalize_requested_products(args.products)
        selected_products = _normalize_selected_products(args.selected_product)
    except ValueError as exc:
        raise RuntimeError(f"❌ Danh sách sản phẩm không hợp lệ: {exc}")
    if selected_products and (requested_products or args.product):
        raise RuntimeError("❌ Không dùng --selected-product cùng --product/--products")

    if args.product:
        normalized_product = args.product.strip() if args.product is not None else ""
        if not re.fullmatch(r"product-\d+", normalized_product):
            raise RuntimeError(f"❌ --product không hợp lệ: {args.product}")
        if normalized_product and normalized_product not in requested_products:
            requested_products.insert(0, normalized_product)

    products, wb, ws, total = read_products(
        batch=args.batch,
        skip=args.skip,
        product_folder=args.product,
        product_folders=requested_products,
        selected_products=selected_products,
        shop_id=args.shop,
    )

    if selected_products:
        found_pairs = {(int(p["row"]), str(p["folder"]).strip()) for p in products}
        missing_pairs = [f"{row}:{folder}" for row, folder in selected_products if (row, folder) not in found_pairs]
        if missing_pairs:
            raise RuntimeError(f"❌ Không tìm thấy đúng row:folder: {', '.join(missing_pairs)}")
    elif requested_products:
        requested_lookup = {str(p).strip() for p in requested_products if str(p).strip()}
        found_folders = {str(p["folder"]).strip() for p in products}
        missing = [folder for folder in requested_products if folder not in found_folders]
        if missing:
            raise RuntimeError(f"❌ Không tìm thấy folder: {', '.join(missing)}")

    if not products:
        print("\n⚠  Không có sản phẩm nào cần đăng.\n")
        return 0

    print(f"\n{'='*55}")
    print(f"  🛍  Etsy Auto Poster")
    print(f"  📦 {total} tổng | Batch: {args.batch} | Skip: {args.skip}")
    print(f"{'='*55}")
    for p in products:
        print(f"   • {p['folder']} | {len(p['image_paths'])} ảnh | {len(p['pdf_paths'])} PDF | {p['title'][:45]}...")

    if FILL_TRANSLATIONS:
        langs = ", ".join(n for _, n, _ in LANGUAGES)
        print(f"\n  🌐 Dịch sang: {langs}")

    BROWSER_DIR = resolve_browser_session_dir(
        args.shop,
        config=SHOPS,
        base_dir=BASE_DIR,
        home_dir=Path.home(),
    )

    print()
    BROWSER_DIR.mkdir(exist_ok=True, parents=True)

    async with async_playwright() as pw:
        ctx, page, owns_context = await _open_poster_context(
            pw, args.shop, BROWSER_DIR
        )

        # Kiểm tra đăng nhập
        await _navigate_to_shop_manager(page)
        await page.wait_for_timeout(4000)
        if _is_etsy_signin_page(page.url, await page.content()):
            raise RuntimeError(
                "❌ Cần đăng nhập Etsy trước khi post. "
                "Profile hiện tại đang ở trang đăng nhập, không thể tiếp tục đăng listing."
            )
        await _assert_shop_manager_identity(page, args.shop)

        print("  ✅ Đã vào Shop Manager!\n")

        success = failed = skipped = 0
        for i, product in enumerate(products, 1):
            print(f"\n[{i}/{len(products)}]", end="")
            try:
                ok = await fill_listing(page, product, edit_url=args.edit_url)
                if ok == "DUPLICATE":
                    skipped += 1
                    save_status(wb, ws, product["row"], "✅ Đã đăng draft")
                    print(f"  ↩️  Bỏ qua (đã có draft trên Etsy) — cập nhật Excel")
                elif ok == UNVERIFIED_DRAFT_URL_SENTINEL:
                    success += 1
                    save_status(wb, ws, product["row"], "✅ Đã đăng draft (URL chưa xác minh)")
                    mark_product_asset_operation_success(product)
                elif isinstance(ok, str) and ok.startswith("http"):
                    success += 1
                    save_status(wb, ws, product["row"], "✅ Đã đăng draft", url=ok)
                    mark_product_asset_operation_success(product)
                elif ok is True:
                    success += 1
                    save_status(wb, ws, product["row"], "✅ Đã đăng draft")
                    mark_product_asset_operation_success(product)
                elif isinstance(ok, tuple) and ok[0] is False:
                    failed += 1
                    err_reason = ok[1]
                    save_status(wb, ws, product["row"], f"❌ Lỗi: {err_reason}")
                else:
                    failed += 1
                    save_status(wb, ws, product["row"], "❌ Lỗi")
            except Exception as e:
                print(f"  ❌ {e}")
                failed += 1
                save_status(wb, ws, product["row"], f"❌ Lỗi: {str(e)[:100]}")

            # Delay giữa các sản phẩm (tránh bị Etsy block)
            if i < len(products):
                wait_sec = 8
                print(f"\n  ⏳ Nghỉ {wait_sec}s trước sản phẩm tiếp theo...")
                await asyncio.sleep(wait_sec)

        print(f"\n{'='*55}")
        print(f"  ✅ Thành công : {success}/{len(products)}")
        print(f"  ↩️  Bỏ qua (trùng): {skipped}/{len(products)}")
        print(f"  ❌ Thất bại  : {failed}/{len(products)}")
        remaining = total - args.skip - len(products)
        if remaining > 0:
            next_skip = args.skip + len(products)
            print(f"\n  📌 Còn {remaining} sản phẩm. Lần sau chạy:")
            print(f"     python3 etsy_auto_post.py --batch {args.batch} --skip {next_skip}")
        print(f"{'='*55}\n")
        if owns_context:
            await ctx.close()
        else:
            await page.close()
        return _exit_code_for_failed_products(failed)

if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
