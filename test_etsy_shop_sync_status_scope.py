"""Regression coverage for all-status snapshots and syncable-only workbook updates."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

import etsy_shop_sync


class TestEtsyShopSyncStatusScope(unittest.TestCase):
    def _make_workbook(self, path: Path, *, listing_url: str = "") -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Listings"
        worksheet.cell(row=4, column=2, value="product-01")
        worksheet.cell(row=4, column=8, value="Only inactive listing")
        worksheet.cell(row=4, column=14, value="⏸ Inactive trên Etsy")
        worksheet.cell(row=4, column=16, value=listing_url or None)
        workbook.save(path)

    def test_inactive_and_expired_are_retained_in_snapshot_but_cannot_remap_workbook(self) -> None:
        crawled = {
            "active": [],
            "draft": [],
            "inactive": [{"id": "901", "title": "Only inactive listing"}],
            "expired": [{"id": "902", "title": "Only expired listing"}],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            shop_dir = root / "shops" / "shop"
            shop_dir.mkdir(parents=True)
            workbook_path = shop_dir / "Etsy_SEO_Generator.xlsx"
            # An existing inactive URL must be preserved, not cleared or
            # replaced through the active/draft reconciliation path.
            self._make_workbook(workbook_path, listing_url="https://www.etsy.com/listing/901")
            (root / "scratch").mkdir()

            with patch.object(etsy_shop_sync, "BASE_DIR", root):
                report = etsy_shop_sync.sync_excel("shop", crawled)

            saved = openpyxl.load_workbook(workbook_path)["Listings"]
            self.assertEqual("⏸ Inactive trên Etsy", saved.cell(row=4, column=14).value)
            self.assertEqual("https://www.etsy.com/listing/901", saved.cell(row=4, column=16).value)
            self.assertEqual(2, report["etsy_total"])
            self.assertEqual(0, report["syncable_etsy_total"])
            self.assertEqual(0, report["matched_total"])
            self.assertEqual(0, report["etsy_unmapped_total"])
            self.assertEqual(
                {"active": 0, "draft": 0, "inactive": 1, "expired": 1},
                report["etsy_counts"],
            )
            snapshot = json.loads(Path(report["crawl"]).read_text(encoding="utf-8"))
            self.assertEqual(crawled, snapshot)


if __name__ == "__main__":
    unittest.main()
