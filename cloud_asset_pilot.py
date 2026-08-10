#!/usr/bin/env python3
"""Fail-closed pilot eligibility planner for the cloud asset rollout.

The default ``plan`` command is read-only.  It never contacts Etsy, Drive, or
rclone and never writes a workbook, state marker, or local asset.  An explicit
``--execute --confirm-cloud-write`` path is provided for a later operator run;
the current implementation still reports cloud-pipeline verification only and
never deletes local assets or claims live Etsy verification.

Two independent pilot modes are supported:

* ``shop-only`` validates and uploads only ``shops/<shop>/<product>``.
* ``shop-and-master`` additionally requires an explicit master mapping and
  validates/uploads ``master_products/<product>``.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from pathlib import Path
from typing import Any, Mapping, Optional, Sequence

import openpyxl

from cloud_asset_store import (
    AssetValidationError,
    CloudAssetError,
    CloudAssetStore,
    build_manifest,
    resolve_product,
)
from cloud_asset_store_config import load_config


UTC = dt.timezone.utc
FOLDER_RE = re.compile(r"^product-(\d+)$")
LISTING_ID_RE = re.compile(r"(?:/listing/|listing-editor/edit/)(\d+)", re.IGNORECASE)
SNAPSHOT_NAME_RE = re.compile(r"etsy_manager_current_[^_]+_(\d{8}_\d{6})\.json$")
MASTER_FIELDS = ("master_product", "master_products", "master")
PILOT_MODES = ("shop-only", "shop-and-master")
DEFAULT_PILOT_MODE = "shop-and-master"


def _utc_now(value: Optional[dt.datetime] = None) -> dt.datetime:
    timestamp = value or dt.datetime.now(UTC)
    if timestamp.tzinfo is None:
        timestamp = timestamp.replace(tzinfo=UTC)
    return timestamp.astimezone(UTC)


def _utc_text(value: dt.datetime) -> str:
    return _utc_now(value).isoformat().replace("+00:00", "Z")


def _parse_time(value: Any) -> Optional[dt.datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = dt.datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    return _utc_now(parsed)


def _snapshot_time(path: Path, payload: Mapping[str, Any]) -> Optional[dt.datetime]:
    for field in ("crawledAt", "snapshotAt", "created_at"):
        parsed = _parse_time(payload.get(field))
        if parsed:
            return parsed
    match = SNAPSHOT_NAME_RE.search(path.name)
    if match:
        try:
            return dt.datetime.strptime(match.group(1), "%Y%m%d_%H%M%S").replace(tzinfo=UTC)
        except ValueError:
            pass
    try:
        return dt.datetime.fromtimestamp(path.stat().st_mtime, tz=UTC)
    except OSError:
        return None


def _latest_snapshot(repo_root: Path, shop_id: str, requested: Optional[Path]) -> Optional[Path]:
    if requested is not None:
        return requested.expanduser().absolute()
    scratch = repo_root / "scratch"
    candidates = sorted(scratch.glob(f"etsy_manager_current_{shop_id}_*.json"))
    return candidates[-1] if candidates else None


def _listing_id(value: Any) -> str:
    text = str(value or "").strip()
    if text.isdigit():
        return text
    match = LISTING_ID_RE.search(text)
    return match.group(1) if match else ""


def _active_snapshot_listings(payload: Mapping[str, Any]) -> dict[str, dict[str, Any]]:
    listings: dict[str, dict[str, Any]] = {}
    raw_active = payload.get("active", [])
    if isinstance(raw_active, list):
        for item in raw_active:
            if not isinstance(item, dict):
                continue
            listing_id = _listing_id(item.get("id") or item.get("listing_id"))
            if listing_id:
                listings[listing_id] = {**item, "managerStatus": "active", "listing_id": listing_id}
    raw_listings = payload.get("listings", [])
    if isinstance(raw_listings, list):
        for item in raw_listings:
            if not isinstance(item, dict):
                continue
            status = str(item.get("managerStatus") or item.get("status") or "").strip().lower()
            listing_id = _listing_id(item.get("id") or item.get("listing_id"))
            if status == "active" and listing_id:
                listings[listing_id] = {**item, "managerStatus": "active", "listing_id": listing_id}
    return listings


def _explicit_master_folder(entry: Any) -> str:
    if not isinstance(entry, dict):
        return ""
    values = []
    for field in MASTER_FIELDS:
        if field in entry:
            raw = entry[field]
            if isinstance(raw, str):
                values.append(raw.strip())
            elif isinstance(raw, list) and len(raw) == 1 and isinstance(raw[0], str):
                values.append(raw[0].strip())
    values = [value for value in values if value]
    if len(set(values)) != 1:
        return ""
    return values[0] if FOLDER_RE.fullmatch(values[0]) else ""


def _source_map_entry(mapping: Mapping[str, Any], shop_id: str, folder: str) -> tuple[str, Any] | None:
    matches = []
    for source_key, raw_entry in mapping.items():
        if not isinstance(raw_entry, dict):
            continue
        if str(raw_entry.get(shop_id) or "").strip() == folder:
            matches.append((str(source_key), raw_entry))
    return matches[0] if len(matches) == 1 else None


def _normalize_mode(value: Any) -> str:
    mode = str(value or DEFAULT_PILOT_MODE).strip().lower()
    if mode not in PILOT_MODES:
        raise ValueError(f"unsupported pilot mode: {mode!r}; expected one of {PILOT_MODES}")
    return mode


def _manifest_for(repo_root: Path, relative_path: str) -> dict[str, Any]:
    product_path, identity = resolve_product(repo_root, relative_path)
    manifest, _data, digest = build_manifest(product_path, identity, "pilot-preview")
    return {
        "path": str(product_path),
        "identity": identity.key,
        "manifest_sha256": digest,
        "counts": dict(manifest.get("counts") or {}),
        "manifest": manifest,
    }


def _folder_number(folder: str) -> int:
    match = FOLDER_RE.fullmatch(folder)
    return int(match.group(1)) if match else 10**12


def _load_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"JSON phải là object: {path}")
    return value


def plan_pilot(
    repo_root: Path,
    shop_id: str = "templystudios",
    snapshot_path: Optional[Path] = None,
    map_path: Optional[Path] = None,
    required: int = 5,
    now: Optional[dt.datetime] = None,
    mode: str = DEFAULT_PILOT_MODE,
) -> dict[str, Any]:
    root = Path(repo_root).expanduser().absolute()
    timestamp = _utc_now(now)
    pilot_mode = _normalize_mode(mode)
    result: dict[str, Any] = {
        "ok": False,
        "state": "BLOCKED_CATALOG_MAPPING",
        "shop": shop_id,
        "mode": pilot_mode,
        "required": int(required),
        "eligible": [],
        "selected": [],
        "candidates": [],
        "snapshot": {},
        "write_performed": False,
        "delete_performed": False,
        "live_etsy_verified": False,
    }
    blockers: list[str] = []

    active_shop_path = root / "active_shop.txt"
    try:
        active_shop = active_shop_path.read_text(encoding="utf-8").strip()
    except OSError:
        active_shop = ""
    if active_shop != shop_id:
        blockers.append(f"active_shop_mismatch:{active_shop or '(missing)'}")

    snapshot_file = _latest_snapshot(root, shop_id, snapshot_path)
    payload: dict[str, Any] = {}
    snapshot_at: Optional[dt.datetime] = None
    if snapshot_file is None or not snapshot_file.is_file():
        blockers.append("snapshot_missing")
    else:
        try:
            payload = _load_json(snapshot_file)
            snapshot_at = _snapshot_time(snapshot_file, payload)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            blockers.append(f"snapshot_invalid:{type(exc).__name__}")
        if str(payload.get("shopId") or payload.get("shop_id") or "").strip() != shop_id:
            blockers.append("snapshot_shop_mismatch")
        age_hours = None if snapshot_at is None else round((timestamp - snapshot_at).total_seconds() / 3600, 2)
        if snapshot_at is None or age_hours is None or age_hours > 24:
            blockers.append("snapshot_stale")
        result["snapshot"] = {
            "path": str(snapshot_file),
            "shop_id": payload.get("shopId") or payload.get("shop_id"),
            "captured_at": _utc_text(snapshot_at) if snapshot_at else None,
            "age_hours": age_hours,
            "max_age_hours": 24,
            "fresh": not any(reason == "snapshot_stale" for reason in blockers),
        }

    snapshot_listings = _active_snapshot_listings(payload)
    workbook_path = root / "shops" / shop_id / "Etsy_SEO_Generator.xlsx"
    source_map: dict[str, Any] = {}
    if pilot_mode == "shop-and-master":
        map_file = map_path.expanduser().absolute() if map_path else root / "product_source_map.json"
        try:
            source_map = _load_json(map_file)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            blockers.append(f"source_map_invalid:{type(exc).__name__}")

    if not workbook_path.is_file():
        blockers.append("workbook_missing")
        rows: list[dict[str, Any]] = []
    else:
        try:
            workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
            sheet = workbook["Listings"]
            rows = []
            for row_number in range(4, sheet.max_row + 1):
                folder = str(sheet.cell(row=row_number, column=2).value or "").strip()
                if folder:
                    rows.append({
                        "row": row_number,
                        "folder": folder,
                        "title": str(sheet.cell(row=row_number, column=8).value or "").strip(),
                        "listing_id": _listing_id(sheet.cell(row=row_number, column=16).value),
                    })
            workbook.close()
        except (OSError, KeyError, ValueError) as exc:
            rows = []
            blockers.append(f"workbook_invalid:{type(exc).__name__}")

    folder_counts: dict[str, int] = {}
    for row in rows:
        folder_counts[row["folder"]] = folder_counts.get(row["folder"], 0) + 1

    for row in rows:
        reasons: list[str] = []
        folder = row["folder"]
        listing_id = row["listing_id"]
        candidate: dict[str, Any] = {
            "row": row["row"],
            "folder": folder,
            "listing_id": listing_id or None,
            "title": row["title"],
            "eligible": False,
            "reasons": reasons,
        }
        if not FOLDER_RE.fullmatch(folder):
            reasons.append("invalid_product_folder")
        if folder_counts.get(folder) != 1:
            reasons.append("local_folder_not_unique")
        if not listing_id or not listing_id.isdigit():
            reasons.append("missing_numeric_workbook_listing_id")
        listing = snapshot_listings.get(listing_id) if listing_id else None
        if listing is None:
            reasons.append("listing_not_active_in_fresh_snapshot")

        shop_manifest = None
        master_manifest = None
        if pilot_mode == "shop-and-master":
            map_entry = _source_map_entry(source_map, shop_id, folder)
            if map_entry is None:
                reasons.append("missing_unambiguous_source_mapping")
            else:
                source_key, entry = map_entry
                candidate["source_key"] = source_key
                master_folder = _explicit_master_folder(entry)
                candidate["master_folder"] = master_folder or None
                if not master_folder:
                    reasons.append("missing_explicit_master_mapping")

        if not reasons:
            try:
                shop_manifest = _manifest_for(root, f"shops/{shop_id}/{folder}")
                if pilot_mode == "shop-and-master":
                    master_manifest = _manifest_for(root, f"master_products/{candidate['master_folder']}")
            except (AssetValidationError, CloudAssetError, OSError, ValueError, TypeError, KeyError) as exc:
                reasons.append(f"asset_validation_failed:{str(exc)[:180]}")
        if shop_manifest:
            shop_bytes = int(shop_manifest["counts"].get("total_bytes", 0))
            candidate.update({
                "shop_manifest_sha256": shop_manifest["manifest_sha256"],
                "shop_bytes": shop_bytes,
            })
        if master_manifest:
            master_bytes = int(master_manifest["counts"].get("total_bytes", 0))
            candidate.update({
                "master_manifest_sha256": master_manifest["manifest_sha256"],
                "master_bytes": master_bytes,
            })
        if shop_manifest and (pilot_mode == "shop-only" or master_manifest):
            candidate["combined_bytes"] = int(candidate.get("shop_bytes", 0)) + int(
                candidate.get("master_bytes", 0)
            )
        if not reasons:
            candidate["eligible"] = True
            result["eligible"].append(candidate)
        result["candidates"].append(candidate)

    result["eligible"] = sorted(
        result["eligible"],
        key=lambda item: (int(item.get("combined_bytes", 0)), _folder_number(str(item.get("folder") or ""))),
    )
    if blockers:
        result["reason"] = ";".join(dict.fromkeys(blockers))
    elif len(result["eligible"]) < required:
        result["reason"] = f"only {len(result['eligible'])}/{required} candidates passed all catalog gates"
    else:
        result["state"] = "PILOT_ELIGIBLE"
        result["ok"] = True
        result["selected"] = result["eligible"][:required]
        basis = "shop bytes" if pilot_mode == "shop-only" else "combined bytes"
        result["reason"] = f"selected {required} candidates by {basis} then numeric folder"
    return result


def execute_pilot(
    plan: Mapping[str, Any], store: CloudAssetStore, mode: Optional[str] = None
) -> dict[str, Any]:
    """Upload and verify an already-approved plan; never deletes local files."""

    plan_mode = _normalize_mode(plan.get("mode", DEFAULT_PILOT_MODE))
    if mode is not None and _normalize_mode(mode) != plan_mode:
        raise CloudAssetError("execution mode does not match the approved pilot plan")
    selected = list(plan.get("selected") or [])
    required = int(plan.get("required") or 5)
    if plan.get("state") != "PILOT_ELIGIBLE" or len(selected) != required:
        raise CloudAssetError("pilot execution requires a fresh fully eligible plan")
    results = []
    for candidate in selected:
        item = {"folder": candidate.get("folder"), "operations": [], "ok": False}
        try:
            operations = [("shop", f"shops/{plan['shop']}/{candidate['folder']}")]
            if plan_mode == "shop-and-master":
                operations.append(("master", f"master_products/{candidate['master_folder']}"))
            for scope, relative in operations:
                uploaded = store.upload(relative)
                if not isinstance(uploaded, dict) or not uploaded.get("ok", False):
                    raise CloudAssetError(f"{scope} upload không đạt cho {candidate['folder']}")
                verified = store.verify(relative)
                if not isinstance(verified, dict) or not verified.get("ok", False):
                    raise CloudAssetError(f"{scope} verify không đạt cho {candidate['folder']}")
                item["operations"].append({"scope": scope, "upload": uploaded, "verify": verified})
            item["ok"] = True
            item["state"] = "CLOUD_PIPELINE_VERIFIED"
        except (CloudAssetError, OSError, ValueError, TypeError, KeyError) as exc:
            item["state"] = "ERROR"
            item["error"] = str(exc)[:300]
        results.append(item)
    return {
        "ok": all(item["ok"] for item in results),
        "state": "CLOUD_PIPELINE_VERIFIED" if all(item["ok"] for item in results) else "ERROR",
        "mode": plan_mode,
        "results": results,
        "write_performed": True,
        "delete_performed": False,
        "live_etsy_verified": False,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("command", choices=("plan",), nargs="?", default="plan")
    parser.add_argument("--repo-root", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--shop", default="templystudios")
    parser.add_argument("--snapshot", type=Path)
    parser.add_argument("--map-file", type=Path)
    parser.add_argument(
        "--mode",
        choices=PILOT_MODES,
        default=DEFAULT_PILOT_MODE,
        help="shop-only uploads only the selected shop scope; shop-and-master also requires master mappings",
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        help="run the already-eligible pilot; still never deletes local assets",
    )
    parser.add_argument(
        "--confirm-cloud-write",
        action="store_true",
        help="explicitly allow the pilot execution path to write Drive revisions",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    result = plan_pilot(args.repo_root, args.shop, args.snapshot, args.map_file, mode=args.mode)
    if args.execute:
        if not args.confirm_cloud_write:
            result["execution"] = {
                "ok": False,
                "state": "BLOCKED_OPERATOR_CONFIRMATION",
                "mode": args.mode,
                "write_performed": False,
                "delete_performed": False,
                "reason": "--execute also requires --confirm-cloud-write",
            }
        elif result.get("ok"):
            config = load_config(args.repo_root)
            store = CloudAssetStore(
                repo_root=config.repo_root,
                remote=config.remote,
                parent_id=config.parent_id,
                rclone_bin=config.rclone_bin,
                cache_root=config.cache_root,
                lock_timeout_seconds=config.lock_timeout_seconds,
                success_ttl_seconds=config.success_ttl_seconds,
                failure_ttl_seconds=config.failure_ttl_seconds,
                offload_age_days=config.offload_age_days,
            )
            result["execution"] = execute_pilot(result, store)
        else:
            result["execution"] = {
                "ok": False,
                "state": "BLOCKED_CATALOG_MAPPING",
                "mode": args.mode,
                "write_performed": False,
                "delete_performed": False,
                "reason": "pilot plan is not fully eligible",
            }
    print(json.dumps(result, ensure_ascii=False, sort_keys=True, indent=2, default=str))
    execution = result.get("execution")
    return 0 if result.get("ok") and (not args.execute or execution and execution.get("ok")) else 2


if __name__ == "__main__":
    raise SystemExit(main())
