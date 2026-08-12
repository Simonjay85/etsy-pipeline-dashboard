#!/usr/bin/env python3
from __future__ import annotations

import tempfile
import unittest
from contextlib import ExitStack
from pathlib import Path
from unittest.mock import patch

import openpyxl
from fastapi.testclient import TestClient

import dashboard_app


class TestRegisterLocalProducts(unittest.TestCase):
    def _make_workbook(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listings"
        wb.save(path)

    def _create_folder(self, path: Path, keep_file: bool = False) -> Path:
        folder = path
        folder.mkdir(parents=True, exist_ok=True)
        (folder / "images").mkdir(exist_ok=True)
        (folder / "files").mkdir(exist_ok=True)
        if keep_file:
            (folder / "images" / "keep.txt").write_text("keep", encoding="utf-8")
        return folder

    def _run_request(self, shop: str, shop_dir: Path, excel_path: Path, payload: dict) -> tuple[int, dict, TestClient]:
        with ExitStack() as stack:
            stack.enter_context(patch.object(dashboard_app, "_active_shop_id", shop))
            stack.enter_context(patch.object(dashboard_app, "SHOP_DIR", lambda: shop_dir))
            stack.enter_context(patch.object(dashboard_app, "EXCEL_FILE", lambda: excel_path))
            client = TestClient(dashboard_app.app)
            response = client.post("/api/local-products/register", json=payload)
            return response.status_code, response.json(), client

    def test_register_allows_bulk_local_folders_and_writes_clean_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            shop = "shop-a"
            shop_dir = base / "shops" / shop
            excel_path = shop_dir / "Etsy_SEO_Generator.xlsx"
            self._make_workbook(excel_path)

            # Existing mapped row should be skipped by allocator.
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Listings"]
            ws.cell(row=4, column=2, value="product-99")
            ws.cell(row=4, column=5, value=9.99)
            wb.save(excel_path)

            self._create_folder(shop_dir / "product-02", keep_file=True)
            self._create_folder(shop_dir / "product-03", keep_file=True)

            status, data, _ = self._run_request(
                shop,
                shop_dir,
                excel_path,
                {"shop": shop, "folders": ["product-02", "product-03"]},
            )

            self.assertEqual(status, 200)
            self.assertTrue(data.get("ok"))
            rows = {item["folder"]: item["row"] for item in data.get("rows", [])}
            self.assertEqual(rows.get("product-02"), 5)
            self.assertEqual(rows.get("product-03"), 6)
            self.assertTrue(any("backup_local_register_" in p.name for p in excel_path.parent.iterdir()))

            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Listings"]

            for folder in ("product-02", "product-03"):
                row = rows[folder]
                self.assertEqual(str(ws.cell(row=row, column=2).value), folder)
                self.assertEqual(float(ws.cell(row=row, column=5).value), 4.99)
                self.assertEqual(int(ws.cell(row=row, column=11).value), 999)
                self.assertEqual(str(ws.cell(row=row, column=13).value), "2020_2026")
                self.assertEqual(str(ws.cell(row=row, column=14).value), "⏳ Chờ đăng")
                self.assertEqual(str(ws.cell(row=row, column=15).value), "Digital Planner")
                self.assertEqual(str(ws.cell(row=row, column=18).value), dashboard_app.generate_sku(shop, folder))
                self.assertEqual(str(ws.cell(row=row, column=16).value or ""), "")

            for folder in ("product-02", "product-03"):
                keep_file = shop_dir / folder / "images" / "keep.txt"
                self.assertTrue(keep_file.exists())

    def test_rejects_invalid_or_nonexistent_or_already_mapped_folders(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            shop = "shop-b"
            shop_dir = base / "shops" / shop
            excel_path = shop_dir / "Etsy_SEO_Generator.xlsx"
            self._make_workbook(excel_path)

            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Listings"]
            ws.cell(row=4, column=2, value="product-01")
            ws.cell(row=4, column=5, value=4.99)
            wb.save(excel_path)

            self._create_folder(shop_dir / "product-01")
            self._create_folder(shop_dir / "product-04")

            status, data, _ = self._run_request(shop, shop_dir, excel_path, {"shop": shop, "folders": ["bad-folder"]})
            self.assertEqual(status, 400)
            self.assertIn("không hợp lệ", str(data.get("detail", "")))

            status, data, _ = self._run_request(shop, shop_dir, excel_path, {"shop": shop, "folders": ["product-04", "product-04"]})
            self.assertEqual(status, 400)
            self.assertIn("lặp", str(data.get("detail", "")).lower())

            status, data, _ = self._run_request(shop, shop_dir, excel_path, {"shop": shop, "folders": ["product-05"]})
            self.assertEqual(status, 404)
            self.assertIn("không tìm thấy", str(data.get("detail", "")).lower())

            status, data, _ = self._run_request(shop, shop_dir, excel_path, {"shop": shop, "folders": ["product-01"]})
            self.assertEqual(status, 409)
            self.assertIn("đã được map", str(data.get("detail", "")).lower())

    def test_rolls_back_and_keeps_assets_if_catalog_write_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            shop = "shop-c"
            shop_dir = base / "shops" / shop
            excel_path = shop_dir / "Etsy_SEO_Generator.xlsx"
            self._make_workbook(excel_path)

            target_folder = self._create_folder(shop_dir / "product-11", keep_file=True)
            keep_file = target_folder / "images" / "keep.txt"

            original_set_cell_value = dashboard_app.set_cell_value

            def fail_on_write(*args, **kwargs):
                raise RuntimeError("forced test failure")

            with ExitStack() as stack:
                stack.enter_context(patch.object(dashboard_app, "_active_shop_id", shop))
                stack.enter_context(patch.object(dashboard_app, "SHOP_DIR", lambda: shop_dir))
                stack.enter_context(patch.object(dashboard_app, "EXCEL_FILE", lambda: excel_path))
                stack.enter_context(patch.object(dashboard_app, "set_cell_value", side_effect=fail_on_write))
                client = TestClient(dashboard_app.app)
                response = client.post("/api/local-products/register", json={"shop": shop, "folders": ["product-11"]})

            self.assertEqual(response.status_code, 500)
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Listings"]
            self.assertIsNone(ws.cell(row=4, column=2).value)
            # Asset files should remain untouched.
            self.assertTrue(keep_file.exists())

            # Restore so later assertions in process (if any) won't be impacted.
            dashboard_app.set_cell_value = original_set_cell_value


if __name__ == "__main__":
    unittest.main()
