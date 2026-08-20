import asyncio
import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from PIL import Image
from starlette.requests import Request
import sync_factory_to_shop as sync_factory

import dashboard_app


FACTORY_ROOT = "shops/templystudios"


def _write_image(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    Image.new("RGB", (8, 8), (32, 96, 128)).save(path, format="PNG")


def _make_factory_source(root: Path, name: str, keyword: str = "Product") -> Path:
    source = root / name
    image_dir = source / "etsy-images-chatgpt-final"
    zip_dir = source / "source-zip"
    image_dir.mkdir(parents=True, exist_ok=True)
    zip_dir.mkdir(parents=True, exist_ok=True)
    _write_image(image_dir / "01_hero_image.png")
    (zip_dir / f"{keyword}.zip").write_bytes(b"source")
    return source


def _make_workbook(path: Path, rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings"
    for row_num, folder, keyword in rows:
        ws.cell(row=row_num, column=2, value=folder)
        ws.cell(row=row_num, column=3, value=keyword)
    wb.save(path)


def _request(payload: dict, shop_id: str = "templystudios") -> Request:
    payload = {"shop_id": shop_id, **payload}
    body = json.dumps(payload).encode("utf-8")
    sent = False

    async def receive():
        nonlocal sent
        if sent:
            return {"type": "http.disconnect"}
        sent = True
        return {"type": "http.request", "body": body, "more_body": False}

    return Request({"type": "http", "method": "POST", "path": "/api/image-factory/import", "headers": []}, receive)


class ImageFactoryImportTests(unittest.TestCase):
    def _source_root(self, temp: str, shop_id: str = "templystudios") -> tuple[Path, Path, Path]:
        root = Path(temp)
        source = root / "shops" / shop_id
        shop = source
        shop.mkdir(parents=True, exist_ok=True)
        excel = shop / "Etsy_SEO_Generator.xlsx"
        return root, source, excel

    def test_scan_uses_temply_layout_sources_and_skips_excluded_folders(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            _make_workbook(excel, [(4, "product-40", "Different keyword")])

            supported = source / "Dinosaur-Bundle-Silhouette-47733169"
            supported.mkdir()
            (supported / "etsy-images-chatgpt-final").mkdir()
            (supported / "source-zip").mkdir()
            _write_image(supported / "etsy-images-chatgpt-final" / "01-hero_image.png")
            (supported / "source-zip" / "Exact Product.zip").write_bytes(b"source")

            legacy = source / "Legacy-Sticker-Factory"
            legacy.mkdir()
            (legacy / "images").mkdir()
            (legacy / "files").mkdir()
            (legacy / "images" / "hero.jpg").write_bytes(b"img")
            (legacy / "files" / "Kawaii Planner.pdf").write_bytes(b"legacy")

            imported_files = source / "product-40" / "files"
            imported_files.mkdir(parents=True)
            (imported_files / "Exact Product.zip").write_bytes(b"existing")

            (source / "product-01").mkdir()
            (source / "_deleted_products").mkdir()
            (source / "_asset_quarantine").mkdir()
            (source / "_failed_local_imports").mkdir()
            (source / "master_products").mkdir()

            folders = dashboard_app._scan_factory_folders(source, source, excel)

            self.assertEqual([item["name"] for item in folders], ["Dinosaur-Bundle-Silhouette-47733169"])
            self.assertTrue(folders[0]["already_imported"])
            self.assertEqual(folders[0]["imported_folder"], "product-40")
            self.assertEqual(folders[0]["matched_by"], "download_file_set")
            self.assertTrue(folders[0]["thumb"].endswith("/01-hero_image.png"))
            self.assertNotIn("etsy-images-chatgpt-final/", folders[0]["thumb"])
            self.assertNotIn("Legacy-Sticker-Factory", [f["name"] for f in folders])
            self.assertNotIn("product-01", [f["name"] for f in folders])
            self.assertNotIn("_deleted_products", [f["name"] for f in folders])
            self.assertNotIn("_asset_quarantine", [f["name"] for f in folders])
            self.assertNotIn("_failed_local_imports", [f["name"] for f in folders])

    def test_scan_treats_polluted_target_file_set_as_pending(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            source_folder = source / "Kawaii-Planner"
            source_folder.mkdir()
            _write_image(source_folder / "etsy-images-chatgpt-final" / "01_hero_image.png")
            (source_folder / "source-zip").mkdir()
            (source_folder / "source-zip" / "Kawaii Planner.zip").write_bytes(b"source")

            target_folder = source / "product-02" / "files"
            target_folder.mkdir(parents=True)
            (target_folder / "Kawaii Planner.zip").write_bytes(b"factory-copy")
            (target_folder / "Old Legitimate Product.zip").write_bytes(b"old-product")

            _make_workbook(excel, [(4, "product-02", "Unrelated old keyword")])

            folders = dashboard_app._scan_factory_folders(source, source, excel)

            self.assertEqual(1, len(folders))
            self.assertFalse(folders[0]["already_imported"])
            self.assertIsNone(folders[0]["imported_folder"])

    def test_daisy_scan_uses_only_daisy_shop_root(self):
        with tempfile.TemporaryDirectory() as temp:
            root, source, excel = self._source_root(temp, "daisyflowdigital")
            _make_workbook(excel)
            temply_source = root / FACTORY_ROOT
            _make_factory_source(temply_source, "Temply-Only", "Temply Product")
            with (
                patch.object(dashboard_app, "_factory_shop_dir", return_value=source),
                patch.object(dashboard_app, "_factory_excel_file", return_value=excel),
                patch.object(dashboard_app, "_active_shop_id", "daisyflowdigital"),
            ):
                result = asyncio.run(dashboard_app.scan_image_factory())

            self.assertTrue(result["ok"])
            self.assertEqual([], result["folders"])
            self.assertEqual("daisyflowdigital", result["shop_id"])
            self.assertIn("factory_path", result)
            self.assertIn("daisyflowdigital", result["factory_path"])
            self.assertNotIn("Temply-Only", json.dumps(result))
            self.assertNotIn("templystudios", result["factory_path"])
            self.assertNotIn("master_products", str(result["factory_path"]))

    def test_scan_reports_aggregate_only_catalog_exclusion_counts(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            _make_workbook(excel)
            _make_factory_source(source, "Ready-For-Import", "Ready Product")
            (source / "product-01").mkdir()
            (source / "product-02").mkdir()
            # A symlink is neither an eligible intake source nor a catalog
            # directory that should be counted in scan context.
            os.symlink(source / "product-01", source / "product-03")

            with (
                patch.object(dashboard_app, "_factory_shop_dir", return_value=source),
                patch.object(dashboard_app, "_factory_excel_file", return_value=excel),
                patch.object(dashboard_app, "_active_shop_id", "templystudios"),
            ):
                result = asyncio.run(dashboard_app.scan_image_factory())

            self.assertTrue(result["ok"])
            self.assertEqual(
                {"eligible_intake_folders": 1, "excluded_catalog_product_folders": 2},
                result["scan_summary"],
            )
            self.assertEqual(["Ready-For-Import"], [item["name"] for item in result["folders"]])
            self.assertNotIn("product-01", json.dumps(result))
            self.assertNotIn("product-02", json.dumps(result))

    def test_allocator_fills_lowest_gap_and_honors_batch_reservations(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            _make_workbook(excel, [
                (4, "product-01", "First"),
                (5, "product-02", "Second"),
            ])
            (source / "product-01").mkdir()
            (source / "product-02").mkdir()
            (source / "product-04").mkdir()
            ws = openpyxl.load_workbook(excel)["Listings"]

            self.assertEqual(3, dashboard_app._next_product_number(ws, source))
            first_slot = dashboard_app._allocate_product_slot(
                ws, source, reusable_slots=[], used_folders=set()
            )
            self.assertEqual("product-03", first_slot["folder"])

            reserved_folders = {"product-03"}
            self.assertEqual(5, dashboard_app._next_product_number(ws, source, reserved_folders))
            reserved_slot = dashboard_app._allocate_product_slot(
                ws, source, reusable_slots=[], used_folders=reserved_folders
            )
            self.assertEqual("product-05", reserved_slot["folder"])

    def test_reusable_slots_exclude_empty_folders_with_etsy_listing_mappings(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            _make_workbook(excel, [
                (4, "product-01", "Mapped by public URL"),
                (5, "product-02", "Mapped by listing ID"),
                (6, "product-03", "Invalid mapping"),
                (7, "product-04", "Unmapped"),
                (8, "product-05", "Mapped by manager URL"),
            ])
            for folder_name in (
                "product-01", "product-02", "product-03", "product-04", "product-05"
            ):
                (source / folder_name / "images").mkdir(parents=True)
                (source / folder_name / "files").mkdir()

            wb = openpyxl.load_workbook(excel)
            ws = wb["Listings"]
            ws.cell(row=4, column=16, value="https://www.etsy.com/listing/4545350918/example")
            ws.cell(row=5, column=16, value="4545350919")
            ws.cell(row=6, column=16, value="https://example.com/listing/4545350920")
            ws.cell(
                row=8,
                column=16,
                value="https://www.etsy.com/your/shops/me/listing-editor/edit/4545350921",
            )

            slots = dashboard_app._find_reusable_empty_product_slots(ws, source)

            self.assertEqual(["product-03", "product-04"], [slot["folder"] for slot in slots])
            wb.close()

    def test_reusable_slots_exclude_cloud_empty_product_slot_mapped_by_locale_public_url(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            _make_workbook(excel, [
                (4, "product-01", "Mapped by locale public URL"),
                (5, "product-02", "Reusable"),
                (6, "product-03", "Not reusable"),
            ])
            for folder_name in ("product-01", "product-02", "product-03"):
                (source / folder_name / "images").mkdir(parents=True)
                (source / folder_name / "files").mkdir()

            wb = openpyxl.load_workbook(excel)
            ws = wb["Listings"]
            ws.cell(
                row=4,
                column=16,
                value="https://www.etsy.com/ca/listing/4555695025/800-ai-commands-for-etsy-sellers-a?ref=listings_manager_grid",
            )

            # product-03 has actual usable assets, so it should not be reusable.
            _write_image(source / "product-03" / "images" / "01_hero_image.png")

            slots = dashboard_app._find_reusable_empty_product_slots(ws, source)

            self.assertEqual(["product-02"], [slot["folder"] for slot in slots])
            wb.close()

    def test_allocator_rechecks_listing_mapping_before_reusing_stale_slot(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            _make_workbook(excel, [
                (4, "product-01", "Mapped after discovery"),
                (5, "product-02", "Still reusable"),
            ])
            wb = openpyxl.load_workbook(excel)
            ws = wb["Listings"]
            slots = []
            for row_num, folder_name in ((4, "product-01"), (5, "product-02")):
                folder_path = source / folder_name
                (folder_path / "images").mkdir(parents=True)
                (folder_path / "files").mkdir()
                slots.append({
                    "folder": folder_name,
                    "number": row_num - 3,
                    "path": folder_path,
                    "row": row_num,
                })

            ws.cell(row=4, column=16, value="https://www.etsy.com/listing/4545350918")
            allocated = dashboard_app._allocate_product_slot(
                ws, source, reusable_slots=slots, used_folders=set()
            )

            self.assertEqual("product-02", allocated["folder"])
            self.assertTrue(allocated["reused"])
            self.assertTrue((source / "product-01" / "images").is_dir())
            self.assertTrue((source / "product-01" / "files").is_dir())
            wb.close()

    def test_import_rechecks_duplicates_without_mutating_shop(self):
        with tempfile.TemporaryDirectory() as temp:
            root, source, excel = self._source_root(temp)
            source.mkdir(parents=True, exist_ok=True)
            source_files = source / "Kawaii-Planner" / "source-zip"
            source_files.mkdir(parents=True)
            _write_image(source / "Kawaii-Planner" / "etsy-images-chatgpt-final" / "01_hero_image.png")
            (source_files / "Exact Product.zip").write_bytes(b"source")

            existing_files = source / "product-40" / "files"
            existing_files.mkdir(parents=True)
            (existing_files / "Exact Product.zip").write_bytes(b"existing")
            _make_workbook(excel, [(4, "product-40", "Different keyword")])

            with (
                patch.object(dashboard_app, "_factory_shop_dir", return_value=source),
                patch.object(dashboard_app, "_factory_excel_file", return_value=excel),
                patch.object(dashboard_app, "_active_shop_id", "templystudios"),
            ):
                result = asyncio.run(dashboard_app.import_from_factory(_request({
                    "folders": ["Kawaii-Planner"],
                    "auto_seo": False,
                })))

            self.assertTrue(result["ok"])
            self.assertEqual(result["imported"], 0)
            self.assertTrue(result["results"][0]["already_imported"])
            self.assertEqual(result["results"][0]["imported_folder"], "product-40")
            self.assertEqual(sorted(p.name for p in source.iterdir()), ["Etsy_SEO_Generator.xlsx", "Kawaii-Planner", "product-40"])

    def test_daisy_import_uses_own_source_and_cannot_import_temply_source(self):
        with tempfile.TemporaryDirectory() as temp:
            root, source, excel = self._source_root(temp, "daisyflowdigital")
            source_files = source / "Kawaii-Planner" / "source-zip"
            source_files.mkdir(parents=True)
            _write_image(source / "Kawaii-Planner" / "etsy-images-chatgpt-final" / "01_hero_image.png")
            (source_files / "New Product.zip").write_bytes(b"source")
            _make_workbook(excel)
            temply_source = root / FACTORY_ROOT
            _make_factory_source(temply_source, "Temply-Only", "Temply Product")

            with (
                patch.object(dashboard_app, "_factory_shop_dir", return_value=source),
                patch.object(dashboard_app, "_factory_excel_file", return_value=excel),
                patch.object(dashboard_app, "_active_shop_id", "daisyflowdigital"),
            ):
                result = asyncio.run(dashboard_app.import_from_factory(_request({
                    "folders": ["Kawaii-Planner"],
                    "auto_seo": False,
                }, shop_id="daisyflowdigital")))
                temply_attempt = asyncio.run(dashboard_app.import_from_factory(_request({
                    "folders": ["Temply-Only"],
                    "auto_seo": False,
                }, shop_id="daisyflowdigital")))

            self.assertTrue(result["ok"])
            self.assertEqual(1, result["imported"])
            self.assertEqual("daisyflowdigital", result["shop_id"])
            self.assertFalse(temply_attempt["results"][0]["ok"])
            self.assertEqual("Source folder not found", temply_attempt["results"][0]["error"])
            self.assertTrue((source / "product-01" / "files" / "New Product.zip").is_file())
            self.assertFalse((temply_source / "product-01").exists())
            wb = openpyxl.load_workbook(excel, data_only=True)
            self.assertEqual("🆕 Mới import · ⏳ Chờ đăng", wb["Listings"].cell(row=4, column=14).value)
            wb.close()
            with (
                patch.object(dashboard_app, "EXCEL_FILE", return_value=excel),
                patch.object(dashboard_app, "SHOP_DIR", return_value=source),
                patch.object(dashboard_app, "_active_shop_id", "daisyflowdigital"),
                patch.object(dashboard_app, "load_social_post_records", return_value={"products": {}}),
            ):
                imported_product = dashboard_app.products_from_excel()[0]
            self.assertEqual(
                "🆕 Mới import · ⏳ Chờ đăng · ⚠ Cần generate SEO",
                imported_product["status"],
            )
            self.assertTrue(imported_product["is_new_import"])

    def test_products_from_excel_preserves_new_import_pending_status_when_title_missing(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            _make_workbook(excel, [(4, "product-01", "Blank title token test")])

            wb = openpyxl.load_workbook(excel)
            ws = wb["Listings"]
            ws.cell(row=4, column=14, value="🆕 Mới import · ⏳ Chờ đăng")
            wb.save(excel)
            wb.close()

            product_dir = source / "product-01" / "images"
            _write_image(product_dir / "01_hero_image.png")

            with (
                patch.object(dashboard_app, "EXCEL_FILE", return_value=excel),
                patch.object(dashboard_app, "SHOP_DIR", return_value=source),
                patch.object(dashboard_app, "_active_shop_id", "templystudios"),
                patch.object(dashboard_app, "load_social_post_records", return_value={"products": {}}),
            ):
                imported_product = dashboard_app.products_from_excel()[0]

            self.assertEqual(
                "🆕 Mới import · ⏳ Chờ đăng · ⚠ Cần generate SEO",
                imported_product["status"],
            )
            self.assertTrue(imported_product["needs_seo"])

    def test_new_import_status_is_additive_and_idempotent(self):
        self.assertEqual(
            "🆕 Mới import · ✅ Đã đăng",
            dashboard_app._new_import_status("✅ Đã đăng"),
        )
        self.assertEqual(
            "🆕 Mới import · ⏳ Chờ đăng",
            dashboard_app._new_import_status("🆕 Mới import · ⏳ Chờ đăng"),
        )

    def test_preserve_new_import_status_only_when_currently_marked(self):
        self.assertEqual(
            "🆕 Mới import · ⚠ Sync lỗi",
            dashboard_app._preserve_new_import_status(
                "⚠ Sync lỗi",
                current_status="🆕 Mới import · ⏳ Chờ đăng",
            ),
        )
        self.assertEqual(
            "⚠ Sync lỗi",
            dashboard_app._preserve_new_import_status(
                "⚠ Sync lỗi",
                current_status="✅ Đã đăng",
            ),
        )

    def test_source_folder_rejects_path_traversal_and_non_source_folders(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, _ = self._source_root(temp)
            valid = source / "Kawaii-Planner"
            _write_image(valid / "etsy-images-chatgpt-final" / "01_hero_image.png")
            (valid / "source-zip").mkdir(parents=True)
            (valid / "source-zip" / "sample.zip").write_bytes(b"valid")
            (source / "product-03").mkdir(parents=True)
            with patch.object(dashboard_app, "_factory_shop_dir", return_value=source):
                self.assertEqual(dashboard_app._resolve_factory_source_folder("Kawaii-Planner"), source / "Kawaii-Planner")
                self.assertIsNone(dashboard_app._resolve_factory_source_folder("../Kawaii-Planner"))
                self.assertIsNone(dashboard_app._resolve_factory_source_folder("../product-03"))
                self.assertIsNone(dashboard_app._resolve_factory_source_folder("product-03"))
                self.assertIsNone(dashboard_app._resolve_factory_source_folder("missing"))

    def test_source_asset_and_catalog_directory_symlinks_are_rejected(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            outside = Path(temp) / "outside"
            outside.mkdir()
            _write_image(outside / "outside.png")
            (outside / "outside.zip").write_bytes(b"outside")

            image_link_source = source / "Image-Link-Source"
            image_link_source.mkdir()
            os.symlink(outside, image_link_source / "etsy-images-chatgpt-final")
            (image_link_source / "source-zip").mkdir()
            (image_link_source / "source-zip" / "safe.zip").write_bytes(b"safe")

            zip_link_source = source / "Zip-Link-Source"
            zip_link_source.mkdir()
            _write_image(zip_link_source / "etsy-images-chatgpt-final" / "01_hero_image.png")
            os.symlink(outside, zip_link_source / "source-zip")

            valid = _make_factory_source(source, "Valid-Source", "Valid Product")
            outside_image = outside / "linked-image.png"
            _write_image(outside_image)
            os.symlink(outside_image, valid / "etsy-images-chatgpt-final" / "linked-image.png")
            outside_zip = outside / "linked.zip"
            outside_zip.write_bytes(b"linked")
            os.symlink(outside_zip, valid / "source-zip" / "linked.zip")

            target = source / "product-02"
            target.mkdir()
            external_files = outside / "files"
            external_files.mkdir()
            (external_files / "Valid Product.zip").write_bytes(b"external")
            os.symlink(external_files, target / "files")
            _make_workbook(excel, [(4, "product-02", "Unrelated keyword")])

            with patch.object(dashboard_app, "_factory_shop_dir", return_value=source):
                self.assertFalse(dashboard_app._is_supported_factory_source(image_link_source))
                self.assertFalse(dashboard_app._is_supported_factory_source(zip_link_source))
                self.assertIsNone(dashboard_app._resolve_factory_source_folder("Image-Link-Source"))
                self.assertEqual(
                    [p.name for p in dashboard_app._factory_image_files(valid)],
                    ["01_hero_image.png"],
                )
                self.assertEqual(
                    [p.name for p in dashboard_app._factory_download_files(valid)],
                    ["Valid Product.zip"],
                )
                folders = dashboard_app._scan_factory_folders(source, source, excel)
                valid_row = next(item for item in folders if item["name"] == "Valid-Source")
                self.assertFalse(valid_row["already_imported"])
                with self.assertRaises(dashboard_app.HTTPException):
                    asyncio.run(dashboard_app.factory_thumb("Image-Link-Source", "outside.png"))

    def test_import_seeds_column_d_with_factory_keyword(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            source_files = source / "Kawaii-Planner" / "source-zip"
            source_files.mkdir(parents=True)
            _write_image(source / "Kawaii-Planner" / "etsy-images-chatgpt-final" / "01_hero_image.png")
            (source_files / "Kawaii-Planner.zip").write_bytes(b"source")
            _make_workbook(excel)

            with (
                patch.object(dashboard_app, "_factory_shop_dir", return_value=source),
                patch.object(dashboard_app, "_factory_excel_file", return_value=excel),
                patch.object(dashboard_app, "_active_shop_id", "templystudios"),
            ):
                result = asyncio.run(dashboard_app.import_from_factory(_request({
                    "folders": ["Kawaii-Planner"],
                    "auto_seo": False,
                })))

            self.assertEqual(1, result["imported"])
            wb = openpyxl.load_workbook(excel, data_only=True)
            self.assertEqual("Kawaii Planner", wb["Listings"].cell(row=4, column=4).value)
            wb.close()

    def test_import_preserves_partial_legacy_row_and_uses_fully_empty_row(self):
        with tempfile.TemporaryDirectory() as temp:
            _, source, excel = self._source_root(temp)
            source_files = source / "Kawaii-Planner" / "source-zip"
            source_files.mkdir(parents=True)
            _write_image(source / "Kawaii-Planner" / "etsy-images-chatgpt-final" / "01_hero_image.png")
            (source_files / "New Product.zip").write_bytes(b"source")
            _make_workbook(excel)
            wb = openpyxl.load_workbook(excel)
            ws = wb["Listings"]
            ws.cell(row=4, column=3, value="legacy keyword")
            ws.cell(row=4, column=14, value="legacy status")
            ws.cell(row=6, column=2, value="product-legacy")
            wb.save(excel)
            wb.close()

            with (
                patch.object(dashboard_app, "_factory_shop_dir", return_value=source),
                patch.object(dashboard_app, "_factory_excel_file", return_value=excel),
                patch.object(dashboard_app, "_active_shop_id", "templystudios"),
            ):
                result = asyncio.run(dashboard_app.import_from_factory(_request({
                    "folders": ["Kawaii-Planner"],
                    "auto_seo": False,
                })))

            self.assertEqual(5, result["results"][0]["row"])
            wb = openpyxl.load_workbook(excel, data_only=True)
            ws = wb["Listings"]
            self.assertIsNone(ws.cell(row=4, column=2).value)
            self.assertEqual("legacy keyword", ws.cell(row=4, column=3).value)
            self.assertEqual("legacy status", ws.cell(row=4, column=14).value)
            self.assertEqual("New Product", ws.cell(row=5, column=4).value)
            wb.close()


class SyncFactoryToShopTests(unittest.TestCase):
    def test_supported_factory_source_guarded_by_layout_and_exclusions(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp) / "shops" / sync_factory.FACTORY_SHOP_ID
            (root / "A-Nested-Layout" / "etsy-images-chatgpt-final").mkdir(parents=True)
            (root / "A-Nested-Layout" / "source-zip").mkdir(parents=True)
            (root / "B-Legacy-Layout" / "images").mkdir(parents=True)
            (root / "C-Legacy-Layout" / "files").mkdir(parents=True)
            (root / "product-01" / "source-zip").mkdir(parents=True)
            (root / "_deleted_products" / "etsy-images-chatgpt-final").mkdir(parents=True)
            (root / "_asset_quarantine" / "source-zip").mkdir(parents=True)
            (root / "master_products" / "source-zip").mkdir(parents=True)
            (root / "dot" / ".hidden").mkdir(parents=True)

            self.assertTrue(sync_factory.is_supported_factory_source(root / "A-Nested-Layout"))
            self.assertFalse(sync_factory.is_supported_factory_source(root / "B-Legacy-Layout"))
            self.assertFalse(sync_factory.is_supported_factory_source(root / "C-Legacy-Layout"))
            self.assertFalse(sync_factory.is_supported_factory_source(root / "product-01"))
            self.assertFalse(sync_factory.is_supported_factory_source(root / "_deleted_products"))
            self.assertFalse(sync_factory.is_supported_factory_source(root / "_asset_quarantine"))
            self.assertFalse(sync_factory.is_supported_factory_source(root / "master_products"))
            self.assertFalse(sync_factory.is_supported_factory_source(root / ".DS_Store"))

    def test_factory_source_collectors_use_nested_layout(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp) / "shops" / sync_factory.FACTORY_SHOP_ID
            source = root / "Kawaii-Planner"
            (source / "etsy-images-chatgpt-final").mkdir(parents=True)
            (source / "source-zip").mkdir(parents=True)
            (source / "images").mkdir()
            (source / "files").mkdir()
            (source / "etsy-images-chatgpt-final" / "01_hero_image.png").write_bytes(b"hero")
            (source / "etsy-images-chatgpt-final" / "01_alt.jpg").write_bytes(b"hero2")
            (source / "images" / "legacy_hero.jpg").write_bytes(b"legacy")
            (source / "source-zip" / "new_source.zip").write_bytes(b"zip")
            (source / "files" / "legacy.pdf").write_bytes(b"legacypdf")

            self.assertEqual(
                [p.name for p in sync_factory.get_factory_images(source)],
                ["01_hero_image.png", "01_alt.jpg"],
            )
            self.assertEqual(
                [p.name for p in sync_factory.get_factory_files(source)],
                ["new_source.zip"],
            )

    def test_main_rejects_non_temply_shop(self):
        with patch.object(sync_factory.argparse.ArgumentParser, "parse_args") as parse_args:
            parse_args.return_value = sync_factory.argparse.Namespace(shop="daisyflowdigital", watch=False, poll=5)
            with patch.object(sync_factory, "sync_shop") as sync_shop:
                with patch.object(sync_factory, "watch_folders") as watch_folders:
                    with patch("builtins.print"):
                        sync_factory.main()
                    sync_shop.assert_not_called()
                    watch_folders.assert_not_called()

    def test_sync_reports_blocked_when_active_shop_is_not_temply(self):
        with tempfile.TemporaryDirectory() as temp:
            active_shop_file = Path(temp) / "active_shop.txt"
            active_shop_file.write_text("daisyflowdigital", encoding="utf-8")
            with patch.object(sync_factory, "ACTIVE_SHOP_FILE", active_shop_file):
                self.assertFalse(sync_factory.sync_shop(sync_factory.FACTORY_SHOP_ID))


if __name__ == "__main__":
    unittest.main()
