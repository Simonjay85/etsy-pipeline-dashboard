from __future__ import annotations

import asyncio
import hashlib
import json
import os
import shutil
import tempfile
import threading
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from openpyxl.workbook.workbook import Workbook

import dashboard_app


class _JsonRequest:
    def __init__(self, payload: dict):
        self.payload = payload

    async def json(self) -> dict:
        return self.payload


class _ResolverStore:
    def __init__(
        self,
        asset_root: Path,
        *,
        error: Exception | None = None,
        manifest: bool = True,
        tamper: bool = False,
        tamper_size: bool = False,
        zero_byte: bool = False,
    ):
        self.asset_root = asset_root
        self.error = error
        self.manifest = manifest
        self.tamper = tamper
        self.tamper_size = tamper_size
        self.zero_byte = zero_byte
        self.threads: list[str] = []
        self.sources: list[Path] = []

    def resolve_asset_root(self, source: Path) -> dict:
        self.threads.append(threading.current_thread().name)
        self.sources.append(source)
        if self.error is not None:
            raise self.error
        images = sorted((self.asset_root / "images").rglob("*"))
        files = sorted((self.asset_root / "files").rglob("*"))
        result = {
            "ok": True,
            "mode": "cloud",
            "source": "cloud-cache",
            "asset_root": str(self.asset_root),
            "image_paths": [str(path) for path in images if path.is_file()],
            "file_paths": [str(path) for path in files if path.is_file()],
            "manifest_sha256": "a" * 64,
        }
        if self.manifest:
            records = []
            for role, paths in (("image", images), ("file", files)):
                for path in paths:
                    if not path.is_file():
                        continue
                    data = path.read_bytes()
                    digest = hashlib.sha256(data).hexdigest()
                    advertised_size = len(data)
                    if self.tamper_size and not records:
                        advertised_size += 1
                    if self.zero_byte and not data:
                        advertised_size = 1
                    if self.tamper and not records:
                        digest = "0" * 64
                    records.append(
                        {
                            "path": path.relative_to(self.asset_root).as_posix(),
                            "role": role,
                            "size": advertised_size,
                            "sha256": digest,
                        }
                    )
            result["manifest"] = {"files": records}
        return result


class DashboardCrossShopSyncTests(unittest.IsolatedAsyncioTestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory(prefix="etsy-cross-shop-sync-")
        self.root = Path(self.temporary.name)
        self.source_shop = self.root / "shops" / "source-shop"
        self.source_product = self.source_shop / "product-01"
        (self.source_product / "images").mkdir(parents=True)
        (self.source_product / "files").mkdir()
        (self.source_product / ".cloud-assets.json").write_text(
            json.dumps({"state": "CLOUD_ONLY", "current_revision": "rev-cloud"}),
            encoding="utf-8",
        )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Listings"
        worksheet.cell(row=4, column=2, value="product-01")
        worksheet.cell(row=4, column=3, value="cloud keywords")
        worksheet.cell(row=4, column=8, value="Cloud source title")
        worksheet.cell(row=4, column=14, value="⏳ Chờ đăng")
        self.source_shop.mkdir(parents=True, exist_ok=True)
        workbook.save(self.source_shop / "Etsy_SEO_Generator.xlsx")

        self.asset_root = self.root / "output" / "cloud-cache" / "source-product-01" / "rev-cloud"
        for index in range(1, 11):
            image = self.asset_root / "images" / "gallery" / f"image-{index:02d}.png"
            image.parent.mkdir(parents=True, exist_ok=True)
            image.write_bytes(f"image-{index}".encode())
        (self.asset_root / "files" / "delivery" ).mkdir(parents=True, exist_ok=True)
        (self.asset_root / "files" / "delivery" / "source.zip").write_bytes(b"cloud-file")

        self.patches = [
            patch.object(dashboard_app, "BASE_DIR", self.root),
            patch.object(dashboard_app, "_active_shop_id", "source-shop"),
            patch.object(
                dashboard_app,
                "SHOPS",
                {"source-shop": {"id": "source-shop"}, "target-shop": {"id": "target-shop"}},
            ),
        ]
        for item in self.patches:
            item.start()

    def tearDown(self) -> None:
        for item in reversed(self.patches):
            item.stop()
        self.temporary.cleanup()

    def _add_source_product(self, *, row: int = 5, folder: str = "product-02") -> None:
        product = self.source_shop / folder
        (product / "images").mkdir(parents=True)
        (product / "files").mkdir()
        workbook_path = self.source_shop / "Etsy_SEO_Generator.xlsx"
        workbook = openpyxl.load_workbook(workbook_path)
        worksheet = workbook["Listings"]
        worksheet.cell(row=row, column=2, value=folder)
        worksheet.cell(row=row, column=3, value=f"keywords-{folder}")
        worksheet.cell(row=row, column=8, value=f"Title {folder}")
        worksheet.cell(row=row, column=14, value="⏳ Chờ đăng")
        workbook.save(workbook_path)

    async def test_cloud_only_source_is_hydrated_off_loop_and_copies_only_assets(self) -> None:
        store = _ResolverStore(self.asset_root)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            result = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": True})
            )

        self.assertTrue(result["ok"])
        self.assertEqual(result["synced"], 1)
        self.assertEqual(result["mode"], "seo+assets")
        self.assertEqual(result["asset_counts"], {"images": 10, "files": 1})
        self.assertEqual(result["items"][0]["asset_counts"], {"images": 10, "files": 1})
        self.assertEqual(store.sources, [self.source_product])
        self.assertTrue(store.threads)
        self.assertTrue(all(name != threading.main_thread().name for name in store.threads))

        destination = self.root / "shops" / "target-shop" / "product-01"
        self.assertEqual(len(list((destination / "images").rglob("*.png"))), 10)
        self.assertEqual((destination / "files" / "delivery" / "source.zip").read_bytes(), b"cloud-file")
        self.assertFalse((destination / ".cloud-assets.json").exists())
        self.assertFalse((destination / ".cloud-assets.lock").exists())
        self.assertFalse((destination / ".cloud-preview.webp").exists())
        self.assertEqual(json.loads((self.source_product / ".cloud-assets.json").read_text())["state"], "CLOUD_ONLY")
        self.assertFalse(any(self.source_product.joinpath(name).exists() for name in ("images/image-01.png", "files/source.zip")))

    async def test_resolver_failure_is_not_reported_as_success(self) -> None:
        store = _ResolverStore(self.asset_root, error=dashboard_app.AssetValidationError("cloud manifest unavailable"))
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": True})
            )

        self.assertEqual(response.status_code, 409)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        self.assertEqual(payload["items"], [])
        self.assertFalse((self.root / "shops" / "target-shop" / "product-01").exists())
        self.assertFalse((self.root / "shops" / "target-shop" / "Etsy_SEO_Generator.xlsx").exists())
        self.assertFalse((self.root / "shops" / "target-shop").exists())

    async def test_copy_files_false_keeps_seo_only_mode_and_does_not_resolve_assets(self) -> None:
        store = _ResolverStore(
            self.asset_root,
            error=AssertionError("SEO-only sync must not resolve assets"),
        )
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            result = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4]})
            )

        self.assertTrue(result["ok"])
        self.assertEqual(result["synced"], 1)
        self.assertEqual(result["mode"], "seo-only")
        self.assertEqual(result["asset_counts"], {"images": 0, "files": 0})
        self.assertEqual(result["items"][0]["asset_counts"], {"images": 0, "files": 0})
        self.assertEqual(store.sources, [])
        destination = self.root / "shops" / "target-shop" / "product-01"
        self.assertEqual(list((destination / "images").iterdir()), [])
        self.assertEqual(list((destination / "files").iterdir()), [])

    async def test_invalid_conflict_resolution_is_rejected_before_target_mutation(self) -> None:
        for invalid in ("overwrite", [], {}):
            with self.subTest(invalid=invalid):
                response = await dashboard_app.sync_to_shop(
                    _JsonRequest(
                        {
                            "target_shop": "target-shop",
                            "rows": [4],
                            "copy_files": False,
                            "conflict_resolution": invalid,
                        }
                    )
                )

                self.assertEqual(response.status_code, 400)
                payload = json.loads(response.body.decode("utf-8"))
                self.assertFalse(payload["ok"])
                self.assertIn("conflict_resolution", payload["error"])
        self.assertFalse((self.root / "shops" / "target-shop").exists())

    async def test_staging_failure_after_resolver_cleans_new_target(self) -> None:
        store = _ResolverStore(self.asset_root)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store), patch.object(
            dashboard_app,
            "_stage_sync_product_assets",
            side_effect=dashboard_app.AssetValidationError("staging failed"),
        ):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": True})
            )

        self.assertEqual(response.status_code, 409)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        self.assertEqual(store.sources, [self.source_product])
        self.assertFalse((self.root / "shops" / "target-shop").exists())

    async def test_later_product_failure_rolls_back_earlier_product_and_workbook(self) -> None:
        self._add_source_product()
        store = _ResolverStore(self.asset_root)
        original_copy = dashboard_app._sync_copy_staged_asset_dirs
        calls = 0

        def fail_on_second_copy(stage_root: Path, destination: Path) -> None:
            nonlocal calls
            calls += 1
            if calls == 2:
                raise dashboard_app.AssetValidationError("second product failed")
            original_copy(stage_root, destination)

        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store), patch.object(
            dashboard_app,
            "_sync_copy_staged_asset_dirs",
            side_effect=fail_on_second_copy,
        ):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4, 5], "copy_files": True})
            )

        self.assertEqual(response.status_code, 409)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        self.assertEqual(calls, 2)
        target_shop = self.root / "shops" / "target-shop"
        self.assertFalse(target_shop.exists())

    async def test_workbook_save_failure_restores_existing_merge_assets_and_workbook(self) -> None:
        target_shop = self.root / "shops" / "target-shop"
        target_product = target_shop / "product-01"
        (target_product / "images").mkdir(parents=True)
        (target_product / "images" / "original.png").write_bytes(b"original")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Listings"
        worksheet.cell(row=4, column=2, value="product-01")
        worksheet.cell(row=4, column=3, value="cloud keywords")
        worksheet.cell(row=4, column=8, value="Cloud source title")
        workbook.save(target_shop / "Etsy_SEO_Generator.xlsx")
        workbook_bytes = (target_shop / "Etsy_SEO_Generator.xlsx").read_bytes()

        store = _ResolverStore(self.asset_root)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store), patch.object(
            Workbook,
            "save",
            side_effect=OSError("simulated workbook save failure"),
        ):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest(
                    {
                        "target_shop": "target-shop",
                        "rows": [4],
                        "copy_files": True,
                        "conflict_resolution": "merge",
                    }
                )
            )

        self.assertEqual(response.status_code, 500)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        self.assertEqual((target_shop / "Etsy_SEO_Generator.xlsx").read_bytes(), workbook_bytes)
        self.assertEqual((target_product / "images" / "original.png").read_bytes(), b"original")
        self.assertFalse((target_product / "images" / "gallery").exists())
        self.assertFalse((target_product / "files").exists())

    async def test_template_copy_failure_cleans_new_target_shop_and_workbook(self) -> None:
        with patch.object(
            dashboard_app,
            "copy_and_clean_template",
            side_effect=OSError("simulated template copy failure"),
        ):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4]})
            )

        self.assertEqual(response.status_code, 500)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        target_shop = self.root / "shops" / "target-shop"
        self.assertFalse((target_shop / "Etsy_SEO_Generator.xlsx").exists())
        self.assertFalse(target_shop.exists())

    async def test_restore_asset_dirs_preserves_absent_content_tree(self) -> None:
        destination = self.root / "shops" / "target-shop" / "product-01"
        (destination / "images").mkdir(parents=True)
        (destination / "images" / "original.png").write_bytes(b"original")
        backup_root, existing = dashboard_app._sync_backup_asset_dirs(destination)
        try:
            (destination / "images" / "replacement.png").write_bytes(b"replacement")
            (destination / "files").mkdir()
            (destination / "files" / "unexpected.zip").write_bytes(b"unexpected")
            dashboard_app._sync_restore_asset_dirs(destination, backup_root, existing)
        finally:
            shutil.rmtree(backup_root, ignore_errors=True)

        self.assertEqual((destination / "images" / "original.png").read_bytes(), b"original")
        self.assertFalse((destination / "images" / "replacement.png").exists())
        self.assertFalse((destination / "files").exists())

    async def test_merge_with_same_hash_but_destination_cloud_identity_fails_closed(self) -> None:
        target_shop = self.root / "shops" / "target-shop"
        target_shop.mkdir(parents=True)
        target_product = target_shop / "product-01"
        (target_product / "images").mkdir(parents=True)
        (target_product / "files").mkdir()
        (target_product / ".cloud-assets.json").write_text(
            json.dumps(
                {
                    "state": "CLOUD_VERIFIED",
                    "current_manifest_sha256": "a" * 64,
                    "product": {"key": "shops/target-shop/product-01"},
                }
            ),
            encoding="utf-8",
        )
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Listings"
        worksheet.cell(row=4, column=2, value="product-01")
        worksheet.cell(row=4, column=3, value="cloud keywords")
        worksheet.cell(row=4, column=8, value="Cloud source title")
        workbook.save(target_shop / "Etsy_SEO_Generator.xlsx")

        store = _ResolverStore(self.asset_root)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest(
                    {
                        "target_shop": "target-shop",
                        "rows": [4],
                        "copy_files": True,
                        "conflict_resolution": "merge",
                    }
                )
            )

        self.assertEqual(response.status_code, 409)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        self.assertIn("cloud identity", payload["error"])
        destination_state = json.loads((target_product / ".cloud-assets.json").read_text())
        self.assertEqual(destination_state["current_manifest_sha256"], "a" * 64)
        self.assertEqual(destination_state["product"]["key"], "shops/target-shop/product-01")

    async def test_merge_replaces_asset_trees_and_removes_stale_destination_files(self) -> None:
        target_shop = self.root / "shops" / "target-shop"
        target_shop.mkdir(parents=True)
        target_product = target_shop / "product-01"
        (target_product / "images").mkdir(parents=True)
        (target_product / "files").mkdir()
        (target_product / "images" / "stale.png").write_bytes(b"stale")
        (target_product / "files" / "stale.zip").write_bytes(b"stale")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Listings"
        worksheet.cell(row=4, column=2, value="product-01")
        worksheet.cell(row=4, column=3, value="cloud keywords")
        worksheet.cell(row=4, column=8, value="Cloud source title")
        workbook.save(target_shop / "Etsy_SEO_Generator.xlsx")

        store = _ResolverStore(self.asset_root)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest(
                    {
                        "target_shop": "target-shop",
                        "rows": [4],
                        "copy_files": True,
                        "conflict_resolution": "merge",
                    }
                )
            )

        self.assertTrue(response["ok"])
        self.assertFalse((target_product / "images" / "stale.png").exists())
        self.assertFalse((target_product / "files" / "stale.zip").exists())
        self.assertEqual(len(list((target_product / "images").rglob("*.png"))), 10)
        self.assertEqual((target_product / "files" / "delivery" / "source.zip").read_bytes(), b"cloud-file")

    async def test_manifest_size_and_hash_mismatch_fails_before_destination_commit(self) -> None:
        store = _ResolverStore(self.asset_root, manifest=True, tamper=True)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": True})
            )

        self.assertEqual(response.status_code, 409)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        self.assertIn("sha256", payload["error"])
        self.assertFalse((self.root / "shops" / "target-shop").exists())

    async def test_manifest_size_only_mismatch_fails_before_destination_commit(self) -> None:
        store = _ResolverStore(self.asset_root, manifest=True, tamper_size=True)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": True})
            )

        self.assertEqual(response.status_code, 409)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertIn("size", payload["error"])
        self.assertFalse((self.root / "shops" / "target-shop").exists())

    async def test_manifest_listed_zero_byte_asset_is_rejected(self) -> None:
        zero_image = self.asset_root / "images" / "gallery" / "zero.png"
        zero_image.write_bytes(b"")
        store = _ResolverStore(self.asset_root, manifest=True, zero_byte=True)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": True})
            )

        self.assertEqual(response.status_code, 409)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertIn("usable", payload["error"])
        self.assertFalse((self.root / "shops" / "target-shop").exists())

    async def test_cloud_cache_resolution_without_manifest_fails_before_destination_commit(self) -> None:
        store = _ResolverStore(self.asset_root, manifest=False)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": True})
            )

        self.assertEqual(response.status_code, 409)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        self.assertIn("manifest.files", payload["error"])
        self.assertFalse((self.root / "shops" / "target-shop").exists())

    async def test_multi_item_success_aggregates_per_item_asset_counts(self) -> None:
        self._add_source_product()
        store = _ResolverStore(self.asset_root)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            result = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4, 5], "copy_files": True})
            )

        self.assertTrue(result["ok"])
        self.assertEqual(result["synced"], 2)
        self.assertEqual(result["asset_counts"], {
            "images": sum(item["asset_counts"]["images"] for item in result["items"]),
            "files": sum(item["asset_counts"]["files"] for item in result["items"]),
        })
        self.assertEqual(result["asset_counts"], {"images": 20, "files": 2})

    async def test_copy_files_requires_json_boolean(self) -> None:
        for invalid in ("false", 0, 1, [], {}):
            with self.subTest(invalid=invalid):
                response = await dashboard_app.sync_to_shop(
                    _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": invalid})
                )
                self.assertEqual(response.status_code, 400)
                self.assertFalse(json.loads(response.body.decode("utf-8"))["ok"])
        self.assertFalse((self.root / "shops" / "target-shop").exists())

    async def test_resolver_failure_rejects_symlinked_target_shop(self) -> None:
        real_target = self.root / "real-target"
        real_target.mkdir()
        symlink_target = self.root / "shops" / "target-shop"
        try:
            os.symlink(real_target, symlink_target)
        except (OSError, NotImplementedError):
            self.skipTest("filesystem does not support symlinks")

        store = _ResolverStore(self.asset_root)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4], "copy_files": True})
            )
        self.assertEqual(response.status_code, 409)
        self.assertIn("symlink", json.loads(response.body.decode("utf-8"))["error"])
        self.assertFalse((real_target / "Etsy_SEO_Generator.xlsx").exists())

    async def test_excel_failure_removes_request_created_shop_workbook_and_product(self) -> None:
        blank_template = self.root / "blank-template.xlsx"
        blank_workbook = Workbook()
        blank_workbook.active.title = "Listings"
        blank_workbook.save(blank_template)

        def copy_template_without_save(source: Path, destination: Path) -> None:
            shutil.copy2(blank_template, destination)

        store = _ResolverStore(self.asset_root)
        with patch.object(dashboard_app, "CLOUD_ASSET_STORE", store), patch.object(
            dashboard_app, "copy_and_clean_template", side_effect=copy_template_without_save
        ), patch.object(Workbook, "save", side_effect=OSError("simulated Excel failure")):
            response = await dashboard_app.sync_to_shop(
                _JsonRequest({"target_shop": "target-shop", "rows": [4]})
            )

        self.assertEqual(response.status_code, 500)
        payload = json.loads(response.body.decode("utf-8"))
        self.assertFalse(payload["ok"])
        self.assertEqual(payload["synced"], 0)
        self.assertFalse((self.root / "shops" / "target-shop" / "product-01").exists())
        self.assertFalse((self.root / "shops" / "target-shop" / "Etsy_SEO_Generator.xlsx").exists())
        self.assertFalse((self.root / "shops" / "target-shop").exists())


if __name__ == "__main__":
    unittest.main()
