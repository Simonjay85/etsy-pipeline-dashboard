#!/usr/bin/env python3
"""Safely compact Daisy Flow's physical ``product-*`` folders.

Dry-run is the default.  Apply is deliberately fail-closed and requires the
dashboard and every other Etsy/rclone writer to be offline.  Remote cloud
revisions are copied and verified under their new product identity before any
local folder, workbook, or source-map mutation is committed.
"""

from __future__ import annotations

import argparse
import copy
import fcntl
import hashlib
import json
import os
import re
import shutil
import socket
import sqlite3
import subprocess
import sys
import tempfile
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook

# Keep direct ``python3 scripts/...`` execution aligned with project modules.
_IMPORT_ROOT = Path(__file__).resolve().parents[1]
if str(_IMPORT_ROOT) not in sys.path:
    sys.path.insert(0, str(_IMPORT_ROOT))

from cloud_asset_store import (
    LocalRemote,
    ProductIdentity,
    RcloneRemote,
    RemoteStore,
    RemoteStoreError,
    canonical_json_bytes,
    verify_manifest_directory,
)
from cloud_asset_store_config import load_config


SHOP_ID = "daisyflowdigital"
EXPECTED_FOLDERS = 198
EXPECTED_BINDINGS = 197
EXPECTED_STALE = 4
EXPECTED_STALE_NAMES = ("product-243", "product-371", "product-374", "product-377")
EXPECTED_SOURCE_MAP_UPDATES = 95
EXPECTED_CLOUD_MOVES = (
    ("product-64", "product-27"), ("product-74", "product-32"),
    ("product-400", "product-190"), ("product-404", "product-193"),
    ("product-410", "product-195"), ("product-418", "product-197"),
)
PRODUCT_RE = re.compile(r"^product-(\d+)$")
STATE_NAME = ".cloud-assets.json"
PRODUCT_LOCK_NAME = ".cloud-assets.lock"
MIGRATION_LOCK_NAME = ".daisy-folder-renumber.lock"
MANIFEST_NAME = "mapping_manifest.json"
JOURNAL_NAME = "migration_journal.json"
CANONICAL_REPO = Path("/Users/aaronnguyen/Developer/Etsy")
ACTIVE_CLOUD_STATES = {"UPLOADING"}
WRITER_PROCESS_NAMES = (
    "dashboard_app.py",
    "etsy_auto_post.py",
    "etsy_push_update.py",
    "etsy_shop_sync.py",
    "etsy_clean_duplicates.py",
    "etsy_repair.py",
    "etsy_scan_errors.py",
    "backup_etsy_to_drive.py",
    "bulk_create_unmapped_listings.py",
    "cloud_asset_cli.py",
    "cloud_asset_pilot.py",
    "image_factory_watcher.py",
    "social_auto_post.py",
    "social_bulk_post.py",
    "gumroad_setup_excel.py",
    "gumroad_auto_post.py",
)


class MigrationError(RuntimeError):
    """A preflight, remote, mutation, or rollback failure."""


@dataclass(frozen=True)
class FolderMapping:
    old_name: str
    new_name: str
    temporary_name: str
    rollback_name: str


@dataclass(frozen=True)
class WorkbookEdit:
    row: int
    old_value: str
    new_value: str | None


@dataclass(frozen=True)
class CloudMove:
    old_name: str
    new_name: str
    old_identity: ProductIdentity
    new_identity: ProductIdentity
    state: dict[str, Any]
    new_state: dict[str, Any]
    revision: str
    manifest_sha256: str


@dataclass
class MigrationPlan:
    repo_root: Path
    shop_dir: Path
    workbook_path: Path
    source_map_path: Path
    mappings: tuple[FolderMapping, ...]
    workbook_edits: tuple[WorkbookEdit, ...]
    mapped_binding_count: int
    stale_binding_count: int
    unregistered_old: tuple[str, ...]
    updated_source_map: dict[str, Any] | None
    source_map_updates: int
    cloud_moves: tuple[CloudMove, ...]
    source_hashes: dict[str, str]
    inventory: dict[str, tuple[int, ...]]
    workbook_snapshot: dict[str, Any]


def utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
    except OSError as exc:
        raise MigrationError(f"cannot hash required file {path}: {exc}") from exc
    return digest.hexdigest()


def _atomic_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, name = tempfile.mkstemp(prefix=f".{path.name}.tmp-", dir=path.parent)
    temporary = Path(name)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(canonical_json_bytes(value))
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        descriptor = os.open(path.parent, os.O_RDONLY)
        try:
            os.fsync(descriptor)
        finally:
            os.close(descriptor)
    finally:
        temporary.unlink(missing_ok=True)


def _atomic_pretty_json(path: Path, value: Any) -> None:
    data = (json.dumps(value, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, name = tempfile.mkstemp(prefix=f".{path.name}.tmp-", dir=path.parent)
    temporary = Path(name)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _atomic_copy(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    fd, name = tempfile.mkstemp(prefix=f".{destination.name}.tmp-", dir=destination.parent)
    os.close(fd)
    temporary = Path(name)
    try:
        shutil.copy2(source, temporary)
        with temporary.open("rb") as handle:
            os.fsync(handle.fileno())
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)


def _atomic_save_workbook(workbook: Any, destination: Path) -> None:
    fd, name = tempfile.mkstemp(prefix=f".{destination.name}.tmp-", suffix=".xlsx", dir=destination.parent)
    os.close(fd)
    temporary = Path(name)
    try:
        workbook.save(temporary)
        with temporary.open("rb") as handle:
            os.fsync(handle.fileno())
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)
        workbook.close()


def product_name(index: int) -> str:
    return f"product-{index:02d}" if index < 100 else f"product-{index}"


def discover_folders(shop_dir: Path, expected_count: int | None = EXPECTED_FOLDERS) -> tuple[str, ...]:
    if not shop_dir.is_dir():
        raise MigrationError(f"shop directory is missing: {shop_dir}")
    unexpected = sorted(
        child.name for child in shop_dir.iterdir()
        if child.is_dir() and child.name.startswith("product-") and not PRODUCT_RE.fullmatch(child.name)
    )
    if unexpected:
        raise MigrationError("unexpected product directory names: " + ", ".join(unexpected))
    parsed: list[tuple[int, str]] = []
    seen_numbers: set[int] = set()
    for child in shop_dir.iterdir():
        if child.name.startswith("product-") and child.is_symlink():
            raise MigrationError(f"product folder symlink is unsafe: {child}")
        match = PRODUCT_RE.fullmatch(child.name) if child.is_dir() else None
        if not match:
            continue
        number = int(match.group(1))
        if number <= 0 or number in seen_numbers:
            raise MigrationError(f"duplicate/invalid numeric product directory: {child.name}")
        seen_numbers.add(number)
        parsed.append((number, child.name))
    parsed.sort()
    if expected_count is not None and len(parsed) != expected_count:
        raise MigrationError(f"expected {expected_count} physical product folders, found {len(parsed)}")
    return tuple(name for _, name in parsed)


def make_mapping(names: Iterable[str], token: str | None = None) -> tuple[FolderMapping, ...]:
    token = token or uuid.uuid4().hex
    result = []
    for index, old_name in enumerate(names, 1):
        result.append(FolderMapping(
            old_name, product_name(index),
            f".daisy-folder-renumber-{token}-{index:03d}",
            f".daisy-folder-renumber-rollback-{token}-{index:03d}",
        ))
    targets = [item.new_name for item in result]
    if len(targets) != len(set(targets)):
        raise MigrationError("generated mapping contains duplicate targets")
    return tuple(result)


def _cell_snapshot(cell: Any, omit_value: bool) -> tuple[Any, ...]:
    return (
        None if omit_value else cell.value,
        None if omit_value else cell.data_type,
        tuple(cell._style) if cell.has_style else None,
        cell.hyperlink.target if cell.hyperlink else None,
        None if cell.comment is None else (cell.comment.text, cell.comment.author),
    )


def workbook_snapshot(workbook: Any, editable_rows: Iterable[int]) -> dict[str, Any]:
    editable = set(editable_rows)
    result: dict[str, Any] = {"sheetnames": tuple(workbook.sheetnames), "sheets": {}}
    for sheet in workbook.worksheets:
        if sheet.title == "Listings":
            # A cleared B cell may be omitted from openpyxl's sparse _cells
            # mapping after save/reload. Materialize all permitted edit cells
            # so before/after snapshots compare the same coordinates.
            for row in editable:
                sheet.cell(row, 2)
        cells: dict[str, tuple[Any, ...]] = {}
        for cell in sheet._cells.values():
            permitted_edit = (
                sheet.title == "Listings" and cell.column == 2 and cell.row in editable
            )
            snapshot = _cell_snapshot(cell, permitted_edit)
            # Some real workbooks contain serialized empty nodes such as
            # ``<c r="B102" t="n"></c>``. openpyxl safely omits these on the
            # next save. Treat only completely contentless, unstyled,
            # unannotated cells as equivalent to absence; meaningful empty
            # cells (style/comment/hyperlink) remain strict invariants.
            if not permitted_edit and snapshot == (None, "n", None, None, None):
                continue
            cells[cell.coordinate] = snapshot
        result["sheets"][sheet.title] = {
            "state": sheet.sheet_state,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            "auto_filter": sheet.auto_filter.ref,
            "merged": tuple(str(value) for value in sheet.merged_cells.ranges),
            "row_dimensions": {
                str(key): (value.hidden, value.outlineLevel, value.collapsed, value.height, value.style)
                for key, value in sorted(sheet.row_dimensions.items())
            },
            "column_dimensions": {
                str(key): (value.hidden, value.outlineLevel, value.collapsed, value.width, value.style)
                for key, value in sorted(sheet.column_dimensions.items())
            },
            "cells": cells,
        }
    return result


def workbook_plan(
    path: Path,
    mapping: dict[str, str],
    *,
    expected_bindings: int | None = EXPECTED_BINDINGS,
) -> tuple[tuple[WorkbookEdit, ...], int, int, tuple[str, ...], dict[str, Any]]:
    try:
        workbook = load_workbook(path, data_only=False)
    except Exception as exc:
        raise MigrationError(f"invalid workbook {path}: {exc}") from exc
    if "Listings" not in workbook.sheetnames:
        workbook.close()
        raise MigrationError("workbook has no Listings sheet")
    sheet = workbook["Listings"]
    edits: list[WorkbookEdit] = []
    seen: dict[str, int] = {}
    mapped = stale = 0
    for row in range(4, sheet.max_row + 1):
        value = sheet.cell(row, 2).value
        if value in (None, ""):
            continue
        if not isinstance(value, str) or not PRODUCT_RE.fullmatch(value):
            workbook.close()
            raise MigrationError(f"Listings!B{row} has invalid folder reference {value!r}")
        if value in seen:
            workbook.close()
            raise MigrationError(f"duplicate workbook folder reference {value}: rows {seen[value]} and {row}")
        seen[value] = row
        if value in mapping:
            edits.append(WorkbookEdit(row, value, mapping[value]))
            mapped += 1
        else:
            edits.append(WorkbookEdit(row, value, None))
            stale += 1
    if expected_bindings is not None and mapped != expected_bindings:
        workbook.close()
        raise MigrationError(f"expected {expected_bindings} physical workbook bindings, found {mapped}")
    unregistered = tuple(sorted(set(mapping) - set(seen), key=lambda name: int(PRODUCT_RE.fullmatch(name).group(1))))
    snapshot = workbook_snapshot(workbook, (edit.row for edit in edits))
    workbook.close()
    return tuple(edits), mapped, stale, unregistered, snapshot


def transform_source_map(data: Any, mapping: dict[str, str]) -> tuple[dict[str, Any], int]:
    if not isinstance(data, dict):
        raise MigrationError("product_source_map.json must contain an object")
    result = copy.deepcopy(data)
    updates = 0

    def visit(value: Any) -> None:
        nonlocal updates
        if isinstance(value, dict):
            for key, child in list(value.items()):
                if key == SHOP_ID and isinstance(child, str) and child in mapping:
                    replacement = mapping[child]
                    if replacement != child:
                        value[key] = replacement
                        updates += 1
                else:
                    visit(child)
        elif isinstance(value, list):
            for child in value:
                visit(child)

    visit(result)
    return result, updates


def inventory_tree(shop_dir: Path, names: Iterable[str]) -> dict[str, tuple[int, ...]]:
    """Record every entry, including zero-byte/cloud-only placeholders."""
    result: dict[str, tuple[int, ...]] = {}
    for name in names:
        root = shop_dir / name
        for current, directories, files in os.walk(root, topdown=True, followlinks=False):
            current_path = Path(current)
            for child_name in sorted([*directories, *files]):
                if child_name in {PRODUCT_LOCK_NAME, STATE_NAME} and current_path == root:
                    # Persistent flock files are coordination metadata. Opening
                    # them to prove idleness must not falsify asset invariants.
                    continue
                child = current_path / child_name
                if child.is_symlink():
                    raise MigrationError(f"symlink is unsafe in product tree: {child}")
                relative = child.relative_to(root).as_posix()
                info = child.lstat()
                kind = 2 if child.is_symlink() else (1 if child.is_dir() else 0)
                result[f"{name}/{relative}"] = (
                    kind, int(info.st_size), int(info.st_mtime_ns), int(info.st_mode & 0o7777),
                    int(info.st_dev), int(info.st_ino), int(info.st_blocks),
                )
    return result


def remap_inventory(inventory: dict[str, tuple[int, ...]], mapping: dict[str, str]) -> dict[str, tuple[int, ...]]:
    result = {}
    for logical, metadata in inventory.items():
        folder, relative = logical.split("/", 1)
        result[f"{mapping[folder]}/{relative}"] = metadata
    return result


def _validate_identity(product: Any, expected: ProductIdentity, context: str) -> None:
    if not isinstance(product, dict) or product != expected.as_dict():
        raise MigrationError(f"{context} has wrong cloud product identity")


def _rewrite_cloud_state(state: dict[str, Any], old: ProductIdentity, new: ProductIdentity) -> tuple[dict[str, Any], str, str]:
    if state.get("schema") != 1 or state.get("state") not in {
        "LOCAL_ONLY", "CLOUD_VERIFIED", "CLOUD_ONLY", "DIRTY_LOCAL", "ERROR",
        "OFFLOAD_SCHEDULED", "RESTORE_VERIFIED",
    }:
        raise MigrationError(f"unsupported/active cloud state for {old.key}: {state.get('state')!r}")
    _validate_identity(state.get("product"), old, str(old.key))
    revision = state.get("current_revision")
    manifest = state.get("current_manifest")
    if revision is None and manifest is None:
        rewritten = copy.deepcopy(state)
        rewritten["product"] = new.as_dict()
        return rewritten, "", ""
    if not isinstance(revision, str) or not revision or not isinstance(manifest, dict):
        raise MigrationError(f"incomplete current cloud revision metadata for {old.key}")
    _validate_identity(manifest.get("product"), old, f"{old.key} current_manifest")
    if manifest.get("revision") != revision:
        raise MigrationError(f"cloud revision mismatch for {old.key}")
    old_digest = hashlib.sha256(canonical_json_bytes(manifest)).hexdigest()
    if state.get("current_manifest_sha256") != old_digest:
        raise MigrationError(f"local current manifest hash mismatch for {old.key}")
    rewritten = copy.deepcopy(state)
    rewritten["product"] = new.as_dict()
    new_manifest = copy.deepcopy(manifest)
    new_manifest["product"] = new.as_dict()
    new_digest = hashlib.sha256(canonical_json_bytes(new_manifest)).hexdigest()
    rewritten["current_manifest"] = new_manifest
    rewritten["current_manifest_sha256"] = new_digest
    if rewritten.get("cloud_manifest_sha256") == old_digest:
        rewritten["cloud_manifest_sha256"] = new_digest
    return rewritten, revision, new_digest


def cloud_plan(shop_dir: Path, mappings: Iterable[FolderMapping]) -> tuple[CloudMove, ...]:
    result = []
    for item in mappings:
        state_path = shop_dir / item.old_name / STATE_NAME
        if not state_path.is_file():
            continue
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise MigrationError(f"invalid local cloud state {state_path}: {exc}") from exc
        if not isinstance(state, dict):
            raise MigrationError(f"invalid local cloud state object: {state_path}")
        old = ProductIdentity("shops", item.old_name, SHOP_ID)
        new = ProductIdentity("shops", item.new_name, SHOP_ID)
        rewritten, revision, digest = _rewrite_cloud_state(state, old, new)
        if state.get("state") == "CLOUD_VERIFIED" and revision:
            try:
                verify_manifest_directory(state_path.parent, state["current_manifest"])
            except Exception as exc:
                raise MigrationError(
                    f"local CLOUD_VERIFIED assets do not match manifest for {old.key}: {exc}"
                ) from exc
        if item.old_name != item.new_name:
            result.append(CloudMove(item.old_name, item.new_name, old, new, state, rewritten, revision, digest))
    return tuple(result)


def build_plan(
    *,
    repo_root: Path,
    shop_dir: Path | None = None,
    workbook_path: Path | None = None,
    source_map_path: Path | None = None,
    expected_folders: int | None = EXPECTED_FOLDERS,
    expected_bindings: int | None = EXPECTED_BINDINGS,
    expected_stale: int | None = EXPECTED_STALE,
    expected_stale_names: tuple[str, ...] | None = EXPECTED_STALE_NAMES,
    expected_source_map_updates: int | None = EXPECTED_SOURCE_MAP_UPDATES,
    expected_cloud_moves: tuple[tuple[str, str], ...] | None = EXPECTED_CLOUD_MOVES,
    token: str | None = None,
) -> MigrationPlan:
    repo_root = repo_root.resolve()
    shop_dir = (shop_dir or repo_root / "shops" / SHOP_ID).resolve()
    workbook_path = (workbook_path or shop_dir / "Etsy_SEO_Generator.xlsx").resolve()
    source_map_path = (source_map_path or repo_root / "product_source_map.json").resolve()
    if shop_dir.name != SHOP_ID or shop_dir.parent != repo_root / "shops":
        raise MigrationError(f"noncanonical shop/path relationship: {shop_dir}")
    leftovers = [p.name for p in shop_dir.iterdir() if p.name.startswith(".daisy-folder-renumber-")]
    if leftovers:
        raise MigrationError("unfinished temporary folders exist: " + ", ".join(sorted(leftovers)))
    names = discover_folders(shop_dir, expected_folders)
    mappings = make_mapping(names, token)
    mapping = {item.old_name: item.new_name for item in mappings}
    edits, bound, stale, unregistered, snapshot = workbook_plan(
        workbook_path, mapping, expected_bindings=expected_bindings
    )
    if expected_stale is not None and stale != expected_stale:
        raise MigrationError(f"expected {expected_stale} stale workbook references, found {stale}")
    stale_names = tuple(sorted(
        (item.old_value for item in edits if item.old_value not in mapping),
        key=lambda name: int(PRODUCT_RE.fullmatch(name).group(1)),
    ))
    if expected_stale_names is not None and stale_names != expected_stale_names:
        raise MigrationError(f"expected stale references {expected_stale_names}, found {stale_names}")
    if expected_folders == EXPECTED_FOLDERS and unregistered != ("product-428",):
        raise MigrationError(f"expected only product-428 unregistered, found {unregistered}")
    updated_map = None
    map_updates = 0
    hashes = {str(workbook_path): sha256_file(workbook_path)}
    if source_map_path.exists():
        try:
            raw_map = json.loads(source_map_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise MigrationError(f"invalid source map {source_map_path}: {exc}") from exc
        updated_map, map_updates = transform_source_map(raw_map, mapping)
        if expected_source_map_updates is not None and map_updates != expected_source_map_updates:
            raise MigrationError(
                f"expected {expected_source_map_updates} source-map updates, found {map_updates}"
            )
        hashes[str(source_map_path)] = sha256_file(source_map_path)
    elif expected_source_map_updates is not None:
        raise MigrationError(f"required source map is missing: {source_map_path}")
    moves = cloud_plan(shop_dir, mappings)
    move_pairs = tuple((item.old_name, item.new_name) for item in moves)
    if expected_cloud_moves is not None and move_pairs != expected_cloud_moves:
        raise MigrationError(f"expected cloud moves {expected_cloud_moves}, found {move_pairs}")
    for item in moves:
        state_path = shop_dir / item.old_name / STATE_NAME
        hashes[str(state_path)] = sha256_file(state_path)
    return MigrationPlan(
        repo_root, shop_dir, workbook_path, source_map_path, mappings, edits,
        bound, stale, unregistered, updated_map, map_updates, moves, hashes,
        inventory_tree(shop_dir, names), snapshot,
    )


class HeldLocks:
    def __init__(self, plan: MigrationPlan) -> None:
        self.plan = plan
        self.handles: list[Any] = []
        self.migration_path = plan.shop_dir / MIGRATION_LOCK_NAME
        self.migration_identity: tuple[int, int] | None = None

    def __enter__(self) -> "HeldLocks":
        try:
            migration = self.migration_path.open("x+")
            info = os.fstat(migration.fileno())
            self.migration_identity = (int(info.st_dev), int(info.st_ino))
            self.handles.append(migration)
            fcntl.flock(migration.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            catalog_path = self.plan.shop_dir / ".catalog_workbook.lock"
            if catalog_path.is_symlink():
                raise MigrationError(f"catalog workbook lock is an unsafe symlink: {catalog_path}")
            if not catalog_path.is_file():
                raise MigrationError(f"required catalog workbook lock is missing: {catalog_path}")
            catalog = catalog_path.open("a+")
            try:
                fcntl.flock(catalog.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError as exc:
                catalog.close()
                raise MigrationError(f"catalog workbook lock is held: {catalog_path}") from exc
            self.handles.append(catalog)
            for item in self.plan.mappings:
                lock_path = self.plan.shop_dir / item.old_name / PRODUCT_LOCK_NAME
                if lock_path.is_symlink():
                    raise MigrationError(f"cloud product lock is an unsafe symlink: {lock_path}")
                if lock_path.exists() and not lock_path.is_file():
                    raise MigrationError(f"cloud product lock is not a regular file: {lock_path}")
                # ProductLock itself uses open("a+").  Acquire/create every
                # per-product lock here so a writer cannot race into a folder
                # that did not already have its persistent lock marker.
                handle = lock_path.open("a+")
                try:
                    fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                except BlockingIOError as exc:
                    handle.close()
                    raise MigrationError(f"cloud product lock is held: {lock_path}") from exc
                self.handles.append(handle)
        except Exception:
            self.__exit__(None, None, None)
            raise
        return self

    def __exit__(self, *_: Any) -> None:
        for handle in reversed(self.handles):
            try:
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
                handle.close()
            except OSError:
                pass
        self.handles.clear()
        if self.migration_identity is not None:
            try:
                info = self.migration_path.lstat()
                if (int(info.st_dev), int(info.st_ino)) == self.migration_identity:
                    self.migration_path.unlink()
            except FileNotFoundError:
                pass
            self.migration_identity = None


def _is_catalog_writer_command(command: str) -> bool:
    if re.search(r"(?:^|[\s/])rclone(?:\s|$)", command):
        return True
    return any(
        re.search(rf"(?:^|[\s/]){re.escape(name)}(?:\s|$)", command)
        for name in WRITER_PROCESS_NAMES
    )


def assert_runtime_offline(plan: MigrationPlan) -> None:
    try:
        with socket.create_connection(("127.0.0.1", 8090), timeout=0.2):
            raise MigrationError("dashboard port 8090 is listening; apply requires it offline")
    except ConnectionRefusedError:
        pass
    except OSError as exc:
        raise MigrationError(f"cannot prove dashboard port 8090 offline: {exc}") from exc
    result = subprocess.run(["ps", "-axo", "pid=,command="], text=True, capture_output=True, check=True)
    own_pid = os.getpid()
    offenders = []
    for line in result.stdout.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        pid_text, _, command = stripped.partition(" ")
        try:
            pid = int(pid_text)
        except ValueError:
            continue
        if pid == own_pid:
            continue
        if _is_catalog_writer_command(command):
            offenders.append(f"{pid}:{command}")
    if offenders:
        raise MigrationError("active dashboard/rclone/Etsy writers: " + "; ".join(offenders[:8]))
    database = plan.repo_root / ".etsy-dashboard-diagnostics" / "etsy_jobs.sqlite"
    if database.is_file():
        try:
            connection = sqlite3.connect(f"file:{database}?mode=ro", uri=True)
            count = connection.execute("SELECT COUNT(*) FROM jobs WHERE status IN ('queued','running')").fetchone()[0]
            connection.close()
        except sqlite3.Error as exc:
            raise MigrationError(f"cannot prove durable job queue idle: {exc}") from exc
        if count:
            raise MigrationError(f"durable job queue has {count} queued/running job(s)")


def _remote_paths(identity: ProductIdentity, revision: str) -> tuple[str, str, str]:
    revision_root = f"{identity.remote_prefix}/revisions/{revision}"
    return revision_root, f"{revision_root}/manifest.json", f"{identity.remote_prefix}/current.json"


def _read_optional_remote_file(remote: RemoteStore, path: str) -> bytes | None:
    try:
        return remote.read_bytes(path)
    except RemoteStoreError as exc:
        message = str(exc).lower()
        if any(marker in message for marker in ("is missing", "not found", "directory not found")):
            return None
        raise MigrationError(f"cannot safely probe remote file {path}: {exc}") from exc


def prepare_cloud_move(remote: RemoteStore, move: CloudMove) -> dict[str, Any]:
    if not move.revision:
        return {"old": move.old_name, "new": move.new_name, "remote": "not_required"}
    old_root, old_manifest_path, old_current_path = _remote_paths(move.old_identity, move.revision)
    new_root, new_manifest_path, new_current_path = _remote_paths(move.new_identity, move.revision)
    try:
        old_pointer_data = remote.read_bytes(old_current_path)
        old_pointer = json.loads(old_pointer_data.decode("utf-8"))
        old_manifest_data = remote.read_bytes(old_manifest_path)
    except Exception as exc:
        raise MigrationError(f"cannot validate old remote cloud prefix {move.old_identity.key}: {exc}") from exc
    old_digest = hashlib.sha256(old_manifest_data).hexdigest()
    if (
        canonical_json_bytes(old_pointer) != old_pointer_data
        or old_pointer.get("schema") != 1
        or old_pointer.get("type") != "etsy-cloud-current-pointer"
        or old_pointer.get("product") != move.old_identity.key
        or old_pointer.get("revision") != move.revision
        or old_pointer.get("revision_path") != old_root
        or old_pointer.get("manifest_sha256") != old_digest
        or old_manifest_data != canonical_json_bytes(move.state["current_manifest"])
    ):
        raise MigrationError(f"old remote current/manifest mismatch for {move.old_identity.key}")
    new_manifest = move.new_state["current_manifest"]
    new_manifest_data = canonical_json_bytes(new_manifest)
    new_pointer = copy.deepcopy(old_pointer)
    new_pointer.update({
        "product": move.new_identity.key,
        "revision_path": new_root,
        "manifest_sha256": move.manifest_sha256,
    })
    new_pointer_data = canonical_json_bytes(new_pointer)
    with tempfile.TemporaryDirectory(prefix="daisy-cloud-copy-") as temporary_name:
        stage = Path(temporary_name) / "revision"
        remote.download_directory(old_root, stage)
        manifest_file = stage / "manifest.json"
        if manifest_file.read_bytes() != old_manifest_data:
            raise MigrationError(f"downloaded old remote manifest differs for {move.old_identity.key}")
        verify_manifest_directory(stage, move.state["current_manifest"])
        manifest_file.write_bytes(new_manifest_data)
        verify_manifest_directory(stage, new_manifest)
        existing_pointer = _read_optional_remote_file(remote, new_current_path)
        root_exists = remote.path_exists(new_root)
        if existing_pointer is not None and root_exists:
            if existing_pointer != new_pointer_data or remote.read_bytes(new_manifest_path) != new_manifest_data:
                raise MigrationError(f"conflicting new remote prefix exists for {move.new_identity.key}")
            remote.verify_directory(stage, new_root)
            status = "reused_verified"
        elif existing_pointer is None and root_exists:
            # A prior attempt may have uploaded and verified the immutable
            # revision, then stopped before publishing current.json.  Complete
            # that exact, fully verified prefix without deleting or replacing
            # any remote asset.
            if remote.read_bytes(new_manifest_path) != new_manifest_data:
                raise MigrationError(f"conflicting partial remote prefix exists for {move.new_identity.key}")
            remote.verify_directory(stage, new_root)
            remote.write_bytes(new_current_path, new_pointer_data, overwrite=False)
            if remote.read_bytes(new_current_path) != new_pointer_data:
                raise MigrationError(f"completed remote pointer read-back differs for {move.new_identity.key}")
            status = "completed_partial_verified"
        elif existing_pointer is not None:
            raise MigrationError(f"partial new remote pointer exists for {move.new_identity.key}")
        else:
            remote.upload_directory(stage, new_root)
            remote.verify_directory(stage, new_root)
            if remote.read_bytes(new_manifest_path) != new_manifest_data:
                raise MigrationError(f"new remote manifest read-back differs for {move.new_identity.key}")
            remote.write_bytes(new_current_path, new_pointer_data, overwrite=False)
            if remote.read_bytes(new_current_path) != new_pointer_data:
                raise MigrationError(f"new remote pointer read-back differs for {move.new_identity.key}")
            status = "copied_verified"
    return {"old": move.old_name, "new": move.new_name, "revision": move.revision, "remote": status}


def _manifest(plan: MigrationPlan, backup_dir: Path) -> dict[str, Any]:
    return {
        "schema": 1,
        "migration": "daisy-folder-renumber",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "repo_root": str(plan.repo_root),
        "shop_dir": str(plan.shop_dir),
        "workbook_path": str(plan.workbook_path),
        "source_map_path": str(plan.source_map_path),
        "backup_dir": str(backup_dir),
        "mapping": [item.__dict__ for item in plan.mappings],
        "workbook_edits": [item.__dict__ for item in plan.workbook_edits],
        "source_hashes": plan.source_hashes,
        "inventory": {key: list(value) for key, value in plan.inventory.items()},
        "cloud_moves": [{
            "old": item.old_name, "new": item.new_name, "revision": item.revision,
            "old_prefix": item.old_identity.remote_prefix,
            "new_prefix": item.new_identity.remote_prefix,
        } for item in plan.cloud_moves],
    }


def create_backup(plan: MigrationPlan, backup_root: Path) -> tuple[Path, dict[str, Any]]:
    backup_dir = backup_root / f"{SHOP_ID}-folder-renumber-{utc_stamp()}-{uuid.uuid4().hex[:8]}"
    backup_dir.mkdir(parents=True)
    _atomic_copy(plan.workbook_path, backup_dir / "Etsy_SEO_Generator.xlsx")
    if plan.source_map_path.exists():
        _atomic_copy(plan.source_map_path, backup_dir / "product_source_map.json")
    cloud_backup = backup_dir / "cloud-state"
    for move in plan.cloud_moves:
        source = plan.shop_dir / move.old_name / STATE_NAME
        _atomic_copy(source, cloud_backup / move.old_name / STATE_NAME)
    backup_pairs = [(plan.workbook_path, backup_dir / "Etsy_SEO_Generator.xlsx")]
    if plan.source_map_path.exists():
        backup_pairs.append((plan.source_map_path, backup_dir / "product_source_map.json"))
    backup_pairs.extend(
        (plan.shop_dir / move.old_name / STATE_NAME,
         cloud_backup / move.old_name / STATE_NAME)
        for move in plan.cloud_moves
    )
    backup_hashes = {}
    for source, destination in backup_pairs:
        source_hash = sha256_file(source)
        destination_hash = sha256_file(destination)
        if source_hash != destination_hash:
            raise MigrationError(f"backup verification failed for {source}")
        backup_hashes[str(destination.relative_to(backup_dir))] = destination_hash
    _atomic_json(backup_dir / MANIFEST_NAME, _manifest(plan, backup_dir))
    manifest = json.loads((backup_dir / MANIFEST_NAME).read_text(encoding="utf-8"))
    manifest["backup_hashes"] = backup_hashes
    _atomic_json(backup_dir / MANIFEST_NAME, manifest)
    journal = {"schema": 1, "status": "prepared", "backup_hashes": backup_hashes,
               "remote_prepared": [], "phase1": [], "phase2": [], "files": []}
    _atomic_json(backup_dir / JOURNAL_NAME, journal)
    return backup_dir, journal


def _journal(path: Path, value: dict[str, Any], **updates: Any) -> None:
    value.update(updates)
    _atomic_json(path, value)


def _rename_forward(plan: MigrationPlan, journal_path: Path, journal: dict[str, Any]) -> None:
    changed = [item for item in plan.mappings if item.old_name != item.new_name]
    for item in changed:
        source = plan.shop_dir / item.old_name
        temporary = plan.shop_dir / item.temporary_name
        if temporary.exists() or not source.is_dir():
            raise MigrationError(f"unsafe phase-1 rename state for {item.old_name}")
        source.rename(temporary)
        journal["phase1"].append(item.old_name)
        _journal(journal_path, journal, status="renaming_phase1")
    for item in changed:
        temporary = plan.shop_dir / item.temporary_name
        target = plan.shop_dir / item.new_name
        if target.exists() or not temporary.is_dir():
            raise MigrationError(f"unsafe phase-2 rename state for {item.old_name}")
        temporary.rename(target)
        journal["phase2"].append(item.old_name)
        _journal(journal_path, journal, status="renaming_phase2")


def rollback_folder_renames(plan: MigrationPlan, journal_path: Path, journal: dict[str, Any]) -> None:
    changed = [item for item in plan.mappings if item.old_name != item.new_name]
    phase1 = set(journal.get("phase1", []))
    phase2 = set(journal.get("phase2", []))
    evacuated: set[str] = set()
    for item in reversed(changed):
        if item.old_name in phase2:
            target = plan.shop_dir / item.new_name
            rollback = plan.shop_dir / item.rollback_name
            if not target.is_dir() or rollback.exists():
                raise MigrationError(f"unsafe rollback evacuation for {item.old_name}")
            target.rename(rollback)
            evacuated.add(item.old_name)
    for item in reversed(changed):
        if item.old_name not in phase1:
            continue
        source = plan.shop_dir / (item.rollback_name if item.old_name in evacuated else item.temporary_name)
        target = plan.shop_dir / item.old_name
        if target.exists() or not source.is_dir():
            raise MigrationError(f"unsafe rollback restore for {item.old_name}")
        source.rename(target)
    _journal(journal_path, journal, status="folders_rolled_back")


def _update_local_files(plan: MigrationPlan) -> None:
    workbook = load_workbook(plan.workbook_path, data_only=False)
    sheet = workbook["Listings"]
    for edit in plan.workbook_edits:
        if sheet.cell(edit.row, 2).value != edit.old_value:
            workbook.close()
            raise MigrationError(f"workbook changed at Listings!B{edit.row}")
        sheet.cell(edit.row, 2).value = edit.new_value
    _atomic_save_workbook(workbook, plan.workbook_path)
    if plan.updated_source_map is not None:
        _atomic_pretty_json(plan.source_map_path, plan.updated_source_map)
    for move in plan.cloud_moves:
        _atomic_json(plan.shop_dir / move.new_name / STATE_NAME, move.new_state)


def _restore_files(plan: MigrationPlan, backup_dir: Path) -> None:
    _atomic_copy(backup_dir / "Etsy_SEO_Generator.xlsx", plan.workbook_path)
    map_backup = backup_dir / "product_source_map.json"
    if map_backup.exists():
        _atomic_copy(map_backup, plan.source_map_path)
    for move in plan.cloud_moves:
        _atomic_copy(
            backup_dir / "cloud-state" / move.old_name / STATE_NAME,
            plan.shop_dir / move.old_name / STATE_NAME,
        )


def revalidate(plan: MigrationPlan) -> None:
    for path_text, expected in plan.source_hashes.items():
        if sha256_file(Path(path_text)) != expected:
            raise MigrationError(f"source changed after planning: {path_text}")
    names = tuple(item.old_name for item in plan.mappings)
    if discover_folders(plan.shop_dir, len(names)) != names:
        raise MigrationError("physical folder set changed after planning")
    if inventory_tree(plan.shop_dir, names) != plan.inventory:
        raise MigrationError("folder contents/metadata changed after planning")


def verify_cloud_move_postflight(remote: RemoteStore, move: CloudMove) -> None:
    new_root, new_manifest_path, new_current_path = _remote_paths(
        move.new_identity, move.revision
    )
    _old_root, _old_manifest_path, old_current_path = _remote_paths(
        move.old_identity, move.revision
    )
    try:
        old_pointer_data = remote.read_bytes(old_current_path)
        old_pointer = json.loads(old_pointer_data.decode("utf-8"))
        expected_pointer = copy.deepcopy(old_pointer)
        expected_pointer.update({
            "product": move.new_identity.key,
            "revision_path": new_root,
            "manifest_sha256": move.manifest_sha256,
        })
        expected_pointer_data = canonical_json_bytes(expected_pointer)
        if canonical_json_bytes(old_pointer) != old_pointer_data:
            raise MigrationError(
                f"old remote pointer is no longer canonical for {move.old_identity.key}"
            )
        if remote.read_bytes(new_current_path) != expected_pointer_data:
            raise MigrationError(
                f"new remote pointer postflight differs for {move.new_identity.key}"
            )
        expected_manifest_data = canonical_json_bytes(move.new_state["current_manifest"])
        if remote.read_bytes(new_manifest_path) != expected_manifest_data:
            raise MigrationError(
                f"new remote manifest postflight differs for {move.new_identity.key}"
            )
        with tempfile.TemporaryDirectory(prefix="daisy-cloud-postflight-") as temporary_name:
            stage = Path(temporary_name) / "revision"
            remote.download_directory(new_root, stage)
            verify_manifest_directory(stage, move.new_state["current_manifest"])
            remote.verify_directory(stage, new_root)
    except MigrationError:
        raise
    except Exception as exc:
        raise MigrationError(
            f"cloud postflight failed for {move.new_identity.key}: {exc}"
        ) from exc


def postflight(plan: MigrationPlan, remote: RemoteStore) -> dict[str, int]:
    names = discover_folders(plan.shop_dir, len(plan.mappings))
    expected_names = tuple(item.new_name for item in plan.mappings)
    if names != expected_names:
        raise MigrationError("postflight folder set is not contiguous")
    _edits, bound, stale, unregistered, _snapshot = workbook_plan(
        plan.workbook_path,
        {item.new_name: item.new_name for item in plan.mappings},
        expected_bindings=plan.mapped_binding_count,
    )
    workbook = load_workbook(plan.workbook_path, data_only=False)
    snapshot = workbook_snapshot(workbook, (item.row for item in plan.workbook_edits))
    workbook.close()
    mapping = {item.old_name: item.new_name for item in plan.mappings}
    expected_unregistered = tuple(mapping[name] for name in plan.unregistered_old)
    if stale != 0 or unregistered != expected_unregistered or snapshot != plan.workbook_snapshot:
        raise MigrationError("postflight workbook bindings or non-column-B cells changed unexpectedly")
    if plan.updated_source_map is not None:
        current = json.loads(plan.source_map_path.read_text(encoding="utf-8"))
        if current != plan.updated_source_map:
            raise MigrationError("postflight source map differs from reviewed transformation")
    actual = inventory_tree(plan.shop_dir, expected_names)
    if actual != remap_inventory(plan.inventory, {item.old_name: item.new_name for item in plan.mappings}):
        raise MigrationError("postflight folder entry metadata/inodes changed")
    for move in plan.cloud_moves:
        state_path = plan.shop_dir / move.new_name / STATE_NAME
        if not state_path.is_file() or state_path.read_bytes() != canonical_json_bytes(move.new_state):
            raise MigrationError(
                f"local cloud state postflight differs for {move.new_identity.key}"
            )
        verify_cloud_move_postflight(remote, move)
    return {
        "physical_folders": len(names), "workbook_bindings": bound,
        "stale_bindings": stale, "unregistered_physical": len(unregistered),
        "source_map_updates": plan.source_map_updates, "cloud_moves": len(plan.cloud_moves),
    }


def apply_plan(
    plan: MigrationPlan,
    backup_root: Path,
    *,
    remote: RemoteStore,
    runtime_checker: Callable[[MigrationPlan], None] = assert_runtime_offline,
) -> tuple[Path, dict[str, int]]:
    runtime_checker(plan)
    with HeldLocks(plan):
        runtime_checker(plan)
        revalidate(plan)
        backup_dir, journal = create_backup(plan, backup_root)
        journal_path = backup_dir / JOURNAL_NAME
        try:
            prepared = []
            for move in plan.cloud_moves:
                prepared.append(prepare_cloud_move(remote, move))
                _journal(journal_path, journal, status="preparing_remote", remote_prepared=prepared)
            _rename_forward(plan, journal_path, journal)
            _update_local_files(plan)
            _journal(journal_path, journal, status="files_updated")
            counts = postflight(plan, remote)
            _journal(journal_path, journal, status="completed", postflight=counts)
            return backup_dir, counts
        except Exception as exc:
            try:
                rollback_folder_renames(plan, journal_path, journal)
                _restore_files(plan, backup_dir)
                if inventory_tree(plan.shop_dir, (item.old_name for item in plan.mappings)) != plan.inventory:
                    raise MigrationError("automatic rollback did not restore exact folder inventory")
                _journal(journal_path, journal, status="rolled_back", failure=str(exc))
            except Exception as rollback_exc:
                _journal(journal_path, journal, status="unfinished", failure=str(exc), rollback_failure=str(rollback_exc))
                raise MigrationError(
                    f"migration failed and rollback is incomplete; journal={journal_path}; "
                    f"error={exc}; rollback={rollback_exc}"
                ) from exc
            raise MigrationError(f"migration failed; automatic rollback completed; journal={journal_path}; {exc}") from exc


def summary(plan: MigrationPlan) -> str:
    changed = sum(item.old_name != item.new_name for item in plan.mappings)
    remote = ",".join(f"{item.old_name}->{item.new_name}" for item in plan.cloud_moves) or "none"
    return (
        f"shop={SHOP_ID} physical={len(plan.mappings)} changed={changed} "
        f"workbook_mapped={plan.mapped_binding_count} stale_to_clear={plan.stale_binding_count} "
        f"unregistered={','.join(plan.unregistered_old) or 'none'} map_updates={plan.source_map_updates} "
        f"cloud_moves={remote}"
    )


def configured_remote(repo: Path) -> RcloneRemote:
    """Build the remote from the repository's resolved routing config."""
    config = load_config(repo)
    return RcloneRemote(
        repo,
        remote=config.remote,
        parent_id=config.parent_id,
        rclone_bin=config.rclone_bin,
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="perform the migration; default is dry-run")
    parser.add_argument("--repo-root", type=Path, default=CANONICAL_REPO)
    parser.add_argument("--backup-root", type=Path)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    repo = args.repo_root.resolve()
    if repo != CANONICAL_REPO.resolve():
        print(f"ERROR: refusing noncanonical repository {repo}", file=sys.stderr)
        return 2
    backup_root = (args.backup_root or repo / "output" / "backup" / "daisy-folder-renumber").resolve()
    try:
        plan = build_plan(repo_root=repo)
        print(f"PREFLIGHT {'APPLY' if args.apply else 'DRY-RUN'} {summary(plan)}")
        if not args.apply:
            print("DRY-RUN: no local or remote writes performed")
            return 0
        backup_dir, counts = apply_plan(plan, backup_root, remote=configured_remote(repo))
    except MigrationError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    print(f"BACKUP {backup_dir}")
    print("POSTFLIGHT " + " ".join(f"{key}={value}" for key, value in sorted(counts.items())))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
