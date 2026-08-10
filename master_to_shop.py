#!/usr/bin/env python3
"""Safely derive selected master products into one Etsy shop.

Dry run is the default.  ``--apply`` is deliberately required to mutate the
shop, workbook, or product source map.  This tool never posts to Etsy.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from PIL import Image

from shop_asset_workflow import IMG_EXTS, copy_image_with_watermark, get_watermark_text


ROOT = Path(__file__).resolve().parent
MASTER_ROOT = ROOT / "master_products"
SHOPS_ROOT = ROOT / "shops"
ACTIVE_SHOP = ROOT / "active_shop.txt"
CONFIG = ROOT / "shops_config.json"
MAP = ROOT / "product_source_map.json"
OUTPUT = ROOT / "output" / "master_to_shop"
PRODUCT_RE = re.compile(r"^product-\d+$")
DOWNLOAD_EXTS = {".pdf", ".zip"}


class TransferError(RuntimeError):
    pass


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def atomic_json(path: Path, value: object) -> None:
    tmp = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    tmp.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    os.replace(tmp, path)


def usable(path: Path, suffixes: set[str]) -> bool:
    if path.is_symlink() or not path.is_file() or path.suffix.lower() not in suffixes:
        return False
    stat = path.stat()
    return stat.st_size > 0 and getattr(stat, "st_blocks", 1) > 0


def direct_assets(directory: Path, suffixes: set[str]) -> list[Path]:
    if directory.is_symlink() or not directory.is_dir():
        return []
    return sorted((p for p in directory.iterdir() if usable(p, suffixes)), key=lambda p: p.name.lower())


def parse_products(raw: list[str]) -> list[str]:
    requested: list[str] = []
    for part in raw:
        for product in part.split(","):
            product = product.strip()
            if product and product not in requested:
                requested.append(product)
    if not requested:
        raise TransferError("--products requires at least one product-NN")
    invalid = [p for p in requested if not PRODUCT_RE.fullmatch(p)]
    if invalid:
        raise TransferError(f"invalid product names: {', '.join(invalid)}")
    return requested


def read_config(shop: str) -> dict:
    try:
        data = json.loads(CONFIG.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise TransferError(f"cannot read shops config: {exc}") from exc
    value = data.get(shop)
    if not isinstance(value, dict):
        raise TransferError(f"shop is not configured: {shop}")
    return value


def preflight(shop: str, products: list[str]) -> tuple[Path, Path, dict, list[dict]]:
    active = ACTIVE_SHOP.read_text(encoding="utf-8").strip()
    if active != shop:
        raise TransferError(f"active shop mismatch: expected {shop!r}, got {active!r}")
    shop_dir = SHOPS_ROOT / shop
    workbook = shop_dir / "Etsy_SEO_Generator.xlsx"
    if not shop_dir.is_dir() or not workbook.is_file():
        raise TransferError("target shop directory or workbook is missing")
    # Confirm the workbook is readable before any mutation.
    try:
        wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
        if "Listings" not in wb.sheetnames:
            raise TransferError("target workbook has no Listings sheet")
        ws = wb["Listings"]
        existing_rows = {str(ws.cell(r, 2).value).strip() for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value}
        wb.close()
    except TransferError:
        raise
    except Exception as exc:
        raise TransferError(f"workbook preflight failed: {exc}") from exc
    plans: list[dict] = []
    for product in products:
        source = MASTER_ROOT / product
        if source.is_symlink() or not source.is_dir():
            raise TransferError(f"invalid source product: {product}")
        images = direct_assets(source / "images", IMG_EXTS)
        files = direct_assets(source / "files", DOWNLOAD_EXTS)
        all_image_names = [p.name for p in (source / "images").iterdir()] if (source / "images").is_dir() else []
        all_file_names = [p.name for p in (source / "files").iterdir()] if (source / "files").is_dir() else []
        if len(images) != 10 or len(all_image_names) != 10:
            raise TransferError(f"{product} must have exactly 10 usable direct images")
        if not files or len(files) != len(all_file_names):
            raise TransferError(f"{product} must have only usable direct PDF/ZIP files")
        plans.append({"source": source, "product": product, "images": images, "files": files})
    return shop_dir, workbook, read_config(shop), plans


def allocate(shop_dir: Path, count: int) -> list[str]:
    values = [int(p.name.split("-", 1)[1]) for p in shop_dir.glob("product-*") if p.is_dir() and p.name.split("-", 1)[1].isdigit()]
    next_value = max(values, default=0) + 1
    folders = [f"product-{n}" for n in range(next_value, next_value + count)]
    existing = [f for f in folders if (shop_dir / f).exists()]
    if existing:
        raise TransferError(f"destination conflict: {', '.join(existing)}")
    return folders


def plan_report(shop: str, plans: list[dict], destinations: list[str], watermark: str) -> dict:
    return {"shop": shop, "watermark": watermark, "operations": [
        {"source": f"master_products/{p['product']}", "destination": f"shops/{shop}/{d}", "images": len(p["images"]), "files": len(p["files"])}
        for p, d in zip(plans, destinations)
    ]}


def copy_stage(plan: dict, stage: Path, watermark: str) -> dict:
    images_dir, files_dir = stage / "images", stage / "files"
    images_dir.mkdir(parents=True)
    files_dir.mkdir(parents=True)
    records = {"source": f"master_products/{plan['product']}", "images": [], "files": []}
    for source in plan["images"]:
        destination = images_dir / source.name
        copy_image_with_watermark(source, destination, watermark)
        with Image.open(destination) as image:
            image.verify()
        if sha256(source) == sha256(destination):
            raise TransferError(f"watermark output did not change image: {source.name}")
        records["images"].append({"name": source.name, "source_sha256": sha256(source), "destination_sha256": sha256(destination)})
    for source in plan["files"]:
        destination = files_dir / source.name
        shutil.copy2(source, destination)
        source_hash, destination_hash = sha256(source), sha256(destination)
        if source_hash != destination_hash:
            raise TransferError(f"downloadable hash mismatch: {source.name}")
        records["files"].append({"name": source.name, "sha256": source_hash})
    if len(direct_assets(images_dir, IMG_EXTS)) != 10 or len(direct_assets(files_dir, DOWNLOAD_EXTS)) != len(plan["files"]):
        raise TransferError(f"staged asset validation failed: {plan['product']}")
    return records


def append_rows(workbook: Path, assignments: list[tuple[dict, str]]) -> None:
    wb = openpyxl.load_workbook(workbook)
    try:
        ws = wb["Listings"]
        existing = {str(ws.cell(r, 2).value).strip() for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value}
        for plan, destination in assignments:
            if destination in existing:
                raise TransferError(f"workbook conflict while applying: {destination}")
            row = next((r for r in range(4, ws.max_row + 2) if not str(ws.cell(r, 2).value or "").strip()), ws.max_row + 1)
            ws.cell(row, 1, row - 3)
            ws.cell(row, 2, destination)
            ws.cell(row, 3, f"Master {plan['product']}")
            ws.cell(row, 5, 4.99)
            ws.cell(row, 11, 999)
            ws.cell(row, 12, "I did")
            ws.cell(row, 13, "2020_2026")
            ws.cell(row, 14, "⏳ Chờ đăng - LOCAL")
            ws.cell(row, 15, "Digital Planner")
            existing.add(destination)
        wb.save(workbook)
    finally:
        wb.close()


def apply(shop: str, shop_dir: Path, workbook: Path, plans: list[dict], destinations: list[str], watermark: str) -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    receipt = OUTPUT / stamp
    receipt.mkdir(parents=True, exist_ok=False)
    workbook_backup = receipt / f"{workbook.stem}.before-{stamp}.xlsx"
    map_backup = receipt / "product_source_map.before.json"
    shutil.copy2(workbook, workbook_backup)
    shutil.copy2(MAP, map_backup)
    original_map = json.loads(MAP.read_text(encoding="utf-8"))
    if not isinstance(original_map, dict):
        raise TransferError("product source map must be a JSON object")
    assignments = list(zip(plans, destinations))
    journal = {"status": "started", "shop": shop, "watermark": watermark, "created": [], "records": [], "workbook_backup": str(workbook_backup.relative_to(ROOT)), "map_backup": str(map_backup.relative_to(ROOT))}
    atomic_json(receipt / "journal.json", journal)
    try:
        for plan, destination in assignments:
            target = shop_dir / destination
            stage = shop_dir / f".{destination}.master-transfer-{stamp}"
            if target.exists() or stage.exists():
                raise TransferError(f"destination appeared during apply: {destination}")
            stage.mkdir()
            record = copy_stage(plan, stage, watermark)
            os.replace(stage, target)
            journal["created"].append(str(target.relative_to(ROOT)))
            journal["records"].append({**record, "destination": str(target.relative_to(ROOT))})
            atomic_json(receipt / "journal.json", journal)
        append_rows(workbook, assignments)
        updated_map = dict(original_map)
        for plan, destination in assignments:
            key = f"master:{plan['product']}"
            if key in updated_map:
                raise TransferError(f"source mapping already exists: {key}")
            updated_map[key] = {shop: destination}
        atomic_json(MAP, updated_map)
        journal["status"] = "complete"
        atomic_json(receipt / "manifest.json", journal)
        atomic_json(receipt / "journal.json", journal)
        return receipt
    except Exception:
        shutil.copy2(workbook_backup, workbook)
        shutil.copy2(map_backup, MAP)
        for created in reversed(journal["created"]):
            path = ROOT / created
            if path.exists() and path.is_dir():
                shutil.rmtree(path)
        journal["status"] = "rolled_back"
        atomic_json(receipt / "journal.json", journal)
        raise


def resume(receipt: Path, shop: str, shop_dir: Path, workbook: Path, plans: list[dict], watermark: str) -> Path:
    """Complete an interrupted apply using only its on-disk journal."""
    journal_path = receipt / "journal.json"
    if not journal_path.is_file():
        raise TransferError("resume receipt has no journal")
    journal = json.loads(journal_path.read_text(encoding="utf-8"))
    if journal.get("status") != "started" or journal.get("shop") != shop:
        raise TransferError("receipt is not a resumable transaction for this shop")
    backup = ROOT / str(journal.get("workbook_backup", ""))
    map_backup = ROOT / str(journal.get("map_backup", ""))
    if not backup.is_file() or not map_backup.is_file():
        raise TransferError("resume receipt backups are missing")
    by_source = {r.get("source"): r for r in journal.get("records", [])}
    assignments: list[tuple[dict, str]] = []
    used = set()
    for plan in plans:
        source_key = f"master_products/{plan['product']}"
        record = by_source.get(source_key)
        if record:
            destination = Path(record["destination"]).name
            target = shop_dir / destination
            if not target.is_dir() or len(direct_assets(target / "images", IMG_EXTS)) != 10 or len(direct_assets(target / "files", DOWNLOAD_EXTS)) != len(plan["files"]):
                raise TransferError(f"interrupted destination is invalid: {destination}")
            used.add(destination)
        else:
            destination = ""
        assignments.append((plan, destination))
    next_number = max(int(x.split("-", 1)[1]) for x in used) + 1 if used else allocate(shop_dir, 1)[0].split("-", 1)[1]
    for index, (plan, destination) in enumerate(assignments):
        if destination:
            continue
        destination = f"product-{next_number}"
        next_number += 1
        if (shop_dir / destination).exists():
            raise TransferError(f"resume destination conflict: {destination}")
        stage = shop_dir / f".{destination}.resume-{receipt.name}"
        if stage.exists():
            raise TransferError(f"resume staging conflict: {stage.name}")
        stage.mkdir()
        record = copy_stage(plan, stage, watermark)
        os.replace(stage, shop_dir / destination)
        journal["created"].append(str((shop_dir / destination).relative_to(ROOT)))
        journal["records"].append({**record, "destination": str((shop_dir / destination).relative_to(ROOT))})
        assignments[index] = (plan, destination)
        atomic_json(journal_path, journal)
    # The original apply had not reached catalog finalization; refuse to resume if it has.
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Listings"]
    existing = {str(ws.cell(r, 2).value).strip() for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value}
    wb.close()
    if any(destination in existing for _, destination in assignments):
        raise TransferError("cannot resume: one or more workbook rows already exist")
    append_rows(workbook, assignments)
    mapping = json.loads(MAP.read_text(encoding="utf-8"))
    if not isinstance(mapping, dict):
        raise TransferError("product source map must be a JSON object")
    for plan, destination in assignments:
        key = f"master:{plan['product']}"
        if key in mapping:
            raise TransferError(f"cannot resume: mapping exists: {key}")
        mapping[key] = {shop: destination}
    atomic_json(MAP, mapping)
    journal["status"] = "complete"
    journal["resumed_at"] = datetime.now(timezone.utc).isoformat()
    atomic_json(receipt / "manifest.json", journal)
    atomic_json(journal_path, journal)
    return receipt


def main() -> int:
    parser = argparse.ArgumentParser(description="Copy selected canonical master products into an active shop")
    parser.add_argument("--shop", required=True)
    parser.add_argument("--products", required=True, nargs="+", help="product-NN names; comma-separated values accepted")
    parser.add_argument("--dry-run", action="store_true", help="print plan only (default)")
    parser.add_argument("--apply", action="store_true", help="perform the staged local transfer")
    parser.add_argument("--resume", type=Path, help="complete a journaled interrupted apply")
    args = parser.parse_args()
    if sum(bool(x) for x in (args.dry_run, args.apply, args.resume)) > 1:
        parser.error("choose one of --dry-run, --apply, or --resume")
    try:
        products = parse_products(args.products)
        shop_dir, workbook, config, plans = preflight(args.shop, products)
        destinations = allocate(shop_dir, len(plans))
        watermark = get_watermark_text(args.shop, config)
        report = plan_report(args.shop, plans, destinations, watermark)
        if args.resume:
            receipt = resume(args.resume.resolve(), args.shop, shop_dir, workbook, plans, watermark)
            print(json.dumps({"mode": "resume", **report, "receipt": str(receipt.relative_to(ROOT))}, indent=2, ensure_ascii=False))
            return 0
        if not args.apply:
            print(json.dumps({"mode": "dry-run", **report}, indent=2, ensure_ascii=False))
            return 0
        receipt = apply(args.shop, shop_dir, workbook, plans, destinations, watermark)
        print(json.dumps({"mode": "apply", **report, "receipt": str(receipt.relative_to(ROOT))}, indent=2, ensure_ascii=False))
        return 0
    except TransferError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
