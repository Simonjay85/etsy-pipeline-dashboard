#!/usr/bin/env python3
from __future__ import annotations

import asyncio
import openpyxl
import tempfile
from pathlib import Path
from typing import Any
from unittest import IsolatedAsyncioTestCase, TestCase
from unittest.mock import AsyncMock, patch
from fastapi import HTTPException

import dashboard_app


class _Request:
    def __init__(self, payload: dict):
        self.payload = payload

    async def json(self):
        return self.payload


def _response_payload(response):
    if hasattr(response, "body"):
        return dashboard_app.json.loads(response.body.decode())
    raise TypeError("Response has no JSON body")


def _create_listing_sheet(excel_path: Path, row_data: dict[int, dict[int, object]]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings"
    for row_num, values in row_data.items():
        for column in range(2, 19):
            ws.cell(row=row_num, column=column, value=values.get(column, ""))
    wb.save(excel_path)


def _cell_value(excel_path: Path, row: int, column: int):
    wb = openpyxl.load_workbook(excel_path)
    return wb["Listings"].cell(row=row, column=column).value


class _TmpShopFixture:
    def __init__(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.shop_id = "daisyflowdigital"
        self.shop_dir = self.tmp / "shops" / self.shop_id
        self.shop_dir.mkdir(parents=True, exist_ok=True)
        self.excel_path = self.shop_dir / "Etsy_SEO_Generator.xlsx"

    def cleanup(self):
        import shutil

        shutil.rmtree(self.tmp, ignore_errors=True)


class TestPostRouteAdmissionGuard(TestCase):
    def setUp(self) -> None:
        self.fixture = _TmpShopFixture()
        _create_listing_sheet(
            self.fixture.excel_path,
            {
                4: {
                    2: "product-259",
                    5: 9.99,
                    8: "Sample title",
                    14: "⏳ Chờ đăng",
                    16: "https://www.etsy.com/listing/4528985419",
                }
            },
        )
        (self.fixture.shop_dir / "product-259").mkdir()

    def tearDown(self) -> None:
        self.fixture.cleanup()
        dashboard_app._OPERATION_QUEUE_COMMANDS.clear()
        dashboard_app._OPERATION_QUEUE_DEDUPE.clear()
        dashboard_app._running_tasks.clear()
        dashboard_app._running_processes.clear()

    def test_post_route_rejects_mapped_product_before_enqueue(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.fixture.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.fixture.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.fixture.shop_id), \
                patch.object(dashboard_app, "_enqueue_operation") as enqueue_operation, \
                patch.object(dashboard_app, "save_to_excel") as save_to_excel:
            response = asyncio.run(dashboard_app.post_to_etsy(4))

        payload = _response_payload(response)
        self.assertEqual(409, response.status_code)
        self.assertEqual(False, payload["ok"])
        self.assertEqual(False, payload["created"])
        self.assertIn("Etsy URL/listing ID", payload["error"])
        enqueue_operation.assert_not_called()
        save_to_excel.assert_not_called()
        self.assertEqual(
            "⏳ Chờ đăng",
            _cell_value(self.fixture.excel_path, 4, 14),
        )


class TestRunPosterFailurePropagation(TestCase):
    def setUp(self) -> None:
        self.fixture = _TmpShopFixture()
        self.folder = self.fixture.shop_dir / "product-259"
        self.folder.mkdir()

    def tearDown(self) -> None:
        self.fixture.cleanup()

    def test_non_zero_poster_exit_raises_and_records_status(self):
        fake_proc = AsyncMock()
        fake_proc.stdout = None
        fake_proc.wait = AsyncMock(return_value=3)

        async def fake_subprocess(*_args, **_kwargs):
            return fake_proc

        with patch.object(dashboard_app.asyncio, "create_subprocess_exec", AsyncMock(side_effect=fake_subprocess)), \
                patch.object(dashboard_app, "_runtime_prefetch_import_check", AsyncMock(return_value=(True, "ok")), \
                ), \
                patch.object(dashboard_app, "_terminate_subprocess", new=AsyncMock(return_value=None)), \
                patch.object(dashboard_app, "save_to_excel") as save_to_excel:
            with self.assertRaises(RuntimeError) as caught:
                asyncio.run(
                    dashboard_app._run_poster(
                        4,
                        "product-259",
                        "daisyflowdigital",
                        lock_key="test-post-lock",
                    )
                )

        self.assertIn("❌ Lỗi exit 3", str(caught.exception))
        self.assertTrue(
            any(
                "❌ Lỗi exit 3" in str(call.args[1].get("status") or "")
                for call in save_to_excel.call_args_list
                if len(call.args) > 1 and isinstance(call.args[1], dict)
            )
        )


class TestRunQueuedOperationFailure(IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        dashboard_app._OPERATION_QUEUE_LOCK = asyncio.Lock()
        dashboard_app._OPERATION_QUEUE_COMMANDS.clear()
        dashboard_app._OPERATION_QUEUE_DEDUPE.clear()
        dashboard_app._running_tasks.clear()
        dashboard_app._running_processes.clear()

    async def asyncTearDown(self) -> None:
        dashboard_app._OPERATION_QUEUE_COMMANDS.clear()
        dashboard_app._OPERATION_QUEUE_DEDUPE.clear()
        dashboard_app._running_tasks.clear()
        dashboard_app._running_processes.clear()

    async def test_run_post_rechecks_eligibility_inside_queue_callback(self) -> None:
        fixture = _TmpShopFixture()
        (fixture.shop_dir / "product-259").mkdir(parents=True, exist_ok=True)

        with patch.object(dashboard_app, "SHOP_DIR", return_value=fixture.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=fixture.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", "daisyflowdigital"), \
                patch.object(
                    dashboard_app,
                    "get_product_by_row",
                    side_effect=[
                        {"folder": "product-259", "status": "⏳ Chờ đăng", "etsy_url": ""},
                        {"folder": "product-259", "status": "✅ Đã đăng", "etsy_url": ""},
                    ],
                ), \
                patch.object(dashboard_app, "save_to_excel") as save_to_excel, \
                patch.object(dashboard_app, "_run_poster", new=AsyncMock(return_value=None)):
            response = await dashboard_app.post_to_etsy(4)

            payload = _response_payload(response)
            self.assertEqual(202, response.status_code)
            self.assertTrue(payload["ok"])
            command_id = payload["command"]["command_id"]

            for _ in range(80):
                state = dashboard_app._OPERATION_QUEUE_COMMANDS.get(command_id)
                if state and state.get("status") in {"failed", "succeeded", "cancelled"}:
                    break
                await asyncio.sleep(0.01)
            state = dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]
            self.assertEqual("failed", state["status"])
            self.assertIn("trạng thái", state["error"])
            save_to_excel.assert_not_called()
        fixture.cleanup()

    async def test_run_post_rejects_folder_drift_after_admission(self) -> None:
        fixture = _TmpShopFixture()
        (fixture.shop_dir / "product-259").mkdir(parents=True, exist_ok=True)
        (fixture.shop_dir / "product-260").mkdir(parents=True, exist_ok=True)

        with patch.object(dashboard_app, "SHOP_DIR", return_value=fixture.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=fixture.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", "daisyflowdigital"), \
                patch.object(
                    dashboard_app,
                    "get_product_by_row",
                    side_effect=[
                        {"folder": "product-259", "status": "⏳ Chờ đăng", "etsy_url": ""},
                        {"folder": "product-260", "status": "⏳ Chờ đăng", "etsy_url": ""},
                    ],
                ), \
                patch.object(dashboard_app, "save_to_excel") as save_to_excel, \
                patch.object(dashboard_app, "_run_poster", new=AsyncMock(return_value=None)):
            response = await dashboard_app.post_to_etsy(4)

            payload = _response_payload(response)
            self.assertEqual(202, response.status_code)
            self.assertTrue(payload["ok"])
            command_id = payload["command"]["command_id"]

            for _ in range(80):
                state = dashboard_app._OPERATION_QUEUE_COMMANDS.get(command_id)
                if state and state.get("status") in {"failed", "succeeded", "cancelled"}:
                    break
                await asyncio.sleep(0.01)
            state = dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]
            self.assertEqual("failed", state["status"])
            self.assertIn("admission='product-259' -> current='product-260'", state["error"])
            save_to_excel.assert_not_called()
            dashboard_app._run_poster.assert_not_called()

        fixture.cleanup()

    async def test_run_post_fails_when_active_shop_changes_before_run(self) -> None:
        fixture = _TmpShopFixture()
        (fixture.shop_dir / "product-259").mkdir(parents=True, exist_ok=True)
        admitted_shop_id = "daisyflowdigital"
        switched_shop_id = "templystudio"

        _create_listing_sheet(
            fixture.excel_path,
            {
                4: {
                    2: "product-259",
                    5: 9.99,
                    8: "Sample title",
                    14: "⏳ Chờ đăng",
                }
            },
        )

        lock = dashboard_app._OPERATION_QUEUE_LOCK
        await lock.acquire()
        try:
            dashboard_app._active_shop_id = admitted_shop_id
            with patch.object(dashboard_app, "SHOP_DIR", return_value=fixture.shop_dir), \
                    patch.object(dashboard_app, "EXCEL_FILE", return_value=fixture.excel_path), \
                    patch.object(
                        dashboard_app,
                        "get_product_by_row",
                        return_value={"folder": "product-259", "status": "⏳ Chờ đăng", "etsy_url": ""},
                    ) as get_product_by_row, \
                    patch.object(dashboard_app, "save_to_excel") as save_to_excel, \
                    patch.object(dashboard_app, "_run_poster", new=AsyncMock(return_value=None)):
                response = await dashboard_app.post_to_etsy(4)

                payload = _response_payload(response)
                self.assertEqual(202, response.status_code)
                self.assertTrue(payload["ok"])
                command_id = payload["command"]["command_id"]
                self.assertEqual(admitted_shop_id, dashboard_app._active_shop_id)

                dashboard_app._active_shop_id = switched_shop_id
                lock.release()

                for _ in range(100):
                    state = dashboard_app._OPERATION_QUEUE_COMMANDS.get(command_id)
                    if state and state.get("status") in {"failed", "succeeded", "cancelled"}:
                        break
                    await asyncio.sleep(0.01)

                state = dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]
                self.assertEqual("failed", state["status"])
                self.assertIn("Shop đổi giữa lúc chờ hàng đợi", state["error"])
                self.assertIn(f"captured='{admitted_shop_id}'", state["error"])
                self.assertIn(f"current='{switched_shop_id}'", state["error"])
                self.assertEqual(1, get_product_by_row.call_count)
                save_to_excel.assert_not_called()
                dashboard_app._run_poster.assert_not_called()
        finally:
            if lock.locked():
                lock.release()
            dashboard_app._active_shop_id = admitted_shop_id
            fixture.cleanup()

    async def test_queued_post_blocks_all_and_selected_when_lock_is_held(self) -> None:
        fixture = _TmpShopFixture()
        (fixture.shop_dir / "product-259").mkdir(parents=True, exist_ok=True)
        _create_listing_sheet(
            fixture.excel_path,
            {
                4: {
                    2: "product-259",
                    5: 9.99,
                    8: "Sample title",
                    14: "⏳ Chờ đăng",
                }
            },
        )

        lock = dashboard_app._OPERATION_QUEUE_LOCK
        await lock.acquire()
        try:
            with patch.object(dashboard_app, "SHOP_DIR", return_value=fixture.shop_dir), \
                    patch.object(dashboard_app, "EXCEL_FILE", return_value=fixture.excel_path), \
                    patch.object(dashboard_app, "_active_shop_id", "daisyflowdigital"), \
                    patch.object(
                        dashboard_app,
                        "get_product_by_row",
                        return_value={"folder": "product-259", "status": "⏳ Chờ đăng", "etsy_url": ""},
                    ), \
                    patch.object(dashboard_app, "_run_poster", new=AsyncMock(return_value=None)):
                response = await dashboard_app.post_to_etsy(4)
                payload = _response_payload(response)
                self.assertEqual(202, response.status_code)
                command_id = payload["command"]["command_id"]
                state = dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]
                self.assertEqual("queued", state["status"])

                run_all = await dashboard_app.run_all_pending()
                self.assertEqual(409, run_all.status_code)

                with self.assertRaises(HTTPException) as context:
                    await dashboard_app.run_selected_products(_Request({
                        "shop": "daisyflowdigital",
                        "items": [{"row": 4, "folder": "product-259"}],
                    }))
                self.assertEqual(409, context.exception.status_code)

            lock.release()

            for _ in range(100):
                state = dashboard_app._OPERATION_QUEUE_COMMANDS.get(command_id)
                if state and state.get("status") in {"succeeded", "failed", "cancelled"}:
                    break
                await asyncio.sleep(0.01)
            state = dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]
            self.assertIn(state["status"], {"succeeded", "failed", "cancelled"})
        finally:
            if lock.locked():
                lock.release()
            fixture.cleanup()

    async def test_queued_run_poster_command_marks_failed_on_non_zero_exit(self) -> None:
        fixture = _TmpShopFixture()
        (fixture.shop_dir / "product-259").mkdir()

        async def callback():
            await dashboard_app._run_poster(
                4,
                "product-259",
                fixture.shop_id,
            )

        fake_proc = AsyncMock()
        fake_proc.stdout = None
        fake_proc.wait = AsyncMock(return_value=3)

        with patch.object(dashboard_app, "SHOP_DIR", return_value=fixture.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=fixture.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", fixture.shop_id), \
                patch.object(dashboard_app, "_runtime_prefetch_import_check", AsyncMock(return_value=(True, "ok")), \
                ), \
                patch.object(dashboard_app.asyncio, "create_subprocess_exec", AsyncMock(return_value=fake_proc)), \
                patch.object(dashboard_app, "_terminate_subprocess", new=AsyncMock(return_value=None)), \
                patch.object(dashboard_app, "save_to_excel") as save_to_excel:
            command, created = await dashboard_app._enqueue_operation(
                operation="etsy-post",
                shop_id=fixture.shop_id,
                target="4:product-259",
                callback=callback,
            )
            self.assertTrue(created)
            command_id = command["command_id"]

            for _ in range(100):
                state = dashboard_app._OPERATION_QUEUE_COMMANDS.get(command_id)
                if state and state.get("status") in {"failed", "succeeded", "cancelled"}:
                    break
                await asyncio.sleep(0.01)

            state = dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]
            self.assertEqual("failed", state["status"])
            self.assertIn("❌ Lỗi exit 3", state["error"])
            save_to_excel.assert_any_call(4, {"status": "❌ Lỗi exit 3"})

        fixture.cleanup()

    async def test_cancelled_queued_run_poster_command_cleans_up_process_and_locks(self) -> None:
        fixture = _TmpShopFixture()
        (fixture.shop_dir / "product-259").mkdir()
        posted = asyncio.Event()
        release = asyncio.Event()

        async def blocking_wait():
            posted.set()
            await release.wait()

        async def fake_create_subprocess(*_args, **_kwargs):
            proc = AsyncMock()
            proc.stdout = None
            proc.wait = AsyncMock(side_effect=blocking_wait)
            proc.kill = AsyncMock()
            proc.communicate = AsyncMock(return_value=(b"", b""))
            return proc

        async def callback():
            await dashboard_app._run_poster(
                4,
                "product-259",
                fixture.shop_id,
            )

        with patch.object(dashboard_app, "SHOP_DIR", return_value=fixture.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=fixture.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", fixture.shop_id), \
                patch.object(dashboard_app, "_runtime_prefetch_import_check", AsyncMock(return_value=(True, "ok")), \
                ), \
                patch.object(dashboard_app.asyncio, "create_subprocess_exec", new=AsyncMock(side_effect=fake_create_subprocess)), \
                patch.object(dashboard_app, "_terminate_subprocess", new=AsyncMock(return_value=None)), \
                patch.object(dashboard_app, "save_to_excel"):
            command, created = await dashboard_app._enqueue_operation(
                operation="etsy-post",
                shop_id=fixture.shop_id,
                target="4:product-259",
                callback=callback,
            )
            self.assertTrue(created)
            command_id = command["command_id"]
            queue_task_key = f"{dashboard_app._OPERATION_QUEUE_TASK_PREFIX}{command_id}"

            for _ in range(100):
                state = dashboard_app._OPERATION_QUEUE_COMMANDS.get(command_id)
                if state and state.get("status") in {"running", "failed", "succeeded", "cancelled"}:
                    break
                await asyncio.sleep(0.01)

            self.assertEqual("running", dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]["status"])
            queue_task = dashboard_app._running_tasks[queue_task_key]
            lock_key = dashboard_app._etsy_post_lock_key(fixture.shop_id)

            await posted.wait()
            queue_task.cancel()
            await asyncio.gather(queue_task, return_exceptions=True)

            for _ in range(100):
                if dashboard_app._OPERATION_QUEUE_COMMANDS.get(command_id, {}).get("status") == "cancelled":
                    break
                await asyncio.sleep(0.01)

            state = dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]
            self.assertEqual("cancelled", state["status"])
            self.assertNotIn(queue_task_key, dashboard_app._running_tasks)
            self.assertNotIn(lock_key, dashboard_app._running_processes)
            self.assertNotIn("product-259", dashboard_app._running_processes)
            self.assertNotIn(command_id, dashboard_app._running_tasks)

            release.set()
            fixture.cleanup()

    async def test_run_queued_operation_marks_failed_when_callback_raises(self) -> None:
        async def exploding_callback():
            raise RuntimeError("callback failed after queue admission")

        command, created = await dashboard_app._enqueue_operation(
            operation="etsy-post",
            shop_id="daisyflowdigital",
            target="4:product-259",
            callback=exploding_callback,
        )

        self.assertTrue(created)
        command_id = command["command_id"]
        for _ in range(50):
            state = dashboard_app._OPERATION_QUEUE_COMMANDS.get(command_id)
            if state and state.get("status") in {"failed", "succeeded", "cancelled"}:
                break
            await asyncio.sleep(0.01)
        state = dashboard_app._OPERATION_QUEUE_COMMANDS[command_id]
        self.assertEqual("failed", state["status"])
        self.assertIn("callback failed after queue admission", state["error"])
