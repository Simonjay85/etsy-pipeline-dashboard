"""
Etsy Alt Text Filler
• Tự động quét các listing từ Excel đã có URL ở cột P
• Mở từng listing trên Etsy dưới dạng Listing Editor
• Tạo Alt Text chuẩn SEO từ Title + Keywords
• Tự động hover qua từng ảnh, click Edit, điền Alt Text, click Apply và Save thay đổi
Chạy: python3 etsy_fill_alt_texts.py [--mode draft/active/all] [--row row_index]
"""
import asyncio, sys, subprocess, argparse
from pathlib import Path

def ensure_deps():
    pkgs = {"openpyxl": "openpyxl", "playwright": "playwright"}
    for mod, pkg in pkgs.items():
        try:
            __import__(mod)
        except ImportError:
            print(f"▶ Cài {pkg}...")
            subprocess.run([sys.executable, "-m", "pip", "install", pkg, "--quiet"], check=True)
    try:
        from google import genai
    except ImportError:
        print("▶ Cài google-genai...")
        subprocess.run([sys.executable, "-m", "pip", "install", "google-genai", "--quiet"], check=True)

ensure_deps()

import openpyxl
from playwright.async_api import async_playwright

try:
    from google import genai
    from google.genai import types
    has_genai = True
except ImportError:
    has_genai = False

BASE_DIR    = Path(__file__).parent
EXCEL_FILE  = BASE_DIR / "shops" / "templystudios" / "Etsy_SEO_Generator.xlsx"
BROWSER_DIR = BASE_DIR / ".browser-session"
CHROME_PATH = Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")

# ── Tab helper ────────────────────────────────────────────────────────────────
async def click_tab(page, *names):
    """Click tab theo tên (thử nhiều selector)."""
    for name in names:
        for sel in [
            f'[role="tab"]:has-text("{name}")',
            f'button:has-text("{name}")',
            f'a:has-text("{name}")',
            f'li:has-text("{name}")',
            f'[class*="tab"]:has-text("{name}")',
        ]:
            el = page.locator(sel).first
            if await el.count() > 0 and await el.is_visible():
                await el.click()
                await page.wait_for_timeout(1500)
                return True
    return False

# ── Alt text generator ────────────────────────────────────────────────────────
def generate_alt_texts(title: str, keywords: str, count: int) -> list:
    """Tạo alt text cho từng ảnh dựa vào title và keywords (tối đa 250 ký tự/ảnh)."""
    clean_title = (title or "")[:200].strip()
    kw_list = [k.strip() for k in (keywords or "").split(",") if k.strip()]

    suffixes = [
        "instant download",
        "printable PDF",
        "digital download",
        "printable template",
        "digital planner",
        "PDF download",
        "printable download",
        "digital file",
        "instant printable",
        "editable template",
    ]

    alt_texts = []
    for i in range(count):
        if i == 0:
            alt_texts.append(clean_title[:250])
        else:
            suffix = suffixes[i % len(suffixes)]
            if kw_list:
                kw = kw_list[i % len(kw_list)]
                text = f"{kw} - {suffix}"
            else:
                text = f"{clean_title[:200]} - {suffix}"
            alt_texts.append(text[:250])
    return alt_texts

# ── Visual Alt Text Generator ──────────────────────────────────────────────────
def generate_visual_alt_text(img_path, title, keywords):
    """
    Sử dụng Gemini 2.5 Flash thông qua Vertex AI để phân tích hình ảnh và sinh Alt Text chuẩn SEO.
    Nếu thất bại hoặc không có thư viện, trả về None để dùng cơ chế fallback dựa trên text.
    """
    if not has_genai:
        print("      ⚠️ Không có thư viện google-genai. Sử dụng Text SEO fallback.")
        return None

    try:
        # Đọc ảnh dưới dạng bytes
        with open(img_path, "rb") as f:
            img_bytes = f.read()

        # Xác định mime_type
        suffix = Path(img_path).suffix.lower()
        mime_type = "image/png"
        if suffix in [".jpg", ".jpeg"]:
            mime_type = "image/jpeg"
        elif suffix == ".webp":
            mime_type = "image/webp"
        elif suffix == ".gif":
            mime_type = "image/gif"

        # Khởi tạo client Vertex AI
        client = genai.Client(vertexai=True, project="temply-ai-lab", location="us-central1")

        prompt = f"""
Analyze the provided product image and generate a highly descriptive, SEO-optimized Alt Text for Etsy.

Instructions:
1. Describe the layout, visual design, text, and specific pages or features shown in the image in detail.
2. Incorporate 1-2 relevant SEO keywords from the following list: {keywords}.
3. The description must be natural, engaging, and extremely helpful for visually impaired buyers.
4. Keep the length strictly under 250 characters.
5. Do NOT use generic phrases like "image of" or "screenshot of".
6. Output ONLY the raw alt text, no quotes, no markdown, no comments, just the plain text.

Listing Title: {title}
"""
        part = types.Part.from_bytes(
            data=img_bytes,
            mime_type=mime_type
        )

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[part, prompt]
        )

        text = response.text
        if text:
            text = text.strip()
            # Loại bỏ các dấu nháy nếu Gemini trả về
            if text.startswith('"') and text.endswith('"'):
                text = text[1:-1].strip()
            if text.startswith("'") and text.endswith("'"):
                text = text[1:-1].strip()

            # Đảm bảo độ dài dưới 250 ký tự
            if len(text) > 250:
                text = text[:247] + "..."
        else:
            text = ""

        return text
    except Exception as e:
        print(f"      ❌ Lỗi khi phân tích ảnh qua Gemini: {e}")
        return None

# ── Fill Alt Text ─────────────────────────────────────────────────────────────
async def fill_image_alt_texts(page, title, keywords, folder=None):
    """Điền alt text cho từng ảnh bằng Playwright modal thumbnail, dùng phân tích ảnh Gemini làm cốt lõi."""
    print("  ✍️ Đang quét các ô ảnh thumbnail trên Etsy...")
    await page.wait_for_timeout(3000)

    # Định vị các nút thumbnail hình vuông đại diện cho từng ảnh trên lưới
    thumb_btns = page.locator('button.le-aspect-ratio--square')
    cnt = await thumb_btns.count()
    print(f"    🔍 Tìm thấy {cnt} ảnh trên giao diện Etsy.")

    if cnt == 0:
        print("    ⚠️ Không tìm thấy ảnh nào trên giao diện để điền Alt Text.")
        return 0

    # Lấy danh sách tệp ảnh cục bộ tương ứng (nếu có folder)
    img_files = []
    if folder:
        img_dir = BASE_DIR / "shops" / "templystudios" / folder / "images"
        if img_dir.exists():
            img_files = sorted([
                f for f in img_dir.iterdir()
                if f.suffix.lower() in ['.png', '.jpg', '.jpeg', '.webp', '.gif']
            ])
            print(f"    📂 Tìm thấy {len(img_files)} tệp ảnh cục bộ tại {folder}/images/")
        else:
            print(f"    ⚠️ Thư mục ảnh cục bộ không tồn tại: {img_dir}")

    # Sinh danh sách Alt Text dạng văn bản dự phòng (fallback)
    fallback_alts = generate_alt_texts(title, keywords, cnt)
    filled = 0

    for i in range(cnt):
        print(f"\n    🎬 Xử lý ảnh #{i+1}...")
        btn = thumb_btns.nth(i)
        
        try:
            # 1. Hover và click thumbnail
            await btn.scroll_into_view_if_needed()
            await page.wait_for_timeout(400)
            await btn.hover()
            await page.wait_for_timeout(400)
            await btn.click(force=True)
            await page.wait_for_timeout(2000)

            # 2. Đợi nút "+ Alt text" hoặc "Alt text" trong dialog trung gian xuất hiện và click
            alt_btn = page.locator('div[role="dialog"] button:has-text("Alt text"), button:has-text("Alt text"), button:has-text("+ Alt text")').first
            try:
                await alt_btn.wait_for(state="visible", timeout=3000)
                await alt_btn.click()
                await page.wait_for_timeout(1500)
            except Exception as alt_btn_ex:
                print(f"      ❌ Không tìm thấy nút mở Alt Text trong dialog: {alt_btn_ex}")
                try:
                    await page.keyboard.press("Escape")
                except:
                    pass
                await page.wait_for_timeout(1000)
                continue

            # 3. Tạo nội dung Alt Text (Thử Gemini trước, nếu lỗi/thiếu ảnh thì dùng Fallback)
            alt_val = None
            if i < len(img_files):
                img_path = img_files[i]
                print(f"      📸 Đang phân tích ảnh cục bộ qua Gemini: {img_path.name}...")
                try:
                    alt_val = await asyncio.wait_for(
                        asyncio.to_thread(generate_visual_alt_text, img_path, title, keywords),
                        timeout=15.0
                    )
                except asyncio.TimeoutError:
                    print("      ⚠️ Gemini API call timed out after 15 seconds. Using Text SEO fallback.")
                    alt_val = None
                except Exception as g_err:
                    print(f"      ⚠️ Lỗi khi gọi Gemini: {g_err}. Using Text SEO fallback.")
                    alt_val = None
            
            if not alt_val:
                alt_val = fallback_alts[i]
                print(f"      ⚠️ Sử dụng Alt Text văn bản SEO (fallback): '{alt_val[:50]}...'")
            else:
                print(f"      ✨ Sinh Alt Text bằng Gemini thành công: '{alt_val}'")

            # 4. Tìm textarea trong Alt Text modal và điền
            alt_input = page.locator('div[role="dialog"]:visible textarea, div[class*="dialog"]:visible textarea, [class*="modal"]:visible textarea').first
            await alt_input.wait_for(state="visible", timeout=3000)
            await alt_input.click()
            await alt_input.fill(alt_val)
            await page.wait_for_timeout(500)

            # 5. Click Apply để lưu Alt Text hiện tại
            apply_btn = page.locator('div[role="dialog"]:visible button:has-text("Apply")').first
            await apply_btn.wait_for(state="visible", timeout=2000)
            await apply_btn.click()
            await page.wait_for_timeout(500)
            
            # Đợi modal điền Alt Text đóng hoàn toàn
            await alt_input.wait_for(state="hidden", timeout=5000)

            # 6. Click Done để lưu và đóng dialog ảnh hiện tại
            done_btn = page.locator('div[role="dialog"]:visible button:has-text("Done")').first
            await done_btn.wait_for(state="visible", timeout=2000)
            await done_btn.click(force=True)
            await page.wait_for_timeout(1000)
            
            print(f"      ✓ Đã điền và lưu Alt Text cho ảnh #{i+1} thành công!")
            filled += 1

        except Exception as e:
            print(f"      ❌ Lỗi khi xử lý ảnh #{i+1}: {e}")
            try:
                # Tránh kẹt modal: Nhấn Escape 2 lần để đóng các dialog đang mở
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(1000)
                await page.keyboard.press("Escape")
            except:
                pass
            await page.wait_for_timeout(1000)

    return filled

# ── Main logic ────────────────────────────────────────────────────────────────
async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", type=str, choices=["draft", "active", "all"], default="draft", 
                        help="Chế độ điền: draft (chỉ nháp), active (chỉ đã đăng), all (tất cả)")
    parser.add_argument("--row", type=int, default=None, 
                        help="Chỉ định dòng cụ thể trong Excel để xử lý")
    args = parser.parse_args()

    print("\n" + "="*60)
    print("  🎨 Etsy Alt Text Automatic Filler")
    print("="*60)

    if not EXCEL_FILE.exists():
        print(f"❌ Không tìm thấy file Excel tại: {EXCEL_FILE}")
        return

    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Listings"]

    # Thu thập hàng cần xử lý
    rows_to_process = []
    for r in range(4, ws.max_row + 1):
        folder = ws.cell(row=r, column=2).value
        title = ws.cell(row=r, column=8).value
        keywords = ws.cell(row=r, column=3).value
        status = ws.cell(row=r, column=14).value
        url = ws.cell(row=r, column=16).value

        # Bỏ qua dòng trống
        if not title:
            continue

        status_str = str(status).strip() if status else ""
        url_str = str(url).strip() if url else ""

        # Kiểm tra chỉ định dòng cụ thể
        if args.row is not None:
            if r != args.row:
                continue
        else:
            # Lọc theo mode
            if args.mode == "draft" and "draft" not in status_str.lower():
                continue
            elif args.mode == "active" and "active" not in status_str.lower():
                continue
            
            # Phải có URL Etsy hợp lệ mới cập nhật được
            if not url_str or url_str == "None" or "etsy.com" not in url_str:
                continue

        rows_to_process.append({
            "row": r,
            "folder": folder,
            "title": title,
            "keywords": keywords or "",
            "url": url_str,
            "status": status_str
        })

    if not rows_to_process:
        print(f"ℹ️ Không tìm thấy sản phẩm nào phù hợp (Mode: {args.mode}, Row: {args.row}).")
        return

    print(f"📋 Tìm thấy {len(rows_to_process)} sản phẩm cần điền Alt Text.")
    for idx, item in enumerate(rows_to_process, 1):
        print(f"  {idx:2}. Row {item['row']:3} | {str(item['folder']):<12} | Status: {item['status']:<16} | {item['title'][:45]}...")

    print("\n🚀 Khởi động Chrome session...")
    BROWSER_DIR.mkdir(exist_ok=True)
    
    async with async_playwright() as pw:
        launch_kw = dict(
            user_data_dir=str(BROWSER_DIR),
            headless=False,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
            viewport=None,
        )
        if CHROME_PATH.exists():
            launch_kw["executable_path"] = str(CHROME_PATH)

        ctx  = await pw.chromium.launch_persistent_context(**launch_kw)
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()

        # Kiểm tra đăng nhập
        await page.goto("https://www.etsy.com/your/shops/me/tools/listings", wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)

        if "signin" in page.url or "join" in page.url:
            print("\n  ⚠ Chưa đăng nhập Etsy! Vui lòng đăng nhập trên trình duyệt vừa mở.")
            input("  → Nhấn Enter sau khi đăng nhập xong trên Etsy...\n")
            await page.goto("https://www.etsy.com/your/shops/me/tools/listings", wait_until="domcontentloaded")
            await page.wait_for_timeout(3000)

        print("  ✅ Đăng nhập thành công! Bắt đầu điền Alt Text...\n")

        for idx, item in enumerate(rows_to_process, 1):
            print(f"\n------------------------------------------------------------")
            print(f"🎬 [{idx}/{len(rows_to_process)}] Xử lý Row {item['row']} | {item['folder']}")
            print(f"   Title: {item['title'][:60]}...")
            
            # Trích xuất listing_id từ URL (ví dụ: https://www.etsy.com/listing/4509070606)
            listing_id = item['url'].split("/listing/")[-1].split("?")[0].strip()
            if not listing_id.isdigit():
                print(f"   ❌ URL Etsy không hợp lệ: {item['url']}")
                continue

            edit_url = f"https://www.etsy.com/your/shops/me/listing-editor/edit/{listing_id}"
            print(f"   🔗 Mở trang sửa listing: {edit_url}")
            
            try:
                # Sử dụng wait_until="commit" và tăng timeout lên 60s để tránh bị nghẽn do tracking scripts của Etsy
                await page.goto(edit_url, wait_until="commit", timeout=60000)
                await page.wait_for_timeout(8000)
                
                # Chờ một số phần tử cơ bản hiển thị
                try:
                    await page.locator('button:has-text("Save draft"), button:has-text("Publish changes"), button[class*="save"]').first.wait_for(state="attached", timeout=15000)
                except Exception as wait_ex:
                    print(f"   ℹ️ Đợi trang hiển thị: {wait_ex}")

                # Chọn tab Photo & Video để hiển thị các ảnh
                print("   📸 Đang mở tab Photo & Video...")
                await click_tab(page, "Photo & Video", "Photos")
                await page.wait_for_timeout(1000)

                # Điền Alt Text
                filled_count = await fill_image_alt_texts(page, item['title'], item['keywords'], item['folder'])
                
                if filled_count > 0:
                    print(f"   💾 Lưu thay đổi cho listing...")
                    saved_listing = False
                    for save_sel in [
                        'button:has-text("Save draft")',
                        'button:has-text("Publish changes")',
                        'button:has-text("Save and continue")',
                        'button:has-text("Publish")',
                        'button[class*="save"]',
                    ]:
                        btn = page.locator(save_sel).first
                        if await btn.count() > 0 and await btn.is_visible():
                            await btn.click()
                            parts = save_sel.split('"')
                            button_name = parts[1] if len(parts) > 1 else save_sel
                            print(f"   ✓ Đã click nút lưu ({button_name})!")
                            saved_listing = True
                            await page.wait_for_timeout(3000)
                            break
                    
                    if not saved_listing:
                        print("   ⚠️ Không tìm thấy nút lưu listing, thử click nút Publish")
                        publish_btn = page.locator('button[data-testid="publish-button"]').first
                        if await publish_btn.count() > 0 and await publish_btn.is_visible():
                            await publish_btn.click()
                            print("   ✓ Đã click nút Publish!")
                            await page.wait_for_timeout(3000)
                else:
                    print("   ℹ️ Không điền được ảnh nào, bỏ qua lưu listing.")

            except Exception as ex:
                print(f"   ❌ Lỗi khi sửa listing Row {item['row']}: {ex}")

        await ctx.close()
        print("\n" + "="*60)
        print("  🎉 Hoàn tất quá trình điền Alt Text cho toàn bộ sản phẩm!")
        print("="*60 + "\n")

if __name__ == "__main__":
    asyncio.run(main())
