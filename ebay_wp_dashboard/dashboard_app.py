"""
eBay + WordPress Dashboard — port 8091
Quản lý sản phẩm (eBay + WooCommerce) và bài viết blog (WordPress).
Chạy: python3 ebay_wp_dashboard/dashboard_app.py
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import shutil
import subprocess
import sys
import time
import urllib.parse
from pathlib import Path
from typing import Optional

import openpyxl
from fastapi import BackgroundTasks, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.base import BaseHTTPMiddleware

# Local imports (same package dir)
DASH_DIR = Path(__file__).parent
BASE_DIR = DASH_DIR.parent
sys.path.insert(0, str(DASH_DIR))

from excel_helpers import (  # noqa: E402
    IMG_EXTS,
    DIGITAL_EXTS,
    add_post_row,
    add_product_row,
    delete_post_row,
    delete_product_row,
    ensure_excel,
    posts_from_excel,
    products_from_excel,
    save_post_row,
    save_product_row,
)
from wp_client import WPClient  # noqa: E402

CONFIG_FILE = DASH_DIR / "ebay_wp_config.json"
SECRETS_FILE = DASH_DIR / "ebay_wp_secrets.json"
ACTIVE_SITE_FILE = DASH_DIR / "active_site.txt"
SHOPS_WP_DIR = BASE_DIR / "shops_wp"
STATIC_DIR = DASH_DIR / "dashboard_static"
PYTHON_BIN = sys.executable
PORT = 8091

_LOG_NOISE = ("UserWarning", "DeprecationWarning")

app = FastAPI(title="eBay + WordPress Dashboard")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        if request.url.path.startswith("/static/") or request.url.path == "/":
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        return response


app.add_middleware(NoCacheMiddleware)

_log_subscribers: list[asyncio.Queue] = []
_running_processes: dict[str, Optional[asyncio.subprocess.Process]] = {}
_post_queue: Optional[asyncio.Queue] = None
_queued_posts: list[dict] = []
_post_worker_task: Optional[asyncio.Task] = None
_stop_requested = False


def _load_json_config(path: Path, label: str) -> dict:
    if not path.exists():
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except Exception as exc:
        print(f"[CONFIG] Failed to load {label} config from {path}: {exc}")
    return {}


async def _terminate_process(key: str, proc: asyncio.subprocess.Process, timeout: float = 1.2) -> str:
    if proc.returncode is not None:
        return key
    try:
        proc.terminate()
    except ProcessLookupError:
        return key
    except Exception:
        return key
    try:
        await asyncio.wait_for(proc.wait(), timeout=timeout)
    except asyncio.TimeoutError:
        try:
            proc.kill()
            await asyncio.wait_for(proc.wait(), timeout=timeout)
        except Exception:
            pass
    return key


def broadcast(msg: str):
    for q in _log_subscribers:
        try:
            q.put_nowait(msg)
        except asyncio.QueueFull:
            pass


def load_sites() -> dict:
    return _load_json_config(CONFIG_FILE, "sites")


def load_secrets() -> dict:
    return _load_json_config(SECRETS_FILE, "secrets")


def get_site_with_secrets(site_id: str) -> dict:
    sites = load_sites()
    site = dict(sites.get(site_id, {}))
    secrets = load_secrets().get(site_id, {})
    site.update(secrets)
    return site


SITES = load_sites()


def _load_active_site_id() -> str:
    if ACTIVE_SITE_FILE.exists():
        saved = ACTIVE_SITE_FILE.read_text(encoding="utf-8").strip()
        if saved in SITES:
            return saved
    return list(SITES.keys())[0] if SITES else ""


_active_site_id = _load_active_site_id()


def save_sites():
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(SITES, f, indent=2, ensure_ascii=False)


def SITE_DIR() -> Path:
    return SHOPS_WP_DIR / _active_site_id


def EXCEL_FILE() -> Path:
    return SITE_DIR() / "Platform_Manager.xlsx"


def make_wp_client(site: dict | None = None) -> WPClient:
    site = site or get_site_with_secrets(_active_site_id)
    return WPClient(
        site_url=site.get("wordpress_url", ""),
        wp_username=site.get("wp_username", ""),
        wp_app_password=site.get("wp_app_password", ""),
        wc_consumer_key=site.get("wc_consumer_key", ""),
        wc_consumer_secret=site.get("wc_consumer_secret", ""),
    )


def enrich_products(products: list[dict]) -> list[dict]:
    site_dir = SITE_DIR()
    for p in products:
        folder = p["folder"]
        img_dir = site_dir / folder / "images"
        images = sorted(
            f.name for f in img_dir.iterdir() if f.suffix.lower() in IMG_EXTS
        ) if img_dir.exists() else []
        p["all_images"] = []
        if images:
            for img in images:
                try:
                    mtime = int((img_dir / img).stat().st_mtime)
                except Exception:
                    mtime = 0
                enc = urllib.parse.quote(img)
                p["all_images"].append(f"/thumb/{folder}/images/{enc}?v={mtime}")
            p["thumb"] = p["all_images"][0]
        else:
            p["thumb"] = None
        p["preflight_ok"] = not p.get("missing_fields")
    return products


def product_stats(products: list[dict]) -> dict:
    return {
        "total": len(products),
        "ebay_active": sum(1 for p in products if p["ebay_status"] in ("active", "draft")),
        "woo_published": sum(1 for p in products if p["woo_status"] in ("published", "publish")),
        "pending": sum(1 for p in products if p["ebay_status"] == "pending" or p["woo_status"] == "pending"),
        "errors": sum(1 for p in products if p["ebay_status"] == "error" or p["woo_status"] == "error"),
    }


def post_stats(posts: list[dict]) -> dict:
    return {
        "total": len(posts),
        "published": sum(1 for p in posts if p["wp_status"] == "publish"),
        "draft": sum(1 for p in posts if p["wp_status"] == "draft"),
    }


def preflight_product(product: dict, site_dir: Path) -> dict:
    issues = list(product.get("missing_fields", []))
    folder_path = site_dir / product["folder"]
    file_dir = folder_path / "files"
    dig = [f for f in file_dir.iterdir() if f.suffix.lower() in DIGITAL_EXTS] if file_dir.exists() else []
    if not dig:
        issues.append("no_digital_file")
    return {"ok": len(issues) == 0, "issues": issues}


# ── Static ──────────────────────────────────────────────────────────────────────
@app.get("/")
def index():
    return FileResponse(STATIC_DIR / "index.html")


app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ── Sites ─────────────────────────────────────────────────────────────────────
@app.get("/api/sites")
def api_sites():
    sites = []
    for sid, cfg in load_sites().items():
        sites.append({
            "id": sid,
            "name": cfg.get("name", sid),
            "emoji": cfg.get("emoji", "🏬"),
            "wordpress_url": cfg.get("wordpress_url", ""),
            "ebay_seller_url": cfg.get("ebay_seller_url", ""),
        })
    return {"sites": sites, "active": _active_site_id}


@app.post("/api/set-site")
async def api_set_site(request: Request):
    global _active_site_id
    body = await request.json()
    site_id = body.get("site_id", "")
    if site_id not in load_sites():
        raise HTTPException(400, "Unknown site")
    _active_site_id = site_id
    ACTIVE_SITE_FILE.write_text(site_id, encoding="utf-8")
    SITE_DIR().mkdir(parents=True, exist_ok=True)
    ensure_excel(EXCEL_FILE())
    return {"ok": True, "active": site_id}


@app.post("/api/sites/update")
async def api_sites_update(request: Request):
    global SITES
    body = await request.json()
    site_id = body.get("id") or _active_site_id
    if site_id not in SITES:
        raise HTTPException(400, "Unknown site")
    for key in ("name", "emoji", "wordpress_url", "ebay_seller_url", "wp_username", "debug_port", "browser_session"):
        if key in body:
            SITES[site_id][key] = body[key]
    save_sites()
    return {"ok": True}


@app.post("/api/test-connection/wp")
async def test_wp():
    client = make_wp_client()
    return await client.test_wp_connection()


@app.post("/api/test-connection/woo")
async def test_woo():
    client = make_wp_client()
    return await client.test_wc_connection()


@app.post("/api/secrets/save")
async def save_secrets(request: Request):
    body = await request.json()
    site_id = body.get("site_id") or _active_site_id
    secrets = load_secrets()
    entry = secrets.setdefault(site_id, {})
    for key in ("wp_app_password", "wc_consumer_key", "wc_consumer_secret"):
        if key in body:
            entry[key] = body[key]
    with open(SECRETS_FILE, "w", encoding="utf-8") as f:
        json.dump(secrets, f, indent=2)
    return {"ok": True}


# ── Products ──────────────────────────────────────────────────────────────────
@app.get("/api/products")
def api_products():
    ensure_excel(EXCEL_FILE())
    products = enrich_products(products_from_excel(SITE_DIR(), EXCEL_FILE()))
    return {"products": products, "stats": product_stats(products)}


@app.get("/api/products/{row}")
def api_product(row: int):
    products = enrich_products(products_from_excel(SITE_DIR(), EXCEL_FILE()))
    p = next((x for x in products if x["row"] == row), None)
    if not p:
        raise HTTPException(404)
    p["preflight"] = preflight_product(p, SITE_DIR())
    return p


@app.patch("/api/products/{row}")
async def api_patch_product(row: int, request: Request):
    body = await request.json()
    allowed = {
        "title", "description", "tags", "price", "qty", "sku", "category", "keywords",
        "ebay_status", "ebay_url", "woo_status", "woo_product_id", "woo_url", "extra",
    }
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates:
        save_product_row(EXCEL_FILE(), row, updates)
    return {"ok": True}


@app.post("/api/products")
async def api_create_product(request: Request):
    body = await request.json()
    folder = body.get("folder", "").strip()
    if not folder:
        raise HTTPException(400, "folder required")
    folder_path = SITE_DIR() / folder
    folder_path.mkdir(parents=True, exist_ok=True)
    (folder_path / "images").mkdir(exist_ok=True)
    (folder_path / "files").mkdir(exist_ok=True)
    row = add_product_row(EXCEL_FILE(), body)
    return {"ok": True, "row": row}


@app.delete("/api/products/{row}")
def api_delete_product(row: int):
    delete_product_row(EXCEL_FILE(), row)
    return {"ok": True}


@app.post("/api/batch-delete")
async def api_batch_delete(request: Request):
    body = await request.json()
    rows = sorted(body.get("rows", []), reverse=True)
    for row in rows:
        delete_product_row(EXCEL_FILE(), row)
    return {"ok": True, "deleted": len(rows)}


@app.get("/api/products/{row}/preflight")
def api_preflight(row: int):
    products = products_from_excel(SITE_DIR(), EXCEL_FILE())
    p = next((x for x in products if x["row"] == row), None)
    if not p:
        raise HTTPException(404)
    return preflight_product(p, SITE_DIR())


# ── Posts ─────────────────────────────────────────────────────────────────────
@app.get("/api/posts")
def api_posts():
    ensure_excel(EXCEL_FILE())
    posts = posts_from_excel(EXCEL_FILE())
    return {"posts": posts, "stats": post_stats(posts)}


@app.get("/api/posts/{row}")
def api_post(row: int):
    posts = posts_from_excel(EXCEL_FILE())
    p = next((x for x in posts if x["row"] == row), None)
    if not p:
        raise HTTPException(404)
    return p


@app.patch("/api/posts/{row}")
async def api_patch_post(row: int, request: Request):
    body = await request.json()
    allowed = {
        "slug", "title", "excerpt", "content", "categories", "tags",
        "featured_image", "wp_status", "wp_post_id", "wp_url",
        "seo_title", "seo_description",
    }
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates:
        save_post_row(EXCEL_FILE(), row, updates)
    return {"ok": True}


@app.post("/api/posts")
async def api_create_post(request: Request):
    body = await request.json()
    row = add_post_row(EXCEL_FILE(), body)
    return {"ok": True, "row": row}


@app.delete("/api/posts/{row}")
def api_delete_post(row: int):
    delete_post_row(EXCEL_FILE(), row)
    return {"ok": True}


# ── WooCommerce publish ───────────────────────────────────────────────────────
async def _publish_woo_row(row: int):
    site_dir = SITE_DIR()
    excel_path = EXCEL_FILE()
    products = products_from_excel(site_dir, excel_path)
    product = next((p for p in products if p["row"] == row), None)
    if not product:
        broadcast(f"[WOO] Row {row} not found")
        return
    save_product_row(excel_path, row, {"woo_status": "posting"})
    try:
        client = make_wp_client()
        folder_path = site_dir / product["folder"]
        img_paths = sorted(
            str(p) for p in (folder_path / "images").iterdir()
            if p.suffix.lower() in IMG_EXTS
        ) if (folder_path / "images").exists() else []
        dl_paths = sorted(
            str(p) for p in (folder_path / "files").iterdir()
            if p.suffix.lower() in DIGITAL_EXTS
        ) if (folder_path / "files").exists() else []
        existing_id = product.get("woo_product_id") or None
        result = await client.create_or_update_product({
            "title": product["title"],
            "description": product["description"],
            "price": product["price"],
            "sku": product["sku"],
            "tags": product["tags"],
            "status": "publish",
            "image_paths": img_paths,
            "download_paths": dl_paths,
        }, existing_id=existing_id if existing_id else None)
        save_product_row(excel_path, row, {
            "woo_status": "published",
            "woo_product_id": str(result["id"]),
            "woo_url": result.get("permalink", ""),
        })
        broadcast(f"[WOO] ✅ Row {row} published: {result.get('permalink')}")
    except Exception as e:
        save_product_row(excel_path, row, {"woo_status": "error", "extra": str(e)[:200]})
        broadcast(f"[WOO] ❌ Row {row}: {e}")


@app.post("/api/products/{row}/post-woo")
async def post_woo(row: int):
    asyncio.create_task(_publish_woo_row(row))
    return {"ok": True, "message": f"Publishing row {row} to WooCommerce"}


@app.post("/api/woo/sync")
async def woo_sync():
    try:
        client = make_wp_client()
        remote = await client.list_products(per_page=50)
        products = products_from_excel(SITE_DIR(), EXCEL_FILE())
        by_id = {p["woo_product_id"]: p for p in products if p["woo_product_id"]}
        matched = 0
        for rp in remote:
            pid = str(rp["id"])
            if pid in by_id:
                save_product_row(EXCEL_FILE(), by_id[pid]["row"], {
                    "woo_url": rp.get("permalink", ""),
                    "woo_status": rp.get("status", "published"),
                })
                matched += 1
        return {"ok": True, "remote_count": len(remote), "matched": matched}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@app.post("/api/run-all-pending-woo")
async def run_all_pending_woo():
    products = products_from_excel(SITE_DIR(), EXCEL_FILE())
    pending = [p for p in products if p["woo_status"] in ("pending", "error")]
    for p in pending:
        asyncio.create_task(_publish_woo_row(p["row"]))
    return {"ok": True, "queued": len(pending)}


# ── WordPress blog publish ────────────────────────────────────────────────────
async def _publish_wp_post_row(row: int, publish: bool = True):
    excel_path = EXCEL_FILE()
    posts = posts_from_excel(excel_path)
    post = next((p for p in posts if p["row"] == row), None)
    if not post:
        broadcast(f"[WP] Post row {row} not found")
        return
    try:
        client = make_wp_client()
        site_dir = SITE_DIR()
        feat = post.get("featured_image", "")
        if feat and not Path(feat).is_absolute():
            feat = str(site_dir / feat)
        result = await client.create_or_update_post({
            "title": post["title"],
            "content": post["content"],
            "excerpt": post["excerpt"],
            "slug": post["slug"],
            "status": "publish" if publish else "draft",
            "featured_image_path": feat if feat and Path(feat).exists() else None,
        }, existing_id=post["wp_post_id"] or None)
        save_post_row(excel_path, row, {
            "wp_post_id": str(result["id"]),
            "wp_url": result.get("link", ""),
            "wp_status": result.get("status", "draft"),
            "slug": result.get("slug", post["slug"]),
        })
        broadcast(f"[WP] ✅ Post row {row}: {result.get('link')}")
    except Exception as e:
        save_post_row(excel_path, row, {"wp_status": "error"})
        broadcast(f"[WP] ❌ Post row {row}: {e}")


@app.post("/api/posts/{row}/publish")
async def publish_post(row: int):
    asyncio.create_task(_publish_wp_post_row(row, publish=True))
    return {"ok": True}


@app.post("/api/posts/{row}/save-draft")
async def save_draft_post(row: int):
    asyncio.create_task(_publish_wp_post_row(row, publish=False))
    return {"ok": True}


@app.post("/api/posts/sync-from-wp")
async def sync_posts_from_wp():
    try:
        client = make_wp_client()
        remote = await client.list_posts(per_page=50)
        existing = {p["wp_post_id"]: p for p in posts_from_excel(EXCEL_FILE()) if p["wp_post_id"]}
        imported = 0
        for rp in remote:
            pid = str(rp["id"])
            title = rp.get("title", {}).get("rendered", "")
            if pid in existing:
                save_post_row(EXCEL_FILE(), existing[pid]["row"], {
                    "title": title,
                    "wp_url": rp.get("link", ""),
                    "wp_status": rp.get("status", ""),
                    "slug": rp.get("slug", ""),
                })
            else:
                add_post_row(EXCEL_FILE(), {
                    "title": title,
                    "slug": rp.get("slug", ""),
                    "content": rp.get("content", {}).get("rendered", ""),
                    "excerpt": rp.get("excerpt", {}).get("rendered", ""),
                    "wp_post_id": pid,
                    "wp_url": rp.get("link", ""),
                    "wp_status": rp.get("status", "draft"),
                })
                imported += 1
        return {"ok": True, "remote": len(remote), "imported": imported}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


# ── eBay subprocess runners ───────────────────────────────────────────────────
async def _run_script(script: str, args: list[str], key: str, env_extra: dict | None = None):
    env = {**os.environ, "PYTHONUNBUFFERED": "1", "ALLOW_EBAY_POST": "1"}
    if env_extra:
        env.update(env_extra)
    cmd = [PYTHON_BIN, str(DASH_DIR / script), *args]
    broadcast(f"[RUN] {' '.join(cmd)}")
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(DASH_DIR), env=env,
        )
        _running_processes[key] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text and not any(n in text for n in _LOG_NOISE):
                    broadcast(text)
        await proc.wait()
    finally:
        _running_processes.pop(key, None)


@app.post("/api/ebay/sync")
async def ebay_sync():
    key = f"ebay_sync_{_active_site_id}"
    if key in _running_processes:
        return JSONResponse({"ok": False, "error": "Sync already running"}, status_code=409)
    asyncio.create_task(_run_script("ebay_shop_sync.py", ["--site", _active_site_id], key))
    return {"ok": True}


@app.post("/api/products/{row}/post-ebay")
async def post_ebay(row: int):
    key = f"ebay_post_{_active_site_id}_{row}"
    if key in _running_processes:
        return JSONResponse({"ok": False, "error": "Already posting"}, status_code=409)
    asyncio.create_task(_run_script("ebay_auto_post.py", ["--site", _active_site_id, "--row", str(row)], key))
    return {"ok": True}


@app.post("/api/products/{row}/push-ebay")
async def push_ebay(row: int):
    key = f"ebay_push_{_active_site_id}_{row}"
    if key in _running_processes:
        return JSONResponse({"ok": False, "error": "Push already running"}, status_code=409)
    asyncio.create_task(_run_script("ebay_push_update.py", ["--site", _active_site_id, "--row", str(row)], key))
    return {"ok": True}


@app.post("/api/run-all-pending-ebay")
async def run_all_pending_ebay():
    products = products_from_excel(SITE_DIR(), EXCEL_FILE())
    pending = [p for p in products if p["ebay_status"] in ("pending", "error")]
    key = f"ebay_batch_{_active_site_id}"
    asyncio.create_task(_run_script(
        "ebay_auto_post.py",
        ["--site", _active_site_id, "--batch", str(len(pending))],
        key,
    ))
    return {"ok": True, "queued": len(pending)}


@app.post("/api/stop-all")
async def stop_all():
    global _stop_requested
    _stop_requested = True
    stopped = []
    stop_jobs = []
    for key, proc in list(_running_processes.items()):
        try:
            stopped.append(key)
            if proc is not None:
                stop_jobs.append(_terminate_process(key, proc))
        except Exception:
            pass
    if stop_jobs:
        await asyncio.gather(*stop_jobs, return_exceptions=True)
    _running_processes.clear()
    broadcast(f"[STOP] All processes terminated ({len(stopped)})")
    return {"ok": True, "stopped": stopped}


# ── Import from Etsy ──────────────────────────────────────────────────────────
@app.post("/api/import-from-etsy")
async def import_from_etsy(request: Request):
    body = await request.json()
    etsy_shop = body.get("etsy_shop") or get_site_with_secrets(_active_site_id).get("etsy_import_source", "")
    if not etsy_shop:
        raise HTTPException(400, "etsy_shop required")
    etsy_dir = BASE_DIR / "shops" / etsy_shop
    etsy_excel = etsy_dir / "Etsy_SEO_Generator.xlsx"
    if not etsy_excel.exists():
        raise HTTPException(404, f"Etsy Excel not found: {etsy_excel}")

    ensure_excel(EXCEL_FILE())
    site_dir = SITE_DIR()
    site_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(etsy_excel, data_only=True)
    ws = wb["Listings"]
    imported = 0
    skipped = 0

    existing_folders = {p["folder"] for p in products_from_excel(site_dir, EXCEL_FILE())}

    for row_num in range(4, ws.max_row + 1):
        folder = ws.cell(row=row_num, column=2).value
        if not folder:
            continue
        folder = str(folder)
        src_folder = etsy_dir / folder
        if not src_folder.exists():
            skipped += 1
            continue
        if folder in existing_folders:
            skipped += 1
            continue

        dst_folder = site_dir / folder
        if not dst_folder.exists():
            shutil.copytree(src_folder, dst_folder)

        title = ws.cell(row=row_num, column=8).value
        add_product_row(EXCEL_FILE(), {
            "folder": folder,
            "keywords": str(ws.cell(row=row_num, column=3).value or ""),
            "price": ws.cell(row=row_num, column=5).value or 4.99,
            "category": str(ws.cell(row=row_num, column=6).value or ""),
            "title": str(title or ""),
            "description": str(ws.cell(row=row_num, column=9).value or ""),
            "tags": str(ws.cell(row=row_num, column=10).value or ""),
            "qty": ws.cell(row=row_num, column=11).value or 999,
            "sku": str(ws.cell(row=row_num, column=18).value or ""),
            "ebay_status": "pending",
            "woo_status": "pending",
        })
        existing_folders.add(folder)
        imported += 1

    broadcast(f"[IMPORT] Imported {imported} products from Etsy shop {etsy_shop} (skipped {skipped})")
    return {"ok": True, "imported": imported, "skipped": skipped}


# ── Media ─────────────────────────────────────────────────────────────────────
@app.get("/thumb/{folder}/{subfolder}/{filename:path}")
def serve_thumb(folder: str, subfolder: str, filename: str):
    decoded = urllib.parse.unquote(filename)
    path = SITE_DIR() / folder / subfolder / decoded
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(path, headers={"Cache-Control": "public, max-age=3600"})


@app.get("/files/{folder}/{subfolder}/{filename:path}")
def serve_file(folder: str, subfolder: str, filename: str):
    decoded = urllib.parse.unquote(filename)
    path = SITE_DIR() / folder / subfolder / decoded
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(path)


# ── Logs & services ───────────────────────────────────────────────────────────
@app.get("/api/logs")
async def sse_logs(request: Request):
    queue: asyncio.Queue = asyncio.Queue(maxsize=500)
    _log_subscribers.append(queue)

    async def event_gen():
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    msg = await asyncio.wait_for(queue.get(), timeout=15)
                    yield f"data: {json.dumps({'msg': msg, 'ts': time.strftime('%H:%M:%S')})}\n\n"
                except asyncio.TimeoutError:
                    yield 'data: {"ping": 1}\n\n'
        finally:
            try:
                _log_subscribers.remove(queue)
            except ValueError:
                pass

    return StreamingResponse(event_gen(), media_type="text/event-stream")


async def _port_open(port: int) -> bool:
    try:
        _, writer = await asyncio.wait_for(asyncio.open_connection("127.0.0.1", port), timeout=0.2)
        writer.close()
        await writer.wait_closed()
        return True
    except Exception:
        return False


@app.get("/api/services")
async def api_services():
    site = get_site_with_secrets(_active_site_id)
    debug_port = int(site.get("debug_port", 9230))
    ebay_browser = await _port_open(debug_port)
    return {
        "ebay_browser": ebay_browser,
        "debug_port": debug_port,
        "running": list(_running_processes.keys()),
        "wordpress_url": site.get("wordpress_url", ""),
    }


@app.get("/api/health")
async def api_health():
    return {
        "ok": True,
        "service": "ebay-wordpress-dashboard",
        "active_site": _active_site_id,
        "running_jobs": len(_running_processes),
        "queued_jobs": len(_queued_posts),
        "stop_requested": _stop_requested,
    }


# ── Startup ───────────────────────────────────────────────────────────────────
@app.on_event("startup")
def on_startup():
    SHOPS_WP_DIR.mkdir(parents=True, exist_ok=True)
    if _active_site_id:
        SITE_DIR().mkdir(parents=True, exist_ok=True)
        ensure_excel(EXCEL_FILE())


if __name__ == "__main__":
    import uvicorn
    print(f"🚀 eBay + WordPress Dashboard → http://127.0.0.1:{PORT}")
    uvicorn.run(app, host="127.0.0.1", port=PORT, log_level="info")
