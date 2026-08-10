"""Per-shop catalog workbook repository primitives for atomic writes.

This module intentionally contains only workbook I/O safety mechanics for
gradual adoption by higher-level writers.
"""

from __future__ import annotations

import hashlib
import os
import shutil
import tempfile
import threading
from contextlib import contextmanager
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Iterator

try:
    import fcntl
except Exception:  # pragma: no cover - optional on non-POSIX.
    fcntl = None

import re
import openpyxl

_SAFE_SHOP_ID = re.compile(r"^[A-Za-z0-9._-]+$")
_DEFAULT_WORKBOOK = "Etsy_SEO_Generator.xlsx"
_WORKBOOK_LOCK_BASENAME = ".catalog_workbook.lock"
_WORKBOOK_EXT = ".xlsx"


class CatalogRepositoryError(RuntimeError):
    """Base class for catalog repository failures."""


class CatalogWriteConflict(CatalogRepositoryError):
    """Raised when the workbook version/hash does not satisfy preconditions."""


class CatalogWriteError(CatalogRepositoryError):
    """Raised when writing a workbook fails and rollback could not be guaranteed."""


@dataclass(frozen=True)
class CatalogWriteReceipt:
    shop_id: str
    operation: str
    workbook_path: str
    backup_path: str
    lock_path: str
    started_at_utc: str
    finished_at_utc: str
    duration_ms: int
    before_version: str
    after_version: str
    before_hash: str
    after_hash: str
    success: bool
    error: str | None


_LOCK_MAP: dict[str, threading.Lock] = {}
_LOCK_MAP_LOCK = threading.Lock()


def _validate_shop_id(shop_id: str) -> str:
    normalized = str(shop_id).strip()
    if not normalized or not _SAFE_SHOP_ID.fullmatch(normalized):
        raise CatalogRepositoryError(f"Invalid shop_id: {shop_id!r}")
    return normalized


def shop_lock_path(repo_root: Path | str, shop_id: str) -> Path:
    safe_shop_id = _validate_shop_id(shop_id)
    return Path(repo_root).resolve() / "shops" / safe_shop_id / _WORKBOOK_LOCK_BASENAME


def workbook_path(repo_root: Path | str, shop_id: str, workbook_name: str = _DEFAULT_WORKBOOK) -> Path:
    safe_shop_id = _validate_shop_id(shop_id)
    return Path(repo_root).resolve() / "shops" / safe_shop_id / str(workbook_name)


def _catalog_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while True:
            chunk = stream.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _catalog_version(path: Path) -> str:
    stat = path.stat()
    return f"{stat.st_size}:{stat.st_mtime_ns}:{stat.st_mode}:{stat.st_ino}"


def _workbook_backup_path(workbook: Path) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S-%f")
    return workbook.with_name(f".{workbook.stem}.catalog_backup_{timestamp}{_WORKBOOK_EXT}")


def _temp_workbook_path(workbook: Path) -> Path:
    with tempfile.NamedTemporaryFile(
        prefix=f".{workbook.stem}.catalog_edit_",
        suffix=workbook.suffix,
        dir=workbook.parent,
        delete=False,
    ) as stream:
        return Path(stream.name)


def _validate_workbook(path: Path) -> None:
    book = openpyxl.load_workbook(path)
    try:
        # Opening + enumerating sheet names validates the workbook shape enough to
        # catch partial writes before swap.
        _ = book.sheetnames
    finally:
        book.close()


def _copy_recovery_artifact(source: Path, backup: Path) -> None:
    backup.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, backup)


def _restore_from_backup(workbook: Path, backup: Path) -> None:
    if not backup.exists():
        raise CatalogWriteError(f"Recovery artifact missing: {backup}")
    shutil.copy2(backup, workbook)


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def _in_process_lock_key(path: Path) -> str:
    return str(path.resolve())


def _get_in_process_lock(path: Path) -> threading.Lock:
    key = _in_process_lock_key(path)
    with _LOCK_MAP_LOCK:
        lock = _LOCK_MAP.get(key)
        if lock is None:
            lock = threading.Lock()
            _LOCK_MAP[key] = lock
        return lock


@contextmanager
def _acquire_shop_lock(lock_path: Path) -> Iterator[None]:
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    in_process = _get_in_process_lock(lock_path)
    lock_file = lock_path.open("a+", encoding="utf-8")
    try:
        in_process.acquire()
        if fcntl is not None:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            if fcntl is not None:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
    finally:
        in_process.release()
        lock_file.close()


def apply_catalog_update(
    repo_root: Path | str,
    shop_id: str,
    operation: str,
    updater: Callable[[openpyxl.Workbook], None],
    *,
    expected_version: str | None = None,
    expected_hash: str | None = None,
    workbook_name: str = _DEFAULT_WORKBOOK,
) -> dict[str, str | int | bool | None]:
    """
    Apply a single catalog mutation through the repository.

    A per-shop lock is acquired, expected version/hash preconditions are enforced,
    a recovery backup is created, and workbook mutation takes place in-memory.
    Changes are saved to a same-filesystem temp file, validated, and swapped in
    with ``os.replace``.
    """
    safe_shop_id = _validate_shop_id(shop_id)
    if not isinstance(operation, str) or not operation.strip():
        raise CatalogRepositoryError("operation is required")

    target = workbook_path(repo_root, safe_shop_id, workbook_name)
    if not target.exists():
        raise CatalogRepositoryError(f"Missing workbook: {target}")

    lock_file = shop_lock_path(repo_root, safe_shop_id)
    started = _now_utc()

    with _acquire_shop_lock(lock_file):
        before_version = _catalog_version(target)
        before_hash = _catalog_hash(target)
        if expected_version is not None and expected_version != before_version:
            raise CatalogWriteConflict("Workbook version does not match expected precondition")
        if expected_hash is not None and expected_hash != before_hash:
            raise CatalogWriteConflict("Workbook hash does not match expected precondition")

        backup_path = _workbook_backup_path(target)
        temp_path: Path | None = None
        try:
            _copy_recovery_artifact(target, backup_path)
            workbook = openpyxl.load_workbook(target)
            try:
                updater(workbook)
                temp_path = _temp_workbook_path(target)
                workbook.save(temp_path)
            finally:
                workbook.close()

            _validate_workbook(temp_path)
            os.replace(temp_path, target)
            _validate_workbook(target)

            after_version = _catalog_version(target)
            after_hash = _catalog_hash(target)
            finished = _now_utc()
            # Keep successful backups out of steady-state artifacts.
            backup_path.unlink(missing_ok=True)
            return asdict(
                CatalogWriteReceipt(
                    shop_id=safe_shop_id,
                    operation=str(operation),
                    workbook_path=str(target),
                    backup_path=str(backup_path),
                    lock_path=str(lock_file),
                    started_at_utc=started,
                    finished_at_utc=finished,
                    duration_ms=_duration_ms(started, finished),
                    before_version=before_version,
                    after_version=after_version,
                    before_hash=before_hash,
                    after_hash=after_hash,
                    success=True,
                    error=None,
                )
            )
        except Exception as exc:
            restore_error: str | None = None
            try:
                _restore_from_backup(target, backup_path)
            except Exception as restore_exc:  # pragma: no cover - rare IO issue.
                restore_error = str(restore_exc)

            if isinstance(exc, CatalogRepositoryError):
                if restore_error is None:
                    raise
                raise CatalogWriteError(
                    f"catalog write failed: {exc}; rollback failed: {restore_error}"
                ) from exc

            if restore_error is None:
                raise CatalogWriteError(f"catalog write failed: {exc}") from exc
            raise CatalogWriteError(
                f"catalog write failed: {exc}; rollback failed: {restore_error}"
            ) from exc
        finally:
            if temp_path is not None:
                temp_path.unlink(missing_ok=True)


def _duration_ms(started_iso: str, finished_iso: str) -> int:
    start = datetime.fromisoformat(started_iso)
    finish = datetime.fromisoformat(finished_iso)
    return int((finish - start).total_seconds() * 1000)
