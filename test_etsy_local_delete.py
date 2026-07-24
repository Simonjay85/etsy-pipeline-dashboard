#!/usr/bin/env python3
from __future__ import annotations

import asyncio
import openpyxl
import shutil
import tempfile
from pathlib import Path
from unittest import TestCase
from unittest.mock import patch

from fastapi import HTTPException

import dashboard_app
import etsy_catalog


class Request:
    def __init__(self, payload):
        self.payload = payload

    async def json(self):
        return self.payload


class InvalidRequest(Request):
    async def json(self):
        raise ValueError("Malformed JSON")


def _create_listing_sheet(excel_path: Path, row_data: dict[int, dict[int, object]]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings"
    for row, values in row_data.items():
        for column in range(2, 19):
            value = values.get(column, "") if isinstance(values, dict) else ""
            ws.cell(row=row, column=column, value=value)
    wb.save(excel_path)


class TestDeleteLocalProducts(TestCase):
    def setUp(self):
        self.shop_id = "daisyflowdigital"
        self.tmp = Path(tempfile.mkdtemp())
        self.shop_dir = self.tmp / "shops" / self.shop_id
        self.shop_dir.mkdir(parents=True, exist_ok=True)
        self.excel_path = self.shop_dir / "Etsy_SEO_Generator.xlsx"

        _create_listing_sheet(
            self.excel_path,
            {
                4: {
                    2: "product-369",
                    5: 9.99,
                    8: "Sample title",
                    9: "sample desc",
                    10: "tag1, tag2",
                    11: 999,
                    13: "2020_2026",
                    14: "⏳ Chờ đăng",
                    15: "Digital Planner",
                    16: "https://www.etsy.com/listing/123",
                    17: "extra",
                    18: "TS_product-369",
                },
                5: {
                    2: "product-370",
                    5: 9.99,
                    8: "Sample title",
                    14: "⏳ Chờ đăng",
                },
            },
        )
        for folder in ("product-369", "product-370"):
            (self.shop_dir / folder / "images").mkdir(parents=True, exist_ok=True)
            (self.shop_dir / folder / "files").mkdir(parents=True, exist_ok=True)
            (self.shop_dir / folder / "files" / "sample.pdf").write_text("pdf", encoding="utf-8")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _row_values_from_excel(self, row: int):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Listings"]
        return [ws.cell(row=row, column=col).value for col in range(2, 19)]

    def _aggregate_local_folders(self):
        return [item.get("folder") for item in etsy_catalog.load_local_catalog(self.tmp, self.shop_id, self.excel_path)]

    def _patch_move_fail_second(self):
        original_move = dashboard_app.shutil.move
        state = {"count": 0}

        def tracked_move(source, destination):
            state["count"] += 1
            if state["count"] == 2:
                raise OSError("simulate move failure")
            return original_move(source, destination)

        return patch.object(dashboard_app.shutil, "move", side_effect=tracked_move)

    def test_single_delete_moves_folder_to_quarantine_and_clears_b_to_r(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            result = asyncio.run(
                dashboard_app.delete_product(4, Request({"shop": self.shop_id, "folder": "product-369"}))
            )

        self.assertTrue(result["ok"])
        self.assertEqual(1, result["deleted"])

        row_values = self._row_values_from_excel(4)
        self.assertTrue(all(value is None for value in row_values))
        self.assertFalse((self.shop_dir / "product-369").exists())

        quarantine_root = self.shop_dir / ".deleted_local_products"
        self.assertTrue(quarantine_root.exists())
        quarantine_dest = Path(result["items"][0]["quarantine_folder"])
        self.assertTrue(quarantine_dest.exists())
        self.assertTrue((quarantine_dest / "files" / "sample.pdf").exists())
        self.assertEqual(
            ".deleted_local_products",
            quarantine_dest.parent.name,
        )
        self.assertFalse((self.shop_dir / "product-369").exists())
        self.assertTrue((self.shop_dir / "product-370").exists())

        local_folders = self._aggregate_local_folders()
        self.assertNotIn("product-369", local_folders)

    def test_batch_delete_success(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            result = asyncio.run(
                dashboard_app.batch_delete(Request({
                    "shop": self.shop_id,
                    "items": [
                        {"row": 4, "folder": "product-369"},
                        {"row": 5, "folder": "product-370"},
                    ],
                }))
            )

        self.assertTrue(result["ok"])
        self.assertEqual(2, result["deleted"])
        self.assertEqual(2, len(result["deleted_folders"]))
        for deleted in result["items"]:
            quarantine_dest = Path(deleted["quarantine_folder"])
            self.assertTrue(quarantine_dest.exists())
            self.assertTrue((quarantine_dest / "files" / "sample.pdf").exists())

        self.assertFalse((self.shop_dir / "product-369").exists())
        self.assertFalse((self.shop_dir / "product-370").exists())
        self.assertTrue(all(value is None for value in self._row_values_from_excel(4)))
        self.assertTrue(all(value is None for value in self._row_values_from_excel(5)))
        local_folders = self._aggregate_local_folders()
        self.assertNotIn("product-369", local_folders)
        self.assertNotIn("product-370", local_folders)

    def test_batch_delete_rejects_duplicate_folders_without_mutation(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            with self.assertRaises(HTTPException) as context:
                asyncio.run(dashboard_app.batch_delete(Request({
                    "shop": self.shop_id,
                    "items": [
                        {"row": 4, "folder": "product-369"},
                        {"row": 5, "folder": "product-369"},
                    ],
                })))
        self.assertEqual(400, context.exception.status_code)
        self.assertTrue((self.shop_dir / "product-369").exists())
        self.assertTrue((self.shop_dir / "product-370").exists())
        self.assertEqual("product-369", self._row_values_from_excel(4)[0])
        self.assertEqual("product-370", self._row_values_from_excel(5)[0])

    def test_batch_delete_rollback_when_workbook_save_fails(self):
        with patch.object(openpyxl.workbook.workbook.Workbook, "save", side_effect=OSError("simulate save failure")), \
                patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            with self.assertRaises(HTTPException) as context:
                asyncio.run(dashboard_app.batch_delete(Request({
                    "shop": self.shop_id,
                    "items": [
                        {"row": 4, "folder": "product-369"},
                        {"row": 5, "folder": "product-370"},
                    ],
                })))
        self.assertEqual(500, context.exception.status_code)
        self.assertTrue((self.shop_dir / "product-369").exists())
        self.assertTrue((self.shop_dir / "product-370").exists())
        self.assertEqual("product-369", self._row_values_from_excel(4)[0])
        self.assertEqual("product-370", self._row_values_from_excel(5)[0])
        quarantine_root = self.shop_dir / ".deleted_local_products"
        if quarantine_root.exists():
            self.assertFalse(any(item.is_dir() for item in quarantine_root.iterdir()))

    def test_single_delete_row_folder_mismatch_does_nothing(self):
        with self.assertRaises(HTTPException) as context:
            with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                    patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                    patch.object(dashboard_app, "_active_shop_id", self.shop_id):
                asyncio.run(
                    dashboard_app.delete_product(4, Request({"shop": self.shop_id, "folder": "product-370"}))
                )
        self.assertEqual(409, context.exception.status_code)

        self.assertTrue((self.shop_dir / "product-369").exists())
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Listings"]
        self.assertEqual("product-369", ws.cell(row=4, column=2).value)

    def test_malformed_json_single_is_400(self):
        with self.assertRaises(HTTPException) as context:
            with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                    patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                    patch.object(dashboard_app, "_active_shop_id", self.shop_id):
                asyncio.run(dashboard_app.delete_product(4, InvalidRequest("not-json")))
        self.assertEqual(400, context.exception.status_code)

    def test_malformed_json_batch_is_400(self):
        with self.assertRaises(HTTPException) as context:
            with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                    patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                    patch.object(dashboard_app, "_active_shop_id", self.shop_id):
                asyncio.run(dashboard_app.batch_delete(InvalidRequest("not-json")))
        self.assertEqual(400, context.exception.status_code)

    def test_ui_delete_flow_awaits_aggregate_refresh(self):
        js_path = Path(__file__).parent / "dashboard_static" / "app.js"
        source = js_path.read_text(encoding="utf-8")
        self.assertIn("await loadAggregateCatalog({ throwOnError });", source)
        self.assertIn("async function loadAggregateCatalog(options = {})", source)
        self.assertIn("if (throwOnError) throw e;", source)
        self.assertIn("await loadProducts({ throwOnError: true });", source)
        self.assertIn("toast('warning', `✅ ${folder} đã bị xoá khỏi dashboard, nhưng làm mới danh sách gặp lỗi:", source)

    def test_batch_delete_transactional_rollback_when_move_fails(self):
        with self._patch_move_fail_second(), \
                patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            with self.assertRaises(HTTPException) as context:
                asyncio.run(dashboard_app.batch_delete(Request({
                    "shop": self.shop_id,
                    "items": [
                        {"row": 4, "folder": "product-369"},
                        {"row": 5, "folder": "product-370"},
                    ],
                })))
            self.assertEqual(500, context.exception.status_code)

        self.assertTrue((self.shop_dir / "product-369").exists())
        self.assertTrue((self.shop_dir / "product-370").exists())
        self.assertEqual(["product-369", "product-370"], [self._row_values_from_excel(4)[0], self._row_values_from_excel(5)[0]])
        quarantine_root = self.shop_dir / ".deleted_local_products"
        if quarantine_root.exists():
            self.assertFalse(any(item.is_dir() for item in quarantine_root.iterdir()))


if __name__ == "__main__":
    import unittest

    unittest.main()
