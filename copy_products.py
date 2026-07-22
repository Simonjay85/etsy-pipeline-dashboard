#!/usr/bin/env python3
"""
copy_products_v2.py — Nâng cấp với duplicate detection theo tên nguồn.

Tính năng mới:
  ✅ Lưu mapping nguồn → đích vào product_source_map.json
  ✅ Skip sản phẩm đã copy dựa trên tên folder nguồn (không chỉ theo số thứ tự)
  ✅ Auto-scan toàn bộ Plantify source — không hardcode danh sách
  ✅ Auto-detect số product tiếp theo trong từng shop
  ✅ Báo cáo chi tiết: đã copy / bỏ qua / mới

Dùng:
  python3 copy_products_v2.py                  # Copy tất cả shop
  python3 copy_products_v2.py --shop daisy     # Chỉ copy daisyflowdigital
  python3 copy_products_v2.py --dry-run        # Xem trước, không copy thật
  python3 copy_products_v2.py --status         # Xem mapping hiện tại
"""

import argparse
import json
import os
import re
import shutil
import sys
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont

# ── Paths ─────────────────────────────────────────────────────────────────────
SOURCE_ROOT   = Path.home() / "Library/CloudStorage/GoogleDrive-trongnghia0805@gmail.com/My Drive/1.Công việc/1. Etsy/Plantify"
PROJECT_ROOT  = Path("/Users/aaronnguyen/Documents/Claude/Projects/Etsy")
SHOPS_ROOT    = PROJECT_ROOT / "shops"
TEMPLATE_INFO = PROJECT_ROOT / "_TEMPLATE/info.txt"
MAP_FILE      = PROJECT_ROOT / "product_source_map.json"   # ← file mapping mới

# ── Shop configs ──────────────────────────────────────────────────────────────
SHOPS = {
    "daisyflowdigital": {
        "dir":   SHOPS_ROOT / "daisyflowdigital",
        "brand": "DaisyFlow Digital",
    },
    "templystudios": {
        "dir":   SHOPS_ROOT / "templystudios",
        "brand": "TemplyStudios",
    },
}

# Các thư mục nguồn cần bỏ qua
SKIP_CATEGORIES = {"99_duplicates", "99_dup"}

# ── Font ──────────────────────────────────────────────────────────────────────
FONT_PATHS = [
    "/System/Library/Fonts/Helvetica.ttc",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "/System/Library/Fonts/Supplemental/Georgia.ttf",
    "/Library/Fonts/Arial.ttf",
]
_FONT_PATH = next((p for p in FONT_PATHS if os.path.exists(p)), None)


# ─────────────────────────────────────────────────────────────────────────────
# Mapping helpers
# ─────────────────────────────────────────────────────────────────────────────

def load_map() -> dict:
    """
    Cấu trúc map:
    {
      "01_planners/.../49_capricorn_weekly_planner": {
        "daisyflowdigital": "product-83",
        "templystudios":    "product-391"
      },
      ...
    }
    Key = đường dẫn tương đối từ SOURCE_ROOT (category/folder_name).
    """
    if MAP_FILE.exists():
        return json.loads(MAP_FILE.read_text(encoding="utf-8"))
    return {}


def save_map(m: dict):
    MAP_FILE.write_text(json.dumps(m, indent=2, ensure_ascii=False), encoding="utf-8")


def source_key(category: str, folder: str) -> str:
    """Tạo key duy nhất cho mỗi folder nguồn."""
    return f"{category}/{folder}"


def already_copied(m: dict, key: str, shop_id: str) -> str | None:
    """Trả về tên product folder nếu đã copy, None nếu chưa."""
    return m.get(key, {}).get(shop_id)


def record_copy(m: dict, key: str, shop_id: str, product_folder: str):
    if key not in m:
        m[key] = {}
    m[key][shop_id] = product_folder


# ─────────────────────────────────────────────────────────────────────────────
# Source scanner
# ─────────────────────────────────────────────────────────────────────────────

def scan_source() -> list[tuple[str, str]]:
    """
    Quét toàn bộ Plantify source, trả về list (category, folder_name)
    của các folder có đủ files/ VÀ images/.
    """
    results = []
    if not SOURCE_ROOT.exists():
        print(f"❌ Không tìm thấy source root: {SOURCE_ROOT}")
        return results

    for cat_dir in sorted(SOURCE_ROOT.iterdir()):
        if not cat_dir.is_dir() or cat_dir.name.startswith("."):
            continue
        # Bỏ qua các category không hợp lệ
        if any(skip in cat_dir.name.lower() for skip in SKIP_CATEGORIES):
            continue

        for folder_dir in sorted(cat_dir.iterdir()):
            if not folder_dir.is_dir() or folder_dir.name.startswith("."):
                continue

            files_dir  = folder_dir / "files"
            images_dir = folder_dir / "images"

            has_files = files_dir.is_dir() and any(
                f for f in files_dir.iterdir() if not f.name.startswith(".")
            )
            has_imgs = images_dir.is_dir() and any(
                f for f in images_dir.iterdir() if not f.name.startswith(".")
            )

            if has_files and has_imgs:
                results.append((cat_dir.name, folder_dir.name))

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Shop helpers
# ─────────────────────────────────────────────────────────────────────────────

def next_product_number(shop_dir: Path, seen_in_run: set | None = None) -> int:
    """Tự động tìm số product tiếp theo trong shop (kể cả các số đã đặt trong dry-run)."""
    nums = set()
    if shop_dir.exists():
        for d in shop_dir.iterdir():
            m = re.match(r"product-(\d+)$", d.name)
            if m:
                nums.add(int(m.group(1)))
    if seen_in_run:
        nums |= seen_in_run
    return (max(nums) + 1) if nums else 1


def check_excel_entry(excel_path: Path, folder_name: str) -> bool:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        if "Listings" not in wb.sheetnames:
            return False
        ws = wb["Listings"]
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == folder_name:
                return True
    except Exception:
        pass
    return False


def append_excel(excel_path: Path, row_data: dict):
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Listings"]
    target_row = ws.max_row + 1
    for r in range(4, ws.max_row + 2):
        if ws.cell(row=r, column=2).value in (None, ""):
            target_row = r
            break
    ws.cell(row=target_row, column=1,  value=target_row - 3)
    ws.cell(row=target_row, column=2,  value=row_data["folder"])
    ws.cell(row=target_row, column=3,  value=row_data["keywords"])
    ws.cell(row=target_row, column=5,  value=4.99)
    ws.cell(row=target_row, column=6,  value=row_data["category"])
    ws.cell(row=target_row, column=7,  value=row_data["listing_images"])
    ws.cell(row=target_row, column=11, value=999)
    ws.cell(row=target_row, column=12, value="I did")
    ws.cell(row=target_row, column=13, value="2020_2026")
    ws.cell(row=target_row, column=14, value=row_data["status"])
    ws.cell(row=target_row, column=15, value=row_data["section"])
    ws.cell(row=target_row, column=18, value=row_data["sku"])
    wb.save(excel_path)
    return target_row


def clean_keyword(folder_name: str) -> str:
    name = re.sub(r"^\d+_", "", folder_name)
    name = name.replace("_", " ").replace("-", " ")
    name = re.sub(r"\s*(noimage|incomplete)\b", "", name, flags=re.IGNORECASE)
    return " ".join(name.split()).title()


def get_sku(shop_id: str, product_folder: str) -> str:
    prefix = "dd" if "daisy" in shop_id.lower() else "TS"
    clean  = re.sub(r"product[-_]?", "", product_folder, flags=re.IGNORECASE)
    clean  = re.sub(r"[^a-z0-9]", "_", clean.lower())
    clean  = re.sub(r"_+", "_", clean).strip("_")
    return f"{prefix}_{clean}"


# ─────────────────────────────────────────────────────────────────────────────
# Watermark
# ─────────────────────────────────────────────────────────────────────────────

def make_overlay(brand: str) -> Image.Image:
    overlay  = Image.new("RGBA", (2000, 2000), (0, 0, 0, 0))
    draw     = ImageDraw.Draw(overlay)
    wm_size  = 160
    logo_size = 44

    if _FONT_PATH:
        wm_font   = ImageFont.truetype(_FONT_PATH, wm_size)
        logo_font = ImageFont.truetype(_FONT_PATH, logo_size)
    else:
        wm_font = logo_font = ImageFont.load_default()

    # Diagonal watermark
    tw = wm_font.getlength(brand)
    pad = int(tw * 1.5)
    txt_img  = Image.new("RGBA", (pad, pad), (0, 0, 0, 0))
    txt_draw = ImageDraw.Draw(txt_img)
    txt_draw.text(((pad - tw) // 2, (pad - wm_size) // 2),
                  brand, fill=(123, 92, 62, 90), font=wm_font)
    rotated = txt_img.rotate(30, resample=Image.Resampling.BILINEAR, expand=False)
    overlay.paste(rotated, ((2000 - pad) // 2, (2000 - pad) // 2), rotated)

    # Top-right logo
    lw  = logo_font.getlength(brand)
    mx, my = 80, 80
    lx, ly = int(2000 - lw - mx), my
    px, py = 30, 16
    draw.rounded_rectangle([lx-px, ly-py, lx+lw+px, ly+logo_size+py],
                            radius=15, fill=(255, 255, 255, 170))
    draw.text((lx, ly), brand, fill=(123, 92, 62, 230), font=logo_font)
    return overlay


def apply_watermark(src: Path, dst: Path, overlay: Image.Image) -> bool:
    try:
        img   = Image.open(src).convert("RGBA").resize((2000, 2000), Image.Resampling.BICUBIC)
        final = Image.alpha_composite(img, overlay).convert("RGB")
        final.save(dst, "PNG", compress_level=1)
        return True
    except Exception as e:
        print(f"    ⚠️  Watermark lỗi {src.name}: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# Main copy logic
# ─────────────────────────────────────────────────────────────────────────────

def copy_for_shop(
    shop_id: str,
    shop_cfg: dict,
    sources: list[tuple[str, str]],
    src_map: dict,
    dry_run: bool,
) -> dict:
    shop_dir  = shop_cfg["dir"]
    brand     = shop_cfg["brand"]
    excel     = shop_dir / "Etsy_SEO_Generator.xlsx"

    stats = {"new": 0, "skipped_map": 0, "skipped_folder": 0, "error": 0}
    seen_nums: set[int] = set()  # Track numbers used in this run (for dry-run)
    print(f"\n{'='*58}")
    print(f"  🏪 Shop: {shop_id}")
    print(f"  📁 Dir:  {shop_dir}")
    print(f"{'='*58}")

    overlay = None if dry_run else make_overlay(brand)

    for category, folder in sources:
        key = source_key(category, folder)

        # ── Kiểm tra 1: Đã có trong mapping chưa? ────────────────────────
        existing = already_copied(src_map, key, shop_id)
        if existing:
            # Xác nhận folder đích còn tồn tại không
            dst_path = shop_dir / existing
            if dst_path.exists() or dry_run:
                print(f"  ⏭  [MAP]    {folder[:45]:<45} → {existing}")
                stats["skipped_map"] += 1
                # Ghi nhận số đã dùng để dry-run tính đúng
                m = re.match(r"product-(\d+)$", existing)
                if m: seen_nums.add(int(m.group(1)))
                continue
            else:
                print(f"  ⚠️  [MAP]    {existing} trong mapping nhưng folder đã bị xoá → copy lại")

        # ── Kiểm tra 2: Tìm số product tiếp theo ─────────────────────────
        next_num     = next_product_number(shop_dir, seen_in_run=seen_nums)
        product_name = f"product-{next_num}"
        seen_nums.add(next_num)  # Đánh dấu ngay để tránh trùng trong cùng 1 lần chạy
        dst_path  = shop_dir / product_name
        files_dst = dst_path / "files"
        images_dst = dst_path / "images"

        # Kiểm tra folder đích (phòng trường hợp map bị thiếu)
        if dst_path.exists():
            has_f = files_dst.exists() and any(files_dst.iterdir())
            has_i = images_dst.exists() and any(images_dst.iterdir())
            excel_ok = check_excel_entry(excel, product_name)
            if has_f and has_i and excel_ok:
                print(f"  ⏭  [DIR]    {folder[:45]:<45} → {product_name} (đã đầy đủ)")
                record_copy(src_map, key, shop_id, product_name)
                stats["skipped_folder"] += 1
                continue

        src_path  = SOURCE_ROOT / category / folder
        files_src = src_path / "files"
        imgs_src  = src_path / "images"

        print(f"  ✨ [NEW]    {folder[:45]:<45} → {product_name}")

        if dry_run:
            stats["new"] += 1
            continue

        # ── Copy files ─────────────────────────────────────────────────
        try:
            files_dst.mkdir(parents=True, exist_ok=True)
            images_dst.mkdir(parents=True, exist_ok=True)

            # info.txt
            if TEMPLATE_INFO.exists():
                shutil.copy(TEMPLATE_INFO, dst_path / "info.txt")

            # Digital files (PDF/ZIP)
            for f in files_src.iterdir():
                if not f.name.startswith("."):
                    shutil.copy(f, files_dst / f.name)

            # Images — watermark + resize
            img_names = []
            for f in sorted(imgs_src.iterdir()):
                if f.name.startswith(".") or f.suffix.lower() not in {".png",".jpg",".jpeg",".webp"}:
                    continue
                dst_img = images_dst / f"watermarked_{f.name}"
                ok = apply_watermark(f, dst_img, overlay)
                if not ok:
                    shutil.copy(f, dst_img)
                img_names.append(dst_img.name)

            # Excel
            if not check_excel_entry(excel, product_name):
                row = append_excel(excel, {
                    "folder":         product_name,
                    "keywords":       clean_keyword(folder),
                    "category":       "Paper & Party Supplies > Paper > Stationery > Design & Templates > Templates > Planner Templates",
                    "listing_images": "; ".join(img_names),
                    "status":         "⏳ Chờ đăng - NEW",
                    "section":        "Digital Planner",
                    "sku":            get_sku(shop_id, product_name),
                })
                print(f"           Excel row {row} | {len(img_names)} ảnh")

            # Lưu vào mapping
            record_copy(src_map, key, shop_id, product_name)
            save_map(src_map)
            stats["new"] += 1

        except Exception as e:
            print(f"    ❌ Lỗi: {e}")
            stats["error"] += 1

    return stats


# ─────────────────────────────────────────────────────────────────────────────
# Status report
# ─────────────────────────────────────────────────────────────────────────────

def print_status():
    m = load_map()
    if not m:
        print("📭 Mapping trống — chưa có sản phẩm nào được copy.")
        return

    print(f"\n{'─'*65}")
    print(f"  {'Folder nguồn':<48} {'daisy':>8}  {'temply':>8}")
    print(f"{'─'*65}")
    for key in sorted(m.keys()):
        shops_data = m[key]
        folder = key.split("/")[-1][:48]
        daisy  = shops_data.get("daisyflowdigital", "—")
        temply = shops_data.get("templystudios", "—")
        print(f"  {folder:<48} {daisy:>8}  {temply:>8}")
    print(f"{'─'*65}")
    print(f"  Tổng: {len(m)} sản phẩm đã được mapping\n")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Copy Etsy products với duplicate detection")
    parser.add_argument("--shop",    type=str, default=None,
                        help="Chỉ chạy 1 shop (daisy / temply / daisyflowdigital / templystudios)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Xem trước — không copy, không sửa file")
    parser.add_argument("--status",  action="store_true",
                        help="Hiển thị mapping hiện tại rồi thoát")
    args = parser.parse_args()

    if args.status:
        print_status()
        return

    # Chuẩn hoá shop ID
    shop_filter = None
    if args.shop:
        s = args.shop.lower()
        if "daisy" in s:
            shop_filter = "daisyflowdigital"
        elif "temply" in s:
            shop_filter = "templystudios"
        else:
            shop_filter = s

    if args.dry_run:
        print("🔍 DRY-RUN mode — không có gì được thay đổi")

    # Scan source
    print(f"🔍 Đang quét {SOURCE_ROOT.name}...")
    sources = scan_source()
    print(f"   Tìm thấy {len(sources)} folder hợp lệ (có files + images)")

    # Load mapping
    src_map = load_map()
    print(f"   Mapping hiện tại: {len(src_map)} entries trong {MAP_FILE.name}")

    # Chạy
    total_stats = {"new": 0, "skipped_map": 0, "skipped_folder": 0, "error": 0}
    for shop_id, shop_cfg in SHOPS.items():
        if shop_filter and shop_id != shop_filter:
            continue
        stats = copy_for_shop(shop_id, shop_cfg, sources, src_map, dry_run=args.dry_run)
        for k, v in stats.items():
            total_stats[k] += v

    # Tổng kết
    print(f"\n{'='*58}")
    print(f"  ✅ Mới copy:         {total_stats['new']}")
    print(f"  ⏭  Bỏ qua (mapping): {total_stats['skipped_map']}")
    print(f"  ⏭  Bỏ qua (folder):  {total_stats['skipped_folder']}")
    print(f"  ❌ Lỗi:              {total_stats['error']}")
    print(f"{'='*58}")

    if not args.dry_run:
        save_map(src_map)
        print(f"\n💾 Mapping đã lưu → {MAP_FILE}")
    print()


if __name__ == "__main__":
    main()
