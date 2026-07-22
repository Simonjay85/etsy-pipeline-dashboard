"""Excel helpers for Platform_Manager.xlsx (Products + BlogPosts sheets)."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl import Workbook

IMG_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
DIGITAL_EXTS = {".pdf", ".zip", ".001", ".002", ".003", ".004", ".005"}

PRODUCT_COLS = {
    "folder": 2,
    "keywords": 3,
    "price": 5,
    "category": 6,
    "title": 8,
    "description": 9,
    "tags": 10,
    "qty": 11,
    "sku": 13,
    "ebay_status": 14,
    "ebay_url": 15,
    "woo_status": 16,
    "woo_product_id": 17,
    "woo_url": 18,
    "extra": 19,
}

POST_COLS = {
    "slug": 2,
    "title": 3,
    "excerpt": 4,
    "content": 5,
    "categories": 6,
    "tags": 7,
    "featured_image": 8,
    "wp_status": 9,
    "wp_post_id": 10,
    "wp_url": 11,
    "seo_title": 12,
    "seo_description": 13,
}

PRODUCT_HEADERS = [
    "#", "Folder", "Keywords", "", "Price", "Category", "",
    "Title", "Description", "Tags", "Qty", "", "SKU",
    "eBay Status", "eBay URL", "Woo Status", "Woo Product ID", "Woo URL", "Notes",
]

POST_HEADERS = [
    "#", "Slug", "Title", "Excerpt", "Content", "Categories", "Tags",
    "Featured Image", "WP Status", "WP Post ID", "WP URL", "SEO Title", "SEO Description",
]


def set_cell(ws, row: int, col: int, val):
    ws.cell(row=row, column=col).value = val


def ensure_excel(excel_path: Path):
    if excel_path.exists():
        return
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "Products"
    for i, h in enumerate(PRODUCT_HEADERS, 1):
        ws_p.cell(row=3, column=i, value=h)
    ws_b = wb.create_sheet("BlogPosts")
    for i, h in enumerate(POST_HEADERS, 1):
        ws_b.cell(row=3, column=i, value=h)
    wb.save(excel_path)


def generate_sku(site_id: str, folder_name: str) -> str:
    prefix = "".join(c for c in site_id[:3].upper() if c.isalnum()) or "WP"
    clean = re.sub(r"product[-_]?", "", folder_name, flags=re.IGNORECASE)
    clean = re.sub(r"[^a-zA-Z0-9]+", "_", clean).strip("_").lower()
    return f"{prefix}_{clean}"


def products_from_excel(site_dir: Path, excel_path: Path) -> list[dict]:
    ensure_excel(excel_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Products"]
    products = []
    existing = {d.name for d in site_dir.iterdir() if d.is_dir()} if site_dir.exists() else set()

    for row_num in range(4, ws.max_row + 1):
        folder = ws.cell(row=row_num, column=PRODUCT_COLS["folder"]).value
        if not folder:
            continue
        folder = str(folder)
        if folder not in existing:
            continue

        folder_path = site_dir / folder
        img_dir = folder_path / "images"
        file_dir = folder_path / "files"
        images = sorted(
            f.name for f in img_dir.iterdir() if f.suffix.lower() in IMG_EXTS
        ) if img_dir.exists() else []
        dig_files = sorted(
            f.name for f in file_dir.iterdir() if f.suffix.lower() in DIGITAL_EXTS
        ) if file_dir.exists() else []

        title = ws.cell(row=row_num, column=PRODUCT_COLS["title"]).value
        sku = str(ws.cell(row=row_num, column=PRODUCT_COLS["sku"]).value or "")
        if not sku.strip():
            sku = generate_sku(site_dir.name, folder)

        missing = []
        if not title:
            missing.append("title")
        if not str(ws.cell(row=row_num, column=PRODUCT_COLS["description"]).value or "").strip():
            missing.append("description")
        if not images:
            missing.append("images")

        price_raw = ws.cell(row=row_num, column=PRODUCT_COLS["price"]).value
        try:
            price = float(str(price_raw)) if price_raw else 4.99
        except ValueError:
            price = 4.99

        ebay_status = str(ws.cell(row=row_num, column=PRODUCT_COLS["ebay_status"]).value or "pending")
        woo_status = str(ws.cell(row=row_num, column=PRODUCT_COLS["woo_status"]).value or "pending")

        products.append({
            "row": row_num,
            "folder": folder,
            "keywords": str(ws.cell(row=row_num, column=PRODUCT_COLS["keywords"]).value or ""),
            "price": price,
            "category": str(ws.cell(row=row_num, column=PRODUCT_COLS["category"]).value or ""),
            "title": str(title or f"[No title] {folder}"),
            "description": str(ws.cell(row=row_num, column=PRODUCT_COLS["description"]).value or ""),
            "tags": str(ws.cell(row=row_num, column=PRODUCT_COLS["tags"]).value or ""),
            "qty": int(ws.cell(row=row_num, column=PRODUCT_COLS["qty"]).value or 999),
            "sku": sku,
            "ebay_status": ebay_status,
            "ebay_url": str(ws.cell(row=row_num, column=PRODUCT_COLS["ebay_url"]).value or ""),
            "woo_status": woo_status,
            "woo_product_id": str(ws.cell(row=row_num, column=PRODUCT_COLS["woo_product_id"]).value or ""),
            "woo_url": str(ws.cell(row=row_num, column=PRODUCT_COLS["woo_url"]).value or ""),
            "extra": str(ws.cell(row=row_num, column=PRODUCT_COLS["extra"]).value or ""),
            "image_count": len(images),
            "pdf_count": len(dig_files),
            "missing_fields": missing,
            "needs_content": not bool(title),
        })

    def folder_num(name: str) -> int:
        try:
            return int(re.search(r"(\d+)", name).group(1))
        except Exception:
            return 9999

    products.sort(key=lambda p: folder_num(p["folder"]))
    return products


def save_product_row(excel_path: Path, row_num: int, updates: dict):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Products"]
    for field, val in updates.items():
        if field in PRODUCT_COLS:
            set_cell(ws, row_num, PRODUCT_COLS[field], val)
    wb.save(excel_path)


def posts_from_excel(excel_path: Path) -> list[dict]:
    ensure_excel(excel_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["BlogPosts"]
    posts = []
    for row_num in range(4, ws.max_row + 1):
        title = ws.cell(row=row_num, column=POST_COLS["title"]).value
        slug = ws.cell(row=row_num, column=POST_COLS["slug"]).value
        if not title and not slug:
            continue
        posts.append({
            "row": row_num,
            "slug": str(slug or ""),
            "title": str(title or ""),
            "excerpt": str(ws.cell(row=row_num, column=POST_COLS["excerpt"]).value or ""),
            "content": str(ws.cell(row=row_num, column=POST_COLS["content"]).value or ""),
            "categories": str(ws.cell(row=row_num, column=POST_COLS["categories"]).value or ""),
            "tags": str(ws.cell(row=row_num, column=POST_COLS["tags"]).value or ""),
            "featured_image": str(ws.cell(row=row_num, column=POST_COLS["featured_image"]).value or ""),
            "wp_status": str(ws.cell(row=row_num, column=POST_COLS["wp_status"]).value or "draft"),
            "wp_post_id": str(ws.cell(row=row_num, column=POST_COLS["wp_post_id"]).value or ""),
            "wp_url": str(ws.cell(row=row_num, column=POST_COLS["wp_url"]).value or ""),
            "seo_title": str(ws.cell(row=row_num, column=POST_COLS["seo_title"]).value or ""),
            "seo_description": str(ws.cell(row=row_num, column=POST_COLS["seo_description"]).value or ""),
        })
    return posts


def save_post_row(excel_path: Path, row_num: int, updates: dict):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["BlogPosts"]
    for field, val in updates.items():
        if field in POST_COLS:
            set_cell(ws, row_num, POST_COLS[field], val)
    wb.save(excel_path)


def add_product_row(excel_path: Path, data: dict) -> int:
    ensure_excel(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Products"]
    row_num = max(ws.max_row, 3) + 1
    for field, default in {
        "folder": "", "title": "", "description": "", "tags": "",
        "price": 4.99, "qty": 999, "ebay_status": "pending", "woo_status": "pending",
    }.items():
        val = data.get(field, default)
        if field in PRODUCT_COLS:
            set_cell(ws, row_num, PRODUCT_COLS[field], val)
    wb.save(excel_path)
    return row_num


def add_post_row(excel_path: Path, data: dict) -> int:
    ensure_excel(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["BlogPosts"]
    row_num = max(ws.max_row, 3) + 1
    defaults = {
        "slug": "", "title": "", "excerpt": "", "content": "",
        "categories": "", "tags": "", "featured_image": "",
        "wp_status": "draft", "wp_post_id": "", "wp_url": "",
        "seo_title": "", "seo_description": "",
    }
    for field, default in defaults.items():
        val = data.get(field, default)
        if field in POST_COLS:
            set_cell(ws, row_num, POST_COLS[field], val)
    wb.save(excel_path)
    return row_num


def delete_product_row(excel_path: Path, row_num: int):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Products"]
    if row_num >= 4:
        ws.delete_rows(row_num, 1)
    wb.save(excel_path)


def delete_post_row(excel_path: Path, row_num: int):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["BlogPosts"]
    if row_num >= 4:
        ws.delete_rows(row_num, 1)
    wb.save(excel_path)
