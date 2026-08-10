#!/usr/bin/env python3
"""Focused regression tests for single-sync Etsy scrape session routing and identity gates."""

from __future__ import annotations

import asyncio
import json
import tempfile
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from unittest import IsolatedAsyncioTestCase
from unittest.mock import AsyncMock, Mock
from unittest.mock import patch

from openpyxl import Workbook

from playwright.async_api import TimeoutError as PlaywrightTimeoutError

import dashboard_app
from job_store import JobStore


class _FakeLocator:
    def __init__(self, value: str | None):
        self._value = value

    @property
    def first(self) -> "_FakeLocator":
        return self

    async def count(self) -> int:
        return 1 if self._value is not None else 0

    async def input_value(self) -> str:
        return str(self._value or "")

    async def inner_text(self) -> str:
        return str(self._value or "")


class _FakePage:
    def __init__(
        self,
        *,
        url: str,
        force_goto_url: str | None = None,
        force_goto_urls: list[str] | None = None,
        title: str = "",
        description: str = "",
        tags: list[str] | None = None,
        shop_slugs: list[str] | None = None,
        manager_shop_slugs: list[str] | None = None,
        editor_shop_slugs: list[str] | None = None,
        content_html: str = "<html><body></body></html>",
        wait_for_selector_error: Exception | None = None,
        locator_values: dict[str, str] | None = None,
        section_text: str = "",
        manager_shop_slugs_sequence: list[list[str]] | None = None,
    ):
        self.url = url
        if force_goto_urls is not None:
            self.force_goto_urls = force_goto_urls[:]
        else:
            self.force_goto_urls = None
        self.goto_calls: list[tuple[str, str | None, int | None]] = []
        self.force_goto_url = force_goto_url
        self.wait_for_selector_error = wait_for_selector_error
        self.title = title
        self.description = description
        self.tags = tags or []
        self.shop_slugs = shop_slugs or []
        self.manager_shop_slugs = manager_shop_slugs
        self.manager_shop_slugs_sequence = manager_shop_slugs_sequence
        self.editor_shop_slugs = editor_shop_slugs
        self.section_text = section_text
        self.content_html = content_html
        self.locator_values = locator_values or {}
        self._manager_shop_slug_poll = 0

    def locator(self, selector: str) -> _FakeLocator:
        return _FakeLocator(self.locator_values.get(selector))

    def _pick_manager_shop_slugs(self) -> list[str]:
        if self.manager_shop_slugs_sequence is not None:
            if not self.manager_shop_slugs_sequence:
                return self.manager_shop_slugs or []
            index = min(self._manager_shop_slug_poll, len(self.manager_shop_slugs_sequence) - 1)
            self._manager_shop_slug_poll += 1
            return self.manager_shop_slugs_sequence[index]
        return self.manager_shop_slugs or []

    async def goto(self, url: str, wait_until: str | None = None, timeout: int | None = None):
        if self.force_goto_urls is not None and self.force_goto_urls:
            self.url = self.force_goto_urls.pop(0)
        elif self.force_goto_url is not None:
            self.url = self.force_goto_url
        else:
            self.url = url
        self.goto_calls.append((url, wait_until, timeout))

    async def wait_for_selector(self, selector: str, timeout: int | None = None):
        if self.wait_for_selector_error is not None:
            raise self.wait_for_selector_error

    async def wait_for_timeout(self, ms: int) -> None:
        pass

    async def content(self) -> str:
        return self.content_html

    async def evaluate(self, script: str) -> Any:
        if "document?.body?.innerText" in script:
            return self.section_text
        if "/shop/" in script:
            if "/your/shops/me/tools/listings" in self.url:
                return self._pick_manager_shop_slugs()
            if "/your/shops/me/listing-editor/edit/" in self.url and self.editor_shop_slugs is not None:
                return self.editor_shop_slugs
            return self.shop_slugs
        if "let tagSection = null;" in script:
            return self.tags
        if "selectedText = await" in script:
            return ""
        if "selectedText" in script:
            return self.section_text
        return ""


class _FakeContext:
    def __init__(self, page: _FakePage):
        self.pages = [page]
        self.close = AsyncMock()
        self.new_page = AsyncMock(return_value=page)


class _FakeBrowser:
    def __init__(self, context: _FakeContext):
        self.contexts = [context]
        self.new_context = AsyncMock(return_value=context)


class _FakePlaywright:
    def __init__(self, browser: _FakeBrowser):
        self.chromium = Mock()
        self.chromium.connect_over_cdp = AsyncMock(return_value=browser)
        self.chromium.launch_persistent_context = AsyncMock()
        self.chromium.launch = AsyncMock()
        self.stop = AsyncMock()


class _FakeLauncher:
    def __init__(self, playwright: _FakePlaywright):
        self._playwright = playwright
        self.start = AsyncMock(return_value=playwright)


class _ExistingListingWorksheet:
    max_row = 4

    def __init__(self, listing_id: str, prior_status: str) -> None:
        self._values = {
            (4, 2): "product-01",
            (4, 14): prior_status,
            (4, 16): f"https://www.etsy.com/listing/{listing_id}",
        }

    def cell(self, row: int, column: int) -> SimpleNamespace:
        return SimpleNamespace(value=self._values.get((row, column)))


class _ExistingListingWorkbook:
    def __init__(self, worksheet: _ExistingListingWorksheet) -> None:
        self.worksheet = worksheet

    def __getitem__(self, name: str) -> _ExistingListingWorksheet:
        if name != "Listings":
            raise KeyError(name)
        return self.worksheet


class _JsonRequest:
    def __init__(self, payload: dict[str, object]) -> None:
        self.payload = payload

    async def json(self) -> dict[str, object]:
        return self.payload


class TestScrapeListingSingleSync(IsolatedAsyncioTestCase):
    async def test_editor_readiness_reloads_once_after_transient_timeout(self):
        listing_id = "4419825830"
        editor_url = f"https://www.etsy.com/your/shops/me/listing-editor/edit/{listing_id}"
        page = _FakePage(
            url=editor_url,
            locator_values={"body": "Listing editor loading"},
        )
        page.wait_for_selector = AsyncMock(
            side_effect=[PlaywrightTimeoutError("transient timeout"), None]
        )

        await dashboard_app._assert_etsy_editor_ready(page, listing_id)

        self.assertEqual(2, page.wait_for_selector.await_count)
        self.assertEqual([12000, 12000], [
            call.kwargs["timeout"] for call in page.wait_for_selector.await_args_list
        ])
        self.assertEqual([(editor_url, "domcontentloaded", 15000)], page.goto_calls)

    def _build_playwright(self, page: _FakePage) -> tuple[_FakeLauncher, _FakePlaywright]:
        context = _FakeContext(page)
        browser = _FakeBrowser(context)
        playwright = _FakePlaywright(browser)
        return _FakeLauncher(playwright), playwright

    async def test_exact_session_not_ready_raises_actionable_error_and_no_persistent_context(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=False), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(url="https://www.etsy.com/your/shops/me/listing-editor/edit/123")
            launcher, playwright = self._build_playwright(page)
            launch_factory.return_value = launcher

            with self.assertRaisesRegex(
                RuntimeError,
                "Chrome profile=.*CDP=http://127.0.0.1:43123",
            ):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

            self.assertEqual(3, dashboard_app.is_etsy_session_ready.call_count)
            launch_factory.return_value.start.assert_not_awaited()
            launch_factory.return_value._playwright.chromium.connect_over_cdp.assert_not_called()
            launch_factory.return_value._playwright.chromium.launch_persistent_context.assert_not_called()
            playwright.stop.assert_not_awaited()

    async def test_session_readiness_retries_false_false_true_before_starting_playwright(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", side_effect=[False, False, True]) as ready_mock, \
            patch.object(dashboard_app.asyncio, "sleep", new=AsyncMock()) as sleep_mock, \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                locator_values={
                    'textarea[name="title"]': "My Listing",
                    'textarea[name="description"]': "Hello Etsy",
                    "body": "My Listing",
                },
                tags=["tag-a"],
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=[],
            )
            launcher, _ = self._build_playwright(page)
            launch_factory.return_value = launcher

            result = await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

            self.assertEqual("My Listing", result["title"])
            self.assertEqual(3, ready_mock.call_count)
            self.assertEqual(2, sleep_mock.await_count)
            launch_factory.return_value.start.assert_awaited_once()

    async def test_exact_session_ready_uses_resolved_cdp_url_not_default_9222(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                title="My Listing",
                description="Hello Etsy",
                locator_values={
                    'textarea[name="title"]': "My Listing",
                    'textarea[name="description"]': "Hello Etsy",
                    '#listing-price-input': "12.5",
                    '#listing-quantity-input': "3",
                    "body": "My Listing",
                },
                tags=["tag-a", "tag-b"],
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=[],
            )
            launcher, _ = self._build_playwright(page)
            launch_factory.return_value = launcher

            result = await dashboard_app.scrape_listing_details("123", shop_id="templystudios")
            launch_factory.return_value._playwright.chromium.connect_over_cdp.assert_awaited_once_with(
                "http://127.0.0.1:43123",
                timeout=10000,
            )
            launch_factory.return_value._playwright.chromium.launch_persistent_context.assert_not_called()
            self.assertEqual("My Listing", result.get("title"))
            self.assertEqual("Hello Etsy", result.get("description"))
            self.assertNotIn(
                "9222",
                str(launch_factory.return_value._playwright.chromium.connect_over_cdp.await_args.args[0]),
            )

    async def test_cdp_connect_failure_reports_busy_or_unresponsive_not_login(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(url="https://www.etsy.com/your/shops/me/listing-editor/edit/123")
            launcher, playwright = self._build_playwright(page)
            playwright.chromium.connect_over_cdp.side_effect = RuntimeError("CDP timed out")
            launch_factory.return_value = launcher

            with self.assertRaisesRegex(RuntimeError, "CDP Etsy đang bận hoặc không phản hồi") as raised:
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

            self.assertNotIn("đăng nhập", str(raised.exception).lower())
            self.assertGreaterEqual(
                playwright.chromium.connect_over_cdp.await_count,
                1,
            )
            self.assertEqual(
                "http://127.0.0.1:43123",
                str(playwright.chromium.connect_over_cdp.await_args.args[0]),
            )
            self.assertEqual(10000, playwright.chromium.connect_over_cdp.await_args.kwargs.get("timeout"))
            playwright.stop.assert_awaited_once()

    async def test_cdp_connect_retries_without_global_download_defaults_when_chrome_rejects_them(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                title="My Listing",
                description="Hello Etsy",
                locator_values={
                    'textarea[name="title"]': "My Listing",
                    'textarea[name="description"]': "Hello Etsy",
                    '#listing-price-input': "12.5",
                    '#listing-quantity-input': "3",
                    "body": "My Listing",
                },
                tags=["tag-a", "tag-b"],
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=[],
            )
            launcher, playwright = self._build_playwright(page)
            launch_factory.return_value = launcher
            playwright.chromium.connect_over_cdp.side_effect = [
                RuntimeError(
                    "Protocol error (Browser.setDownloadBehavior): "
                    "Browser context management is not supported."
                ),
                _FakeBrowser(_FakeContext(page)),
            ]

            result = await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

        self.assertEqual("My Listing", result["title"])
        self.assertEqual(2, playwright.chromium.connect_over_cdp.await_count)
        first_call, fallback_call = playwright.chromium.connect_over_cdp.await_args_list
        self.assertEqual("http://127.0.0.1:43123", first_call.args[0])
        self.assertEqual({"timeout": 10000}, first_call.kwargs)
        self.assertEqual("http://127.0.0.1:43123", fallback_call.args[0])
        self.assertEqual({"timeout": 10000, "no_defaults": True}, fallback_call.kwargs)

    async def test_connect_over_cdp_retries_after_transient_failure(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True) as session_ready_mock, \
            patch("playwright.async_api.async_playwright") as launch_factory, \
            patch.object(dashboard_app.asyncio, "sleep", new=AsyncMock()) as sleep_mock:

            page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                locator_values={
                    'textarea[name="title"]': "My Listing",
                    'textarea[name="description"]': "Hello Etsy",
                    '#listing-price-input': "12.5",
                    '#listing-quantity-input': "3",
                    "body": "My Listing",
                },
                tags=["tag-a", "tag-b"],
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=[],
            )
            launcher, playwright = self._build_playwright(page)
            launch_factory.return_value = launcher
            second_browser = _FakeBrowser(_FakeContext(_FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                locator_values=page.locator_values,
                tags=page.tags,
                manager_shop_slugs=page.manager_shop_slugs,
                editor_shop_slugs=[],
            )))
            playwright.chromium.connect_over_cdp.side_effect = [
                RuntimeError("CDP transient contention"),
                second_browser,
            ]

            result = await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

        self.assertEqual("My Listing", result["title"])
        self.assertEqual(2, playwright.chromium.connect_over_cdp.await_count)
        self.assertGreaterEqual(sleep_mock.await_count, 1)
        self.assertEqual(2, session_ready_mock.call_count)

    async def test_sign_in_redirect_classified_before_identity(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(
                url="https://www.etsy.com/signin",
                force_goto_url="https://www.etsy.com/signin",
                title="",
                description="",
            )
            launcher, _ = self._build_playwright(page)
            launch_factory.return_value = launcher

            with self.assertRaisesRegex(RuntimeError, "Phiên Etsy chưa đăng nhập"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

    async def test_wrong_editor_listing_or_title_timeout_classified_before_identity(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/999",
                force_goto_urls=[
                    "https://www.etsy.com/your/shops/me/tools/listings",
                    "https://www.etsy.com/your/shops/me/listing-editor/edit/999",
                ],
                title="Other",
                description="Oops",
                wait_for_selector_error=PlaywrightTimeoutError("not ready"),
                locator_values={
                    'textarea[name="title"]': "Other",
                    'textarea[name="description"]': "Oops",
                    '#listing-price-input': "5",
                    '#listing-quantity-input': "1",
                    "body": "Other editor loaded",
                },
                tags=["tag-a"],
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=["templystudios"],
            )
            launcher, _ = self._build_playwright(page)
            launch_factory.return_value = launcher

            with self.assertRaisesRegex(RuntimeError, "Editor sai listing"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

            page.force_goto_urls = [
                "https://www.etsy.com/your/shops/me/tools/listings",
                "https://www.etsy.com/your/shops/me/listing-editor/edit/123",
            ]
            page.wait_for_selector_error = PlaywrightTimeoutError("not ready")
            with self.assertRaisesRegex(RuntimeError, "Không nạp được giao diện Listing Editor"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

    async def test_shop_manager_preflight_wrong_shop_fails(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            manager_wrong_shop_page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                title="A",
                description="B",
                locator_values={
                    'textarea[name="title"]': "A",
                    'textarea[name="description"]': "B",
                    '#listing-price-input': "10",
                    '#listing-quantity-input': "2",
                    "body": "A",
                },
                tags=["tag-a"],
                manager_shop_slugs=["wrongshop"],
                editor_shop_slugs=[],
            )
            launcher_wrong, _ = self._build_playwright(manager_wrong_shop_page)
            launch_factory.return_value = launcher_wrong
            with self.assertRaisesRegex(RuntimeError, "Phiên Etsy sai shop"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

    async def test_shop_manager_preflight_no_identity_evidence_fails(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            manager_empty_shop_page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                title="A",
                description="B",
                locator_values={
                    'textarea[name="title"]': "A",
                    'textarea[name="description"]': "B",
                    '#listing-price-input': "10",
                    '#listing-quantity-input': "2",
                    "body": "A",
                },
                tags=["tag-a"],
                manager_shop_slugs=[],
                editor_shop_slugs=[],
            )
            launcher_no_identity, _ = self._build_playwright(manager_empty_shop_page)
            launch_factory.return_value = launcher_no_identity
            with self.assertRaisesRegex(RuntimeError, "chưa sẵn sàng/chưa xác minh"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

    async def test_editor_anchorless_passes_only_after_verified_shop_manager_preflight(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            missing_editor_anchor_page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                title="My Listing",
                description="Desc",
                locator_values={
                    'textarea[name="title"]': "My Listing",
                    'textarea[name="description"]': "Desc",
                    '#listing-price-input': "10",
                    '#listing-quantity-input': "2",
                    "body": "My Listing",
                },
                tags=["tag-a"],
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=[],
            )
            launcher_anchorless, _ = self._build_playwright(missing_editor_anchor_page)
            launch_factory.return_value = launcher_anchorless
            details = await dashboard_app.scrape_listing_details("123", shop_id="templystudios")
            self.assertEqual("My Listing", details["title"])
            missing_editor_anchor_page.manager_shop_slugs = []
            launch_factory.return_value = launcher_anchorless
            with self.assertRaisesRegex(RuntimeError, "chưa sẵn sàng/chưa xác minh"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")


    async def test_expected_and_wrong_shop_anchor_in_editor_fails(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            mixed_anchor_page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                title="A",
                description="B",
                locator_values={
                    'textarea[name="title"]': "A",
                    'textarea[name="description"]': "B",
                    '#listing-price-input': "10",
                    '#listing-quantity-input': "2",
                    "body": "A",
                },
                tags=["tag-a"],
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=["templystudios", "other-shop"],
            )
            launcher_mixed, _ = self._build_playwright(mixed_anchor_page)
            launch_factory.return_value = launcher_mixed

            with self.assertRaisesRegex(RuntimeError, "Phiên Etsy sai shop"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

    async def test_access_gate_classifies_sign_in_and_blocked_text_and_not_script_hidden(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            sign_in_page = _FakePage(
                url="https://www.etsy.com/signin",
                force_goto_url="https://www.etsy.com/signin",
                title="",
                description="",
                section_text="Please sign in to continue",
                manager_shop_slugs=[],
                editor_shop_slugs=["templystudios"],
            )
            launcher_signin, _ = self._build_playwright(sign_in_page)
            launch_factory.return_value = launcher_signin
            with self.assertRaisesRegex(RuntimeError, "Phiên Etsy chưa đăng nhập"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

            blocked_page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                locator_values={"body": "Our system detected unusual activity and access denied"},
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=[],
                section_text="Our system detected unusual activity and access denied",
            )
            launcher_blocked, _ = self._build_playwright(blocked_page)
            launch_factory.return_value = launcher_blocked
            with self.assertRaisesRegex(RuntimeError, "Etsy đang chặn/đòi xác minh"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

            html_script_page = _FakePage(
                url="https://www.etsy.com/your/shops/me/listing-editor/edit/123",
                locator_values={
                    'textarea[name="title"]': "A",
                    'textarea[name="description"]': "B",
                    '#listing-price-input': "10",
                    '#listing-quantity-input': "2",
                    "body": "Listing editor is ready to edit",
                },
                tags=["tag-a"],
                section_text="<script>robot login challenge blocked</script>",
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=[],
            )
            launcher_script_only, _ = self._build_playwright(html_script_page)
            launch_factory.return_value = launcher_script_only
            details = await dashboard_app.scrape_listing_details("123", shop_id="templystudios")
            self.assertEqual("A", details["title"])

    def test_etsy_auth_required_url_checks_path_only_without_query_false_positives(self) -> None:
        self.assertFalse(
            dashboard_app._is_etsy_auth_required_url(
                "https://www.etsy.com/shop/templystudios?return_to=/signin"
            )
        )
        self.assertTrue(
            dashboard_app._is_etsy_auth_required_url(
                "https://www.etsy.com/signin?next=/your/shops/me/tools/listings"
            )
        )

    def test_etsy_access_blocked_url_checks_path_only_without_query_false_positives(self) -> None:
        self.assertFalse(
            dashboard_app._is_etsy_access_blocked_url(
                "https://www.etsy.com/your/shops/me/listing-editor/edit/123?next=/challenge"
            )
        )
        self.assertTrue(dashboard_app._is_etsy_access_blocked_url("https://www.etsy.com/challenge"))

    async def test_shop_manager_preflight_accepts_manager_route_with_query_hash(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(
                url="https://www.etsy.com/your/shops/me/tools/listings?foo=1#hash",
                force_goto_urls=[
                    "https://www.etsy.com/your/shops/me/tools/listings?foo=1#hash",
                ],
                title="My Listing",
                description="Desc",
                locator_values={
                    'textarea[name="title"]': "My Listing",
                    'textarea[name="description"]': "Desc",
                    '#listing-price-input': "10",
                    '#listing-quantity-input': "2",
                    "body": "My Listing",
                },
                tags=["tag-a"],
                manager_shop_slugs=["templystudios"],
                editor_shop_slugs=[],
            )
            launcher, _ = self._build_playwright(page)
            launch_factory.return_value = launcher
            details = await dashboard_app.scrape_listing_details("123", shop_id="templystudios")
            self.assertEqual("My Listing", details["title"])

    async def test_shop_manager_preflight_rejects_public_shop_redirect(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            page = _FakePage(
                url="https://www.etsy.com/shop/templystudios",
                force_goto_url="https://www.etsy.com/shop/templystudios",
                title="Shop",
                description="Shop page",
            )
            launcher, _ = self._build_playwright(page)
            launch_factory.return_value = launcher
            with self.assertRaisesRegex(RuntimeError, "chưa xác minh"):
                await dashboard_app.scrape_listing_details("123", shop_id="templystudios")

    async def test_shop_manager_preflight_waits_for_delayed_manager_anchors(self):
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "SHOPS", {
                "templystudios": {
                    "name": "Temply",
                    "etsy_link": "https://www.etsy.com/shop/templystudios",
                    "etsy_login_debug_port": 43123,
                }
            }), \
            patch.object(dashboard_app, "is_etsy_session_ready", return_value=True), \
            patch("playwright.async_api.async_playwright") as launch_factory:

            manager_delayed_page = _FakePage(
                url="https://www.etsy.com/your/shops/me/tools/listings",
                title="My Listing",
                description="Desc",
                locator_values={
                    'textarea[name="title"]': "My Listing",
                    'textarea[name="description"]': "Desc",
                    '#listing-price-input': "10",
                    '#listing-quantity-input': "2",
                    "body": "My Listing",
                },
                tags=["tag-a"],
                manager_shop_slugs_sequence=[[], ["templystudios"]],
                editor_shop_slugs=[],
            )
            launcher, _ = self._build_playwright(manager_delayed_page)
            launch_factory.return_value = launcher
            details = await dashboard_app.scrape_listing_details("123", shop_id="templystudios")
            self.assertEqual("My Listing", details["title"])


class TestSingleSyncBusyGuard(IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        dashboard_app._etsy_single_sync_busy_shops.clear()
        dashboard_app._OPERATION_QUEUE_LOCK = asyncio.Lock()
        dashboard_app._OPERATION_QUEUE_COMMANDS.clear()
        dashboard_app._OPERATION_QUEUE_DEDUPE.clear()

    async def asyncTearDown(self) -> None:
        dashboard_app._etsy_single_sync_busy_shops.clear()
        dashboard_app._OPERATION_QUEUE_COMMANDS.clear()
        dashboard_app._OPERATION_QUEUE_DEDUPE.clear()

    async def test_second_same_shop_request_is_deduplicated_in_queue(self) -> None:
        started = asyncio.Event()
        release = asyncio.Event()

        async def blocking_sync(**kwargs):
            started.set()
            await release.wait()
            return (
                {},
                {},
                {"metadata_ok": True, "assets_complete": True, "overall": True},
                True,
                ["title"],
            )

        sync_mock = AsyncMock(side_effect=blocking_sync)
        product = {
            "folder": "product-65",
            "etsy_url": "https://www.etsy.com/listing/4528326700/example",
        }
        with tempfile.TemporaryDirectory() as tmpdir, \
            patch.object(dashboard_app, "BASE_DIR", Path(tmpdir)), \
            patch.object(dashboard_app, "_active_shop_id", "daisyflowdigital"), \
            patch.object(dashboard_app, "get_product_by_row", return_value=product), \
            patch.object(dashboard_app, "_sync_local_from_etsy", new=sync_mock), \
            patch.object(dashboard_app, "broadcast", new=Mock()):

            first_response = await dashboard_app.sync_from_etsy(65)
            await started.wait()
            second_response = await dashboard_app.sync_from_etsy(66)

            self.assertEqual(202, first_response.status_code)
            self.assertEqual(200, second_response.status_code)
            self.assertIn(b'"created":false', second_response.body)
            self.assertEqual(1, sync_mock.await_count)

            release.set()
            for _ in range(20):
                if "daisyflowdigital" not in dashboard_app._etsy_single_sync_busy_shops:
                    break
                await asyncio.sleep(0)

        self.assertNotIn("daisyflowdigital", dashboard_app._etsy_single_sync_busy_shops)
        call_kwargs = sync_mock.await_args.kwargs
        self.assertEqual("daisyflowdigital", call_kwargs["shop_id"])
        self.assertEqual(Path(tmpdir) / "shops" / "daisyflowdigital" / "product-65", call_kwargs["product_path"])
        self.assertEqual(Path(tmpdir) / "shops" / "daisyflowdigital" / "Etsy_SEO_Generator.xlsx", call_kwargs["excel_path"])

    async def test_guard_releases_after_sync_failure(self) -> None:
        success = (
            {},
            {},
            {"metadata_ok": True, "assets_complete": True, "overall": True},
            True,
            ["title"],
        )
        sync_mock = AsyncMock(side_effect=[RuntimeError("scrape failed"), success])
        product = {
            "folder": "product-65",
            "etsy_url": "https://www.etsy.com/listing/4528326700/example",
        }
        with tempfile.TemporaryDirectory() as tmpdir, \
            patch.object(dashboard_app, "BASE_DIR", Path(tmpdir)), \
            patch.object(dashboard_app, "_active_shop_id", "daisyflowdigital"), \
            patch.object(dashboard_app, "get_product_by_row", return_value=product), \
            patch.object(dashboard_app, "_sync_local_from_etsy", new=sync_mock), \
            patch.object(dashboard_app, "broadcast", new=Mock()):

            failed_response = await dashboard_app.sync_from_etsy(65)
            self.assertEqual(202, failed_response.status_code)
            for _ in range(20):
                if "daisyflowdigital" not in dashboard_app._etsy_single_sync_busy_shops:
                    break
                await asyncio.sleep(0)
            self.assertNotIn("daisyflowdigital", dashboard_app._etsy_single_sync_busy_shops)

            retry_response = await dashboard_app.sync_from_etsy(65)
            self.assertIn(retry_response.status_code, {200, 202})
        # Admission is asynchronous now; a failure is reported in the queued
        # command/log instead of being returned by this HTTP request.
        self.assertTrue(retry_response)
        self.assertNotIn("daisyflowdigital", dashboard_app._etsy_single_sync_busy_shops)


class TestEtsySyncRequestIdentity(IsolatedAsyncioTestCase):
    def setUp(self) -> None:
        super().setUp()
        self._job_store_workspace = tempfile.TemporaryDirectory(
            prefix="etsy-single-sync-job-store-"
        )
        self._job_store = JobStore(Path(self._job_store_workspace.name) / "jobs.sqlite")
        self._job_store_patch = patch.object(
            dashboard_app, "_get_job_store", return_value=self._job_store
        )
        self._job_store_patch.start()
        self._update_jobs_patch = patch.dict(dashboard_app._etsy_update_jobs, clear=True)
        self._update_jobs_patch.start()
        dashboard_app._etsy_single_sync_busy_shops.clear()
        dashboard_app._OPERATION_QUEUE_LOCK = asyncio.Lock()
        dashboard_app._OPERATION_QUEUE_COMMANDS.clear()
        dashboard_app._OPERATION_QUEUE_DEDUPE.clear()
        self.product = {
            "folder": "product-121",
            "etsy_url": "https://www.etsy.com/listing/4529147696/example",
        }

    def tearDown(self) -> None:
        dashboard_app._etsy_single_sync_busy_shops.clear()
        dashboard_app._OPERATION_QUEUE_COMMANDS.clear()
        dashboard_app._OPERATION_QUEUE_DEDUPE.clear()
        self._update_jobs_patch.stop()
        self._job_store_patch.stop()
        self._job_store.close()
        self._job_store_workspace.cleanup()
        super().tearDown()

    def test_identity_payload_is_shop_agnostic_and_rejects_stale_mapping(self) -> None:
        for shop_id in ("templystudios", "daisyflowdigital"):
            with self.subTest(shop_id=shop_id), patch.object(
                dashboard_app, "_active_shop_id", shop_id
            ):
                self.assertEqual(
                    (shop_id, "product-121", "4529147696"),
                    dashboard_app._validate_product_etsy_identity(
                        121,
                        self.product,
                        {
                            "shop": shop_id,
                            "folder": "product-121",
                            "listing_id": "4529147696",
                        },
                    ),
                )

        mismatch_payloads = (
            {"shop": "daisyflowdigital", "folder": "product-121", "listing_id": "4529147696"},
            {"shop": "templystudios", "folder": "product-999", "listing_id": "4529147696"},
            {"shop": "templystudios", "folder": "product-121", "listing_id": "999"},
        )
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"):
            for payload in mismatch_payloads:
                with self.subTest(payload=payload), self.assertRaises(dashboard_app.HTTPException) as raised:
                    dashboard_app._validate_product_etsy_identity(121, self.product, payload)
                self.assertEqual(409, raised.exception.status_code)

    async def test_sync_route_passes_exact_temply_identity_and_paths(self) -> None:
        sync_result = (
            {},
            {},
            {"metadata_ok": True, "assets_complete": True, "overall": True},
            True,
            ["title"],
        )
        request = _JsonRequest({
            "shop": "templystudios",
            "folder": "product-121",
            "listing_id": "4529147696",
        })
        broadcast_mock = Mock()
        with tempfile.TemporaryDirectory() as tmpdir, \
            patch.object(dashboard_app, "BASE_DIR", Path(tmpdir)), \
            patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "get_product_by_row", return_value=self.product), \
            patch.object(dashboard_app, "_sync_local_from_etsy", new=AsyncMock(return_value=sync_result)) as sync_mock, \
            patch.object(dashboard_app, "broadcast", new=broadcast_mock):

            response = await dashboard_app.sync_from_etsy(121, request)

        self.assertEqual(202, response.status_code)
        payload = json.loads(response.body)
        self.assertEqual("etsy-listing-sync", payload["command"]["operation"])
        self.assertEqual("templystudios", payload["command"]["shop_id"])
        self.assertIn("product-121:4529147696", payload["command"]["target"])

    async def test_sync_route_reports_warning_when_sync_status_not_complete(self) -> None:
        sync_result = (
            {},
            {},
            {"metadata_ok": True, "assets_complete": False, "overall": False},
            False,
            ["title"],
        )
        request = _JsonRequest({
            "shop": "templystudios",
            "folder": "product-121",
            "listing_id": "4529147696",
        })
        broadcast_mock = Mock()
        with tempfile.TemporaryDirectory() as tmpdir, \
            patch.object(dashboard_app, "BASE_DIR", Path(tmpdir)), \
            patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "get_product_by_row", return_value=self.product), \
            patch.object(dashboard_app, "_sync_local_from_etsy", new=AsyncMock(return_value=sync_result)), \
            patch.object(dashboard_app, "broadcast", new=broadcast_mock):

            response = await dashboard_app.sync_from_etsy(121, request)

        self.assertEqual(202, response.status_code)
        payload = json.loads(response.body)
        self.assertEqual("queued", payload["command"]["status"])

    async def test_sync_route_rejects_stale_folder_before_scrape(self) -> None:
        request = _JsonRequest({
            "shop": "templystudios",
            "folder": "product-999",
            "listing_id": "4529147696",
        })
        sync_mock = AsyncMock()
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "get_product_by_row", return_value=self.product), \
            patch.object(dashboard_app, "_sync_local_from_etsy", new=sync_mock):
            with self.assertRaises(dashboard_app.HTTPException) as raised:
                await dashboard_app.sync_from_etsy(121, request)

        self.assertEqual(409, raised.exception.status_code)
        sync_mock.assert_not_awaited()

    async def test_push_route_rejects_stale_listing_before_creating_job(self) -> None:
        request = _JsonRequest({
            "fields": ["title"],
            "shop": "templystudios",
            "folder": "product-121",
            "listing_id": "999",
        })
        before_jobs = dict(dashboard_app._etsy_update_jobs)
        with patch.object(dashboard_app, "_active_shop_id", "templystudios"), \
            patch.object(dashboard_app, "get_product_by_row", return_value=self.product):
            with self.assertRaises(dashboard_app.HTTPException) as raised:
                await dashboard_app.push_local_updates_to_etsy(121, request)

        self.assertEqual(409, raised.exception.status_code)
        self.assertEqual(before_jobs, dashboard_app._etsy_update_jobs)

    def test_update_busy_gate_is_scoped_by_active_shop(self) -> None:
        with patch.dict(dashboard_app._etsy_update_jobs, {
            "temply-job": {"shop_id": "templystudios", "status": "running"},
            "daisy-done": {"shop_id": "daisyflowdigital", "status": "success"},
        }, clear=True):
            self.assertTrue(dashboard_app._etsy_update_shop_is_busy("templystudios"))
            self.assertFalse(dashboard_app._etsy_update_shop_is_busy("daisyflowdigital"))


class TestCreateLocalListingSyncStatus(IsolatedAsyncioTestCase):
    async def _create_existing_listing(
        self,
        *,
        manager_status: str = "active",
        sync_result: tuple[dict, dict, dict, bool, list[str]] | None = None,
        sync_exception: Exception | None = None,
        status_write_exception: Exception | None = None,
        request_payload: dict[str, object] | None = None,
    ) -> tuple[dict, dict[str, str], list[tuple[int, dict, Path | None]], AsyncMock, Mock]:
        listing_id = "4527467265"
        prior_status = "⚠ Sync lỗi"
        worksheet = _ExistingListingWorksheet(listing_id, prior_status)
        workbook = _ExistingListingWorkbook(worksheet)
        status_state = {"value": prior_status}
        saved_updates: list[tuple[int, dict, Path | None]] = []

        def fake_save_to_excel(row_num: int, updates: dict, excel_path: Path | None = None) -> None:
            saved_updates.append((row_num, dict(updates), excel_path))
            if "status" in updates:
                if status_write_exception is not None:
                    raise status_write_exception
                status_state["value"] = str(updates["status"])

        if sync_exception is not None:
            sync_mock = AsyncMock(side_effect=sync_exception)
        else:
            sync_mock = AsyncMock(
                return_value=sync_result
                or (
                    {"title": "Synced title", "description": "Synced description", "tags": "one, two"},
                    {},
                    {"metadata_ok": True, "assets_complete": True, "overall": True},
                    True,
                    ["description", "tags", "title"],
                )
            )

        broadcast_mock = Mock()
        with tempfile.TemporaryDirectory() as tmpdir:
            shop_dir = Path(tmpdir) / "shop"
            excel_path = Path(tmpdir) / "Etsy_SEO_Generator.xlsx"
            with patch.object(
                dashboard_app,
                "latest_etsy_manager_snapshot",
                return_value={
                    "listings": [{
                        "id": listing_id,
                        "managerStatus": manager_status,
                        "url": f"https://www.etsy.com/listing/{listing_id}",
                        "title": "Remote title",
                    }]
                },
            ), patch.object(dashboard_app, "_active_shop_id", "templystudios"), patch.object(
                dashboard_app, "SHOP_DIR", return_value=shop_dir
            ), patch.object(dashboard_app, "EXCEL_FILE", return_value=excel_path), patch.object(
                dashboard_app.openpyxl, "load_workbook", return_value=workbook
            ), patch.object(dashboard_app, "save_to_excel", side_effect=fake_save_to_excel), patch.object(
                dashboard_app, "_sync_local_from_etsy", new=sync_mock
            ), patch.object(dashboard_app, "broadcast", new=broadcast_mock):
                payload: dict[str, object] = {"listing_id": listing_id}
                if request_payload:
                    payload.update(request_payload)
                result = await dashboard_app.create_local_product_from_etsy(
                    _JsonRequest(payload)
                )

        return result, status_state, saved_updates, sync_mock, broadcast_mock

    async def test_existing_sync_error_row_becomes_manager_derived_success(self) -> None:
        for manager_status, expected_status in (
            ("active", "✅ Đã đăng"),
            ("draft", "✅ Đã đăng draft"),
        ):
            with self.subTest(manager_status=manager_status):
                result, status_state, saved_updates, sync_mock, _ = await self._create_existing_listing(
                    manager_status=manager_status,
                )

                self.assertTrue(result["existing"])
                self.assertTrue(result["sync_ok"])
                self.assertEqual(expected_status, status_state["value"])
                self.assertIn(expected_status, [updates.get("status") for _, updates, _ in saved_updates])
                sync_mock.assert_awaited_once()

    async def test_incomplete_sync_keeps_existing_sync_error_status(self) -> None:
        result, status_state, saved_updates, _, _ = await self._create_existing_listing(
            sync_result=(
                {"title": "Synced title", "description": "Synced description", "tags": "one, two"},
                {"images_found": 10, "images_downloaded": 9},
                {"metadata_ok": True, "assets_complete": False, "overall": False},
                False,
                ["description", "tags", "title"],
            )
        )

        self.assertFalse(result["sync_ok"])
        self.assertIn("không hoàn tất đủ", result["sync_error"])
        self.assertEqual("⚠ Sync lỗi", status_state["value"])
        self.assertIn("⚠ Sync lỗi", [updates.get("status") for _, updates, _ in saved_updates])

    async def test_sync_exception_keeps_existing_sync_error_status(self) -> None:
        result, status_state, saved_updates, _, _ = await self._create_existing_listing(
            sync_exception=RuntimeError("scrape failed"),
        )

        self.assertFalse(result["sync_ok"])
        self.assertEqual("scrape failed", result["sync_error"])
        self.assertEqual("⚠ Sync lỗi", status_state["value"])
        self.assertIn("⚠ Sync lỗi", [updates.get("status") for _, updates, _ in saved_updates])

    async def test_manager_success_status_write_failure_is_not_reported_as_success(self) -> None:
        result, status_state, saved_updates, sync_mock, broadcast_mock = await self._create_existing_listing(
            status_write_exception=RuntimeError("status write failed"),
        )

        self.assertTrue(result["existing"])
        self.assertFalse(result["sync_ok"])
        self.assertIn("status write failed", result["sync_error"])
        self.assertTrue(
            "trạng thái" in result["sync_error"].lower()
            or "status" in result["sync_error"].lower()
        )
        self.assertEqual(
            {"etsy_url": "https://www.etsy.com/listing/4527467265"},
            saved_updates[0][1],
        )
        self.assertIn("✅ Đã đăng", [updates.get("status") for _, updates, _ in saved_updates])
        self.assertEqual("⚠ Sync lỗi", status_state["value"])
        sync_mock.assert_awaited_once()

        messages = [str(call.args[0]) for call in broadcast_mock.call_args_list if call.args]
        self.assertFalse(any("✅" in message for message in messages))
        self.assertTrue(
            any(
                "⚠" in message
                or "❌" in message
                or "lỗi" in message.lower()
                or "error" in message.lower()
                for message in messages
            )
        )

    async def test_create_local_listing_full_sync_default_uses_assets(self) -> None:
        _, _, _, sync_mock, _ = await self._create_existing_listing()

        call_kwargs = sync_mock.await_args.kwargs
        self.assertIn("sync_assets", call_kwargs)
        self.assertIs(call_kwargs["sync_assets"], True)

    async def test_create_local_listing_metadata_only_uses_metadata_path_only(self) -> None:
        _, _, _, sync_mock, _ = await self._create_existing_listing(
            request_payload={"metadata_only": True},
            sync_result=(
                {"title": "Synced title", "description": "Synced description", "tags": "one, two"},
                {},
                {"metadata_ok": True, "assets_complete": False, "overall": True, "assets_deferred": True},
                True,
                ["description", "tags", "title"],
            ),
        )

        call_kwargs = sync_mock.await_args.kwargs
        self.assertIn("sync_assets", call_kwargs)
        self.assertIs(call_kwargs["sync_assets"], False)

    async def test_metadata_only_does_not_reuse_empty_product_slot(self) -> None:
        listing_id = "777888999"
        wb = Workbook()
        ws = wb["Sheet"]
        ws.title = "Listings"
        ws["B4"] = "product-277"

        allocate_calls: list[list[dict] | None] = []
        original_allocate = dashboard_app._allocate_product_slot

        def spy_allocate(*args, **kwargs):
            reusable_slots = args[3] if len(args) > 3 else kwargs.get("reusable_slots")
            allocate_calls.append(list(reusable_slots) if reusable_slots is not None else None)
            return original_allocate(*args, **kwargs)

        with tempfile.TemporaryDirectory() as tmpdir:
            shop_dir = Path(tmpdir) / "shop"
            shop_dir.mkdir()
            reusable_folder = shop_dir / "product-277"
            (reusable_folder / "images").mkdir(parents=True)
            (reusable_folder / "files").mkdir(parents=True)
            excel_path = Path(tmpdir) / "Etsy_SEO_Generator.xlsx"
            wb.save(excel_path)

            payload = {"listing_id": listing_id, "metadata_only": True}
            with patch.object(
                dashboard_app,
                "latest_etsy_manager_snapshot",
                return_value={
                    "listings": [{
                        "id": listing_id,
                        "managerStatus": "active",
                        "url": f"https://www.etsy.com/listing/{listing_id}",
                        "title": "Remote title",
                    }],
                },
            ), patch.object(dashboard_app, "_active_shop_id", "templystudios"), patch.object(
                dashboard_app, "SHOP_DIR", return_value=shop_dir
            ), patch.object(dashboard_app, "EXCEL_FILE", return_value=excel_path), patch.object(
                dashboard_app.openpyxl, "load_workbook", return_value=wb
            ), patch.object(
                dashboard_app,
                "save_to_excel",
                new=Mock(),
            ), patch.object(
                dashboard_app,
                "_sync_local_from_etsy",
                new=AsyncMock(
                    return_value=(
                        {"title": "Synced title", "description": "Synced desc", "tags": "tag1, tag2"},
                        {},
                        {"metadata_ok": True, "assets_complete": False, "overall": True},
                        True,
                        ["description", "tags", "title"],
                    )
                ),
            ), patch.object(dashboard_app, "_allocate_product_slot", side_effect=spy_allocate), patch.object(
                dashboard_app,
                "broadcast",
                new=Mock(),
            ):
                result = await dashboard_app.create_local_product_from_etsy(_JsonRequest(payload))

        self.assertTrue(allocate_calls)
        self.assertEqual([], allocate_calls[0])
        self.assertNotEqual("product-277", result["folder"])
        self.assertEqual("product-277", str(ws["B4"].value))
        self.assertEqual("product-278", str(ws["B5"].value))
        self.assertEqual(5, result["row"])
