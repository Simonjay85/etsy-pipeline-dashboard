"""
Etsy Pipeline Dashboard — port 8090
Quản lý & điều khiển pipeline tạo/đăng Etsy listing.
Chạy: python3 dashboard_app.py
"""
import asyncio, hashlib, html, io, json, logging, os, re, secrets, shutil, socket, subprocess, sys, tempfile, time, urllib.parse
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
# httpx lazy-imported to avoid hanging at startup on macOS (SSL init issue)
_HAS_HTTPX = True  # assume available; will fail gracefully if not
httpx = None       # lazy-loaded on first use
from pathlib import Path
from difflib import SequenceMatcher
from typing import Any, Callable, Mapping, NoReturn, Optional, cast

import runtime_identity
import openpyxl
from job_store import (
    JOB_STATUS_CANCELLED,
    JOB_STATUS_FAILED,
    JOB_STATUS_QUEUED,
    JOB_STATUS_RUNNING,
    JOB_STATUS_SUCCEEDED,
    JobStore,
)
from operation_context import OperationContext, OperationContextError
from fastapi import FastAPI, File, HTTPException, Request, UploadFile, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from etsy_catalog import build_unified_catalog, merge_safe_duplicates, normalize_etsy_manager_snapshot
from shops_config_validation import (
    DEFAULT_EXAMPLE_NAME,
    ShopsConfigError,
    load_shops_config as validate_and_load_shops_config,
)
from shop_asset_workflow import copy_image_with_watermark, get_watermark_text
from etsy_browser_session import (
    EtsyProfileLockedError,
    is_session_ready as is_etsy_session_ready,
    SHOP_MANAGER_URL,
    open_etsy_login_browser,
    resolve_etsy_session,
)
from social_browser_session import (
    SOCIAL_URLS,
    is_session_ready,
    open_social_browser,
    resolve_social_session,
)
from social_post_store import get_product_social_statuses, load_social_post_records
from medium_content import make_medium_research_article
from cloud_asset_store import (
    AssetValidationError,
    PREVIEW_FILE_NAME as CLOUD_PREVIEW_FILE_NAME,
    CloudAssetError,
    CloudAssetStore,
    discover_product_roots,
    load_local_state,
    resolve_product,
)
from cloud_asset_store_config import load_config as load_cloud_asset_config
import catalog_repository
from asset_readiness import AssetReadinessEngine, AssetReadinessError
from catalog_repository import (
    CatalogRepositoryError,
    CatalogWriteConflict,
    workbook_path as catalog_workbook_path,
)

# ── Logging ─────────────────────────────────────────────────────────────────────
logger = logging.getLogger("etsy_dashboard")
if not logger.handlers:
    _stream_handler = logging.StreamHandler()
    _stream_handler.setFormatter(
        logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s")
    )
    logger.addHandler(_stream_handler)
    logger.setLevel(logging.INFO)


class _CorrelationLogAdapter(logging.LoggerAdapter):
    """Prefix log records with correlation context (shop, listing_id, operation)."""

    def process(self, msg, kwargs):
        ctx = " ".join(f"{key}={value}" for key, value in (self.extra or {}).items())
        return f"[{ctx}] {msg}", kwargs


def _ctx_logger(operation: str, **context) -> logging.LoggerAdapter:
    return _CorrelationLogAdapter(logger, {"operation": operation, **context})

# ── Paths ───────────────────────────────────────────────────────────────────────
BASE_DIR       = runtime_identity.current_runtime_root(__file__)
# ── Multi-Shop Config ───────────────────────────────────────────────────────────
CONFIG_FILE = BASE_DIR / "shops_config.json"

def load_shops(config_path: str | Path = CONFIG_FILE) -> dict[str, dict]:
    """Load and validate shops config with non-secret, actionable failures."""

    try:
        return validate_and_load_shops_config(config_path=config_path, base_dir=BASE_DIR)
    except FileNotFoundError as exc:
        raise RuntimeError(
            "Missing required local config file: shops_config.json. "
            f"Copy {DEFAULT_EXAMPLE_NAME} to shops_config.json before starting."
        ) from exc
    except ShopsConfigError as exc:
        raise RuntimeError(f"Malformed local shop config: {exc}") from exc

SHOPS = load_shops()
ACTIVE_SHOP_FILE = BASE_DIR / "active_shop.txt"

def load_active_shop_id() -> str:
    try:
        saved_shop = ACTIVE_SHOP_FILE.read_text(encoding="utf-8").strip()
        if saved_shop in SHOPS:
            return saved_shop
    except OSError:
        pass
    return "templystudios" if "templystudios" in SHOPS else list(SHOPS.keys())[0] if SHOPS else ""

_active_shop_id = load_active_shop_id()

def get_active_shop() -> dict:
    return SHOPS.get(_active_shop_id, {})

def save_shops():
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(SHOPS, f, indent=2, ensure_ascii=False)

def SHOP_DIR() -> Path:
    return BASE_DIR / "shops" / _active_shop_id

def EXCEL_FILE() -> Path:
    return SHOP_DIR() / "Etsy_SEO_Generator.xlsx"

# ── Paths ───────────────────────────────────────────────────────────────────────
STATIC_DIR     = BASE_DIR / "dashboard_static"
_OVERRIDDEN_PYTHON_BIN = os.environ.get("ETSY_AUTOMATION_PYTHON", "").strip()
_PROJECT_VENV_BIN = BASE_DIR / ".venv" / "bin" / "python"
PYTHON_BIN     = (
    _OVERRIDDEN_PYTHON_BIN
    if _OVERRIDDEN_PYTHON_BIN
    else str(_PROJECT_VENV_BIN)
    if _PROJECT_VENV_BIN.is_file()
    else str(sys.executable)
)
ETSY_POSTER    = str(BASE_DIR / "etsy_auto_post.py")
ETSY_UPDATER   = str(BASE_DIR / "etsy_push_update.py")
TRIGGER_SCRIPT = str(Path.home() / "TN/App/Truyen ngan/automation/trigger_planner.py")
PROCESSED_DIR  = Path.home() / "Desktop" / "Auto_Etsy_Planner" / ".processed"
VERTEX_URL        = "http://127.0.0.1:8080"
MLX_URL           = "http://127.0.0.1:8000"        # rapid-mlx Qwen (SEO AI)
VERTEX_IMG_URL    = "http://127.0.0.1:8080"         # Vertex Etsy Listing Studio (image gen)
VERTEX_OUTPUT_DIR = Path("/Users/aaronnguyen/vertex_etsy_listing/output")
VERTEX_INPUT_DIR  = Path("/Users/aaronnguyen/vertex_etsy_listing/input")

# ── Security boundary ───────────────────────────────────────────────────────────
_DASHBOARD_MUTATION_TOKEN_HEADER = "x-dashboard-mutation-token"
_DASHBOARD_LOOPBACK_HOSTS = {"127.0.0.1", "localhost", "::1", "127.0.0.1:8090", "localhost:8090", "::1"}
_DASHBOARD_LAN_MODE_ENV = "ETSY_DASHBOARD_ALLOW_LAN"
_DASHBOARD_ALLOWED_HOSTS_ENV = "ETSY_DASHBOARD_ALLOWED_HOSTS"
_DASHBOARD_MUTATION_TOKEN = secrets.token_urlsafe(32)


def _env_flag_truthy(name: str) -> bool:
    return (os.environ.get(name) or "").strip().lower() in {"1", "true", "yes", "on", "y"}


def _dashboard_lan_mode() -> bool:
    return _env_flag_truthy(_DASHBOARD_LAN_MODE_ENV)


def _dashboard_allowlist_hosts() -> set[str]:
    hosts = set(_DASHBOARD_LOOPBACK_HOSTS)
    raw = os.environ.get(_DASHBOARD_ALLOWED_HOSTS_ENV) or ""
    for item in raw.split(","):
        host = item.strip().lower()
        if host:
            hosts.add(host)
    if os.environ.get("PYTEST_CURRENT_TEST"):
        hosts.add("testserver")
    return hosts


def _normalize_host(raw: str) -> str:
    host = (raw or "").strip().lower()
    if not host:
        return ""
    host = host.split("/", 1)[0]
    if host.startswith("[") and "]" in host:
        host = host[1 : host.index("]")]
        return host
    if host.count(":") > 1:
        return host
    return host.split(":", 1)[0]


def _request_host(request: Request) -> str:
    return _normalize_host(request.headers.get("host") or (request.url.hostname or ""))


def _origin_host(request: Request) -> str:
    return _normalize_host(urllib.parse.urlparse(request.headers.get("origin", "")).hostname or "")


def _host_allowed(request: Request) -> bool:
    return _request_host(request) in _dashboard_allowlist_hosts()


def _origin_allowed(request: Request) -> bool:
    origin = request.headers.get("origin")
    if not origin:
        return True
    return _origin_host(request) in _dashboard_allowlist_hosts() and _origin_host(request) == _request_host(request)


def _mutation_token_matches(request: Request) -> bool:
    token = request.headers.get(_DASHBOARD_MUTATION_TOKEN_HEADER, "")
    if not token:
        return False
    return secrets.compare_digest(token, _DASHBOARD_MUTATION_TOKEN)


def _dashboard_listen_host() -> str:
    return "0.0.0.0" if _dashboard_lan_mode() else "127.0.0.1"


def _dashboard_index_html() -> str:
    index_html = (STATIC_DIR / "index.html").read_text(encoding="utf-8")
    marker = '<meta name="etsy-dashboard-mutation-token"'
    token = html.escape(_DASHBOARD_MUTATION_TOKEN, quote=True)
    if marker in index_html:
        token_tag = f'<meta name="etsy-dashboard-mutation-token" content="{token}">'
        return re.sub(
            r'<meta\s+name="etsy-dashboard-mutation-token"\s+content="[^"]*"\s*>',
            token_tag,
            index_html,
            count=1,
        )
    return index_html.replace(
        "</head>",
        f'  <meta name="etsy-dashboard-mutation-token" content="{token}">\n</head>',
        1,
    )


# ── App ─────────────────────────────────────────────────────────────────────────
app = FastAPI(title="Etsy Dashboard")
@app.middleware("http")
async def _dashboard_security_middleware(request: Request, call_next):
    if not _host_allowed(request):
        return JSONResponse({"detail": "host not allowed"}, status_code=403)
    if not _origin_allowed(request):
        return JSONResponse({"detail": "origin not allowed"}, status_code=403)
    if request.method in {"POST", "PATCH", "DELETE"}:
        if not _mutation_token_matches(request):
            return JSONResponse({"detail": "missing or invalid mutation token"}, status_code=403)
    response = await call_next(request)
    try:
        del response.headers["access-control-allow-origin"]
    except KeyError:
        pass
    return response

# Active SSE subscribers
_log_subscribers: list[asyncio.Queue] = []
_running_processes: dict[str, asyncio.subprocess.Process | None] = {}
_running_tasks: dict[str, asyncio.Task] = {}
_product_create_lock = asyncio.Lock()
_etsy_compare_lock = asyncio.Lock()
_etsy_draft_delete_lock = asyncio.Lock()
_batch_delete_lock = asyncio.Lock()
_etsy_post_run_lock = asyncio.Lock()
_etsy_update_jobs: dict[str, dict] = {}
_etsy_single_sync_busy_shops: set[str] = set()
_DIAGNOSTICS_DIR = BASE_DIR / ".etsy-dashboard-diagnostics"
_JOB_STORE_PATH = _DIAGNOSTICS_DIR / "etsy_jobs.sqlite"
_ETSY_JOB_STORE: JobStore | None = None
_ETSY_POST_LOCK_PREFIX = "__ETSY_POST__"
_RUNTIME_PRECHECK_MODULES_POSTER = ("openpyxl", "playwright.async_api", "deep_translator", "google.genai.types")
_RUNTIME_PRECHECK_MODULES_UPDATER = ("openpyxl", "playwright.async_api")

_RUNTIME_PRECHECK_TIMEOUT = 15

# Cloud asset operations are deliberately lazy so importing the dashboard does
# not contact Drive or require an rclone invocation. Tests and local tooling
# can replace this public seam with a deterministic LocalRemote-backed store.
CLOUD_ASSET_STORE: CloudAssetStore | None = None
_CLOUD_ASSET_MUTATION_LOCKS: dict[str, asyncio.Lock] = {}
_CLOUD_ASSET_MUTATION_LOCKS_GUARD = asyncio.Lock()
_CLOUD_ASSET_UPLOAD_SCHEDULES: dict[str, dict[str, Any]] = {}
_CLOUD_ASSET_UPLOAD_QUEUE_BY_SHOP: dict[str, list[str]] = {}
_CLOUD_ASSET_UPLOAD_WORKERS: dict[str, asyncio.Task] = {}
_CLOUD_ASSET_UPLOAD_QUEUE_GUARD = asyncio.Lock()
_CLOUD_ASSET_UPLOAD_IDLE_POLL_SECONDS = 0.75

# One FIFO for commands that touch Etsy's shared Chrome profile or mutate the
# cloud asset copy.  The previous implementation had separate locks for
# listing sync, posting, updating and cloud uploads; that meant a valid new
# command was rejected merely because another type of command was active.
# Keep this queue in memory deliberately: after a dashboard restart we must not
# silently resume a browser/cloud mutation without the operator seeing it.
_OPERATION_QUEUE_LOCK = asyncio.Lock()
_OPERATION_QUEUE_GUARD = asyncio.Lock()
_OPERATION_QUEUE_COMMANDS: dict[str, dict[str, Any]] = {}
_OPERATION_QUEUE_DEDUPE: dict[str, str] = {}
_OPERATION_QUEUE_TASK_PREFIX = "__OPERATION_QUEUE__:"


def _operation_queue_public(command: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in command.items() if key not in {"callback", "dedupe_key"}}


async def _enqueue_operation(
    *,
    operation: str,
    shop_id: str,
    target: str,
    callback: Callable[[], Any],
) -> tuple[dict[str, Any], bool]:
    """Accept an explicit operator command and run it in global FIFO order."""
    dedupe_key = f"{operation}:{shop_id}:{target}"
    async with _OPERATION_QUEUE_GUARD:
        existing_id = _OPERATION_QUEUE_DEDUPE.get(dedupe_key)
        existing = _OPERATION_QUEUE_COMMANDS.get(existing_id or "")
        if existing and existing.get("status") in {"queued", "running"}:
            return existing, False

        command_id = secrets.token_hex(8)
        command = {
            "command_id": command_id,
            "operation": operation,
            "shop_id": shop_id,
            "target": target,
            "status": "queued",
            "enqueued_at": time.time(),
            "updated_at": time.time(),
            "dedupe_key": dedupe_key,
            "callback": callback,
        }
        _OPERATION_QUEUE_COMMANDS[command_id] = command
        _OPERATION_QUEUE_DEDUPE[dedupe_key] = command_id
        task = asyncio.create_task(_run_queued_operation(command_id))
        _register_background_task(f"{_OPERATION_QUEUE_TASK_PREFIX}{command_id}", task)
        queued_before = sum(1 for item in _OPERATION_QUEUE_COMMANDS.values() if item.get("status") == "queued") - 1
        command["position"] = max(0, queued_before) + 1

    broadcast(f"[QUEUE] 🗓️ Đã thêm {operation}: {target} (vị trí {command['position']})")
    return command, True


async def _run_queued_operation(command_id: str) -> None:
    command = _OPERATION_QUEUE_COMMANDS.get(command_id)
    if command is None:
        return
    try:
        async with _OPERATION_QUEUE_LOCK:
            async with _OPERATION_QUEUE_GUARD:
                if command.get("status") != "queued":
                    return
                command.update({"status": "running", "started_at": time.time(), "updated_at": time.time()})
            broadcast(f"[QUEUE] ▶ Đang chạy {command['operation']}: {command['target']}")
            result = command["callback"]()
            if asyncio.iscoroutine(result):
                await result
            async with _OPERATION_QUEUE_GUARD:
                command.update({"status": "succeeded", "finished_at": time.time(), "updated_at": time.time()})
            broadcast(f"[QUEUE] ✅ Hoàn tất {command['operation']}: {command['target']}")
    except asyncio.CancelledError:
        async with _OPERATION_QUEUE_GUARD:
            command.update({"status": "cancelled", "finished_at": time.time(), "updated_at": time.time()})
        broadcast(f"[QUEUE] ⚠️ Đã hủy {command['operation']}: {command['target']}")
        raise
    except Exception as exc:
        async with _OPERATION_QUEUE_GUARD:
            command.update({"status": "failed", "error": str(exc), "finished_at": time.time(), "updated_at": time.time()})
        broadcast(f"[QUEUE] ❌ Lỗi {command['operation']}: {command['target']}: {exc}")
    finally:
        async with _OPERATION_QUEUE_GUARD:
            _OPERATION_QUEUE_DEDUPE.pop(str(command.get("dedupe_key") or ""), None)
        _pop_background_task(f"{_OPERATION_QUEUE_TASK_PREFIX}{command_id}")


@app.get("/api/operation-queue")
async def operation_queue_status():
    async with _OPERATION_QUEUE_GUARD:
        active = [item for item in _OPERATION_QUEUE_COMMANDS.values() if item.get("status") in {"queued", "running"}]
        active.sort(key=lambda item: float(item.get("enqueued_at") or 0))
        return {"ok": True, "commands": [_operation_queue_public(item) for item in active]}


def _build_cloud_asset_store(config) -> CloudAssetStore:
    """Build the configured store without performing any remote operation."""

    return CloudAssetStore(
        repo_root=config.repo_root,
        remote=config.remote,
        parent_id=config.parent_id,
        rclone_bin=config.rclone_bin,
        cache_root=config.cache_root,
        lock_timeout_seconds=config.lock_timeout_seconds,
        success_ttl_seconds=config.success_ttl_seconds,
        failure_ttl_seconds=config.failure_ttl_seconds,
        offload_age_days=config.offload_age_days,
    )


def get_cloud_asset_store() -> CloudAssetStore:
    global CLOUD_ASSET_STORE
    store = CLOUD_ASSET_STORE
    if store is None:
        config = load_cloud_asset_config(BASE_DIR)
        store = _build_cloud_asset_store(config)
        CLOUD_ASSET_STORE = store
    return store


async def _get_cloud_asset_mutation_lock(product_key: str) -> asyncio.Lock:
    async with _CLOUD_ASSET_MUTATION_LOCKS_GUARD:
        lock = _CLOUD_ASSET_MUTATION_LOCKS.get(product_key)
        if lock is None:
            lock = asyncio.Lock()
            _CLOUD_ASSET_MUTATION_LOCKS[product_key] = lock
        return lock


async def _terminate_subprocess(process: asyncio.subprocess.Process | None) -> None:
    if process is None:
        return
    try:
        if process.returncode is None:
            try:
                process.kill()
            except ProcessLookupError:
                pass
            except Exception:
                pass
    except Exception:
        pass
    try:
        await asyncio.wait_for(process.communicate(), timeout=2.0)
    except asyncio.TimeoutError:
        try:
            if process.returncode is None:
                try:
                    process.kill()
                except ProcessLookupError:
                    pass
                except Exception:
                    pass
        except Exception:
            pass
        try:
            await process.wait()
        except Exception:
            pass
    except Exception:
        try:
            await process.wait()
        except Exception:
            pass


def _build_runtime_prefetch_code(modules: tuple[str, ...]) -> str:
    safe_modules = ", ".join(repr(mod) for mod in modules)
    return (
        "import importlib\n"
        f"for module_name in ({safe_modules},):\n"
        "    importlib.import_module(module_name)\n"
    )


def _register_background_task(key: str, task: asyncio.Task):
    _running_tasks[key] = task


def _pop_background_task(key: str):
    _running_tasks.pop(key, None)


def _etsy_post_lock_key(shop_id: str) -> str:
    return f"{_ETSY_POST_LOCK_PREFIX}:{shop_id}"


def _is_poster_locked_for_shop(shop_id: str) -> bool:
    lock_key = _etsy_post_lock_key(shop_id)
    return lock_key in _running_processes or lock_key in _running_tasks


def _acquire_poster_lock(shop_id: str) -> str:
    lock_key = _etsy_post_lock_key(shop_id)
    if _is_poster_locked_for_shop(shop_id):
        raise HTTPException(409, f"Một bài đăng đang chạy cho shop {shop_id}. Vui lòng đợi hoàn tất rồi thử lại")
    _running_processes[lock_key] = None
    return lock_key


def _release_poster_lock(shop_id: str):
    lock_key = _etsy_post_lock_key(shop_id)
    _running_processes.pop(lock_key, None)
    _running_tasks.pop(lock_key, None)


_LISTING_URL_RE = re.compile(r"/(?:listing/|listing-editor/edit/)(\d+)")
_LISTING_ID_RE = re.compile(r"^\s*\d+\s*$")
_ETSY_PUBLIC_LISTING_RE = re.compile(r"^https://(?:www\.)?etsy\.com/listing/(\d+)(?:[/?#].*)?$", re.IGNORECASE)
_ETSY_MANAGER_LISTING_RE = re.compile(r"^https://(?:www\.)?etsy\.com/(?:your/shops/me/)?listing-editor/edit/(\d+)(?:[/?#].*)?$", re.IGNORECASE)


def _listing_url_has_id(value: str) -> bool:
    text = str(value or "").strip()
    return bool(_LISTING_URL_RE.search(text) or _LISTING_ID_RE.fullmatch(text))


def _safe_etsy_public_url(value: object, listing_id: str) -> str:
    match = _ETSY_PUBLIC_LISTING_RE.fullmatch(str(value or "").strip())
    return str(value).strip() if match and match.group(1) == listing_id else ""


def _safe_etsy_manager_url(value: object, listing_id: str) -> str:
    match = _ETSY_MANAGER_LISTING_RE.fullmatch(str(value or "").strip())
    return str(value).strip() if match and match.group(1) == listing_id else ""

def broadcast(msg: str):
    for q in _log_subscribers:
        try:
            q.put_nowait(msg)
        except asyncio.QueueFull:
            logger.warning("broadcast message dropped (subscriber queue full): %s", msg)


async def _runtime_prefetch_import_check(python_bin: str, label: str, modules: tuple[str, ...]) -> tuple[bool, str]:
    process: asyncio.subprocess.Process | None = None
    try:
        process = await asyncio.create_subprocess_exec(
            python_bin,
            "-c",
            _build_runtime_prefetch_code(modules),
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
            cwd=str(BASE_DIR),
        )
        try:
            stdout, _ = await asyncio.wait_for(process.communicate(), timeout=_RUNTIME_PRECHECK_TIMEOUT)
        except asyncio.TimeoutError:
            await _terminate_subprocess(process)
            return False, f"{label} preflight timeout ({_RUNTIME_PRECHECK_TIMEOUT}s)"
    except asyncio.CancelledError:
        await _terminate_subprocess(process)
        raise
    except Exception as exc:
        return False, f"{label} preflight failed: {type(exc).__name__}: {exc}"

    output = stdout.decode("utf-8", errors="replace").strip() if stdout else ""
    if process.returncode != 0:
        return False, f"{label} preflight failed (exit {process.returncode}): {output[:220]}"
    return True, output

# ── Excel helpers ────────────────────────────────────────────────────────────────
IMG_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}

def _renderable_image_url(folder: str, image_path: Path) -> dict | None:
    """Resolve an image without opening content or hydrating an iCloud placeholder."""
    try:
        original_stat = image_path.stat()
    except OSError:
        return None

    original_has_blocks = getattr(original_stat, "st_blocks", 1) > 0
    encoded_folder = urllib.parse.quote(str(folder), safe="")
    encoded_name = urllib.parse.quote(image_path.name, safe="")
    original_url = f"/files/{encoded_folder}/images/{encoded_name}"

    # Real hydrated image (has backing blocks) is preferred and can be treated as
    # fully renderable.
    if original_stat.st_size > 0 and original_has_blocks:
        return {
            "url": original_url,
            "full_url": original_url,
            "preview_only": False,
            "hydration_needed": False,
            "availability": "local",
        }

    cache_name = f"{hashlib.md5(image_path.name.encode('utf-8')).hexdigest()[:12]}_180.webp"
    cache_path = image_path.parent / ".thumbcache" / cache_name

    # Dataless/sparse images can still be previewed from cache when present.
    try:
        cache_stat = cache_path.stat()
        cache_ready = cache_stat.st_size > 0 and getattr(cache_stat, "st_blocks", 1) > 0
    except OSError:
        cache_ready = False
    if cache_ready:
        cache_url = f"/files/{encoded_folder}/images/.thumbcache/{urllib.parse.quote(cache_name, safe='')}"
        return {
            "url": cache_url,
            "full_url": original_url,
            "preview_only": True,
            "hydration_needed": original_stat.st_size > 0 and not original_has_blocks,
            "availability": "cached_preview",
        }

    # As a final fallback for dataless/sparse files with metadata, return the
    # original URL marked as preview-only so the UI/API can surface a hydration
    # path without needing to read image bytes.
    if original_stat.st_size > 0 and not original_has_blocks:
        return {
            "url": original_url,
            "full_url": original_url,
            "preview_only": True,
            "hydration_needed": True,
            "availability": "hydration_required",
        }

    return None


def _cloud_preview_descriptor(
    folder: str,
    folder_path: Path,
    cloud_manifest: dict | None = None,
) -> dict | None:
    """Describe the retained root preview without hydrating full assets.

    ``full_url`` points at a read-only image endpoint.  The endpoint hydrates
    only the requested manifest-verified image into the Phase 2 cache when the
    product is cloud-only; it never installs bytes back into ``images/``.
    """

    preview_path = folder_path / CLOUD_PREVIEW_FILE_NAME
    try:
        stat_result = preview_path.stat()
        if preview_path.is_symlink() or stat_result.st_size <= 0 or getattr(stat_result, "st_blocks", 1) <= 0:
            return None
    except OSError:
        return None

    encoded_folder = urllib.parse.quote(str(folder), safe="")
    preview_url = f"/api/cloud-assets/preview/{encoded_folder}"
    full_url = preview_url
    image_name = CLOUD_PREVIEW_FILE_NAME
    files = cloud_manifest.get("files", []) if isinstance(cloud_manifest, dict) else []
    image_records = []
    for item in files if isinstance(files, list) else []:
        if not isinstance(item, dict) or item.get("role") != "image":
            continue
        relative = str(item.get("path") or "").replace("\\", "/").strip("/")
        parts = relative.split("/")
        if not relative or not parts or parts[0] != "images" or any(part in {"", ".", ".."} for part in parts):
            continue
        image_records.append(relative)
    if image_records:
        relative = sorted(image_records)[0]
        encoded_relative = urllib.parse.quote(relative, safe="/")
        full_url = f"/api/cloud-assets/image/{encoded_folder}/{encoded_relative}"
        image_name = Path(relative).name
    return {
        "url": preview_url,
        "full_url": full_url,
        "name": image_name,
        "preview_only": True,
        "hydration_needed": full_url != preview_url,
        "availability": "cloud_preview",
    }

def set_cell_value(ws, row_num: int, col_num: int, val):
    cell = ws.cell(row=row_num, column=col_num)
    if hasattr(cell, "value"):
        cell.value = val

def copy_and_clean_template(src_path: Path, dst_path: Path):
    """
    Copies the Excel template and deletes all listings (rows 4 onwards)
    to ensure the destination starts with a clean slate.
    """
    import shutil
    shutil.copy(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)
    ws = wb["Listings"]
    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row - 3)
    wb.save(dst_path)

def get_sku_prefix(shop_id: str) -> str:
    shop_id_lower = (shop_id or "templystudios").lower()
    if "temply" in shop_id_lower:
        return "TS"
    elif "daisy" in shop_id_lower:
        return "dd"
    else:
        return (shop_id or "TS")[:2].upper()

def generate_sku(shop_id: str, folder_name: str) -> str:
    prefix = get_sku_prefix(shop_id)
    import re
    clean_folder = "".join(c if c.isalnum() else "_" for c in folder_name).lower()
    clean_folder = re.sub(r'_+', '_', clean_folder).strip('_')
    return f"{prefix}_{clean_folder}"

EMPTY_PRODUCT_TITLE = "San pham trong - can bo sung"

def _asset_file_is_usable(path: Path) -> bool:
    if not path.is_file() or path.name == ".DS_Store":
        return False
    try:
        stat = path.stat()
    except OSError:
        return False
    return stat.st_size > 0 and getattr(stat, "st_blocks", 1) > 0

def _folder_has_usable_assets(folder_path: Path) -> bool:
    for subfolder in ("images", "files"):
        asset_dir = folder_path / subfolder
        if not asset_dir.is_dir():
            continue
        try:
            if any(_asset_file_is_usable(path) for path in asset_dir.iterdir()):
                return True
        except OSError:
            continue
    return False

def _clear_product_asset_folder(folder_path: Path):
    for subfolder in ("images", "files"):
        asset_dir = folder_path / subfolder
        if asset_dir.exists():
            shutil.rmtree(asset_dir)
        asset_dir.mkdir(parents=True, exist_ok=True)

def _product_folder_number(folder_name: str) -> Optional[int]:
    match = re.fullmatch(r"product-(\d+)", str(folder_name or "").strip())
    return int(match.group(1)) if match else None

def _validate_product_numbered_folder_name(folder: object) -> str:
    normalized = str(folder or "").strip()
    if not normalized:
        raise HTTPException(400, "Tên folder không được để trống")
    if not re.fullmatch(r"product-\d+", normalized):
        raise HTTPException(400, f"Tên folder không hợp lệ (yêu cầu định dạng product-N): {normalized}")
    return normalized

def _find_mapped_catalog_rows(ws) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row_num in range(4, ws.max_row + 1):
        folder = str(ws.cell(row=row_num, column=2).value or "").strip()
        if not folder:
            continue
        if any(ws.cell(row=row_num, column=column).value not in (None, "") for column in range(2, 19)):
            rows[folder] = row_num
    return rows

def _make_local_product_slot_rows(ws, start_row: int, total: int) -> list[int]:
    rows: list[int] = []
    cursor = start_row
    for _ in range(total):
        row_num = _next_empty_catalog_row(ws, cursor)
        rows.append(row_num)
        cursor = row_num + 1
    return rows

def _product_folder_name(number: int) -> str:
    return f"product-{number:02d}"

def _find_reusable_empty_product_slots(ws, shop_dir: Path) -> list[dict]:
    slots = []
    for row_num in range(4, ws.max_row + 1):
        folder_name = str(ws.cell(row=row_num, column=2).value or "").strip()
        number = _product_folder_number(folder_name)
        if number is None:
            continue
        folder_path = shop_dir / folder_name
        if not folder_path.is_dir() or _folder_has_usable_assets(folder_path):
            continue
        slots.append({
            "folder": folder_name,
            "number": number,
            "path": folder_path,
            "row": row_num,
        })
    return sorted(slots, key=lambda slot: slot["number"])

def _next_product_number(ws, shop_dir: Path, used_folders: Optional[set[str]] = None) -> int:
    if used_folders is None:
        used_folders = set()
    max_product_num = 0
    for row_num in range(4, ws.max_row + 1):
        number = _product_folder_number(ws.cell(row=row_num, column=2).value)
        if number is not None:
            max_product_num = max(max_product_num, number)
    if shop_dir.exists():
        try:
            for item in shop_dir.iterdir():
                if item.name in used_folders:
                    continue
                number = _product_folder_number(item.name) if item.is_dir() else None
                if number is not None:
                    max_product_num = max(max_product_num, number)
        except OSError:
            pass
    return max_product_num + 1

def _allocate_product_slot(
    ws,
    shop_dir: Path,
    start_row: int = 4,
    reusable_slots: Optional[list[dict]] = None,
    used_folders: Optional[set[str]] = None,
) -> dict:
    if used_folders is None:
        used_folders = set()
    reusable_slots = reusable_slots if reusable_slots is not None else _find_reusable_empty_product_slots(ws, shop_dir)
    while reusable_slots:
        slot = reusable_slots.pop(0)
        if slot["folder"] in used_folders:
            continue
        _clear_product_asset_folder(slot["path"])
        used_folders.add(slot["folder"])
        return {**slot, "reused": True}

    number = _next_product_number(ws, shop_dir, used_folders)
    while True:
        folder_name = _product_folder_name(number)
        if folder_name not in used_folders and not (shop_dir / folder_name).exists():
            break
        number += 1
    folder_path = shop_dir / folder_name
    used_folders.add(folder_name)
    return {
        "folder": folder_name,
        "number": number,
        "path": folder_path,
        "row": _next_empty_catalog_row(ws, start_row),
        "reused": False,
    }

def _clear_catalog_row(ws, row_num: int):
    for column in range(2, 19):
        set_cell_value(ws, row_num, column, None)


def _quarantine_root(shop_dir: Path) -> Path:
    """Recoverable folder for removed local product folders."""
    root = shop_dir / ".deleted_local_products"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _quarantine_destination(shop_dir: Path, folder: str) -> Path:
    root = _quarantine_root(shop_dir)
    safe_folder = re.sub(r"[^A-Za-z0-9._-]+", "_", folder)
    marker = f"{time.strftime('%Y%m%d_%H%M%S')}_{time.time_ns()}"
    return root / f"{safe_folder}_deleted_{marker}"


def _build_local_delete_metadata(row: int, folder: str, target_folder: Path) -> dict[str, int | str]:
    return {
        "row": row,
        "folder": folder,
        "quarantine_folder": str(target_folder),
    }


def _parse_delete_request_payload(data: dict, *, require_folder: bool = True) -> tuple[str, str]:
    if not isinstance(data, dict):
        raise HTTPException(400, "Payload phải là JSON object")
    shop = str(data.get("shop") or "").strip()
    if not shop:
        raise HTTPException(400, "Thiếu thông tin shop")
    try:
        _assert_shop_identity(shop)
    except RuntimeError as error:
        raise HTTPException(409, str(error)) from error
    folder = str(data.get("folder") or "").strip() if require_folder else ""
    if require_folder:
        folder = _validate_product_numbered_folder_name(folder)
    return shop, folder


async def _read_delete_payload(req: Request) -> dict:
    try:
        data = await req.json()
    except Exception as error:
        raise HTTPException(400, "Payload phải là JSON object") from error
    if not isinstance(data, dict):
        raise HTTPException(400, "Payload phải là JSON object")
    return data


def _parse_run_selected_request_payload(data: dict) -> tuple[str, list[tuple[int, str]]]:
    if not isinstance(data, dict):
        raise HTTPException(400, "Payload phải là JSON object")

    shop = str(data.get("shop") or "").strip()
    if not shop:
        raise HTTPException(400, "Thiếu thông tin shop")
    try:
        _assert_shop_identity(shop)
    except RuntimeError as error:
        raise HTTPException(409, str(error)) from error

    raw_items = data.get("items")
    if not isinstance(raw_items, list) or not raw_items:
        raise HTTPException(400, "items phải là danh sách không rỗng")

    normalized: list[tuple[int, str]] = []
    seen_rows: set[int] = set()
    seen_folders: set[str] = set()
    for item in raw_items:
        if not isinstance(item, dict):
            raise HTTPException(400, "Mỗi phần tử trong items phải có dạng {row, folder}")
        row = item.get("row")
        if isinstance(row, bool) or not isinstance(row, int):
            raise HTTPException(400, "Mỗi item phải có row dạng số nguyên")
        if row < 4:
            raise HTTPException(400, f"Row {row} không hợp lệ (phải >= 4)")
        if row in seen_rows:
            raise HTTPException(400, f"row {row} bị lặp trong danh sách")
        folder = _validate_product_numbered_folder_name(item.get("folder"))
        if folder in seen_folders:
            raise HTTPException(400, f"folder bị lặp trong danh sách: {folder}")

        normalized.append((row, folder))
        seen_rows.add(row)
        seen_folders.add(folder)

    return shop, normalized


async def _read_optional_json_payload(request: Request | None) -> dict | None:
    """Read an optional JSON body while keeping older body-less callers working."""
    if request is None:
        return None
    try:
        data = await request.json()
    except Exception as error:
        try:
            raw_body = await request.body()
        except Exception:
            raw_body = b""
        if not raw_body:
            return None
        raise HTTPException(400, "Payload phải là JSON object") from error
    if data is None:
        return None
    if not isinstance(data, dict):
        raise HTTPException(400, "Payload phải là JSON object")
    return data


def _get_job_store() -> JobStore:
    global _ETSY_JOB_STORE
    store = _ETSY_JOB_STORE
    if store is None:
        _DIAGNOSTICS_DIR.mkdir(parents=True, exist_ok=True)
        store = JobStore(_JOB_STORE_PATH)
        _ETSY_JOB_STORE = store
        recovered = store.recover_running_jobs()
        if recovered:
            broadcast(f"[JOB-STORE] ♻️ Đã đánh dấu {recovered} job đang chạy trước đó là failed")
    return store


def _op_status_to_legacy(status: str | None) -> str:
    normalized = str(status or "").strip().lower()
    if normalized in {"queued", "starting", "preflight"}:
        return "starting"
    if normalized in {"running"}:
        return "running"
    if normalized in {"succeeded", "success"}:
        return "success"
    if normalized in {"failed", "error"}:
        return "error"
    if normalized == "cancelled":
        return "cancelled"
    return normalized


def _cache_job_status(job_id: str, update: dict | None = None) -> None:
    if update is None:
        return
    existing = _etsy_update_jobs.get(job_id, {})
    merged = dict(existing)
    merged.update(update)
    merged["job_id"] = job_id
    merged.setdefault("logs", [])
    merged.setdefault("fields", [])
    _etsy_update_jobs[job_id] = merged


def _job_payload_for_frontend(job_record: dict[str, object] | None) -> dict | None:
    if not job_record:
        return None
    payload = dict(job_record)
    legacy_status = _op_status_to_legacy(str(payload.get("status")))
    payload["status"] = legacy_status
    payload["last_message"] = str(
        payload.get("last_message") or payload.get("latest_log_excerpt") or ""
    )
    logs = payload.get("logs")
    if not isinstance(logs, list):
        logs = [value for value in [payload.get("last_message")] if value]
    payload["logs"] = logs[-12:]
    return payload


def _coerce_int(value: object, default: int = 0) -> int:
    """Convert persisted JSON/SQLite scalar values without widening types."""
    if isinstance(value, (int, float, str, bytes, bytearray)):
        try:
            return int(value)
        except (TypeError, ValueError, OverflowError):
            pass
    return default


def _coerce_float(value: object, default: float = 0.0) -> float:
    """Convert persisted JSON/SQLite scalar values without raising on corrupt data."""
    if isinstance(value, (int, float, str, bytes, bytearray)):
        try:
            return float(value)
        except (TypeError, ValueError, OverflowError):
            pass
    return default


def _job_center_payload(job_record: dict[str, object] | None) -> dict | None:
    """Return a compact, path-free record for the operator Job Center."""
    if not job_record:
        return None
    fields = job_record.get("fields")
    if not isinstance(fields, list):
        fields = []
    message = str(job_record.get("last_message") or job_record.get("latest_log_excerpt") or "")
    # Job Center messages are intentionally one-line excerpts; never expose
    # operation receipts, raw logs, or local filesystem paths in this endpoint.
    message = re.sub(r"(?:/Users|/private/var|/tmp)/[^\s]+", "[local path]", message)
    return {
        "job_id": str(job_record.get("job_id") or ""),
        "shop_id": str(job_record.get("shop_id") or ""),
        "operation": str(job_record.get("operation") or ""),
        "row": job_record.get("row"),
        "folder": str(job_record.get("folder") or ""),
        "listing_id": str(job_record.get("listing_id") or ""),
        "status": str(job_record.get("status") or ""),
        "attempt_count": _coerce_int(job_record.get("attempt_count"), 1),
        "created_at": job_record.get("created_at"),
        "updated_at": job_record.get("updated_at"),
        "started_at": job_record.get("started_at"),
        "finished_at": job_record.get("finished_at"),
        "pid": job_record.get("pid"),
        "exit_code": job_record.get("exit_code"),
        "retry_parent_job_id": job_record.get("retry_parent_job_id"),
        "fields": [str(field) for field in fields if isinstance(field, str)],
        "last_message": message[:500],
    }


async def _queue_etsy_update_job(
    context: OperationContext,
    fields: list[str],
    job_record: dict[str, object],
) -> str:
    """Attach one durable job to the existing updater runtime."""
    job_id = str(job_record["job_id"])
    if len(_etsy_update_jobs) > 100:
        oldest = sorted(
            _etsy_update_jobs,
            key=lambda key: _etsy_update_jobs[key].get("created_at", 0),
        )[:25]
        for old_job_id in oldest:
            _etsy_update_jobs.pop(old_job_id, None)
    _cache_job_status(
        job_id,
        {
            "status": "starting",
            "operation": context.operation,
            "folder": context.folder,
            "listing_id": context.listing_id,
            "fields": fields,
            "shop_id": context.shop_id,
            "created_at": _coerce_float(job_record.get("created_at"), time.time()),
            "row": context.row,
            "request_id": context.request_id,
            "last_message": "Runtime started; waiting for Chrome marker",
            "logs": [],
        },
    )
    async def run_update():
        await _run_etsy_updater(context=context, fields=fields, job_id=job_id)

    command, _ = await _enqueue_operation(
        operation="etsy-update",
        shop_id=context.shop_id,
        target=f"{context.row}:{context.folder}:{context.listing_id}:{','.join(fields)}",
        callback=run_update,
    )
    _cache_job_status(job_id, {"queue_command_id": command["command_id"], "last_message": "Đã thêm vào hàng chờ chung"})
    return job_id


def _validate_product_etsy_identity(
    row: int,
    product: dict,
    payload: dict | None,
) -> tuple[str, str, str]:
    """Fail closed when a caller supplies a stale shop/folder/listing mapping."""
    try:
        context = OperationContext.from_request(
            operation="etsy_sync_from",
            row=row,
            payload=payload or {},
            active_shop_id=_active_shop_id,
            current_folder=product.get("folder"),
            current_etsy_url=product.get("etsy_url"),
        )
    except OperationContextError as error:
        _raise_context_error(error)
    return context.shop_id, context.folder, context.listing_id


def _raise_context_error(error: OperationContextError) -> NoReturn:
    message = str(error)
    if any(
        key in message
        for key in (
            "Payload định danh phải có đủ",
            "Sản phẩm chưa có Etsy listing ID hợp lệ",
            "không hợp lệ",
            "dữ liệu listing_id",
            "Row không hợp lệ",
            "Dữ liệu sản phẩm thiếu folder",
        )
    ):
        raise HTTPException(400, message) from error
    raise HTTPException(409, message) from error


def _resolve_update_context(
    row: int,
    product: dict,
    payload: dict | None,
    operation: str = "etsy_push_update",
) -> OperationContext:
    try:
        return OperationContext.from_request(
            operation=operation,
            row=row,
            payload=payload or {},
            active_shop_id=_active_shop_id,
            current_folder=product.get("folder"),
            current_etsy_url=product.get("etsy_url"),
        )
    except OperationContextError as error:
        _raise_context_error(error)
        raise  # pragma: no cover


def _etsy_update_shop_is_busy(shop_id: str) -> bool:
    requested_shop_id = str(shop_id)
    active_jobs = _get_job_store().get_active_jobs_for_shop(shop_id)
    if any(
        str(job.get("shop_id")) == requested_shop_id
        and str(job.get("status")) in {JOB_STATUS_QUEUED, JOB_STATUS_RUNNING}
        for job in active_jobs
    ):
        return True

    # Keep the legacy in-memory cache as a compatibility overlay for callers
    # and older tests that still populate it directly.  ``starting`` is the
    # legacy representation of a queued durable job.
    return any(
        str(job.get("shop_id")) == requested_shop_id
        and str(job.get("status")) in {JOB_STATUS_QUEUED, JOB_STATUS_RUNNING, "starting"}
        for job in _etsy_update_jobs.values()
    )


def _partition_selected_local_products(
    shop_id: str,
    selected_items: list[tuple[int, str]],
) -> tuple[list[tuple[int, str]], list[dict]]:
    """Split selected items into (valid, rejected).

    Per-item problems (đã có URL, Đã đăng, thiếu title, ...) are rejected with a
    reason so the batch can continue with the remaining products.
    """
    excel_path = EXCEL_FILE()
    if not excel_path.exists():
        raise HTTPException(404, "Chưa có file Excel của shop đang hoạt động")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Listings"]
    shop_dir = SHOP_DIR()
    valid: list[tuple[int, str]] = []
    rejected: list[dict] = []
    for row, folder in selected_items:
        current_folder = str(ws.cell(row=row, column=2).value or "").strip()
        if current_folder != folder:
            rejected.append({
                "row": row,
                "folder": folder,
                "reason": (
                    f"Row {row} bị thay đổi: folder hiện tại "
                    f"'{current_folder or '[trống]'}' không khớp '{folder}'"
                ),
            })
            continue
        title = ws.cell(row=row, column=8).value
        if not str(title or "").strip():
            rejected.append({
                "row": row,
                "folder": folder,
                "reason": f"Row {row} chưa có đủ dữ liệu: thiếu title",
            })
            continue
        status = str(ws.cell(row=row, column=14).value or "").strip()
        if "Đã đăng" in status:
            rejected.append({
                "row": row,
                "folder": folder,
                "reason": f"Row {row} đang có trạng thái '{status}', không nên đăng lại",
            })
            continue
        if _listing_url_has_id(str(ws.cell(row=row, column=16).value or "")):
            rejected.append({
                "row": row,
                "folder": folder,
                "reason": f"Row {row} đã có Etsy URL/listing ID tại cột P, không nên đăng lại",
            })
            continue
        folder_path = shop_dir / folder
        if not folder_path.exists():
            rejected.append({
                "row": row,
                "folder": folder,
                "reason": f"Không tìm thấy folder local: {folder}",
            })
            continue
        valid.append((row, folder))
    return valid, rejected


def _validate_selected_local_products(
    shop_id: str,
    selected_items: list[tuple[int, str]],
) -> list[tuple[int, str]]:
    """Legacy hard-fail wrapper — prefer _partition_selected_local_products."""
    valid, rejected = _partition_selected_local_products(shop_id, selected_items)
    if rejected:
        raise HTTPException(400, rejected[0]["reason"])
    return valid


def _mark_selected_row_errors(excel_path: Path, rejected: list[dict]) -> None:
    """Write ❌ Lỗi status for skipped batch items so the card shows the reason."""
    if not rejected:
        return
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Listings"]
    for item in rejected:
        reason = str(item.get("reason") or "Không đủ điều kiện đăng").strip()
        set_cell_value(ws, int(item["row"]), 14, f"❌ Lỗi: {reason}")
    wb.save(excel_path)


def _set_selected_rows_pending(excel_path: Path, selected_rows: list[int]) -> None:
    if not selected_rows:
        return
    backup = excel_path.with_name(f"{excel_path.stem}.run_selected_backup_{time.strftime('%Y%m%d_%H%M%S')}{excel_path.suffix}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Listings"]

    backup_created = False
    try:
        shutil.copy2(excel_path, backup)
        backup_created = True
        for row in selected_rows:
            set_cell_value(ws, row, 14, "⏳ Chờ đăng")
        wb.save(excel_path)
    except Exception as error:
        if backup_created:
            try:
                shutil.copy2(backup, excel_path)
            except Exception as restore_error:
                raise HTTPException(
                    500,
                    "Đặt trạng thái Chờ đăng thất bại và không khôi phục được workbook",
                ) from restore_error
        raise HTTPException(500, "Đặt trạng thái Chờ đăng thất bại, đã rollback workbook") from error


def products_from_excel() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_FILE(), data_only=True)
    ws = wb["Listings"]
    products = []
    DIGITAL_EXTS = {".pdf", ".zip"}

    # Cache directory listing to avoid hundreds of disk I/O calls
    shop_dir = SHOP_DIR()
    existing_folders = set()
    social_products = load_social_post_records(
        BASE_DIR, _active_shop_id
    ).get("products", {})
    if shop_dir.exists():
        try:
            existing_folders = {d.name for d in shop_dir.iterdir() if d.is_dir()}
        except Exception:
            pass

    for row_num in range(4, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=c).value for c in range(1, 19)]
        folder = row[1]   # col B
        title  = row[7]   # col H (generated Etsy title)
        factory_seed_title = str(row[3] or "").strip()  # col D fallback seeded by Image Factory
        if not folder or str(folder) not in existing_folders:
            continue

        folder_path = shop_dir / str(folder)

        img_dir  = folder_path / "images"
        file_dir = folder_path / "files"

        images = sorted([f.name for f in img_dir.iterdir()
                         if f.suffix.lower() in IMG_EXTS]) if img_dir.exists() else []
        dig_files = sorted([f.name for f in file_dir.iterdir()
                            if f.suffix.lower() in DIGITAL_EXTS]) if file_dir.exists() else []
        planner_files = ([f.name for f in file_dir.iterdir()
                          if f.suffix.lower() in IMG_EXTS | {".pdf", ".zip"}]
                         if file_dir.exists() else [])

        # Check what fields are missing for SEO
        missing_fields = []
        if not title:
            missing_fields.append("title")
        if not str(row[8] or "").strip():
            missing_fields.append("description")

        raw_tags = str(row[9] or "").strip()
        if not raw_tags:
            missing_fields.append("tags")
        else:
            tags_list = [t.strip() for t in raw_tags.split(",") if t.strip()]
            if len(tags_list) < 13:
                missing_fields.append("tags_count")

        # If no title: show only if folder has actual files/images OR folder exists on disk
        has_content = len(images) > 0 or len(dig_files) > 0 or len(planner_files) > 0
        if not title:
            if not has_content and not folder_path.exists():
                continue   # truly empty and no folder on disk — hide
            # Show with placeholder so user can trigger SEO generation
            display_title = factory_seed_title or f"[Cần SEO] {folder}"
            status = "⚠ Cần generate SEO"
        else:
            display_title = str(title)
            status = str(row[13] or "⏳ Chờ đăng")

        sku_val = str(row[17] or "") if len(row) > 17 else ""
        if not sku_val or not sku_val.strip():
            sku_val = generate_sku(_active_shop_id, str(folder))

        renderable_images = [resolved for img in images
                             if (resolved := _renderable_image_url(str(folder), img_dir / img))]
        cloud_manifest = None
        cloud_state_path = folder_path / ".cloud-assets.json"
        if cloud_state_path.is_file() and not cloud_state_path.is_symlink():
            try:
                cloud_state = load_local_state(folder_path)
                candidate_manifest = cloud_state.get("current_manifest")
                if isinstance(candidate_manifest, dict):
                    cloud_manifest = candidate_manifest
            except (OSError, ValueError, TypeError, KeyError):
                cloud_manifest = None
        cloud_preview = _cloud_preview_descriptor(str(folder), folder_path, cloud_manifest)
        thumb_image = next((image for image in renderable_images if image.get("availability") != "hydration_required"), None)
        gallery_images = renderable_images
        if thumb_image is None and cloud_preview:
            gallery_images = [cloud_preview, *renderable_images]
        products.append({
            "row":         row_num,
            "folder":      str(folder),
            "keywords":    str(row[2]  or ""),
            "seed_title":   factory_seed_title,
            "price":       float(str(row[4])) if row[4] else 4.99,
            "category":    str(row[5]  or ""),
            "title":       display_title,
            "description": str(row[8]  or ""),
            "tags":        str(row[9]  or ""),
            "qty":         int(str(row[10])) if row[10] else 999,
            "when_made":   str(row[12] or "2020_2026"),
            "status":      status,
            "section":     str(row[14] or "").strip() or "Digital Planner",
            "image_count": len(images),
            "thumb":       thumb_image["url"] if thumb_image else (cloud_preview["url"] if cloud_preview else (renderable_images[0]["url"] if renderable_images else None)),
            "all_images":  [
                image["url"] for image in renderable_images
                if image.get("availability") != "hydration_required"
            ],
            "image_previews": gallery_images,
            "cloud_preview": cloud_preview,
            "pdf_count":   len(dig_files),
            "has_planner": len(planner_files) > 0,
            "needs_seo":   not bool(title),
            "missing_fields": missing_fields,
            "etsy_url":    str(row[15] or "") if len(row) > 15 else "",
            "extra":       str(row[16] or "") if len(row) > 16 else "",
            "sku":         sku_val,
            "social_statuses": (
                social_products.get(str(folder), {}).get("channels", {})
                if isinstance(social_products.get(str(folder)), dict)
                else {}
            ),
        })

    # Sắp xếp theo số thứ tự của folder (VD: product-01 -> 1)
    def extract_num(folder_name):
        try:
            return int(folder_name.split("-")[1])
        except:
            return 9999

    products.sort(key=lambda p: extract_num(p["folder"]))

    return products


_CATALOG_UPDATE_COLUMNS = {
    "title": 8, "description": 9, "tags": 10, "keywords": 3,
    "price": 5, "qty": 11, "status": 14, "section": 15,
    "category": 6, "when_made": 13, "etsy_url": 16, "extra": 17,
    "sku": 18,
}


def _apply_excel_updates(workbook: openpyxl.Workbook, row_num: int, updates: dict) -> None:
    ws = workbook["Listings"]
    for field, val in updates.items():
        if field in _CATALOG_UPDATE_COLUMNS:
            if field == "title" and val and str(val).startswith("[Cần SEO]"):
                val = ""
            set_cell_value(ws, row_num, _CATALOG_UPDATE_COLUMNS[field], val)


def _is_canonical_catalog_workbook(target_excel: Path) -> bool:
    try:
        return target_excel.resolve() == catalog_workbook_path(BASE_DIR, _active_shop_id).resolve()
    except (OSError, ValueError, CatalogRepositoryError):
        return False


def save_to_excel(
    row_num: int,
    updates: dict,
    excel_path: Optional[Path] = None,
    *,
    expected_version: str | None = None,
    expected_hash: str | None = None,
):
    """Persist product fields, using the atomic repository for the active catalog.

    Explicit non-canonical workbook paths retain the legacy writer for callers
    that operate on imported/snapshot workbooks.  Preconditions are accepted as
    keyword-only arguments so older callers remain source-compatible.
    """
    target_excel = Path(excel_path) if excel_path is not None else EXCEL_FILE()
    if _is_canonical_catalog_workbook(target_excel):
        return catalog_repository.apply_catalog_update(
            BASE_DIR,
            _active_shop_id,
            "dashboard_product_update",
            lambda workbook: _apply_excel_updates(workbook, row_num, updates),
            expected_version=expected_version,
            expected_hash=expected_hash,
            workbook_name=target_excel.name,
        )

    wb = openpyxl.load_workbook(target_excel)
    try:
        _apply_excel_updates(wb, row_num, updates)
        wb.save(target_excel)
    finally:
        wb.close()


def _catalog_preconditions(data: dict, request: Request) -> tuple[dict, str | None, str | None]:
    updates = dict(data)
    expected_version = updates.pop("_expected_version", None)
    expected_hash = updates.pop("_expected_hash", None)
    if expected_version is None:
        expected_version = request.headers.get("x-catalog-expected-version")
    if expected_hash is None:
        expected_hash = request.headers.get("x-catalog-expected-hash")
    if expected_hash is None:
        expected_hash = request.headers.get("if-match")

    def _normalize(value: object) -> str | None:
        if value is None:
            return None
        normalized = str(value).strip().strip('"')
        return normalized or None

    return updates, _normalize(expected_version), _normalize(expected_hash)


def _safe_asset_filename(value: str, fallback: str) -> str:
    name = Path(urllib.parse.unquote(str(value or ""))).name.strip()
    name = re.sub(r"[^A-Za-z0-9._() -]+", "_", name).strip(" ._")
    return name or fallback


def _to_il_fullxfull_url(value: str) -> str | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    cleaned = raw.split("#", 1)[0].split("?", 1)[0].strip()
    if "i.etsystatic.com" not in cleaned or "/r/il/" not in cleaned:
        return None
    if not cleaned.startswith(("http://", "https://")):
        return None

    normalized = re.sub(
        r"il_[^/.]+(?P<suffix>\.[^/.]+)?\.(?P<ext>jpe?g|png|webp|gif)(?![^/]*\.)",
        lambda m: f"il_fullxfull{m.group('suffix') or ''}.{m.group('ext')}",
        cleaned,
        flags=re.IGNORECASE,
        count=1,
    )
    if "il_fullxfull" not in normalized:
        return None
    return normalized


def _resolve_browser_session_dir(target_shop_id: str) -> Path:
    shop_cfg = SHOPS.get(target_shop_id, {})
    configured_session = str(shop_cfg.get("browser_session") or "").strip()
    if configured_session:
        return Path(os.path.expanduser(configured_session))
    if target_shop_id == "templystudios":
        return BASE_DIR / ".browser-session"
    return Path.home() / f".etsy_browser_session_{target_shop_id}"


def _assert_shop_identity(target_shop_id: str) -> None:
    if target_shop_id != _active_shop_id:
        raise RuntimeError(
            f"Shop không khớp: đang hoạt động={_active_shop_id or 'chưa chọn'}, yêu cầu={target_shop_id}"
        )


def _etsy_session_payload(shop_id: str) -> dict:
    session = resolve_etsy_session(BASE_DIR, SHOPS, shop_id)
    return {
        "shop_id": shop_id,
        "shop_name": SHOPS.get(shop_id, {}).get("name", shop_id),
        "profile_dir": str(session.profile_dir),
        "debug_port": session.debug_port,
        "browser_ready": is_etsy_session_ready(session),
        "purpose": "etsy_posting",
    }


def _extract_manager_listing_id(value: str) -> str:
    match = _ETSY_MANAGER_LISTING_RE.fullmatch(str(value or "").strip())
    return match.group(1) if match else ""


def _normalize_etsy_shop_slug(raw: str) -> str:
    slug = str(raw or "").strip().lower().strip("/")
    if not slug:
        return ""
    slug = re.sub(r"\?.*$", "", slug)
    slug = re.sub(r"/.*$", "", slug)
    return slug.strip()


def _expected_shop_slug(target_shop_id: str) -> str:
    shop_url = str(SHOPS.get(target_shop_id, {}).get("etsy_link") or "").strip()
    match = re.search(r"/shop/([^/?#]+)", shop_url, re.I)
    return _normalize_etsy_shop_slug(match.group(1)) if match else ""


def _dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        deduped.append(value)
    return deduped


def _parse_visible_shop_slugs(values: list[str] | None) -> list[str]:
    normalized: list[str] = []
    for raw in values or []:
        slug = _normalize_etsy_shop_slug(raw)
        if not slug or slug == "me":
            continue
        normalized.append(slug)
    return _dedupe_preserve_order(normalized)


def _is_etsy_auth_required_text(value: str) -> bool:
    normalized = str(value or "").lower()
    return any(
        token in normalized
        for token in (
            "sign in to continue",
            "please sign in",
            "create an account",
            "đăng nhập để",
            "you need to sign in",
        )
    )


def _is_etsy_access_blocked_text(value: str) -> bool:
    normalized = str(value or "").lower()
    return any(
        token in normalized
        for token in (
            "verify you are human",
            "access denied",
            "unusual activity",
            "hcaptcha",
            "recaptcha",
        )
    )


def _is_etsy_auth_required_url(url: str) -> bool:
    try:
        parsed = urllib.parse.urlparse(url)
        pathname = (parsed.path or "").lower()
        return any(
            pathname.startswith(segment)
            for segment in (
                "/join",
                "/signin",
                "/sign-in",
            )
        )
    except Exception:
        return False


def _normalize_url_path(path: str) -> str:
    normalized = str(path or "").strip()
    if not normalized:
        return "/"
    normalized = re.sub(r"/+", "/", normalized.strip())
    if not normalized.startswith("/"):
        normalized = f"/{normalized}"
    return normalized.rstrip("/") or "/"


def _is_etsy_host(url_or_host: str) -> bool:
    try:
        parsed = urllib.parse.urlparse(
            url_or_host if "://" in url_or_host else f"//{url_or_host}"
        )
        return (parsed.hostname or "").lower() in {"etsy.com", "www.etsy.com"}
    except Exception:
        return False


def _is_etsy_shop_manager_route(url: str) -> bool:
    try:
        parsed = urllib.parse.urlparse(url)
        if not _is_etsy_host(parsed.netloc):
            return False
        manager_path = _normalize_url_path(urllib.parse.urlparse(SHOP_MANAGER_URL).path)
        return _normalize_url_path(parsed.path) == manager_path
    except Exception:
        return False


def _is_etsy_shop_public_route(url: str) -> bool:
    try:
        parsed = urllib.parse.urlparse(url)
        if not _is_etsy_host(parsed.netloc):
            return False
        return _normalize_url_path(parsed.path).startswith("/shop/")
    except Exception:
        return False


def _is_etsy_access_blocked_url(url: str) -> bool:
    try:
        parsed = urllib.parse.urlparse(url)
        pathname = (parsed.path or "").lower()
        blocked_paths = (
            "/access-denied",
            "/access_denied",
            "/challenge",
            "/session/challenge",
            "/captcha",
        )
        return any(pathname.startswith(item) for item in blocked_paths)
    except Exception:
        return False


async def _read_visible_text(page) -> str:
    for selector in ("main", "[role='main']", "#content", "body"):
        try:
            element = page.locator(selector).first
            if await element.count() > 0:
                text = await element.inner_text()
                if text:
                    return str(text)
        except Exception:
            pass
    try:
        return str(await page.evaluate("() => (document?.body?.innerText || '')")) or ""
    except Exception:
        return ""


async def _extract_shop_anchors(page) -> list[str]:
    return _parse_visible_shop_slugs(
        list(
            await page.evaluate(
                r'''() => {
                const selectors = [
                    "header a[href*='/shop/']",
                    "a[href*='/shop/']",
                    ".shops a[href*='/shop/']",
                    "#content a[href*='/shop/']",
                ];
                const anchors = [];
                for (const selector of selectors) {
                    for (const node of document.querySelectorAll(selector)) {
                        anchors.push(node.href || node.getAttribute('href') || '');
                    }
                }
                return anchors.map((href) => {
                    const match = (href || '').match(/\/shop\/([^/?#]+)/i);
                    return match ? match[1] : null;
                }).filter(Boolean);
            }'''
            )
        )
    )


async def _assert_etsy_page_shop_identity(
    page,
    target_shop_id: str,
    *,
    shop_identity_verified: bool = False,
) -> None:
    expected = _expected_shop_slug(target_shop_id)
    if not expected:
        raise RuntimeError(f"Shop {target_shop_id} chưa có Etsy URL hợp lệ để xác minh phiên đăng nhập")

    anchors = await _extract_shop_anchors(page)
    if not anchors and not shop_identity_verified:
        raise RuntimeError(
            f"Phiên Etsy chưa xác minh: không đọc được shop trên editor, yêu cầu={expected}"
        )
    if not anchors:
        return
    if len(anchors) != 1 or anchors[0] != expected:
        raise RuntimeError(
            f"Phiên Etsy sai shop: editor={', '.join(anchors)} , yêu cầu={expected}"
        )


def _classify_etsy_session_state(url: str, visible_text: str) -> tuple[bool, bool]:
    auth_required = _is_etsy_auth_required_url(url) or _is_etsy_auth_required_text(visible_text)
    access_blocked = _is_etsy_access_blocked_url(url) or _is_etsy_access_blocked_text(visible_text)
    return auth_required, access_blocked


async def _assert_etsy_editor_access(page, listing_id: str) -> None:
    final_url = str(page.url or "").strip()
    text = await _read_visible_text(page)
    auth_required, access_blocked = _classify_etsy_session_state(final_url, text)
    if auth_required:
        raise RuntimeError(
            "Phiên Etsy chưa đăng nhập: hãy mở nút Browser Etsy, đăng nhập đúng shop, sau đó nhấn Sync lại"
        )
    if access_blocked:
        raise RuntimeError("Etsy đang chặn/đòi xác minh phiên này, hãy xác minh thủ công rồi Sync lại")

    opened_listing_id = _extract_manager_listing_id(final_url)
    if not opened_listing_id:
        raise RuntimeError(
            f"Không mở đúng editor listing cho ID {listing_id}. URL hiện tại: {final_url or 'không có'}"
        )
    if opened_listing_id != listing_id:
        raise RuntimeError(
            f"Editor sai listing: editor={opened_listing_id}, yêu cầu={listing_id}"
        )


async def _assert_etsy_shop_manager_preflight(page, target_shop_id: str) -> bool:
    await page.goto(SHOP_MANAGER_URL, wait_until="domcontentloaded", timeout=20000)
    expected = _expected_shop_slug(target_shop_id)
    if not expected:
        raise RuntimeError(f"Shop {target_shop_id} chưa có Etsy URL hợp lệ để xác minh phiên đăng nhập")

    timeout_ms = 8000
    poll_ms = 250
    deadline = time.monotonic() + timeout_ms / 1000.0

    while True:
        final_url = str(page.url or "").strip()
        text = await _read_visible_text(page)
        auth_required, access_blocked = _classify_etsy_session_state(final_url, text)
        if auth_required:
            raise RuntimeError(
                "Phiên Etsy chưa đăng nhập: hãy mở nút Browser Etsy, đăng nhập đúng shop, sau đó nhấn Sync lại"
            )
        if access_blocked:
            raise RuntimeError("Etsy đang chặn/đòi xác minh phiên này, hãy xác minh thủ công rồi Sync lại")
        if _is_etsy_shop_public_route(final_url):
            raise RuntimeError(
                "Phiên Etsy chưa xác minh: không mở đúng Shop Manager, đang ở trang cửa hàng công khai."
            )

        if not _is_etsy_shop_manager_route(final_url):
            if time.monotonic() >= deadline:
                raise RuntimeError(
                    f"Shop Manager chưa sẵn sàng/chưa xác minh: url hiện tại={final_url or 'không có'}"
                )
            await asyncio.sleep(poll_ms / 1000.0)
            continue

        manager_slugs = await _extract_shop_anchors(page)
        if manager_slugs:
            if len(manager_slugs) != 1 or manager_slugs[0] != expected:
                raise RuntimeError(
                    f"Phiên Etsy sai shop: editor={', '.join(manager_slugs)} , yêu cầu={expected}"
                )
            return True

        if time.monotonic() >= deadline:
            raise RuntimeError(
                "Shop Manager chưa sẵn sàng/chưa xác minh: không đọc được shop trên Shop Manager"
            )

        await asyncio.sleep(poll_ms / 1000.0)


async def _assert_etsy_editor_ready(page, listing_id: str) -> None:
    from playwright.async_api import TimeoutError as PlaywrightTimeoutError

    editor_url = f"https://www.etsy.com/your/shops/me/listing-editor/edit/{listing_id}"
    selector = 'textarea[name="title"], input[name="title"], #title-input'

    async def _assert_current_state(exc: Exception | None = None) -> None:
        text = await _read_visible_text(page)
        current_url = str(page.url or "").strip()
        auth_required, access_blocked = _classify_etsy_session_state(current_url, text)
        if auth_required:
            raise RuntimeError(
                "Phiên Etsy chưa đăng nhập: hãy mở nút Browser Etsy, đăng nhập đúng shop, sau đó nhấn Sync lại"
            ) from exc
        if access_blocked:
            raise RuntimeError("Etsy đang chặn/đòi xác minh phiên này, hãy xác minh thủ công rồi Sync lại") from exc
        opened_listing_id = _extract_manager_listing_id(current_url)
        if opened_listing_id != listing_id:
            raise RuntimeError(
                f"Editor sai listing: editor={opened_listing_id or 'không có'}, yêu cầu={listing_id}"
            ) from exc

    last_timeout: Exception | None = None
    for attempt in range(2):
        await _assert_current_state()
        try:
            # Two bounded waits give a transiently slow editor one clean reload
            # while keeping the worst-case selector wait close to the old 20s.
            await page.wait_for_selector(selector, timeout=12000)
            await _assert_current_state()
            return
        except PlaywrightTimeoutError as exc:
            last_timeout = exc
            await _assert_current_state(exc)
            if attempt == 0:
                try:
                    await page.goto(editor_url, wait_until="domcontentloaded", timeout=15000)
                except Exception:
                    # A navigation timeout can still leave a usable hydrated DOM;
                    # the second readiness pass validates route and session again.
                    pass

    raise RuntimeError(
        f"Không nạp được giao diện Listing Editor cho listing {listing_id}; có thể đăng nhập sai shop hoặc trang bị chặn"
    ) from last_timeout


def _validate_etsy_image_bytes(image_bytes: bytes) -> tuple[bool, tuple[int, int] | None]:
    if not image_bytes:
        return False, None
    try:
        from PIL import Image
    except Exception:
        return False, None

    with io.BytesIO(image_bytes) as fp:
        try:
            with Image.open(fp) as img:
                img.load()
                width, height = img.size
                return width > 0 and height > 0, (width, height)
        except Exception:
            return False, None


def _extract_asset_sync_status(
    asset_report: dict,
    *,
    metadata_ok: bool,
    include_assets: bool = True,
) -> dict:
    images_found = int(asset_report.get("images_found") or 0)
    images_downloaded = int(asset_report.get("images_downloaded") or 0)
    files_found = int(asset_report.get("files_found") or 0)
    files_downloaded = int(asset_report.get("files_downloaded") or 0)
    files_section_observed = bool(asset_report.get("files_section_observed", False))

    images_warning = str(asset_report.get("image_warning") or "")
    files_warning = str(asset_report.get("file_warning") or "")

    images_complete = images_found > 0 and images_downloaded == images_found and not bool(images_warning)
    if not files_section_observed:
        files_complete = False
    elif files_found == 0:
        files_complete = not bool(files_warning)
    else:
        files_complete = files_downloaded == files_found and not bool(files_warning)

    return {
        "metadata_ok": bool(metadata_ok),
        "images_complete": bool(images_complete),
        "files_complete": bool(files_complete),
        "assets_complete": bool(images_complete and files_complete),
        "assets_deferred": not bool(include_assets),
        "overall": bool(metadata_ok and (images_complete and files_complete if include_assets else True)),
        "files_section_observed": bool(files_section_observed),
        "image_warning": images_warning or None,
        "file_warning": files_warning or None,
    }


def _parse_request_bool(payload: dict, key: str, default: bool) -> bool:
    """Read optional boolean request flags safely (accept common boolean strings)."""
    if key not in payload:
        return default
    value = payload.get(key)
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        if value in {0, 1}:
            return bool(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "yes", "on"}:
            return True
        if normalized in {"0", "false", "no", "off"}:
            return False
    raise ValueError(f"{key} phải là boolean")


def _size_text_to_bytes(value: str) -> Optional[int]:
    match = re.search(r"([\d.,]+)\s*(B|KB|MB|GB)\b", str(value or ""), re.I)
    if not match:
        return None
    try:
        number = float(match.group(1).replace(",", "."))
    except ValueError:
        return None
    # Etsy displays seller file sizes with decimal units (MB = 1,000,000 B).
    multiplier = {"B": 1, "KB": 1000, "MB": 1000 ** 2, "GB": 1000 ** 3}[match.group(2).upper()]
    return int(number * multiplier)


def _etsy_size_matches(actual_bytes: int, expected_bytes: Optional[int]) -> bool:
    if expected_bytes is None or actual_bytes <= 0:
        return False
    tolerance = max(100_000, int(expected_bytes * 0.05))
    return abs(actual_bytes - expected_bytes) <= tolerance


async def _click_etsy_editor_tab(page, *labels: str) -> bool:
    for label in labels:
        button = page.get_by_role("button", name=label, exact=False).first
        try:
            if await button.count() > 0 and await button.is_visible():
                await button.click()
                await page.wait_for_timeout(900)
                return True
        except Exception:
            continue
    return False


async def _inspect_etsy_listing_assets(page) -> dict:
    """Read Etsy image/file metadata without downloading or changing the listing."""
    await _click_etsy_editor_tab(page, "Photo & Video", "Photos")
    image_counts = []
    for selector in [
        'button[data-testid="image-delete-button"]',
        '[data-testid*="photo" i] button[aria-label*="Remove" i]',
        '[data-testid*="photo" i] button[aria-label*="Delete" i]',
    ]:
        try:
            image_counts.append(await page.locator(selector).count())
        except Exception:
            pass
    image_count = max(image_counts or [0])

    await _click_etsy_editor_tab(page, "Item Details", "Details")
    file_items = await page.evaluate(r'''() => {
        const container = document.getElementById('field-digitalFiles');
        if (!container) return [];
        const items = Array.from(container.querySelectorAll('[data-clg-id="WtUploadItem"], .wt-upload__item'));
        return items.map((item, index) => {
            const text = (item.innerText || item.textContent || '').replace(/\s+/g, ' ').trim();
            const nameNode = item.querySelector('.wt-text-truncate');
            let name = nameNode ? (nameNode.textContent || '').trim() : '';
            if (!name) {
                const match = text.match(/([^\\/]+\.(?:pdf|zip|png|jpe?g|webp|[0-9]{3}))/i);
                name = match ? match[1].trim() : `File ${index + 1}`;
            }
            const sizeMatch = text.match(/([\d.,]+\s*(?:B|KB|MB|GB))\b/i);
            return {name, size_text: sizeMatch ? sizeMatch[1] : ''};
        });
    }''')
    files = []
    for item in file_items:
        size_bytes = _size_text_to_bytes(item.get("size_text", ""))
        files.append({
            "name": str(item.get("name") or ""),
            "size_text": str(item.get("size_text") or ""),
            "size_bytes": size_bytes,
        })
    known_sizes = [item["size_bytes"] for item in files if item["size_bytes"] is not None]
    return {
        "images": {"count": image_count},
        "files": {
            "count": len(files),
            "total_bytes": sum(known_sizes) if len(known_sizes) == len(files) else None,
            "known_size_count": len(known_sizes),
            "items": files,
        },
    }


async def _extract_editor_listing_images(
    editor_page,
    *,
    timeout_ms: int = 8000,
    poll_ms: int = 250,
) -> dict | list[Any]:
    """Prefer authenticated editor HTML for image URLs, fallback later if empty."""
    await _click_etsy_editor_tab(editor_page, "Photo & Video", "Photos")
    try:
        media_button = editor_page.locator('button:has-text("Media"), button:has-text("Photos"), [role="tab"]:has-text("Media")').first
        if await media_button.count() > 0:
            await media_button.click()
            await editor_page.wait_for_timeout(600)
    except Exception:
        pass

    deadline = time.monotonic() + timeout_ms / 1000.0
    last_signature: tuple[int, tuple[str, ...], tuple[str, ...], int] | None = None
    stable_count = 0

    while True:
        payload = await editor_page.evaluate(r'''() => {
            const media_root = document.querySelector('#media') || document.body;
            const deleteButtons = Array.from(media_root.querySelectorAll(
                'button[data-testid="image-delete-button"], [data-testid*="photo" i] button[aria-label*="Delete" i], [data-testid*="photo" i] button[aria-label*="Remove" i]'
            ));
            const uploadImageSelector =
                'img.le-media-preview-image-hide-edges, source.le-media-preview-image-hide-edges, picture.le-media-preview-image-hide-edges';
            const candidateSet = new Set();
            const imageNodes = [];

            const addNode = (node, allowGeneric = false) => {
                if (!node || node.nodeType !== 1) return;
                const testId = node.getAttribute ? (node.getAttribute('data-testid') || '').trim() : '';
                if (testId === 'thumbnail_image') return;
                const tagName = node.tagName.toLowerCase();
                if (tagName === 'source' || allowGeneric) {
                    if (testId === 'thumbnail_image') return;
                    candidateSet.add(node);
                    imageNodes.push(node);
                    return;
                }
                if (node.classList && node.classList.contains('le-media-preview-image-hide-edges')) {
                    candidateSet.add(node);
                    imageNodes.push(node);
                }
            };

            // Prefer the dedicated uploaded-media preview nodes.
            for (const node of Array.from(media_root.querySelectorAll(uploadImageSelector))) {
                addNode(node);
            }

            // Fallback: for robustness, include preview nodes that belong to
            // the same upload item container as a detected delete control.
            if (imageNodes.length === 0 && deleteButtons.length > 0) {
                for (const button of deleteButtons) {
                    const scope = button.closest('[data-photo-id], [data-listing-image-id], [data-clg-id], [class*="photo" i], li, div');
                    if (!scope) continue;
                    for (const node of Array.from(scope.querySelectorAll('img.le-media-preview-image-hide-edges, source.le-media-preview-image-hide-edges, picture.le-media-preview-image-hide-edges, img[src], img[srcset], picture source[src], picture source[srcset]'))) {
                        const testId = (node.getAttribute ? (node.getAttribute('data-testid') || '').trim() : '');
                        if (testId === 'thumbnail_image') continue;
                        addNode(node, true);
                    }
                }
            }

            const fallbackNodes = Array.from(media_root.querySelectorAll(
                'img[src], img[srcset], picture source[src], picture source[srcset]'
            ));
            if (imageNodes.length === 0 && fallbackNodes.length > 0) {
                for (const node of fallbackNodes) {
                    const testId = (node.getAttribute ? (node.getAttribute('data-testid') || '').trim() : '');
                    if (testId === 'thumbnail_image') continue;
                    if (candidateSet.has(node)) {
                        continue;
                    }
                    addNode(node, true);
                }
            }

            return {
                image_nodes: imageNodes.map((node, position) => {
                    const index = (
                        (node.getAttribute('data-index') ||
                        node.getAttribute('data-photo-id') ||
                        node.getAttribute('data-listing-image-id') ||
                        String(position)).trim()
                    );
                    const candidates = [];
                    const add = (value) => {
                        if (!value) return;
                        for (const token of String(value).split(',').map((value) => value.trim()).filter(Boolean)) {
                            const first = token.split(/\s+/)[0].trim();
                            if (first) {
                                candidates.push(first);
                            }
                        }
                    };

                    add(node.currentSrc || '');
                    add(node.getAttribute('src') || '');
                    add(node.getAttribute('srcset') || '');
                    add(node.getAttribute('data-src') || '');
                    add(node.getAttribute('data-srcset') || '');
                    return {index, candidates};
                }),
                delete_count: deleteButtons.length,
            };
        }''')
        payload = payload if isinstance(payload, dict) else {}
        image_nodes = payload.get("image_nodes")
        image_nodes = image_nodes if isinstance(image_nodes, list) else []

        normalized_urls: list[str] = []
        seen_urls: set[str] = set()
        indices: list[str] = []
        for position, node in enumerate(image_nodes):
            index = str(position).strip()
            candidates: list[str] = []
            if isinstance(node, dict):
                index = str(node.get("index") or position).strip()
                candidates = list(node.get("candidates") or [])

            if index:
                indices.append(index)
            for raw in candidates:
                clean = str(raw).split("?")[0].split("#")[0].strip()
                if not clean or "/r/il/" not in clean:
                    continue
                normalized = _to_il_fullxfull_url(clean)
                if not normalized or normalized in seen_urls:
                    continue
                seen_urls.add(normalized)
                normalized_urls.append(normalized)
                if len(normalized_urls) >= 20:
                    break
            if len(normalized_urls) >= 20:
                break

        gallery_node_count = len(image_nodes)
        unique_indices = sorted(set(indices))
        delete_count = int(payload.get("delete_count", 0) or 0)
        editor_image_count = delete_count if delete_count > 0 else len(unique_indices)
        complete = (
            gallery_node_count > 0
            and len(normalized_urls) == len(unique_indices) == delete_count == gallery_node_count
        )
        signature = (
            gallery_node_count,
            tuple(unique_indices),
            tuple(normalized_urls),
            delete_count,
        )

        if complete:
            if signature == last_signature:
                stable_count += 1
            else:
                stable_count = 1
                last_signature = signature
            if stable_count >= 2:
                return {
                    "images": normalized_urls,
                    "complete": True,
                    "image_count": editor_image_count,
                    "gallery_node_count": gallery_node_count,
                    "gallery_index_count": len(unique_indices),
                    "delete_count": delete_count,
                }
        else:
            stable_count = 0
            last_signature = None

        if time.monotonic() >= deadline:
            return {
                "images": normalized_urls,
                "complete": False,
                "image_count": editor_image_count,
                "gallery_node_count": gallery_node_count,
                "gallery_index_count": len(unique_indices),
                "delete_count": delete_count,
                "reason": (
                    "incomplete_editor_hydration"
                    if len(unique_indices) != len(normalized_urls)
                    or delete_count != gallery_node_count
                    or gallery_node_count != len(normalized_urls)
                    else "timeout"
                ),
            }

        await asyncio.sleep(poll_ms / 1000.0)


async def _extract_public_listing_images(
    public_page,
    *,
    timeout_ms: int = 8000,
    poll_ms: int = 250,
) -> dict | list[Any]:
    deadline = time.monotonic() + timeout_ms / 1000.0
    last_signature: tuple[int, tuple[str, ...], tuple[str, ...]] | None = None
    stable_count = 0

    while True:
        gallery_payload = await public_page.evaluate(r"""() => {
            const photos = document.querySelector("#photos");
            if (!photos) return [];
            const gallery_root = photos.querySelector(".listing-page-image-carousel-component") || photos;
            const gallery_nodes = Array.from(gallery_root.querySelectorAll("img.carousel-image"));

            return gallery_nodes.map((node, position) => {
                const index = (node.getAttribute("data-index") || String(position)).trim();
                const candidates = [];

                const add = (raw) => {
                    if (!raw) return;
                    for (const token of String(raw).split(",").map((value) => value.trim()).filter(Boolean)) {
                        const first = token.split(/\s+/)[0].trim();
                        if (first) {
                            candidates.push(first);
                        }
                    }
                };

                add(node.currentSrc || "");
                add(node.getAttribute("src") || "");
                add(node.getAttribute("data-src") || "");
                add(node.getAttribute("srcset") || "");
                add(node.getAttribute("data-srcset") || "");
                return { index, candidates };
            });
        }""")

        gallery_payload = gallery_payload if isinstance(gallery_payload, list) else []

        normalized_urls: list[str] = []
        seen_urls: set[str] = set()
        indices: list[str] = []
        for position, node in enumerate(gallery_payload):
            index = str(position).strip()
            candidates: list[str] = []
            if isinstance(node, dict):
                index = str((node.get("index") or position)).strip()
                candidates = list(node.get("candidates") or [])
            elif isinstance(node, str):
                candidates = [node]
            else:
                index = str((node.get("index") if hasattr(node, "get") else position)).strip()
                candidates = list(getattr(node, "candidates", []) or [])

            if index:
                indices.append(index)
            for raw in candidates:
                clean = str(raw).split("?")[0].split("#")[0].strip()
                if not clean or "/r/il/" not in clean:
                    continue
                normalized = _to_il_fullxfull_url(clean)
                if not normalized or normalized in seen_urls:
                    continue
                seen_urls.add(normalized)
                normalized_urls.append(normalized)
                if len(normalized_urls) >= 20:
                    break
            if len(normalized_urls) >= 20:
                break

        gallery_node_count = len(gallery_payload)
        unique_indices = sorted(set(indices))
        complete = (
            gallery_node_count > 0
            and len(normalized_urls) == len(unique_indices) == gallery_node_count
        )
        signature = (gallery_node_count, tuple(unique_indices), tuple(normalized_urls))

        if complete:
            if signature == last_signature:
                stable_count += 1
            else:
                stable_count = 1
                last_signature = signature
            if stable_count >= 2:
                return {
                    "urls": normalized_urls,
                    "complete": True,
                    "gallery_node_count": gallery_node_count,
                    "gallery_index_count": len(unique_indices),
                }
        else:
            stable_count = 0
            last_signature = None

        if time.monotonic() >= deadline:
            return {
                "urls": normalized_urls,
                "complete": False,
                "gallery_node_count": gallery_node_count,
                "gallery_index_count": len(unique_indices),
                "reason": (
                    "incomplete_gallery_hydration"
                    if len(unique_indices) != len(normalized_urls)
                    else "timeout"
                ),
            }

        await asyncio.sleep(poll_ms / 1000.0)


@asynccontextmanager
async def _redirect_cdp_downloads_to_staging(browser_ctx, page, download_dir: Path):
    """Keep seller-button downloads inside product-local staging.

    A browser attached over CDP otherwise follows the Chrome profile's normal
    download directory even when Playwright later copies the artifact with
    ``download.save_as``.  Scope the override to the click and always restore
    Chrome's default behavior before detaching the temporary CDP session.
    """
    download_dir.mkdir(parents=True, exist_ok=True)
    cdp_session = await browser_ctx.new_cdp_session(page)
    redirect_enabled = False
    try:
        await cdp_session.send(
            "Page.setDownloadBehavior",
            {
                "behavior": "allow",
                "downloadPath": str(download_dir.resolve()),
            },
        )
        redirect_enabled = True
        yield
    finally:
        try:
            if redirect_enabled:
                await cdp_session.send(
                    "Page.setDownloadBehavior",
                    {"behavior": "default"},
                )
        finally:
            await cdp_session.detach()


def _snapshot_staged_files(directory: Path) -> dict[Path, tuple[int, int]]:
    """Capture stable file signatures for a staging directory."""
    result: dict[Path, tuple[int, int]] = {}
    for candidate in directory.iterdir():
        if not candidate.is_file():
            continue
        try:
            stat = candidate.stat()
        except OSError:
            continue
        result[candidate.resolve()] = (stat.st_size, stat.st_mtime_ns)
    return result


_STAGED_DOWNLOAD_POLL_MS = 250
_STAGED_DOWNLOAD_MAX_ATTEMPTS = 50  # 50 x 250ms ~= 12.5s bounded wait for large assets
_STAGED_DOWNLOAD_MIN_SIZE_BYTES = 100
_STAGED_DOWNLOAD_STABLE_ROUNDS = 2


async def _await_stable_matching_staged_download(
    staging_dir: Path,
    *,
    baseline: dict[Path, tuple[int, int]],
    expected_size: int | None,
    max_attempts: int = _STAGED_DOWNLOAD_MAX_ATTEMPTS,
    poll_ms: int = _STAGED_DOWNLOAD_POLL_MS,
    min_size: int = _STAGED_DOWNLOAD_MIN_SIZE_BYTES,
    stable_rounds: int = _STAGED_DOWNLOAD_STABLE_ROUNDS,
) -> list[Path]:
    """Wait for a new/changed staged file that is stable and size-matched."""
    if expected_size is None:
        return []
    previous_sizes: dict[Path, int] = {}
    stable_rounds_seen: dict[Path, int] = {}

    for _ in range(max_attempts):
        candidates: list[Path] = []
        for candidate in staging_dir.iterdir():
            if not candidate.is_file():
                continue
            try:
                stat = candidate.stat()
            except OSError:
                continue
            current_signature = (stat.st_size, stat.st_mtime_ns)
            candidate_key = candidate.resolve()
            if baseline.get(candidate_key) == current_signature:
                continue
            if stat.st_size < max(min_size, 1):
                continue
            if not _etsy_size_matches(stat.st_size, expected_size):
                continue
            previous_size = previous_sizes.get(candidate_key)
            if previous_size is None or previous_size != stat.st_size:
                previous_sizes[candidate_key] = stat.st_size
                stable_rounds_seen[candidate_key] = 1
                continue
            stable_rounds_seen[candidate_key] = stable_rounds_seen.get(candidate_key, 1) + 1
            if stable_rounds_seen[candidate_key] >= stable_rounds:
                candidates.append(candidate)

        if candidates:
            return sorted(candidates, key=lambda candidate: candidate.name)

        await asyncio.sleep(poll_ms / 1000.0)

    return []


async def _sync_listing_assets(browser_ctx, editor_page, listing_id: str, product_path: Path) -> dict:
    """Download listing photos (editor-first) and seller-downloadable digital files."""
    images_dir = product_path / "images"
    files_dir = product_path / "files"

    sync_staging = product_path / ".sync_staging"
    image_staging = sync_staging / "images"
    file_staging = sync_staging / "files"

    images_dir.mkdir(parents=True, exist_ok=True)
    files_dir.mkdir(parents=True, exist_ok=True)
    if sync_staging.exists():
        shutil.rmtree(sync_staging, ignore_errors=True)
    image_staging.mkdir(parents=True, exist_ok=True)
    file_staging.mkdir(parents=True, exist_ok=True)

    def _commit_staged_entries(staged_entries: list[tuple[Path, Path]]) -> None:
        backups: list[tuple[Path, Path]] = []
        created: list[Path] = []
        try:
            for stage_src, final_dst in staged_entries:
                final_dst.parent.mkdir(parents=True, exist_ok=True)
                if final_dst.exists():
                    backup = final_dst.with_name(f"{final_dst.name}.merge-backup")
                    if not backup.exists():
                        shutil.move(final_dst, backup)
                        backups.append((backup, final_dst))
                shutil.move(stage_src, final_dst)
                created.append(final_dst)
        except Exception:
            for final_dst in reversed(created):
                if final_dst.exists():
                    final_dst.unlink()
            for backup, final_dst in reversed(backups):
                if final_dst.exists():
                    final_dst.unlink()
                backup.rename(final_dst)
            raise
        finally:
            for backup, _ in backups:
                if backup.exists():
                    backup.unlink()

    def _unstage_entries(staged_entries: list[tuple[Path, Path]]) -> None:
        for stage_src, _ in staged_entries:
            if stage_src.exists():
                stage_src.unlink(missing_ok=True)

    def _stage_file(stage_src: Path, final_name: str, staged: list[tuple[Path, Path]]) -> None:
        final_dst = files_dir / final_name
        staged.append((stage_src, final_dst))

    def _stage_image(stage_src: Path, final_name: str, staged: list[tuple[Path, Path]]) -> None:
        final_dst = images_dir / final_name
        staged.append((stage_src, final_dst))

    report = {
        "images_found": 0,
        "images_downloaded": 0,
        "files_found": 0,
        "files_downloaded": 0,
        "files_recovered_local": 0,
        "file_names": [],
        "file_sizes": {},
        "file_warning": "",
        "image_warning": "",
        "images_source": "editor",
        "public_image_status": None,
        "files_section_observed": False,
    }

    image_payload = {}
    image_urls: list[str] = []
    editor_expected_images = 0
    staged_image_entries: list[tuple[Path, Path]] = []
    editor_complete = False
    editor_payload_had_urls = False

    try:
        image_payload = await _extract_editor_listing_images(editor_page)
        if isinstance(image_payload, dict):
            image_urls = [str(v) for v in image_payload.get("images", []) if v]
            editor_expected_images = int(image_payload.get("image_count") or 0)
            if "complete" in image_payload:
                editor_complete = bool(image_payload.get("complete"))
            elif editor_expected_images and editor_expected_images == len(image_urls):
                editor_complete = True
        else:
            image_urls = [str(v) for v in image_payload or [] if v]
            editor_expected_images = len(image_urls)
            editor_complete = True

        editor_payload_had_urls = bool(image_urls)

        if image_urls and not editor_complete:
            editor_warning = image_payload.get("reason") if isinstance(image_payload, dict) else None
            report["image_warning"] = (
                editor_warning
                or (
                    f"Ảnh editor chưa hydrate đủ cho listing {listing_id}: "
                    f"{len(image_urls)} ảnh / {editor_expected_images or len(image_urls)} nút/ảnh."
                )
            )
            if not editor_warning:
                report["image_warning"] += " Không commit ảnh một phần."
            image_urls = []
        if image_urls:
            report["images_found"] = max(editor_expected_images, len(image_urls))
            if editor_expected_images and editor_expected_images != len(image_urls):
                report["image_warning"] = (
                    f"Số ảnh trên editor ({editor_expected_images}) không khớp số URL duy nhất ({len(image_urls)}). "
                    "Không commit ảnh một phần."
                )
        elif not editor_payload_had_urls:
            report["image_warning"] = "Không đọc được ảnh từ editor; chuyển qua public page."
    except Exception as exc:
        report["image_warning"] = f"Không đọc được ảnh từ editor: {exc}"

    if image_urls:
        for index, url in enumerate(image_urls, 1):
            try:
                response = await browser_ctx.request.get(url, timeout=30000)
                body = await response.body()
                content_type = str(response.headers.get("content-type") or "").lower()
                if not response.ok or "image/" not in content_type:
                    broadcast(f"[ETSY-ASSET] ⚠️ Ảnh {index}: phản hồi không hợp lệ (status={response.status})")
                    continue
                is_valid_image, _ = _validate_etsy_image_bytes(body)
                if not is_valid_image:
                    broadcast(f"[ETSY-ASSET] ⚠️ Ảnh {index}: ảnh tải về hỏng/không decode được")
                    continue
                ext = ".png" if "png" in content_type else ".webp" if "webp" in content_type else ".jpg"
                stage_file = image_staging / f"etsy_{index:02d}{ext}.part"
                stage_file.write_bytes(body)
                _stage_image(stage_file, f"etsy_{index:02d}{ext}", staged_image_entries)
                report["images_downloaded"] += 1
            except Exception as exc:
                broadcast(f"[ETSY-ASSET] ⚠️ Ảnh {index}: {exc}")
            await asyncio.sleep(0)

        if report["images_downloaded"] == report["images_found"] and not report["image_warning"]:
            _commit_staged_entries(staged_image_entries)
            committed_targets = {target for _, target in staged_image_entries}
            for stale in sorted(images_dir.glob("etsy_*")):
                if stale not in committed_targets:
                    stale.unlink(missing_ok=True)
        else:
            _unstage_entries(staged_image_entries)
            report["images_downloaded"] = 0

    public_page = None
    if not image_urls:
        public_page = await browser_ctx.new_page()
        try:
            public_url = f"https://www.etsy.com/listing/{listing_id}"
            response = await public_page.goto(
                public_url,
                wait_until="domcontentloaded",
                timeout=30000,
            )
            if response is not None:
                report["public_image_status"] = int(response.status) if hasattr(response, "status") else None
            final_url = str(public_page.url or "").strip()
            if not _safe_etsy_public_url(final_url, listing_id):
                report["image_warning"] = "Public page mở sai listing so với request, không cho phép fallback."
                raise RuntimeError(
                    f"Không đọc được ảnh public: public url khác listing {listing_id} -> {final_url or public_url}"
                )
            if response is None:
                report["image_warning"] = "Không đọc được phản hồi từ public page."
                raise RuntimeError("Public page navigation failed")
            if response.status == 403:
                report["image_warning"] = "Public page bị chặn bởi DataDome (403). Không thể kiểm tra ảnh công khai."
                raise RuntimeError("DataDome challenge on public listing page")
            elif response.status >= 400:
                report["image_warning"] = f"Public page lỗi {response.status}, không thể kiểm tra ảnh."
            else:
                report["images_source"] = "public"
                public_payload = await _extract_public_listing_images(public_page)
                if isinstance(public_payload, dict):
                    gallery_node_count = int(public_payload.get("gallery_node_count", 0))
                    image_urls = [str(v) for v in public_payload.get("urls", []) if v]
                    if public_payload.get("complete"):
                        report["image_warning"] = ""
                    elif gallery_node_count > len(image_urls):
                        report["image_warning"] = (
                            f"Gallery ảnh công khai chưa hydrate đủ cho listing {listing_id}: "
                            f"{len(image_urls)} ảnh / {gallery_node_count} nút."
                        )
                    else:
                        report["image_warning"] = (
                            "Không đọc được gallery ảnh công khai trong #photos của listing "
                            f"{listing_id}: {public_payload.get('reason') or 'không thu thập đủ'}"
                        )
                else:
                    image_urls = [str(v) for v in public_payload or [] if v]
                    report["image_warning"] = (
                        "Không đọc được gallery ảnh công khai trong #photos của listing "
                        f"{listing_id}."
                    )
                report["images_found"] = len(image_urls)
                staged_image_entries = []
                for index, url in enumerate(image_urls, 1):
                    try:
                        response = await browser_ctx.request.get(url, timeout=30000)
                        body = await response.body()
                        content_type = str(response.headers.get("content-type") or "").lower()
                        if response.ok and "image/" in content_type and len(body) > 1024:
                            is_valid_image, _ = _validate_etsy_image_bytes(body)
                            if not is_valid_image:
                                raise RuntimeError("Ảnh công khai tải về hỏng")
                            ext = ".png" if "png" in content_type else ".webp" if "webp" in content_type else ".jpg"
                            stage_file = image_staging / f"etsy_{index:02d}{ext}.part"
                            stage_file.write_bytes(body)
                            _stage_image(stage_file, f"etsy_{index:02d}{ext}", staged_image_entries)
                            report["images_downloaded"] += 1
                        else:
                            broadcast(f"[ETSY-ASSET] ⚠️ Public Ảnh {index}: phản hồi không hợp lệ (status={response.status})")
                    except Exception as exc:
                        broadcast(f"[ETSY-ASSET] ⚠️ Public Ảnh {index}: {exc}")
                if report["images_downloaded"] == report["images_found"] and not report["image_warning"]:
                    _commit_staged_entries(staged_image_entries)
                    committed_targets = {target for _, target in staged_image_entries}
                    for stale in sorted(images_dir.glob("etsy_*")):
                        if stale not in committed_targets:
                            stale.unlink(missing_ok=True)
                else:
                    _unstage_entries(staged_image_entries)
                    report["images_downloaded"] = 0
        except Exception as exc:
            if not report.get("image_warning"):
                report["image_warning"] = f"Không đọc được ảnh public: {exc}"
        finally:
            if public_page:
                await public_page.close()

    # Etsy only exposes digital files inside the seller editor. Download links
    # differ between editor versions, so support both direct links and buttons.
    staged_file_entries: list[tuple[Path, Path]] = []
    try:
        file_container = editor_page.locator("#field-digitalFiles")
        if await file_container.count() == 0:
            details_tab = editor_page.locator('button:has-text("Item Details"), button:has-text("Details")')
            if await details_tab.count() > 0:
                await details_tab.first.click()
                await editor_page.wait_for_timeout(1200)
        file_items = editor_page.locator(
            '#field-digitalFiles [data-clg-id="WtUploadItem"], #field-digitalFiles .wt-upload__item'
        )
        item_count = await file_items.count()
        report["files_section_observed"] = await file_container.count() > 0
        report["files_found"] = item_count

        resolved_file_names = set()
        expected_file_sizes: dict[str, Optional[int]] = {}
        for index in range(item_count):
            item = file_items.nth(index)
            file_info = await item.evaluate(r'''item => {
                const nameEl = item.querySelector('.wt-text-truncate');
                let name = nameEl ? (nameEl.textContent || '').trim() : '';
                if (!name) {
                    const text = (item.textContent || '').split('\n').map(v => v.trim()).find(v => /\.[a-z0-9]{2,5}$/i.test(v));
                    name = text || '';
                }
                const text = (item.innerText || item.textContent || '').replace(/\s+/g, ' ').trim();
                const sizeMatch = text.match(/([\d.,]+\s*(?:B|KB|MB|GB))\b/i);
                const link = item.querySelector('a[download], a[href*="download" i], a[href*="file" i]');
                return {name, href: link ? link.href : '', size_text: sizeMatch ? sizeMatch[1] : ''};
            }''')
            file_name = _safe_asset_filename(file_info.get("name"), f"etsy_file_{index + 1:02d}")
            expected_size = _size_text_to_bytes(file_info.get("size_text", ""))
            report["file_names"].append(file_name)
            report["file_sizes"][file_name] = expected_size
            expected_file_sizes[file_name] = expected_size
            downloaded = False

            href = str(file_info.get("href") or "")
            if href.startswith("https://"):
                try:
                    response = await browser_ctx.request.get(href, timeout=60000)
                    body = await response.body()
                    content_type = str(response.headers.get("content-type") or "").lower()
                    if (response.ok and len(body) > 100 and "text/html" not in content_type
                            and (expected_size is None or _etsy_size_matches(len(body), expected_size))):
                        stage_file = file_staging / f"{file_name}.part"
                        stage_file.write_bytes(body)
                        _stage_file(stage_file, file_name, staged_file_entries)
                        report["files_downloaded"] += 1
                        resolved_file_names.add(file_name)
                        downloaded = True
                except Exception as exc:
                    broadcast(f"[ETSY-ASSET] ⚠️ File {file_name}: {exc}")

            if not downloaded:
                download_control = item.locator('button[data-testid="digital_file_action_download"]')
                if await download_control.count() == 0:
                    download_control = item.locator(
                        'button[aria-label*="download" i], a[aria-label*="download" i], button[title*="download" i]'
                    )
                if await download_control.count() == 1:
                    try:
                        staging_before = _snapshot_staged_files(file_staging)
                        async with _redirect_cdp_downloads_to_staging(
                            browser_ctx,
                            editor_page,
                            file_staging,
                        ):
                            async with editor_page.expect_download(timeout=10000) as download_info:
                                await download_control.click()
                            download = await download_info.value
                            suggested = _safe_asset_filename(download.suggested_filename, file_name)
                            stage_file = file_staging / f"{suggested}.part"
                            await download.save_as(str(stage_file))

                        redirected_candidates = await _await_stable_matching_staged_download(
                            file_staging,
                            baseline=staging_before,
                            expected_size=expected_size,
                        )
                        if len(redirected_candidates) != 1:
                            stage_file.unlink(missing_ok=True)
                            raise RuntimeError(
                                "Không xác định duy nhất file Chrome tải hoàn chỉnh "
                                f"(save_as={stage_file.stat().st_size if stage_file.exists() else 0}, "
                                f"Etsy={expected_size}, "
                                f"candidates={len(redirected_candidates)})"
                            )

                        redirected = redirected_candidates[0]
                        if redirected != stage_file:
                            stage_file.unlink(missing_ok=True)
                            recovered_stage = file_staging / f".redirected-{index + 1:02d}-{file_name}.part"
                            if recovered_stage.exists():
                                raise RuntimeError(f"Staging collision cho {file_name}")
                            redirected.rename(recovered_stage)
                            stage_file = recovered_stage

                        actual_size = stage_file.stat().st_size if stage_file.exists() else 0
                        if actual_size <= 100 or not _etsy_size_matches(actual_size, expected_size):
                            stage_file.unlink(missing_ok=True)
                            raise RuntimeError(
                                f"Dung lượng tải về không khớp Etsy: local={actual_size}, Etsy={expected_size}"
                            )
                        _stage_file(stage_file, file_name, staged_file_entries)
                        report["files_downloaded"] += 1
                        resolved_file_names.add(file_name)
                    except Exception as exc:
                        broadcast(f"[ETSY-ASSET] ⚠️ Nút tải {file_name}: {exc}")

        for file_name in report["file_names"]:
            if file_name in resolved_file_names:
                continue
            expected_size = expected_file_sizes.get(file_name)
            target = files_dir / file_name
            if (target.exists() and target.stat().st_size > 100
                    and _etsy_size_matches(target.stat().st_size, expected_size)):
                stage_file = file_staging / f"{file_name}.part"
                stage_file.write_bytes(target.read_bytes())
                _stage_file(stage_file, file_name, staged_file_entries)
                report["files_downloaded"] += 1
                resolved_file_names.add(file_name)
                continue

            candidates = []
            for candidate in BASE_DIR.glob(f"shops/*/product-*/files/{file_name}") if expected_size is not None else []:
                try:
                    if (candidate.resolve() != target.resolve()
                            and _etsy_size_matches(candidate.stat().st_size, expected_size)):
                        candidates.append(candidate)
                except OSError:
                    continue
            if candidates:
                digests: dict[str, list[Path]] = {}
                for candidate in candidates:
                    digest = hashlib.sha256(candidate.read_bytes()).hexdigest()
                    digests.setdefault(digest, []).append(candidate)
                if len(digests) == 1:
                    source = candidates[0]
                    stage_file = file_staging / f"{file_name}.part"
                    shutil.copy2(source, stage_file)
                    _stage_file(stage_file, file_name, staged_file_entries)
                    report["files_downloaded"] += 1
                    report["files_recovered_local"] += 1
                    resolved_file_names.add(file_name)
                    broadcast(f"[ETSY-ASSET] ✅ Khôi phục {file_name} từ bản local trùng hash: {source}")

        if report["files_section_observed"]:
            if report["files_found"] == report["files_downloaded"] and not report["file_warning"]:
                _commit_staged_entries(staged_file_entries)
        else:
            _unstage_entries(staged_file_entries)
            report["files_downloaded"] = 0

        if not report.get("file_warning") and report["files_found"] > report["files_downloaded"]:
            report["file_warning"] = (
                "Etsy editor chỉ hiển thị tên file, không cung cấp nút/link tải xuống cho seller; "
                "không tìm thấy bản local trùng khớp để khôi phục."
            )
    except Exception as exc:
        report["file_warning"] = f"Không đọc được vùng Digital files: {exc}"

    # Prevent false-OK when listing has no extractable source due blocked/editor miss.
    if report.get("images_source") == "editor" and not image_urls:
        report["images_found"] = report.get("images_found") or 0

    if sync_staging.exists():
        shutil.rmtree(sync_staging, ignore_errors=True)

    return report




async def scrape_listing_details(
    listing_id: str,
    shop_id: Optional[str] = None,
    product_path: Optional[Path] = None,
    include_asset_summary: bool = False,
) -> dict:
    from playwright.async_api import async_playwright

    target_shop_id = shop_id or _active_shop_id
    if shop_id:
        _assert_shop_identity(target_shop_id)
    session = resolve_etsy_session(BASE_DIR, SHOPS, target_shop_id)
    scrape_log = _ctx_logger("scrape_listing_details", shop=target_shop_id, listing_id=listing_id)
    session_verified = False
    for attempt in range(1, 4):
        session_verified = await asyncio.to_thread(is_etsy_session_ready, session)
        if session_verified:
            break
        if attempt < 3:
            await asyncio.sleep(0.2 * attempt)
    scrape_log.info("scrape start (session_ready=%s)", session_verified)

    # 1. Chỉ dùng đúng session theo resolve_etsy_session.
    if not session_verified:
        scrape_log.error(
            "exact Etsy Chrome profile/CDP unavailable (profile=%s, cdp=%s)",
            session.profile_dir,
            session.cdp_url,
        )
        raise RuntimeError(
            f"Không truy cập được Chrome profile={session.profile_dir} "
            f"tại CDP={session.cdp_url} cho shop={target_shop_id} sau 3 lần kiểm tra. "
            "Hãy bảo đảm đúng cửa sổ Browser Etsy vẫn đang mở, rồi Sync lại."
        )

    pw = await async_playwright().start()
    browser_ctx = None
    connected_cdp = False

    def _requires_cdp_no_defaults_fallback(error: Exception) -> bool:
        """Detect Chrome builds that reject Playwright's global download setup.

        The single-sync downloader configures downloads per page immediately
        before clicking a seller file.  Some current Chrome builds reject the
        global ``Browser.setDownloadBehavior`` sent while Playwright attaches,
        even though the exact CDP session itself is healthy.
        """
        message = str(error)
        return (
            "Browser.setDownloadBehavior" in message
            and "Browser context management is not supported" in message
        )

    # 2. Connect chính xác vào CDP của phiên đã xác thực.
    #    Retry a few short times if the CDP endpoint is busy/transiently
    #    unavailable. Recheck session readiness between retries to avoid
    #    continuing against a no-longer-ready browser profile.
    connect_attempts = 0
    last_error = None
    for connect_attempts in range(1, 4):
        if connect_attempts > 1:
            session_verified = await asyncio.to_thread(is_etsy_session_ready, session)
            if not session_verified:
                await asyncio.sleep(0.2 * connect_attempts)
                continue
        try:
            browser = await pw.chromium.connect_over_cdp(session.cdp_url, timeout=10000)
            browser_ctx = browser.contexts[0] if browser.contexts else await browser.new_context()
            connected_cdp = True
            break
        except Exception as exc:
            last_error = exc
            if _requires_cdp_no_defaults_fallback(exc):
                scrape_log.info(
                    "CDP attach rejected Playwright global download defaults; retrying exact session without them"
                )
                try:
                    browser = await pw.chromium.connect_over_cdp(
                        session.cdp_url,
                        timeout=10000,
                        no_defaults=True,
                    )
                    browser_ctx = browser.contexts[0] if browser.contexts else await browser.new_context()
                    connected_cdp = True
                    break
                except Exception as fallback_exc:
                    last_error = fallback_exc
            scrape_log.warning(
                "CDP connect attempt %s failed for shop=%s cdp=%s: %s",
                connect_attempts,
                target_shop_id,
                session.cdp_url,
                last_error,
            )
            if connect_attempts < 3:
                await asyncio.sleep(0.25 * connect_attempts)
                continue
            scrape_log.error("failed to connect Etsy CDP session via %s", session.cdp_url)

    if not connected_cdp:
        await pw.stop()
        if last_error is not None:
            raise RuntimeError(
                f"CDP Etsy đang bận hoặc không phản hồi cho shop={target_shop_id} "
                f"tại {session.cdp_url} (profile={session.profile_dir}). "
                "Hãy chờ lượt Sync hiện tại xong hoặc đóng bớt tab Etsy, rồi thử lại."
            ) from last_error
        raise RuntimeError(
            f"CDP Etsy đang bận hoặc không phản hồi cho shop={target_shop_id} "
            f"tại {session.cdp_url} (profile={session.profile_dir}). "
            "Hãy chờ lượt Sync hiện tại xong hoặc đóng bớt tab Etsy, rồi thử lại."
        )

    if browser_ctx is None:
        await pw.stop()
        raise RuntimeError("Không khởi tạo được browser context cho phiên Etsy đã xác thực")

    # 3. Dùng profile + cdp này rồi xác nhận không được fallback sang bundled/headless.
    #   (Không mở context local khi CDP chưa sẵn sàng)
    if not connected_cdp:
        scrape_log.error("no active CDP session available for scraping")
        raise RuntimeError(
            "Không có phiên CDP hoạt động cho scraping. Hãy mở Browser Etsy đúng shop, rồi thử lại."
        )

    try:
        page = None
        for p in browser_ctx.pages:
            if "etsy.com" in p.url:
                page = p
                break
        if page is None:
            page = await browser_ctx.new_page()
        # 3. Shop-manager identity preflight trong cùng browser context/page.
        #    Tránh fail-open khi profile đúng nhưng account/shop không khớp.
        shop_identity_verified = await _assert_etsy_shop_manager_preflight(page, target_shop_id)
        edit_url = f"https://www.etsy.com/your/shops/me/listing-editor/edit/{listing_id}"
        try:
            await page.goto(edit_url, wait_until="commit", timeout=20000)
        except Exception as e:
            scrape_log.warning("navigation timeout/error opening listing editor: %s", e)
        await _assert_etsy_editor_access(page, listing_id)
        # Verify editor route chính xác cho listing_id trước khi chốt metadata.
        opened_listing_id = _extract_manager_listing_id(str(page.url or "").strip())
        if not opened_listing_id:
            scrape_log.error("editor opened without a listing id in URL: %s", page.url or "")
            raise RuntimeError(
                f"Không mở đúng editor listing cho ID {listing_id}. URL hiện tại: {page.url or 'không có'}"
            )
        if opened_listing_id != listing_id:
            scrape_log.error("editor opened wrong listing: editor=%s requested=%s", opened_listing_id, listing_id)
            raise RuntimeError(
                f"Editor sai listing: editor={opened_listing_id}, yêu cầu={listing_id}"
            )
        await _assert_etsy_editor_ready(page, listing_id)
        await _assert_etsy_page_shop_identity(
            page,
            target_shop_id,
            shop_identity_verified=shop_identity_verified,
        )

        # Scrape Title
        title = ""
        for sel in ['textarea[name="title"]', 'input[name="title"]', '#title-input', 'textarea[id*="title"]']:
            el = page.locator(sel).first
            if await el.count() > 0:
                title = (await el.input_value()).strip()
                if title:
                    break

        # Scrape Description
        desc = ""
        for sel in ['textarea[name="description"]', '#description-textarea', 'textarea[id*="description"]']:
            el = page.locator(sel).first
            if await el.count() > 0:
                desc = (await el.input_value()).strip()
                if desc:
                    break

        # Scrape Price
        price_val = None
        for sel in ['#listing-price-input', '[data-testid="price-input"]', 'input[name="price"]']:
            el = page.locator(sel).first
            if await el.count() > 0:
                try:
                    price_str = await el.input_value()
                    price_val = float(price_str.replace("$", "").strip())
                    break
                except:
                    pass

        # Scrape Tags
        raw_tags = await page.evaluate(r'''() => {
            let tagSection = null;
            let allEls = Array.from(document.querySelectorAll("legend, label, h2, h3, p, span"));
            for (let el of allEls) {
                if ((el.innerText || "").trim() === "Tags") {
                    tagSection = el.closest("fieldset") || el.parentElement;
                    break;
                }
            }
            if (!tagSection) return [];
            let tags = [];
            let seen = new Set();
            for (let el of Array.from(tagSection.querySelectorAll("button, span, div, li"))) {
                let txt = (el.innerText || "").replace(/[×✕]/g, "").trim();
                if (txt.length >= 2 && txt.length <= 40 && !txt.includes("\n") && /[a-z]/i.test(txt) && txt.split(" ").length <= 5 && !seen.has(txt.toLowerCase()) && !["Add", "Tags", "Remove", "Add tag", "used", "left", "Add up to", "Shape, color"].some(k => txt.toLowerCase().startsWith(k.toLowerCase()))) {
                    seen.add(txt.toLowerCase());
                    tags.push(txt);
                }
            }
            return tags;
        }''')
        tags = ", ".join([t for t in raw_tags if t][:13])

        # Scrape Quantity
        qty_val = None
        for sel in ['#listing-quantity-input', 'input[name="quantity"]', 'input[id*="quantity"]']:
            el = page.locator(sel).first
            if await el.count() > 0:
                try:
                    qty_str = await el.input_value()
                    qty_val = int(qty_str.strip())
                    break
                except:
                    pass

        # Scrape Section
        section_val = None
        try:
            selected_text = await page.evaluate(r'''() => {
                let selects = Array.from(document.querySelectorAll('select[name*="section" i], select[id*="section" i]'));
                for (let s of selects) {
                    if (s.selectedIndex >= 0) {
                        return s.options[s.selectedIndex].text;
                    }
                }
                let buttons = Array.from(document.querySelectorAll('button[id*="section" i], [data-testid*="section"] button'));
                for (let b of buttons) {
                    let txt = (b.innerText || "").trim();
                    if (txt && !txt.toLowerCase().includes("select") && !txt.toLowerCase().includes("choose")) {
                        return txt;
                    }
                }
                return "";
            }''')
            if selected_text:
                section_val = selected_text.strip()
        except:
            pass

        result = {
            "ok": True,
            "title": title,
            "description": desc,
            "tags": tags,
        }
        if price_val is not None:
            result["price"] = price_val
        if qty_val is not None:
            result["qty"] = qty_val
        if section_val:
            result["section"] = section_val
        if include_asset_summary:
            result["assets"] = await _inspect_etsy_listing_assets(page)
        if product_path is not None:
            result["_asset_sync"] = await _sync_listing_assets(
                browser_ctx, page, listing_id, product_path
            )
        scrape_log.info("scrape complete (fields=%s)", sorted(k for k in result if k != "ok"))
        return result
    finally:
        if browser_ctx is not None and not connected_cdp:
            await browser_ctx.close()
        await pw.stop()


def latest_etsy_manager_snapshot() -> dict:
    """Return counts from the latest Etsy Manager crawl, if one exists."""
    scratch_dir = BASE_DIR / "scratch"
    candidates = []
    for path in scratch_dir.glob(f"etsy_manager_current_{_active_shop_id}_*.json"):
        ts = _snapshot_path_timestamp(path)
        if ts is None:
            ts = datetime.fromtimestamp(path.stat().st_mtime)
        candidates.append((ts, path))
    if not candidates:
        return {}

    _, latest = sorted(candidates, key=lambda item: item[0])[-1]
    try:
        data = json.loads(latest.read_text(encoding="utf-8"))
    except Exception:
        return {}

    normalized = normalize_etsy_manager_snapshot(data)
    counts = dict(normalized.get("counts", {}))
    raw_counts = dict(normalized.get("raw_counts", {}))
    if not counts and not any(raw_counts.values()):
        return {}
    raw_counts["total"] = sum(raw_counts.values())
    counts["total"] = sum(counts.values())
    freshness = _snapshot_freshness(snapshot_source=str(latest), snapshot_at=_snapshot_path_timestamp(latest))
    listings = []
    for listing in normalized.get("listings", []):
        item = dict(listing)
        listings.append(item)
    snapshot_at = _snapshot_path_timestamp(latest)
    return {
        "source": str(latest),
        "snapshotAt": snapshot_at.isoformat(timespec="seconds") if snapshot_at else None,
        "counts": counts,
        "raw_counts": raw_counts,
        "duplicate_count": normalized.get("duplicate_count", 0),
        "freshness": freshness,
        "stale": freshness["stale"],
        "listings": listings,
    }


def _snapshot_path_timestamp(path: Path) -> datetime | None:
    ts = _snapshot_path_timestamp_from_name(path.name)
    if ts is not None:
        return ts
    try:
        return datetime.fromtimestamp(path.stat().st_mtime)
    except Exception:
        return None


def _snapshot_path_timestamp_from_name(filename: str) -> datetime | None:
    match = re.search(r"_(\d{8}_\d{6})\.json$", filename)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y%m%d_%H%M%S")
    except Exception:
        return None


def _snapshot_freshness(
    *,
    snapshot_source: str,
    snapshot_at: datetime | None,
    max_age_hours: int = 24,
) -> dict:
    age_hours = None
    if snapshot_at is not None:
        age_hours = round(max(0.0, (datetime.now() - snapshot_at).total_seconds() / 3600), 2)
    return {
        "source": snapshot_source or None,
        "snapshotAt": snapshot_at.isoformat(timespec="seconds") if snapshot_at else None,
        "ageHours": age_hours,
        "maxAgeHours": max_age_hours,
        "stale": age_hours is None or age_hours > max_age_hours,
    }


def _snapshot_too_old(snapshot: dict, *, max_age_hours: int = 24) -> bool:
    source = str(snapshot.get("source") or "")
    if not source:
        return True
    ts = _snapshot_path_timestamp(Path(source))
    if ts is None:
        return True
    return datetime.now() - ts > timedelta(hours=max_age_hours)


def enrich_products_with_etsy_manager(products: list[dict], snapshot: dict) -> list[dict]:
    """Add status-aware Etsy links without changing the workbook mapping field.

    ``etsy_url`` remains the Excel mapping source.  The dashboard may expose a
    public URL only for an exact listing found as active in the newest manager
    snapshot; all known non-active listings require a verified manager/editor
    URL.  Missing evidence is represented as an unavailable link.
    """
    listings_by_id = {
        str(item.get("listing_id") or item.get("id") or "").strip(): item
        for item in snapshot.get("listings", [])
        if isinstance(item, dict)
    } if isinstance(snapshot, dict) else {}
    freshness = snapshot.get("freshness", {}) if isinstance(snapshot, dict) else {}

    enriched: list[dict] = []
    for product in products:
        item = dict(product)
        listing_id = _extract_etsy_listing_id(str(item.get("etsy_url") or ""))
        listing = listings_by_id.get(listing_id) if listing_id else None
        manager_status = str((listing or {}).get("managerStatus") or (listing or {}).get("status") or "").strip().lower()
        public_url = ""
        edit_url = ""
        manager_url = ""
        link_type = "unavailable"
        unavailable_reason = "listing_not_in_snapshot" if listing_id else "no_listing_id"

        if listing:
            manager_url = _safe_etsy_manager_url(
                listing.get("editUrl") or listing.get("manageUrl"), listing_id
            )
            edit_url = manager_url
            if manager_status == "active":
                public_url = _safe_etsy_public_url(listing.get("url"), listing_id)
                if not public_url:
                    # Exact active ID + manager status is sufficient to
                    # form Etsy's canonical public URL.
                    public_url = f"https://www.etsy.com/listing/{listing_id}"
                link_type = "public"
                unavailable_reason = ""
                manager_url = ""
            elif manager_status in {"draft", "inactive", "expired"}:
                if manager_url:
                    link_type = "manager"
                    unavailable_reason = ""
                else:
                    unavailable_reason = "manager_url_missing"

        item.update({
            "etsy_listing_id": listing_id or None,
            "etsy_manager_status": manager_status or None,
            "etsy_public_url": public_url or None,
            "etsy_manage_url": manager_url or None,
            "etsy_edit_url": edit_url or None,
            "etsy_link_type": link_type,
            "etsy_link_verified": bool(public_url or manager_url),
            "etsy_link_unavailable_reason": unavailable_reason or None,
            "etsy_snapshot_at": freshness.get("snapshotAt"),
            "etsy_snapshot_source": freshness.get("source"),
            "etsy_snapshot_stale": bool(freshness.get("stale", True)),
        })
        enriched.append(item)
    return enriched


def _validate_draft_listing_ids(raw_ids: object, snapshot: dict) -> list[str]:
    if not isinstance(raw_ids, list) or not raw_ids:
        raise HTTPException(400, "Cần chọn ít nhất một Etsy draft")
    ids: list[str] = []
    for value in raw_ids:
        if isinstance(value, bool):
            raise HTTPException(400, "Listing ID phải là số")
        listing_id = str(value).strip()
        if not listing_id.isdigit():
            raise HTTPException(400, f"Listing ID không hợp lệ: {listing_id or '(trống)'}")
        ids.append(listing_id)
    if len(set(ids)) != len(ids):
        raise HTTPException(400, "Listing ID bị trùng trong yêu cầu")
    by_id = {str(item.get("id") or item.get("listing_id") or ""): item
             for item in snapshot.get("listings", []) if isinstance(item, dict)}
    missing = [listing_id for listing_id in ids if listing_id not in by_id]
    if missing:
        raise HTTPException(409, f"Listing không có trong snapshot mới nhất: {', '.join(missing)}")
    non_drafts = [listing_id for listing_id in ids
                  if str(by_id[listing_id].get("managerStatus") or "").lower() != "draft"]
    if non_drafts:
        raise HTTPException(409, f"Chỉ được xoá draft; listing không phải draft: {', '.join(non_drafts)}")
    return ids


def _remove_deleted_drafts_from_snapshot(snapshot: dict, listing_ids: list[str]) -> None:
    source = Path(str(snapshot.get("source") or ""))
    if not source.is_file():
        return
    data = json.loads(source.read_text(encoding="utf-8"))
    deleted = set(listing_ids)
    drafts = data.get("draft", [])
    if isinstance(drafts, list):
        data["draft"] = [item for item in drafts
                         if str(item.get("id") or item.get("listing_id") or "") not in deleted]
    temp = source.with_suffix(source.suffix + ".tmp")
    temp.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temp.replace(source)


async def _delete_selected_etsy_drafts_unlocked(req: Request):
    data = await req.json()
    target_shop = str(data.get("shop") or _active_shop_id)
    try:
        _assert_shop_identity(target_shop)
    except RuntimeError as error:
        raise HTTPException(409, str(error)) from error
    snapshot = latest_etsy_manager_snapshot()
    if not snapshot:
        raise HTTPException(409, "Chưa có Etsy snapshot cho shop đang hoạt động; hãy đồng bộ lại")
    if _snapshot_too_old(snapshot):
        snapshot_time = snapshot.get("snapshotAt")
        raise HTTPException(
            409,
            f"Snapshot listing của shop quá cũ (hơn 24 giờ). Hãy Sync Etsy Shop trước để lấy snapshot mới. "
            f"snapshotAt={snapshot_time or 'không xác định'}",
        )
    listing_ids = _validate_draft_listing_ids(data.get("listing_ids"), snapshot)
    command = [sys.executable, str(BASE_DIR / "etsy_clean_duplicates.py"), "--shop", target_shop]
    for listing_id in listing_ids:
        command.extend(["--listing-id", listing_id])
    process = await asyncio.create_subprocess_exec(
        *command, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.STDOUT
    )
    stdout, _ = await process.communicate()
    output = stdout.decode("utf-8", errors="replace")
    if process.returncode != 0:
        raise HTTPException(502, f"Etsy không xoá draft: {output.strip()[-1200:]}")
    result = None
    for line in reversed(output.splitlines()):
        try:
            candidate = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(candidate, dict) and candidate.get("ok"):
            result = candidate
            break
    if not result or set(map(str, result.get("deleted_listing_ids", []))) != set(listing_ids):
        raise HTTPException(502, "Cleaner không xác nhận đầy đủ listing đã xoá; snapshot chưa thay đổi")
    _remove_deleted_drafts_from_snapshot(snapshot, listing_ids)
    return {"ok": True, "shop": target_shop, "deleted": len(listing_ids),
            "deleted_listing_ids": listing_ids, "output": output.strip()[-1200:]}


@app.post("/api/etsy/delete-drafts")
async def delete_selected_etsy_drafts(req: Request):
    if _etsy_draft_delete_lock.locked():
        raise HTTPException(409, "Một lệnh xoá Etsy draft khác đang chạy")
    async with _etsy_draft_delete_lock:
        return await _delete_selected_etsy_drafts_unlocked(req)


def attach_local_products_to_etsy_snapshot(snapshot: dict, products: list[dict]) -> dict:
    """Attach an actionable local product only when the Etsy listing ID matches exactly.

    Shop crawls contain remote-only metadata.  Reusing a local card without an
    exact URL match can operate on the wrong folder, so title matching is
    deliberately not used here.
    """
    if not snapshot or not snapshot.get("listings"):
        return snapshot

    by_listing_id = {}
    for product in products:
        listing_id = str(product.get("etsy_listing_id") or _extract_etsy_listing_id(str(product.get("etsy_url", ""))))
        if listing_id:
            by_listing_id[listing_id] = product

    for listing in snapshot["listings"]:
        listing_id = str(listing.get("listing_id") or listing.get("id") or "")
        local_product = by_listing_id.get(listing_id)
        if local_product:
            listing["localProduct"] = local_product
    return snapshot

# ── API: Shops ───────────────────────────────────────────────────────────────────
@app.get("/api/shops")
async def get_shops():
    return {
        "shops": list(SHOPS.values()),
        "active": _active_shop_id
    }

@app.post("/api/set-shop")
async def set_shop(request: Request):
    global _active_shop_id
    data = await request.json()
    new_id = data.get("shop_id")
    if new_id in SHOPS:
        _active_shop_id = new_id
        ACTIVE_SHOP_FILE.write_text(new_id, encoding="utf-8")
    return {"ok": True, "active": _active_shop_id}

@app.post("/api/shops/update")
async def update_shop(request: Request):
    data = await request.json()
    shop_id = data.get("id")
    if not shop_id: return {"ok": False, "error": "Missing shop id"}
    if shop_id not in SHOPS:
        SHOPS[shop_id] = {"id": shop_id}

    SHOPS[shop_id]["name"] = data.get("name", SHOPS[shop_id].get("name", ""))
    SHOPS[shop_id]["emoji"] = data.get("emoji", SHOPS[shop_id].get("emoji", ""))
    SHOPS[shop_id]["etsy_link"] = data.get("etsy_link", SHOPS[shop_id].get("etsy_link", ""))
    SHOPS[shop_id]["social_links"] = data.get("social_links", SHOPS[shop_id].get("social_links", ""))
    SHOPS[shop_id]["shop_info"] = data.get("shop_info", SHOPS[shop_id].get("shop_info", ""))

    save_shops()
    return {"ok": True, "shop": SHOPS[shop_id]}


# ── API: Cloud Asset Store (local/cloud only; never Etsy writes) ───────────────
_CLOUD_SCOPE_PATHS = {"shop": "shops", "master": "master_products"}
_CLOUD_SAFE_SHOP_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,127}$")
_CLOUD_SECRET_RE = re.compile(
    r"(?i)(access[_-]?token|refresh[_-]?token|client[_-]?secret|authorization)\s*[=:]\s*[^\s,;]+"
)


def _normalize_cloud_scope(value: object, *, allow_empty: bool = False) -> str:
    """Accept only the public scope names; never expose internal path aliases."""

    scope = str(value or "").strip().lower()
    if not scope and allow_empty:
        return ""
    if scope not in _CLOUD_SCOPE_PATHS:
        raise HTTPException(400, "scope phải là shop hoặc master")
    return scope


def _normalize_cloud_folder(value: object, *, required: bool = True) -> str:
    if not isinstance(value, str):
        if required:
            raise HTTPException(400, "folder phải là chuỗi product-N")
        return ""
    folder = value.strip()
    if not folder and not required:
        return ""
    return _validate_product_numbered_folder_name(folder)


def _validate_cloud_shop_id(value: object, *, required: bool) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise HTTPException(400, "Thiếu shop_id")
        shop_id = str(_active_shop_id or "").strip()
        if not shop_id:
            raise HTTPException(409, "Chưa có shop đang hoạt động hợp lệ")
    elif not isinstance(value, str):
        raise HTTPException(400, "shop_id phải là chuỗi")
    else:
        shop_id = value.strip()
    if not shop_id or not _CLOUD_SAFE_SHOP_RE.fullmatch(shop_id):
        raise HTTPException(400, "shop_id không hợp lệ")
    try:
        _assert_shop_identity(shop_id)
    except RuntimeError as exc:
        raise HTTPException(409, str(exc)) from exc
    if shop_id not in SHOPS:
        raise HTTPException(409, f"Shop đang hoạt động chưa được cấu hình: {shop_id}")
    return shop_id


def _cloud_product_identity(scope: str, shop_id: str, folder: str) -> dict:
    path_scope = _CLOUD_SCOPE_PATHS[scope]
    shop = shop_id if scope == "shop" else None
    key = f"{path_scope}/{shop_id}/{folder}" if scope == "shop" else f"{path_scope}/{folder}"
    return {
        "scope": scope,
        "shop": shop,
        "product": folder,
        "key": key,
    }


def _cloud_asset_product_root(scope: str, shop_id: str, folder: str) -> Path:
    """Resolve one product only through the canonical shop/master layout."""

    normalized_scope = _normalize_cloud_scope(scope)
    normalized_folder = _normalize_cloud_folder(folder)
    # Every cloud API context is anchored to the currently active shop,
    # including master products.  The shop component is not part of a master
    # path, but the request still must carry the active-shop identity.
    shop_id = _validate_cloud_shop_id(shop_id, required=True)
    if normalized_scope == "shop":
        candidate = BASE_DIR / "shops" / shop_id / normalized_folder
    else:
        candidate = BASE_DIR / "master_products" / normalized_folder
    if candidate.is_symlink():
        raise HTTPException(400, "product folder không được là symlink")
    if not candidate.exists():
        raise HTTPException(404, f"Không tìm thấy product folder: {normalized_folder}")
    try:
        resolved, identity = resolve_product(BASE_DIR, candidate)
    except (AssetValidationError, OSError, ValueError) as exc:
        raise HTTPException(400, f"product folder không hợp lệ hoặc không tồn tại: {normalized_folder}") from exc
    expected_scope = _CLOUD_SCOPE_PATHS[normalized_scope]
    if identity.scope != expected_scope or (
        normalized_scope == "shop" and identity.shop != shop_id
    ):
        raise HTTPException(400, "product folder không thuộc scope/shop yêu cầu")
    return resolved


def _cloud_status_targets(scope: str, shop_id: str, folder: str = "") -> list[Path]:
    normalized_scope = _normalize_cloud_scope(scope)
    normalized_shop_id = _validate_cloud_shop_id(shop_id, required=True)
    normalized_folder = _normalize_cloud_folder(folder, required=False)
    if normalized_folder:
        return [_cloud_asset_product_root(normalized_scope, normalized_shop_id, normalized_folder)]

    targets: list[Path] = []
    for product_path, identity in discover_product_roots(BASE_DIR):
        if normalized_scope == "shop":
            if identity.scope == "shops" and identity.shop == normalized_shop_id:
                targets.append(product_path)
        elif identity.scope == "master_products":
            targets.append(product_path)
    return targets


async def _read_cloud_asset_request(request: Request) -> dict:
    try:
        data = await request.json()
    except Exception as exc:
        raise HTTPException(400, "Payload cloud assets phải là JSON object") from exc
    if not isinstance(data, dict):
        raise HTTPException(400, "Payload cloud assets phải là JSON object")
    for field in ("shop_id", "scope", "folder"):
        if field not in data:
            raise HTTPException(400, f"Thiếu {field}")

    shop_id = _validate_cloud_shop_id(data.get("shop_id"), required=True)
    scope_value = data.get("scope")
    if not isinstance(scope_value, str) or not scope_value.strip():
        raise HTTPException(400, "Thiếu scope")
    scope = _normalize_cloud_scope(scope_value)
    folder = _normalize_cloud_folder(data.get("folder"))
    target = await asyncio.to_thread(_cloud_asset_product_root, scope, shop_id, folder)
    identity = _cloud_product_identity(scope, shop_id, folder)
    normalized = dict(data)
    normalized.update(
        {
            "shop_id": shop_id,
            "scope": scope,
            "folder": folder,
            "target": target,
            "product_identity": identity,
            "product_key": identity["key"],
        }
    )
    return normalized


def _cloud_error_text(exc: Exception) -> str:
    message = str(exc).strip() or exc.__class__.__name__
    return _CLOUD_SECRET_RE.sub(r"\1=[REDACTED]", message)[:800]


def _decorate_cloud_status(item: dict, identity: dict) -> dict:
    decorated = dict(item)
    decorated["scope"] = identity["scope"]
    decorated["shop"] = identity["shop"]
    decorated["folder"] = identity["product"]
    decorated["product_identity"] = identity
    decorated["product_key"] = identity["key"]
    return decorated


def _cloud_operation_error(operation: str, data: dict, exc: Exception) -> JSONResponse:
    message = _cloud_error_text(exc)
    identity = data["product_identity"]
    logger.error(
        "cloud asset operation failed operation=%s product=%s error=%s",
        operation,
        identity["key"],
        message,
    )
    broadcast(f"[CLOUD-ASSETS] ❌ {operation} {identity['key']}: {message}")
    return JSONResponse(
        {
            "ok": False,
            "operation": operation,
            "product": identity,
            "state": "ERROR",
            "error": message,
        },
        status_code=502 if isinstance(exc, CloudAssetError) else 500,
    )


@app.get("/api/cloud-assets/status")
async def cloud_assets_status(
    shop_id: str = "",
    scope: str = "shop",
    folder: str = "",
):
    requested_shop = _validate_cloud_shop_id(shop_id, required=False)
    normalized_scope = _normalize_cloud_scope(scope or "shop")
    normalized_folder = _normalize_cloud_folder(folder, required=False)
    targets = await asyncio.to_thread(
        _cloud_status_targets,
        normalized_scope,
        requested_shop,
        normalized_folder,
    )
    store = get_cloud_asset_store()
    items: list[dict] = []
    for target in targets:
        identity = _cloud_product_identity(normalized_scope, requested_shop, target.name)
        try:
            # CloudAssetStore.status defaults to local/cache inspection.  Do
            # not expose a remote-check flag on this dashboard route.
            item = await asyncio.to_thread(store.status, target)
            items.append(_decorate_cloud_status(item, identity))
        except Exception as exc:
            message = _cloud_error_text(exc)
            items.append(
                _decorate_cloud_status(
                    {
                        "ok": False,
                        "state": "ERROR",
                        "path": str(target),
                        "error": message,
                        "last_error": message,
                    },
                    identity,
                )
            )
    async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
        schedules = {
            key: {field: value for field, value in value.items() if field != "data"}
            for key, value in _CLOUD_ASSET_UPLOAD_SCHEDULES.items()
        }
    for item in items:
        schedule = schedules.get(str(item.get("product_key") or ""))
        if schedule is None:
            continue
        item["upload_schedule"] = schedule
        schedule_status = str(schedule.get("status") or "queued")
        if schedule_status == "queued":
            item["state"] = "UPLOAD_SCHEDULED"
        elif schedule_status == "running":
            item["state"] = "UPLOADING"
        elif schedule_status == "error":
            item["state"] = "ERROR"
            item["last_error"] = str(schedule.get("error") or "Cloud upload theo lịch thất bại")
    return {
        "ok": all(bool(item.get("ok", False)) for item in items) if items else True,
        "active_shop": _active_shop_id,
        "shop_id": requested_shop,
        "scope": normalized_scope,
        "folder": normalized_folder or None,
        "cloud_network_contacted": False,
        "items": items,
        "products": items,
    }


def _cloud_image_relative_path(filename: str) -> str:
    relative = urllib.parse.unquote(str(filename or "")).replace("\\", "/").strip("/")
    parts = relative.split("/")
    if not relative or not parts or parts[0] != "images" or any(part in {"", ".", ".."} for part in parts):
        raise HTTPException(400, "filename phải trỏ tới một image path hợp lệ")
    return "/".join(parts)


def _safe_regular_cloud_file(path: Path) -> bool:
    try:
        stat_result = path.stat()
    except OSError:
        return False
    return (
        not path.is_symlink()
        and path.is_file()
        and stat_result.st_size > 0
        and getattr(stat_result, "st_blocks", 1) > 0
    )


@app.get("/api/cloud-assets/preview/{folder}")
async def cloud_assets_preview(folder: str):
    """Serve only the retained, small shop preview marker."""

    target = _cloud_asset_product_root("shop", _active_shop_id, folder)
    preview = target / CLOUD_PREVIEW_FILE_NAME
    if not _safe_regular_cloud_file(preview):
        raise HTTPException(404, "Không có cloud preview cho product này")
    return FileResponse(
        preview,
        media_type="image/webp",
        headers={"Cache-Control": "private, max-age=300"},
    )


@app.get("/api/cloud-assets/image/{folder}/{filename:path}")
async def cloud_assets_image(folder: str, filename: str):
    """Serve one manifest-verified image, hydrating to cache when needed."""

    target = _cloud_asset_product_root("shop", _active_shop_id, folder)
    relative = _cloud_image_relative_path(filename)
    local_candidate = target / Path(relative)
    if _safe_regular_cloud_file(local_candidate):
        return FileResponse(local_candidate, headers={"Cache-Control": "private, max-age=300"})

    store = get_cloud_asset_store()
    try:
        resolution = await asyncio.to_thread(
            store.resolve_asset_root,
            target,
        )
    except (CloudAssetError, AssetValidationError, OSError, ValueError, TypeError, KeyError) as exc:
        raise HTTPException(502, f"Không hydrate được image cloud: {_cloud_error_text(exc)}") from exc

    manifest = resolution.get("manifest") if isinstance(resolution, dict) else None
    records = manifest.get("files", []) if isinstance(manifest, dict) else []
    if not any(
        isinstance(item, dict)
        and item.get("role") == "image"
        and str(item.get("path") or "").replace("\\", "/") == relative
        for item in (records if isinstance(records, list) else [])
    ):
        raise HTTPException(404, "Image không có trong manifest cloud hiện hành")

    asset_root = Path(str(resolution.get("asset_root") or "")).absolute()
    candidate = (asset_root / Path(relative)).absolute()
    try:
        candidate.relative_to(asset_root)
    except ValueError as exc:
        raise HTTPException(400, "image path thoát khỏi asset root") from exc
    if not _safe_regular_cloud_file(candidate):
        raise HTTPException(404, "Không tìm thấy image đã hydrate")
    return FileResponse(candidate, headers={"Cache-Control": "private, max-age=300"})


def _cloud_store_operation(store: CloudAssetStore, operation: str, data: dict):
    target = data["target"]
    if operation == "upload":
        return store.upload(target, revision=data.get("revision"))
    if operation == "verify":
        return store.verify(target)
    if operation == "restore":
        return store.restore(target, force=data.get("force", False))
    if operation == "cancel-offload":
        return store.cancel_offload(target)
    raise RuntimeError(f"unsupported cloud asset operation: {operation}")


async def _perform_cloud_asset_mutation(data: dict, operation: str) -> dict:
    """Run one cloud mutation under the existing product-level lock."""
    identity = data["product_identity"]
    product_key = identity["key"]
    store = get_cloud_asset_store()
    mutation_lock = await _get_cloud_asset_mutation_lock(product_key)
    if mutation_lock.locked():
        broadcast(f"[CLOUD-ASSETS] ⏳ Chờ mutation cloud đang chạy: {product_key}")

    async with mutation_lock:
        broadcast(f"[CLOUD-ASSETS] 🔄 {operation} {product_key}")
        result = await asyncio.to_thread(_cloud_store_operation, store, operation, data)

    broadcast(f"[CLOUD-ASSETS] ✅ {operation} {product_key}")
    return result if isinstance(result, dict) else {"ok": True, "state": None}


def _cloud_upload_wait_reason(shop_id: str) -> str:
    """Return why a scheduled cloud upload must wait, scoped to its shop."""
    if shop_id in _etsy_single_sync_busy_shops:
        return "đang Sync listing Etsy"
    if shop_id == _active_shop_id and "__ETSY_SYNC__" in _running_processes:
        return "đang Sync Shop Manager"
    if _is_poster_locked_for_shop(shop_id):
        return "đang Post Etsy"
    if _etsy_update_shop_is_busy(shop_id):
        return "đang Update Etsy"
    return ""


async def _run_scheduled_cloud_uploads(shop_id: str) -> None:
    """Drain one shop's explicit cloud queue only while Etsy work is idle."""
    try:
        while True:
            async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
                queue = _CLOUD_ASSET_UPLOAD_QUEUE_BY_SHOP.get(shop_id, [])
                while queue and queue[0] not in _CLOUD_ASSET_UPLOAD_SCHEDULES:
                    queue.pop(0)
                if not queue:
                    _CLOUD_ASSET_UPLOAD_QUEUE_BY_SHOP.pop(shop_id, None)
                    return

                product_key = queue[0]
                scheduled = _CLOUD_ASSET_UPLOAD_SCHEDULES[product_key]
                reason = _cloud_upload_wait_reason(shop_id)
                if reason:
                    if scheduled.get("wait_reason") != reason:
                        scheduled["wait_reason"] = reason
                        scheduled["updated_at"] = time.time()
                        broadcast(f"[CLOUD-ASSETS] ⏳ Upload đã xếp lịch {product_key}: chờ {reason}")
                else:
                    queue.pop(0)
                    scheduled["status"] = "running"
                    scheduled["wait_reason"] = ""
                    scheduled["updated_at"] = time.time()

            if reason:
                await asyncio.sleep(_CLOUD_ASSET_UPLOAD_IDLE_POLL_SECONDS)
                continue

            data = scheduled["data"]
            try:
                upload_result = await _perform_cloud_asset_mutation(data, "upload")
                if not bool(upload_result.get("ok", True)):
                    raise RuntimeError(upload_result.get("error") or "upload cloud không thành công")
                verify_result = await _perform_cloud_asset_mutation(data, "verify")
                if not bool(verify_result.get("ok", True)):
                    raise RuntimeError(verify_result.get("error") or "verify cloud không thành công")
            except Exception as exc:
                message = _cloud_error_text(exc)
                async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
                    current = _CLOUD_ASSET_UPLOAD_SCHEDULES.get(product_key)
                    if current is not None:
                        current.update({"status": "error", "error": message, "updated_at": time.time()})
                broadcast(f"[CLOUD-ASSETS] ❌ Upload đã xếp lịch {product_key}: {message}")
            else:
                async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
                    _CLOUD_ASSET_UPLOAD_SCHEDULES.pop(product_key, None)
                broadcast(f"[CLOUD-ASSETS] ✅ Upload & verify theo lịch {product_key}")
    finally:
        async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
            if _CLOUD_ASSET_UPLOAD_WORKERS.get(shop_id) is asyncio.current_task():
                _CLOUD_ASSET_UPLOAD_WORKERS.pop(shop_id, None)


async def _schedule_cloud_upload_and_verify(data: dict) -> tuple[dict, bool]:
    identity = data["product_identity"]
    product_key = identity["key"]
    shop_id = str(data["shop_id"])
    async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
        existing = _CLOUD_ASSET_UPLOAD_SCHEDULES.get(product_key)
        if existing is not None and existing.get("status") != "error":
            return existing, False
        if existing is not None:
            _CLOUD_ASSET_UPLOAD_SCHEDULES.pop(product_key, None)
        scheduled = {
            "product_key": product_key,
            "shop_id": shop_id,
            "scope": data["scope"],
            "folder": data["folder"],
            "status": "queued",
            "wait_reason": _cloud_upload_wait_reason(shop_id),
            "enqueued_at": time.time(),
            "updated_at": time.time(),
            "data": dict(data),
        }
        _CLOUD_ASSET_UPLOAD_SCHEDULES[product_key] = scheduled

    async def run_upload_verify():
        # Preserve the status contract for a cloud command that was admitted
        # while a legacy/in-flight listing sync is finishing. Normally the
        # shared FIFO already provides this ordering, but this guard also
        # covers a sync that pre-dated the queue migration.
        while (reason := _cloud_upload_wait_reason(shop_id)):
            async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
                current = _CLOUD_ASSET_UPLOAD_SCHEDULES.get(product_key)
                if current is None:
                    return
                current.update({"status": "queued", "wait_reason": reason, "updated_at": time.time()})
            await asyncio.sleep(_CLOUD_ASSET_UPLOAD_IDLE_POLL_SECONDS)
        async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
            current = _CLOUD_ASSET_UPLOAD_SCHEDULES.get(product_key)
            if current is None:
                return
            current.update({"status": "running", "wait_reason": "", "updated_at": time.time()})
        try:
            upload_result = await _perform_cloud_asset_mutation(data, "upload")
            if not bool(upload_result.get("ok", True)):
                raise RuntimeError(upload_result.get("error") or "upload cloud không thành công")
            verify_result = await _perform_cloud_asset_mutation(data, "verify")
            if not bool(verify_result.get("ok", True)):
                raise RuntimeError(verify_result.get("error") or "verify cloud không thành công")
        except Exception as exc:
            message = _cloud_error_text(exc)
            async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
                current = _CLOUD_ASSET_UPLOAD_SCHEDULES.get(product_key)
                if current is not None:
                    current.update({"status": "error", "error": message, "updated_at": time.time()})
            broadcast(f"[CLOUD-ASSETS] ❌ Upload đã xếp lịch {product_key}: {message}")
            raise
        else:
            async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
                _CLOUD_ASSET_UPLOAD_SCHEDULES.pop(product_key, None)
            broadcast(f"[CLOUD-ASSETS] ✅ Upload & verify theo lịch {product_key}")

    command, created = await _enqueue_operation(
        operation="cloud-upload-verify",
        shop_id=shop_id,
        target=product_key,
        callback=run_upload_verify,
    )
    scheduled["command_id"] = command["command_id"]
    scheduled["queue_position"] = command.get("position")
    scheduled["wait_reason"] = "đang chờ hàng lệnh chung" if created else "đã có trong hàng chờ"
    broadcast(f"[CLOUD-ASSETS] 🗓️ Đã thêm Upload & verify {product_key} vào hàng chờ chung")
    return scheduled, created


async def _run_cloud_asset_mutation(request: Request, operation: str):
    data = await _read_cloud_asset_request(request)
    identity = data["product_identity"]
    try:
        result = await _perform_cloud_asset_mutation(data, operation)
    except Exception as exc:
        return _cloud_operation_error(operation, data, exc)

    ok = bool(result.get("ok", True)) if isinstance(result, dict) else True
    response = {
        "ok": ok,
        "operation": operation,
        "product": identity,
        "state": result.get("state") if isinstance(result, dict) else None,
        "result": result,
    }
    if not ok and isinstance(result, dict):
        response["error"] = result.get("error") or result.get("local_error") or (
            "local assets do not match the cloud revision"
        )
    return response


@app.post("/api/cloud-assets/upload")
async def cloud_assets_upload(request: Request):
    return await _run_cloud_asset_mutation(request, "upload")


@app.post("/api/cloud-assets/schedule-upload-verify")
async def cloud_assets_schedule_upload_verify(request: Request):
    data = await _read_cloud_asset_request(request)
    scheduled, created = await _schedule_cloud_upload_and_verify(data)
    return JSONResponse(
        status_code=202 if created else 200,
        content={
            "ok": True,
            "created": created,
            "operation": "schedule-upload-verify",
            "product": data["product_identity"],
            "schedule": {key: value for key, value in scheduled.items() if key != "data"},
        },
    )


@app.post("/api/cloud-assets/verify")
async def cloud_assets_verify(request: Request):
    return await _run_cloud_asset_mutation(request, "verify")


@app.post("/api/cloud-assets/restore")
async def cloud_assets_restore(request: Request):
    return await _run_cloud_asset_mutation(request, "restore")


@app.post("/api/cloud-assets/cancel-offload")
async def cloud_assets_cancel_offload(request: Request):
    return await _run_cloud_asset_mutation(request, "cancel-offload")


@app.post("/api/cloud-assets/maintain")
async def cloud_assets_maintain(request: Request):
    data = await _read_cloud_asset_request(request)
    identity = data["product_identity"]
    product_key = identity["key"]

    raw_dry_run = data.get("dry_run", True)
    if not isinstance(raw_dry_run, bool):
        raise HTTPException(400, "dry_run phải là boolean")
    dry_run = raw_dry_run
    apply_requested = data.get("apply") is True and dry_run is False
    policy_enabled = data.get("policy_enabled") is True
    raw_allowlist = data.get("allowlist", [])
    if isinstance(raw_allowlist, str):
        allowlist = [item.strip() for item in raw_allowlist.split(",") if item.strip()]
    elif isinstance(raw_allowlist, list):
        allowlist = [str(item).strip() for item in raw_allowlist if str(item).strip()]
    else:
        raise HTTPException(400, "allowlist phải là danh sách hoặc chuỗi")

    if apply_requested:
        if not policy_enabled:
            raise HTTPException(409, "Xoá offload cần policy_enabled=true trong request")
        if product_key not in allowlist:
            raise HTTPException(409, f"Xoá offload cần allowlist exact: {product_key}")

    # A plain request stays policy-disabled.  An explicit dry-run may opt into
    # the same allowlist checks without ever permitting deletion.
    dry_run_policy_enabled = policy_enabled and product_key in allowlist
    offload_enabled = policy_enabled if apply_requested or dry_run_policy_enabled else False
    allowlist_for_store = allowlist if offload_enabled else ()
    store = get_cloud_asset_store()
    broadcast(f"[CLOUD-ASSETS] 🔎 maintain {'apply' if apply_requested else 'dry-run'} {product_key}")
    try:
        results = await asyncio.to_thread(
            store.maintain,
            [data["target"]],
            apply=apply_requested,
            offload_enabled=offload_enabled,
            allowlist=allowlist_for_store,
            older_than_days=data.get("older_than_days"),
        )
    except Exception as exc:
        return _cloud_operation_error("maintain", data, exc)
    return {
        "ok": all(bool(item.get("ok", False)) for item in results),
        "operation": "maintain",
        "dry_run": not apply_requested,
        "apply": apply_requested,
        "policy_enabled": offload_enabled,
        "allowlist": list(allowlist_for_store),
        "product": identity,
        "results": results,
    }

# ── API: Products ────────────────────────────────────────────────────────────────
@app.get("/api/products")
def get_products(
    page: int = 1,
    page_size: int | None = None,
    q: str = "",
    search: str = "",
    status: str = "",
    sort: str = "row",
):
    try:
        products = products_from_excel()
        snapshot = latest_etsy_manager_snapshot()
        products = enrich_products_with_etsy_manager(products, snapshot)
        if page_size is None:
            return JSONResponse({
                "products": products,
                "etsy_manager": attach_local_products_to_etsy_snapshot(
                    snapshot, products
                ),
            })

        if page < 1 or page_size < 1 or page_size > 200:
            raise HTTPException(400, "page phải >= 1 và page_size phải trong khoảng 1..200")
        query = str(q or search or "").strip().casefold()
        status_filter = str(status or "").strip().casefold()
        filtered = products
        if query:
            filtered = [
                product for product in filtered
                if query in " ".join(
                    str(product.get(field) or "")
                    for field in ("title", "folder", "tags", "sku", "etsy_url")
                ).casefold()
            ]
        if status_filter:
            filtered = [
                product for product in filtered
                if status_filter in str(product.get("status") or "").casefold()
            ]
        sort_key = str(sort or "row").strip().casefold()
        descending = sort_key.startswith("-")
        sort_key = sort_key.lstrip("-")
        allowed_sort = {"row", "folder", "title", "status", "price"}
        if sort_key not in allowed_sort:
            raise HTTPException(400, "sort không hợp lệ")
        filtered = sorted(
            filtered,
            key=lambda product: (
                float(product.get("price") or 0)
                if sort_key == "price"
                else str(product.get(sort_key) or "").casefold()
            ),
            reverse=descending,
        )
        total = len(filtered)
        start = (page - 1) * page_size
        page_products = filtered[start:start + page_size]
        return JSONResponse({
            "products": page_products,
            "etsy_manager": attach_local_products_to_etsy_snapshot(snapshot, page_products),
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "pages": (total + page_size - 1) // page_size,
                "has_next": start + page_size < total,
                "query": query,
                "status": status_filter,
                "sort": sort,
            },
        })
    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        raise HTTPException(500, str(e))


@app.get("/api/aggregate-products")
def get_aggregate_products():
    """Return one catalog containing Etsy listings and local folders.

    A mapped Etsy listing + local folder is emitted as one record. Unmapped
    Etsy listings and folder-only local products remain visible as separate
    records so duplicate review has the complete picture.
    """
    try:
        snapshot = latest_etsy_manager_snapshot()
        catalog = build_unified_catalog(BASE_DIR, _active_shop_id, EXCEL_FILE())
        for record in catalog.get("records", []):
            if not isinstance(record, dict) or record.get("source") not in {"local", "both"}:
                continue
            decorated = enrich_products_with_etsy_manager([record], snapshot)[0]
            record.update({
                key: value for key, value in decorated.items()
                if key.startswith("etsy_") and key != "etsy_url"
            })
        return JSONResponse(catalog)
    except Exception as exc:
        raise HTTPException(500, str(exc))


def _register_local_folders_in_catalog(
    ws,
    folders: list[str],
    shop_id: str,
) -> list[dict[str, int | str]]:
    section_default = "Digital Planner"
    assignments: list[dict[str, int | str]] = []
    slot_rows = _make_local_product_slot_rows(ws, 4, len(folders))
    for folder, target_row in zip(folders, slot_rows):
        _clear_catalog_row(ws, target_row)
        set_cell_value(ws, target_row, 2, folder)              # B = folder
        set_cell_value(ws, target_row, 5, 4.99)                # E = price
        set_cell_value(ws, target_row, 11, 999)                # K = qty
        set_cell_value(ws, target_row, 13, "2020_2026")        # M = when made
        set_cell_value(ws, target_row, 14, "⏳ Chờ đăng")      # N = status
        set_cell_value(ws, target_row, 15, section_default)    # O = section
        set_cell_value(ws, target_row, 18, generate_sku(shop_id, folder))  # R = sku
        assignments.append({"folder": folder, "row": target_row})
    return assignments


@app.post("/api/local-products/register")
async def register_local_products(request: Request):
    """Allocate existing local folders into catalog rows (without touching assets)."""
    data = await request.json()
    shop = str(data.get("shop") if isinstance(data, dict) else "").strip()
    folders_raw = data.get("folders") if isinstance(data, dict) else None

    if not shop:
        raise HTTPException(400, "Thiếu shop")
    try:
        _assert_shop_identity(shop)
    except RuntimeError as error:
        raise HTTPException(409, str(error)) from error

    if not isinstance(folders_raw, list) or not folders_raw:
        raise HTTPException(400, "folders phải là mảng tên folder không rỗng")

    folders = []
    normalized_seen: set[str] = set()
    for folder in folders_raw:
        if not isinstance(folder, str):
            raise HTTPException(400, "Mỗi folder phải là chuỗi")
        normalized = _validate_product_numbered_folder_name(folder)
        if normalized in normalized_seen:
            raise HTTPException(400, f"Folder bị lặp trong yêu cầu: {normalized}")
        normalized_seen.add(normalized)
        folders.append(normalized)

    if not folders:
        return {"ok": True, "shop": shop, "rows": []}

    shop_dir = SHOP_DIR()
    excel_path = EXCEL_FILE()
    if not excel_path.exists():
        raise HTTPException(404, "Chưa có file Excel của shop đang hoạt động")

    for folder in folders:
        folder_path = shop_dir / folder
        if not folder_path.is_dir():
            raise HTTPException(404, f"Không tìm thấy folder: {folder}")

    async with _product_create_lock:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Listings"]
        mapped_rows = _find_mapped_catalog_rows(ws)
        for folder in folders:
            if folder in mapped_rows:
                raise HTTPException(409, f"Folder {folder} đã được map trước đó tại row {mapped_rows[folder]}")
        backup_path = excel_path.with_name(
            f"{excel_path.stem}.backup_local_register_{time.strftime('%Y%m%d_%H%M%S')}{excel_path.suffix}"
        )
        shutil.copy2(excel_path, backup_path)
        try:
            assignments = _register_local_folders_in_catalog(ws, folders, shop)
            wb.save(excel_path)
        except Exception as error:
            try:
                shutil.copy2(backup_path, excel_path)
            except Exception:
                pass
            raise HTTPException(500, "Lưu file Excel thất bại, đã khôi phục bản sao dự phòng") from error

    return {"ok": True, "shop": shop, "rows": assignments}


@app.post("/api/aggregate-products/merge-duplicates")
async def merge_aggregate_duplicates(request: Request):
    """Merge only high-confidence local duplicates and keep a rollback quarantine."""
    data = await request.json()
    group_ids = data.get("group_ids") if isinstance(data, dict) else None
    if group_ids is not None and not isinstance(group_ids, list):
        raise HTTPException(400, "group_ids phải là danh sách")
    try:
        result = merge_safe_duplicates(BASE_DIR, _active_shop_id, EXCEL_FILE(), group_ids)
        for item in result.get("merged", []):
            broadcast(
                f"[CATALOG-DEDUPE] ✅ {item.get('canonical_folder')} ← "
                f"{', '.join(item.get('moved_folders', []))}"
            )
        return JSONResponse(result)
    except Exception as exc:
        raise HTTPException(500, str(exc))

# ── API: Import Etsy CSV ──────────────────────────────────────────────────────────
@app.post("/api/import-csv")
async def import_etsy_csv(file: UploadFile, background_tasks: BackgroundTasks, target_shop: str = ""):
    """
    Import listings from an Etsy-exported CSV.
    Etsy CSV columns of interest: TITLE, DESCRIPTION, TAGS, PRICE, QUANTITY, SECTION
    Creates product-XX folders + populates Excel for the active (or target) shop.
    """
    import csv, io, shutil

    shop_id = target_shop or _active_shop_id
    shop_dir = BASE_DIR / "shops" / shop_id
    excel_path = shop_dir / "Etsy_SEO_Generator.xlsx"

    if not excel_path.exists():
        # Copy template from another shop or create blank, and clean all listings
        src = BASE_DIR / "shops" / "templystudios" / "Etsy_SEO_Generator.xlsx"
        if src.exists():
            copy_and_clean_template(src, excel_path)
        else:
            raise HTTPException(400, "No Excel template found for target shop")

    raw = await file.read()
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Listings"]

    next_row = _next_empty_catalog_row(ws)
    reusable_slots = _find_reusable_empty_product_slots(ws, shop_dir)
    used_folders = set()

    imported = []
    for i, row in enumerate(reader):
        # Etsy CSV field names (case-insensitive match)
        row_ci = {k.strip().upper(): v.strip() for k, v in row.items()}

        title = row_ci.get("TITLE", "")
        description = row_ci.get("DESCRIPTION", "")
        tags_raw = row_ci.get("TAGS", "")
        price = row_ci.get("PRICE", "4.99")
        qty = row_ci.get("QUANTITY", "999")
        section = row_ci.get("SECTION", "Digital Planner")

        # Skip empty rows
        if not title:
            continue

        slot = _allocate_product_slot(ws, shop_dir, next_row, reusable_slots, used_folders)
        folder_name = slot["folder"]
        folder_path = slot["path"]
        (folder_path / "images").mkdir(parents=True, exist_ok=True)
        (folder_path / "files").mkdir(parents=True, exist_ok=True)

        # Write to Excel
        target_row = slot["row"]
        _clear_catalog_row(ws, target_row)
        set_cell_value(ws, target_row, 2, folder_name)    # B = folder
        set_cell_value(ws, target_row, 8, title)          # H = title
        set_cell_value(ws, target_row, 9, description)    # I = description
        set_cell_value(ws, target_row, 10, tags_raw)      # J = tags
        set_cell_value(ws, target_row, 5, float(price) if price else 4.99)  # E = price
        set_cell_value(ws, target_row, 11, int(qty) if qty.isdigit() else 999)  # K = qty
        set_cell_value(ws, target_row, 14, "✅ Đã đăng")  # N = status
        set_cell_value(ws, target_row, 15, section)        # O = section

        # Extract image URLs
        image_urls = []
        for img_col in ["IMAGE1", "IMAGE2", "IMAGE3", "IMAGE4", "IMAGE5", "IMAGE6", "IMAGE7", "IMAGE8", "IMAGE9", "IMAGE10"]:
            url = row_ci.get(img_col)
            if url and url.startswith("http"):
                image_urls.append(url)

        if image_urls:
            background_tasks.add_task(download_images_background, folder_path / "images", image_urls)

        imported.append(folder_name)
        next_row = _next_empty_catalog_row(ws, target_row + 1)

    wb.save(excel_path)
    return {"ok": True, "imported": len(imported), "folders": imported, "shop": shop_id}

async def download_images_background(images_dir: Path, urls: list[str]):
    """Background task to download images from URLs."""
    try:
        import httpx as _httpx
    except Exception:
        return
    async with _httpx.AsyncClient(timeout=30.0) as client:
        for i, url in enumerate(urls, 1):
            try:
                resp = await client.get(url)
                if resp.status_code == 200:
                    ext = url.split('.')[-1].split('?')[0]
                    if ext.lower() not in ['jpg', 'jpeg', 'png', 'webp', 'gif']: ext = 'jpg'
                    filepath = images_dir / f"image_{i:02d}.{ext}"
                    with open(filepath, 'wb') as f:
                        f.write(resp.content)
            except Exception as e:
                logger.warning("Failed to download image %s: %s", url, e, extra={"operation": "download_images_background"})


# ── API: Sync products to another shop ────────────────────────────────────────────
@app.post("/api/products/sync-to-shop")
async def sync_to_shop(request: Request):
    """
    Copy selected products (Excel rows) to another shop.
    Optionally copies the physical product folder (images, files).
    Detects duplicates and handles merge/skip conflicts.
    """
    import shutil
    data = await request.json()
    target_shop_id = data.get("target_shop")
    rows = data.get("rows", [])                  # list of row numbers from current shop
    copy_files = data.get("copy_files", False)          # also copy images/files folders
    conflict_resolution = data.get("conflict_resolution")  # None, "merge", "skip"

    if not target_shop_id or target_shop_id not in SHOPS:
        return JSONResponse({"ok": False, "error": "Invalid target shop"}, 400)
    if target_shop_id == _active_shop_id:
        return JSONResponse({"ok": False, "error": "Cannot sync to same shop"}, 400)

    src_shop_dir = SHOP_DIR()
    dst_shop_dir = BASE_DIR / "shops" / target_shop_id
    dst_excel = dst_shop_dir / "Etsy_SEO_Generator.xlsx"

    # Ensure destination shop folder exists
    dst_shop_dir.mkdir(parents=True, exist_ok=True)

    # Load source Excel
    src_wb = openpyxl.load_workbook(EXCEL_FILE(), data_only=True)
    src_ws = src_wb["Listings"]

    # Load/create destination Excel
    if not dst_excel.exists():
        copy_and_clean_template(EXCEL_FILE(), dst_excel)
        dst_wb = openpyxl.load_workbook(dst_excel)
    else:
        dst_wb = openpyxl.load_workbook(dst_excel)
    dst_ws = dst_wb["Listings"]

    # ── Check for duplicates in the destination shop ──
    existing_dst_products = []
    dst_wb_read = openpyxl.load_workbook(dst_excel, data_only=True)
    dst_ws_read = dst_wb_read["Listings"]
    for r in range(4, dst_ws_read.max_row + 1):
        f_val = dst_ws_read.cell(row=r, column=2).value
        k_val = dst_ws_read.cell(row=r, column=3).value
        t_val = dst_ws_read.cell(row=r, column=8).value
        if not f_val:
            continue
        existing_dst_products.append({
            "row": r,
            "folder": str(f_val).strip(),
            "keywords": str(k_val or "").strip(),
            "title": str(t_val or "").strip()
        })

    conflicts = []
    for row_num in rows:
        src_folder = src_ws.cell(row=row_num, column=2).value
        src_keywords = str(src_ws.cell(row=row_num, column=3).value or "").strip()
        src_title = str(src_ws.cell(row=row_num, column=8).value or "").strip()

        if not src_folder:
            continue

        for dp in existing_dst_products:
            kw_match = src_keywords and src_keywords.lower() not in ("none", "not provided", "") and src_keywords.lower() == dp["keywords"].lower()
            title_match = src_title and not src_title.startswith("[Cần SEO]") and src_title.lower() == dp["title"].lower()

            if kw_match or title_match:
                conflicts.append({
                    "src_row": row_num,
                    "src_folder": str(src_folder),
                    "src_title": src_title or f"Keywords: {src_keywords}",
                    "dst_row": dp["row"],
                    "dst_folder": dp["folder"],
                    "dst_title": dp["title"] or f"Keywords: {dp['keywords']}",
                })
                break

    # If conflicts found and no resolution specified, return the conflict list to let user choose
    if conflicts and not conflict_resolution:
        return {"ok": True, "has_conflicts": True, "conflicts": conflicts}

    # Find next empty row in destination
    dst_next_row = 4
    while dst_ws.cell(row=dst_next_row, column=2).value:
        dst_next_row += 1

    # Find next folder number in destination
    existing_dst = sorted([
        d.name for d in dst_shop_dir.iterdir()
        if d.is_dir() and d.name.startswith("product-")
    ])
    try:
        last_dst_num = max(int(n.split("-")[1]) for n in existing_dst) if existing_dst else 0
    except:
        last_dst_num = 0

    synced = []
    conflict_map = {c["src_row"]: c for c in conflicts}
    new_folder_count = 0

    for row_num in rows:
        src_folder = src_ws.cell(row=row_num, column=2).value
        if not src_folder:
            continue

        is_conflict = row_num in conflict_map

        if is_conflict:
            if conflict_resolution == "skip":
                continue
            elif conflict_resolution == "merge":
                target_row = conflict_map[row_num]["dst_row"]
                target_folder = conflict_map[row_num]["dst_folder"]

                # Overwrite columns in Excel
                for col in range(2, 16):
                    val = src_ws.cell(row=row_num, column=col).value
                    if col == 2:
                        val = target_folder
                    if col == 14:
                        val = "⏳ Chờ đăng"
                    set_cell_value(dst_ws, target_row, col, val)

                # Optionally copy files
                if copy_files:
                    src_folder_path = src_shop_dir / str(src_folder)
                    dst_folder_path = dst_shop_dir / target_folder
                    if src_folder_path.exists():
                        shutil.copytree(src_folder_path, dst_folder_path, dirs_exist_ok=True)

                synced.append({"src": str(src_folder), "dst": target_folder, "action": "merged"})
                continue

        # Normal sync (no conflict)
        new_num = last_dst_num + new_folder_count + 1
        new_folder = f"product-{new_num:02d}"
        new_folder_count += 1

        # Write to destination Excel
        for col in range(2, 16):
            val = src_ws.cell(row=row_num, column=col).value
            if col == 2:
                val = new_folder
            if col == 14:
                val = "⏳ Chờ đăng"
            set_cell_value(dst_ws, dst_next_row, col, val)

        dst_next_row += 1

        # Copy folders
        src_folder_path = src_shop_dir / str(src_folder)
        dst_folder_path = dst_shop_dir / new_folder
        if copy_files:
            if src_folder_path.exists():
                shutil.copytree(src_folder_path, dst_folder_path, dirs_exist_ok=True)
            else:
                (dst_folder_path / "images").mkdir(parents=True, exist_ok=True)
                (dst_folder_path / "files").mkdir(parents=True, exist_ok=True)
        else:
            (dst_shop_dir / new_folder / "images").mkdir(parents=True, exist_ok=True)
            (dst_shop_dir / new_folder / "files").mkdir(parents=True, exist_ok=True)

        synced.append({"src": str(src_folder), "dst": new_folder, "action": "created"})

    dst_wb.save(dst_excel)
    return {"ok": True, "synced": len(synced), "items": synced, "target": target_shop_id}

@app.patch("/api/products/{row}")
async def update_product(row: int, request: Request):
    data = await request.json()
    if not isinstance(data, dict):
        raise HTTPException(400, "Payload phải là JSON object")
    updates, expected_version, expected_hash = _catalog_preconditions(data, request)
    try:
        save_to_excel(
            row,
            updates,
            expected_version=expected_version,
            expected_hash=expected_hash,
        )
    except CatalogWriteConflict as error:
        raise HTTPException(409, "Catalog changed; refresh before retrying") from error
    except CatalogRepositoryError as error:
        raise HTTPException(500, "Catalog update failed safely") from error
    return {"ok": True}

@app.post("/api/products/{row}/reset-status")
async def reset_status(row: int):
    save_to_excel(row, {"status": "⏳ Chờ đăng"})
    return {"ok": True, "status": "⏳ Chờ đăng"}

@app.post("/api/products")
async def create_product(request: Request):
    """
    Tạo folder sản phẩm mới (dạng product-XX) cùng với các folder con (files, images).
    Đồng thời thêm dòng tương ứng vào file Excel với các cấu hình mặc định.
    """
    # 1. Tìm số thứ tự tiếp theo của product folder
    shop_dir = SHOP_DIR()
    nums = []
    if shop_dir.exists():
        for item in shop_dir.iterdir():
            if item.is_dir() and item.name.startswith("product-"):
                try:
                    num = int(item.name.split("-")[1])
                    nums.append(num)
                except ValueError:
                    pass

    next_num = max(nums) + 1 if nums else 1
    folder_name = f"product-{next_num:02d}"

    # 2. Tạo folder và folder con
    prod_path = shop_dir / folder_name
    img_path = prod_path / "images"
    file_path = prod_path / "files"

    prod_path.mkdir(parents=True, exist_ok=True)
    img_path.mkdir(parents=True, exist_ok=True)
    file_path.mkdir(parents=True, exist_ok=True)

    # 3. Tìm dòng trống đầu tiên hoặc thêm ở cuối Excel
    wb = openpyxl.load_workbook(EXCEL_FILE())
    ws = wb["Listings"]

    target_row = None
    # Quét từ dòng 4 trở đi để tìm dòng trống (không có folder name ở cột B/2)
    for r in range(4, ws.max_row + 2):
        val = ws.cell(row=r, column=2).value
        if val is None or str(val).strip() == "":
            target_row = r
            break

    if not target_row:
        target_row = ws.max_row + 1

    # Ghi dữ liệu mặc định vào dòng trống này
    ws.cell(row=target_row, column=2, value=folder_name)  # B: Folder name
    ws.cell(row=target_row, column=5, value=4.99)         # E: Price
    ws.cell(row=target_row, column=11, value=999)         # K: Qty
    ws.cell(row=target_row, column=13, value="2020_2026") # M: When made
    ws.cell(row=target_row, column=14, value="⏳ Chờ đăng") # N: Status
    ws.cell(row=target_row, column=15, value="Digital Planner") # O: Section

    wb.save(EXCEL_FILE())

    broadcast(f"[System] 📁 Đã tạo folder mới: {folder_name} tại row {target_row}")
    return {
        "ok": True,
        "folder": folder_name,
        "row": target_row,
        "message": f"Đã tạo thành công folder {folder_name} (row {target_row})"
    }

@app.delete("/api/products/{row}")
async def delete_product(row: int, request: Request):
    """Xoá sản phẩm local khỏi cả catalog và folder."""
    if _batch_delete_lock.locked():
        raise HTTPException(409, "Một lệnh xoá đang chạy")

    async with _batch_delete_lock:
        return await _delete_local_products(request, row)

@app.post("/api/batch-delete")
async def batch_delete(req: Request):
    if _batch_delete_lock.locked():
        raise HTTPException(409, "Một lệnh xoá hàng loạt khác đang chạy")
    async with _batch_delete_lock:
        return await _batch_delete_unlocked(req)


async def _batch_delete_unlocked(req: Request):
    """Xoá nhiều sản phẩm khỏi Excel cùng lúc."""
    data = await _read_delete_payload(req)
    shop_id, _ = _parse_delete_request_payload(data, require_folder=False)

    raw_items = data.get("items")
    if not isinstance(raw_items, list) or not raw_items:
        raise HTTPException(400, "items phải là danh sách không rỗng")

    normalized_items = []
    seen_rows = set()
    seen_folders = set()
    for item in raw_items:
        if not isinstance(item, dict):
            raise HTTPException(400, "Mỗi phần tử trong items phải có định dạng {row, folder}")
        row = item.get("row")
        if isinstance(row, bool) or not isinstance(row, int):
            raise HTTPException(400, "Mỗi item phải có row dạng số nguyên")
        if row < 4:
            raise HTTPException(400, f"row {row} không hợp lệ (phải >= 4)")
        if row in seen_rows:
            raise HTTPException(400, f"row {row} bị lặp trong danh sách")
        folder = _validate_product_numbered_folder_name(item.get("folder"))
        if folder in seen_folders:
            raise HTTPException(400, f"folder bị lặp trong danh sách: {folder}")
        seen_folders.add(folder)
        seen_rows.add(row)
        normalized_items.append((row, folder))

    if not normalized_items:
        return {
            "ok": True,
            "shop": shop_id,
            "deleted": 0,
            "items": [],
            "deleted_folders": [],
        }

    excel_path = EXCEL_FILE()
    try:
        result = await _delete_local_products_unlocked(shop_id, normalized_items, excel_path)
        return result
    except HTTPException:
        raise
    except Exception as error:
        raise HTTPException(500, "Xoá hàng loạt thất bại") from error


async def _delete_local_products(request: Request, row: int):
    data = await _read_delete_payload(request)
    shop_id, folder = _parse_delete_request_payload(data)
    normalized_items = [(row, folder)]

    return await _delete_local_products_unlocked(shop_id, normalized_items, EXCEL_FILE())


async def _delete_local_products_unlocked(shop_id: str, normalized_items: list[tuple[int, str]], excel_path: Path):
    if not excel_path.exists():
        raise HTTPException(404, "Chưa có file Excel của shop đang hoạt động")
    if not normalized_items:
        return {"ok": True, "shop": shop_id, "deleted": 0, "items": [], "deleted_folders": []}

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Listings"]
    shop_dir = SHOP_DIR()
    for row, folder in normalized_items:
        if row < 4:
            raise HTTPException(400, "Row phải từ 4 trở lên")
        current_folder = str(ws.cell(row=row, column=2).value or "").strip()
        if not current_folder:
            raise HTTPException(409, f"Không tìm thấy sản phẩm hợp lệ tại row {row}")
        if current_folder != folder:
            raise HTTPException(
                409,
                f"Row {row} bị thay đổi: folder hiện tại '{current_folder}' không khớp '{folder}'",
            )
        if not (shop_dir / folder).exists():
            raise HTTPException(404, f"Không tìm thấy folder local: {folder}")

    moved_items: list[tuple[int, str, Path, Path]] = []
    deleted_items: list[dict[str, int | str]] = []

    backup_path = excel_path.with_name(
        f"{excel_path.stem}.backup_delete_{time.strftime('%Y%m%d_%H%M%S')}{excel_path.suffix}"
    )
    mutation_started = False
    try:
        shutil.copy2(excel_path, backup_path)
        mutation_started = True

        for row, folder in normalized_items:
            folder_path = shop_dir / folder
            if not folder_path.exists():
                raise HTTPException(404, f"Không tìm thấy folder local: {folder}")
            target_folder = _quarantine_destination(shop_dir, folder)
            shutil.move(str(folder_path), str(target_folder))
            moved_items.append((row, folder, folder_path, target_folder))

        for row, _folder in normalized_items:
            _clear_catalog_row(ws, row)

        wb.save(excel_path)
        for row, folder, _source, target in moved_items:
            deleted_items.append(_build_local_delete_metadata(row, folder, target))
        return {
            "ok": True,
            "shop": shop_id,
            "deleted": len(moved_items),
            "items": deleted_items,
            "deleted_folders": [item["quarantine_folder"] for item in deleted_items],
        }
    except Exception as error:
        restore_errors: list[str] = []
        if mutation_started:
            for row, folder, source_folder, target_folder in reversed(moved_items):
                try:
                    if target_folder.exists():
                        target_folder.parent.mkdir(parents=True, exist_ok=True)
                        shutil.move(str(target_folder), str(source_folder))
                except Exception as restore_error:
                    restore_errors.append(f"{folder}@row {row}: {restore_error}")

            try:
                shutil.copy2(backup_path, excel_path)
            except Exception as restore_error:
                restore_errors.append(f"khôi phục workbook: {restore_error}")

            if restore_errors:
                raise HTTPException(
                    500,
                    "Xoá local thất bại và khôi phục không hoàn toàn: "
                    + "; ".join(restore_errors),
                ) from error

            if isinstance(error, HTTPException):
                raise HTTPException(500, "Xoá local thất bại sau khi bắt đầu thao tác") from error
            raise HTTPException(500, "Lưu file Excel thất bại, đã khôi phục bản sao dự phòng") from error

        if isinstance(error, HTTPException):
            raise error
        raise HTTPException(500, str(error)) from error

@app.post("/api/fix-sections")
async def fix_sections():
    """Bất đầu set 'Digital Planner' cho tất cả row trong Excel đang có folder nhưng không có section."""
    wb = openpyxl.load_workbook(EXCEL_FILE())
    ws = wb["Listings"]
    updated = 0
    for row_num in range(4, ws.max_row + 1):
        folder  = ws.cell(row=row_num, column=2).value
        section = ws.cell(row=row_num, column=15).value
        if folder and not str(section or "").strip():
            set_cell_value(ws, row_num, 15, "Digital Planner")
            updated += 1
    wb.save(EXCEL_FILE())
    return {"ok": True, "updated": updated}

# ── API: Images ──────────────────────────────────────────────────────────────────
def get_product_by_row(row: int) -> dict:
    products = products_from_excel()
    p = next((p for p in products if p["row"] == row), None)
    if not p:
        raise HTTPException(404, "Product not found")
    return p

@app.get("/api/products/{row}/images")
async def list_images(row: int):
    p = get_product_by_row(row)
    img_dir = SHOP_DIR() / p["folder"] / "images"
    if not img_dir.exists():
        return {"images": []}
    images = []
    for image_path in img_dir.iterdir():
        if image_path.suffix.lower() not in IMG_EXTS:
            continue
        resolved = _renderable_image_url(p["folder"], image_path)
        if resolved:
            images.append({"name": image_path.name, "size": image_path.stat().st_size, **resolved})
    images.sort(key=lambda x: x["name"])
    return {"images": images}

@app.post("/api/products/{row}/images")
async def upload_images(row: int, files: list[UploadFile] = File(...)):
    p = get_product_by_row(row)
    img_dir = SHOP_DIR() / p["folder"] / "images"
    img_dir.mkdir(parents=True, exist_ok=True)
    staged: list[tuple[Path, Path]] = []
    used_names: set[str] = set()
    try:
        with tempfile.TemporaryDirectory(prefix=".image_upload_", dir=str(img_dir)) as stage_dir_name:
            stage_dir = Path(stage_dir_name)
            for index, upload in enumerate(files, start=1):
                safe_name = _safe_asset_filename(upload.filename or "", f"image-{index:02d}.png")
                if safe_name in used_names:
                    stem = Path(safe_name).stem or f"image-{index:02d}"
                    suffix = Path(safe_name).suffix
                    duplicate = 2
                    candidate = f"{stem}-{duplicate}{suffix}"
                    while candidate in used_names:
                        duplicate += 1
                        candidate = f"{stem}-{duplicate}{suffix}"
                    safe_name = candidate
                used_names.add(safe_name)
                source_suffix = Path(safe_name).suffix.lower()
                readiness_suffix = source_suffix if source_suffix in IMG_EXTS else ".png"
                stage_path = stage_dir / f"{index:04d}{readiness_suffix}"
                stage_path.write_bytes(await upload.read())
                staged.append((stage_path, img_dir / safe_name))

            try:
                AssetReadinessEngine().assert_ready([stage_path for stage_path, _ in staged])
            except AssetReadinessError as error:
                blocked_statuses = {item.status for item in error.report.blocked_items}
                status_code = 409 if blocked_statuses & {"checksum-mismatch", "dataless", "cloud-only"} else 400
                raise HTTPException(status_code, "Image upload blocked by asset readiness checks") from error

            for stage_path, destination in staged:
                os.replace(stage_path, destination)
    except HTTPException:
        raise
    except (OSError, ValueError) as error:
        raise HTTPException(400, "Image upload could not be staged safely") from error

    saved = [destination.name for _, destination in staged]
    return {"ok": True, "saved": saved}

@app.delete("/api/products/{row}/images/{filename}")
async def delete_image(row: int, filename: str):
    p = get_product_by_row(row)
    decoded_filename = urllib.parse.unquote(filename)
    path = SHOP_DIR() / p["folder"] / "images" / decoded_filename
    if path.exists():
        path.unlink()
    return {"ok": True}

@app.post("/api/products/{row}/open-folder")
async def open_folder(row: int, request: Request):
    """Open files/ or images/ folder in Finder."""
    p = get_product_by_row(row)
    data = await request.json()
    subfolder = data.get("type", "files")   # 'files' or 'images'
    folder_path = SHOP_DIR() / p["folder"] / subfolder
    folder_path.mkdir(parents=True, exist_ok=True)
    subprocess.Popen(["open", "-R", str(folder_path)])
    return {"ok": True, "path": str(folder_path)}


# ── File serving ─────────────────────────────────────────────────────────────────
@app.get("/files/{folder}/{subfolder}/{filename:path}")
def serve_file(folder: str, subfolder: str, filename: str):
    decoded_filename = urllib.parse.unquote(filename)
    path = SHOP_DIR() / folder / subfolder / decoded_filename
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(path)

# ── API: SSE Logs ────────────────────────────────────────────────────────────────
@app.get("/api/logs")
async def sse_logs(request: Request):
    queue: asyncio.Queue = asyncio.Queue(maxsize=500)
    _log_subscribers.append(queue)

    async def event_gen():
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    msg = await asyncio.wait_for(queue.get(), timeout=15)
                    yield f"data: {json.dumps({'msg': msg, 'ts': time.strftime('%H:%M:%S')})}\n\n"
                except asyncio.TimeoutError:
                    yield "data: {\"ping\": 1}\n\n"
        finally:
            try: _log_subscribers.remove(queue)
            except ValueError: pass

    return StreamingResponse(event_gen(), media_type="text/event-stream")

# ── API: Service Status ──────────────────────────────────────────────────────────
async def _check_watcher_async() -> bool:
    try:
        proc = await asyncio.create_subprocess_exec(
            "pgrep", "-f", "planner_watcher.py",
            stdout=asyncio.subprocess.DEVNULL,
            stderr=asyncio.subprocess.DEVNULL
        )
        await proc.wait()
        return proc.returncode == 0
    except:
        return False


# These status probes run on the dashboard's initial render and on a short
# polling cadence.  They must never make an unavailable *optional* local AI
# service feel like the dashboard itself is unavailable.
_RUNTIME_SERVICE_CHECK_TIMEOUT_SECONDS = 0.35
_RUNTIME_SERVICE_SNAPSHOT_TIMEOUT_SECONDS = 0.5


def _normalize_http_url(base_url: str) -> str:
    return base_url[:-1] if base_url.endswith("/") else base_url


async def _check_http_json_identity(
    base_url: str,
    path: str,
    *,
    required_fields: tuple[str, ...] = (),
    response_validator: Callable[[Any], bool] | None = None,
    timeout_seconds: float = _RUNTIME_SERVICE_CHECK_TIMEOUT_SECONDS,
) -> bool:
    try:
        import httpx as _httpx
    except Exception:
        return False

    url = f"{_normalize_http_url(base_url)}{path}"
    timeout = _httpx.Timeout(timeout=timeout_seconds)
    try:
        # The targets are fixed loopback services.  Do not consult proxy
        # environment variables here: a missing/broken proxy can otherwise
        # turn a local readiness probe into a multi-second network wait.
        async with _httpx.AsyncClient(timeout=timeout, trust_env=False) as client:
            response = await client.get(url)
        if response.status_code != 200:
            return False
        payload = response.json()
        if not isinstance(payload, Mapping):
            return False
        for field in required_fields:
            if field not in payload:
                return False
        if response_validator is not None:
            return bool(response_validator(payload))
        return True
    except Exception:
        return False


def _has_core_service_identity(
    payload: Mapping[str, Any],
    *,
    required_paths: tuple[str, ...] = (),
) -> bool:
    if payload.get("openapi") is None:
        return False
    paths = payload.get("paths")
    if not isinstance(paths, Mapping):
        return False
    return any(path in paths for path in required_paths)


def _has_openai_identity(payload: Mapping[str, Any]) -> bool:
    if payload.get("object") != "list":
        return False
    data = payload.get("data")
    return isinstance(data, list)


async def _check_vertex_service_async() -> bool:
    openapi_check, health_check = await asyncio.gather(
        _check_http_json_identity(
            VERTEX_URL,
            "/openapi.json",
            required_fields=("openapi", "paths"),
            response_validator=lambda payload: _has_core_service_identity(
                payload, required_paths=("/api/upload-and-analyze", "/api/generate")
            ),
        ),
        _check_http_json_identity(
            VERTEX_URL,
            "/api/health",
            required_fields=("status",),
            response_validator=lambda payload: isinstance(payload.get("status"), (str, int, bool)),
        ),
    )
    return openapi_check or health_check


async def _check_mlx_service_async() -> bool:
    models_check, openapi_check = await asyncio.gather(
        _check_http_json_identity(
            MLX_URL,
            "/v1/models",
            required_fields=("object", "data"),
            response_validator=_has_openai_identity,
        ),
        _check_http_json_identity(
            MLX_URL,
            "/health/openapi",
            required_fields=("openapi", "paths"),
            response_validator=lambda payload: _has_core_service_identity(
                payload, required_paths=("/v1/models",)
            ),
        ),
    )
    return models_check or openapi_check


@app.get("/api/services")
async def service_status():
    return await _service_status_snapshot()


async def _service_status_snapshot() -> dict[str, Any]:
    watcher_task = asyncio.create_task(_check_watcher_async())
    vertex_task = asyncio.create_task(_check_vertex_service_async())
    mlx_task = asyncio.create_task(_check_mlx_service_async())

    try:
        watcher, vertex_app, mlx_ai = await asyncio.wait_for(
            asyncio.gather(watcher_task, vertex_task, mlx_task),
            timeout=_RUNTIME_SERVICE_SNAPSHOT_TIMEOUT_SECONDS,
        )
    except (asyncio.TimeoutError, Exception):
        # A readiness indicator is strictly best-effort.  Return all services
        # as offline rather than holding up the main dashboard API if a future
        # probe regresses or an optional process becomes wedged.
        for task in (watcher_task, vertex_task, mlx_task):
            if not task.done():
                task.cancel()
        await asyncio.gather(watcher_task, vertex_task, mlx_task, return_exceptions=True)
        watcher, vertex_app, mlx_ai = False, False, False

    return {
        "vertex_app": vertex_app,
        "mlx_ai":     mlx_ai,
        "watcher":    watcher,
        "running":    list(dict.fromkeys(list(_running_processes.keys()) + list(_running_tasks.keys()))),
        "running_tasks": list(_running_tasks.keys()),
    }


@app.get("/api/runtime-health")
async def runtime_health() -> dict[str, Any]:
    services = await _service_status_snapshot()
    active_shop = get_active_shop()
    optional_service_readiness = {
        key: bool(value)
        for key, value in services.items()
        if isinstance(value, bool) and key in {"vertex_app", "mlx_ai", "watcher"}
    }
    return runtime_identity.runtime_health_payload(
        runtime_root=BASE_DIR,
        listen_host=_dashboard_listen_host(),
        listen_port=8090,
        service_readiness=services,
        core_service_readiness={"ok": True, "checks": {"dashboard_endpoint": True}},
        optional_service_readiness={
            "ok": all(optional_service_readiness.values())
            if optional_service_readiness
            else False,
            "checks": optional_service_readiness,
        },
        active_shop_id=_active_shop_id,
        active_shop_name=str(active_shop.get("name", "")).strip(),
    )

@app.post("/api/services/watcher/start")
async def start_watcher():
    watcher_py = str(Path.home() / "TN/App/Truyen ngan/automation/planner_watcher.py")
    pid_file = Path.home() / "Desktop" / "Auto_Etsy_Planner" / ".watcher.pid"
    pid_file.unlink(missing_ok=True)
    proc = subprocess.Popen([PYTHON_BIN, watcher_py],
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    broadcast(f"[WATCHER] 🚀 Started PID {proc.pid}")
    return {"ok": True, "pid": proc.pid}


# ── API: Sync Etsy Shop Manager ─────────────────────────────────────────────────
@app.post("/api/etsy/sync")
async def sync_etsy_shop():
    shop_id = _active_shop_id
    command, created = await _enqueue_operation(
        operation="etsy-shop-sync",
        shop_id=shop_id,
        target=shop_id,
        callback=lambda: _run_etsy_shop_sync(shop_id),
    )
    return JSONResponse(
        status_code=202 if created else 200,
        content={"ok": True, "created": created, "command": _operation_queue_public(command), "message": f"Đã thêm Sync Etsy shop vào hàng chờ ({shop_id})"},
    )


async def _run_etsy_shop_sync(shop_id: str):
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}
    cmd = [sys.executable, "-u", str(BASE_DIR / "etsy_shop_sync.py"), "--shop", shop_id]
    broadcast(f"[ETSY-SYNC] 🔄 Bắt đầu đồng bộ shop {shop_id} từ Etsy Shop Manager...")
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
        _running_processes["__ETSY_SYNC__"] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text and not any(n in text for n in _LOG_NOISE):
                    broadcast(f"[ETSY-SYNC] {text}")
        code = await proc.wait()
        if code == 0:
            broadcast("[ETSY-SYNC] ✅ Đồng bộ xong. Dashboard sẽ đọc lại Excel mới.")
        else:
            broadcast(f"[ETSY-SYNC] ❌ Lỗi exit {code}")
    except Exception as e:
        broadcast(f"[ETSY-SYNC] ❌ {e}")
    finally:
        _running_processes.pop("__ETSY_SYNC__", None)

# ── API: Sync Listing Metadata from Etsy ─────────────────────────────────────────
_MATCH_STOP_WORDS = {
    "a", "an", "and", "for", "from", "in", "of", "the", "to", "with",
    "digital", "download", "instant", "printable",
}


def _normalized_match_title(value: str) -> str:
    clean = html.unescape(str(value or "")).lower()
    return " ".join(re.findall(r"[a-z0-9]+", clean))


def _listing_match_score(remote_title: str, local_title: str) -> float:
    remote = _normalized_match_title(remote_title)
    local = _normalized_match_title(local_title)
    if not remote or not local:
        return 0.0
    ratio = SequenceMatcher(None, remote, local).ratio()
    remote_tokens = {token for token in remote.split() if token not in _MATCH_STOP_WORDS}
    local_tokens = {token for token in local.split() if token not in _MATCH_STOP_WORDS}
    common = len(remote_tokens & local_tokens)
    token_f1 = (2 * common / (len(remote_tokens) + len(local_tokens))) if remote_tokens and local_tokens else 0.0
    containment = (common / min(len(remote_tokens), len(local_tokens))) if remote_tokens and local_tokens else 0.0
    return round((ratio * 0.65) + (token_f1 * 0.25) + (containment * 0.10), 4)


def _extract_etsy_listing_id(url_text: str) -> str:
    match = re.search(r"/listing/(\d+)", str(url_text or ""))
    return match.group(1) if match else ""


def _is_metadata_sync_complete(details: dict) -> bool:
    title_ok = bool(str(details.get("title") or "").strip())
    desc_ok = bool(str(details.get("description") or "").strip())
    tags_ok = bool(str(details.get("tags") or "").strip())
    return title_ok and desc_ok and tags_ok


async def _sync_local_from_etsy(
    *,
    listing_id: str,
    row: int,
    shop_id: str,
    product_path: Path,
    excel_path: Path,
    sync_assets: bool = True,
) -> tuple[dict, dict, dict, bool, list[str]]:
    """Pull metadata/assets from Etsy for a local row and persist synced fields."""
    sync_log = _ctx_logger("sync_local_from_etsy", shop=shop_id, listing_id=listing_id)
    sync_log.info("sync start (row=%s, product=%s)", row, product_path.name)
    synced_fields: list[str] = []
    details: dict = {}
    asset_sync: dict = {}
    sync_status: dict = _extract_asset_sync_status({}, metadata_ok=False)

    details = await scrape_listing_details(
        listing_id,
        shop_id=shop_id,
        product_path=product_path if sync_assets else None,
    )
    if not details.get("ok"):
        sync_log.error("scrape returned no usable listing details")
        raise RuntimeError("Không đọc được thông tin từ Etsy")

    details.pop("ok", None)
    asset_sync = details.pop("_asset_sync", {})
    metadata_ok = _is_metadata_sync_complete(details)
    sync_status = _extract_asset_sync_status(
        asset_sync,
        metadata_ok=metadata_ok,
        include_assets=sync_assets,
    )

    details = {
        key: value for key, value in details.items()
        if value is not None
        and (key not in {"title", "description", "tags"} or str(value or "").strip())
    }
    synced_fields = sorted(details)
    if details:
        save_to_excel(row, details, excel_path=excel_path)

    sync_log.info(
        "sync complete (row=%s, fields=%s, overall=%s)",
        row,
        synced_fields,
        bool(sync_status.get("overall")),
    )
    return details, asset_sync, sync_status, bool(sync_status.get("overall")), synced_fields


def _mapped_etsy_listing_ids(products: list[dict] | None = None) -> set[str]:
    mapped: set[str] = set()
    for product in products if products is not None else products_from_excel():
        listing_id = _extract_etsy_listing_id(str(product.get("etsy_url") or ""))
        if listing_id:
            mapped.add(listing_id)
    return mapped


def _status_for_linked_etsy_listing(listing: dict | None, fallback_status: str = "") -> str:
    status_labels = {
        "active": "✅ Đã đăng",
        "draft": "✅ Đã đăng draft",
        "inactive": "⏸ Inactive trên Etsy",
        "expired": "⌛ Expired trên Etsy",
    }
    manager_status = str((listing or {}).get("managerStatus") or "").lower()
    if manager_status in status_labels:
        return status_labels[manager_status]
    current = str(fallback_status or "").strip()
    if "URL chưa xác minh" in current or ("Đã đăng" in current and "draft" in current.lower()):
        return "✅ Đã đăng draft"
    return current or "✅ Đã đăng draft"


@app.get("/api/etsy/match-suggestions/{listing_id}")
async def etsy_match_suggestions(listing_id: str, limit: int = 5):
    if not re.fullmatch(r"\d+", listing_id):
        raise HTTPException(400, "Listing ID Etsy không hợp lệ")
    limit = max(1, min(int(limit), 10))
    snapshot = latest_etsy_manager_snapshot()
    listing = next(
        (item for item in snapshot.get("listings", []) if str(item.get("id", "")) == listing_id),
        None,
    )
    if not listing:
        raise HTTPException(404, "Listing không có trong bản đồng bộ Etsy mới nhất")

    remote_title = str(listing.get("title") or "")
    candidates = []
    for product in products_from_excel():
        if str(product.get("etsy_url") or "").strip():
            continue
        score = _listing_match_score(remote_title, str(product.get("title") or ""))
        candidates.append({
            "row": product.get("row"),
            "folder": product.get("folder"),
            "title": product.get("title"),
            "keywords": product.get("keywords"),
            "status": product.get("status"),
            "price": product.get("price"),
            "thumb": product.get("thumb"),
            "image_count": product.get("image_count", 0),
            "pdf_count": product.get("pdf_count", 0),
            "score": score,
            "confidence": "high" if score >= 0.90 else "medium" if score >= 0.75 else "low",
        })
    candidates.sort(key=lambda item: (-item["score"], str(item.get("folder") or "")))
    suggestions = candidates[:limit]
    top_score = suggestions[0]["score"] if suggestions else 0.0
    second_score = suggestions[1]["score"] if len(suggestions) > 1 else 0.0
    auto_fill_folder = None
    if suggestions and top_score >= 0.90 and (top_score >= 0.96 or top_score - second_score >= 0.06):
        auto_fill_folder = suggestions[0]["folder"]

    return {
        "ok": True,
        "listing": {
            "id": listing_id,
            "title": remote_title,
            "status": listing.get("managerStatus"),
            "url": listing.get("url") or f"https://www.etsy.com/listing/{listing_id}",
        },
        "suggestions": suggestions,
        "auto_fill_folder": auto_fill_folder,
        "scanned_local_total": len(candidates),
    }


@app.get("/api/etsy/link-suggestions-for-folder/{folder}")
async def etsy_link_suggestions_for_folder(folder: str, limit: int = 5):
    """Suggest unmapped Etsy listings that likely belong to a local product folder."""
    folder = str(folder or "").strip()
    if not folder:
        raise HTTPException(400, "Thiếu folder local")
    limit = max(1, min(int(limit), 10))

    products = products_from_excel()
    local_product = next((p for p in products if p.get("folder") == folder), None)
    if not local_product:
        raise HTTPException(404, f"Không tìm thấy folder local: {folder}")
    if str(local_product.get("etsy_url") or "").strip():
        raise HTTPException(409, f"{folder} đã có link Etsy")

    snapshot = latest_etsy_manager_snapshot()
    mapped_ids = _mapped_etsy_listing_ids(products)
    local_title = str(local_product.get("title") or "")
    candidates = []
    for listing in snapshot.get("listings", []):
        listing_id = str(listing.get("id") or "").strip()
        if not listing_id.isdigit() or listing_id in mapped_ids:
            continue
        manager_status = str(listing.get("managerStatus") or listing.get("status") or "").lower()
        if manager_status and manager_status not in {"draft", "active"}:
            continue
        score = _listing_match_score(str(listing.get("title") or ""), local_title)
        candidates.append({
            "id": listing_id,
            "title": listing.get("title") or "",
            "status": listing.get("managerStatus") or listing.get("status") or "",
            "url": listing.get("url") or f"https://www.etsy.com/listing/{listing_id}",
            "score": score,
            "confidence": "high" if score >= 0.90 else "medium" if score >= 0.75 else "low",
        })
    candidates.sort(key=lambda item: (-item["score"], str(item.get("id") or "")))
    suggestions = candidates[:limit]
    top_score = suggestions[0]["score"] if suggestions else 0.0
    second_score = suggestions[1]["score"] if len(suggestions) > 1 else 0.0
    auto_fill_listing_id = None
    if suggestions and top_score >= 0.90 and (top_score >= 0.96 or top_score - second_score >= 0.06):
        auto_fill_listing_id = suggestions[0]["id"]

    return {
        "ok": True,
        "folder": folder,
        "row": local_product.get("row"),
        "title": local_title,
        "status": local_product.get("status"),
        "suggestions": suggestions,
        "auto_fill_listing_id": auto_fill_listing_id,
        "scanned_etsy_total": len(candidates),
        "snapshot_total": len(snapshot.get("listings", [])),
    }


@app.post("/api/etsy/map-listing")
async def map_etsy_listing(request: Request):
    data = await request.json()
    listing_id = str(data.get("listing_id") or "").strip()
    folder = str(data.get("folder") or "").strip()
    raw_url = str(data.get("etsy_url") or "").strip()
    if not listing_id and raw_url:
        listing_id = _extract_etsy_listing_id(raw_url)
    if not re.fullmatch(r"\d+", listing_id):
        raise HTTPException(400, "Listing ID Etsy không hợp lệ")
    if not folder:
        raise HTTPException(400, "Thiếu folder local")

    products = products_from_excel()
    local_product = next((p for p in products if p.get("folder") == folder), None)
    if not local_product:
        raise HTTPException(404, f"Không tìm thấy folder local: {folder}")

    for product in products:
        match = re.search(r"/listing/(\d+)", str(product.get("etsy_url", "")))
        if match and match.group(1) == listing_id and product.get("row") != local_product.get("row"):
            raise HTTPException(409, f"Listing này đã ghép với {product.get('folder')}")

    snapshot = latest_etsy_manager_snapshot()
    listing = next(
        (item for item in snapshot.get("listings", []) if str(item.get("id", "")) == listing_id),
        None,
    )
    allow_manual = bool(data.get("allow_manual")) or bool(raw_url)
    if not listing and not allow_manual:
        raise HTTPException(404, "Listing không có trong bản đồng bộ Etsy mới nhất")

    etsy_url = str((listing or {}).get("url") or raw_url or f"https://www.etsy.com/listing/{listing_id}")
    if not _extract_etsy_listing_id(etsy_url):
        etsy_url = f"https://www.etsy.com/listing/{listing_id}"
    status_value = _status_for_linked_etsy_listing(listing, local_product.get("status") or "")
    save_to_excel(local_product["row"], {
        "etsy_url": etsy_url,
        "status": status_value,
    })
    broadcast(f"[ETSY-MAP] ✅ {listing_id} → {folder}")
    return {
        "ok": True,
        "listing_id": listing_id,
        "folder": folder,
        "row": local_product["row"],
        "etsy_url": etsy_url,
        "status": status_value,
        "from_snapshot": bool(listing),
    }


@app.post("/api/etsy/create-local-listing")
async def create_local_product_from_etsy(request: Request):
    """Create a new local product row/folder, map it, then pull Etsy metadata."""
    data = await request.json()
    listing_id = str(data.get("listing_id") or "").strip()
    if not re.fullmatch(r"\d+", listing_id):
        raise HTTPException(400, "Listing ID Etsy không hợp lệ")

    try:
        metadata_only = _parse_request_bool(data, "metadata_only", False)
        sync_assets = _parse_request_bool(data, "sync_assets", True)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    if metadata_only:
        sync_assets = False

    snapshot = latest_etsy_manager_snapshot()
    listing = next(
        (item for item in snapshot.get("listings", []) if str(item.get("id", "")) == listing_id),
        None,
    )
    if not listing:
        raise HTTPException(404, "Listing không có trong bản đồng bộ Etsy mới nhất")

    manager_status = str(listing.get("managerStatus") or "").lower()
    status_labels = {
        "active": "✅ Đã đăng",
        "draft": "✅ Đã đăng draft",
        "inactive": "⏸ Inactive trên Etsy",
        "expired": "⌛ Expired trên Etsy",
    }
    etsy_url = str(listing.get("url") or f"https://www.etsy.com/listing/{listing_id}")
    remote_title = str(listing.get("title") or "").strip()

    async with _product_create_lock:
        shop_id = _active_shop_id
        shop_dir = SHOP_DIR()
        excel_path = EXCEL_FILE()
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Listings"]

        target_row = None
        existing_row = None
        existing_folder = None
        existing = False
        for row_num in range(4, ws.max_row + 2):
            folder_value = str(ws.cell(row=row_num, column=2).value or "").strip()
            url_value = str(ws.cell(row=row_num, column=16).value or "").strip()
            existing_match = _extract_etsy_listing_id(url_value)
            if existing_match and existing_match == listing_id:
                existing_row = row_num
                existing_folder = folder_value or None
                break
            if target_row is None and not folder_value:
                target_row = row_num

        reusable_slots = [] if not sync_assets else _find_reusable_empty_product_slots(ws, shop_dir)
        used_folders = set()

        if existing_row is not None:
            target_row = existing_row
            existing = True
            if existing_folder:
                folder_name = existing_folder
                product_path = shop_dir / folder_name
                product_path.mkdir(parents=True, exist_ok=True)
                (product_path / "images").mkdir(exist_ok=True)
                (product_path / "files").mkdir(exist_ok=True)
            else:
                slot = _allocate_product_slot(ws, shop_dir, existing_row, reusable_slots, used_folders)
                slot_row = slot["row"]
                target_row = existing_row
                folder_name = slot["folder"]
                product_path = slot["path"]
                backup_path = excel_path.with_name(
                    f"{excel_path.stem}.backup_create_from_etsy_{time.strftime('%Y%m%d_%H%M%S')}_{listing_id}{excel_path.suffix}"
                )
                shutil.copy2(excel_path, backup_path)

                try:
                    (product_path / "images").mkdir(parents=True, exist_ok=slot["reused"])
                    (product_path / "files").mkdir(parents=True, exist_ok=slot["reused"])
                    if slot_row != target_row:
                        _clear_catalog_row(ws, slot_row)
                    _clear_catalog_row(ws, target_row)
                    set_cell_value(ws, target_row, 2, folder_name)
                    set_cell_value(ws, target_row, 5, 4.99)
                    set_cell_value(ws, target_row, 8, remote_title)
                    set_cell_value(ws, target_row, 11, 999)
                    set_cell_value(ws, target_row, 13, "2020_2026")
                    set_cell_value(ws, target_row, 14, status_labels.get(manager_status, "⏳ Chờ đăng"))
                    set_cell_value(ws, target_row, 15, "Digital Planner")
                    set_cell_value(ws, target_row, 16, etsy_url)
                    set_cell_value(ws, target_row, 18, generate_sku(shop_id, folder_name))
                    wb.save(excel_path)
                except Exception:
                    if slot["reused"]:
                        _clear_product_asset_folder(product_path)
                    else:
                        shutil.rmtree(product_path, ignore_errors=True)
                    try:
                        shutil.copy2(backup_path, excel_path)
                    except Exception:
                        pass
                    existing = False
                    raise
        else:
            slot = _allocate_product_slot(ws, shop_dir, target_row or (ws.max_row + 1), reusable_slots, used_folders)
            target_row = slot["row"]
            folder_name = slot["folder"]
            product_path = slot["path"]
            backup_path = excel_path.with_name(
                f"{excel_path.stem}.backup_create_from_etsy_{time.strftime('%Y%m%d_%H%M%S')}_{listing_id}{excel_path.suffix}"
            )
            shutil.copy2(excel_path, backup_path)

            try:
                (product_path / "images").mkdir(parents=True, exist_ok=slot["reused"])
                (product_path / "files").mkdir(parents=True, exist_ok=slot["reused"])
                _clear_catalog_row(ws, target_row)
                set_cell_value(ws, target_row, 2, folder_name)
                set_cell_value(ws, target_row, 5, 4.99)
                set_cell_value(ws, target_row, 8, remote_title)
                set_cell_value(ws, target_row, 11, 999)
                set_cell_value(ws, target_row, 13, "2020_2026")
                set_cell_value(ws, target_row, 14, status_labels.get(manager_status, "⏳ Chờ đăng"))
                set_cell_value(ws, target_row, 15, "Digital Planner")
                set_cell_value(ws, target_row, 16, etsy_url)
                set_cell_value(ws, target_row, 18, generate_sku(shop_id, folder_name))
                wb.save(excel_path)
            except Exception:
                if slot["reused"]:
                    _clear_product_asset_folder(product_path)
                else:
                    shutil.rmtree(product_path, ignore_errors=True)
                try:
                    shutil.copy2(backup_path, excel_path)
                except Exception:
                    pass
                raise

        if existing:
            # Keep idempotency: if listing already has local row, reuse it.
            # Ensure etsy_url is set (could be missing in stale rows).
            save_to_excel(target_row, {"etsy_url": etsy_url}, excel_path=excel_path)

    broadcast(f"[ETSY-CREATE] {'[GHÉP LẠI] Đã có' if existing else 'Đã tạo'} {folder_name} và ghép Etsy {listing_id}; đang sync thông tin...")
    sync_ok = False
    sync_error = ""
    synced_fields = []
    asset_sync = {}
    sync_status = _extract_asset_sync_status({}, metadata_ok=False)
    expected_status = status_labels.get(manager_status, "⏳ Chờ đăng")
    try:
        _, asset_sync, sync_status, sync_ok, synced_fields = await _sync_local_from_etsy(
            listing_id=listing_id,
            row=target_row,
            shop_id=shop_id,
            product_path=product_path,
            excel_path=excel_path,
            sync_assets=sync_assets,
        )
    except Exception as exc:
        sync_error = str(exc)

    if not sync_error and not sync_ok:
        sync_error = "Sync Etsy không hoàn tất đủ (thiếu metadata hoặc assets)."
    if sync_error:
        try:
            save_to_excel(target_row, {"status": "⚠ Sync lỗi"}, excel_path=excel_path)
        except Exception as status_exc:
            sync_error = f"{sync_error}; không ghi được trạng thái lỗi: {status_exc}"
        broadcast(f"[ETSY-CREATE] ⚠️ Đã tạo {folder_name}, nhưng sync metadata chưa hoàn tất: {sync_error}")
        sync_ok = False
    else:
        try:
            save_to_excel(target_row, {"status": expected_status}, excel_path=excel_path)
        except Exception as status_exc:
            sync_error = f"Không ghi được trạng thái thành công từ Etsy Manager: {status_exc}"
            broadcast(f"[ETSY-CREATE] ⚠️ Đã tạo {folder_name}, nhưng ghi trạng thái thất bại: {sync_error}")
            try:
                save_to_excel(target_row, {"status": "⚠ Sync lỗi"}, excel_path=excel_path)
            except Exception as status_write_exc:
                sync_error = f"{sync_error}; không ghi được trạng thái lỗi: {status_write_exc}"
            sync_ok = False
        else:
            broadcast(f"[ETSY-CREATE] ✅ {'Ghép lại' if existing else 'Đã sync'} Etsy {listing_id} → {folder_name} (row {target_row})")

    return {
        "ok": True,
        "listing_id": listing_id,
        "folder": folder_name,
        "row": target_row,
        "etsy_url": etsy_url,
        "sync_ok": sync_ok,
        "sync_error": sync_error,
        "synced_fields": synced_fields,
        "assets": asset_sync,
        "sync_status": sync_status,
        "existing": existing,
        "metadata_ok": sync_status.get("metadata_ok", False),
        "assets_complete": sync_status.get("assets_complete", False),
    }


@app.post("/api/products/{row}/sync-from-etsy")
async def sync_from_etsy(row: int, request: Request = cast(Any, None)):
    p = get_product_by_row(row)
    payload = await _read_optional_json_payload(request)
    context = _resolve_update_context(row, p, payload, operation="etsy_sync_from")
    shop_id, folder, listing_id = context.shop_id, context.folder, context.listing_id
    product_path = BASE_DIR / "shops" / shop_id / folder
    excel_path = BASE_DIR / "shops" / shop_id / "Etsy_SEO_Generator.xlsx"

    async def run_listing_sync():
        sync_status = _extract_asset_sync_status({}, metadata_ok=False)
        _etsy_single_sync_busy_shops.add(shop_id)
        try:
            broadcast(f"[ETSY-SINGLE-SYNC] 🔄 Đang cào dữ liệu cho Listing {listing_id} ({folder})...")
            _, asset_sync, sync_status, sync_ok, synced_fields = await _sync_local_from_etsy(
                listing_id=listing_id,
                row=row,
                shop_id=shop_id,
                product_path=product_path,
                excel_path=excel_path,
            )
            sync_complete = bool(sync_ok and sync_status.get("overall"))
            if sync_complete:
                broadcast(f"[ETSY-SINGLE-SYNC] ✅ Hoàn tất đồng bộ Listing {listing_id} vào Excel dòng {row}!")
            else:
                broadcast(f"[ETSY-SINGLE-SYNC] ⚠️ Chưa hoàn tất đồng bộ Listing {listing_id}: metadata_ok={sync_status.get('metadata_ok')}, assets_complete={sync_status.get('assets_complete')}")
        except Exception as e:
            broadcast(f"[ETSY-SINGLE-SYNC] ❌ Thất bại: {str(e)}")
            raise
        finally:
            _etsy_single_sync_busy_shops.discard(shop_id)

    command, created = await _enqueue_operation(
        operation="etsy-listing-sync",
        shop_id=shop_id,
        target=f"{folder}:{listing_id}",
        callback=run_listing_sync,
    )
    return JSONResponse(
        status_code=202 if created else 200,
        content={"ok": True, "created": created, "queued": True, "command": _operation_queue_public(command)},
    )


def _local_asset_summary(folder_path: Path, subfolder: str, allowed_exts: set[str]) -> dict:
    asset_dir = folder_path / subfolder
    items = []
    if asset_dir.exists():
        for path in sorted(asset_dir.iterdir()):
            if not path.is_file() or path.name.startswith(".") or path.suffix.lower() not in allowed_exts:
                continue
            try:
                size_bytes = path.stat().st_size
            except OSError:
                size_bytes = 0
            items.append({"name": path.name, "size_bytes": size_bytes})
    return {
        "count": len(items),
        "total_bytes": sum(item["size_bytes"] for item in items),
        "items": items,
    }


def _seo_filename_words(name: str) -> str:
    stem = Path(name).stem
    stem = re.sub(r"[_-]+", " ", stem)
    stem = re.sub(r"\s+", " ", stem).strip()
    return stem


def _seo_asset_items(asset_dir: Path, allowed_exts: set[str], *, limit: int = 16) -> tuple[list[dict], int]:
    items = []
    total = 0
    if not asset_dir.exists():
        return items, total
    for path in sorted(asset_dir.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file() or path.name.startswith(".") or path.suffix.lower() not in allowed_exts:
            continue
        total += 1
        if len(items) >= limit:
            continue
        try:
            stat = path.stat()
            size_bytes = stat.st_size
            local_available = size_bytes > 0 and getattr(stat, "st_blocks", 1) > 0
        except OSError:
            size_bytes = 0
            local_available = False
        items.append({
            "path": path,
            "name": path.name,
            "size_bytes": size_bytes,
            "local_available": local_available,
        })
    return items, total


def _seo_zip_member_samples(zip_path: Path, *, limit: int = 24) -> list[str]:
    if zip_path.suffix.lower() != ".zip" or not _asset_file_is_usable(zip_path):
        return []
    try:
        import zipfile
        with zipfile.ZipFile(zip_path) as archive:
            samples = []
            for name in archive.namelist():
                normalized = name.strip("/")
                base = Path(normalized).name
                if not normalized or normalized.endswith("/") or "__MACOSX" in normalized:
                    continue
                if base.startswith(".") or base.startswith("._"):
                    continue
                samples.append(normalized)
                if len(samples) >= limit:
                    break
            return samples
    except Exception:
        return []


def _build_seo_asset_context(p: dict) -> str:
    folder = str(p.get("folder") or "")
    captured_shop_dir = p.get("_shop_dir")
    shop_dir = Path(captured_shop_dir) if captured_shop_dir else SHOP_DIR()
    folder_path = shop_dir / folder
    image_items, image_total = _seo_asset_items(folder_path / "images", IMG_EXTS)
    file_items, file_total = _seo_asset_items(
        folder_path / "files",
        {".pdf", ".zip", ".001", ".002", ".003", ".004", ".005"},
        limit=10,
    )

    evidence_names = [folder, str(p.get("seed_title") or "")]
    evidence_names.extend(item["name"] for item in image_items)
    evidence_names.extend(item["name"] for item in file_items)

    zip_samples = []
    for item in file_items:
        samples = _seo_zip_member_samples(item["path"])
        if samples:
            zip_samples.append((item["name"], samples))
            evidence_names.extend(samples)

    tokens = []
    for name in evidence_names:
        for token in re.findall(r"[A-Za-z][A-Za-z0-9]+", _seo_filename_words(name).lower()):
            if len(token) >= 3 and token not in {"product", "image", "images", "bundle", "file", "files", "png", "jpg", "jpeg", "pdf", "zip"}:
                tokens.append(token)
    keyword_preview = ", ".join(dict.fromkeys(tokens[:40]))

    existing_title = str(p.get("title") or "").strip()
    title_looks_planner = bool(re.search(r"\b(planner|goodnotes|ipad|neurodivergent|daily|weekly|monthly)\b", existing_title, re.I))
    asset_looks_design_bundle = bool(re.search(
        r"\b(svg|dxf|eps|ai|cricut|silhouette|sublimation|clipart|wildflower|kindness|teacher)\b",
        " ".join(evidence_names),
        re.I,
    ))

    lines = [
        f"Folder: {folder}",
        f"Factory seed title: {p.get('seed_title') or 'Not provided'}",
        f"Existing SEO title in form/Excel: {existing_title or 'Not provided'}",
        f"Local listing image count: {image_total}",
        "Local listing image filenames: " + (
            "; ".join(item["name"] for item in image_items) if image_items else "None"
        ),
        f"Local digital file count: {file_total}",
        "Local digital filenames: " + (
            "; ".join(f"{item['name']} ({round(item['size_bytes'] / 1_000_000, 2)} MB)" for item in file_items) if file_items else "None"
        ),
    ]
    if zip_samples:
        for zip_name, samples in zip_samples[:3]:
            lines.append(f"ZIP sample contents for {zip_name}: " + "; ".join(samples))
    if keyword_preview:
        lines.append(f"Asset-derived terms: {keyword_preview}")
    if title_looks_planner and asset_looks_design_bundle:
        lines.append(
            "Conflict warning: existing SEO title looks like a planner listing, but local assets look like an SVG/PNG design bundle. Treat the existing title as stale."
        )
    return "\n".join(lines)


@app.get("/api/products/{row}/etsy-comparison")
async def compare_local_product_with_etsy(row: int):
    """Return a read-only Local vs Etsy preview before selected fields are pushed."""
    p = get_product_by_row(row)
    etsy_url = str(p.get("etsy_url") or "").strip()
    listing_match = re.search(r"/listing/(\d+)", etsy_url)
    if not listing_match:
        raise HTTPException(400, "Sản phẩm chưa có Etsy listing ID hợp lệ")

    shop_id = _active_shop_id
    folder_path = BASE_DIR / "shops" / shop_id / p["folder"]
    local_tags = [tag.strip() for tag in str(p.get("tags") or "").split(",") if tag.strip()]
    local = {
        "title": {"value": str(p.get("title") or ""), "chars": len(str(p.get("title") or ""))},
        "description": {
            "value": str(p.get("description") or ""),
            "chars": len(str(p.get("description") or "")),
        },
        "tags": {"value": str(p.get("tags") or ""), "count": len(local_tags)},
        "price": {"value": float(p.get("price") or 0)},
        "qty": {"value": int(p.get("qty") or 0)},
        "images": _local_asset_summary(folder_path, "images", IMG_EXTS),
        "files": _local_asset_summary(
            folder_path,
            "files",
            {".pdf", ".zip", ".001", ".002", ".003", ".004", ".005"},
        ),
    }

    listing_id = listing_match.group(1)
    broadcast(f"[ETSY-COMPARE] 🔎 Đang đọc Etsy {listing_id} để so sánh với {p['folder']}...")
    try:
        async with _etsy_compare_lock:
            remote_details = await scrape_listing_details(
                listing_id,
                shop_id=shop_id,
                include_asset_summary=True,
            )
    except Exception as exc:
        broadcast(f"[ETSY-COMPARE] ❌ {p['folder']}: {exc}")
        raise HTTPException(502, f"Không đọc được listing Etsy: {exc}")

    remote_tags = [tag.strip() for tag in str(remote_details.get("tags") or "").split(",") if tag.strip()]
    remote_assets = remote_details.get("assets") or {}
    remote = {
        "title": {
            "value": str(remote_details.get("title") or ""),
            "chars": len(str(remote_details.get("title") or "")),
        },
        "description": {
            "value": str(remote_details.get("description") or ""),
            "chars": len(str(remote_details.get("description") or "")),
        },
        "tags": {"value": str(remote_details.get("tags") or ""), "count": len(remote_tags)},
        "price": {"value": float(remote_details.get("price") or 0)},
        "qty": {"value": int(remote_details.get("qty") or 0)},
        "images": remote_assets.get("images") or {"count": 0},
        "files": remote_assets.get("files") or {"count": 0, "total_bytes": None, "items": []},
    }
    broadcast(f"[ETSY-COMPARE] ✅ Đã đọc dữ liệu Etsy {listing_id}")
    return {
        "ok": True,
        "shop_id": shop_id,
        "folder": p["folder"],
        "listing_id": listing_id,
        "local": local,
        "etsy": remote,
    }


# ── API: Post to Etsy ────────────────────────────────────────────────────────────
@app.get("/api/etsy/session")
async def get_etsy_session():
    shop_id = _active_shop_id
    try:
        return {"ok": True, "session": _etsy_session_payload(shop_id)}
    except (KeyError, ValueError) as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=409)


@app.post("/api/etsy/session/open")
async def open_etsy_session(request: Request):
    """Open the active shop's exact poster profile at Shop Manager."""
    shop_id = _active_shop_id
    try:
        body = await request.json()
    except (ValueError, json.JSONDecodeError):
        body = {}
    requested_shop_id = str(body.get("shop_id") or "").strip()
    if not requested_shop_id:
        return JSONResponse(
            {"ok": False, "error": "Thiếu shop_id để mở session Etsy"},
            status_code=400,
        )
    if requested_shop_id != shop_id:
        return JSONResponse(
            {
                "ok": False,
                "error": (
                    f"Shop không khớp: dashboard={shop_id or 'chưa chọn'}, "
                    f"yêu cầu={requested_shop_id}"
                ),
            },
            status_code=409,
        )
    try:
        _assert_shop_identity(shop_id)
        session = resolve_etsy_session(BASE_DIR, SHOPS, shop_id)
        ready = await asyncio.to_thread(open_etsy_login_browser, session)
        _assert_shop_identity(shop_id)
    except EtsyProfileLockedError as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=409)
    except (FileNotFoundError, KeyError, OSError, ValueError, RuntimeError) as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=409)
    if not ready:
        return JSONResponse(
            {
                "ok": False,
                "error": "Chrome không mở được đúng profile Etsy của shop",
            },
            status_code=503,
        )
    payload = _etsy_session_payload(shop_id)
    payload["browser_ready"] = True
    return {
        "ok": True,
        "session": payload,
        "message": (
            f"Đã mở Shop Manager bằng profile Etsy dùng để Post cho {shop_id}. "
            "Sau khi đăng nhập, có thể giữ cửa sổ này mở khi nhấn Post."
        ),
    }


@app.post("/api/products/{row}/post")
async def post_to_etsy(row: int):
    p = get_product_by_row(row)
    folder = p["folder"]
    shop_id = _active_shop_id

    async def run_post():
        save_to_excel(row, {"status": "⏳ Chờ đăng"})
        broadcast(f"[DASH] 🔄 Reset → ⏳ Chờ đăng: {folder}")
        await _run_poster(row, folder, shop_id)

    command, created = await _enqueue_operation(
        operation="etsy-post",
        shop_id=shop_id,
        target=f"{row}:{folder}",
        callback=run_post,
    )
    return JSONResponse(
        status_code=202 if created else 200,
        content={"ok": True, "created": created, "command": _operation_queue_public(command), "message": f"Đã thêm Post {folder} vào hàng chờ"},
    )


@app.post("/api/run-selected-products")
async def run_selected_products(req: Request):
    data = await _read_delete_payload(req)
    shop_id, items = _parse_run_selected_request_payload(data)
    valid_items, rejected = _partition_selected_local_products(shop_id, items)
    excel_path = EXCEL_FILE()

    # Mark skipped products with per-card error reasons; do not abort the batch.
    if rejected:
        try:
            _mark_selected_row_errors(excel_path, rejected)
        except Exception as error:
            raise HTTPException(500, "Không ghi được trạng thái lỗi cho sản phẩm bị bỏ qua") from error
        for item in rejected:
            broadcast(f"[BATCH] ⏭ Bỏ qua {item['folder']}: {item['reason']}")

    if not valid_items:
        raise HTTPException(
            400,
            f"Không có sản phẩm hợp lệ để đăng trong {len(items)} đã chọn"
            + (f" (đã bỏ qua {len(rejected)})" if rejected else ""),
        )

    if _is_poster_locked_for_shop(shop_id):
        raise HTTPException(
            409,
            "Một bài đăng đang chạy cho shop này. Vui lòng đợi hoàn tất rồi thử lại",
        )

    rows = [row for row, _ in valid_items]
    lock_key = _acquire_poster_lock(shop_id)
    try:
        _set_selected_rows_pending(excel_path, rows)
    except HTTPException:
        _release_poster_lock(shop_id)
        raise
    except Exception as error:
        _release_poster_lock(shop_id)
        raise HTTPException(500, "Đặt trạng thái Chờ đăng thất bại") from error

    try:
        task = asyncio.create_task(_run_selected_poster(shop_id, valid_items, lock_key=lock_key))
    except Exception:
        _release_poster_lock(shop_id)
        raise
    _register_background_task(lock_key, task)
    skipped_payload = [
        {"row": item["row"], "folder": item["folder"], "reason": item["reason"]}
        for item in rejected
    ]
    message = f"Đã xếp hàng {len(valid_items)} sản phẩm vào queue, chạy 1 Chrome lần lượt"
    if rejected:
        message += f" (bỏ qua {len(rejected)} sản phẩm không đủ điều kiện)"
    return JSONResponse(
        status_code=202,
        content={
            "ok": True,
            "shop": shop_id,
            "queued": len(valid_items),
            "skipped": len(rejected),
            "items": [{"row": row, "folder": folder} for row, folder in valid_items],
            "folders": [folder for _, folder in valid_items],
            "rejected": skipped_payload,
            "job": lock_key,
            "message": message,
        },
    )


def _assert_product_images_ready_for_push(context: OperationContext) -> None:
    shop_root = (BASE_DIR / "shops" / context.shop_id).resolve()
    product_root = (shop_root / context.folder).resolve()
    if product_root.parent != shop_root:
        raise HTTPException(400, "Product asset path is invalid")

    images_dir = product_root / "images"
    try:
        image_paths = sorted(
            path for path in images_dir.iterdir()
            if path.is_file() and path.suffix.lower() in IMG_EXTS
        )
    except FileNotFoundError:
        image_paths = []
    except OSError as error:
        raise HTTPException(409, "Local image assets could not be inspected") from error

    if not image_paths:
        raise HTTPException(400, "No local image assets are available for image push")
    try:
        AssetReadinessEngine().assert_ready(image_paths)
    except AssetReadinessError as error:
        blocked_statuses = {item.status for item in error.report.blocked_items}
        status_code = 409 if blocked_statuses & {"checksum-mismatch", "dataless", "cloud-only"} else 400
        raise HTTPException(status_code, "Image push blocked by asset readiness checks") from error


@app.post("/api/products/{row}/push-to-etsy")
async def push_local_updates_to_etsy(row: int, request: Request):
    """Push selected local fields into the product's already-mapped Etsy listing."""
    p = get_product_by_row(row)
    data = await _read_optional_json_payload(request)
    if data is None:
        raise HTTPException(400, "Payload phải là JSON object")
    context = _resolve_update_context(
        row=row,
        product=p,
        payload=data,
        operation="etsy_push_update",
    )
    requested_fields = data.get("fields", []) if isinstance(data, dict) else []
    if not isinstance(requested_fields, list):
        raise HTTPException(400, "fields phải là danh sách")

    valid_fields = {"title", "description", "tags", "price", "qty", "images", "files"}
    fields = [field for field in requested_fields if field in valid_fields]
    fields = list(dict.fromkeys(fields))
    if not fields:
        raise HTTPException(400, "Chưa chọn nội dung cần cập nhật")
    if "images" in fields:
        _assert_product_images_ready_for_push(context)

    dedupe_key = context.context_key
    job_record, created = _get_job_store().create_or_get_deduplicated_job(
        shop_id=context.shop_id,
        operation=context.operation,
        row=context.row,
        folder=context.folder,
        listing_id=context.listing_id,
        request_id=context.request_id,
        dedupe_key=dedupe_key,
        operation_receipt={
            "listing_id": context.listing_id,
            "fields": fields,
            "requested_shop": context.shop_id,
            "request_id": context.request_id,
        },
        fields=fields,
    )
    if not created and str(job_record.get("status")) in {JOB_STATUS_QUEUED, JOB_STATUS_RUNNING}:
        return {
            "ok": True,
            "created": False,
            "message": f"Update {context.folder} đã có trong hàng chờ",
            "listing_id": context.listing_id,
            "fields": fields,
            "job_id": str(job_record.get("job_id") or ""),
        }

    job_id = await _queue_etsy_update_job(context, fields, job_record)
    return {
        "ok": True,
        "message": f"Đang cập nhật {context.folder} lên Etsy {context.listing_id}",
        "listing_id": context.listing_id,
        "fields": fields,
        "job_id": job_id,
    }


@app.get("/api/etsy/update-status")
async def etsy_update_status(job_id: str):
    store_job = _get_job_store().get_job(job_id)
    if store_job:
        job = _job_payload_for_frontend(store_job)
        _cache_job_status(job_id, job)
    else:
        job = _etsy_update_jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Không tìm thấy tiến trình cập nhật")
    return {"ok": True, **(_job_payload_for_frontend(job) or job)}


@app.get("/api/etsy/jobs")
async def list_etsy_jobs(status: str | None = None, limit: int = 50):
    """Read the active shop's durable job inbox without exposing raw receipts."""
    try:
        bounded_limit = max(1, min(int(limit), 100))
    except (TypeError, ValueError) as error:
        raise HTTPException(400, "limit phải là số nguyên") from error
    requested_status: set[str] | None = None
    if status:
        requested_status = {item.strip().lower() for item in status.split(",") if item.strip()}
        allowed = {
            JOB_STATUS_QUEUED,
            JOB_STATUS_RUNNING,
            JOB_STATUS_SUCCEEDED,
            JOB_STATUS_FAILED,
            JOB_STATUS_CANCELLED,
        }
        if not requested_status.issubset(allowed):
            raise HTTPException(400, "status job không hợp lệ")
    records = _get_job_store().list_jobs(
        shop_id=_active_shop_id,
        status=requested_status,
        limit=bounded_limit,
    )
    return {
        "ok": True,
        "shop_id": _active_shop_id,
        "jobs": [payload for record in records if (payload := _job_center_payload(record))],
    }


def _job_context_from_record(job: dict[str, object]) -> OperationContext:
    listing_id = str(job.get("listing_id") or "").strip()
    folder = str(job.get("folder") or "").strip()
    shop_id = str(job.get("shop_id") or "").strip()
    row = _coerce_int(job.get("row"), 0)
    if not listing_id or not folder or not shop_id or row < 1:
        raise HTTPException(409, "Job thiếu định danh an toàn để thao tác")
    try:
        return OperationContext.from_request(
            operation=str(job.get("operation") or "etsy_push_update"),
            row=row,
            payload={
                "shop": shop_id,
                "folder": folder,
                "listing_id": listing_id,
            },
            active_shop_id=_active_shop_id,
            current_folder=folder,
            current_etsy_url=f"https://www.etsy.com/listing/{listing_id}",
        )
    except OperationContextError as error:
        _raise_context_error(error)
        raise  # pragma: no cover


@app.post("/api/etsy/jobs/{job_id}/cancel")
async def cancel_etsy_job(job_id: str):
    store = _get_job_store()
    job = store.get_job(job_id)
    if not job:
        raise HTTPException(404, "Không tìm thấy job")
    if str(job.get("shop_id")) != str(_active_shop_id):
        raise HTTPException(409, "Job không thuộc shop đang hoạt động")
    if str(job.get("status")) not in {JOB_STATUS_QUEUED, JOB_STATUS_RUNNING}:
        raise HTTPException(409, "Job đã kết thúc, không thể huỷ lại")

    task = _running_tasks.get(job_id)
    if task is not None and not task.done():
        task.cancel()
        await asyncio.gather(task, return_exceptions=True)
    proc = _running_processes.get(job_id)
    if proc is not None:
        try:
            proc.kill()
        except (ProcessLookupError, OSError):
            pass
    store.cancel_job(job_id)
    _cache_job_status(
        job_id,
        {
            "status": "cancelled",
            "finished_at": time.time(),
            "last_message": "Đã hủy bởi Job Center",
        },
    )
    return {"ok": True, "job_id": job_id, "status": "cancelled"}


@app.post("/api/etsy/jobs/{job_id}/retry")
async def retry_etsy_job(job_id: str):
    store = _get_job_store()
    previous = store.get_job(job_id)
    if not previous:
        raise HTTPException(404, "Không tìm thấy job")
    if str(previous.get("shop_id")) != str(_active_shop_id):
        raise HTTPException(409, "Job không thuộc shop đang hoạt động")
    previous_status = str(previous.get("status") or "")
    if previous_status in {JOB_STATUS_QUEUED, JOB_STATUS_RUNNING}:
        raise HTTPException(409, "Job đang chạy, không thể retry")
    if previous_status == JOB_STATUS_SUCCEEDED:
        raise HTTPException(409, "Job đã thành công, không cần retry")
    if previous_status not in {JOB_STATUS_FAILED, JOB_STATUS_CANCELLED}:
        raise HTTPException(409, "Job không ở trạng thái retry được")

    context = _job_context_from_record(previous)
    fields = previous.get("fields")
    if not isinstance(fields, list) or not fields:
        raise HTTPException(409, "Job không lưu được fields để retry an toàn")
    normalized_fields = [str(field) for field in fields if str(field) in {"title", "description", "tags", "price", "qty", "images", "files"}]
    if not normalized_fields:
        raise HTTPException(409, "Job không có field hợp lệ để retry")
    if "images" in normalized_fields:
        _assert_product_images_ready_for_push(context)

    retry_record, created = store.create_or_get_deduplicated_job(
        shop_id=context.shop_id,
        operation=context.operation,
        row=context.row,
        folder=context.folder,
        listing_id=context.listing_id,
        request_id=context.request_id,
        dedupe_key=context.context_key,
        operation_receipt=previous.get("operation_receipt") or {"retry_parent_job_id": job_id},
        fields=normalized_fields,
        parent_job_id=job_id,
    )
    if not created:
        raise HTTPException(409, "Đã có một retry đang chờ hoặc đang chạy")
    new_job_id = await _queue_etsy_update_job(context, normalized_fields, retry_record)
    return {
        "ok": True,
        "job_id": new_job_id,
        "retry_parent_job_id": job_id,
        "status": JOB_STATUS_QUEUED,
    }


async def _run_etsy_updater(
    context: OperationContext,
    fields: list[str],
    job_id: str,
):
    job = _etsy_update_jobs.get(job_id, {})
    folder = context.folder
    shop_id = context.shop_id
    listing_id = context.listing_id
    row = context.row
    proc: asyncio.subprocess.Process | None = None
    terminal_status: str | None = None
    terminal_message: str | None = None
    store = _get_job_store()
    runtime_key = job_id
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}
    try:
        _cache_job_status(
            job_id,
            {
                "operation": context.operation,
                "status": "starting",
                "folder": folder,
                "listing_id": listing_id,
                "shop_id": shop_id,
                "row": row,
            },
        )
        preflight_ok, preflight_msg = await _runtime_prefetch_import_check(
            PYTHON_BIN,
            f"[ETSY-UPDATE] {folder}",
            _RUNTIME_PRECHECK_MODULES_UPDATER,
        )
        if not preflight_ok:
            terminal_status = "error"
            terminal_message = preflight_msg
            broadcast(f"[ETSY-UPDATE] ❌ {preflight_msg}")
            return
        cmd = [
            PYTHON_BIN, "-u", ETSY_UPDATER,
            "--listing-id", listing_id,
            "--row", str(row),
            "--shop", shop_id,
            "--fields", ",".join(fields),
        ]
        broadcast(
            f"[ETSY-UPDATE] Runtime started: {folder} → {listing_id} "
            f"({', '.join(fields)})"
        )
        _cache_job_status(
            job_id,
            {
                "status": "preflight",
                "last_message": "Runtime started; creating Chrome process",
                "operation": context.operation,
            },
        )
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
            cwd=str(BASE_DIR),
            env=env,
        )
        _running_processes[runtime_key] = proc
        _cache_job_status(
            job_id,
            {
                "status": "running",
                "pid": proc.pid,
                "last_message": "Runtime started; waiting for Chrome marker",
                "operation": context.operation,
            },
        )
        store.mark_running(job_id, pid=proc.pid)
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text and not any(noise in text for noise in _LOG_NOISE):
                    broadcast(f"[ETSY-UPDATE] {text}")
                    _cache_job_status(
                        job_id,
                        {
                            "last_message": text,
                            "operation": context.operation,
                        },
                    )
                    store.append_log_excerpt(job_id, text)
        code = await proc.wait()
        result = "✅ Xong" if code == 0 else f"❌ Lỗi exit {code}"
        terminal_status = "success" if code == 0 else "error"
        _cache_job_status(job_id, {"exit_code": code, "operation": context.operation})
        terminal_message = result if code == 0 else (job.get("last_message") or result)
        broadcast(f"[ETSY-UPDATE] {result}: {folder} → {listing_id}")
    except asyncio.CancelledError:
        terminal_status = "cancelled"
        terminal_message = "Đã hủy cập nhật"
        broadcast(f"[ETSY-UPDATE] ⚠️ Hủy: {folder} → {listing_id}")
        raise
    except Exception as exc:
        terminal_status = "error"
        terminal_message = str(exc)
        broadcast(f"[ETSY-UPDATE] ❌ {folder}: {exc}")
    finally:
        await _terminate_subprocess(proc)
        _pop_background_task(runtime_key)
        _running_processes.pop(runtime_key, None)
        terminal_status = terminal_status or "error"
        if terminal_status == "success":
            store.mark_succeeded(job_id, exit_code=0 if proc is None else proc.returncode or 0, log_excerpt=terminal_message or "")
        elif terminal_status == "cancelled":
            store.cancel_job(job_id)
            _cache_job_status(job_id, {"operation": context.operation, "status": "cancelled", "finished_at": time.time(), "last_message": terminal_message or ""})
        else:
            store.mark_failed(job_id, exit_code=(proc.returncode if proc is not None else None) or 1, log_excerpt=terminal_message or "")
        _cache_job_status(
            job_id,
            {
                "operation": context.operation,
                "status": terminal_status,
                "finished_at": time.time(),
                "last_message": terminal_message or "",
            },
        )

async def _run_poster(row: int, folder: str, shop_id: str, lock_key: str | None = None):
    if lock_key is None:
        lock_key = _etsy_post_lock_key(shop_id)
    proc: asyncio.subprocess.Process | None = None
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}
    try:
        preflight_ok, preflight_msg = await _runtime_prefetch_import_check(
            PYTHON_BIN,
            f"[POSTER] {folder}",
            _RUNTIME_PRECHECK_MODULES_POSTER,
        )
        if not preflight_ok:
            status_msg = f"❌ Lỗi: {preflight_msg}"
            save_to_excel(row, {"status": status_msg})
            broadcast(f"[POSTER] ❌ {preflight_msg}")
            return
        cmd = [PYTHON_BIN, "-u", ETSY_POSTER, "--product", folder, "--shop", shop_id]
        broadcast(f"[POSTER] 🚀 Bắt đầu: {folder}")
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
        _running_processes[folder] = proc
        _running_processes[lock_key] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text and not any(n in text for n in _LOG_NOISE):
                    broadcast(f"[POSTER] {text}")
        code = await proc.wait()
        broadcast(f"[POSTER] {'✅ Xong' if code == 0 else f'❌ Lỗi exit {code}'}: {folder}")
    except Exception as e:
        broadcast(f"[POSTER] ❌ {e}")
    finally:
        await _terminate_subprocess(proc)
        _running_processes.pop(lock_key, None)
        _running_processes.pop(folder, None)
        _pop_background_task(folder)
        _pop_background_task(lock_key)

# ── API: Run ALL pending — 1 Chrome duy nhất, tuần tự ────────────────────────────
@app.post("/api/run-all-pending")
async def run_all_pending():
    shop_id = _active_shop_id
    if _is_poster_locked_for_shop(shop_id):
        return JSONResponse({"ok": False, "error": "Batch đang chạy rồi — chờ xong mới chạy tiếp"}, status_code=409)
    lock_key = _acquire_poster_lock(shop_id)
    try:
        task = asyncio.create_task(_run_batch_poster(shop_id, lock_key=lock_key))
    except Exception:
        _release_poster_lock(shop_id)
        raise
    _register_background_task(lock_key, task)
    return {"ok": True, "message": "Đã khởi động batch poster (1 Chrome)"}

# Noise patterns to filter from live logs
_LOG_NOISE = (
    "NotOpenSSLWarning", "urllib3", "warnings.warn(",
    "LibreSSL", "site-packages/urllib",
)

async def _run_batch_poster(shop_id: str, lock_key: str):
    proc: asyncio.subprocess.Process | None = None
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}
    try:
        preflight_ok, preflight_msg = await _runtime_prefetch_import_check(
            PYTHON_BIN,
            "[BATCH]",
            _RUNTIME_PRECHECK_MODULES_POSTER,
        )
        if not preflight_ok:
            broadcast(f"[BATCH] ❌ {preflight_msg}")
            return
        cmd = [PYTHON_BIN, "-u", ETSY_POSTER, "--batch", "200", "--shop", shop_id]
        broadcast("[BATCH] 🚀 Khởi động batch poster — 1 Chrome, xử lý tuần tự...")
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
        _running_processes[lock_key] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text and not any(n in text for n in _LOG_NOISE):
                    broadcast(f"[BATCH] {text}")
        code = await proc.wait()
        broadcast(f"[BATCH] {'✅ Hoàn tất tất cả sản phẩm!' if code == 0 else f'❌ Lỗi exit {code}'}")
    except Exception as e:
        broadcast(f"[BATCH] ❌ {e}")
    finally:
        await _terminate_subprocess(proc)
        _running_processes.pop(lock_key, None)
        _pop_background_task(lock_key)


async def _run_selected_poster(shop_id: str, items: list[tuple[int, str]], lock_key: str):
    cmd = [PYTHON_BIN, "-u", ETSY_POSTER, "--shop", shop_id]
    for row, folder in items:
        cmd.extend(["--selected-product", f"{row}:{folder}"])

    proc: asyncio.subprocess.Process | None = None
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}
    try:
        preflight_ok, preflight_msg = await _runtime_prefetch_import_check(
            PYTHON_BIN,
            "[BATCH]",
            _RUNTIME_PRECHECK_MODULES_POSTER,
        )
        if not preflight_ok:
            broadcast(f"[BATCH] ❌ {preflight_msg}")
            return
        broadcast("[BATCH] 🚀 Khởi động đăng hàng loạt local đã chọn — 1 Chrome, xử lý tuần tự...")
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
        _running_processes[lock_key] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text and not any(n in text for n in _LOG_NOISE):
                    broadcast(f"[BATCH] {text}")
        code = await proc.wait()
        broadcast(f"[BATCH] {'✅ Hoàn tất hàng loạt đã chọn!' if code == 0 else f'❌ Lỗi exit {code}'}")
    except Exception as e:
        broadcast(f"[BATCH] ❌ {e}")
    finally:
        await _terminate_subprocess(proc)
        _running_processes.pop(lock_key, None)
        _pop_background_task(lock_key)

# ── API: Stop all running posters ────────────────────────────────────────────────
@app.post("/api/stop-all")
async def stop_all():
    killed = []
    cancelled = []
    store = _get_job_store()
    process_cleanup = []
    for key, proc in list(_running_processes.items()):
        if proc is None:
            continue
        try:
            proc.kill()
            killed.append(key)
        except ProcessLookupError:
            pass
        except Exception:
            pass
        process_cleanup.append(_terminate_subprocess(proc))
    for key, task in list(_running_tasks.items()):
        if not task.done():
            task.cancel()
            cancelled.append(key)
    if process_cleanup:
        await asyncio.gather(*process_cleanup, return_exceptions=True)
    if cancelled:
        await asyncio.gather(*[_running_tasks[k] for k in cancelled if k in _running_tasks], return_exceptions=True)
    for job in store.list_jobs(status={JOB_STATUS_QUEUED, JOB_STATUS_RUNNING}):
        job_id = str(job.get("job_id") or "")
        if not job_id:
            continue
        store.cancel_job(job_id)
        _cache_job_status(
            job_id,
            {
                "status": "cancelled",
                "finished_at": time.time(),
                "last_message": "Đã hủy bởi stop-all",
            },
        )
    # A scheduled cloud upload is an external write too.  Clear both legacy
    # per-shop workers and the unified-queue cards before returning success, so
    # no command accepted before Stop all can begin afterwards.
    cloud_workers = list(_CLOUD_ASSET_UPLOAD_WORKERS.values())
    for worker in cloud_workers:
        if not worker.done():
            worker.cancel()
    if cloud_workers:
        await asyncio.gather(*cloud_workers, return_exceptions=True)
    async with _CLOUD_ASSET_UPLOAD_QUEUE_GUARD:
        _CLOUD_ASSET_UPLOAD_WORKERS.clear()
        _CLOUD_ASSET_UPLOAD_QUEUE_BY_SHOP.clear()
        _CLOUD_ASSET_UPLOAD_SCHEDULES.clear()
    async with _OPERATION_QUEUE_GUARD:
        for command in _OPERATION_QUEUE_COMMANDS.values():
            if command.get("status") == "queued":
                command.update({"status": "cancelled", "finished_at": time.time(), "updated_at": time.time()})
    _running_tasks.clear()
    _running_processes.clear()
    broadcast(f"[DASH] 🛑 Đã dừng {len(killed)} tiến trình, hủy {len(cancelled)} nhiệm vụ nền")
    return {"ok": True, "stopped": killed, "cancelled": cancelled}


# ── API: Regenerate images ───────────────────────────────────────────────────────
@app.post("/api/products/{row}/regenerate")
async def regenerate_images(row: int):
    p = get_product_by_row(row)
    folder = p["folder"]
    file_dir = SHOP_DIR() / folder / "files"

    # Find planner source
    planner_path = None
    if file_dir.exists():
        for f in sorted(file_dir.iterdir()):
            if f.suffix.lower() in IMG_EXTS | {".pdf"}:
                planner_path = str(f); break

    if not planner_path:
        return JSONResponse({"ok": False, "error": "Không tìm thấy file planner trong files/"}, 400)

    broadcast(f"[REGEN] 🖼 Bắt đầu regenerate: {folder}")
    asyncio.create_task(_run_regenerate(p, planner_path))
    return {"ok": True}

async def _run_regenerate(p: dict, planner_path: str):
    folder = p["folder"]
    img_dir = str(SHOP_DIR() / folder / "images")
    cmd = [PYTHON_BIN, TRIGGER_SCRIPT, planner_path, "--mode", "Single",
           "--output-dir", img_dir]
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.STDOUT)
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text: broadcast(f"[REGEN] {text}")
        code = await proc.wait()
        broadcast(f"[REGEN] {'✅ Ảnh mới xong' if code == 0 else f'❌ Lỗi exit {code}'}: {folder}")
    except Exception as e:
        broadcast(f"[REGEN] ❌ {e}")

# ── API: SEO regenerate via MLX ──────────────────────────────────────────────────
@app.post("/api/products/{row}/regen-seo")
async def regen_seo(row: int, request: Request):
    p = get_product_by_row(row)
    body = {}
    try: body = await request.json()
    except: pass

    # Use values from form if provided, else fall back to Excel
    p["title"]    = body.get("title",    p.get("title", ""))    or p["folder"]
    p["keywords"] = body.get("keywords", p.get("keywords", "")) or ""
    extra_hint    = body.get("extra", "")
    field         = body.get("field", "all")  # "all" | "title" | "tags" | "description"

    broadcast(f"[SEO] 🤖 Generate {field.upper()}: {p['folder']} — {p['keywords'][:60]}")
    result = await _run_seo(p, extra_hint, field=field)
    if result:
        return {"ok": True, "seo": result}
    return JSONResponse({"ok": False, "error": "Vertex AI không phản hồi"}, 500)

async def _run_seo(p: dict, extra_hint: str = "", field: str = "all"):
    folder = p["folder"]
    title  = p["title"] or folder
    asset_context = _build_seo_asset_context(p)

    def smart_trim(raw: str, max_len: int = 140) -> str:
        if len(raw) <= max_len:
            return raw
        parts = raw.split(" | ")
        result = parts[0]
        for part in parts[1:]:
            candidate = result + " | " + part
            if len(candidate) <= max_len:
                result = candidate
            else:
                break
        return result

    try:
        captured_shop_config = p.get("_shop_config")
        shop_config = captured_shop_config if isinstance(captured_shop_config, dict) else get_active_shop()
        shop_info = shop_config.get("shop_info", "")
        social_links = shop_config.get("social_links", "")
        etsy_link = shop_config.get("etsy_link", f"https://{shop_config.get('id', 'shop')}.etsy.com")

        shop_context_str = f"SHOP INFO: {shop_info}\n" if shop_info else ""

        # ── Single-field focused prompts ──────────────────────────────────────
        if field == "title":
            prompt = f"""You are an Etsy SEO expert. Generate ONLY an Etsy product title.
PRODUCT CONTEXT FROM LOCAL ASSETS:
{asset_context}
CURRENT TITLE FROM FORM: {title}
KEYWORD_LIST: {p.get('keywords', '')}
EXTRA INFO: {extra_hint or 'None'}
{shop_context_str}
Source-of-truth rules:
- Prefer PRODUCT CONTEXT FROM LOCAL ASSETS and Factory seed title over stale existing SEO text.
- If filenames or ZIP contents show SVG/PNG/DXF/EPS/AI design assets, write for a craft/design bundle, not a planner.
- Do NOT mention planner, GoodNotes, iPad, PDF planner, ADHD, daily, weekly, or monthly unless the asset context or user instructions explicitly support it.
Rules: Pipe-separated ( ' | ' ). First phrase = most important keyword. <=140 chars. 5-7 segments.
CRITICAL: Do NOT use the characters '&', '%' or ':' more than once in the title. If used, use them at most once (e.g., 'Goodnotes & iPad' is fine, but do not add another '&'). For additional/redundant occurrences, you MUST use 'and' instead of '&', 'percent' instead of '%', and a hyphen '-' instead of ':'.
CRITICAL: MAXIMIZE the 140 character limit. You MUST use between 135 and 140 characters to pack in as many relevant keywords as possible. Do not output short titles.
Return ONLY: <etsy_title>...</etsy_title>"""

        elif field == "tags":
            prompt = f"""You are an Etsy SEO expert. Generate ONLY Etsy tags.
PRODUCT CONTEXT FROM LOCAL ASSETS:
{asset_context}
CURRENT TITLE FROM FORM: {title}
KEYWORDS: {p.get('keywords', '')}
EXTRA INFO: {extra_hint or 'None'}
Source-of-truth rules:
- Prefer PRODUCT CONTEXT FROM LOCAL ASSETS and Factory seed title over stale existing SEO text.
- If filenames or ZIP contents show SVG/PNG/DXF/EPS/AI design assets, write tags for a craft/design bundle, not a planner.
- Do NOT mention planner, GoodNotes, iPad, PDF planner, ADHD, daily, weekly, or monthly unless the asset context or user instructions explicitly support it.
Rules: Exactly 13 tags, comma-separated. Each tag <=20 chars. No duplicates. Highly relevant only.
Return ONLY: <etsy_tags>...</etsy_tags>"""

        elif field == "description":
            prompt = f"""You are an Etsy SEO copywriter. Generate ONLY a product description.
PRODUCT CONTEXT FROM LOCAL ASSETS:
{asset_context}
CURRENT TITLE FROM FORM: {title}
KEYWORDS: {p.get('keywords', '')}
EXTRA INFO: {extra_hint or 'None'}
{shop_context_str}
Rules: Follow the full description template exactly — opening, Product Details, Key Features, Perfect For, Usage Ideas, What's Included, Compatible With, FAQ, INSTANT DOWNLOAD, T&C, Store link.
Base the product type on PRODUCT CONTEXT FROM LOCAL ASSETS first. Do NOT invent technical specs not mentioned in PRODUCT CONTEXT, KEYWORDS, or EXTRA INFO.
If filenames or ZIP contents show SVG/PNG/DXF/EPS/AI design assets, write for a craft/design bundle and mention compatible craft/design use such as Cricut/Silhouette only when supported by context.
Do NOT mention planner, GoodNotes, iPad, PDF planner, ADHD, daily, weekly, or monthly unless the asset context or user instructions explicitly support it.
Store link: {etsy_link}
{('Follow Us: ' + social_links) if social_links else ''}
Return ONLY: <description>...</description>"""

        else:
            # Full generation
            prompt = f"""You are a professional Etsy SEO copywriter creating a complete, high-converting product listing for a digital download product.

STRICT RULES:
	- ONLY write about the product type defined by the PRODUCT CONTEXT, KEYWORDS, and USER INSTRUCTIONS. Do NOT mix niches.
	- Base all content STRICTLY on PRODUCT CONTEXT FROM LOCAL ASSETS, KEYWORDS, and USER INSTRUCTIONS below.
	- PRODUCT CONTEXT FROM LOCAL ASSETS and Factory seed title are the source of truth. If they conflict with stale existing SEO title/description/tags, ignore the stale SEO.
	- If filenames or ZIP contents show SVG/PNG/DXF/EPS/AI design assets, the product is a craft/design bundle, not a planner.
	- Do NOT mention planner, GoodNotes, iPad, PDF planner, ADHD, daily, weekly, or monthly unless the asset context or user instructions explicitly support it.
	- If USER INSTRUCTIONS are provided, follow them exactly. Do NOT invent features that contradict them.
	- If USER INSTRUCTIONS are NOT provided, generate a realistic set of features typical for this product type.
	- For technical specs (file format, resolution, sizes, hyperlinks, etc.): ONLY mention what is explicitly stated in PRODUCT CONTEXT, KEYWORDS, or USER INSTRUCTIONS. Do NOT invent specs.
- Keep the SHOP INFO in mind as brand context but don't repeat it verbatim.
- Return ONLY valid XML with the exact tags shown below. No extra text outside XML tags.
- CRITICAL: Under <description>, the template contains instruction lines enclosed in square brackets (e.g. [GOOGLE-OPTIMIZED OPENING...], [Paragraph 1...], [Feature 1], [Audience 1], [Included item 1], etc.). You MUST completely replace these instructions with the actual generated text. Do NOT include the square brackets or the instruction text within them in your final output under any circumstances.

	PRODUCT CONTEXT FROM LOCAL ASSETS:
	{asset_context}

	CURRENT TITLE FROM FORM / EXCEL: {title}
	KEYWORDS TO TARGET: {p.get('keywords', 'Not provided')}
USER INSTRUCTIONS / EXTRA INFO: {extra_hint if extra_hint else 'Not provided'}
{shop_context_str}

Return ONLY this exact XML structure:

<etsy_title>
	CRITICAL: The very first phrase before the first ' | ' MUST be the most important, most-searched keyword for this product based on the asset context.
This opening keyword will be used by Google as the search snippet — make it count.
Pipe-separated. MUST use ' | ' between each phrase.
CRITICAL: Do NOT use the characters '&', '%' or ':' more than once in the title under any circumstances. If used, use them at most once (e.g., 'Goodnotes & iPad' is fine, but do not use '&' again). For additional/redundant occurrences, you MUST use the word 'and' instead of '&', the word 'percent' instead of '%', and a hyphen '-' instead of ':'.
CRITICAL: MAXIMIZE the 140 character limit. You MUST use between 135 and 140 characters to pack in as many relevant keywords as possible. Do not output short titles.
</etsy_title>

<etsy_tags>Comma-separated list of exactly 13 tags. Each tag <=20 chars. No duplicates. Use only highly relevant search terms.</etsy_tags>

<description>
[GOOGLE-OPTIMIZED OPENING — CRITICAL: Pack the top 3 most searched keywords into the first 160 characters. This is what Google shows as meta description. Start with the main product keyword, mention the format and 1-2 key features. End with ✨]

[Paragraph 1: 2-3 sentences describing the product's core value proposition. Who is it for? What problem does it solve? Mention the product keyword 2-3 times naturally.]

[Paragraph 2: 2-3 sentences expanding on why this specific product stands out. Mention the format, style, or unique angle. Include secondary keywords from the KEYWORDS list.]

Whether you're [use case 1], [use case 2], or [use case 3], this [product keyword] makes [benefit statement].

📄 Product Details

	[1-2 sentences describing the product structure, design style, formats, and scope shown by the asset context.]

	Key Features:

[Feature 1 — specific, keyword-rich, based on product type]
[Feature 2]
[Feature 3]
[Feature 4]
[Feature 5]
[Feature 6]
[Feature 7]
[Feature 8]

🎯 Perfect For

✔ [Audience 1 — specific persona or use case]
✔ [Audience 2]
✔ [Audience 3]
✔ [Audience 4]
✔ [Audience 5]
✔ [Audience 6]

💡 Usage Ideas

[Full sentence idea 1 — practical day-to-day use case]
[Full sentence idea 2]
[Full sentence idea 3]
	[Full sentence idea 4 — mention compatible apps/software only if relevant to the asset context, such as Cricut/Silhouette for SVG bundles or GoodNotes/PDF readers for planner PDFs]
[Full sentence idea 5]

📦 What's Included

	[IMPORTANT: Only list items explicitly mentioned in PRODUCT CONTEXT, KEYWORDS, or USER INSTRUCTIONS. If no specific items are mentioned, generate realistic items typical for this product type.]
✔ [Included item 1 — be specific with file name, size, or format]
✔ [Included item 2]
✔ [Included item 3]
✔ [Included item 4]
✔ [Included item 5]
✔ [Included item 6]

	Everything you need for one complete [product keyword].

🖥️ Compatible With

	[List only software/platforms supported by PRODUCT CONTEXT FROM LOCAL ASSETS. For SVG/PNG/DXF/EPS/AI bundles, use craft/design software such as Cricut Design Space, Silhouette Studio, Adobe Illustrator, or sublimation/print workflows when supported. For planner PDFs, use PDF readers or GoodNotes only when supported.]

❓ Frequently Asked Questions

	Q: [Generate a relevant compatibility question based on the product type]
	A: [Answer using only formats/software supported by the asset context]

Q: [Generate a relevant Q&A based on the product type — e.g. about printing, dating, editing, file format, refund policy, or app compatibility]
A: [Answer naturally and helpfully]

Q: [Another relevant Q&A specific to this product]
A: [Answer naturally and helpfully]

	Q: [Generate a relevant usage question based on the product type]
	A: [Answer using only formats and use cases supported by the asset context]

⚡ INSTANT DOWNLOAD

• Digital product only — no physical item shipped
• Files available immediately after purchase
• Download once and start using right away

💬 Questions or Issues?

We're here to help 💖
If you have any questions about this [product keyword], need help setting it up, or run into any issues, please feel free to message us. We're always happy to assist and ensure you love your product.

📜 TERMS & CONDITIONS

Permitted Usage:
• Personal use of the [product keyword]
• Use on your own devices
• Customize pages for your personal needs

Prohibited Usage:
• Do not resell, redistribute, or share the files
• Do not offer this product as a free download or giveaway

If you enjoy using this [product keyword], please consider leaving a review — it truly helps support our shop ✨

🛍️ FOR MORE DESIGNS LIKE THIS AND OTHER BEAUTIFUL BUNDLES PLEASE VISIT OUR STORE:

{etsy_link}
{(chr(10) + '📱 FOLLOW US:' + chr(10) + social_links + chr(10)) if social_links else ''}
</description>"""

        from google import genai
        client = genai.Client(vertexai=True, project="temply-ai-lab", location="us-central1")
        response = await asyncio.to_thread(
            client.models.generate_content,
            model="gemini-2.5-pro",
            contents=prompt,
            config=genai.types.GenerateContentConfig(
                temperature=0.3,
                max_output_tokens=8192,
            )
        )
        content = response.text or ""

        import re
        title_match = re.search(r'<etsy_title>(.*?)</etsy_title>', content, re.DOTALL | re.IGNORECASE)
        tags_match = re.search(r'<etsy_tags>(.*?)</etsy_tags>', content, re.DOTALL | re.IGNORECASE)
        desc_match = re.search(r'<description>(.*?)</description>', content, re.DOTALL | re.IGNORECASE)

        has_match = False
        if field == "all" and (title_match or tags_match or desc_match):
            has_match = True
        elif field == "title" and title_match:
            has_match = True
        elif field == "tags" and tags_match:
            has_match = True
        elif field == "description" and desc_match:
            has_match = True

        if not has_match:
            broadcast(f"[SEO] ❌ Lỗi Parse: Không tìm thấy thẻ XML phù hợp cho trường '{field}' trong phản hồi của AI.\nRaw (trích xuất): {content[:100]}...")
            return {}

        seo_title = smart_trim((title_match.group(1) if title_match else "").strip())
        raw_tags = (tags_match.group(1) if tags_match else "").strip()
        if "[" in raw_tags:
            raw_tags = re.sub(r'[\[\]"\'\n]', '', raw_tags)
        seo_tags = ", ".join([t.strip()[:20] for t in raw_tags.split(",") if t.strip()][:13])
        seo_desc = (desc_match.group(1) if desc_match else "").strip()
        if seo_desc:
            # Clean up leftover instruction blocks in square brackets from the AI generation
            lines = seo_desc.split("\n")
            cleaned_lines = []
            for line in lines:
                stripped = line.strip()
                if stripped.startswith("[") and stripped.endswith("]"):
                    lower_stripped = stripped.lower()
                    if any(word in lower_stripped for word in [
                        "critical", "paragraph", "sentences", "feature", "audience",
                        "idea", "important", "included item", "google-optimized", "write"
                    ]):
                        continue
                cleaned_lines.append(line)
            seo_desc = "\n".join(cleaned_lines).strip()

        # Save ONLY the requested field(s) to Excel — don't overwrite other fields
        excel_updates = {"keywords": p.get("keywords", ""), "extra": extra_hint}
        if field in ("all", "title") and seo_title:
            excel_updates["title"] = seo_title
        if field in ("all", "tags") and seo_tags:
            excel_updates["tags"] = seo_tags
        if field in ("all", "description") and seo_desc:
            excel_updates["description"] = seo_desc
        captured_excel_file = p.get("_excel_file")
        save_to_excel(
            p["row"],
            excel_updates,
            excel_path=Path(captured_excel_file) if captured_excel_file else None,
        )

        label = {"title": "Title", "tags": "Tags", "description": "Description"}.get(field, "SEO")
        display_val = seo_title or (seo_desc[:60] if seo_desc else "") or (seo_tags[:60] if seo_tags else "")
        broadcast(f"[SEO] ✅ {label} xong — {folder}: {display_val[:60]}...")
        if field == "all":
            broadcast(f"[SEO] 📊 {len(seo_title)} ký tự | {len(seo_tags.split(','))} tags")
        return {"title": seo_title, "tags": seo_tags, "description": seo_desc}
    except Exception as e:
        broadcast(f"[SEO] ❌ {e}")
    return {}


# ── API: Batch SEO for products with files but no title ─────────────────────────
@app.post("/api/batch-seo")
async def batch_seo():
    """Generate SEO for all products that have files but no title in Excel."""
    if "__BATCH_SEO__" in _running_processes:
        return JSONResponse({"ok": False, "error": "Batch SEO đang chạy rồi — chờ xong giúp em"}, status_code=409)
    wb = openpyxl.load_workbook(EXCEL_FILE(), data_only=True)
    ws = wb["Listings"]
    targets = []
    for row_num in range(4, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=c).value for c in range(1, 16)]
        folder = row[1]
        title  = row[7]
        if not folder: continue
        if title:       continue  # already has title, skip

        # Check if folder has any files
        folder_path = SHOP_DIR() / str(folder)
        files_dir   = folder_path / "files"
        images_dir  = folder_path / "images"
        has_files   = (files_dir.exists() and any(files_dir.iterdir())) or \
                      (images_dir.exists() and any(images_dir.iterdir()))
        if not has_files: continue

        # Guess product name from files
        name_guess = str(folder)
        if files_dir.exists():
            for f in files_dir.iterdir():
                stem = f.stem.replace("-", " ").replace("_", " ")
                if len(stem) > 8:
                    name_guess = stem; break

        targets.append({
            "row": row_num,
            "folder": str(folder),
            "title": name_guess,
            "description": "", "tags": "", "keywords": "",
            "price": float(str(row[4])) if row[4] else 4.99,
            "section": str(row[14] or ""),
        })

    if not targets:
        return {"ok": False, "message": "Không có folder nào cần generate SEO"}

    broadcast(f"[BATCH-SEO] 🚀 Bắt đầu generate SEO cho {len(targets)} sản phẩm...")
    _running_processes["__BATCH_SEO__"] = None
    asyncio.create_task(_run_batch_seo(targets))
    return {"ok": True, "count": len(targets), "folders": [t["folder"] for t in targets]}

async def _run_batch_seo(targets: list):
    try:
        for i, p in enumerate(targets, 1):
            broadcast(f"[BATCH-SEO] [{i}/{len(targets)}] Đang xử lý {p['folder']}...")
            await _run_seo(p, "")
            await asyncio.sleep(2)  # throttle to avoid Vertex overload
        broadcast(f"[BATCH-SEO] ✅ Hoàn thành {len(targets)} sản phẩm!")
    finally:
        _running_processes.pop("__BATCH_SEO__", None)


@app.get("/api/products/{row}/social-posts")
async def get_social_posts(row: int):
    p = get_product_by_row(row)
    if not p:
        raise HTTPException(404, "Không tìm thấy sản phẩm")

    title = p.get("title", "")
    desc = p.get("description", "")
    tags = p.get("tags", "")
    folder = p.get("folder", "")
    etsy_url = p.get("etsy_url", "")

    if not title or str(title).startswith("[Cần SEO]"):
        return {
            "ok": False,
            "error": "Sản phẩm này chưa được làm SEO. Hãy tạo SEO trước khi lấy Caption Social."
        }

    if not etsy_url or not str(etsy_url).strip():
        # Lấy etsy_link của shop đang active làm tiền tố
        shop_id = _active_shop_id
        shop_link = SHOPS.get(shop_id, {}).get("etsy_link", "")
        if shop_link:
            etsy_url = f"{shop_link.rstrip('/')}/listing/{folder}"
        else:
            etsy_url = f"https://www.etsy.com/shop/YourShop"

    # Xử lý caption Instagram
    sentences = [s.strip() for s in str(desc or "").replace("\n", " ").split(".") if s.strip()]
    hook = ". ".join(sentences[:2]) + "." if sentences else str(title)
    tag_list = [t.strip() for t in str(tags or "").split(",") if t.strip()]
    common_hashtags = "#digitaldownload #printable #instantdownload #etsyshop #etsyseller #digitalart"
    ig_hashtags = " ".join(f"#{t.replace(' ', '')}" for t in tag_list[:10])
    ig_hashtags += f" {common_hashtags}"
    ig_caption = f"{hook}\n\n✨ Get it instantly as a digital download!\n👇 Link in bio or search on Etsy: \"{str(title)[:40]}\"\n\n{ig_hashtags}"

    # Pinterest
    short_desc = ". ".join(sentences[:3]) + "." if sentences else str(desc)[:200]
    keywords = " | ".join(tag_list[:5])
    pinterest_desc = f"{title}\n\n{short_desc}\n\n{keywords} | Instant Digital Download | Printable PDF\n\n🛒 Shop now → {etsy_url}"

    # Facebook
    fb_body = " ".join(sentences[:4]) + "." if sentences else str(desc)[:300]
    fb_hashtags = " ".join(f"#{t.replace(' ', '')}" for t in tag_list[:6])
    facebook_post = f"🆕 New listing just dropped!\n\n📌 {title}\n\n{fb_body}\n\n✅ Instant digital download — print at home or at any print shop!\n🔗 Get it here: {etsy_url}\n\n{fb_hashtags} {common_hashtags}"

    # Twitter
    short_title = str(title)[:60] if len(str(title)) > 60 else str(title)
    twitter_post = f"🆕 {short_title} — instant digital download! ✨\n\n🛒 {etsy_url}\n\n#printable #digitaldownload #etsyshop"
    twitter_post = twitter_post[:280]

    # Medium research note / how-to article
    medium_intro = make_medium_research_article(title, desc, tags, etsy_url)

    return {
        "ok": True,
        "folder": folder,
        "title": title,
        "etsy_url": etsy_url,
        "social_statuses": get_product_social_statuses(
            BASE_DIR, _active_shop_id, str(folder), row
        ),
        "posts": {
            "instagram": ig_caption,
            "pinterest": pinterest_desc,
            "facebook": facebook_post,
            "twitter": twitter_post,
            "medium": medium_intro
        }
    }


def _social_session_payload(shop_id: str) -> dict:
    session = resolve_social_session(BASE_DIR, SHOPS, shop_id)
    return {
        "shop_id": shop_id,
        "shop_name": SHOPS.get(shop_id, {}).get("name", shop_id),
        "profile_dir": str(session.profile_dir),
        "debug_port": session.debug_port,
        "browser_ready": is_session_ready(session),
        "authentication": "unknown",
        "platforms": SOCIAL_URLS,
    }


@app.get("/api/social/session")
async def get_social_session():
    return {"ok": True, "session": _social_session_payload(_active_shop_id)}


@app.post("/api/social/session/open")
async def open_social_session(request: Request):
    shop_id = _active_shop_id
    body = await request.json()
    platform = str(body.get("platform", "")).lower()
    if platform not in SOCIAL_URLS:
        return JSONResponse(
            {"ok": False, "error": "Nền tảng social không được hỗ trợ"},
            status_code=400,
        )
    session = resolve_social_session(BASE_DIR, SHOPS, shop_id)
    try:
        ready = await asyncio.to_thread(open_social_browser, session, platform)
    except (FileNotFoundError, OSError, ValueError) as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)
    if not ready:
        return JSONResponse(
            {
                "ok": False,
                "error": "Chrome social không mở được tab trong đúng session của shop",
            },
            status_code=503,
        )
    payload = _social_session_payload(shop_id)
    payload["browser_ready"] = ready
    payload["opened_platform"] = platform
    return {
        "ok": True,
        "session": payload,
        "message": (
            f"Đã mở {platform.upper()} bằng session riêng của shop {shop_id}. "
            "Trạng thái đăng nhập chưa được xác nhận."
        ),
    }


@app.post("/api/products/{row}/post-social")
async def post_to_social(row: int, request: Request):
    shop_id = _active_shop_id
    p = get_product_by_row(row)
    if not p:
        raise HTTPException(404, "Không tìm thấy sản phẩm")

    body = await request.json()
    platform = body.get("platform")
    if platform not in SOCIAL_URLS:
        return JSONResponse({"ok": False, "error": "Thiếu nền tảng (platform)"}, status_code=400)

    folder = p["folder"]
    process_key = f"social_{shop_id}_{folder}_{platform}"
    if process_key in _running_processes:
        return JSONResponse({"ok": False, "error": f"Tiến trình đăng {platform.upper()} đang chạy rồi!"}, status_code=409)

    asyncio.create_task(_run_social_poster(shop_id, row, folder, platform))
    return {"ok": True, "message": f"Đang khởi động auto-post lên {platform.upper()}"}

async def _run_social_poster(shop_id: str, row: int, folder: str, platform: str):
    process_key = f"social_{shop_id}_{folder}_{platform}"
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}

    script_path = str(BASE_DIR / "social_auto_post.py")
    cmd = [PYTHON_BIN, "-u", script_path, "--row", str(row), "--platform", platform, "--shop", shop_id]

    broadcast(f"[SOCIAL] 🚀 [{shop_id}] Bắt đầu đăng lên {platform.upper()}: {folder}")
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
        _running_processes[process_key] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text:
                    broadcast(f"[SOCIAL] {text}")
        code = await proc.wait()
        broadcast(f"[SOCIAL] {'✅ Đăng xong' if code == 0 else f'❌ Lỗi đăng social exit {code}'}: {platform.upper()} - {folder}")
    except Exception as e:
        broadcast(f"[SOCIAL] ❌ {e}")
    finally:
        _running_processes.pop(process_key, None)


@app.post("/api/products/{row}/post-social-all")
async def post_to_all_social(row: int):
    """Post one product to ALL social platforms sequentially."""
    shop_id = _active_shop_id
    p = get_product_by_row(row)
    if not p:
        raise HTTPException(404, "Không tìm thấy sản phẩm")

    folder = p["folder"]
    process_key = f"social_all_{shop_id}_{folder}"
    if process_key in _running_processes:
        return JSONResponse(
            {"ok": False, "error": "Tiến trình đăng tất cả nền tảng đang chạy rồi!"},
            status_code=409,
        )

    # Also check if any individual platform poster is running for this product
    for platform in SOCIAL_URLS:
        individual_key = f"social_{shop_id}_{folder}_{platform}"
        if individual_key in _running_processes:
            return JSONResponse(
                {"ok": False, "error": f"Tiến trình đăng {platform.upper()} đang chạy cho sản phẩm này. Vui lòng đợi hoàn tất."},
                status_code=409,
            )

    asyncio.create_task(_run_social_poster_all(shop_id, row, folder))
    return {"ok": True, "message": f"Đang khởi động auto-post lên TẤT CẢ nền tảng cho {folder}"}


async def _run_social_poster_all(shop_id: str, row: int, folder: str):
    """Post one product to all social platforms sequentially."""
    process_key = f"social_all_{shop_id}_{folder}"
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}

    script_path = str(BASE_DIR / "social_auto_post.py")
    platforms = list(SOCIAL_URLS.keys())
    total = len(platforms)
    success_count = 0
    fail_count = 0

    broadcast(f"[SOCIAL] 🚀 [{shop_id}] BẮT ĐẦU ĐĂNG LÊN TẤT CẢ {total} NỀN TẢNG: {folder}")

    try:
        for i, platform in enumerate(platforms, 1):
            broadcast(f"[SOCIAL] 📢 [{i}/{total}] Đang đăng lên {platform.upper()}...")
            cmd = [PYTHON_BIN, "-u", script_path, "--row", str(row), "--platform", platform, "--shop", shop_id]

            try:
                proc = await asyncio.create_subprocess_exec(
                    *cmd, stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
                # Track under individual key too so stop-all can find it
                individual_key = f"social_{shop_id}_{folder}_{platform}"
                _running_processes[individual_key] = proc
                if proc.stdout:
                    async for line in proc.stdout:
                        text = line.decode("utf-8", errors="replace").rstrip()
                        if text:
                            broadcast(f"[SOCIAL] {text}")
                code = await proc.wait()
                if code == 0:
                    success_count += 1
                    broadcast(f"[SOCIAL] ✅ [{i}/{total}] Đăng xong {platform.upper()} - {folder}")
                else:
                    fail_count += 1
                    broadcast(f"[SOCIAL] ❌ [{i}/{total}] Lỗi đăng {platform.upper()} (exit {code}) - {folder}")
                _running_processes.pop(individual_key, None)
            except Exception as e:
                fail_count += 1
                broadcast(f"[SOCIAL] ❌ [{i}/{total}] Lỗi đăng {platform.upper()}: {e}")
                _running_processes.pop(f"social_{shop_id}_{folder}_{platform}", None)

        broadcast(
            f"[SOCIAL] {'🎉' if fail_count == 0 else '⚠️'} HOÀN TẤT đăng tất cả nền tảng cho {folder}: "
            f"✅ {success_count}/{total} thành công"
            + (f", ❌ {fail_count} lỗi" if fail_count else "")
        )
    except Exception as e:
        broadcast(f"[SOCIAL] ❌ Lỗi đăng tất cả nền tảng: {e}")
    finally:
        _running_processes.pop(process_key, None)


@app.post("/api/social/bulk-post")
async def bulk_post_social(request: Request):
    shop_id = _active_shop_id
    body = await request.json()
    platform = body.get("platform")
    start_row = body.get("start")
    end_row = body.get("end")
    delay = body.get("delay", 180)

    if platform not in {"instagram", "pinterest", "facebook", "twitter", "medium"} or not start_row or not end_row:
        return JSONResponse({"ok": False, "error": "Thiếu dữ liệu (platform, start, end)"}, status_code=400)

    process_key = f"social_bulk_{shop_id}_{platform}"
    if process_key in _running_processes:
        return JSONResponse({"ok": False, "error": "Tiến trình đăng hàng loạt đang chạy rồi!"}, status_code=409)

    asyncio.create_task(_run_social_bulk_poster(shop_id, start_row, end_row, platform, delay))
    return {"ok": True, "message": f"Đang khởi động auto-post hàng loạt lên {platform.upper()}"}


async def _run_social_bulk_poster(shop_id: str, start_row: int, end_row: int, platform: str, delay: int):
    process_key = f"social_bulk_{shop_id}_{platform}"
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}

    script_path = str(BASE_DIR / "social_bulk_post.py")
    cmd = [
        PYTHON_BIN, "-u", script_path,
        "--shop", shop_id,
        "--platform", platform,
        "--start", str(start_row),
        "--end", str(end_row),
        "--delay", str(delay)
    ]

    broadcast(f"[SOCIAL] 🚀 [{shop_id}] BẮT ĐẦU ĐĂNG HÀNG LOẠT LÊN {platform.upper()} (Dòng {start_row} ➔ {end_row})")
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
        _running_processes[process_key] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text:
                    broadcast(f"[SOCIAL] {text}")
        code = await proc.wait()
        broadcast(f"[SOCIAL] ✅ Đã hoàn thành đăng bài hàng loạt lên {platform.upper()} với Exit Code: {code}")
    except Exception as e:
        broadcast(f"[SOCIAL] ❌ Lỗi đăng hàng loạt: {e}")
    finally:
        _running_processes.pop(process_key, None)


# ── API: Gumroad Integration ──────────────────────────────────────────────────

@app.post("/api/gumroad/setup")
async def setup_gumroad(request: Request):
    process_key = "gumroad_setup"
    if process_key in _running_processes:
        return JSONResponse({"ok": False, "error": "Tiến trình thiết lập Gumroad đang chạy rồi!"}, status_code=409)

    asyncio.create_task(_run_gumroad_setup())
    return {"ok": True, "message": "Đang khởi động cấu hình Excel Gumroad..."}

async def _run_gumroad_setup():
    process_key = "gumroad_setup"
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}
    script_path = str(BASE_DIR / "gumroad_setup_excel.py")
    cmd = [PYTHON_BIN, "-u", script_path, "--shop", _active_shop_id]

    broadcast(f"[GUMROAD] ⚙️ Khởi động thiết lập sheet Gumroad cho shop: {_active_shop_id}")
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
        _running_processes[process_key] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text:
                    broadcast(f"[GUMROAD] {text}")
        code = await proc.wait()
        broadcast(f"[GUMROAD] {'✅ Thiết lập hoàn tất!' if code == 0 else f'❌ Gặp lỗi exit {code}'}")
    except Exception as e:
        broadcast(f"[GUMROAD] ❌ {e}")
    finally:
        _running_processes.pop(process_key, None)

@app.post("/api/gumroad/post")
async def post_gumroad(request: Request):
    body = await request.json()
    batch = body.get("batch", 5)
    skip = body.get("skip", 0)

    process_key = "gumroad_post"
    if process_key in _running_processes:
        return JSONResponse({"ok": False, "error": "Tiến trình đăng nháp Gumroad đang chạy rồi!"}, status_code=409)

    asyncio.create_task(_run_gumroad_poster(batch, skip))
    return {"ok": True, "message": f"Đang khởi động đăng {batch} sản phẩm lên Gumroad..."}

async def _run_gumroad_poster(batch: int, skip: int):
    process_key = "gumroad_post"
    env = {**__import__('os').environ, "PYTHONUNBUFFERED": "1"}
    script_path = str(BASE_DIR / "gumroad_auto_post.py")
    cmd = [
        PYTHON_BIN, "-u", script_path,
        "--shop", _active_shop_id,
        "--batch", str(batch),
        "--skip", str(skip)
    ]

    broadcast(f"[GUMROAD] 🚀 BẮT ĐẦU ĐĂNG LÊN GUMROAD (Batch: {batch}, Skip: {skip})")
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT, cwd=str(BASE_DIR), env=env)
        _running_processes[process_key] = proc
        if proc.stdout:
            async for line in proc.stdout:
                text = line.decode("utf-8", errors="replace").rstrip()
                if text:
                    broadcast(f"[GUMROAD] {text}")
        code = await proc.wait()
        broadcast(f"[GUMROAD] {'✅ Đã hoàn thành tiến trình đăng Gumroad!' if code == 0 else f'❌ Gặp lỗi exit {code}'}")
    except Exception as e:
        broadcast(f"[GUMROAD] ❌ Lỗi đăng Gumroad: {e}")
    finally:
        _running_processes.pop(process_key, None)


# ── API: Image Generator proxy ──────────────────────────────────────────────────

@app.get("/api/products/{row}/planner-info")
async def get_planner_info(row: int):
    """Return URL of the first planner image/PDF in the product's files/ dir."""
    p = get_product_by_row(row)
    file_dir = SHOP_DIR() / p["folder"] / "files"
    if not file_dir.exists():
        return JSONResponse({"planner_url": None, "name": None})
    for f in sorted(file_dir.iterdir()):
        if f.suffix.lower() in IMG_EXTS | {".pdf"}:
            return JSONResponse({
                "planner_url": f"/files/{p['folder']}/files/{urllib.parse.quote(f.name)}",
                "name": f.name,
            })
    return JSONResponse({"planner_url": None, "name": None})


@app.get("/api/products/{row}/pdf-page-count")
async def pdf_page_count(row: int):
    """Return number of pages in the product's PDF file."""
    p = get_product_by_row(row)
    file_dir = SHOP_DIR() / p["folder"] / "files"
    for f in sorted(file_dir.iterdir()) if file_dir.exists() else []:
        if f.suffix.lower() == ".pdf":
            try:
                import fitz
                doc = fitz.open(str(f))
                return {"pages": len(doc), "name": f.name}
            except Exception:
                return {"pages": None}
    return {"pages": None}


@app.get("/api/products/{row}/planner-png")
async def get_planner_png(row: int):
    """Return the product's planner as PNG (auto-converts PDF first page)."""
    p = get_product_by_row(row)
    file_dir = SHOP_DIR() / p["folder"] / "files"
    if not file_dir.exists():
        raise HTTPException(404, "No files/ folder")

    planner_path = None
    for f in sorted(file_dir.iterdir()):
        if f.suffix.lower() in IMG_EXTS | {".pdf"}:
            planner_path = f
            break
    if not planner_path:
        raise HTTPException(404, "No planner file found in files/")

    raw = planner_path.read_bytes()

    if planner_path.suffix.lower() == ".pdf":
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=raw, filetype="pdf")
            page = doc[0]
            mat = fitz.Matrix(150 / 72, 150 / 72)  # 150 DPI
            pix = page.get_pixmap(matrix=mat, alpha=False)
            png_bytes = pix.tobytes("png")
            return Response(
                content=png_bytes, media_type="image/png",
                headers={"Content-Disposition": "inline; filename=planner.png"}
            )
        except Exception as e:
            raise HTTPException(500, f"PDF convert failed: {e}")
    else:
        mime = "image/png" if planner_path.suffix.lower() == ".png" else "image/jpeg"
        return Response(content=raw, media_type=mime)


@app.post("/api/imagegen/pdf-to-png")
async def pdf_to_png(file: UploadFile = File(...)):
    """Convert first page of an uploaded PDF to PNG bytes."""
    content = await file.read()
    if content[:4] != b"%PDF":
        # Not a PDF — return as-is
        return Response(content=content, media_type="image/png")
    try:
        import fitz
        doc = fitz.open(stream=content, filetype="pdf")
        page = doc[0]
        mat = fitz.Matrix(150 / 72, 150 / 72)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        return Response(
            content=pix.tobytes("png"), media_type="image/png",
            headers={"Content-Disposition": "inline; filename=planner.png"}
        )
    except Exception as e:
        raise HTTPException(500, f"PDF convert: {e}")


@app.post("/api/products/{row}/convert-pdf")
async def convert_pdf_to_images(row: int):
    """Convert ALL pages of product's PDF → PNG, save into images/ folder."""
    p = get_product_by_row(row)
    file_dir = SHOP_DIR() / p["folder"] / "files"
    img_dir  = SHOP_DIR() / p["folder"] / "images"

    if not file_dir.exists():
        raise HTTPException(404, "No files/ folder found")

    # Find PDF (pick the first one)
    pdf_path = None
    for f in sorted(file_dir.iterdir()):
        if f.suffix.lower() == ".pdf":
            pdf_path = f
            break
    if not pdf_path:
        raise HTTPException(404, f"No PDF found in {p['folder']}/files/")

    img_dir.mkdir(parents=True, exist_ok=True)

    try:
        import fitz
        doc = fitz.open(str(pdf_path))
        saved = []
        broadcast(f"[PDF] 🔄 Converting {pdf_path.name} ({len(doc)} pages) → {p['folder']}/images/")
        for i in range(len(doc)):
            page = doc[i]
            mat = fitz.Matrix(2.0, 2.0)   # 144 DPI — high quality
            pix = page.get_pixmap(matrix=mat, alpha=False)
            out_name = f"image_{i + 1:02d}.png"
            out_path = img_dir / out_name
            pix.save(str(out_path))
            saved.append(out_name)
            broadcast(f"[PDF] ✅ Page {i + 1}/{len(doc)} → {out_name}")
        broadcast(f"[PDF] 🎉 Done — {len(saved)} ảnh lưu vào {p['folder']}/images/")
        return {"ok": True, "count": len(saved), "saved": saved, "folder": p["folder"]}
    except Exception as e:
        raise HTTPException(500, f"PDF convert failed: {e}")



@app.post("/api/imagegen/analyze")
async def imagegen_analyze(file: UploadFile = File(...)):
    """Proxy: upload + analyze planner → Vertex image studio."""
    content = await file.read()
    try:
        import httpx as _httpx
        async with _httpx.AsyncClient(timeout=120) as client:
            r = await client.post(
                f"{VERTEX_IMG_URL}/api/upload-and-analyze",
                files={"file": (file.filename, content, file.content_type or "image/png")},
            )
        return JSONResponse(r.json())
    except Exception as e:
        raise HTTPException(502, f"Vertex Image Studio không phản hồi: {e}")


@app.post("/api/imagegen/analyze-bundle")
async def imagegen_analyze_bundle(files: list[UploadFile] = File(...)):
    """Proxy: upload + analyze bundle → Vertex image studio."""
    form_files = []
    for f in files:
        content = await f.read()
        form_files.append(("files", (f.filename, content, f.content_type or "image/png")))
    try:
        import httpx as _httpx
        async with _httpx.AsyncClient(timeout=120) as client:
            r = await client.post(f"{VERTEX_IMG_URL}/api/upload-bundle", files=form_files)
        return JSONResponse(r.json())
    except Exception as e:
        raise HTTPException(502, f"Vertex Image Studio không phản hồi: {e}")


@app.post("/api/imagegen/generate")
async def imagegen_generate(request: Request):
    """Proxy SSE stream from Vertex image studio generate endpoint."""
    data = await request.json()

    async def stream():
        try:
            import httpx as _httpx
            async with _httpx.AsyncClient(timeout=700) as client:
                async with client.stream(
                    "POST", f"{VERTEX_IMG_URL}/api/generate",
                    json=data,
                    headers={"Accept": "text/event-stream"},
                ) as r:
                    async for line in r.aiter_lines():
                        yield f"{line}\n"
        except Exception as e:
            yield f"data: {json.dumps({'s': 'error', 'err': str(e)})}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


@app.get("/api/imagegen/output/{filename}")
async def imagegen_output(filename: str):
    """Serve generated images from Vertex output directory."""
    path = VERTEX_OUTPUT_DIR / filename
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(path)


@app.post("/api/products/{row}/import-gen-images")
async def import_gen_images(row: int):
    """Copy generated images from Vertex output/ → product images/ folder."""
    p = get_product_by_row(row)
    img_dir = SHOP_DIR() / p["folder"] / "images"
    img_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for f in sorted(VERTEX_OUTPUT_DIR.glob("image_*.png")):
        dest = img_dir / f.name
        shutil.copy2(f, dest)
        copied.append(f.name)
    broadcast(f"[IMG-GEN] 📥 Import {len(copied)} ảnh → {p['folder']}/images/")
    return {"ok": True, "copied": copied, "folder": p["folder"]}


STATIC_DIR.mkdir(exist_ok=True)
@app.post("/api/repair-listing")
async def api_repair_listing(req: Request):
    try:
        body = await req.json()
        listing_id = body.get("listing_id")
        if not listing_id:
            return {"ok": False, "error": "Thiếu listing_id"}

        args = ["--id", listing_id]
        if body.get("fix_tabs"): args.append("--tabs")
        if body.get("fix_desc"): args.append("--desc")
        if body.get("fix_tags"): args.append("--tags")

        script_path = BASE_DIR / "etsy_repair.py"
        cmd = [sys.executable, str(script_path)] + args

        process = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT
        )
        stdout, _ = await process.communicate()
        output = stdout.decode('utf-8')

        if process.returncode == 0:
            return {"ok": True, "output": output}
        else:
            return {"ok": False, "error": f"Script error:\n{output}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

from fastapi.responses import StreamingResponse

@app.get("/api/scan-listings")
async def api_scan_listings():
    async def event_stream():
        script_path = BASE_DIR / "etsy_scan_errors.py"
        process = await asyncio.create_subprocess_exec(
            sys.executable, str(script_path),
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT
        )
        if process.stdout:
            while True:
                try:
                    line = await process.stdout.readline()
                    if not line:
                        break
                except Exception:
                    break
                try:
                    text = line.decode('utf-8').strip()
                    if text:
                        yield f"data: {text}\n\n"
                except Exception:
                    pass
        await process.wait()
        yield "data: {\"type\": \"finished\"}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.get("/api/clean-duplicates")
async def api_clean_duplicates():
    async def event_stream():
        script_path = BASE_DIR / "etsy_clean_duplicates.py"
        process = await asyncio.create_subprocess_exec(
            sys.executable, str(script_path),
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT
        )
        if process.stdout:
            while True:
                try:
                    line = await process.stdout.readline()
                    if not line:
                        break
                except Exception:
                    break
                try:
                    text = line.decode('utf-8').strip()
                    if text:
                        broadcast(f"[🧹 DỌN TRÙNG] {text}")
                        yield f"data: {text}\n\n"
                except Exception:
                    pass
        await process.wait()
        yield "data: {\"type\": \"finished\"}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


# ── API: Import from Image Factory ──────────────────────────────────────────────
# Image Factory intake and catalog state always belong to the active shop.  This
# prevents a shop from scanning or importing another shop's source folders.

_FACTORY_SOURCE_EXCLUDE_DIRS = {
    "_deleted_products",
    "_asset_quarantine",
    "_failed_local_imports",
    "master_products",
}

_FACTORY_IMAGE_DIR_NAME = "etsy-images-chatgpt-final"
_FACTORY_ZIP_DIR_NAME = "source-zip"
_FACTORY_PRODUCT_RE = re.compile(r"^product-\d+$")

def _factory_shop_active() -> bool:
    return bool(_active_shop_id) and _active_shop_id in SHOPS

def _factory_shop_dir() -> Path:
    return SHOP_DIR()

def _factory_excel_file() -> Path:
    return _factory_shop_dir() / "Etsy_SEO_Generator.xlsx"

def _is_supported_factory_source(folder: Path) -> bool:
    """Return True only for a complete direct-child active-shop intake folder."""
    if folder.is_symlink() or not folder.is_dir():
        return False
    if folder.name.startswith(".") or folder.name.startswith("_"):
        return False
    if _FACTORY_PRODUCT_RE.fullmatch(folder.name) or folder.name in _FACTORY_SOURCE_EXCLUDE_DIRS:
        return False
    image_dir = folder / _FACTORY_IMAGE_DIR_NAME
    zip_dir = folder / _FACTORY_ZIP_DIR_NAME
    if image_dir.is_symlink() or zip_dir.is_symlink():
        return False
    try:
        folder_resolved = folder.resolve()
        return (
            image_dir.is_dir()
            and image_dir.resolve().parent == folder_resolved
            and zip_dir.is_dir()
            and zip_dir.resolve().parent == folder_resolved
        )
    except OSError:
        return False


def _factory_asset_is_usable(path: Path, allowed_suffixes: set[str]) -> bool:
    if path.is_symlink() or not path.is_file() or path.suffix.lower() not in allowed_suffixes:
        return False
    try:
        stat = path.stat()
    except OSError:
        return False
    return stat.st_size > 0 and getattr(stat, "st_blocks", 1) > 0

def _factory_asset_dir_is_contained(parent: Path) -> bool:
    if parent.is_symlink() or not parent.is_dir():
        return False
    try:
        return parent.resolve().parent == parent.parent.resolve()
    except OSError:
        return False

def _collect_image_nodes(parent: Path) -> list[Path]:
    if not _factory_asset_dir_is_contained(parent):
        return []
    return sorted([
        item for item in parent.iterdir()
        if _factory_asset_is_usable(item, IMG_EXTS)
    ], key=lambda p: (p.name != "01_hero_image.png", p.name.lower()))

def _collect_download_nodes(parent: Path) -> list[Path]:
    if not _factory_asset_dir_is_contained(parent):
        return []
    return sorted([
        item for item in parent.iterdir()
        if _factory_asset_is_usable(item, {".zip"})
    ], key=lambda p: p.name.lower())

def _normalize_factory_identity(value: object) -> str:
    text = str(value or "").lower().replace("-", " ").replace("_", " ").strip()
    return re.sub(r"\s+", " ", text)

def _factory_image_files(folder: Path) -> list[Path]:
    """Return usable preview images from the exact Image Factory image folder."""
    return _collect_image_nodes(folder / _FACTORY_IMAGE_DIR_NAME)

def _factory_download_files(folder: Path) -> list[Path]:
    return _collect_download_nodes(folder / _FACTORY_ZIP_DIR_NAME)

def _factory_source_keyword(folder: Path, files: Optional[list[Path]] = None) -> str:
    download_files = files if files is not None else _factory_download_files(folder)
    source_name = download_files[0].stem if download_files else folder.name
    return source_name.replace("-", " ").replace("_", " ").strip()

def _next_empty_catalog_row(ws, start_row: int = 4) -> int:
    """Return a row whose complete catalog payload B:R is empty."""
    for row_num in range(start_row, ws.max_row + 1):
        if all(not str(ws.cell(row=row_num, column=column).value or "").strip()
               for column in range(2, 19)):
            return row_num
    return max(start_row, ws.max_row + 1)

def _factory_shop_import_index(shop_dir: Path, excel_file: Path) -> tuple[dict, dict, dict]:
    """Index shop products by keyword and complete downloadable filename set."""
    keyword_targets: dict[str, tuple[str, int]] = {}
    folder_rows: dict[str, int] = {}
    if excel_file.exists():
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb["Listings"]
            for row_num in range(4, ws.max_row + 1):
                folder = str(ws.cell(row=row_num, column=2).value or "").strip()
                keyword = _normalize_factory_identity(ws.cell(row=row_num, column=3).value)
                if not _FACTORY_PRODUCT_RE.fullmatch(folder):
                    continue
                product_dir = shop_dir / folder
                try:
                    if (
                        product_dir.is_symlink()
                        or not product_dir.is_dir()
                        or product_dir.resolve().parent != shop_dir.resolve()
                    ):
                        continue
                except OSError:
                    continue
                folder_rows[folder] = row_num
                if keyword:
                    keyword_targets[keyword] = (folder, row_num)
            wb.close()
        except (OSError, KeyError, ValueError):
            pass

    file_targets: dict[frozenset[str], str] = {}
    if shop_dir.exists():
        try:
            for product_dir in shop_dir.iterdir():
                files_dir = product_dir / "files"
                # Only catalog-backed folders belong to the active shop's
                # imported state. Quarantine/orphan folders must not hide a
                # source product from the importer.
                if (
                    product_dir.name not in folder_rows
                    or not _FACTORY_PRODUCT_RE.fullmatch(product_dir.name)
                    or product_dir.is_symlink()
                    or not product_dir.is_dir()
                    or not files_dir.is_dir()
                    or files_dir.is_symlink()
                    or files_dir.resolve().parent != product_dir.resolve()
                ):
                    continue
                names = frozenset(
                    source_file.name.casefold()
                    for source_file in files_dir.iterdir()
                    if _factory_asset_is_usable(source_file, {".zip", ".pdf", ".001", ".002", ".003", ".004", ".005"})
                )
                if names:
                    file_targets.setdefault(names, product_dir.name)
        except OSError:
            pass
    return keyword_targets, file_targets, folder_rows

def _factory_import_match(
    folder: Path,
    keyword_targets: dict,
    file_targets: dict,
    folder_rows: dict,
) -> Optional[dict]:
    files = _factory_download_files(folder)
    source_names = frozenset(source_file.name.casefold() for source_file in files)
    target_folder = file_targets.get(source_names) if source_names else None
    if target_folder:
        return {
            "folder": target_folder,
            "row": folder_rows.get(target_folder),
            "matched_by": "download_file_set",
        }
    keyword = _normalize_factory_identity(_factory_source_keyword(folder, files))
    matched = keyword_targets.get(keyword)
    if matched:
        return {"folder": matched[0], "row": matched[1], "matched_by": "keyword"}
    return None

def _scan_factory_folders(source_dir: Path, shop_dir: Path, excel_file: Path) -> list[dict]:
    keyword_targets, file_targets, folder_rows = _factory_shop_import_index(shop_dir, excel_file)
    folders = []
    for item in sorted(source_dir.iterdir(), key=lambda p: p.name.lower()):
        try:
            if item.is_symlink() or item.resolve().parent != source_dir.resolve():
                continue
        except OSError:
            continue
        if not _is_supported_factory_source(item):
            continue

        images_in_folder = _factory_image_files(item)
        files_in_folder = _factory_download_files(item)
        # Incomplete intake folders are not importable products.
        if not images_in_folder or not files_in_folder:
            continue

        thumb_url = None
        if images_in_folder:
            # The thumbnail endpoint serves only a direct basename from the
            # exact image directory; do not expose the internal subdirectory
            # name in the URL.
            rel_thumb = images_in_folder[0].name
            thumb_url = (
                f"/api/image-factory/thumb/"
                f"{urllib.parse.quote(item.name)}/"
                f"{urllib.parse.quote(rel_thumb)}"
            )

        match = _factory_import_match(item, keyword_targets, file_targets, folder_rows)
        folders.append({
            "name": item.name,
            "keyword_guess": _factory_source_keyword(item, files_in_folder),
            "image_count": len(images_in_folder),
            "file_count": len(files_in_folder),
            "file_names": [f.stem for f in files_in_folder],
            "thumb": thumb_url,
            "already_imported": bool(match),
            "imported_folder": match.get("folder") if match else None,
            "imported_row": match.get("row") if match else None,
            "matched_by": match.get("matched_by") if match else None,
        })
    return folders


def _factory_scan_summary(source_dir: Path, folders: list[dict]) -> dict[str, int]:
    """Return aggregate-only scan context without exposing excluded paths.

    ``product-NN`` directories are the local catalog targets created by import,
    not Image Factory intake sources.  The UI needs their count to explain why
    they are absent, but must never receive their names or make them importable.
    """
    excluded_product_folder_count = 0
    try:
        source_root = source_dir.resolve()
        for item in source_dir.iterdir():
            if item.is_symlink() or not item.is_dir() or not _FACTORY_PRODUCT_RE.fullmatch(item.name):
                continue
            try:
                if item.resolve().parent == source_root:
                    excluded_product_folder_count += 1
            except OSError:
                continue
    except OSError:
        pass
    return {
        "eligible_intake_folders": len(folders),
        "excluded_catalog_product_folders": excluded_product_folder_count,
    }

def _resolve_factory_source_folder(folder_name: object) -> Optional[Path]:
    if not isinstance(folder_name, str) or not folder_name or Path(folder_name).name != folder_name:
        return None
    shop_dir = _factory_shop_dir()
    candidate = shop_dir / folder_name
    try:
        if candidate.is_symlink() or not candidate.is_dir():
            return None
        if candidate.resolve().parent != shop_dir.resolve():
            return None
        if not _is_supported_factory_source(candidate):
            return None
        if not _factory_image_files(candidate) or not _factory_download_files(candidate):
            return None
    except OSError:
        return None
    return candidate

@app.get("/api/image-factory/scan")
async def scan_image_factory():
    """
    Scan canonical Image Factory products for the active shop and include both
    imported and not-yet-imported products for filtering in the UI.
    """
    if not _factory_shop_active():
        return {
            "ok": False,
            "error": "Image Factory cần một active shop hợp lệ.",
            "factory_path": str(_factory_shop_dir()),
            "shop_id": _active_shop_id,
            "shop_name": _active_shop_id,
        }
    shop_dir = _factory_shop_dir()
    if not shop_dir.exists():
        return {"ok": False, "error": f"Image Factory output not found: {shop_dir}", "folders": []}

    try:
        folders = _scan_factory_folders(shop_dir, shop_dir, _factory_excel_file())
    except OSError as exc:
        return {"ok": False, "error": f"Could not scan Image Factory: {exc}", "folders": []}
    shop = get_active_shop()
    return {
        "ok": True,
        "folders": folders,
        "scan_summary": _factory_scan_summary(shop_dir, folders),
        "factory_path": str(shop_dir),
        "source_label": shop.get("name") or _active_shop_id,
        "shop_id": _active_shop_id,
        "shop_name": shop.get("name") or _active_shop_id,
    }


@app.get("/api/image-factory/thumb/{folder_name}/{filename:path}")
async def factory_thumb(folder_name: str, filename: str):
    """Serve only direct preview images from an active-shop Image Factory source."""
    if not _factory_shop_active():
        raise HTTPException(409, "Image Factory cần một active shop hợp lệ.")
    folder = _resolve_factory_source_folder(folder_name)
    if not folder:
        raise HTTPException(404)
    if not filename or Path(filename).name != filename:
        raise HTTPException(404)
    image_dir = folder / _FACTORY_IMAGE_DIR_NAME
    path = image_dir / filename
    try:
        if (
            not _factory_asset_is_usable(path, IMG_EXTS)
            or path.resolve().parent != image_dir.resolve()
        ):
            raise HTTPException(404)
    except OSError:
        raise HTTPException(404)
    return FileResponse(path)


@app.post("/api/image-factory/import")
async def import_from_factory(request: Request):
    """
    Import one or more folders from the Image Factory output into the dashboard.
    For each folder:
      1. Create product-XX folder with files/ and images/ subfolders
      2. Copy all files from factory output
      3. Write a new row in Excel with keyword from filename/folder name
      4. Trigger AI SEO generation (title, tags, description)
    """
    import shutil
    data = await request.json()
    requested_shop_id = str(data.get("shop_id") or "").strip()
    folder_names = data.get("folders", [])   # list of folder names to import
    auto_seo     = data.get("auto_seo", True)

    if not isinstance(folder_names, list) or not folder_names:
        return {"ok": False, "error": "No folders specified"}

    if requested_shop_id != _active_shop_id or not _factory_shop_active():
        return {
            "ok": False,
            "error": "Image Factory import cần shop_id khớp active shop.",
        }
    shop_dir = _factory_shop_dir()
    if not shop_dir.exists():
        return {"ok": False, "error": f"Image Factory output not found: {shop_dir}"}

    shop_id = _active_shop_id
    excel_file = _factory_excel_file()
    shop_config = dict(SHOPS.get(shop_id, {}))

    # Find next empty Excel row
    wb        = openpyxl.load_workbook(excel_file)
    ws        = wb["Listings"]
    next_row = _next_empty_catalog_row(ws)

    results = []
    seo_targets = []

    keyword_targets, file_targets, folder_rows = _factory_shop_import_index(shop_dir, excel_file)
    watermark = get_watermark_text(shop_id, shop_config)
    workbook_modified = False
    reusable_slots = _find_reusable_empty_product_slots(ws, shop_dir)
    used_folders = set()

    for factory_folder_name in dict.fromkeys(folder_names):
        src = _resolve_factory_source_folder(factory_folder_name)
        if not src:
            results.append({"folder": factory_folder_name, "ok": False, "error": "Source folder not found"})
            continue

        images = _factory_image_files(src)
        files = _factory_download_files(src)
        if not images or not files:
            results.append({"folder": factory_folder_name, "ok": False, "error": "Source folder is incomplete"})
            continue

        existing = _factory_import_match(src, keyword_targets, file_targets, folder_rows)
        if existing:
            results.append({
                "folder": factory_folder_name,
                "ok": False,
                "already_imported": True,
                "imported_folder": existing["folder"],
                "error": f"Already imported as {existing['folder']}",
            })
            continue

        slot = _allocate_product_slot(ws, shop_dir, next_row, reusable_slots, used_folders)
        folder_name = slot["folder"]
        prod_path   = slot["path"]
        img_dst     = prod_path / "images"
        file_dst    = prod_path / "files"

        try:
            img_dst.mkdir(parents=True, exist_ok=slot["reused"])
            file_dst.mkdir(parents=True, exist_ok=slot["reused"])

            copied_imgs = 0
            for f in images:
                copy_image_with_watermark(f, img_dst / f.name, watermark)
                copied_imgs += 1

            copied_files = 0
            file_stems = []
            for f in files:
                shutil.copy2(f, file_dst / f.name)
                copied_files += 1
                file_stems.append(f.stem)
        except (OSError, ValueError) as exc:
            if slot["reused"]:
                _clear_product_asset_folder(prod_path)
            else:
                # This destination was allocated by this request only. Remove a
                # partial copy so it cannot masquerade as an imported product.
                shutil.rmtree(prod_path, ignore_errors=True)
            results.append({
                "folder": factory_folder_name,
                "ok": False,
                "error": f"Could not copy product assets: {exc}",
            })
            continue

        # Build keyword from file name (prefer file stem), else from folder name
        if file_stems:
            # Use first file stem, clean up dashes/underscores
            keyword = file_stems[0].replace("-", " ").replace("_", " ").strip()
        else:
            keyword = factory_folder_name.replace("-", " ").replace("_", " ").strip()

        # Write Excel row
        target_row = slot["row"]
        _clear_catalog_row(ws, target_row)
        set_cell_value(ws, target_row, 2, folder_name)         # B: folder
        set_cell_value(ws, target_row, 3, keyword)             # C: keywords
        set_cell_value(ws, target_row, 4, keyword)             # D: visible factory seed title
        set_cell_value(ws, target_row, 5, 4.99)                # E: price
        set_cell_value(ws, target_row, 11, 999)                 # K: qty
        set_cell_value(ws, target_row, 13, "2020_2026")         # M: when_made
        set_cell_value(ws, target_row, 14, "⏳ Chờ đăng")       # N: status
        set_cell_value(ws, target_row, 15, "Digital Planner")   # O: section

        broadcast(f"[FACTORY] 📁 Imported {factory_folder_name} → {folder_name} ({copied_imgs} ảnh, {copied_files} file) | Keyword: {keyword}")

        seo_targets.append({
            "row":      target_row,
            "folder":   folder_name,
            "title":    keyword,
            "keywords": keyword,
            "description": "", "tags": "",
            "price":    4.99,
            "section":  "Digital Planner",
            "_shop_dir": shop_dir,
            "_excel_file": excel_file,
            "_shop_config": shop_config,
        })

        results.append({
            "ok":            True,
            "source":        factory_folder_name,
            "folder":        folder_name,
            "row":           target_row,
            "keyword":       keyword,
            "images_copied": copied_imgs,
            "files_copied":  copied_files,
        })

        keyword_targets[_normalize_factory_identity(keyword)] = (folder_name, target_row)
        folder_rows[folder_name] = target_row
        source_names = frozenset(source_file.name.casefold() for source_file in files)
        if source_names:
            file_targets[source_names] = folder_name
        workbook_modified = True

        next_row = _next_empty_catalog_row(ws, target_row + 1)

    if workbook_modified:
        wb.save(excel_file)
    wb.close()

    # Auto-generate SEO in background
    if auto_seo and seo_targets:
        broadcast(f"[FACTORY] 🤖 Bắt đầu gen SEO cho {len(seo_targets)} sản phẩm vừa import...")
        asyncio.create_task(_run_batch_seo(seo_targets))

    return {
        "ok":      True,
        "imported": len([r for r in results if r.get("ok")]),
        "skipped": len([r for r in results if r.get("already_imported")]),
        "failed": len([r for r in results if not r.get("ok") and not r.get("already_imported")]),
        "shop_id": shop_id,
        "results":  results,
        "auto_seo": auto_seo,
    }


app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

@app.get("/")
async def root():
    return Response(_dashboard_index_html(), media_type="text/html")

if __name__ == "__main__":
    startup_guard = runtime_identity.startup_guard_message(BASE_DIR)
    if startup_guard is not None:
        raise SystemExit(startup_guard)

    import uvicorn
    listen_host = _dashboard_listen_host()
    print(f"🚀 Etsy Dashboard → http://{listen_host}:8090")
    uvicorn.run("dashboard_app:app", host=listen_host, port=8090, reload=False)
