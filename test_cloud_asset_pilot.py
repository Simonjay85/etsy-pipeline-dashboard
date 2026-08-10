from __future__ import annotations

import datetime as dt
import tempfile
import unittest
from pathlib import Path

import openpyxl

from cloud_asset_pilot import build_parser, execute_pilot, plan_pilot


UTC = dt.timezone.utc
NOW = dt.datetime(2026, 8, 4, 12, 0, tzinfo=UTC)


def make_repo(tmp: str, count: int = 5, *, explicit_master: bool = True, age_hours: int = 1) -> Path:
    root = Path(tmp) / "repo"
    (root / "shops" / "templystudios").mkdir(parents=True)
    (root / "master_products").mkdir()
    (root / "active_shop.txt").write_text("templystudios\n", encoding="utf-8")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Listings"
    mapping = {}
    active = []
    for index in range(1, count + 1):
        folder = f"product-{index:02d}"
        listing_id = str(7000000000 + index)
        sheet.cell(row=index + 3, column=2, value=folder)
        sheet.cell(row=index + 3, column=8, value=f"Product {index}")
        sheet.cell(row=index + 3, column=16, value=f"https://www.etsy.com/listing/{listing_id}")
        shop = root / "shops" / "templystudios" / folder
        master = root / "master_products" / folder
        for product in (shop, master):
            (product / "images").mkdir(parents=True)
            (product / "files").mkdir()
            (product / "images" / "hero.png").write_bytes(b"image" * index)
            (product / "files" / "source.zip").write_bytes(b"file" * index)
        entry = {"templystudios": folder}
        if explicit_master:
            entry["master_product"] = folder
        mapping[f"source/category-{index}"] = entry
        active.append({"id": listing_id, "title": f"Product {index}"})
    workbook.save(root / "shops" / "templystudios" / "Etsy_SEO_Generator.xlsx")
    snapshot = {
        "shopId": "templystudios",
        "crawledAt": (NOW - dt.timedelta(hours=age_hours)).isoformat().replace("+00:00", "Z"),
        "active": active,
        "draft": [],
    }
    (root / "snapshot.json").write_text(__import__("json").dumps(snapshot), encoding="utf-8")
    (root / "product_source_map.json").write_text(__import__("json").dumps(mapping), encoding="utf-8")
    return root


class FakeStore:
    def __init__(self):
        self.calls = []

    def upload(self, path):
        self.calls.append(("upload", path))
        return {"ok": True, "revision": "rev-1"}

    def verify(self, path):
        self.calls.append(("verify", path))
        return {"ok": True, "state": "OFFLOAD_SCHEDULED"}


class CloudAssetPilotTests(unittest.TestCase):
    def test_stale_snapshot_blocks_without_selection(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = make_repo(temporary, age_hours=25)
            result = plan_pilot(root, snapshot_path=root / "snapshot.json", now=NOW)
        self.assertEqual("BLOCKED_CATALOG_MAPPING", result["state"])
        self.assertEqual([], result["selected"])
        self.assertIn("snapshot_stale", result["reason"])
        self.assertFalse(result["write_performed"])

    def test_missing_master_mapping_blocks_candidate(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = make_repo(temporary, count=1, explicit_master=False)
            result = plan_pilot(root, snapshot_path=root / "snapshot.json", now=NOW)
        self.assertEqual([], result["eligible"])
        self.assertIn("missing_explicit_master_mapping", result["candidates"][0]["reasons"])

    def test_shop_only_does_not_require_source_or_master_mapping(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = make_repo(temporary, count=5, explicit_master=False)
            (root / "product_source_map.json").write_text("{not valid json", encoding="utf-8")
            result = plan_pilot(
                root,
                snapshot_path=root / "snapshot.json",
                now=NOW,
                mode="shop-only",
            )
        self.assertTrue(result["ok"])
        self.assertEqual("shop-only", result["mode"])
        self.assertEqual(5, len(result["selected"]))
        self.assertTrue(all("master_product" not in item for item in result["selected"]))

    def test_unknown_mode_is_rejected(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = make_repo(temporary, count=1)
            with self.assertRaises(ValueError):
                plan_pilot(root, snapshot_path=root / "snapshot.json", now=NOW, mode="invalid")

    def test_cli_exposes_independent_modes_and_keeps_legacy_default(self):
        parser = build_parser()
        self.assertEqual("shop-and-master", parser.parse_args([]).mode)
        self.assertEqual("shop-only", parser.parse_args(["--mode", "shop-only"]).mode)

    def test_fewer_than_five_never_substitutes_ineligible_candidates(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = make_repo(temporary, count=4)
            result = plan_pilot(root, snapshot_path=root / "snapshot.json", now=NOW)
        self.assertEqual(4, len(result["eligible"]))
        self.assertEqual([], result["selected"])
        self.assertEqual("BLOCKED_CATALOG_MAPPING", result["state"])

    def test_selects_exactly_five_by_combined_bytes_then_numeric_folder(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = make_repo(temporary, count=6)
            result = plan_pilot(root, snapshot_path=root / "snapshot.json", now=NOW)
        self.assertTrue(result["ok"])
        self.assertEqual("PILOT_ELIGIBLE", result["state"])
        self.assertEqual(5, len(result["selected"]))
        self.assertEqual(
            ["product-01", "product-02", "product-03", "product-04", "product-05"],
            [item["folder"] for item in result["selected"]],
        )

    def test_execute_path_only_calls_upload_and_verify_and_never_delete(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = make_repo(temporary, count=5)
            plan = plan_pilot(root, snapshot_path=root / "snapshot.json", now=NOW)
            store = FakeStore()
            result = execute_pilot(plan, store)
        self.assertTrue(result["ok"])
        self.assertEqual("CLOUD_PIPELINE_VERIFIED", result["state"])
        self.assertFalse(result["delete_performed"])
        self.assertEqual(20, len(store.calls))

    def test_shop_only_execute_never_calls_master_scope(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = make_repo(temporary, count=5, explicit_master=False)
            plan = plan_pilot(
                root,
                snapshot_path=root / "snapshot.json",
                now=NOW,
                mode="shop-only",
            )
            store = FakeStore()
            result = execute_pilot(plan, store)
        self.assertTrue(result["ok"])
        self.assertEqual("shop-only", result["mode"])
        self.assertEqual(10, len(store.calls))
        self.assertTrue(all(path.startswith("shops/") for _, path in store.calls))


if __name__ == "__main__":
    unittest.main()
