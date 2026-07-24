#!/usr/bin/env python3
from __future__ import annotations

import asyncio
import json
import openpyxl
import shutil
import sys
import tempfile
import types
from pathlib import Path
from typing import Any
from unittest import TestCase
from unittest.mock import AsyncMock, patch

from fastapi import HTTPException

if "deep_translator" not in sys.modules:
    fake_google_translator = types.ModuleType("deep_translator")

    class _FakeGoogleTranslator:
        def __init__(self, source: str, target: str):
            self.source = source
            self.target = target

        def translate(self, text: str):
            return str(text)

    fake_google_translator.GoogleTranslator = _FakeGoogleTranslator
    sys.modules["deep_translator"] = fake_google_translator

if "google" not in sys.modules:
    fake_google = types.ModuleType("google")
    fake_google_genai = types.ModuleType("google.genai")
    fake_google_types = types.ModuleType("google.genai.types")
    fake_google_genai.types = fake_google_types
    fake_google.genai = fake_google_genai
    sys.modules["google"] = fake_google
    sys.modules["google.genai"] = fake_google_genai
    sys.modules["google.genai.types"] = fake_google_types

import dashboard_app
import etsy_auto_post


class Request:
    def __init__(self, payload: Any):
        self.payload = payload

    async def json(self):
        return self.payload


def _response_json(response):
    if isinstance(response, dict):
        return response
    if isinstance(response, tuple):
        _, body = response
        if isinstance(body, (bytes, bytearray)):
            return json.loads(body.decode())
        return json.loads(body) if isinstance(body, str) else body
    if hasattr(response, "status_code") and hasattr(response, "body"):
        return json.loads(response.body.decode())
    raise TypeError(f"Unsupported response type: {type(response)!r}")


def _create_listing_sheet(excel_path: Path, row_data: dict[int, dict[int, object]]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings"
    for row_num, values in row_data.items():
        for column in range(2, 19):
            ws.cell(row=row_num, column=column, value=values.get(column, ""))
    wb.save(excel_path)


def _cell_value(excel_path: Path, row: int, column: int):
    wb = openpyxl.load_workbook(excel_path)
    return wb["Listings"].cell(row=row, column=column).value


class TestParseSelectedPayload(TestCase):
    def setUp(self):
        self.shop_id = "daisyflowdigital"

    def test_parse_run_selected_payload_validates_payload(self):
        with patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            with self.assertRaises(HTTPException):
                dashboard_app._parse_run_selected_request_payload({"shop": self.shop_id, "items": {}})

            with self.assertRaises(HTTPException):
                dashboard_app._parse_run_selected_request_payload(
                    {"shop": self.shop_id, "items": [{"row": True, "folder": "product-1"}]}
                )

            with self.assertRaises(HTTPException):
                dashboard_app._parse_run_selected_request_payload(
                    {"shop": self.shop_id, "items": [{"row": 3, "folder": "product-1"}]}
                )

    def test_parse_run_selected_payload_rejects_duplicates(self):
        with patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            with self.assertRaises(HTTPException):
                dashboard_app._parse_run_selected_request_payload(
                    {
                        "shop": self.shop_id,
                        "items": [
                            {"row": 4, "folder": "product-1"},
                            {"row": 4, "folder": "product-2"},
                        ],
                    }
                )

            with self.assertRaises(HTTPException):
                dashboard_app._parse_run_selected_request_payload(
                    {
                        "shop": self.shop_id,
                        "items": [
                            {"row": 4, "folder": "product-1"},
                            {"row": 5, "folder": "product-1"},
                        ],
                    }
                )


class TestSelectedRouteValidation(TestCase):
    def setUp(self):
        self.shop_id = "daisyflowdigital"
        self.tmp = Path(tempfile.mkdtemp())
        self.shop_dir = self.tmp / "shops" / self.shop_id
        self.shop_dir.mkdir(parents=True, exist_ok=True)
        self.excel_path = self.shop_dir / "Etsy_SEO_Generator.xlsx"
        _create_listing_sheet(
            self.excel_path,
            {
                4: {2: "product-369", 5: 9.99, 8: "Sample title", 14: "⏳ Chờ đăng"},
                5: {2: "product-370", 5: 9.99, 8: "Another title", 14: "⏳ Chờ đăng"},
            },
        )
        for folder in ("product-369", "product-370"):
            (self.shop_dir / folder / "images").mkdir(parents=True, exist_ok=True)
            (self.shop_dir / folder / "files").mkdir(parents=True, exist_ok=True)

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)
        dashboard_app._running_processes.clear()
        dashboard_app._running_tasks.clear()

    def _run_selected(self, payload):
        return asyncio.run(dashboard_app.run_selected_products(Request(payload)))

    def test_route_rejects_missing_local_data_before_any_update(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb["Listings"]
            ws.cell(row=4, column=16, value="https://www.etsy.com/listing/9999999999")
            wb.save(self.excel_path)

            with self.assertRaises(HTTPException) as context:
                self._run_selected({"shop": self.shop_id, "items": [{"row": 4, "folder": "product-369"}]})
            self.assertEqual(400, context.exception.status_code)
            self.assertIn("Không có sản phẩm hợp lệ", str(context.exception.detail))
            # Invalid-only batch still marks the card with an error reason.
            self.assertIn("Etsy URL/listing ID", str(_cell_value(self.excel_path, 4, 14)))
            self.assertEqual("Sample title", _cell_value(self.excel_path, 4, 8))
            self.assertEqual("https://www.etsy.com/listing/9999999999", _cell_value(self.excel_path, 4, 16))

            wb2 = openpyxl.load_workbook(self.excel_path)
            ws = wb2["Listings"]
            ws.cell(row=5, column=16, value="1234567890")
            ws.cell(row=5, column=8, value="")
            wb2.save(self.excel_path)
            wb2.close()

            with self.assertRaises(HTTPException):
                self._run_selected({"shop": self.shop_id, "items": [{"row": 5, "folder": "product-370"}]})
            self.assertIsNone(_cell_value(self.excel_path, 5, 8))
            self.assertIn("thiếu title", str(_cell_value(self.excel_path, 5, 14)))

            self.assertEqual("product-369", _cell_value(self.excel_path, 4, 2))
            self.assertEqual("product-370", _cell_value(self.excel_path, 5, 2))

    def test_route_skips_invalid_items_and_queues_valid_rest(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id), \
                patch.object(dashboard_app, "_run_selected_poster", new=AsyncMock(return_value=None)) as run_poster:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb["Listings"]
            ws.cell(row=4, column=16, value="https://www.etsy.com/listing/9999999999")
            wb.save(self.excel_path)

            response = self._run_selected(
                {
                    "shop": self.shop_id,
                    "items": [
                        {"row": 4, "folder": "product-369"},
                        {"row": 5, "folder": "product-370"},
                    ],
                }
            )
            data = _response_json(response)
            self.assertEqual(202, response.status_code)
            self.assertEqual(True, data["ok"])
            self.assertEqual(1, data["queued"])
            self.assertEqual(1, data["skipped"])
            self.assertEqual(["product-370"], data["folders"])
            self.assertEqual(1, len(data["rejected"]))
            self.assertEqual("product-369", data["rejected"][0]["folder"])
            self.assertIn("Etsy URL/listing ID", data["rejected"][0]["reason"])
            self.assertIn("Etsy URL/listing ID", str(_cell_value(self.excel_path, 4, 14)))
            self.assertEqual("⏳ Chờ đăng", _cell_value(self.excel_path, 5, 14))
            run_poster.assert_called_once()
            queued_items = run_poster.call_args.args[1]
            self.assertEqual([(5, "product-370")], queued_items)

    def test_route_rejects_folder_mismatch_and_missing_folder(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            with self.assertRaises(HTTPException) as context:
                self._run_selected({"shop": self.shop_id, "items": [{"row": 4, "folder": "product-370"}]})
            self.assertEqual(400, context.exception.status_code)
            self.assertIn("Không có sản phẩm hợp lệ", str(context.exception.detail))

            shutil.rmtree(self.shop_dir / "product-370", ignore_errors=True)
            with self.assertRaises(HTTPException):
                self._run_selected({"shop": self.shop_id, "items": [{"row": 5, "folder": "product-370"}]})

    def test_route_rejects_posted_status_and_listing_editor_url(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb["Listings"]
            ws.cell(row=4, column=14, value="✅ Đã đăng draft")
            ws.cell(
                row=5,
                column=16,
                value="https://www.etsy.com/your/shops/me/listing-editor/edit/1234567890",
            )
            wb.save(self.excel_path)

            with self.assertRaises(HTTPException):
                self._run_selected({"shop": self.shop_id, "items": [{"row": 4, "folder": "product-369"}]})
            with self.assertRaises(HTTPException):
                self._run_selected({"shop": self.shop_id, "items": [{"row": 5, "folder": "product-370"}]})
            self.assertIn("Đã đăng", str(_cell_value(self.excel_path, 4, 14)))
            self.assertIn("Etsy URL/listing ID", str(_cell_value(self.excel_path, 5, 14)))

    def test_route_sets_rows_pending_in_single_save_and_returns_queue(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id), \
                patch.object(dashboard_app, "_run_selected_poster", new=AsyncMock(return_value=None)) as run_poster:
            response = self._run_selected(
                {
                    "shop": self.shop_id,
                    "items": [
                        {"row": 4, "folder": "product-369"},
                        {"row": 5, "folder": "product-370"},
                    ],
                }
            )
            data = _response_json(response)
            self.assertEqual(202, response.status_code)
            self.assertEqual(True, data["ok"])
            self.assertEqual(2, data["queued"])
            self.assertEqual(["product-369", "product-370"], data["folders"])
            self.assertEqual("⏳ Chờ đăng", _cell_value(self.excel_path, 4, 14))
            self.assertEqual("⏳ Chờ đăng", _cell_value(self.excel_path, 5, 14))
            run_poster.assert_called_once()

    def test_route_rolls_back_when_pending_update_fails(self):
        with patch.object(openpyxl.workbook.workbook.Workbook, "save", side_effect=OSError("simulate save fail")), \
                patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            with self.assertRaises(HTTPException) as context:
                self._run_selected({"shop": self.shop_id, "items": [{"row": 4, "folder": "product-369"}]})
            self.assertEqual(500, context.exception.status_code)
            self.assertEqual("Sample title", _cell_value(self.excel_path, 4, 8))


class TestSelectedRouteCollision(TestCase):
    def setUp(self):
        self.shop_id = "daisyflowdigital"
        self.tmp = Path(tempfile.mkdtemp())
        self.shop_dir = self.tmp / "shops" / self.shop_id
        self.shop_dir.mkdir(parents=True, exist_ok=True)
        self.excel_path = self.shop_dir / "Etsy_SEO_Generator.xlsx"
        _create_listing_sheet(
            self.excel_path,
            {4: {2: "product-369", 5: 9.99, 8: "Sample title", 14: "⏳ Chờ đăng"}},
        )
        (self.shop_dir / "product-369" / "images").mkdir(parents=True, exist_ok=True)
        (self.shop_dir / "product-369" / "files").mkdir(parents=True, exist_ok=True)

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)
        dashboard_app._running_processes.clear()
        dashboard_app._running_tasks.clear()

    def test_run_selected_blocks_when_shop_lock_exists(self):
        lock_key = dashboard_app._etsy_post_lock_key(self.shop_id)
        dashboard_app._running_processes[lock_key] = object()
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id):
            with self.assertRaises(HTTPException):
                asyncio.run(
                    dashboard_app.run_selected_products(Request({
                        "shop": self.shop_id,
                        "items": [{"row": 4, "folder": "product-369"}],
                    }))
                )

    def test_single_post_blocks_runall_and_selected(self):
        with patch.object(dashboard_app, "SHOP_DIR", return_value=self.shop_dir), \
                patch.object(dashboard_app, "EXCEL_FILE", return_value=self.excel_path), \
                patch.object(dashboard_app, "_active_shop_id", self.shop_id), \
                patch.object(dashboard_app, "get_product_by_row", return_value={"folder": "product-369"}), \
                patch.object(dashboard_app, "save_to_excel", return_value=None), \
                patch.object(dashboard_app, "_run_poster", new=AsyncMock(return_value=None)):
            single_resp = asyncio.run(dashboard_app.post_to_etsy(4))
            self.assertTrue(single_resp["ok"])

            run_all = asyncio.run(dashboard_app.run_all_pending())
            self.assertEqual(409, run_all.status_code)

            with self.assertRaises(HTTPException):
                asyncio.run(
                    dashboard_app.run_selected_products(Request({
                        "shop": self.shop_id,
                        "items": [{"row": 4, "folder": "product-369"}],
                    }))
                )


class TestSelectedPosterCommand(TestCase):
    def test_run_selected_poster_launches_single_subprocess_with_products_mode(self):
        async def fake_create_subprocess(*args, **kwargs):
            proc = AsyncMock()
            proc.stdout = None
            proc.wait = AsyncMock(return_value=0)
            return proc

        with patch.object(
            dashboard_app,
            "_runtime_prefetch_import_check",
            new=AsyncMock(return_value=(True, "ok")),
        ), patch.object(
            dashboard_app.asyncio,
            "create_subprocess_exec",
            new=AsyncMock(side_effect=fake_create_subprocess),
        ):
            args = [
                dashboard_app.PYTHON_BIN,
                "-u",
                dashboard_app.ETSY_POSTER,
                "--shop",
                "daisyflowdigital",
                "--selected-product",
                "4:product-02",
                "--selected-product",
                "5:product-01",
            ]
            asyncio.run(
                dashboard_app._run_selected_poster(
                    "daisyflowdigital",
                    [(4, "product-02"), (5, "product-01")],
                    dashboard_app._etsy_post_lock_key("daisyflowdigital"),
                )
            )
            self.assertEqual(1, dashboard_app.asyncio.create_subprocess_exec.await_count)
            actual_args = dashboard_app.asyncio.create_subprocess_exec.await_args.args
            self.assertEqual(args, list(actual_args[: len(args)]))
            self.assertIn("cwd", dashboard_app.asyncio.create_subprocess_exec.await_args.kwargs)


class TestEtsyAutoPostCLISelection(TestCase):
    def _mk_shop_and_excel(self, root: Path):
        shop_dir = root / "shops" / "daisyflowdigital"
        shop_dir.mkdir(parents=True)
        excel_path = shop_dir / "Etsy_SEO_Generator.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listings"
        rows = [
            (4, "product-01", "Title 1", 9.99, "", "Category", "When made", "", "", "2020_2026"),
            (5, "product-02", "Title 2", 9.99, "", "Category", "When made", "", "", "2020_2026"),
            (6, "product-03", "Title 3", 9.99, "✅ Đã đăng", "Category", "When made", "", "", "2020_2026"),
        ]
        for row, folder, title, price, status, category, when_made, _a, _b, _c in rows:
            ws.cell(row=row, column=2, value=folder)
            ws.cell(row=row, column=5, value=price)
            ws.cell(row=row, column=6, value=category)
            ws.cell(row=row, column=8, value=title)
            ws.cell(row=row, column=14, value=status)
            ws.cell(row=row, column=13, value=when_made)
            ws.cell(row=row, column=15, value="Digital Planner")
        wb.save(excel_path)

        for folder in ("product-01", "product-02", "product-03"):
            (shop_dir / folder / "images").mkdir(parents=True, exist_ok=True)
            (shop_dir / folder / "files").mkdir(parents=True, exist_ok=True)
        return shop_dir, excel_path

    def test_normalize_requested_products_supports_repeat_and_comma(self):
        self.assertEqual(
            ["product-1", "product-2", "product-3"],
            etsy_auto_post._normalize_requested_products(["product-1, product-2", "product-2", "product-3"]),
        )

    def test_normalize_requested_products_rejects_invalid(self):
        with self.assertRaises(ValueError):
            etsy_auto_post._normalize_requested_products(["abc", "product-12"])

    def test_normalize_selected_products_preserves_exact_pairs(self):
        self.assertEqual(
            [(5, "product-02"), (4, "product-01")],
            etsy_auto_post._normalize_selected_products(["5:product-02", "4:product-01"]),
        )
        with self.assertRaises(ValueError):
            etsy_auto_post._normalize_selected_products(["5:product-02", "5:product-01"])

    def test_read_products_keeps_exact_selected_order(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            shop_dir, excel = self._mk_shop_and_excel(root)
            orig_shop_dir = etsy_auto_post.SHOP_DIR
            orig_excel = etsy_auto_post.EXCEL_FILE
            etsy_auto_post.SHOP_DIR = shop_dir
            etsy_auto_post.EXCEL_FILE = excel
            try:
                products, _, _, total = etsy_auto_post.read_products(
                    batch=10,
                    skip=0,
                    product_folder=None,
                    shop_id="daisyflowdigital",
                    product_folders=["product-02", "product-01"],
                )
                self.assertEqual(2, total)
                self.assertEqual(["product-02", "product-01"], [p["folder"] for p in products])
            finally:
                etsy_auto_post.SHOP_DIR = orig_shop_dir
                etsy_auto_post.EXCEL_FILE = orig_excel

    def test_read_products_rechecks_posted_status_and_listing_reference(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            shop_dir, excel = self._mk_shop_and_excel(root)
            orig_shop_dir = etsy_auto_post.SHOP_DIR
            orig_excel = etsy_auto_post.EXCEL_FILE
            etsy_auto_post.SHOP_DIR = shop_dir
            etsy_auto_post.EXCEL_FILE = excel
            try:
                with self.assertRaises(RuntimeError):
                    etsy_auto_post.read_products(
                        shop_id="daisyflowdigital",
                        product_folders=["product-03"],
                    )

                wb = openpyxl.load_workbook(excel)
                ws = wb["Listings"]
                ws.cell(row=4, column=16, value="https://www.etsy.com/your/shops/me/listing-editor/edit/9876543210")
                wb.save(excel)
                with self.assertRaises(RuntimeError):
                    etsy_auto_post.read_products(
                        shop_id="daisyflowdigital",
                        product_folders=["product-01"],
                    )
            finally:
                etsy_auto_post.SHOP_DIR = orig_shop_dir
                etsy_auto_post.EXCEL_FILE = orig_excel

    def test_read_products_rechecks_exact_selected_row_folder_pair(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            shop_dir, excel = self._mk_shop_and_excel(root)
            orig_shop_dir = etsy_auto_post.SHOP_DIR
            orig_excel = etsy_auto_post.EXCEL_FILE
            etsy_auto_post.SHOP_DIR = shop_dir
            etsy_auto_post.EXCEL_FILE = excel
            try:
                products, _, _, _ = etsy_auto_post.read_products(
                    shop_id="daisyflowdigital",
                    selected_products=[(5, "product-02"), (4, "product-01")],
                )
                self.assertEqual(
                    [(5, "product-02"), (4, "product-01")],
                    [(p["row"], p["folder"]) for p in products],
                )
                with self.assertRaises(RuntimeError):
                    etsy_auto_post.read_products(
                        shop_id="daisyflowdigital",
                        selected_products=[(4, "product-02")],
                    )
            finally:
                etsy_auto_post.SHOP_DIR = orig_shop_dir
                etsy_auto_post.EXCEL_FILE = orig_excel

    def test_shop_manager_identity_requires_expected_public_shop_link(self):
        class FakeLocator:
            def __init__(self, hrefs):
                self.hrefs = hrefs

            async def evaluate_all(self, _script):
                return self.hrefs

        class FakePage:
            def __init__(self, hrefs):
                self.url = "https://www.etsy.com/your/shops/me/tools/listings"
                self.hrefs = hrefs

            def locator(self, _selector):
                return FakeLocator(self.hrefs)

        with patch.object(
            etsy_auto_post,
            "SHOPS",
            {"daisyflowdigital": {"etsy_link": "https://www.etsy.com/shop/Daisyflowdigital"}},
        ):
            asyncio.run(
                etsy_auto_post._assert_shop_manager_identity(
                    FakePage(["https://www.etsy.com/shop/Daisyflowdigital?ref=seller-platform"]),
                    "daisyflowdigital",
                )
            )
            with self.assertRaises(RuntimeError):
                asyncio.run(
                    etsy_auto_post._assert_shop_manager_identity(
                        FakePage(["https://www.etsy.com/shop/Templystudios"]),
                        "daisyflowdigital",
                    )
                )

    def test_main_rejects_missing_selected_folders(self):
        with patch.object(
            etsy_auto_post,
            "read_products",
            return_value=([], Path("/dev/null"), Path("/dev/null"), 0),
        ):
            with patch.object(sys, "argv", [
                "etsy_auto_post.py",
                "--shop",
                "daisyflowdigital",
                "--products",
                "product-99999,product-01",
            ]):
                with self.assertRaises(RuntimeError) as context:
                    asyncio.run(etsy_auto_post.main())
                self.assertIn("Không tìm thấy folder", str(context.exception))


if __name__ == "__main__":
    import unittest

    unittest.main()
