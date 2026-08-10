"""
Sync Etsy Shop Manager listings into the dashboard Excel file.

This intentionally uses the logged-in Chrome/Playwright session instead of the
Etsy API because new Etsy API access is not available for this workflow.
"""
import argparse
import asyncio
import json
import re
import shutil
from collections import Counter
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright

from cloud_asset_store import CloudAssetError, CloudAssetStore
from etsy_browser_session import is_session_ready, resolve_etsy_session

BASE_DIR = Path(__file__).parent
SHOP_MANAGER_URL = "https://www.etsy.com/your/shops/me/tools/listings"
SHOPS_CONFIG_FILE = BASE_DIR / "shops_config.json"
# Only these statuses are eligible to change local workbook mappings/statuses.
# The snapshot crawl deliberately includes every Manager status below so the
# dashboard total remains an honest shop total rather than active+draft only.
SYNC_TARGET_STATUSES = ("active", "draft")
# Manager statuses retained in the read-only snapshot/report.
ALL_MANAGER_STATUSES = ("active", "draft", "inactive", "expired")
STATUS_LABELS = {
    "active": "✅ Đã đăng",
    "draft": "✅ Đã đăng draft",
    "inactive": "⏸ Inactive trên Etsy",
    "expired": "⌛ Expired trên Etsy",
}


def record_local_sync_candidate(
    shop_id: str,
    folder: str,
    store: CloudAssetStore,
) -> dict:
    """Record a local candidate only after an explicit asset-writing sync.

    ``sync_excel`` only reconciles Etsy listings into the workbook; it does not
    write product assets, so it deliberately does not call this helper. An
    asset-writing caller may invoke it after its write succeeds. The strict
    path checks prevent a shop copy, traversal, or incomplete product root
    from being treated as a canonical candidate.
    """

    shop_text = str(shop_id or "").strip()
    folder_text = str(folder or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]{0,127}", shop_text):
        return {"ok": True, "marked": False, "reason": "shop is not a safe identifier"}
    if not re.fullmatch(r"product-\d+", folder_text):
        return {"ok": True, "marked": False, "reason": "folder is not a product root"}

    shop_dir = BASE_DIR / "shops" / shop_text
    product_root = shop_dir / folder_text
    if (
        shop_dir.is_symlink()
        or not shop_dir.is_dir()
        or product_root.is_symlink()
        or not product_root.is_dir()
        or product_root.resolve().parent != shop_dir.resolve()
        or any(
            (product_root / dirname).is_symlink()
            or not (product_root / dirname).is_dir()
            for dirname in ("images", "files")
        )
    ):
        return {"ok": True, "marked": False, "reason": "product root is not clearly resolved"}

    try:
        result = store.record_local_candidate(product_root)
    except (CloudAssetError, OSError, ValueError, TypeError, KeyError) as exc:
        return {"ok": False, "marked": False, "reason": str(exc)}
    return dict(result) if isinstance(result, dict) else {"ok": False, "marked": False, "reason": "invalid core result"}


def normalize_title(value: str) -> str:
    text = str(value or "").lower()
    text = text.replace("&amp;", "&").replace("&quot;", '"').replace("&#39;", "'")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _as_status_list(payload: dict, status: str) -> list[dict]:
    """Return listing list for a status while tolerating older/partial snapshots."""
    raw = payload.get(status, [])
    return raw if isinstance(raw, list) else []


def title_score(a: str, b: str) -> float:
    na = normalize_title(a)
    nb = normalize_title(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return 0.95
    return SequenceMatcher(None, na, nb).ratio()


def is_ds_store_row(*values: str) -> bool:
    combined = " ".join(str(v or "") for v in values).lower()
    return ".ds_store" in combined or "ds store" in combined.replace(".", " ")


def extract_listing_id(url: str) -> str:
    match = re.search(r"/listing/([0-9]+)", str(url or ""))
    return match.group(1) if match else ""


def configured_shop_slug(shop_id: str) -> str:
    try:
        config = json.loads(SHOPS_CONFIG_FILE.read_text(encoding="utf-8"))
        shop_url = str(config.get(shop_id, {}).get("etsy_link", ""))
        match = re.search(r"/shop/([^/?#]+)", shop_url, re.I)
        return match.group(1).strip().lower() if match else ""
    except Exception:
        return ""


async def verify_active_etsy_shop(page, shop_id: str) -> str:
    expected = configured_shop_slug(shop_id)
    shop_links = await page.evaluate(r"""() => Array.from(document.querySelectorAll('a[href*="/shop/"]'))
      .map(a => a.href || a.getAttribute('href') || '')
      .map(href => {
        const match = href.match(/\/shop\/([^/?#]+)/i);
        return match ? match[1] : '';
      })
      .filter(Boolean)""")
    actual = next((slug for slug in shop_links if slug.lower() != "me"), "")
    if expected and actual.lower() != expected:
        raise RuntimeError(
            f"Chrome Etsy đang ở shop {actual or 'không xác định'}, "
            f"nhưng dashboard yêu cầu {expected}. Hãy chọn/đăng nhập đúng shop rồi sync lại."
        )
    return actual or expected


async def scrape_visible_listings(page, status: str, page_num: int) -> list[dict]:
    items = await page.evaluate(
        """({ status, pageNum }) => {
          const clean = (text) => (text || '')
            .replace(/\\s+/g, ' ')
            .replace(/\\b(Digital|Video|Physical item|Auto-renews|Manual renew)\\b/gi, ' ')
            .trim();
          const anchors = Array.from(document.querySelectorAll('a[href*="/listing-editor/edit/"]'));
          const out = [];
          const seen = new Set();
          for (const a of anchors) {
            const href = a.href || a.getAttribute('href') || '';
            const idMatch = href.match(/\\/listing-editor\\/edit\\/(\\d+)/);
            if (!idMatch) continue;
            const id = idMatch[1];
            if (seen.has(id)) continue;
            seen.add(id);

            const card = a.closest('[data-listing-id]')
              || a.closest('tr')
              || a.closest('li')
              || a.closest('[class*="listing"]')
              || a.closest('[class*="card"]')
              || a.parentElement;

            const titleCandidates = [];
            if (card) {
              for (const sel of [
                'h1', 'h2', 'h3',
                '[data-test-id*="title"]',
                '[class*="title"]',
                'a[href*="/listing/"]:not([href*="/listing-editor/"])',
                'p', 'span'
              ]) {
                for (const el of Array.from(card.querySelectorAll(sel))) {
                  const text = clean(el.innerText);
                  if (text && text.length > 8) titleCandidates.push(text);
                }
              }
            }
            const anchorText = clean(a.innerText);
            if (anchorText && anchorText.length > 8) titleCandidates.push(anchorText);

            let title = '';
            for (const candidate of titleCandidates) {
              if (/^(edit|editing options|preview|renew|copy|delete|stats)$/i.test(candidate)) continue;
              if (/\\$\\d|\\bin stock\\b|\\bviews\\b|\\bfavorites\\b/i.test(candidate)) continue;
              title = candidate;
              break;
            }
            if (!title) title = `Listing ${id}`;
            out.push({
              id,
              title,
              editUrl: href.startsWith('http') ? href : `https://www.etsy.com${href}`,
              url: `https://www.etsy.com/listing/${id}`,
              managerStatus: status,
              page: pageNum,
            });
          }
          return out;
        }""",
        {"status": status, "pageNum": page_num},
    )
    deduped = []
    seen = set()
    for item in items:
        if item["id"] in seen:
            continue
        seen.add(item["id"])
        deduped.append(item)
    return deduped


async def choose_status_filter(page, status: str) -> None:
    await page.goto(SHOP_MANAGER_URL, wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(2500)
    selector = f'input[name="item_status"][value="{status}"]'
    try:
        await page.locator(selector).first.click(force=True, timeout=8000)
        await page.wait_for_timeout(2500)
    except Exception:
        # Etsy changes markup often; the URL fallback still works for some states.
        await page.goto(f"{SHOP_MANAGER_URL}?item_status={status}", wait_until="domcontentloaded", timeout=60000)
        await page.wait_for_timeout(2500)


async def crawl_status(page, status: str) -> list[dict]:
    print(f"[CRAWL] Chọn filter {status}...")
    await choose_status_filter(page, status)
    all_items = []
    seen_ids = set()

    for page_num in range(1, 31):
        if page_num > 1:
            # Etsy's manager route keeps filters in the path, e.g.
            # /tools/listings/page:2,state:inactive. A slash before state drops
            # the filter and silently returns active/all listings.
            page_url = f"{SHOP_MANAGER_URL}/page:{page_num},state:{status}"
            await page.goto(page_url, wait_until="domcontentloaded", timeout=60000)
            await page.wait_for_timeout(2000)

        items = await scrape_visible_listings(page, status, page_num)
        fresh = [item for item in items if item["id"] not in seen_ids]
        if page_num > 1 and not fresh:
            break
        for item in fresh:
            seen_ids.add(item["id"])
            all_items.append(item)
        print(f"[CRAWL] {status} page {page_num}: +{len(fresh)} listing")
        if len(items) == 0:
            break

    print(f"[CRAWL] {status}: {len(all_items)} listing")
    return all_items


async def crawl_etsy_shop(shop_id: str) -> dict:
    shops = json.loads(SHOPS_CONFIG_FILE.read_text(encoding="utf-8")) if SHOPS_CONFIG_FILE.exists() else {}
    session = resolve_etsy_session(BASE_DIR, shops, shop_id)
    if not is_session_ready(session):
        raise RuntimeError(
            "Phiên Etsy chưa sẵn sàng cho shop này (chưa đăng nhập/không đúng Chrome CDP profile). "
            "Mở đúng cửa sổ Chrome đăng nhập Etsy trước, rồi chạy lại."
        )
    async with async_playwright() as pw:
        browser = await pw.chromium.connect_over_cdp(session.cdp_url, timeout=5000)
        if not browser.contexts:
            raise RuntimeError(
                "Không thấy browser context cho phiên Etsy. Mở cửa sổ Chrome đăng nhập cho shop trước rồi thử lại."
            )
        context = browser.contexts[0]
        page = None
        try:
            page = await context.new_page()
            page.set_default_timeout(30000)
            await page.goto(SHOP_MANAGER_URL, wait_until="domcontentloaded", timeout=60000)
            await page.wait_for_timeout(3000)
            if "signin" in page.url or "join" in page.url:
                raise RuntimeError("Chrome session chưa đăng nhập Etsy. Anh mở Chrome sync rồi login Etsy trước giúp em.")

            actual_shop = await verify_active_etsy_shop(page, shop_id)
            result = {
                "crawledAt": datetime.now().isoformat(timespec="seconds"),
                "shopId": shop_id,
                "shopSlug": actual_shop,
            }
            for status in ALL_MANAGER_STATUSES:
                result[status] = await crawl_status(page, status)
            return result
        finally:
            if page is not None:
                try:
                    await page.close()
                except Exception:
                    pass


def sync_excel(shop_id: str, crawled: dict) -> dict:
    shop_dir = BASE_DIR / "shops" / shop_id
    excel_path = shop_dir / "Etsy_SEO_Generator.xlsx"
    if not excel_path.exists():
        raise FileNotFoundError(f"Không thấy Excel cho shop {shop_id}: {excel_path}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = excel_path.with_name(f"Etsy_SEO_Generator.backup_etsy_shop_sync_{ts}.xlsx")
    report_path = shop_dir / f"etsy_shop_sync_report_{ts}.json"
    crawl_path = BASE_DIR / "scratch" / f"etsy_manager_current_{shop_id}_{ts}.json"
    shutil.copy2(excel_path, backup)
    crawl_path.write_text(json.dumps(crawled, ensure_ascii=False, indent=2), encoding="utf-8")

    # Do not make inactive/expired records candidates for Excel matching.  They
    # are preserved in ``crawled``/the snapshot for correct dashboard totals,
    # but cannot overwrite a local URL or status during sync reconciliation.
    listings = []
    for status in SYNC_TARGET_STATUSES:
        for item in _as_status_list(crawled, status):
            listing = dict(item)
            listing["managerStatus"] = status
            listing["url"] = f"https://www.etsy.com/listing/{listing['id']}"
            listings.append(listing)

    by_id = {item["id"]: item for item in listings}
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Listings"]
    used_ids = set()
    matched = []
    unmatched = []
    changed = []
    conflicts = []

    for row_num in range(4, ws.max_row + 1):
        folder = ws.cell(row=row_num, column=2).value
        keywords = ws.cell(row=row_num, column=3).value
        title = ws.cell(row=row_num, column=8).value
        old_status = ws.cell(row=row_num, column=14).value
        old_url = ws.cell(row=row_num, column=16).value
        if not folder and not title:
            continue

        current_id = extract_listing_id(old_url)
        existing = by_id.get(current_id) if current_id else None
        if current_id and existing is None:
            unmatched.append({
                "row": row_num,
                "folder": folder,
                "title": title,
                "old_status": old_status,
                "old_url": old_url,
                "existing_id": current_id,
                "preserved": True,
                "reason": "existing_listing_outside_sync_scope",
            })
            continue

        existing_score = title_score(title, existing["title"]) if existing else 0.0
        best = max(listings, key=lambda item: title_score(title, item["title"])) if listings else None
        best_score = title_score(title, best["title"]) if best else 0.0

        selected = None
        method = None
        score = 0.0
        if existing and existing_score >= 0.82:
            selected = existing
            method = "existing_url_id"
            score = existing_score
        elif best and best_score >= 0.86:
            selected = best
            method = "title_match"
            score = best_score

        if selected and selected["id"] in used_ids:
            conflicts.append({
                "row": row_num,
                "folder": folder,
                "title": title,
                "candidate_id": selected["id"],
                "score": round(score, 3),
                "method": method,
            })
            selected = None

        if selected:
            used_ids.add(selected["id"])
            new_status = STATUS_LABELS[selected["managerStatus"]]
            new_url = selected["url"]
            ws.cell(row=row_num, column=14, value=new_status)
            ws.cell(row=row_num, column=16, value=new_url)
            record = {
                "row": row_num,
                "folder": folder,
                "title": title,
                "listing_id": selected["id"],
                "etsy_title": selected["title"],
                "manager_status": selected["managerStatus"],
                "method": method,
                "score": round(score, 3),
                "old_status": old_status,
                "new_status": new_status,
                "old_url": old_url,
                "new_url": new_url,
            }
            matched.append(record)
            if old_status != new_status or old_url != new_url:
                changed.append(record)
        else:
            preserve_existing_url = bool(current_id)
            if preserve_existing_url:
                unmatched.append({
                    "row": row_num,
                    "folder": folder,
                    "title": title,
                    "old_status": old_status,
                    "old_url": old_url,
                    "best_listing_id": best["id"] if best else None,
                    "best_title": best["title"] if best else None,
                    "best_score": round(best_score, 3),
                    "existing_id": current_id,
                    "existing_score": round(existing_score, 3),
                    "preserved": True,
                    "reason": "existing_valid_etsy_listing_url",
                })
                continue

            new_status = "🗑 Rác .DS_Store" if is_ds_store_row(folder, keywords, title) else "⏳ Chờ đăng"
            ws.cell(row=row_num, column=14, value=new_status)
            ws.cell(row=row_num, column=16, value=None)
            unmatched.append({
                "row": row_num,
                "folder": folder,
                "title": title,
                "old_status": old_status,
                "old_url": old_url,
                "new_status": new_status,
                "best_listing_id": best["id"] if best else None,
                "best_title": best["title"] if best else None,
                "best_score": round(best_score, 3),
                "existing_id": current_id,
                "existing_score": round(existing_score, 3),
            })

    wb.save(excel_path)

    etsy_unmapped = [
        item for item in listings
        if item["id"] not in used_ids
    ]
    status_counts = Counter()
    url_count = 0
    for row_num in range(4, ws.max_row + 1):
        folder = ws.cell(row=row_num, column=2).value
        title = ws.cell(row=row_num, column=8).value
        if not folder and not title:
            continue
        status_counts[ws.cell(row=row_num, column=14).value] += 1
        if ws.cell(row=row_num, column=16).value:
            url_count += 1

    snapshot_total = sum(
        len(_as_status_list(crawled, status)) for status in ALL_MANAGER_STATUSES
    )
    report = {
        "created_at": ts,
        "shop": shop_id,
        "excel": str(excel_path),
        "backup": str(backup),
        "crawl": str(crawl_path),
        "etsy_counts": {status: len(_as_status_list(crawled, status)) for status in ALL_MANAGER_STATUSES},
        "etsy_total": snapshot_total,
        "syncable_etsy_total": len(listings),
        "matched_total": len(matched),
        "changed_total": len(changed),
        "unmatched_dashboard_total": len(unmatched),
        "etsy_unmapped_total": len(etsy_unmapped),
        "conflicts_total": len(conflicts),
        "dashboard_status_counts": dict(status_counts),
        "dashboard_url_count": url_count,
        "matched": matched,
        "unmatched_dashboard": unmatched,
        "etsy_unmapped": etsy_unmapped,
        "conflicts": conflicts,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--shop", default="templystudios")
    parser.add_argument("--crawl-file", default="")
    parser.add_argument(
        "--crawl-only",
        action="store_true",
        help="Chỉ cào Etsy Shop Manager và lưu snapshot; không sửa workbook.",
    )
    args = parser.parse_args()

    if args.crawl_file:
        crawled = json.loads(Path(args.crawl_file).read_text(encoding="utf-8"))
        if crawled.get("shopId") != args.shop:
            raise RuntimeError(f"Crawl file không thuộc shop {args.shop}; không ghi để tránh trộn dữ liệu shop.")
    else:
        crawled = await crawl_etsy_shop(args.shop)

    if args.crawl_only:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot_dir = BASE_DIR / "scratch"
        snapshot_dir.mkdir(parents=True, exist_ok=True)
        snapshot = snapshot_dir / f"etsy_manager_current_{args.shop}_{ts}.json"
        snapshot.write_text(json.dumps(crawled, ensure_ascii=False, indent=2), encoding="utf-8")
        print("[CRAWL] Snapshot:", snapshot)
        for status in SYNC_TARGET_STATUSES:
            print(f"[CRAWL] {status}: {len(_as_status_list(crawled, status))}")
        return

    report = sync_excel(args.shop, crawled)
    print("[SYNC] Etsy counts:", report["etsy_counts"])
    print("[SYNC] Matched dashboard rows:", report["matched_total"])
    print("[SYNC] Changed rows:", report["changed_total"])
    print("[SYNC] Dashboard unmatched:", report["unmatched_dashboard_total"])
    print("[SYNC] Etsy not mapped to dashboard:", report["etsy_unmapped_total"])
    print("[SYNC] Dashboard statuses:", report["dashboard_status_counts"])
    print("[SYNC] Report:", report["backup"])
    print("[SYNC] Report:", report["crawl"])
    print("[SYNC] Report:", str(Path(report["excel"]).parent / f"etsy_shop_sync_report_{report['created_at']}.json"))


if __name__ == "__main__":
    asyncio.run(main())
