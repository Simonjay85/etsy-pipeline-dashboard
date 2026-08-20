#!/usr/bin/env python3
"""Unified Etsy + local catalog and conservative local-folder deduplication."""

from __future__ import annotations

import hashlib
import json
import stat
import re
import shutil
import time
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
DIGITAL_EXTS = {".pdf", ".zip"}
HASH_EXTS = {".pdf", ".zip", ".svg", ".eps", ".dxf", ".ai", ".psd"}
IGNORED_NAMES = {".DS_Store"}


def _natural_sort_key(value: str) -> tuple[tuple[int, Any], ...]:
    """Return a deterministic, case-insensitive key with numeric chunks as ints."""
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", str(value or ""))
    )


def _catalog_record_sort_key(record: dict[str, Any]) -> tuple[Any, ...]:
    """Keep folder-backed products naturally ordered, followed by Etsy-only rows."""
    folder = str(record.get("folder") or "").strip()
    listing_id = str(record.get("listing_id") or "").strip()
    record_id = str(record.get("record_id") or "")
    if folder:
        return (0, _natural_sort_key(folder), folder.casefold(), folder, record_id)
    return (
        1,
        _natural_sort_key(listing_id),
        normalize_title(str(record.get("etsy_title") or record.get("title") or "")),
        record_id,
    )


def _canonical_sku(shop_id: str, folder_name: str) -> str:
    shop_key = str(shop_id or "").lower()
    if "daisy" in shop_key:
        prefix = "dd"
    elif "temply" in shop_key:
        prefix = "TS"
    else:
        prefix = (shop_id or "TS")[:2].upper() if shop_id else "TS"
    normalized_folder = str(folder_name or "").strip().lower().replace("-", "_")
    return f"{prefix}_{normalized_folder}"


def normalize_title(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    return " ".join(re.findall(r"[a-z0-9]+", text))


def extract_listing_id(url: str) -> str:
    match = re.search(r"/listing/([0-9]+)", str(url or ""))
    return match.group(1) if match else ""


def _normalize_etsy_listing_id(value: Any) -> str:
    text = str(value or "").strip()
    return text if text.isdigit() else ""


def normalize_etsy_manager_snapshot(raw: dict[str, Any]) -> dict[str, Any]:
    """Normalize and dedupe an Etsy manager snapshot payload.

    Status precedence is active > draft > inactive > expired.
    Malformed/missing listing IDs are kept as independent records.
    """
    status_order = ("active", "draft", "inactive", "expired")
    raw_counts = {status: 0 for status in status_order}
    winners_by_id: dict[str, dict[str, Any]] = {}
    winner_order: list[str] = []
    malformed_records: list[dict[str, Any]] = []
    duplicate_count = 0

    for status in status_order:
        values = raw.get(status, [])
        if not isinstance(values, list):
            continue
        raw_counts[status] = len(values)
        for item in values:
            if not isinstance(item, dict):
                continue
            listing = dict(item)
            listing["managerStatus"] = status
            listing_id = _normalize_etsy_listing_id(item.get("id"))
            if not listing_id:
                malformed_records.append(listing)
                continue
            if listing_id in winners_by_id:
                duplicate_count += 1
                continue
            winners_by_id[listing_id] = listing
            winner_order.append(listing_id)

    listings: list[dict[str, Any]] = []
    for listing_id in winner_order:
        listing = dict(winners_by_id[listing_id])
        listing["listing_id"] = listing_id
        listings.append(listing)

    for listing in malformed_records:
        listing = dict(listing)
        listing["listing_id"] = str(listing.get("id") or "")
        listings.append(listing)

    counts = {status: 0 for status in status_order}
    for listing in listings:
        status = str(listing.get("managerStatus", "")).lower()
        if status in counts:
            counts[status] += 1

    return {
        "raw_counts": raw_counts,
        "counts": counts,
        "duplicate_count": duplicate_count,
        "listings": listings,
    }


def _safe_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _quick_signature(path: Path) -> str:
    """Fast sparse-file signature; full hashes are reserved for merge validation.

    A large part of this workspace is iCloud-backed and many files report zero
    allocated blocks until opened. Metadata-only indexing keeps the catalog
    scan bounded; candidate groups are byte-verified only during merge.
    """
    stat = path.stat()
    digest = hashlib.sha256()
    digest.update(str(stat.st_size).encode())
    digest.update(path.name.lower().encode("utf-8", "ignore"))
    return digest.hexdigest()


def _digital_file_sha256_set(folder_path: Path) -> tuple[set[str], list[str]]:
    """
    Return full SHA-256 hashes for all local PDF/ZIP files under a folder.
    Empty folders (no supported files) return an empty set and no failures.
    """
    hashes: set[str] = set()
    failures: list[str] = []
    if not folder_path.is_dir():
        return hashes, failures

    for path in folder_path.rglob("*"):
        if path.name in IGNORED_NAMES or path.suffix.lower() not in DIGITAL_EXTS:
            continue
        try:
            relative_path = str(path.relative_to(folder_path))
        except ValueError:
            relative_path = path.name
        try:
            if not stat.S_ISREG(path.stat().st_mode):
                continue
        except OSError:
            failures.append(relative_path)
            continue
        try:
            hashes.add(_safe_hash(path))
        except OSError:
            failures.append(relative_path)
            continue

    return hashes, failures


def _digital_hash_match_for_folders(folder_paths: list[Path]) -> tuple[bool, dict[str, list[str]], dict[str, list[str]]]:
    """
    Return whether PDF/ZIP hash sets are fully equal across candidate folders.
    Any hashing/stat failure fails the check and returns read-failure evidence.
    """
    hash_sets: dict[str, set[str]] = {}
    failed_paths: dict[str, list[str]] = {}
    for path in folder_paths:
        hashes, failures = _digital_file_sha256_set(path)
        hash_sets[str(path)] = hashes
        if failures:
            failed_paths[str(path)] = failures

    if not hash_sets:
        return False, {}, {}
    if failed_paths:
        return False, {k: sorted(list(v)) for k, v in hash_sets.items()}, {k: sorted(v) for k, v in failed_paths.items()}
    first: set[str] | None = None
    for hashes in hash_sets.values():
        if not hashes:
            return False, {k: sorted(list(v)) for k, v in hash_sets.items()}, {}
        if first is None:
            first = hashes
            continue
        if hashes != first:
            return False, {k: sorted(list(v)) for k, v in hash_sets.items()}, {}
    return True, {k: sorted(list(v)) for k, v in hash_sets.items()}, {}


def _deliverable_hash_multiset(folder_path: Path) -> tuple[Counter[str], list[str]]:
    """Hash every customer-deliverable file under files/, preserving duplicates."""
    hashes: Counter[str] = Counter()
    failures: list[str] = []
    files_dir = folder_path / "files"
    if not files_dir.is_dir():
        return hashes, failures
    for path in sorted(files_dir.rglob("*")):
        if not path.is_file() or path.name in IGNORED_NAMES:
            continue
        relative = str(path.relative_to(folder_path))
        try:
            if path.stat().st_size <= 0:
                failures.append(relative)
                continue
            hashes[_safe_hash(path)] += 1
        except OSError:
            failures.append(relative)
    return hashes, failures


def _deliverable_match_for_folders(folder_paths: list[Path]) -> tuple[bool, dict[str, dict[str, int]], dict[str, list[str]]]:
    multisets: dict[str, Counter[str]] = {}
    failures: dict[str, list[str]] = {}
    for folder_path in folder_paths:
        hashes, failed = _deliverable_hash_multiset(folder_path)
        multisets[str(folder_path)] = hashes
        if failed:
            failures[str(folder_path)] = failed
    evidence = {path: dict(counter) for path, counter in multisets.items()}
    if failures or not multisets or any(not counter for counter in multisets.values()):
        return False, evidence, failures
    values = list(multisets.values())
    return all(counter == values[0] for counter in values[1:]), evidence, failures


def _safe_sha256_match(path_a: Path, path_b: Path) -> bool:
    """
    Compare file content using full SHA-256. Any read/open/stat failure is treated
    as a mismatch.
    """
    try:
        return _safe_hash(path_a) == _safe_hash(path_b)
    except OSError:
        return False


def _folder_assets(folder_path: Path) -> dict[str, Any]:
    image_count = 0
    file_count = 0
    total_bytes = 0
    digital_files: list[dict[str, Any]] = []
    asset_hashes: list[dict[str, Any]] = []

    if not folder_path.is_dir():
        return {
            "image_count": 0,
            "file_count": 0,
            "total_bytes": 0,
            "digital_files": [],
            "asset_hashes": [],
        }

    for path in folder_path.rglob("*"):
        if not path.is_file() or path.name in IGNORED_NAMES or ".thumbcache" in path.parts:
            continue
        try:
            size = path.stat().st_size
        except OSError:
            continue
        rel = str(path.relative_to(folder_path))
        suffix = path.suffix.lower()
        total_bytes += size
        if suffix in IMAGE_EXTS and "images" in path.parts:
            image_count += 1
        if "files" in path.parts and suffix in DIGITAL_EXTS:
            file_count += 1
            digital_files.append({"path": rel, "name": path.name, "size": size})
        if suffix in HASH_EXTS and size > 1024:
            try:
                asset_hashes.append({
                    "path": rel,
                    "name": path.name,
                    "size": size,
                    "sha256": _quick_signature(path),
                })
            except OSError:
                pass

    return {
        "image_count": image_count,
        "file_count": file_count,
        "total_bytes": total_bytes,
        "digital_files": digital_files,
        "asset_hashes": asset_hashes,
    }


def _load_workbook_rows(excel_path: Path) -> dict[str, dict[str, Any]]:
    if not excel_path.exists():
        return {}
    # The active workbook may be an iCloud sparse file. Reading the small XML
    # payload directly avoids openpyxl waiting for a full workbook hydration.
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(excel_path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall(f"{namespace}si"):
                shared_strings.append("".join(node.text or "" for node in item.iter(f"{namespace}t")))
        root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))

        def cell_value(cell: ET.Element) -> str:
            value = cell.find(f"{namespace}v")
            if value is None:
                inline = cell.find(f"{namespace}is/{namespace}t")
                return inline.text if inline is not None and inline.text else ""
            raw = value.text or ""
            if cell.attrib.get("t") == "s":
                try:
                    return shared_strings[int(raw)]
                except (ValueError, IndexError):
                    return raw
            return raw

    rows: dict[str, dict[str, Any]] = {}
    for row_node in root.findall(f"{namespace}sheetData/{namespace}row"):
        try:
            row_num = int(row_node.attrib.get("r", "0"))
        except ValueError:
            continue
        if row_num < 4:
            continue
        values: dict[str, str] = {}
        for cell in row_node.findall(f"{namespace}c"):
            ref = str(cell.attrib.get("r", ""))
            column = re.match(r"([A-Z]+)", ref)
            if column:
                values[column.group(1)] = cell_value(cell)
        folder = str(values.get("B", "") or "").strip()
        if not folder:
            continue
        rows[folder] = {
            "row": row_num,
            "title": str(values.get("H", "") or "").strip(),
            "description": str(values.get("I", "") or "").strip(),
            "status": str(values.get("N", "") or "⏳ Chờ đăng").strip(),
            "is_new_import": "Mới import" in str(values.get("N", "") or ""),
            "etsy_url": str(values.get("P", "") or "").strip(),
            "extra": str(values.get("Q", "") or "").strip(),
            "sku": str(values.get("R", "") or "").strip(),
        }
    return rows


def load_local_catalog(base_dir: Path, shop_id: str, excel_path: Path) -> list[dict[str, Any]]:
    shop_dir = base_dir / "shops" / shop_id
    workbook_rows = _load_workbook_rows(excel_path)
    folder_names = {
        path.name
        for path in shop_dir.iterdir()
        if path.is_dir() and path.name.startswith("product-")
    } if shop_dir.is_dir() else set()

    records: list[dict[str, Any]] = []
    for folder in sorted(folder_names, key=lambda value: (_natural_sort_key(value), value.casefold(), value)):
        folder_path = shop_dir / folder
        row = workbook_rows.get(folder, {})
        assets = _folder_assets(folder_path)
        title = row.get("title", "")
        records.append({
            "record_id": f"local:{shop_id}:{folder}",
            "source": "local",
            "shop": shop_id,
            "folder": folder,
            "row": row.get("row"),
            "exists": folder_path.is_dir(),
            "title": title or folder,
            "normalized_title": normalize_title(title),
            "status": row.get("status", "📁 Local chưa đăng ký") if row else "📁 Local chưa đăng ký",
            "is_new_import": bool(row.get("is_new_import", False)) if row else False,
            "etsy_url": row.get("etsy_url", ""),
            "listing_id": extract_listing_id(row.get("etsy_url", "")),
            "sku": row.get("sku", ""),
            "extra": row.get("extra", ""),
            "image_count": assets["image_count"],
            "file_count": assets["file_count"],
            "total_bytes": assets["total_bytes"],
            "digital_files": assets["digital_files"],
            "asset_hashes": assets["asset_hashes"],
        })
    return records


def load_etsy_snapshot(base_dir: Path, shop_id: str) -> dict[str, Any]:
    candidates = sorted((base_dir / "scratch").glob(f"etsy_manager_current_{shop_id}_*.json"))
    if not candidates:
        return {"source": "", "counts": {}, "listings": []}
    latest = candidates[-1]
    try:
        data = json.loads(latest.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"source": str(latest), "counts": {}, "listings": []}

    normalized = normalize_etsy_manager_snapshot(data)
    counts = dict(normalized.get("counts", {}))
    raw_counts = dict(normalized.get("raw_counts", {}))
    raw_counts["total"] = sum(raw_counts.values())
    if counts:
        counts["total"] = sum(counts.values())
    return {
        "source": str(latest),
        "counts": counts,
        "raw_counts": raw_counts,
        "duplicate_count": normalized.get("duplicate_count", 0),
        "listings": normalized.get("listings", []),
    }


def build_unified_catalog(base_dir: Path, shop_id: str, excel_path: Path) -> dict[str, Any]:
    local = load_local_catalog(base_dir, shop_id, excel_path)
    snapshot = load_etsy_snapshot(base_dir, shop_id)
    by_listing: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in local:
        if record.get("listing_id"):
            by_listing[str(record["listing_id"])].append(record)
    records: list[dict[str, Any]] = []
    mapped_listing_ids: set[str] = set()
    mapped_local_record_ids: set[str] = set()

    etsy_only_hidden_non_syncable_total = 0
    for listing in snapshot["listings"]:
        listing_id = str(listing.get("listing_id") or listing.get("id") or "")
        listing_status = str(listing.get("managerStatus") or "").strip().lower()
        local_records = by_listing.get(listing_id, [])
        if local_records:
            mapped_listing_ids.add(listing_id)
            for local_record in local_records:
                mapped_local_record_ids.add(str(local_record["record_id"]))
                records.append({
                    **local_record,
                    "source": "both",
                    "source_label": "Etsy + local",
                    "etsy_title": str(listing.get("title") or ""),
                    "etsy_status": str(listing.get("managerStatus") or ""),
                    "etsy_url": str(listing.get("url") or local_record.get("etsy_url") or ""),
                    "listing_id": listing_id,
                    "record_id": f"both:{shop_id}:{listing_id}:{local_record.get('folder') or local_record['record_id']}",
                })
        elif listing_status not in {"active", "draft"}:
            etsy_only_hidden_non_syncable_total += 1
            continue
        else:
            records.append({
                "record_id": f"etsy:{shop_id}:{listing_id}",
                "source": "etsy",
                "source_label": "Etsy",
                "shop": shop_id,
                "folder": "",
                "row": None,
                "exists": False,
                "title": str(listing.get("title") or f"Listing {listing_id}"),
                "normalized_title": normalize_title(listing.get("title", "")),
                "status": str(listing.get("managerStatus") or ""),
                "etsy_status": str(listing.get("managerStatus") or ""),
                "etsy_title": str(listing.get("title") or ""),
                "etsy_url": str(listing.get("url") or f"https://www.etsy.com/listing/{listing_id}"),
                "listing_id": listing_id,
                "sku": "",
                "extra": "",
                "image_count": int(listing.get("image_count") or 0),
                "file_count": 0,
                "total_bytes": 0,
                "digital_files": [],
                "asset_hashes": [],
            })

    for local_record in local:
        if str(local_record.get("record_id") or "") in mapped_local_record_ids:
            continue
        local_record = dict(local_record)
        local_record["source_label"] = "Local"
        # A local row can retain an old Etsy URL after that listing no longer
        # appears in the newest manager snapshot.  Keep the workbook status
        # intact (it is historical metadata), but expose the reconciliation
        # mismatch so the dashboard does not present it as confirmed posted.
        listing_id = str(local_record.get("listing_id") or "")
        if snapshot.get("source") and listing_id and listing_id not in mapped_listing_ids:
            local_record["reconciliation_status"] = "unmatched_local_listing"
            local_record["reconciliation_note"] = "Listing ID is absent from the latest Etsy Manager snapshot"
        records.append(local_record)

    records.sort(key=_catalog_record_sort_key)
    duplicates = detect_duplicate_groups(records)
    return {
        "shop": shop_id,
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "source": snapshot.get("source", ""),
        "counts": {
            "unified_total": len(records),
            "etsy_total": len(snapshot.get("listings", [])),
            "local_total": len(local),
            "mapped_total": len(mapped_local_record_ids),
            "mapped_listing_total": len(mapped_listing_ids),
            "etsy_only_total": sum(1 for item in records if item["source"] == "etsy"),
            "local_only_total": sum(1 for item in records if item["source"] == "local"),
            "etsy_only_hidden_non_syncable_total": etsy_only_hidden_non_syncable_total,
            "duplicate_groups": len(duplicates),
            "safe_merge_groups": sum(1 for group in duplicates if group["safe_to_merge"]),
        },
        "records": records,
        "duplicate_groups": duplicates,
    }


def detect_duplicate_groups(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    evidence: dict[str, dict[str, Any]] = {}

    for record in records:
        if record.get("source") == "etsy":
            continue
        folder = str(record.get("folder") or "")
        if not folder:
            continue
        for asset in record.get("asset_hashes", []):
            key = f"asset:{asset['sha256']}"
            by_key[key].append(record)
            evidence[key] = {
                "match_type": "matching digital asset signature",
                "match_value": asset["sha256"],
                "asset_names": set(),
            }
            evidence[key]["asset_names"].add(asset["name"])
        title = record.get("normalized_title") or ""
        if title and len(title.split()) >= 3:
            key = f"title:{title}"
            by_key[key].append(record)
            evidence.setdefault(key, {
                "match_type": "exact normalized title",
                "match_value": title,
                "asset_names": set(),
            })

    groups: list[dict[str, Any]] = []
    seen_pairs: set[tuple[str, ...]] = set()
    for key, values in by_key.items():
        unique = {item["record_id"]: item for item in values}
        if len(unique) < 2:
            continue
        records_in_group = list(unique.values())
        folders = sorted({str(item.get("folder")) for item in records_in_group if item.get("folder")})
        pair_key = tuple(folders)
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)
        ev = evidence[key]
        is_asset = key.startswith("asset:")
        listing_ids = sorted({str(item.get("listing_id")) for item in records_in_group if item.get("listing_id")})
        safe_to_merge = bool(is_asset and len(listing_ids) <= 1 and len(folders) >= 2)
        groups.append({
            "group_id": hashlib.sha1("|".join(folders).encode()).hexdigest()[:12],
            "match_type": ev["match_type"],
            "match_value": ev["match_value"],
            "asset_names": sorted(ev["asset_names"]),
            "folders": folders,
            "listing_ids": listing_ids,
            "safe_to_merge": safe_to_merge,
            "records": records_in_group,
        })
    groups.sort(key=lambda item: (not item["safe_to_merge"], item["match_type"], item["folders"]))
    return groups


def _canonical_record(records: list[dict[str, Any]]) -> dict[str, Any]:
    def score(item: dict[str, Any]) -> tuple[int, int, int, int, str]:
        registered = 1 if item.get("row") else 0
        return (
            registered,
            int(item.get("total_bytes") or 0),
            int(item.get("image_count") or 0),
            int(item.get("file_count") or 0),
            str(item.get("folder") or ""),
        )
    return sorted(records, key=score, reverse=True)[0]


def merge_safe_duplicates(base_dir: Path, shop_id: str, excel_path: Path, requested_group_ids: list[str] | None = None) -> dict[str, Any]:
    catalog = build_unified_catalog(base_dir, shop_id, excel_path)
    groups = [group for group in catalog["duplicate_groups"] if group["safe_to_merge"]]
    if requested_group_ids is not None:
        allowed = set(requested_group_ids)
        groups = [group for group in groups if group["group_id"] in allowed]
    if not groups:
        return {"ok": True, "message": "Không có nhóm trùng an toàn để dồn.", "merged": [], "catalog": catalog}

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    shop_dir = base_dir / "shops" / shop_id
    quarantine = shop_dir / f".quarantine_duplicates_{timestamp}"
    quarantine.mkdir(parents=True, exist_ok=True)
    backup = excel_path.with_name(f"{excel_path.stem}.backup_aggregate_merge_{timestamp}{excel_path.suffix}")
    shutil.copy2(excel_path, backup)

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook["Listings"]
    merged: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    moved_source_folders: set[str] = set()
    for group in groups:
        local_records = [item for item in group["records"] if item.get("folder")]
        # Verify that the candidate assets are identical before mutating anything
        # using full PDF/ZIP hash comparison.
        verification_mode = "full SHA-256 multiset equality of every file under files/"
        available_records = [
            item for item in local_records
            if str(item.get("folder") or "") not in moved_source_folders
            and (shop_dir / str(item.get("folder") or "")).is_dir()
        ]
        if len(available_records) < 2:
            continue
        folder_paths = [(shop_dir / str(item["folder"])) for item in available_records]
        is_digital_match, folder_hashes, failed_paths = _deliverable_match_for_folders(folder_paths)
        if not is_digital_match:
            skipped.append({
                "group_id": group["group_id"],
                "match_type": group["match_type"],
                "verification_mode": verification_mode,
                "reason": "Deliverable hash set mismatch (multiset) or hash/read failure",
                "folders": sorted({str(item.get("folder")) for item in available_records}),
                "folder_digital_hashes": folder_hashes,
                "failed_relative_paths": failed_paths,
            })
            continue
        canonical = _canonical_record(available_records)
        canonical_folder = str(canonical["folder"])
        canonical_path = shop_dir / canonical_folder
        moved_folders: list[str] = []
        conflicts: list[str] = []
        destination_entries: dict[str, list[tuple[str, Path]]] = defaultdict(list)
        source_files: list[tuple[str, Path, Path]] = []
        for record in available_records:
            folder = str(record["folder"])
            root = shop_dir / folder
            for source_file in sorted(root.rglob("*")):
                if not source_file.is_file() or source_file.name in IGNORED_NAMES:
                    continue
                relative = source_file.relative_to(root)
                key = str(relative)
                destination_entries[key].append((folder, source_file))
                if folder != canonical_folder and not str(relative).startswith("files/"):
                    source_files.append((folder, source_file, relative))

        # Unique relative paths cannot collide, so avoid hydrating every iCloud
        # image during preflight. Only duplicate destination paths need byte
        # comparison; files that must be copied are verified again during the
        # atomic copy itself.
        for relative, entries in destination_entries.items():
            if len(entries) < 2:
                continue
            digests: list[tuple[str, str]] = []
            for folder, source_file in entries:
                try:
                    digests.append((folder, _safe_hash(source_file)))
                except OSError:
                    conflicts.append(f"{folder}/{relative} (read failure)")
            if digests and len({digest for _, digest in digests}) > 1:
                conflicts.append(" <> ".join(f"{folder}/{relative}" for folder, _ in digests))

        if conflicts:
            skipped.append({
                "group_id": group["group_id"],
                "match_type": group["match_type"],
                "verification_mode": verification_mode,
                "reason": "Phát hiện collision/read failure trong manifest đích trước khi dồn",
                "folders": sorted({str(item.get("folder")) for item in available_records}),
                "conflicts": sorted(conflicts),
                "folder_digital_hashes": folder_hashes,
                "failed_relative_paths": failed_paths,
            })
            continue

        created_targets: list[Path] = []
        temporary_targets: list[Path] = []
        quarantined: list[tuple[Path, Path]] = []
        row_backups: list[tuple[int, Any, Any, Any, Any, Any]] = []
        mapped_record = next((record for record in available_records if record.get("listing_id")), None)
        mapped_row = mapped_record.get("row") if mapped_record else None
        mapped_url = str(sheet.cell(row=mapped_row, column=16).value or "").strip() if mapped_row else ""
        mapped_status = str(sheet.cell(row=mapped_row, column=14).value or "").strip() if mapped_row else ""
        canonical_row = canonical.get("row")
        workbook_stage = excel_path.with_name(
            f".{excel_path.stem}.merge_{timestamp}_{group['group_id']}.tmp{excel_path.suffix}"
        )
        try:
            if canonical_row:
                row_backups.append((
                    canonical_row,
                    sheet.cell(row=canonical_row, column=2).value,
                    sheet.cell(row=canonical_row, column=14).value,
                    sheet.cell(row=canonical_row, column=16).value,
                    sheet.cell(row=canonical_row, column=17).value,
                    sheet.cell(row=canonical_row, column=18).value,
                ))
                if mapped_record and not canonical.get("listing_id") and mapped_row and mapped_row != canonical_row:
                    sheet.cell(row=canonical_row, column=14).value = mapped_status or sheet.cell(row=canonical_row, column=14).value
                    sheet.cell(row=canonical_row, column=16).value = mapped_url
                sheet.cell(row=canonical_row, column=18).value = _canonical_sku(shop_id, canonical_folder)

            for _folder, source_file, relative in source_files:
                target = canonical_path / relative
                if target.exists():
                    continue
                target.parent.mkdir(parents=True, exist_ok=True)
                source_digest = _safe_hash(source_file)
                temporary = target.with_name(f".{target.name}.merge_{group['group_id']}.tmp")
                temporary_targets.append(temporary)
                shutil.copy2(source_file, temporary)
                if _safe_hash(temporary) != source_digest:
                    raise OSError(f"Hash mismatch after copy: {source_file} -> {target}")
                temporary.replace(target)
                temporary_targets.remove(temporary)
                created_targets.append(target)

            for record in available_records:
                folder = str(record["folder"])
                if folder == canonical_folder:
                    continue
                source_path = shop_dir / folder
                destination = quarantine / folder
                destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.move(str(source_path), str(destination))
                quarantined.append((destination, source_path))
                moved_folders.append(folder)

            for record in available_records:
                folder = str(record["folder"])
                if folder == canonical_folder:
                    continue
                row_num = record.get("row")
                if row_num:
                    row_backups.append((
                        row_num,
                        sheet.cell(row=row_num, column=2).value,
                        sheet.cell(row=row_num, column=14).value,
                        sheet.cell(row=row_num, column=16).value,
                        sheet.cell(row=row_num, column=17).value,
                        sheet.cell(row=row_num, column=18).value,
                    ))
                    sheet.cell(row=row_num, column=2).value = None
                    sheet.cell(row=row_num, column=14).value = f"🔁 Dồn vào {canonical_folder}"
                    sheet.cell(row=row_num, column=16).value = None
                    sheet.cell(row=row_num, column=18).value = None
                    old_extra = str(sheet.cell(row=row_num, column=17).value or "").strip()
                    note = f"merged_into={canonical_folder}; quarantine={quarantine.name}"
                    sheet.cell(row=row_num, column=17).value = f"{old_extra} | {note}".strip(" |")

            # Persist the workbook inside this group's transaction. Saving to a
            # sibling and replacing atomically keeps the previous workbook valid
            # if serialization fails.
            workbook.save(workbook_stage)
            workbook_stage.replace(excel_path)
            for folder in moved_folders:
                moved_source_folders.add(folder)
        except Exception as exc:
            for row_num, folder_value, status_value, url_value, extra_value, sku_value in row_backups:
                sheet.cell(row=row_num, column=2).value = folder_value
                sheet.cell(row=row_num, column=14).value = status_value
                sheet.cell(row=row_num, column=16).value = url_value
                sheet.cell(row=row_num, column=17).value = extra_value
                sheet.cell(row=row_num, column=18).value = sku_value
            for destination, source_path in reversed(quarantined):
                if destination.exists() and not source_path.exists():
                    shutil.move(str(destination), str(source_path))
            for target in reversed(created_targets):
                target.unlink(missing_ok=True)
            for temporary in temporary_targets:
                temporary.unlink(missing_ok=True)
            workbook_stage.unlink(missing_ok=True)
            skipped.append({
                "group_id": group["group_id"],
                "match_type": group["match_type"],
                "verification_mode": verification_mode,
                "reason": f"Lỗi khi dồn; đã rollback group: {exc}",
                "folders": sorted({str(item.get("folder")) for item in available_records}),
                "folder_digital_hashes": folder_hashes,
                "failed_relative_paths": failed_paths,
            })
            continue
        finally:
            workbook_stage.unlink(missing_ok=True)
            for temporary in temporary_targets:
                temporary.unlink(missing_ok=True)

        if not moved_folders:
            continue

        merged.append({
            "group_id": group["group_id"],
            "match_type": group["match_type"],
            "verification": verification_mode,
            "asset_names": group["asset_names"],
            "canonical_folder": canonical_folder,
            "moved_folders": moved_folders,
            "conflicts": conflicts,
        })

    manifest = {
        "created_at": timestamp,
        "shop": shop_id,
        "quarantine": str(quarantine),
        "excel_backup": str(backup),
        "skipped": skipped,
        "merged": merged,
        "pre_merge_counts": catalog["counts"],
    }
    manifest_path = shop_dir / f"aggregate_duplicate_merge_{timestamp}.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "ok": True,
        "manifest": str(manifest_path),
        "backup": str(backup),
        "merged": merged,
        "skipped": skipped,
        "catalog": build_unified_catalog(base_dir, shop_id, excel_path),
    }
