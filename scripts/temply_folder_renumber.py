#!/usr/bin/env python3
"""Reviewable, reversible Temply catalog-folder renumber migration.

The command is deliberately dry-run by default.  It only updates the workbook's
Listings column B, the Temply-specific fields in product_source_map.json, and
the two current Temply social stores when ``--apply`` is supplied.

No product assets are copied.  Folder moves are recorded in a timestamped
backup manifest and performed through hidden, same-directory temporary names so
that source/target name overlap cannot cause a destructive overwrite.
"""

from __future__ import annotations

import argparse
import copy
import fcntl
import hashlib
import json
import os
import re
import shutil
import sys
import tempfile
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook


DEFAULT_SHOP_ID = "templystudios"
DEFAULT_EXPECTED_CATALOG_COUNT = 275
SOCIAL_STORE_NAMES = ("social_post_status.json", "social_posts.json")
SHOP_MANIFEST_NAME = "2027_PRODUCT_MANIFEST.md"
PRODUCT_RE = re.compile(r"^product-(\d+)$")
JOURNAL_NAME = "migration_journal.json"
MANIFEST_NAME = "mapping_manifest.json"
LOCK_NAME = ".temply-folder-renumber.lock"
BACKUP_EXCEL_NAME = "Etsy_SEO_Generator.xlsx"
BACKUP_MAP_NAME = "product_source_map.json"
BACKUP_STATUS_NAME = "social_post_status.json"
BACKUP_SOCIAL_NAME = "social_posts.json"
BACKUP_SHOP_MANIFEST_NAME = SHOP_MANIFEST_NAME
TERMINAL_JOURNAL_STATUSES = {"completed", "rolled_back"}


class MigrationError(RuntimeError):
    """A fail-closed preflight, mutation, or rollback error."""


@dataclass(frozen=True)
class FolderMapping:
    old_name: str
    new_name: str
    temporary_name: str
    rollback_temporary_name: str
    excel_rows: tuple[int, ...]


@dataclass
class MigrationPlan:
    shop_id: str
    shop_dir: Path
    excel_path: Path
    map_path: Path
    social_paths: tuple[Path, ...]
    shop_manifest_path: Path
    catalog_rows: tuple[tuple[int, str], ...]
    mappings: tuple[FolderMapping, ...]
    updated_map: dict[str, Any]
    updated_social: tuple[tuple[Path, Any], ...]
    map_key_updates: int
    map_value_updates: int
    social_key_updates: int
    social_folder_updates: int
    shop_manifest_updates: int
    updated_shop_manifest: str
    source_hashes: dict[str, str]
    asset_inventory: dict[str, tuple[int, ...]]
    workbook_snapshot: dict[str, Any]


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def sha256_file(path: Path) -> str:
    if not path.is_file():
        raise MigrationError(f"required source file is missing: {path}")
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _dimension_snapshot(dimension: Any) -> tuple[Any, ...]:
    return (
        dimension.hidden,
        dimension.outlineLevel,
        dimension.collapsed,
        dimension.width if hasattr(dimension, "width") else None,
        dimension.height if hasattr(dimension, "height") else None,
        dimension.style if hasattr(dimension, "style") else None,
        dimension.number_format if hasattr(dimension, "number_format") else None,
    )


def _cell_snapshot(cell: Any, *, omit_value: bool) -> tuple[Any, ...]:
    hyperlink = cell.hyperlink.target if cell.hyperlink else None
    comment = None if cell.comment is None else (cell.comment.text, cell.comment.author)
    style = tuple(cell._style) if cell.has_style else None
    return (
        None if omit_value else cell.value,
        None if omit_value else cell.data_type,
        hyperlink,
        comment,
        style,
    )


def workbook_snapshot(workbook: Any, target_rows: Iterable[int]) -> dict[str, Any]:
    """Capture workbook structure and every cell except target Listings!B values."""
    target = {("Listings", row, 2) for row in target_rows}
    sheets: dict[str, Any] = {}
    for sheet in workbook.worksheets:
        cells = {
            cell.coordinate: _cell_snapshot(
                cell,
                omit_value=(sheet.title, cell.row, cell.column) in target,
            )
            for (_, _), cell in sorted(sheet._cells.items())
        }
        row_dimensions = {
            str(key): _dimension_snapshot(value)
            for key, value in sorted(sheet.row_dimensions.items())
        }
        column_dimensions = {
            str(key): _dimension_snapshot(value)
            for key, value in sorted(sheet.column_dimensions.items())
        }
        sheets[sheet.title] = {
            "state": sheet.sheet_state,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            "auto_filter": sheet.auto_filter.ref,
            "merged_cells": tuple(str(value) for value in sheet.merged_cells.ranges),
            "row_dimensions": row_dimensions,
            "column_dimensions": column_dimensions,
            "cells": cells,
        }
    return {
        "sheetnames": tuple(workbook.sheetnames),
        "sheets": sheets,
    }


def inventory_assets(shop_dir: Path, folder_names: Iterable[str]) -> dict[str, tuple[int, ...]]:
    """Inventory files without reading/copying their contents.

    Each record is ``(bytes, mtime_ns, permission_bits, kind, device, inode,
    blocks)``.  The logical
    key includes the current product folder name and relative asset path; this
    makes it possible to compare the same assets after a folder-only rename.
    """
    inventory: dict[str, tuple[int, ...]] = {}
    for folder_name in folder_names:
        folder = folder_path(shop_dir, folder_name)
        if not folder.is_dir():
            raise MigrationError(f"asset inventory source folder is missing: {folder}")
        for root, _, files in os.walk(folder, topdown=True, followlinks=False):
            root_path = Path(root)
            for file_name in sorted(files):
                path = root_path / file_name
                info = path.lstat()
                relative = path.relative_to(folder).as_posix()
                logical = f"{folder_name}/{relative}"
                kind = 1 if path.is_symlink() else 0
                if kind == 0 and (info.st_size == 0 or info.st_blocks == 0):
                    raise MigrationError(
                        f"asset is empty or not hydrated; refusing migration: {path} "
                        f"(bytes={info.st_size}, blocks={info.st_blocks})"
                    )
                inventory[logical] = (
                    int(info.st_size),
                    int(info.st_mtime_ns),
                    int(info.st_mode & 0o7777),
                    kind,
                    int(info.st_dev),
                    int(info.st_ino),
                    int(info.st_blocks),
                )
    return inventory


def remap_asset_inventory(
    inventory: dict[str, tuple[int, ...]], mapping: dict[str, str]
) -> dict[str, tuple[int, ...]]:
    expected: dict[str, tuple[int, ...]] = {}
    for logical, metadata in inventory.items():
        folder, relative = logical.split("/", 1)
        target = f"{mapping.get(folder, folder)}/{relative}"
        if target in expected:
            raise MigrationError(f"asset inventory collision after folder mapping: {target}")
        expected[target] = metadata
    return expected


def asset_counts(inventory: dict[str, tuple[int, int, int, int]]) -> tuple[int, int]:
    return len(inventory), sum(metadata[0] for metadata in inventory.values())


class ExclusiveShopLock:
    """Exclusive create-only lock for apply/rollback operations."""

    def __init__(self, shop_dir: Path) -> None:
        self.path = shop_dir / LOCK_NAME
        self.handle: Any = None

    def __enter__(self) -> "ExclusiveShopLock":
        try:
            self.handle = self.path.open("x", encoding="utf-8")
            self.handle.write(json.dumps({"pid": os.getpid(), "created_at": utc_timestamp()}))
            self.handle.write("\n")
            self.handle.flush()
            os.fsync(self.handle.fileno())
            fcntl.flock(self.handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except (BlockingIOError, FileExistsError, OSError) as exc:
            if self.handle is not None:
                self.handle.close()
                self.handle = None
            raise MigrationError(f"could not acquire exclusive shop lock {self.path}: {exc}") from exc
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, traceback: Any) -> None:
        if self.handle is not None:
            fcntl.flock(self.handle.fileno(), fcntl.LOCK_UN)
            self.handle.close()
            self.handle = None
            self.path.unlink(missing_ok=True)
            self.handle = None


def parse_product_name(value: Any, *, context: str) -> tuple[str, int]:
    if not isinstance(value, str):
        raise MigrationError(f"{context}: expected product-N text, got {value!r}")
    match = PRODUCT_RE.fullmatch(value)
    if not match:
        raise MigrationError(f"{context}: invalid product-N value {value!r}")
    return value, int(match.group(1))


def iter_dicts(value: Any) -> Iterator[dict[str, Any]]:
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from iter_dicts(child)
    elif isinstance(value, list):
        for child in value:
            yield from iter_dicts(child)


def fsync_directory(path: Path) -> None:
    """Best-effort directory durability after an atomic replace."""
    try:
        fd = os.open(path, os.O_RDONLY)
    except OSError:
        return
    try:
        os.fsync(fd)
    finally:
        os.close(fd)


def atomic_write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{path.name}.tmp-", suffix=".json", dir=path.parent)
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as handle:
            json.dump(value, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(tmp_path, path)
        fsync_directory(path.parent)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def atomic_copy(path: Path, destination: Path) -> None:
    """Copy one file into place without exposing a partial destination."""
    if not path.is_file():
        raise MigrationError(f"required file is missing: {path}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{destination.name}.tmp-", dir=destination.parent)
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "wb") as handle, path.open("rb") as source:
            shutil.copyfileobj(source, handle)
            handle.flush()
            os.fsync(handle.fileno())
        shutil.copystat(path, tmp_path)
        os.replace(tmp_path, destination)
        fsync_directory(destination.parent)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def atomic_save_workbook(workbook: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{path.name}.tmp-", suffix=".xlsx", dir=path.parent)
    tmp_path = Path(tmp_name)
    try:
        os.close(fd)
        workbook.save(tmp_path)
        with tmp_path.open("rb") as handle:
            os.fsync(handle.fileno())
        os.replace(tmp_path, path)
        fsync_directory(path.parent)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def find_listings_header(ws: Any) -> int:
    for row in range(1, ws.max_row + 1):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        if str(a).strip() == "STT" and isinstance(b, str) and "Folder" in b:
            return row
    raise MigrationError("Listings sheet has no recognizable STT/Folder header")


def read_catalog(excel_path: Path, *, expected_count: int | None = DEFAULT_EXPECTED_CATALOG_COUNT) -> tuple[Any, tuple[tuple[int, str], ...]]:
    if not excel_path.is_file():
        raise MigrationError(f"workbook is missing: {excel_path}")
    workbook = load_workbook(excel_path, data_only=False)
    if "Listings" not in workbook.sheetnames:
        raise MigrationError("workbook is missing the Listings sheet")
    ws = workbook["Listings"]
    header_row = find_listings_header(ws)
    rows: list[tuple[int, str]] = []
    seen: set[str] = set()
    numeric_seen: dict[int, str] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row, 2).value
        if value is None or value == "":
            continue
        name, number = parse_product_name(value, context=f"Listings!B{row}")
        if name in seen:
            raise MigrationError(f"duplicate catalog folder {name!r} at Listings!B{row}")
        if number in numeric_seen and numeric_seen[number] != name:
            raise MigrationError(
                f"numeric folder suffix {number} has multiple names: "
                f"{numeric_seen[number]!r} and {name!r}"
            )
        seen.add(name)
        numeric_seen[number] = name
        rows.append((row, name))
    if expected_count is not None and len(rows) != expected_count:
        raise MigrationError(f"expected {expected_count} catalog folders, found {len(rows)}")
    if not rows:
        raise MigrationError("Listings column B contains no catalog folders")
    return workbook, tuple(rows)


def discover_current_folders(shop_dir: Path, catalog_names: Iterable[str]) -> tuple[str, ...]:
    if not shop_dir.is_dir():
        raise MigrationError(f"shop directory is missing: {shop_dir}")
    catalog = set(catalog_names)
    current_dirs = {
        child.name
        for child in shop_dir.iterdir()
        if child.is_dir() and PRODUCT_RE.fullmatch(child.name)
    }
    missing = sorted(catalog - current_dirs, key=lambda name: (int(name.split("-")[1]), name))
    extra = sorted(current_dirs - catalog, key=lambda name: (int(name.split("-")[1]), name))
    if missing or extra:
        pieces = []
        if missing:
            pieces.append(f"missing folders: {', '.join(missing[:12])}{' ...' if len(missing) > 12 else ''}")
        if extra:
            pieces.append(f"extra current folders: {', '.join(extra[:12])}{' ...' if len(extra) > 12 else ''}")
        raise MigrationError("folder/catalog one-to-one validation failed; " + "; ".join(pieces))
    return tuple(sorted(current_dirs, key=lambda name: (int(name.split("-")[1]), name)))


def make_mapping(current_names: Iterable[str], rows: Iterable[tuple[int, str]], *, token: str | None = None) -> tuple[FolderMapping, ...]:
    current = list(current_names)
    if not current:
        raise MigrationError("cannot create a mapping for an empty catalog")
    # Match the existing Temply convention: two digits for the leading range
    # (product-01..product-99), then ordinary decimal names (product-100+).
    width = 2
    row_by_name: dict[str, list[int]] = {}
    for row, name in rows:
        row_by_name.setdefault(name, []).append(row)
    token = token or uuid.uuid4().hex
    mappings = []
    for rank, old_name in enumerate(current, start=1):
        new_name = f"product-{rank:0{width}d}"
        mappings.append(
            FolderMapping(
                old_name=old_name,
                new_name=new_name,
                temporary_name=f".temply-folder-renumber-{token}-{rank:03d}",
                rollback_temporary_name=f".temply-folder-renumber-rollback-{token}-{rank:03d}",
                excel_rows=tuple(row_by_name.get(old_name, ())),
            )
        )
    return tuple(mappings)


def mapping_dict(mappings: Iterable[FolderMapping]) -> dict[str, str]:
    return {item.old_name: item.new_name for item in mappings}


def transform_map_data(data: Any, mapping: dict[str, str], *, shop_id: str) -> tuple[Any, int, int]:
    """Change only exact nested ``<shop_id>`` values.

    ``sync:<shop_id>/...`` keys are intentionally left byte/structure-wise
    unchanged, including keys whose suffix is a current folder.
    """
    if not isinstance(data, dict):
        raise MigrationError("product_source_map.json must contain a JSON object")
    result = copy.deepcopy(data)
    value_updates = 0

    def transform_dict(obj: dict[str, Any]) -> None:
        nonlocal value_updates
        for key, value in list(obj.items()):
            new_value = value
            if key == shop_id and isinstance(value, str) and value in mapping:
                new_value = mapping[value]
                if new_value != value:
                    value_updates += 1
            obj[key] = new_value
        for value in obj.values():
            if isinstance(value, dict):
                transform_dict(value)
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        transform_dict(item)

    transform_dict(result)
    return result, 0, value_updates


def transform_social_data(
    data: Any,
    mapping: dict[str, str],
    *,
    path: Path,
) -> tuple[Any, int, int]:
    """Update only exact current keys in a ``products`` object.

    The row field and arbitrary report/folder fields are deliberately ignored.
    A stale key such as ``product-414`` must remain attached to its original
    social record even if its row now points at a different catalog product.
    """
    if not isinstance(data, dict):
        raise MigrationError(f"{path} must contain a JSON object")
    result = copy.deepcopy(data)
    key_updates = 0
    folder_updates = 0

    def visit(obj: Any) -> None:
        nonlocal key_updates, folder_updates
        if isinstance(obj, dict):
            for key, value in list(obj.items()):
                if key == "products" and isinstance(value, dict):
                    rewritten: list[tuple[Any, Any]] = []
                    for product_key, record in value.items():
                        new_key = product_key
                        new_record = record
                        if isinstance(product_key, str) and product_key in mapping:
                            new_key = mapping[product_key]
                            if new_key != product_key:
                                key_updates += 1
                            if isinstance(record, dict) and record.get("folder") == product_key:
                                new_record = copy.deepcopy(record)
                                new_record["folder"] = new_key
                                folder_updates += 1
                        rewritten.append((new_key, new_record))
                    keys = [product_key for product_key, _ in rewritten]
                    if len(keys) != len(set(keys)):
                        raise MigrationError(
                            f"{path} update would collide with an existing social product key"
                        )
                    obj[key] = dict(rewritten)
                    for record in obj[key].values():
                        visit(record)
                else:
                    visit(value)
        elif isinstance(obj, list):
            for item in obj:
                visit(item)

    visit(result)
    return result, key_updates, folder_updates


def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{path.name}.tmp-", dir=path.parent)
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(tmp_path, path)
        fsync_directory(path.parent)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def load_json(path: Path) -> Any:
    if not path.is_file():
        raise MigrationError(f"required JSON file is missing: {path}")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise MigrationError(f"invalid JSON in {path}: {exc}") from exc


def check_unfinished_journal(backup_root: Path) -> None:
    if not backup_root.exists():
        return
    for candidate_dir in sorted(backup_root.iterdir()):
        journal_path = candidate_dir / JOURNAL_NAME
        if not candidate_dir.is_dir() or not journal_path.is_file():
            continue
        try:
            journal = json.loads(journal_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise MigrationError(f"cannot inspect journal {journal_path}: {exc}") from exc
        if journal.get("status") not in TERMINAL_JOURNAL_STATUSES:
            raise MigrationError(f"unfinished migration journal exists: {journal_path}")


def manifest_for_plan(plan: MigrationPlan, backup_dir: Path) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "migration": "temply-folder-renumber",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "shop_id": plan.shop_id,
        "shop_dir": str(plan.shop_dir),
        "excel_path": str(plan.excel_path),
        "map_path": str(plan.map_path),
        "social_paths": [str(path) for path in plan.social_paths],
        "shop_manifest_path": str(plan.shop_manifest_path),
        "backup_dir": str(backup_dir),
        "catalog_count": len(plan.mappings),
        "mapping": [
            {
                "old_folder": item.old_name,
                "new_folder": item.new_name,
                "temporary_folder": item.temporary_name,
                "rollback_temporary_folder": item.rollback_temporary_name,
                "excel_rows": list(item.excel_rows),
            }
            for item in plan.mappings
        ],
        "source_hashes": plan.source_hashes,
        "asset_inventory": {
            key: list(value) for key, value in plan.asset_inventory.items()
        },
        "counts": {
            "map_key_updates": plan.map_key_updates,
            "map_value_updates": plan.map_value_updates,
            "social_key_updates": plan.social_key_updates,
            "social_folder_updates": plan.social_folder_updates,
            "shop_manifest_updates": plan.shop_manifest_updates,
        },
    }


def reject_leftover_temporary_folders(shop_dir: Path) -> None:
    leftovers = sorted(
        child.name
        for child in shop_dir.iterdir()
        if child.name.startswith(".temply-folder-renumber-")
    )
    if leftovers:
        raise MigrationError(
            "temporary renumber folders already exist; inspect the previous migration before retrying: "
            + ", ".join(leftovers[:12])
        )


def revalidate_plan_sources(plan: MigrationPlan) -> None:
    """Prove the dry-run inputs did not change before the first mutation."""
    reject_leftover_temporary_folders(plan.shop_dir)
    for path in (
        plan.excel_path,
        plan.map_path,
        *plan.social_paths,
        plan.shop_manifest_path,
    ):
        key = str(path)
        expected_hash = plan.source_hashes.get(key)
        if expected_hash is None:
            raise MigrationError(f"plan has no source hash for {path}")
        actual_hash = sha256_file(path)
        if actual_hash != expected_hash:
            raise MigrationError(
                f"source changed after preflight; refusing migration: {path} "
                f"(expected {expected_hash}, found {actual_hash})"
            )

    workbook, rows = read_catalog(plan.excel_path, expected_count=len(plan.catalog_rows))
    current_snapshot = workbook_snapshot(workbook, (row for row, _ in rows))
    workbook.close()
    if rows != plan.catalog_rows or current_snapshot != plan.workbook_snapshot:
        raise MigrationError("workbook structure or catalog rows changed after preflight")
    current_names = discover_current_folders(
        plan.shop_dir, (name for _, name in plan.catalog_rows)
    )
    expected_names = tuple(item.old_name for item in plan.mappings)
    if current_names != expected_names:
        raise MigrationError("current folder set/order changed after preflight")
    current_assets = inventory_assets(plan.shop_dir, current_names)
    if current_assets != plan.asset_inventory:
        raise MigrationError("product asset metadata changed after preflight")


def build_plan(
    *,
    shop_dir: Path,
    excel_path: Path,
    map_path: Path,
    social_paths: Iterable[Path] | None = None,
    shop_manifest_path: Path | None = None,
    expected_count: int | None = DEFAULT_EXPECTED_CATALOG_COUNT,
    token: str | None = None,
) -> MigrationPlan:
    shop_id = shop_dir.name
    reject_leftover_temporary_folders(shop_dir)
    workbook, catalog_rows = read_catalog(excel_path, expected_count=expected_count)
    workbook_state = workbook_snapshot(workbook, (row for row, _ in catalog_rows))
    workbook.close()
    catalog_names = [name for _, name in catalog_rows]
    current_names = discover_current_folders(shop_dir, catalog_names)
    mappings = make_mapping(current_names, catalog_rows, token=token)
    if any(not item.excel_rows for item in mappings):
        raise MigrationError("mapping contains a current folder without an Excel row")
    if any(len(item.excel_rows) != 1 for item in mappings):
        raise MigrationError("each current folder must map to exactly one Excel row")

    map_data = load_json(map_path)
    updated_map, map_key_updates, map_value_updates = transform_map_data(
        map_data, mapping_dict(mappings), shop_id=shop_id
    )
    paths = tuple(social_paths or (shop_dir / name for name in SOCIAL_STORE_NAMES))
    if len(paths) != len(SOCIAL_STORE_NAMES):
        raise MigrationError(
            f"expected exactly {len(SOCIAL_STORE_NAMES)} social stores, found {len(paths)}"
        )
    updated_social: list[tuple[Path, Any]] = []
    social_key_updates = 0
    social_folder_updates = 0
    for path in paths:
        data = load_json(path)
        transformed, key_count, folder_count = transform_social_data(
            data,
            mapping_dict(mappings),
            path=path,
        )
        updated_social.append((path, transformed))
        social_key_updates += key_count
        social_folder_updates += folder_count

    shop_manifest_path = shop_manifest_path or (shop_dir / SHOP_MANIFEST_NAME)
    if not shop_manifest_path.is_file():
        raise MigrationError(f"required shop product manifest is missing: {shop_manifest_path}")
    try:
        manifest_text = shop_manifest_path.read_text(encoding="utf-8")
    except OSError as exc:
        raise MigrationError(f"cannot read shop product manifest {shop_manifest_path}: {exc}") from exc
    # This is a historical packaging report, not a live catalog index.  Keep it
    # byte-for-byte unchanged and record its hash as an immutable postflight
    # invariant.
    updated_manifest_text = manifest_text
    shop_manifest_updates = 0

    authoritative_paths = (excel_path, map_path, *paths, shop_manifest_path)
    source_hashes = {str(path): sha256_file(path) for path in authoritative_paths}
    asset_inventory = inventory_assets(shop_dir, current_names)

    return MigrationPlan(
        shop_id=shop_id,
        shop_dir=shop_dir,
        excel_path=excel_path,
        map_path=map_path,
        social_paths=paths,
        shop_manifest_path=shop_manifest_path,
        catalog_rows=catalog_rows,
        mappings=mappings,
        updated_map=updated_map,
        updated_social=tuple(updated_social),
        map_key_updates=map_key_updates,
        map_value_updates=map_value_updates,
        social_key_updates=social_key_updates,
        social_folder_updates=social_folder_updates,
        shop_manifest_updates=shop_manifest_updates,
        updated_shop_manifest=updated_manifest_text,
        source_hashes=source_hashes,
        asset_inventory=asset_inventory,
        workbook_snapshot=workbook_state,
    )


def create_backup(plan: MigrationPlan, backup_root: Path) -> tuple[Path, dict[str, Any]]:
    backup_root.mkdir(parents=True, exist_ok=True)
    backup_dir = backup_root / f"{plan.shop_id}-folder-renumber-{utc_timestamp()}-{uuid.uuid4().hex[:8]}"
    backup_dir.mkdir()
    atomic_copy(plan.excel_path, backup_dir / BACKUP_EXCEL_NAME)
    atomic_copy(plan.map_path, backup_dir / BACKUP_MAP_NAME)
    for path, backup_name in zip(plan.social_paths, (BACKUP_STATUS_NAME, BACKUP_SOCIAL_NAME)):
        atomic_copy(path, backup_dir / backup_name)
    atomic_copy(plan.shop_manifest_path, backup_dir / BACKUP_SHOP_MANIFEST_NAME)
    manifest = manifest_for_plan(plan, backup_dir)
    atomic_write_json(backup_dir / MANIFEST_NAME, manifest)
    journal = {
        "schema_version": 1,
        "status": "prepared",
        "manifest": str(backup_dir / MANIFEST_NAME),
        "phase1_completed": [],
        "phase2_completed": [],
        "files_updated": [],
    }
    atomic_write_json(backup_dir / JOURNAL_NAME, journal)
    return backup_dir, journal


def write_journal(path: Path, journal: dict[str, Any], **updates: Any) -> None:
    journal.update(updates)
    atomic_write_json(path, journal)


def folder_path(shop_dir: Path, name: str) -> Path:
    if name in {".", ".."} or "/" in name or "\\" in name:
        raise MigrationError(f"unsafe folder name in manifest: {name!r}")
    return shop_dir / name


def apply_folder_renames(plan: MigrationPlan, journal_path: Path, journal: dict[str, Any]) -> None:
    phase1: list[str] = list(journal.get("phase1_completed", []))
    phase2: list[str] = list(journal.get("phase2_completed", []))
    changed = [item for item in plan.mappings if item.old_name != item.new_name]
    changed_old = {item.old_name for item in changed}

    for item in changed:
        target = folder_path(plan.shop_dir, item.new_name)
        temporary = folder_path(plan.shop_dir, item.temporary_name)
        if item.old_name not in phase2 and item.new_name not in changed_old and target.exists():
            raise MigrationError(f"refusing to overwrite existing target folder: {item.new_name}")
        if item.old_name in phase2:
            if not target.is_dir() or temporary.exists():
                raise MigrationError(
                    f"journal says phase 2 is complete but target state is unsafe: {item.old_name}"
                )
        elif item.old_name in phase1:
            if not temporary.is_dir():
                raise MigrationError(f"journal says phase 1 is complete but temporary is missing: {item.old_name}")
        elif temporary.exists():
            raise MigrationError(f"temporary folder already exists: {item.temporary_name}")
    write_journal(journal_path, journal, status="renaming_phase1")
    for item in changed:
        if item.old_name in phase1:
            continue
        source = folder_path(plan.shop_dir, item.old_name)
        temporary = folder_path(plan.shop_dir, item.temporary_name)
        if not source.is_dir() or temporary.exists():
            raise MigrationError(f"phase 1 source/temporary state is unsafe for {item.old_name}")
        source.rename(temporary)
        phase1.append(item.old_name)
        write_journal(journal_path, journal, phase1_completed=phase1)

    write_journal(journal_path, journal, status="renaming_phase2")
    for item in changed:
        if item.old_name in phase2:
            continue
        temporary = folder_path(plan.shop_dir, item.temporary_name)
        target = folder_path(plan.shop_dir, item.new_name)
        if not temporary.is_dir() or target.exists():
            raise MigrationError(f"phase 2 source/target state is unsafe for {item.old_name} -> {item.new_name}")
        temporary.rename(target)
        phase2.append(item.old_name)
        write_journal(journal_path, journal, phase2_completed=phase2)


def rollback_folder_renames(plan: MigrationPlan, journal_path: Path, journal: dict[str, Any]) -> None:
    changed = [item for item in plan.mappings if item.old_name != item.new_name]
    phase1_done = set(journal.get("phase1_completed", []))
    phase2_done = set(journal.get("phase2_completed", []))
    if not phase2_done.issubset(phase1_done):
        raise MigrationError("rollback journal is inconsistent: phase 2 contains an incomplete phase 1 item")
    completed = [item for item in changed if item.old_name in phase1_done]
    evacuated = set(journal.get("rollback_phase1_completed", []))
    restored = set(journal.get("rollback_phase2_completed", []))
    write_journal(journal_path, journal, status="rolling_back")

    # First evacuate all final targets to unique names.  This is required when
    # a target name is also an old name elsewhere in the mapping.
    for item in reversed(completed):
        if item.old_name not in phase2_done or item.old_name in evacuated:
            continue
        target = folder_path(plan.shop_dir, item.new_name)
        rollback_temporary = folder_path(plan.shop_dir, item.rollback_temporary_name)
        if rollback_temporary.exists():
            if target.exists():
                raise MigrationError(f"rollback temporary and target both exist: {item.old_name}")
        elif not target.is_dir():
            raise MigrationError(f"rollback target is missing: {item.new_name}")
        else:
            target.rename(rollback_temporary)
        evacuated.add(item.old_name)
        write_journal(journal_path, journal, rollback_phase1_completed=sorted(evacuated))

    # Then restore each original name from either the rollback temporary (if
    # phase 2 had completed) or the normal phase-1 temporary.
    for item in reversed(completed):
        if item.old_name in restored:
            continue
        source_name = (
            item.rollback_temporary_name
            if item.old_name in phase2_done
            else item.temporary_name
        )
        source = folder_path(plan.shop_dir, source_name)
        old_path = folder_path(plan.shop_dir, item.old_name)
        if old_path.exists():
            raise MigrationError(f"rollback would overwrite existing old folder: {item.old_name}")
        if not source.is_dir():
            raise MigrationError(f"rollback source is missing: {source_name}")
        source.rename(old_path)
        restored.add(item.old_name)
        write_journal(journal_path, journal, rollback_phase2_completed=sorted(restored))
    write_journal(journal_path, journal, status="folders_rolled_back")


def restore_backup_files(plan: MigrationPlan, backup_dir: Path) -> list[str]:
    restored: list[str] = []
    for source_name, destination in (
        (BACKUP_EXCEL_NAME, plan.excel_path),
        (BACKUP_MAP_NAME, plan.map_path),
    ):
        atomic_copy(backup_dir / source_name, destination)
        restored.append(str(destination))
    for path, backup_name in zip(plan.social_paths, (BACKUP_STATUS_NAME, BACKUP_SOCIAL_NAME)):
        atomic_copy(backup_dir / backup_name, path)
        restored.append(str(path))
    atomic_copy(backup_dir / BACKUP_SHOP_MANIFEST_NAME, plan.shop_manifest_path)
    restored.append(str(plan.shop_manifest_path))
    return restored


def postflight_rollback(plan: MigrationPlan, backup_dir: Path) -> None:
    expected_names = {item.old_name for item in plan.mappings}
    actual_names = {
        child.name
        for child in plan.shop_dir.iterdir()
        if child.is_dir() and PRODUCT_RE.fullmatch(child.name)
    }
    if actual_names != expected_names:
        raise MigrationError(
            f"rollback folder set mismatch: expected {len(expected_names)}, found {len(actual_names)}"
        )
    reject_leftover_temporary_folders(plan.shop_dir)
    if plan.asset_inventory:
        actual_assets = inventory_assets(plan.shop_dir, expected_names)
        if actual_assets != plan.asset_inventory:
            raise MigrationError("rollback changed product asset metadata")
    backup_pairs = (
        (backup_dir / BACKUP_EXCEL_NAME, plan.excel_path),
        (backup_dir / BACKUP_MAP_NAME, plan.map_path),
        *zip(
            (backup_dir / BACKUP_STATUS_NAME, backup_dir / BACKUP_SOCIAL_NAME),
            plan.social_paths,
        ),
        (backup_dir / BACKUP_SHOP_MANIFEST_NAME, plan.shop_manifest_path),
    )
    for backup, destination in backup_pairs:
        if sha256_file(backup) != sha256_file(destination):
            raise MigrationError(f"rollback file does not match its exact backup: {destination}")


def update_workbook_column_b(plan: MigrationPlan) -> None:
    workbook = load_workbook(plan.excel_path, data_only=False)
    ws = workbook["Listings"]
    mapping = mapping_dict(plan.mappings)
    for row, old_name in plan.catalog_rows:
        ws.cell(row, 2).value = mapping[old_name]
    atomic_save_workbook(workbook, plan.excel_path)


def postflight_counts(plan: MigrationPlan) -> dict[str, int]:
    names = {
        child.name
        for child in plan.shop_dir.iterdir()
        if child.is_dir() and PRODUCT_RE.fullmatch(child.name)
    }
    expected = {item.new_name for item in plan.mappings}
    if names != expected:
        raise MigrationError(f"postflight folder set mismatch: expected {len(expected)}, found {len(names)}")
    mapping = mapping_dict(plan.mappings)
    reject_leftover_temporary_folders(plan.shop_dir)

    workbook, rows = read_catalog(plan.excel_path, expected_count=len(plan.catalog_rows))
    expected_rows = tuple((row, mapping[old_name]) for row, old_name in plan.catalog_rows)
    if rows != expected_rows:
        raise MigrationError("postflight workbook catalog rows do not match the renumber mapping")
    current_snapshot = workbook_snapshot(workbook, (row for row, _ in rows))
    workbook.close()
    if current_snapshot != plan.workbook_snapshot:
        raise MigrationError(
            "postflight workbook structure, non-folder fields, or styles changed unexpectedly"
        )

    map_data = load_json(plan.map_path)
    if map_data != plan.updated_map:
        raise MigrationError("postflight product_source_map.json does not match the reviewed transformation")
    expected_social = {str(path): transformed for path, transformed in plan.updated_social}
    for path in plan.social_paths:
        if load_json(path) != expected_social[str(path)]:
            raise MigrationError(f"postflight social store does not match the reviewed transformation: {path}")
    manifest_hash = sha256_file(plan.shop_manifest_path)
    if manifest_hash != plan.source_hashes[str(plan.shop_manifest_path)]:
        raise MigrationError("postflight historical product manifest changed unexpectedly")

    actual_assets = inventory_assets(plan.shop_dir, expected)
    expected_assets = remap_asset_inventory(plan.asset_inventory, mapping)
    if actual_assets != expected_assets:
        raise MigrationError("postflight asset metadata changed during folder renumber")
    asset_file_count, asset_bytes = asset_counts(actual_assets)
    return {
        "catalog_folders": len(expected),
        "excel_column_b_updates": len(plan.catalog_rows),
        "map_key_updates": plan.map_key_updates,
        "map_value_updates": plan.map_value_updates,
        "social_key_updates": plan.social_key_updates,
        "social_folder_updates": plan.social_folder_updates,
        "shop_manifest_updates": plan.shop_manifest_updates,
        "asset_files": asset_file_count,
        "asset_bytes": asset_bytes,
    }


def apply_plan(plan: MigrationPlan, backup_root: Path) -> tuple[Path, dict[str, int]]:
    with ExclusiveShopLock(plan.shop_dir):
        check_unfinished_journal(backup_root)
        revalidate_plan_sources(plan)
        backup_dir, journal = create_backup(plan, backup_root)
        journal_path = backup_dir / JOURNAL_NAME
        try:
            apply_folder_renames(plan, journal_path, journal)
            write_journal(journal_path, journal, status="updating_files")
            update_workbook_column_b(plan)
            write_journal(journal_path, journal, files_updated=[str(plan.excel_path)])
            atomic_write_json(plan.map_path, plan.updated_map)
            write_journal(journal_path, journal, files_updated=[str(plan.excel_path), str(plan.map_path)])
            for path, transformed in plan.updated_social:
                atomic_write_json(path, transformed)
                write_journal(
                    journal_path,
                    journal,
                    files_updated=[*journal["files_updated"], str(path)],
                )
            counts = postflight_counts(plan)
            write_journal(journal_path, journal, status="completed", postflight=counts)
            return backup_dir, counts
        except Exception as exc:
            try:
                rollback_folder_renames(plan, journal_path, journal)
                restored = restore_backup_files(plan, backup_dir)
                write_journal(
                    journal_path,
                    journal,
                    status="rolled_back",
                    failure=str(exc),
                    restored_files=restored,
                )
            except Exception as rollback_exc:
                write_journal(
                    journal_path,
                    journal,
                    status="unfinished",
                    failure=str(exc),
                    rollback_failure=str(rollback_exc),
                )
                raise MigrationError(
                    f"migration failed and automatic rollback is incomplete; journal: {journal_path}; "
                    f"original error: {exc}; rollback error: {rollback_exc}"
                ) from exc
            raise MigrationError(
                f"migration failed; automatic rollback completed; journal: {journal_path}; {exc}"
            ) from exc


def plan_summary(plan: MigrationPlan) -> str:
    changed = sum(item.old_name != item.new_name for item in plan.mappings)
    first = plan.mappings[0]
    last = plan.mappings[-1]
    return (
        f"shop={plan.shop_id} folders={len(plan.mappings)} changed={changed} "
        f"excel_rows={len(plan.catalog_rows)} map_keys={plan.map_key_updates} "
        f"map_values={plan.map_value_updates} social_keys={plan.social_key_updates} "
        f"social_folders={plan.social_folder_updates} manifest={plan.shop_manifest_updates} "
        f"range={first.old_name}->{first.new_name}, {last.old_name}->{last.new_name}"
    )


def load_plan_from_manifest(manifest_path: Path) -> tuple[MigrationPlan, dict[str, Any], Path]:
    manifest = load_json(manifest_path)
    if manifest.get("migration") != "temply-folder-renumber" or manifest.get("schema_version") != 1:
        raise MigrationError(f"unsupported migration manifest: {manifest_path}")
    mappings = tuple(
        FolderMapping(
            old_name=item["old_folder"],
            new_name=item["new_folder"],
            temporary_name=item["temporary_folder"],
            rollback_temporary_name=item.get(
                "rollback_temporary_folder",
                f"{item['temporary_folder']}-rollback",
            ),
            excel_rows=tuple(item.get("excel_rows", [])),
        )
        for item in manifest["mapping"]
    )
    social_paths = tuple(Path(value) for value in manifest["social_paths"])
    plan = MigrationPlan(
        shop_id=manifest["shop_id"],
        shop_dir=Path(manifest["shop_dir"]),
        excel_path=Path(manifest["excel_path"]),
        map_path=Path(manifest["map_path"]),
        social_paths=social_paths,
        catalog_rows=tuple((row, item.old_name) for item in mappings for row in item.excel_rows),
        mappings=mappings,
        updated_map={},
        updated_social=(),
        map_key_updates=manifest.get("counts", {}).get("map_key_updates", 0),
        map_value_updates=manifest.get("counts", {}).get("map_value_updates", 0),
        social_key_updates=manifest.get("counts", {}).get("social_key_updates", 0),
        social_folder_updates=manifest.get("counts", {}).get("social_folder_updates", 0),
        shop_manifest_path=Path(manifest["shop_manifest_path"]),
        shop_manifest_updates=manifest.get("counts", {}).get("shop_manifest_updates", 0),
        updated_shop_manifest="",
        source_hashes={str(key): str(value) for key, value in manifest.get("source_hashes", {}).items()},
        asset_inventory={
            str(key): tuple(int(value) for value in metadata)
            for key, metadata in manifest.get("asset_inventory", {}).items()
        },
        workbook_snapshot={},
    )
    return plan, manifest, manifest_path.parent


def rollback_backup(backup_dir: Path, *, allow_non_temply_shop: bool = False) -> dict[str, Any]:
    manifest_path = backup_dir / MANIFEST_NAME
    journal_path = backup_dir / JOURNAL_NAME
    plan, manifest, _ = load_plan_from_manifest(manifest_path)
    if plan.shop_id != DEFAULT_SHOP_ID and not allow_non_temply_shop:
        raise MigrationError(
            f"refusing rollback for shop {plan.shop_id!r}; "
            "use --allow-non-temply-shop for explicit safety override"
        )
    journal = load_json(journal_path)
    if journal.get("status") == "rolled_back":
        return {"status": "already_rolled_back", "backup_dir": str(backup_dir)}
    if journal.get("status") not in {
        "completed",
        "unfinished",
        "renaming_phase1",
        "renaming_phase2",
        "updating_files",
        "folders_rolled_back",
    }:
        raise MigrationError(f"journal is not rollback-ready: {journal.get('status')!r}")
    with ExclusiveShopLock(plan.shop_dir):
        if journal.get("status") == "completed":
            journal["phase1_completed"] = [
                item.old_name for item in plan.mappings if item.old_name != item.new_name
            ]
            journal["phase2_completed"] = [
                item.old_name for item in plan.mappings if item.old_name != item.new_name
            ]
        if journal.get("status") != "folders_rolled_back":
            rollback_folder_renames(plan, journal_path, journal)
        restored = restore_backup_files(plan, backup_dir)
        postflight_rollback(plan, backup_dir)
        write_journal(journal_path, journal, status="rolled_back", restored_files=restored)
        return {"status": "rolled_back", "backup_dir": str(backup_dir), "restored_files": len(restored)}


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    repo_root = Path(__file__).resolve().parents[1]
    parser.add_argument("--shop-dir", type=Path, default=repo_root / "shops" / DEFAULT_SHOP_ID)
    parser.add_argument("--excel", type=Path, default=None)
    parser.add_argument("--map-file", type=Path, default=repo_root / "product_source_map.json")
    parser.add_argument(
        "--backup-root",
        type=Path,
        default=repo_root / "output" / "backup" / "temply-folder-renumber",
    )
    parser.add_argument("--apply", action="store_true", help="apply the reviewed plan; dry-run is the default")
    parser.add_argument("--rollback", type=Path, help="restore from a timestamped migration backup directory")
    parser.add_argument(
        "--allow-non-temply-shop",
        action="store_true",
        help="explicitly allow a shop directory whose name is not templystudios",
    )
    args = parser.parse_args(argv)
    if args.excel is None:
        args.excel = args.shop_dir / "Etsy_SEO_Generator.xlsx"
    if args.apply and args.rollback:
        parser.error("--apply and --rollback are mutually exclusive")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.rollback:
        try:
            result = rollback_backup(
                args.rollback.resolve(),
                allow_non_temply_shop=args.allow_non_temply_shop,
            )
        except MigrationError as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 2
        print(f"ROLLBACK {json.dumps(result, ensure_ascii=False, sort_keys=True)}")
        return 0

    shop_dir = args.shop_dir.resolve()
    excel_path = args.excel.resolve()
    map_path = args.map_file.resolve()
    if shop_dir.name != DEFAULT_SHOP_ID and not args.allow_non_temply_shop:
        print(
            f"ERROR: refusing shop {shop_dir.name!r}; use --allow-non-temply-shop for explicit safety override",
            file=sys.stderr,
        )
        return 2
    try:
        plan = build_plan(shop_dir=shop_dir, excel_path=excel_path, map_path=map_path)
        check_unfinished_journal(args.backup_root.resolve())
    except MigrationError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    print(f"PREFLIGHT {'APPLY' if args.apply else 'DRY-RUN'} {plan_summary(plan)}")
    if not args.apply:
        print(f"POSTFLIGHT (planned) folders={len(plan.mappings)} files=6 assets_copied=0")
        print("DRY-RUN: no files or folders changed")
        return 0
    try:
        backup_dir, counts = apply_plan(plan, args.backup_root.resolve())
    except MigrationError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    print(f"BACKUP {backup_dir}")
    print("POSTFLIGHT " + " ".join(f"{key}={value}" for key, value in sorted(counts.items())))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
