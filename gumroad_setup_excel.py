"""
Gumroad Excel Setup
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Chạy 1 lần duy nhất để tạo sheet "Gumroad" trong file
Etsy_SEO_Generator.xlsx.

Script sẽ:
  1. Tạo sheet "Gumroad" với đầy đủ headers
  2. Copy data từ sheet "Listings" sang (folder, keywords,
     price, title, description, tags) — tiết kiệm thời gian
     không cần nhập lại
  3. Các cột Gumroad-specific (summary, permalink, status)
     để trống để anh/Claude điền sau

Cách dùng:
  python3 gumroad_setup_excel.py
"""

import sys, subprocess
from pathlib import Path
import re

def ensure_deps():
    try:
        import openpyxl
    except ImportError:
        print("▶ Cài openpyxl...")
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"], check=True)

ensure_deps()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR   = Path(__file__).parent

# ── Màu sắc header ─────────────────────────────────────────────────────────────
COLOR_HEADER_BG  = "1DA462"   # Gumroad green
COLOR_HEADER_FG  = "FFFFFF"
COLOR_AUTO_BG    = "E8F5E9"   # Xanh nhạt = cột script tự điền
COLOR_MANUAL_BG  = "FFF9C4"   # Vàng nhạt = cột anh điền
COLOR_STATUS_BG  = "E3F2FD"   # Xanh nhạt = status

def make_permalink(folder: str, keywords: str) -> str:
    kw = str(keywords or "").strip()
    if kw:
        first = kw.split(",")[0].strip().lower()
        slug  = re.sub(r"[^a-z0-9\-]", "-", first)
        slug  = re.sub(r"-+", "-", slug).strip("-")
        if slug:
            return slug[:60]
    return folder.lower().replace("_", "-")

def style_cell(cell, bold=False, bg=None, fg="000000", wrap=False, align="left"):
    cell.font      = Font(bold=bold, color=fg, name="Calibri", size=10)
    cell.alignment = Alignment(wrap_text=wrap, vertical="top",
                                horizontal=align)
    if bg:
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def setup_gumroad_sheet(shop_id: str):
    excel_file = BASE_DIR / "shops" / shop_id / "Etsy_SEO_Generator.xlsx"
    if not excel_file.exists():
        print(f"❌ Không tìm thấy Excel của shop tại: {excel_file}")
        return
        
    wb = openpyxl.load_workbook(excel_file)

    # ── Kiểm tra / tạo sheet ─────────────────────────────────────────────────
    if "Gumroad" in wb.sheetnames:
        print(f"ℹ  Sheet 'Gumroad' đã tồn tại trong Excel của shop '{shop_id}'.")
        ans = input("   Ghi đè? (y/N): ").strip().lower()
        if ans != "y":
            print("✅ Giữ nguyên, không thay đổi.")
            return
        del wb["Gumroad"]
        print("   → Đã xóa sheet cũ, tạo lại...")

    ws = wb.create_sheet("Gumroad")

    # ── Row 1: Tiêu đề dự án ─────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = "🟢  GUMROAD LISTINGS  —  Digital Products Auto Poster"
    c.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    c.fill      = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Row 2: Chú thích màu ─────────────────────────────────────────────────
    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = ("🟡 Vàng = Anh điền  |  🟢 Xanh nhạt = Claude/Script tự điền  |  "
               "🔵 Xanh = Script cập nhật tự động")
    c.font      = Font(italic=True, size=9, color="555555")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: Headers ────────────────────────────────────────────────────────
    headers = [
        ("A", "STT",         COLOR_HEADER_BG, 6),
        ("B", "Folder",      COLOR_MANUAL_BG, 12),
        ("C", "Keywords",    COLOR_MANUAL_BG, 30),
        ("D", "Ghi chú",     COLOR_MANUAL_BG, 25),
        ("E", "Giá (USD)",   COLOR_MANUAL_BG, 11),
        ("F", "—",           "F5F5F5",         8),   # unused
        ("G", "—",           "F5F5F5",         8),   # unused
        ("H", "Title",       COLOR_AUTO_BG,   45),
        ("I", "Description", COLOR_AUTO_BG,   50),
        ("J", "Tags",        COLOR_AUTO_BG,   40),
        ("K", "Summary",     COLOR_AUTO_BG,   35),
        ("L", "Permalink",   COLOR_AUTO_BG,   25),
        ("M", "—",           "F5F5F5",         8),   # unused
        ("N", "Status",      COLOR_STATUS_BG, 18),
    ]

    ws.row_dimensions[3].height = 22
    for col_letter, label, bg, width in headers:
        c = ws[f"{col_letter}3"]
        c.value = label
        style_cell(c, bold=True, bg=bg, fg="FFFFFF" if bg == COLOR_HEADER_BG else "222222",
                   align="center")
        ws.column_dimensions[col_letter].width = width

    # ── Copy data từ Listings ─────────────────────────────────────────────────
    if "Listings" not in wb.sheetnames:
        print("⚠  Không tìm thấy sheet 'Listings' — tạo sheet Gumroad trống.")
        wb.save(excel_file)
        print(f"✅ Đã tạo sheet 'Gumroad' trong {excel_file.name}")
        return

    listings_ws = wb["Listings"]
    copied = 0

    for row_num, row in enumerate(
        listings_ws.iter_rows(min_row=4, max_row=100, values_only=True), start=4
    ):
        cols = (list(row) + [None] * 15)[:15]
        _, folder, keywords, notes, price, category, _, title, description, tags, qty, _, _, status, section = cols

        if not folder:
            continue

        dest_row = row_num  # giữ cùng số row với Listings

        # Cột A: STT
        ws.cell(row=dest_row, column=1, value=copied + 1)

        # Copy dữ liệu anh điền
        ws.cell(row=dest_row, column=2, value=folder)
        ws.cell(row=dest_row, column=3, value=keywords)
        ws.cell(row=dest_row, column=4, value=notes)
        ws.cell(row=dest_row, column=5, value=price)

        # Copy SEO content (H/I/J) từ Listings
        if title and not str(title).startswith("←"):
            ws.cell(row=dest_row, column=8, value=title)
        if description:
            ws.cell(row=dest_row, column=9, value=description)
        if tags:
            ws.cell(row=dest_row, column=10, value=tags)

        # Cột L: Permalink tự động
        permalink = make_permalink(str(folder), str(keywords or ""))
        ws.cell(row=dest_row, column=12, value=permalink)

        # Cột N: Status (chưa đăng)
        ws.cell(row=dest_row, column=14, value="⏳ Chưa đăng")

        # Style cho data rows
        for col in range(1, 15):
            c = ws.cell(row=dest_row, column=col)
            thin = Side(style="thin", color="DDDDDD")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            c.font   = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="top", wrap_text=(col in (9, 10)))

        ws.row_dimensions[dest_row].height = 18
        copied += 1

    # Freeze header rows
    ws.freeze_panes = "A4"

    wb.save(excel_file)
    print(f"\n✅ Hoàn tất!")
    print(f"   • Tạo sheet 'Gumroad' trong {excel_file.name}")
    print(f"   • Copy {copied} sản phẩm từ sheet 'Listings'")
    print(f"   • Permalink tự động từ keywords")
    print(f"\n📌 Bước tiếp theo:")
    print(f"   1. Mở Excel → sheet 'Gumroad'")
    print(f"   2. Kiểm tra cột K (Summary) — thêm nếu muốn")
    print(f"   3. Chỉnh Permalink (cột L) nếu muốn URL khác")
    print(f"   4. Chạy: python3 gumroad_auto_post.py --shop {shop_id}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Gumroad Excel Setup Helper")
    parser.add_argument("--shop", type=str, required=True, help="ID của shop hiện tại (ví dụ: templystudios)")
    args = parser.parse_args()
    setup_gumroad_sheet(args.shop)
